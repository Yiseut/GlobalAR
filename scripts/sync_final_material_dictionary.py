#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_BOOK = PROJECT_DIR.parent / "全球医美企业库_标准化版v4.xlsx"
DEFAULT_DICTIONARY = Path(r"E:\shared\Downloads\final_dictionary.json")
AUDIT_DIR = PROJECT_DIR / "data" / "audits"

TAXONOMY_FIELDS = [
    "Material_Taxonomy_L1_CN",
    "Material_Taxonomy_L2_CN",
    "Material_Taxonomy_L3_CN",
    "Material_Taxonomy_Path_CN",
]

DEFINITION_HEADERS = [
    "L1 一级类",
    "L2 二级类",
    "L3 三级 / 具体材料",
    "代表产品 / 品牌(示例)",
    "给药术式 / 形态",
    "使用场景 Setting",
    "US-FDA",
    "CN-NMPA",
    "EU-CE",
    "作用机制 / 备注",
    "source_sheet",
]

L1_RENAMES = {
    "外用/护肤": "功效性护肤品",
    "外用产品": "功效性护肤品",
    "耗材器械": "耗材/器械",
}

L2_RENAMES = {
    "医学护肤": "医学护肤活性",
    "外用药": "医学护肤活性",
    "光 / IPL": "光/IPL",
    "温控 / 其他": "温控/其他",
    "皮肤清洁": "皮肤清洁/护理设备",
    "吸脂器械": "手术器械/吸脂器械",
    "设备配件": "设备配件/组件",
}

L3_RENAMES = {
    "活性成分护肤": "功效活性成分",
    "外用处方药": "功效活性成分",
    "LED光疗 / 光动力 PDT": "LED光疗/光动力 PDT",
    "等离子束 / 等离子笔": "等离子束/等离子笔",
    "剥脱点阵（CO₂/铒）": "剥脱点阵(CO₂/铒)",
    "射频微针（有创）": "射频微针(有创)",
    "聚焦超声减脂（体雕）": "聚焦超声减脂",
    "微聚焦超声 HIFU/MFU（提升）": "微聚焦超声 HIFU/MFU",
    "注射针 / 钝头套管": "注射针/钝头套管",
    "无针注射/电穿孔": "无针注射/电穿孔等",
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", norm(value))


def split_path(path: str) -> tuple[str, str, str]:
    parts = [norm(part) for part in path.split(">")]
    while len(parts) < 3:
        parts.append("")
    return parts[0], parts[1], parts[2]


def canonical_parts(l1: Any, l2: Any, l3: Any) -> tuple[str, str, str]:
    out_l1 = L1_RENAMES.get(norm(l1), norm(l1))
    out_l2 = L2_RENAMES.get(norm(l2), norm(l2))
    out_l3 = L3_RENAMES.get(norm(l3), norm(l3))
    return out_l1, out_l2, out_l3


def canonical_path(path: Any) -> str:
    l1, l2, l3 = split_path(norm(path))
    l1, l2, l3 = canonical_parts(l1, l2, l3)
    return " > ".join(part for part in [l1, l2, l3] if part)


def worksheet_headers(ws) -> dict[str, int]:
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def ensure_headers(ws, headers: list[str]) -> dict[str, int]:
    colmap = worksheet_headers(ws)
    for header in headers:
        if header not in colmap:
            col = ws.max_column + 1
            ws.cell(1, col).value = header
            colmap[header] = col
    return colmap


def load_dictionary(path: Path) -> list[tuple[str, str, str, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("final dictionary must be a list of taxonomy paths")
    rows: list[tuple[str, str, str, str]] = []
    for item in data:
        path_text = canonical_path(item)
        l1, l2, l3 = split_path(path_text)
        if l1 and l2 and l3:
            rows.append((path_text, l1, l2, l3))
    return rows


def style_definition_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="E7EEF8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [16, 22, 30, 26, 20, 18, 14, 14, 14, 42, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def sync_definition_sheet(wb, dictionary_rows: list[tuple[str, str, str, str]]) -> int:
    sheet_name = "Material_Taxonomy_Definitions"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    if ws.max_row == 0:
        ws.append(DEFINITION_HEADERS)
    colmap = ensure_headers(ws, DEFINITION_HEADERS)
    existing = set()
    for row in range(2, ws.max_row + 1):
        l1 = norm(ws.cell(row, colmap["L1 一级类"]).value)
        l2 = norm(ws.cell(row, colmap["L2 二级类"]).value)
        l3 = norm(ws.cell(row, colmap["L3 三级 / 具体材料"]).value)
        if l1 and l2 and l3:
            existing.add((compact(l1), compact(l2), compact(l3)))
    added = 0
    for _path, l1, l2, l3 in dictionary_rows:
        key = (compact(l1), compact(l2), compact(l3))
        if key in existing:
            continue
        new_row = ws.max_row + 1
        ws.cell(new_row, colmap["L1 一级类"]).value = l1
        ws.cell(new_row, colmap["L2 二级类"]).value = l2
        ws.cell(new_row, colmap["L3 三级 / 具体材料"]).value = l3
        ws.cell(new_row, colmap["作用机制 / 备注"]).value = "final_dictionary_20260527"
        ws.cell(new_row, colmap["source_sheet"]).value = "final_dictionary"
        existing.add(key)
        added += 1
    style_definition_sheet(ws)
    return added


def normalize_sheet_taxonomy(ws, record_field_candidates: list[str]) -> tuple[int, Counter]:
    colmap = ensure_headers(ws, TAXONOMY_FIELDS)
    status = Counter()
    for row in range(2, ws.max_row + 1):
        if not any(norm(ws.cell(row, colmap.get(field, 1)).value) for field in TAXONOMY_FIELDS):
            continue
        old_l1 = norm(ws.cell(row, colmap["Material_Taxonomy_L1_CN"]).value)
        old_l2 = norm(ws.cell(row, colmap["Material_Taxonomy_L2_CN"]).value)
        old_l3 = norm(ws.cell(row, colmap["Material_Taxonomy_L3_CN"]).value)
        old_path = norm(ws.cell(row, colmap["Material_Taxonomy_Path_CN"]).value)
        if old_path and not (old_l1 and old_l2 and old_l3):
            old_l1, old_l2, old_l3 = split_path(old_path)
        new_l1, new_l2, new_l3 = canonical_parts(old_l1, old_l2, old_l3)
        new_path = " > ".join(part for part in [new_l1, new_l2, new_l3] if part)
        if (new_l1, new_l2, new_l3, new_path) == (old_l1, old_l2, old_l3, old_path):
            continue
        ws.cell(row, colmap["Material_Taxonomy_L1_CN"]).value = new_l1
        ws.cell(row, colmap["Material_Taxonomy_L2_CN"]).value = new_l2
        ws.cell(row, colmap["Material_Taxonomy_L3_CN"]).value = new_l3
        ws.cell(row, colmap["Material_Taxonomy_Path_CN"]).value = new_path
        status["normalized_rows"] += 1
        if old_l1 != new_l1:
            status[f"l1:{old_l1}->{new_l1}"] += 1
        if old_l2 != new_l2:
            status[f"l2:{old_l2}->{new_l2}"] += 1
        if old_l3 != new_l3:
            status[f"l3:{old_l3}->{new_l3}"] += 1
    return status["normalized_rows"], status


def run(source_book: Path, dictionary_path: Path) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = source_book.with_name(f"{source_book.stem}_backup_before_final_material_dictionary_{timestamp}{source_book.suffix}")
    shutil.copy2(source_book, backup)

    dictionary_rows = load_dictionary(dictionary_path)
    wb = openpyxl.load_workbook(source_book)
    definition_added = sync_definition_sheet(wb, dictionary_rows)
    sheet_status: dict[str, dict[str, int]] = {}
    for sheet_name, record_fields in {
        "Product_Lines": ["Record_ID"],
        "Product_Master": ["seed_record_id"],
        "Product_SKU_Master": ["seed_record_id"],
    }.items():
        if sheet_name not in wb.sheetnames:
            continue
        _changed, counter = normalize_sheet_taxonomy(wb[sheet_name], record_fields)
        sheet_status[sheet_name] = dict(counter)
    wb.save(source_book)
    wb.close()

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    audit_path = AUDIT_DIR / "final_material_dictionary_sync_latest.json"
    payload = {
        "source_book": str(source_book),
        "dictionary": str(dictionary_path),
        "backup": str(backup),
        "dictionary_paths": len(dictionary_rows),
        "definition_rows_added": definition_added,
        "sheet_status": sheet_status,
    }
    audit_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    sys.stdout.buffer.write((json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8"))
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge reviewed material dictionary into source workbook.")
    parser.add_argument("--source-book", type=Path, default=DEFAULT_SOURCE_BOOK)
    parser.add_argument("--dictionary", type=Path, default=DEFAULT_DICTIONARY)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    raise SystemExit(run(args.source_book, args.dictionary))
