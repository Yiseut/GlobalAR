#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import csv
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
AUDIT_DIR = ROOT / "data" / "audits"
MANUAL_FACT_CSV = ROOT / "data" / "manual_product_fact_evidence.csv"

SEED_RECORD_ID = "REC_0080"
PRODUCT_ID = "prod_52b69899fa6b"
PRODUCT_FAMILY_ID = "pf_b4578605d281"
COMPANY_ID = "co_c79c80f0dc0e"


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(norm(part).lower() for part in parts if norm(part))
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12] if raw else "0" * 12
    return f"{prefix}_{digest}"


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def set_cell(ws, row_idx: int, colmap: dict[str, int], field: str, value: str, changes: list[dict[str, Any]]) -> None:
    if field not in colmap:
        return
    cell = ws.cell(row=row_idx, column=colmap[field])
    old = norm(cell.value)
    if old == value:
        return
    cell.value = value
    changes.append({"sheet": ws.title, "row": row_idx, "field": field, "old": old, "new": value})


def update_product_master(wb, changes: list[dict[str, Any]]) -> None:
    ws = wb["Product_Master"]
    colmap = headers(ws)
    for row_idx in range(2, ws.max_row + 1):
        if norm(ws.cell(row=row_idx, column=colmap["seed_record_id"]).value) != SEED_RECORD_ID:
            continue
        set_cell(ws, row_idx, colmap, "verification_status", "official_commercial_fact_corrected", changes)
        set_cell(ws, row_idx, colmap, "review_status", "reviewed_nonregulated_skincare", changes)
        set_cell(ws, row_idx, colmap, "source_status", "official_seller_storefront_cross_checked", changes)
        set_cell(ws, row_idx, colmap, "r_and_d_origin_status", "commercial_identity_verified", changes)
        set_cell(
            ws,
            row_idx,
            colmap,
            "claim_text",
            "GoBizKorea/EC21 storefronts identify NEXGEN Biotechnologies as a skincare/recombinant-protein supplier and list Spider Cream / Spider Toxin style anti-wrinkle skincare products. Treat as non-regulated cosmeceutical identity evidence only; not an injectable botulinum toxin product.",
            changes,
        )
        return
    raise RuntimeError(f"Product_Master row not found for {SEED_RECORD_ID}")


def append_manual_facts(wb, changes: list[dict[str, Any]]) -> int:
    ws = wb["Manual_Product_Fact_Evidence"]
    colmap = headers(ws)
    existing = {
        norm(ws.cell(row=row_idx, column=colmap["fact_id"]).value)
        for row_idx in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row_idx, column=colmap["fact_id"]).value)
    }
    captured_at = datetime.now().astimezone().isoformat(timespec="seconds")
    facts = [
        {
            "fact_group": "official_product_page",
            "field_name": "official_product_page",
            "field_value": "https://dasisj.gobizkorea.com/mini/site/productList.do",
            "source_url": "https://dasisj.gobizkorea.com/mini/site/productList.do",
            "evidence_title": "All products of NEXGEN Biotechnologies, Inc. - GoBizKorea seller store",
            "evidence_excerpt": "Seller-store product list includes Preservative Free Cream - Dry Defender Spider Cream and Anti-Wrinkle Cream with Spider Toxin & Fiber Proteins - SPIDER CREAM.",
            "source_type": "official_product_page",
            "confidence": "official_seller_storefront_cross_checked",
            "note": "Commercial skincare identity evidence only; no medical-device registration or botulinum-toxin injectable claim promoted.",
        },
        {
            "fact_group": "official_specification_candidate",
            "field_name": "product_positioning",
            "field_value": "Non-regulated anti-wrinkle skincare/cosmeceutical using spider toxin/fiber protein wording; not an injectable botulinum toxin product.",
            "source_url": "https://nexgenbio.en.ec21.com/",
            "evidence_title": "NEXGEN Biotechnologies, Inc. - EC21 storefront",
            "evidence_excerpt": "Company storefront lists cosmetics, skin care, recombinant protein and Anti-Wrinkle Cream with Spider Toxin & Fiber Proteins SPIDER CREAM among main products.",
            "source_type": "official_product_page",
            "confidence": "official_seller_storefront_cross_checked",
            "note": "Use for product identity/spec positioning only; EC21 itself warns it does not guarantee seller credentials.",
        },
    ]
    added = 0
    for fact in facts:
        fact_id = stable_id("pfact", SEED_RECORD_ID, fact["fact_group"], fact["field_name"], fact["source_url"])
        if fact_id in existing:
            continue
        row_idx = ws.max_row + 1
        row = {
            "fact_id": fact_id,
            "product_id": PRODUCT_ID,
            "seed_record_id": SEED_RECORD_ID,
            "company_id": COMPANY_ID,
            "company": "Nexgen Bio",
            "brand": "Biometics",
            "product_family_id": PRODUCT_FAMILY_ID,
            "standard_product_name": "Spider Cream / Toxin",
            "priority": "P2-close",
            "captured_at": captured_at,
            "promoted_at": captured_at,
            "review_status": "manual_verified_nonregulated_skincare",
            **fact,
        }
        for field, value in row.items():
            if field in colmap:
                ws.cell(row=row_idx, column=colmap[field]).value = value
        changes.append({"sheet": ws.title, "row": row_idx, "field": "fact_id", "old": "", "new": fact_id})
        existing.add(fact_id)
        added += 1
    return added


def fact_payloads(captured_at: str) -> list[dict[str, str]]:
    facts = [
        {
            "fact_group": "official_product_page",
            "field_name": "official_product_page",
            "field_value": "https://dasisj.gobizkorea.com/mini/site/productList.do",
            "source_url": "https://dasisj.gobizkorea.com/mini/site/productList.do",
            "evidence_title": "All products of NEXGEN Biotechnologies, Inc. - GoBizKorea seller store",
            "evidence_excerpt": "Seller-store product list includes Preservative Free Cream - Dry Defender Spider Cream and Anti-Wrinkle Cream with Spider Toxin & Fiber Proteins - SPIDER CREAM.",
            "source_type": "official_product_page",
            "confidence": "official_seller_storefront_cross_checked",
            "note": "Commercial skincare identity evidence only; no medical-device registration or botulinum-toxin injectable claim promoted.",
        },
        {
            "fact_group": "official_specification_candidate",
            "field_name": "product_positioning",
            "field_value": "Non-regulated anti-wrinkle skincare/cosmeceutical using spider toxin/fiber protein wording; not an injectable botulinum toxin product.",
            "source_url": "https://nexgenbio.en.ec21.com/",
            "evidence_title": "NEXGEN Biotechnologies, Inc. - EC21 storefront",
            "evidence_excerpt": "Company storefront lists cosmetics, skin care, recombinant protein and Anti-Wrinkle Cream with Spider Toxin & Fiber Proteins SPIDER CREAM among main products.",
            "source_type": "official_product_page",
            "confidence": "official_seller_storefront_cross_checked",
            "note": "Use for product identity/spec positioning only; EC21 itself warns it does not guarantee seller credentials.",
        },
    ]
    output: list[dict[str, str]] = []
    for fact in facts:
        output.append(
            {
                "fact_id": stable_id("pfact", SEED_RECORD_ID, fact["fact_group"], fact["field_name"], fact["source_url"]),
                "product_id": PRODUCT_ID,
                "seed_record_id": SEED_RECORD_ID,
                "company_id": COMPANY_ID,
                "company": "Nexgen Bio",
                "brand": "Biometics",
                "product_family_id": PRODUCT_FAMILY_ID,
                "standard_product_name": "Spider Cream / Toxin",
                "priority": "P2-close",
                "captured_at": captured_at,
                "promoted_at": captured_at,
                "review_status": "manual_verified_nonregulated_skincare",
                **fact,
            }
        )
    return output


def append_manual_fact_csv() -> int:
    if not MANUAL_FACT_CSV.exists():
        raise FileNotFoundError(MANUAL_FACT_CSV)
    with MANUAL_FACT_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    if not fieldnames:
        raise RuntimeError(f"No header found in {MANUAL_FACT_CSV}")
    existing = {norm(row.get("fact_id")) for row in rows if norm(row.get("fact_id"))}
    captured_at = datetime.now().astimezone().isoformat(timespec="seconds")
    additions = [row for row in fact_payloads(captured_at) if row["fact_id"] not in existing]
    if not additions:
        return 0
    backup = AUDIT_DIR / f"manual_product_fact_evidence_backup_before_nexgen_p2_close_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    shutil.copy2(MANUAL_FACT_CSV, backup)
    with MANUAL_FACT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        writer.writerows(additions)
    return len(additions)


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    backup = AUDIT_DIR / f"source_workbook_backup_before_nexgen_p2_close_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    changes: list[dict[str, Any]] = []
    update_product_master(wb, changes)
    added = append_manual_facts(wb, changes)
    wb.save(SOURCE_BOOK)
    csv_added = append_manual_fact_csv()
    print({"backup": str(backup), "changes": len(changes), "manual_facts_added": added, "manual_fact_csv_added": csv_added})


if __name__ == "__main__":
    main()
