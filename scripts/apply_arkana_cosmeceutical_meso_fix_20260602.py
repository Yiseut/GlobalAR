from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TZ = timezone(timedelta(hours=8))
RUN_TS = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = REPO_ROOT.parent
WORKBOOK_PATH = DATA_ROOT / "全球医美企业库_标准化版v4.xlsx"
AUDIT_DIR = REPO_ROOT / "data" / "audits"

CORRECTION_TAG = "arkana_cosmeceutical_meso_fix_20260602"

FAMILY_UPDATES = {
    "REC_0258": {
        "product_id": "prod_62cba0e9d6a5",
        "old_family_id": "pf_4957194648b8",
        "company": "Arkana Cosmetics",
        "brand": "Exo Complex",
        "product_family": "外泌体再生疗法",
        "category_l1": "Skincare",
        "category_l2": "Cosmeceutical",
        "tech_type": "Exosome",
    },
    "REC_0571": {
        "product_id": "prod_ba776e888b9e",
        "old_family_id": "pf_b1625795ed59",
        "company": "Arkana Cosmetics",
        "brand": "PRP-like Therapy",
        "product_family": "仿生肽与微针",
        "category_l1": "Injectables",
        "category_l2": "Mesotherapy",
        "tech_type": "Biomimetic Peptide Mesotherapy",
    },
}

PRODUCT_LINE_UPDATES = {
    "REC_0060": {
        "Verified_Product_Type_CN": "院线功效护肤 / 复合酸治疗",
        "Market_Channel": "professional cosmeceutical / beauty salon",
        "Backfill_Audit": "Arkana skincare line display normalized to Cosmeceutical/professional functional skincare.",
    },
    "REC_0490": {
        "Verified_Product_Type_CN": "院线功效护肤 / 神经美容酸焕肤",
        "Market_Channel": "professional cosmeceutical / beauty salon",
        "Backfill_Audit": "Arkana peel line display normalized to Cosmeceutical/professional functional skincare.",
    },
    "REC_0258": {
        "Category_L1": "Skincare",
        "Category_L2": "Cosmeceutical",
        "Tech_Type_Std": "Exosome",
        "Verified_Product_Type_CN": "院线功效护肤 / 外泌体再生护理",
        "Market_Channel": "professional cosmeceutical / beauty salon / regenerative skincare",
        "Material_Taxonomy_L1_CN": "功效性护肤品",
        "Material_Taxonomy_L2_CN": "医学护肤活性",
        "Material_Taxonomy_L3_CN": "功效活性成分",
        "Material_Taxonomy_Path_CN": "功效性护肤品 > 医学护肤活性 > 功效活性成分",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": (
            "Arkana Exo Complex is a professional dermocosmetic/exosome skincare line; "
            "kept Exosome as technology but normalized commercial classification to Cosmeceutical."
        ),
        "Material_Family": "外泌体 Exosome",
        "Backfill_Audit": "Arkana Exosome line moved from Regenerative/Exosome display bucket to Skincare/Cosmeceutical; Exosome retained as technology.",
    },
    "REC_0571": {
        "Category_L1": "Injectables",
        "Category_L2": "Mesotherapy",
        "Tech_Type_Std": "Biomimetic Peptide Mesotherapy",
        "Tech_Type_Original": "PRP-like peptide microneedling / no-needle mesotherapy; product-led, not RF or energy device.",
        "Verified_Product_Type_CN": "仿生肽美素 / 微针导入疗法",
        "Market_Channel": "professional cosmeceutical / beauty salon / mesotherapy",
        "Material_Taxonomy_L1_CN": "注射类",
        "Material_Taxonomy_L2_CN": "美塑成分",
        "Material_Taxonomy_L3_CN": "生长因子/多肽鸡尾酒",
        "Material_Taxonomy_Path_CN": "注射类 > 美塑成分 > 生长因子/多肽鸡尾酒",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": (
            "PRP-like Arkana line uses W3 Peptide/GHK-Cu peptide complex with microneedling/no-needle mesotherapy delivery; "
            "the needle is a consumable delivery route, not radiofrequency or other energy-based device, and it is not autologous PRP."
        ),
        "Material_Family": "生长因子/多肽鸡尾酒",
        "Backfill_Audit": "PRP-like Therapy corrected from EBD/Radiofrequency to Injectables/Mesotherapy; product-led peptide mesotherapy, not an energy device.",
    },
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(norm(part).lower() for part in parts if norm(part))
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12] if raw else "0" * 12
    return f"{prefix}_{digest}"


for item in FAMILY_UPDATES.values():
    item["new_family_id"] = stable_id(
        "pf",
        item["company"],
        item["brand"],
        item["product_family"],
        item["category_l1"],
        item["category_l2"],
        item["tech_type"],
    )


def backup_file(path: Path) -> Path:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".xlsx":
        backup = path.with_name(f"{path.stem}.backup_before_{CORRECTION_TAG}_{RUN_TS}{path.suffix}")
    else:
        backup = AUDIT_DIR / f"{path.stem}.backup_before_{CORRECTION_TAG}_{RUN_TS}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def sheet_header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = norm(cell.value)
        if key:
            headers[key] = idx
    return headers


def append_audit(existing: Any, message: str) -> str:
    before = norm(existing)
    entry = f"{CORRECTION_TAG}: {message}"
    if entry in before:
        return before
    return " | ".join(part for part in [before, entry] if part)


def update_workbook() -> dict[str, Any]:
    backup = backup_file(WORKBOOK_PATH)
    wb = load_workbook(WORKBOOK_PATH)
    if "Product_Lines" not in wb.sheetnames:
        raise RuntimeError("Workbook has no Product_Lines sheet")
    ws = wb["Product_Lines"]
    headers = sheet_header_map(ws)
    if "Record_ID" not in headers:
        raise RuntimeError("Product_Lines has no Record_ID column")

    changed: dict[str, dict[str, dict[str, str]]] = {}
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        record_id = norm(row[headers["Record_ID"] - 1].value)
        updates = PRODUCT_LINE_UPDATES.get(record_id)
        if not updates:
            continue
        seen.add(record_id)
        row_changes: dict[str, dict[str, str]] = {}
        for field, value in updates.items():
            if field not in headers:
                raise KeyError(f"Missing Product_Lines column: {field}")
            cell = row[headers[field] - 1]
            before = "" if cell.value is None else str(cell.value)
            if field == "Backfill_Audit":
                after = append_audit(before, value)
            else:
                after = value
            cell.value = after
            after_text = "" if after is None else str(after)
            if before != after_text:
                row_changes[field] = {"before": before, "after": after_text}
        if row_changes:
            changed[record_id] = row_changes

    missing = sorted(set(PRODUCT_LINE_UPDATES) - seen)
    if missing:
        raise RuntimeError(f"Missing target rows in Product_Lines: {', '.join(missing)}")

    wb.save(WORKBOOK_PATH)
    return {
        "path": str(WORKBOOK_PATH),
        "backup": str(backup),
        "changed_records": changed,
    }


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise RuntimeError(f"{path} has no header")
        return list(reader.fieldnames), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def matching_update(row: dict[str, str]) -> dict[str, str] | None:
    company = norm(row.get("company"))
    brand = norm(row.get("brand"))
    for record_id, item in FAMILY_UPDATES.items():
        if company != item["company"] or brand != item["brand"]:
            continue
        if (
            norm(row.get("product_family_id")) in {item["old_family_id"], item["new_family_id"]}
            or norm(row.get("product_id")) == item["product_id"]
            or norm(row.get("seed_record_id")) == record_id
        ):
            return item
    return None


def official_source_query(row: dict[str, str], item: dict[str, str]) -> str:
    base = (
        f"\"{item['company']}\" \"{item['brand']}\" {item['product_family']} "
        f"{item['category_l1']} {item['category_l2']} {item['tech_type']}"
    )
    query_type = norm(row.get("query_type"))
    if query_type == "product_ifu_labeling":
        return f"{base} IFU instructions for use intended purpose official PDF"
    if query_type == "product_certificate_registration":
        return f"{base} certificate declaration of conformity registration official"
    return f"{base} official product page professional aesthetic source"


def mdr_ce_query(row: dict[str, str], item: dict[str, str]) -> str:
    base = f"{item['company']} {item['brand']} {item['product_family']}"
    source_key = norm(row.get("source_key"))
    evidence_target = norm(row.get("evidence_target"))
    if source_key == "eu_eudamed" or "EUDAMED" in evidence_target:
        return f"{base} EUDAMED Basic UDI-DI SRN certificate device"
    if source_key == "ce_notified_body_certificate" or "Notified-body" in evidence_target:
        return f"{base} CE MDR certificate notified body declaration conformity scope"
    if source_key == "company_ce_documents" or "IFU" in evidence_target:
        return f"{base} IFU instructions for use intended purpose declaration of conformity official PDF"
    return f"{base} CE MDR official evidence"


def update_classification_plan_row(row: dict[str, str], query_builder) -> bool:
    item = matching_update(row)
    if not item:
        return False
    before = dict(row)
    row["product_family_id"] = item["new_family_id"]
    row["category_l1"] = item["category_l1"]
    row["category_l2"] = item["category_l2"]
    row["tech_type"] = item["tech_type"]
    row["query"] = query_builder(row, item)
    return row != before


def update_family_only(row: dict[str, str]) -> bool:
    item = matching_update(row)
    if not item:
        return False
    before = row.get("product_family_id", "")
    row["product_family_id"] = item["new_family_id"]
    return row.get("product_family_id", "") != before


def update_csv(path: Path, updater) -> dict[str, Any]:
    fieldnames, rows = read_csv(path)
    changed = 0
    for row in rows:
        if updater(row):
            changed += 1
    backup = ""
    if changed:
        backup = str(backup_file(path))
        write_csv(path, fieldnames, rows)
    return {"path": str(path), "backup": backup, "changed_rows": changed}


def main() -> None:
    expected_ids = {
        "REC_0258": "pf_e48c8f7b8d7c",
        "REC_0571": "pf_93756b2d7567",
    }
    for record_id, expected in expected_ids.items():
        actual = FAMILY_UPDATES[record_id]["new_family_id"]
        if actual != expected:
            raise RuntimeError(f"Unexpected family id for {record_id}: {actual}")

    workbook_result = update_workbook()
    csv_results = [
        update_csv(
            REPO_ROOT / "data" / "company_official_source_plan.csv",
            lambda row: update_classification_plan_row(row, official_source_query),
        ),
        update_csv(
            REPO_ROOT / "data" / "mdr_ce_search_plan.csv",
            lambda row: update_classification_plan_row(row, mdr_ce_query),
        ),
        update_csv(REPO_ROOT / "data" / "manual_product_fact_evidence.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "manual_evidence_promotion_log.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "product_specification_evidence.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "company_media_asset_index.csv", update_family_only),
    ]

    audit = {
        "run_ts": RUN_TS,
        "correction": CORRECTION_TAG,
        "records": {
            record_id: {
                "product_id": item["product_id"],
                "old_family_id": item["old_family_id"],
                "new_family_id": item["new_family_id"],
                "category_l1": item["category_l1"],
                "category_l2": item["category_l2"],
                "tech_type": item["tech_type"],
            }
            for record_id, item in FAMILY_UPDATES.items()
        },
        "workbook": workbook_result,
        "csv_updates": csv_results,
    }
    audit_path = AUDIT_DIR / f"{CORRECTION_TAG}_{RUN_TS}.json"
    latest_path = AUDIT_DIR / f"{CORRECTION_TAG}_latest.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    latest_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
