from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"

NEXT_QUEUE = AUDIT_DIR / "next_user_confirmation_batch_latest.csv"
NEXT_QUEUE_FALLBACKS = [
    AUDIT_DIR / "next_user_confirmation_batch_20260601_040653.csv",
    AUDIT_DIR / "next_user_confirmation_batch_20260601_033852.csv",
    AUDIT_DIR / "next_user_confirmation_batch_20260601_033508.csv",
    AUDIT_DIR / "next_user_confirmation_batch_20260601_032747.csv",
]
PRODUCT_MASTER = DATA_DIR / "product_master.csv"
MANUAL_INDICATION = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_FACT = DATA_DIR / "manual_product_fact_evidence.csv"


KEEP_OFFICIAL = {
    "NEXT-001",
    "NEXT-002",
    "NEXT-003",
    "NEXT-004",
    "NEXT-005",
    "NEXT-006",
    "NEXT-007",
    "NEXT-008",
    "NEXT-022",
    "NEXT-023",
    "NEXT-035",
    "NEXT-036",
    "NEXT-037",
    "NEXT-038",
}

KEEP_KOREA = {
    "NEXT-009",
    "NEXT-010",
    "NEXT-011",
    "NEXT-012",
    "NEXT-013",
    "NEXT-015",
}

# These rows are duplicate descriptions or accessories. Rows whose target is
# themselves are kept as the canonical line and normalized by adding evidence.
MERGE_DECISIONS = {
    "NEXT-016": {
        "target_record_id": "REC_0500",
        "target_label": "NOVAPlus needle-free injector",
        "action": "delete_duplicate",
        "note": "NOVAPlus Nozzle is an accessory/nozzle kit for the NOVAPlus injector.",
    },
    "NEXT-017": {
        "target_record_id": "REC_0733",
        "target_label": "BTL Vanquish ME",
        "action": "keep_canonical",
        "note": "Use this row as the canonical Vanquish ME line and fold regional wording into it.",
    },
    "NEXT-018": {
        "target_record_id": "REC_0733",
        "target_label": "BTL Vanquish ME",
        "action": "delete_duplicate",
        "note": "Non-Contact Selective RF is another description of BTL Vanquish ME.",
    },
    "NEXT-024": {
        "target_record_id": "REC_0724",
        "target_label": "BTL Unison",
        "action": "keep_canonical",
        "note": "Use this row as the canonical BTL Unison line.",
    },
}

EXCLUDE_DECISIONS = {
    "NEXT-014": "Kysense Body HA filler is not retained as an independent upstream product line in this review batch.",
    "NEXT-019": "Mosquito Needle / carbon nanotube needle is excluded by user decision in this review batch.",
    "NEXT-020": "E-Finger needle-free transdermal delivery system is excluded by user decision in this review batch.",
    "NEXT-021": "Universal IPS surgical implants are generic orthopedic/dental/surgical implant scope, not aesthetic-dedicated upstream products.",
    "NEXT-025": "Rejulight LED device is excluded by user decision in this review batch.",
    "NEXT-026": "Dalyance RF + ultrasound device is excluded by user decision in this review batch.",
    "NEXT-027": "Epil One diode hair-removal device is excluded by user decision in this review batch.",
    "NEXT-028": "Safyre RF device is excluded by user decision in this review batch.",
    "NEXT-029": "Daejoo Medical Cosmetics ODM is broad ODM/service capability without a concrete independent terminal product line.",
    "NEXT-030": "Dongbang acupuncture needles are TCM/rehabilitation consumables, not aesthetic microneedling products.",
    "NEXT-031": "Doum fat transfer cannulas are hard surgical/liposuction instruments outside the current light-aesthetic upstream scope.",
    "NEXT-032": "Doum infiltration pump is general liposuction surgery hardware outside the current scope.",
    "NEXT-033": "Doum PAL liposuction system is large surgical hardware outside the current scope.",
    "NEXT-034": "MicroGlide is excluded by user decision in this review batch and is not retained as an injectable filler/product line.",
}

CANONICAL_MERGE_KEEP = {"NEXT-017", "NEXT-024"}


INDICATIONS = {
    "NEXT-001": "Fractional CO2 laser platform for ablative skin resurfacing, wrinkles, acne scars, texture improvement and professional skin rejuvenation.",
    "NEXT-002": "808 nm diode laser platform for professional hair reduction/removal across treatment areas.",
    "NEXT-003": "Q-switched laser platform for pigment, tattoo, toning and skin-rejuvenation applications.",
    "NEXT-004": "Professional oxygen-infusion and skin-management treatment for hydration, cleansing and facial rejuvenation.",
    "NEXT-005": "Cryolipolysis/cold body-contouring device for non-invasive localized fat reduction and body shaping.",
    "NEXT-006": "Hydrodermabrasion and oxygen-infusion skin-management system for cleansing, hydration and medspa facial rejuvenation.",
    "NEXT-007": "HIFU/LIPUS multi-function platform for lifting, skin tightening and non-invasive body/face contouring applications.",
    "NEXT-008": "RF/LIPUS professional device for skin tightening, contouring and body-shaping applications.",
    "NEXT-009": "PLLA injectable biostimulator/filler retained as a Korean regenerative-aesthetic product line; used for collagen stimulation, wrinkles and volume restoration.",
    "NEXT-010": "PDLLA/HA collagen stimulator retained as a Korean regenerative injectable product line for collagen stimulation and tissue-volume support.",
    "NEXT-011": "PLLA biostimulator retained as a Korean regenerative injectable line for collagen stimulation and skin/volume improvement.",
    "NEXT-012": "PLLA skin-booster/biostimulator retained as a Korean regenerative injectable line for skin quality and collagen-stimulation applications.",
    "NEXT-013": "PLLA collagen-stimulator retained as a Korean regenerative injectable line for collagen stimulation, wrinkle correction and volume support.",
    "NEXT-015": "Deoxycholic-acid/cholic-acid lipolytic injectable product line retained for localized fat reduction and contouring applications.",
    "NEXT-017": "BTL Vanquish ME is a non-contact selective RF body-contouring device for circumferential reduction and localized fat reduction.",
    "NEXT-019": "Carbon-nanotube micro-needle consumable retained as a Korean minimally invasive delivery/needle technology line.",
    "NEXT-020": "Needle-free electroporation/transdermal delivery system for professional mesotherapy active delivery and skin-rejuvenation workflows.",
    "NEXT-022": "Meso-Eye C71 is an eye-area mesotherapy product for periorbital concerns such as dark circles, puffiness, microcirculation and fine lines.",
    "NEXT-023": "Meso-Genesis BP3 is a professional hair-restoration mesotherapy product for scalp and follicle revitalization workflows.",
    "NEXT-024": "BTL Unison combines RF and acoustic/shockwave-style energy for cellulite, skin tightening and body-contouring workflows.",
    "NEXT-025": "Professional LED phototherapy device for post-procedure repair, skin rejuvenation, anti-aging and acne-oriented light therapy.",
    "NEXT-026": "RF and ultrasound professional aesthetic device for skin tightening, face/body contouring and tissue-rejuvenation workflows.",
    "NEXT-027": "Professional diode laser hair-removal platform for long-term hair reduction.",
    "NEXT-028": "RF professional device for skin tightening, cellulite/body contouring and aesthetic tissue remodeling.",
    "NEXT-035": "Q-switched Nd:YAG laser platform for tattoo removal, pigmented-lesion treatment and skin toning/rejuvenation applications.",
    "NEXT-036": "Professional aesthetic dermal/mesocare product line for clinic-led skin quality, peels and professional skin-treatment workflows.",
    "NEXT-037": "Professional intimate-care aesthetic product line for intimate rejuvenation, skin quality and pigmentation-oriented workflows.",
    "NEXT-038": "Professional post-procedure/transdermal cocktail product line for skin repair, regeneration support and procedure-adjunct care.",
}


def clean(value: object) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(clean(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_fields: list[str], new_rows: list[dict[str, str]]) -> int:
    existing = {tuple(clean(row.get(field)) for field in key_fields) for row in rows}
    added = 0
    for row in new_rows:
        key = tuple(clean(row.get(field)) for field in key_fields)
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def load_next_queue() -> dict[str, dict[str, str]]:
    for path in [NEXT_QUEUE, *NEXT_QUEUE_FALLBACKS]:
        if not path.exists():
            continue
        _, rows = read_csv(path)
        indexed = {clean(row.get("confirm_id")): row for row in rows if clean(row.get("confirm_id"))}
        if indexed:
            return indexed
    return {}


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER)
    return {clean(row.get("seed_record_id")): row for row in rows if clean(row.get("seed_record_id"))}


def headers(ws) -> dict[str, int]:
    return {clean(cell.value): cell.column for cell in ws[1] if clean(cell.value)}


def append_note(existing: str, marker: str) -> str:
    existing = clean(existing)
    if marker in existing:
        return existing
    return f"{existing}; {marker}".strip("; ")


def source_url_for(next_row: dict[str, str]) -> str:
    return clean(next_row.get("lead_url"))


def indication_row(product: dict[str, str], next_row: dict[str, str], checked_at: str, mode: str) -> dict[str, str]:
    confirm_id = clean(next_row.get("confirm_id"))
    indication = INDICATIONS[confirm_id]
    is_korea = mode == "korea"
    is_skintech = clean(product.get("commercial_path_l1")) == "Skincare"
    source_url = source_url_for(next_row)
    if is_korea:
        jurisdiction = "KR"
        regulator = "KFDA/MFDS"
        pathway = "medical device or product registration follow-up"
        status = "User confirmed retain; KFDA/MFDS lookup required. Public certificate number not captured; mark as not publicly available if no official number is found."
        review_status = "user_confirmed_mfds_pending_public_number"
        confidence = "user_confirmed_keep_mfds_pending"
    else:
        jurisdiction = "US / EU"
        regulator = "FDA / CE-MDR"
        pathway = "FDA/CE follow-up"
        status = "User confirmed retain; FDA/CE evidence to be supplemented when public official records are available."
        review_status = "user_confirmed_keep_fda_ce_pending"
        confidence = "user_confirmed_keep_fda_ce_followup"
    if confirm_id == "NEXT-035":
        jurisdiction = "US"
        regulator = "FDA"
        pathway = "FDA public device record / 510(k) lead"
        status = "FDA public device lead retained; exact registration fields pending source parsing."
        review_status = "user_confirmed_fda_lead"
        confidence = "official_fda_lead_user_confirmed"

    registered_name = clean(next_row.get("brand")) or clean(product.get("brand")) or clean(next_row.get("original_product_name"))
    return {
        "product_id": clean(product.get("product_id")),
        "seed_record_id": clean(product.get("seed_record_id")),
        "company_id": clean(product.get("company_id")),
        "company": clean(product.get("company")),
        "brand": clean(product.get("brand")),
        "jurisdiction": jurisdiction,
        "regulator": regulator,
        "regulatory_pathway": pathway,
        "status": status,
        "registration_no": "",
        "approval_date": "",
        "expiry_date": "",
        "registered_name": registered_name,
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": clean(product.get("legal_manufacturer")) or clean(product.get("company")),
        "local_holder": "",
        "source_key": stable_id("user_next_20260601", confirm_id, product.get("seed_record_id"), jurisdiction, regulator),
        "source_url": source_url,
        "source_type": "user_confirmed_next_batch_regulatory_status",
        "evidence_title": f"{registered_name} user-confirmed product line status",
        "evidence_excerpt": f"{confirm_id}: user confirmed retain and requested official indication/regulatory completion. {status}",
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": "Exact certificate number remains blank unless supplied by official regulator/IFU/certificate source.",
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601_next_batch",
        "review_status": review_status,
        "confidence": confidence,
    }


def fact_rows(product: dict[str, str], next_row: dict[str, str], checked_at: str) -> list[dict[str, str]]:
    confirm_id = clean(next_row.get("confirm_id"))
    indication = INDICATIONS[confirm_id]
    source_url = source_url_for(next_row)
    rows: list[dict[str, str]] = []
    if source_url:
        rows.append(
            {
                "fact_id": stable_id("pfact", confirm_id, product.get("product_id"), "official_product_page", source_url),
                "product_id": clean(product.get("product_id")),
                "seed_record_id": clean(product.get("seed_record_id")),
                "company_id": clean(product.get("company_id")),
                "company": clean(product.get("company")),
                "brand": clean(product.get("brand")),
                "product_family_id": "",
                "standard_product_name": clean(product.get("standard_product_name")),
                "priority": clean(next_row.get("priority")) or "P1",
                "fact_group": "official_product_page",
                "field_name": "official_product_page",
                "field_value": source_url,
                "source_url": source_url,
                "evidence_title": f"{clean(product.get('brand'))} official/product lead",
                "evidence_excerpt": f"{confirm_id}: user accepted this source lead for product-line retention and follow-up.",
                "source_type": "official_product_page",
                "confidence": "official_source_lead_user_confirmed",
                "captured_at": checked_at,
                "promoted_at": checked_at,
                "review_status": "user_confirmed",
                "note": "next_user_confirmation_batch_20260601",
            }
        )
    rows.append(
        {
            "fact_id": stable_id("pfact", confirm_id, product.get("product_id"), "official_indication_or_positioning", indication),
            "product_id": clean(product.get("product_id")),
            "seed_record_id": clean(product.get("seed_record_id")),
            "company_id": clean(product.get("company_id")),
            "company": clean(product.get("company")),
            "brand": clean(product.get("brand")),
            "product_family_id": "",
            "standard_product_name": clean(product.get("standard_product_name")),
            "priority": clean(next_row.get("priority")) or "P1",
            "fact_group": "official_specification_candidate",
            "field_name": "official_indication_or_positioning",
            "field_value": indication,
            "source_url": source_url,
            "evidence_title": f"{clean(product.get('brand'))} user-confirmed indication/positioning",
            "evidence_excerpt": f"{confirm_id}: {indication}",
            "source_type": "user_confirmed_product_positioning",
            "confidence": "user_confirmed",
            "captured_at": checked_at,
            "promoted_at": checked_at,
            "review_status": "user_confirmed",
            "note": "next_user_confirmation_batch_20260601",
        }
    )
    return rows


def update_workbook(next_rows: dict[str, dict[str, str]], products: dict[str, dict[str, str]], stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_next_confirmation_batch_20260601_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)

    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    col = headers(ws)
    row_by_id = {clean(ws.cell(row=i, column=col["Record_ID"]).value): i for i in range(2, ws.max_row + 1)}
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: str) -> None:
        row_idx = row_by_id.get(record_id)
        col_idx = col.get(field)
        if not row_idx or not col_idx:
            return
        old = clean(ws.cell(row=row_idx, column=col_idx).value)
        new = clean(value)
        if old == new:
            return
        ws.cell(row=row_idx, column=col_idx, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row_idx = row_by_id.get(record_id)
        col_idx = col.get("Backfill_Audit")
        if not row_idx or not col_idx:
            return
        old = clean(ws.cell(row=row_idx, column=col_idx).value)
        new = append_note(old, note)
        if old == new:
            return
        ws.cell(row=row_idx, column=col_idx, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    def keep_record(confirm_id: str, mode: str) -> None:
        row = next_rows[confirm_id]
        record_id = clean(row.get("seed_record_id"))
        status = (
            "KFDA/MFDS retained; public certificate number pending/not public"
            if mode == "korea"
            else "User confirmed retained; FDA/CE evidence pending/not public"
        )
        set_cell(record_id, "Inclusion_Status", "active")
        set_cell(record_id, "Is_Primary_Record", "True")
        set_cell(record_id, "V4_1_Registered_Name", clean(row.get("brand")) or clean(row.get("original_product_name")))
        set_cell(record_id, "V4_1_Registration_Review_Status", status)
        note = (
            f"{confirm_id}: user confirmed retain; KFDA/MFDS or exact certificate number remains blank unless public official evidence is found."
            if mode == "korea"
            else f"{confirm_id}: user confirmed retain; FDA/CE or exact certificate number remains blank unless public official evidence is found."
        )
        set_cell(record_id, "V4_1_Registration_Note", note)
        set_cell(record_id, "Verified_Product_Type_CN", "用户确认保留的医美上游产品线")
        set_cell(record_id, "Market_Channel", "医美机构/专业渠道")
        if mode == "korea":
            set_cell(record_id, "KFDA_Status", "KFDA/MFDS to verify; public number not captured")
        append_audit(record_id, f"next_confirmation_20260601:{confirm_id}: confirmed retain; supplement indication/regulatory status.")

    for confirm_id in sorted(KEEP_OFFICIAL):
        keep_record(confirm_id, "official")
    for confirm_id in sorted(KEEP_KOREA):
        keep_record(confirm_id, "korea")
    for confirm_id in sorted(CANONICAL_MERGE_KEEP):
        keep_record(confirm_id, "official")

    for confirm_id, decision in MERGE_DECISIONS.items():
        row = next_rows[confirm_id]
        record_id = clean(row.get("seed_record_id"))
        target = decision["target_record_id"]
        if decision["action"] == "delete_duplicate":
            set_cell(record_id, "Inclusion_Status", "deleted")
            set_cell(record_id, "Is_Primary_Record", "False")
            set_cell(record_id, "Duplicate_Note", f"duplicate_of:{target}; {decision['note']}")
            set_cell(record_id, "V4_1_Registration_Review_Status", "merged_to_existing_product")
            set_cell(record_id, "V4_1_Registration_Note", f"{confirm_id}: {decision['note']} Target={target} ({decision['target_label']}).")
            append_audit(record_id, f"next_confirmation_20260601:{confirm_id}: merged into {target}; not an independent product line.")
            append_audit(target, f"next_confirmation_20260601:{confirm_id}: absorbed duplicate/accessory row {record_id}.")
        else:
            set_cell(record_id, "Inclusion_Status", "active")
            set_cell(record_id, "Is_Primary_Record", "True")
            set_cell(record_id, "V4_1_Registration_Review_Status", "merged_to_existing_canonical_product_line")
            set_cell(record_id, "V4_1_Registration_Note", f"{confirm_id}: {decision['note']}")
            append_audit(record_id, f"next_confirmation_20260601:{confirm_id}: kept as canonical line; regional wording folded into this row.")

    for confirm_id, reason in EXCLUDE_DECISIONS.items():
        row = next_rows[confirm_id]
        record_id = clean(row.get("seed_record_id"))
        set_cell(record_id, "Inclusion_Status", "excluded")
        set_cell(record_id, "Is_Primary_Record", "False")
        set_cell(record_id, "Duplicate_Note", f"excluded_by_user_20260601:{confirm_id}; {reason}")
        set_cell(record_id, "V4_1_Registration_Review_Status", "excluded_scope")
        set_cell(record_id, "V4_1_Registration_Note", f"{confirm_id}: {reason}")
        append_audit(record_id, f"next_confirmation_20260601:{confirm_id}: excluded; {reason}")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def unlink_deleted_product_evidence(product_ids: dict[str, str], stamp: str) -> list[dict[str, object]]:
    files = [
        (MANUAL_FACT, "note", "review_status"),
        (DATA_DIR / "manual_evidence_promotion_log.csv", "note", None),
        (DATA_DIR / "product_specification_evidence.csv", "notes", "review_status"),
        (MANUAL_INDICATION, "field_note", "review_status"),
        (DATA_DIR / "manual_nmpa_registration_evidence.csv", "field_note", "review_status"),
    ]
    results: list[dict[str, object]] = []
    for path, note_field, status_field in files:
        if not path.exists():
            continue
        fields, rows = read_csv(path)
        if "product_id" not in fields:
            continue
        backup = path.with_name(f"{path.stem}_backup_before_next_confirmation_unlink_{stamp}{path.suffix}")
        shutil.copy2(path, backup)
        counts: Counter[str] = Counter()
        for row in rows:
            original = clean(row.get("product_id"))
            if original not in product_ids:
                continue
            marker = (
                "next_confirmation_unlink_20260601: "
                f"{product_ids[original]} Retained as company/source-history evidence only, not active product evidence."
            )
            row["product_id"] = ""
            if "product_family_id" in row:
                row["product_family_id"] = ""
            if status_field and status_field in row:
                row[status_field] = "excluded_scope_unlinked"
            if note_field in row:
                row[note_field] = append_note(row.get(note_field, ""), marker)
            counts[original] += 1
        write_csv(path, fields, rows)
        results.append(
            {
                "file": str(path),
                "backup": str(backup),
                "rows_unlinked": sum(counts.values()),
                "by_original_product_id": dict(counts),
            }
        )
    return results


def build_new_evidence_rows(next_rows: dict[str, dict[str, str]], products: dict[str, dict[str, str]], checked_at: str) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    indication_rows: list[dict[str, str]] = []
    fact_output: list[dict[str, str]] = []
    for confirm_id in sorted(KEEP_OFFICIAL | KEEP_KOREA | CANONICAL_MERGE_KEEP):
        next_row = next_rows[confirm_id]
        product = products.get(clean(next_row.get("seed_record_id")))
        if not product:
            continue
        mode = "korea" if confirm_id in KEEP_KOREA else "official"
        indication_rows.append(indication_row(product, next_row, checked_at, mode))
        fact_output.extend(fact_rows(product, next_row, checked_at))
    return indication_rows, fact_output


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    next_rows = load_next_queue()
    expected = KEEP_OFFICIAL | KEEP_KOREA | set(MERGE_DECISIONS) | set(EXCLUDE_DECISIONS)
    missing = sorted(confirm_id for confirm_id in expected if confirm_id not in next_rows)
    if missing:
        raise SystemExit(f"Missing NEXT ids in latest queue: {missing}")

    products = product_lookup()
    backup, workbook_changes = update_workbook(next_rows, products, stamp)

    indication_fields, indication_existing = read_csv(MANUAL_INDICATION)
    fact_fields, fact_existing = read_csv(MANUAL_FACT)
    new_indications, new_facts = build_new_evidence_rows(next_rows, products, checked_at)
    added_indications = append_unique(
        indication_existing,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        new_indications,
    )
    added_facts = append_unique(fact_existing, ["fact_id"], new_facts)
    write_csv(MANUAL_INDICATION, indication_fields, indication_existing)
    write_csv(MANUAL_FACT, fact_fields, fact_existing)

    deleted_record_ids = []
    deleted_record_ids.extend(clean(next_rows[cid].get("seed_record_id")) for cid, decision in MERGE_DECISIONS.items() if decision["action"] == "delete_duplicate")
    deleted_record_ids.extend(clean(next_rows[cid].get("seed_record_id")) for cid in EXCLUDE_DECISIONS)
    deleted_product_ids = {
        clean(products[record_id].get("product_id")): f"{record_id} {clean(products[record_id].get('company'))} / {clean(products[record_id].get('brand'))}"
        for record_id in deleted_record_ids
        if record_id in products and clean(products[record_id].get("product_id"))
    }
    unlink_results = unlink_deleted_product_evidence(deleted_product_ids, stamp)

    decisions_csv = AUDIT_DIR / "next_user_confirmation_decisions_20260601_latest.csv"
    with decisions_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = ["confirm_id", "seed_record_id", "company", "brand", "product", "decision", "target_record_id", "note"]
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for confirm_id in sorted(KEEP_OFFICIAL):
            row = next_rows[confirm_id]
            writer.writerow(
                {
                    "confirm_id": confirm_id,
                    "seed_record_id": row.get("seed_record_id", ""),
                    "company": row.get("company", ""),
                    "brand": row.get("brand", ""),
                    "product": row.get("original_product_name", ""),
                    "decision": "confirm_keep_supplement_official_indication_regulatory",
                    "target_record_id": row.get("seed_record_id", ""),
                    "note": INDICATIONS.get(confirm_id, ""),
                }
            )
        for confirm_id in sorted(KEEP_KOREA):
            row = next_rows[confirm_id]
            writer.writerow(
                {
                    "confirm_id": confirm_id,
                    "seed_record_id": row.get("seed_record_id", ""),
                    "company": row.get("company", ""),
                    "brand": row.get("brand", ""),
                    "product": row.get("original_product_name", ""),
                    "decision": "confirm_keep_supplement_korea_mfds_or_mark_not_public",
                    "target_record_id": row.get("seed_record_id", ""),
                    "note": INDICATIONS.get(confirm_id, ""),
                }
            )
        for confirm_id, decision in sorted(MERGE_DECISIONS.items()):
            row = next_rows[confirm_id]
            writer.writerow(
                {
                    "confirm_id": confirm_id,
                    "seed_record_id": row.get("seed_record_id", ""),
                    "company": row.get("company", ""),
                    "brand": row.get("brand", ""),
                    "product": row.get("original_product_name", ""),
                    "decision": decision["action"],
                    "target_record_id": decision["target_record_id"],
                    "note": decision["note"],
                }
            )
        for confirm_id, reason in sorted(EXCLUDE_DECISIONS.items()):
            row = next_rows[confirm_id]
            writer.writerow(
                {
                    "confirm_id": confirm_id,
                    "seed_record_id": row.get("seed_record_id", ""),
                    "company": row.get("company", ""),
                    "brand": row.get("brand", ""),
                    "product": row.get("original_product_name", ""),
                    "decision": "exclude_out_of_scope",
                    "target_record_id": "",
                    "note": reason,
                }
            )

    summary = {
        "generated_at": checked_at,
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_indications,
        "manual_product_fact_rows_added": added_facts,
        "keep_official": sorted(KEEP_OFFICIAL),
        "keep_korea": sorted(KEEP_KOREA),
        "merge_decisions": MERGE_DECISIONS,
        "exclude_decisions": EXCLUDE_DECISIONS,
        "deleted_product_ids_unlinked": deleted_product_ids,
        "unlink_results": unlink_results,
        "decisions_csv": str(decisions_csv),
        "changed_fields_sample": workbook_changes[:160],
    }
    out = AUDIT_DIR / f"next_user_confirmation_apply_20260601_{stamp}.json"
    latest = AUDIT_DIR / "next_user_confirmation_apply_20260601_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
