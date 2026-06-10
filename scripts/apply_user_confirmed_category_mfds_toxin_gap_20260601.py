"""Apply user-confirmed category, MFDS attribution, toxin, and P0 gap feedback."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_FACT_PATH = DATA_DIR / "manual_product_fact_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_fields: list[str], new_rows: list[dict[str, str]]) -> int:
    existing = {tuple(norm(row.get(field)) for field in key_fields) for row in rows}
    added = 0
    for row in new_rows:
        key = tuple(norm(row.get(field)) for field in key_fields)
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def tax(l1: str, l2: str, l3: str, family: str, status: str = "active") -> dict[str, str]:
    return {
        "Material_Taxonomy_L1_CN": l1,
        "Material_Taxonomy_L2_CN": l2,
        "Material_Taxonomy_L3_CN": l3,
        "Material_Taxonomy_Path_CN": " > ".join(x for x in [l1, l2, l3] if x),
        "Material_Taxonomy_Source": "user_feedback_20260601",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "user_confirmed",
        "Material_Taxonomy_Note": "用户确认分类/归属后写回；用于清理美素、肉毒、韩国 MFDS 队列中的错配。",
        "Material_Family": family,
        "Inclusion_Status": status,
    }


HA_FILLER = tax("透明质酸填充剂", "交联HA", "面部填充/轮廓塑形", "透明质酸")
HA_SKINBOOSTER = tax("皮肤动能素/水光", "非交联/微交联HA复配", "浅层微滴注射", "HA皮肤动能素")
PN_SKINBOOSTER = tax("皮肤动能素/水光", "PN/PDRN", "再生水光/组织修复", "PN/PDRN")
INJECTION_DEVICE = tax("器械/给药设备", "注射辅助/微通道", "自动注射或微通道给药", "给药设备")
CARBOXY_DEVICE = tax("器械/给药设备", "气体治疗/二氧化碳", "Carboxytherapy", "二氧化碳治疗设备")
SKIN_MANAGEMENT = tax("皮肤管理设备", "水氧/负压清洁导入", "HydraFacial/Vortex-Fusion", "皮肤管理设备")
HAIR_EBD = tax("能量源设备", "毛发复苏/透皮导入", "超声透皮给药", "毛发复苏设备")
SURGICAL_TOOL = tax("手术器械", "整形外科工具", "鼻整形/脂肪移植/剪刀套件", "整形外科手术器械")
LIPOLYTIC = tax("溶脂注射", "脱氧胆酸", "下颌下脂肪", "脱氧胆酸")
NEUROTOXIN = tax("肉毒毒素", "A型肉毒毒素", "医美及治疗适应症", "A型肉毒毒素")
PRO_SKINCARE = tax("功效性护肤品", "专业线/术后修护", "生长因子/精华导入", "专业护肤")
RAW_EXCLUDED = tax("原料/非终端产品", "医药原料药/API", "PDRN/PN 原料", "原料药", "excluded")
WRONG_ATTR_EXCLUDED = tax("错配记录", "公司/产品归属错误", "合并到正确主体", "错配排除", "excluded")


def merge_fields(*chunks: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for chunk in chunks:
        out.update(chunk)
    return out


WORKBOOK_UPDATES: dict[str, dict[str, str]] = {
    # Mesotherapy category cleanup.
    "REC_0387": merge_fields(
        LIPOLYTIC,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Lipolytic",
            "Tech_Type_Std": "Deoxycholic Acid",
            "Tech_Type_Original": "Deoxycholic acid lipolytic injection",
            "Core_Product": "Deoxycholic Acid (submental fat injection)",
            "FDA_Status": "FDA approved NDA 206333",
            "CE_Status": "Europe authorization/market access confirmed by user; exact pathway pending",
            "NMPA_Status": "China NMPA clinical/review status noted by user; no approval number captured",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[脱氧胆酸溶脂注射] Kybella/Belkyra 用于改善成人中度至重度下颌下脂肪外观，属于组织破坏性化学去脂药物，不归入美素营养成分。",
            "Feature_Tags": "injectables, lipolytic, deoxycholic-acid, submental-fat",
        },
    ),
    "REC_0734": merge_fields(
        HA_FILLER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Tech_Type_Original": "Cross-linked monophasic HA filler",
            "Core_Product": "HA Dermal Fillers",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[高交联 HA 深层填充] Varioderm 为 Adoderm 的深层皱纹静态填充和面部轮廓容量重塑产品，不归入美素浅层微滴注射。",
            "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, facial-contouring",
        },
    ),
    "REC_0707": merge_fields(
        HA_SKINBOOSTER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type_Std": "Hyaluronic Acid Skin Booster",
            "Tech_Type_Original": "Micro-crosslinked / non-crosslinked HA mesolift",
            "Core_Product": "Skin Boosters",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[浅层 HA 水光/美素] Touch My Skin / Varioderm Mesolift 用于面部、颈部、手背浅表微滴注射，改善细小干纹和真皮水合状态。",
            "Feature_Tags": "injectables, skin-booster, hyaluronic-acid, mesolift",
        },
    ),
    "REC_0103": merge_fields(
        CARBOXY_DEVICE,
        {
            "Category_L1": "EBD",
            "Category_L2": "Injection Devices",
            "Tech_Type_Std": "CO2 Carboxytherapy Device",
            "Tech_Type_Original": "Carboxytherapy gas injection pen",
            "Core_Product": "Carboxytherapy Devices",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[二氧化碳气体治疗设备] Carboxy-Pen 是卡波西疗法/CO2 气体治疗设备，不是 CO2 激光，也不是美素成分。",
            "Feature_Tags": "ebd, injection-device, carboxytherapy, co2",
        },
    ),
    "REC_0126": merge_fields(
        INJECTION_DEVICE,
        {
            "Category_L1": "EBD",
            "Category_L2": "Injection Devices",
            "Tech_Type_Std": "Electronic Mesotherapy Injection Gun",
            "Tech_Type_Original": "Microcomputer mesotherapy injector",
            "Core_Product": "Mesogun",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[美素注射枪] Concerto 是微电脑美素/给药设备，应作为给药载体而非注射成分管理。",
            "Feature_Tags": "ebd, injection-device, mesogun, mesotherapy-device",
        },
    ),
    "REC_0720": merge_fields(
        INJECTION_DEVICE,
        {
            "Category_L1": "EBD",
            "Category_L2": "Injection Devices",
            "Tech_Type_Std": "Electronic Mesotherapy Injection Gun",
            "Tech_Type_Original": "Microcomputer mesotherapy injector",
            "Core_Product": "Mesogun",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[美素注射枪] Ultim 是微电脑美素/给药设备，应作为给药载体而非注射成分管理。",
            "Feature_Tags": "ebd, injection-device, mesogun, mesotherapy-device",
        },
    ),
    "REC_0745": merge_fields(
        CARBOXY_DEVICE,
        {
            "Category_L1": "EBD",
            "Category_L2": "Injection Devices",
            "Tech_Type_Std": "CO2 Carboxytherapy Device",
            "Tech_Type_Original": "CO2 gas flow injection system",
            "Core_Product": "Carboxytherapy (CO2 Therapy)",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[二氧化碳气流注射仪] Venusian CO2 Therapy 通过皮下注入 CO2 促进局部循环，属于气体治疗/给药设备，不归入美素成分。",
            "Feature_Tags": "ebd, injection-device, carboxytherapy, co2",
        },
    ),
    "REC_0029": merge_fields(
        HAIR_EBD,
        {
            "Category_L1": "EBD",
            "Category_L2": "Hair Restoration",
            "Tech_Type_Std": "Ultrasound Trans-Epidermal Drug Delivery",
            "Tech_Type_Original": "Alma TED ultrasound hair restoration",
            "Core_Product": "Trans-Epidermal Drug Delivery",
            "FDA_Status": "FDA clearance confirmed by user; clearance number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[超声透皮给药生发设备] Alma TED 通过超声辅助非侵入性透皮给药用于毛发复苏，属于设备载体，不归入美素注射成分。",
            "Feature_Tags": "ebd, hair-restoration, ultrasound, trans-epidermal-drug-delivery",
        },
    ),
    "REC_0848": merge_fields(
        HAIR_EBD,
        {
            "Category_L1": "EBD",
            "Category_L2": "Hair Restoration",
            "Tech_Type_Std": "Ultrasound Trans-Epidermal Drug Delivery",
            "Tech_Type_Original": "Trans-epidermal drug delivery platform",
            "Core_Product": "Trans-Epidermal Drug Delivery",
            "FDA_Status": "FDA clearance confirmed by user; clearance number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[超声透皮给药生发设备] Alma TED 是非侵入式透皮给药/毛发复苏平台，作为设备载体管理。",
            "Feature_Tags": "ebd, hair-restoration, ultrasound, trans-epidermal-drug-delivery",
        },
    ),
    "REC_0044": merge_fields(
        INJECTION_DEVICE,
        {
            "Category_L1": "Consumables",
            "Category_L2": "Injection Devices",
            "Tech_Type_Std": "Microchannel Injection Device",
            "Tech_Type_Original": "AQUAGOLD fine touch microchannel device",
            "FDA_Status": "FDA listing/clearance noted by user; exact number pending",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[手动微通道给药器械] Aquagold Fine Touch 属于黄金微针印章/微通道输送工具，不是注射成分。",
            "Feature_Tags": "consumables, injection-device, microchannel, aquagold",
        },
    ),
    "REC_0453": merge_fields(
        INJECTION_DEVICE,
        {
            "Category_L1": "Consumables",
            "Category_L2": "Injection Devices",
            "Tech_Type_Std": "Microchannel Injection Device",
            "Tech_Type_Original": "Microchannel drug-delivery technology",
            "Core_Product": "Microchannel Injection Device",
            "FDA_Status": "FDA listing/clearance noted by user; exact number pending",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[手动微通道输送器械] Aquavit Microchannel Technology/Aquagold 属于微型给药耗材工具，不归入再生/美素成分。",
            "Feature_Tags": "consumables, injection-device, microchannel, aquagold",
        },
    ),
    "REC_0381": merge_fields(
        SKIN_MANAGEMENT,
        {
            "Category_L1": "EBD",
            "Category_L2": "Skin Management",
            "Tech_Type_Std": "HydraFacial Scalp Health Treatment",
            "Tech_Type_Original": "Vortex-Fusion scalp protocol",
            "Core_Product": "HydraFacial Keravive",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[海菲秀头皮护理设备流程] Keravive 属于 HydraFacial 非侵入式头皮护理/导入流程，不归入真皮层注射。",
            "Feature_Tags": "ebd, skin-management, hydrafacial, scalp-health",
        },
    ),
    "REC_0334": merge_fields(
        SKIN_MANAGEMENT,
        {
            "Category_L1": "EBD",
            "Category_L2": "Skin Management",
            "Tech_Type_Std": "HydraFacial Vortex-Fusion",
            "Tech_Type_Original": "Vacuum hydrodermabrasion and serum delivery",
            "FDA_Status": "FDA access/registration noted by user; exact number pending",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "NMPA_Status": "NMPA access confirmed by user; exact certificate pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[海菲秀负压皮肤管理仪] HydraFacial Syndeo/Elite 通过微流涡旋清洁与导入精华，属于非侵入式皮肤管理设备。",
            "Feature_Tags": "ebd, skin-management, hydrafacial, hydrodermabrasion",
        },
    ),
    "REC_0956": merge_fields(
        SKIN_MANAGEMENT,
        {
            "Category_L1": "EBD",
            "Category_L2": "Skin Management",
            "Tech_Type_Std": "HydraFacial Vortex-Fusion",
            "Tech_Type_Original": "Vortex-Fusion hydradermabrasion",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[海菲秀负压皮肤管理仪] HydraFacial Syndeo 属于非侵入式水氧/负压清洁导入设备。",
            "Feature_Tags": "ebd, skin-management, hydrafacial, hydrodermabrasion",
        },
    ),
    "REC_0957": merge_fields(
        SKIN_MANAGEMENT,
        {
            "Category_L1": "EBD",
            "Category_L2": "Skin Management",
            "Tech_Type_Std": "HydraFacial Keravive Scalp Treatment",
            "Tech_Type_Original": "Vortex-Fusion scalp protocol",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[海菲秀 Keravive 头皮流程] HydraFacial Keravive 是非侵入式头皮护理/导入系统，不是注射水光。",
            "Feature_Tags": "ebd, skin-management, hydrafacial, scalp-health",
        },
    ),
    "REC_0958": merge_fields(
        PRO_SKINCARE,
        {
            "Category_L1": "Skincare",
            "Category_L2": "Professional",
            "Tech_Type_Std": "Topical Booster Serum",
            "Tech_Type_Original": "HydraFacial topical boosters",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[海菲秀外用加强剂] HydraFacial Boosters 为皮肤管理设备配套外用精华，不归入注射成分。",
            "Feature_Tags": "skincare, professional, hydrafacial, topical-booster",
        },
    ),
    "REC_0071": merge_fields(
        PRO_SKINCARE,
        {
            "Category_L1": "Skincare",
            "Category_L2": "Professional",
            "Tech_Type_Std": "Growth Factor Sterile Topical",
            "Tech_Type_Original": "GF-DNA growth-factor serum; microneedling/TED adjunct",
            "Core_Product": "Growth Factors",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[生长因子术后修护/导入精华] Benev GF Series 通常作为医疗级无菌外用粉/精华配合微针、点阵或 Alma TED 导入，不按手针注射产品管理。",
            "Feature_Tags": "skincare, professional, growth-factor, post-procedure, hair-restoration-adjunct",
        },
    ),
    "REC_0032": merge_fields(
        HA_SKINBOOSTER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type_Std": "HA + Amino Acid Biorevitalization",
            "Tech_Type_Original": "Non-crosslinked HA + amino-acid matrix",
            "Core_Product": "Mesotherapy Solution",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[非交联 HA 氨基酸动能素] AMINO-JAL/JALUCOMPLEX 用于生物重组再生、浅表细纹改善与真皮层水分补充。",
            "Feature_Tags": "injectables, skin-booster, hyaluronic-acid, amino-acids, biorevitalization",
        },
    ),
    "REC_0178": merge_fields(
        HA_SKINBOOSTER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type_Std": "Multicomponent Mesotherapy Skin Booster",
            "Tech_Type_Original": "Vitamins + amino acids + minerals",
            "CE_Status": "CE Class III certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[多成分营养动能素] Dermastir 16/32 EBF 用于全脸真皮中层多点微注射，提高成熟期皮肤组织密度、激活代谢并改善静态细纹。",
            "Feature_Tags": "injectables, skin-booster, mesotherapy, vitamins, amino-acids",
        },
    ),
    "REC_0179": merge_fields(
        HA_SKINBOOSTER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type_Std": "Scalp Mesotherapy Skin Booster",
            "Tech_Type_Original": "EGF and peptide scalp booster",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[头皮美素动能素] Dermastir H53EGF 面向头皮衰老和脱发，通过浅层微滴注射作用于毛囊基质，延缓退行期并强化发根。",
            "Feature_Tags": "injectables, skin-booster, scalp, egf, hair-restoration",
        },
    ),
    # Botulinum-toxin category corrections and true toxin confirmations.
    "REC_0094": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "OnabotulinumtoxinA",
            "Tech_Type_Original": "Botulinum Toxin Type A",
            "FDA_Status": "FDA BLA approved",
            "CE_Status": "EU authorization confirmed by user; exact country/pathway pending",
            "NMPA_Status": "NMPA approved",
            "KFDA_Status": "KFDA/MFDS approval confirmed by user; number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[OnabotulinumtoxinA 全球肉毒标杆] Botox/Botox Cosmetic 覆盖眉间纹、鱼尾纹、额纹及慢性偏头痛、多汗、痉挛、斜视、眼睑痉挛等多治疗适应症。",
            "Feature_Tags": "injectables, neurotoxin, onabotulinumtoxina, botox",
        },
    ),
    "REC_0211": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "AbobotulinumtoxinA",
            "Tech_Type_Original": "Botulinum Toxin Type A",
            "FDA_Status": "FDA approved",
            "CE_Status": "EMA/EU authorization confirmed by user; exact number pending",
            "NMPA_Status": "NMPA approved",
            "KFDA_Status": "KFDA/MFDS registration status pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[AbobotulinumtoxinA] Dysport/Azzalure 为 Ipsen 生产、Galderma 医美渠道推广的 A 型肉毒，用于中重度眉间纹及颈部肌张力障碍、成人/儿童肢体痉挛等治疗适应症。",
            "Feature_Tags": "injectables, neurotoxin, abobotulinumtoxina, dysport, azzalure",
        },
    ),
    "REC_0772": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "IncobotulinumtoxinA",
            "Tech_Type_Original": "Pure Neurotoxin (no complexing proteins)",
            "FDA_Status": "FDA approved",
            "CE_Status": "EMA/EU authorization confirmed by user; exact number pending",
            "NMPA_Status": "NMPA approved",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[IncobotulinumtoxinA 纯化肉毒] Xeomin/Bocouture 主打去除复合蛋白，医美用于眉间纹，治疗端覆盖颈部肌张力障碍、眼睑痉挛、上肢痉挛和慢性流涎。",
            "Feature_Tags": "injectables, neurotoxin, incobotulinumtoxina, xeomin, bocouture",
        },
    ),
    "REC_0151": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "DaxibotulinumtoxinA-lanm",
            "Tech_Type_Original": "Peptide-formulated long-duration toxin",
            "FDA_Status": "FDA approved",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[DaxibotulinumtoxinA 长效肉毒] Daxxify 采用专有肽技术，医美适应症为中重度眉间纹，治疗适应症包括颈部肌张力障碍。",
            "Feature_Tags": "injectables, neurotoxin, daxibotulinumtoxina, daxxify",
        },
    ),
    "REC_0473": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "PrabotulinumtoxinA",
            "Tech_Type_Original": "Botulinum Toxin Type A",
            "FDA_Status": "FDA approved for Jeuveau",
            "CE_Status": "EMA/EU authorization confirmed by user for Nuceiva; exact number pending",
            "KFDA_Status": "KFDA/MFDS approval confirmed by user; number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[PrabotulinumtoxinA 绿毒] Nabota/Jeuveau/Nuceiva 的原研和注册厂商为 Daewoong；医美用于中重度眉间纹，韩国治疗端覆盖脑卒中后上肢痉挛和眼睑痉挛。",
            "Feature_Tags": "injectables, neurotoxin, prabotulinumtoxina, nabota, jeuveau, nuceiva",
        },
    ),
    "REC_0364": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "PrabotulinumtoxinA",
            "Tech_Type_Original": "900kDa purified botulinum toxin",
            "FDA_Status": "FDA approved for Jeuveau",
            "CE_Status": "EMA/EU authorization confirmed by user for Nuceiva; exact number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[PrabotulinumtoxinA 商业化渠道] Evolus 负责 Jeuveau/Nuceiva 等市场渠道，产品原研/制造归属 Daewoong。",
            "Feature_Tags": "injectables, neurotoxin, prabotulinumtoxina, jeuveau, nuceiva, daewoong",
        },
    ),
    "REC_0395": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "LetibotulinumtoxinA",
            "Tech_Type_Original": "Botulinum Toxin Type A",
            "FDA_Status": "FDA approved in 2024 confirmed by user; exact BLA number pending",
            "CE_Status": "Europe approvals confirmed by user; exact numbers pending",
            "NMPA_Status": "NMPA approved",
            "KFDA_Status": "KFDA/MFDS approval confirmed by user; number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[LetibotulinumtoxinA 白毒/乐提葆] Letybo/Botulax 医美用于中重度眉间纹，治疗端包括良性特发性眼睑痉挛和小儿脑瘫相关马蹄内翻足畸形。",
            "Feature_Tags": "injectables, neurotoxin, letibotulinumtoxina, letybo, botulax",
        },
    ),
    "REC_0805": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "AbobotulinumtoxinA Ready-to-Use Liquid",
            "Tech_Type_Original": "Ready-to-use liquid BoNT-A",
            "CE_Status": "Europe authorization confirmed by user; exact number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[即用型液态 A 型肉毒] Alluzience 由 Ipsen 制造、Galderma 经销，无需复溶，用于中重度眉间纹。",
            "Feature_Tags": "injectables, neurotoxin, liquid-bont-a, alluzience",
        },
    ),
    "REC_0962": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "AbobotulinumtoxinA Ready-to-Use Liquid",
            "Tech_Type_Original": "Liquid Ready-to-Use",
            "CE_Status": "Europe authorization confirmed by user; exact number pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[即用型液态 A 型肉毒] Alluzience 为 Ipsen/Galderma 液态肉毒产品，用于中重度眉间纹。",
            "Feature_Tags": "injectables, neurotoxin, liquid-bont-a, alluzience, ipsen",
        },
    ),
    "REC_0806": merge_fields(
        NEUROTOXIN,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "RelabotulinumtoxinA",
            "Tech_Type_Original": "Ready-to-use liquid BoNT-A / QM1114",
            "CE_Status": "Australia and other approvals confirmed by user; exact pathway pending",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[RelabotulinumtoxinA 新一代液态肉毒] Relfydess/QM1114 为 Galderma 自研即用型液态肉毒，用于改善眉间纹和鱼尾纹。",
            "Feature_Tags": "injectables, neurotoxin, liquid-bont-a, relfydess, relabotulinumtoxina",
        },
    ),
    "REC_0213": merge_fields(
        HA_FILLER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Tech_Type_Original": "Cross-linked HA filler",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[Jetema 交联 HA 填充剂] e.p.t.q. S100/300/500 是玻尿酸填充剂，不是肉毒素；Jetema 肉毒管线另为 JTM201。",
            "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, eptq, not-neurotoxin",
        },
    ),
    "REC_0090": merge_fields(
        SURGICAL_TOOL,
        {
            "Category_L1": "Surgical",
            "Category_L2": "Surgical Tools",
            "Tech_Type_Std": "Fat Grafting Surgical Kit",
            "Tech_Type_Original": "Plastic surgery instrument kit",
            "Core_Product": "B.A.F.F.I. Kit",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[脂肪移植/整形外科手术器械] B.A.F.F.I. Kit 属于 Bontempi/BMED 外科工具套件，不是肉毒素或注射材料。",
            "Feature_Tags": "surgical, surgical-tools, fat-grafting, plastic-surgery",
        },
    ),
    "REC_0091": merge_fields(
        SURGICAL_TOOL,
        {
            "Category_L1": "Surgical",
            "Category_L2": "Surgical Tools",
            "Tech_Type_Std": "Plastic Surgery Tools",
            "Tech_Type_Original": "Supercut scissors",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[整形外科剪刀器械] Supercut Scissors 属于 Bontempi/BMED 外科器械，不是肉毒素。",
            "Feature_Tags": "surgical, surgical-tools, plastic-surgery, scissors",
        },
    ),
    "REC_0092": merge_fields(
        SURGICAL_TOOL,
        {
            "Category_L1": "Surgical",
            "Category_L2": "Surgical Tools",
            "Tech_Type_Std": "Rhinoplasty Instruments",
            "Tech_Type_Original": "Rhinoplasty set",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[鼻整形手术器械] Rhinoplasty Instruments 属于 Bontempi/BMED 外科工具，不是肉毒素。",
            "Feature_Tags": "surgical, surgical-tools, rhinoplasty",
        },
    ),
    "REC_0224": merge_fields(
        HA_FILLER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Tech_Type_Original": "HA Filler (Fine/Deep/Sub-Q)",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[HA 填充剂规格矩阵] Eloquence Fine/Deep/Sub-Q 为玻尿酸填充剂不同规格/深度，不是肉毒素。",
            "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, not-neurotoxin",
        },
    ),
    "REC_0973": merge_fields(
        HA_SKINBOOSTER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type_Std": "Hyaluronic Acid Skin Booster",
            "Tech_Type_Original": "HA filler / skin booster",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[BNC Korea HA 填充/水光产品] BONIVA 为玻尿酸填充/水光类产品，不是肉毒素。",
            "Feature_Tags": "injectables, skin-booster, hyaluronic-acid, boniva, not-neurotoxin",
        },
    ),
    # Korea MFDS queue hard corrections.
    "REC_0963": merge_fields(
        WRONG_ATTR_EXCLUDED,
        {
            "Status": "Excluded",
            "Category_L1": "Injectables",
            "Category_L2": "Neurotoxin",
            "Tech_Type_Std": "PrabotulinumtoxinA",
            "Duplicate_Note": "Wrong attribution: Nabota belongs to Daewoong; use REC_0473 as canonical active record.",
            "Is_Primary_Record": "False",
            "Data_Source": "taxonomy_conflict_correction",
            "Introduction": "[错配排除] Dongkook Pharma/Nabota 为错误归属；Nabota 正确原研/注册主体是 Daewoong。",
            "Feature_Tags": "excluded, wrong-attribution, nabota, daewoong",
        },
    ),
    "REC_0965": merge_fields(
        WRONG_ATTR_EXCLUDED,
        {
            "Status": "Excluded",
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Duplicate_Note": "Wrong attribution: Yvoire belongs to LG Chem; use REC_0778 as canonical active record.",
            "Is_Primary_Record": "False",
            "Data_Source": "taxonomy_conflict_correction",
            "Introduction": "[错配排除] Jetema/Yvoire 为错误归属；Yvoire 正确主体是 LG Chem，Jetema HA 品牌为 e.p.t.q.。",
            "Feature_Tags": "excluded, wrong-attribution, yvoire, lg-chem",
        },
    ),
    "REC_0884": merge_fields(
        HA_FILLER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Tech_Type_Original": "HA Filler for superficial wrinkles",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[Neuramis Light 浅层 HA 填充剂] Medytox Neuramis Light 属于交联透明质酸填充剂，不是 EBD/射频设备。",
            "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, neuramis",
        },
    ),
    "REC_0759": merge_fields(
        HA_FILLER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Tech_Type_Original": "VOM HA filler O/M/V",
            "Data_Source": "official_company_fact_override",
            "Introduction": "[VOM 分层 HA 玻尿酸填充剂] VOM O/M/V 为分层定制 HA 填充剂，不是射频设备。",
            "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, vom",
        },
    ),
    "REC_0533": merge_fields(
        RAW_EXCLUDED,
        {
            "Status": "Excluded",
            "Category_L1": "Consumables",
            "Category_L2": "Raw Materials",
            "Tech_Type_Std": "PDRN / PN API Raw Material",
            "Tech_Type_Original": "API raw material",
            "Duplicate_Note": "Excluded: upstream API raw material, not terminal aesthetic product line.",
            "Is_Primary_Record": "False",
            "Data_Source": "taxonomy_conflict_correction",
            "Introduction": "[原料药排除] Daejoo PDRN/PN API 属于医药原料药，不是面向医美终端的医疗器械/注射耗材产品线。",
            "Feature_Tags": "excluded, api, raw-material, pdrn, pn",
        },
    ),
    "REC_0005": {
        "KFDA_Status": "MFDS/KFDA registration confirmed by user; certificate number pending",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[韩国 HIFU 设备] 10Thera/TenThera 为 Tentech 韩国合规 HIFU 设备，需后续反查 MFDS 具体注册号。",
        "Feature_Tags": "ebd, hifu, korea, mfds-pending-number",
    },
    "REC_0006": {
        "KFDA_Status": "MFDS/KFDA registration confirmed by user; certificate number pending",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[韩国单极射频设备] 10Therma/TenTherma 为 Tentech 韩国本土高活跃单极 RF 设备，需后续反查 MFDS 具体注册号。",
        "Feature_Tags": "ebd, radiofrequency, monopolar-rf, korea, mfds-pending-number",
    },
    "TOPLINE_MEDYTOX_NEWLUX_20260530": {
        "KFDA_Status": "MFDS/KFDA approved; certificate number pending",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[Medytox 新一代 A 型肉毒] NEWLUX 是 Medytox 获韩国 MFDS 批准的新一代肉毒素产品，具体注册号待反查。",
        "Feature_Tags": "injectables, neurotoxin, botulinum-toxin-a, mfds-pending-number",
    },
    "REC_1055": {
        "KFDA_Status": "MFDS/KFDA registration confirmed by user; certificate number pending",
        "CE_Status": "CE certification confirmed by user; certificate number pending",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[PLA/HA 再生晶球] Etrebelle 为 FacePharm Korea 的 PLA/HA 再生填充产品，韩国本土及出口活跃，MFDS 具体证号待反查。",
        "Feature_Tags": "injectables, biostimulator, pla, ha, korea, mfds-pending-number",
    },
    "REC_0932": merge_fields(
        HA_SKINBOOSTER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type_Std": "Peptide-enriched HA Skin Booster",
            "Tech_Type_Original": "Aquashine skin booster",
            "KFDA_Status": "MFDS/KFDA registration confirmed by user; certificate number pending",
            "CE_Status": "CE certification confirmed by user; certificate number pending",
            "Data_Source": "official_product_fact_promoted",
            "Introduction": "[Caregen 肽复配 HA 水光] Aquashine 为 Caregen 当家玻尿酸/水光线，MFDS 历史证件丰富，具体证号待反查。",
            "Feature_Tags": "injectables, skin-booster, hyaluronic-acid, peptide, korea",
        },
    ),
    "REC_0964": merge_fields(
        HA_FILLER,
        {
            "Category_L1": "Injectables",
            "Category_L2": "Dermal Filler",
            "Tech_Type_Std": "Hyaluronic Acid Dermal Filler",
            "Tech_Type_Original": "Elravie HA filler",
            "KFDA_Status": "MFDS/KFDA registration confirmed by user; certificate number pending",
            "Data_Source": "official_product_fact_promoted",
            "Introduction": "[Dongkook HA 填充剂] Elravie 为 Dongkook Pharma 当家玻尿酸填充线，MFDS 具体证号待反查。",
            "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, korea",
        },
    ),
    # Product gap P0/P1 examples: lock in registration and differentiator wording.
    "TOPLINE_ALLERGAN_JUVEDERM_ULTRA_PLUS_XC_20260530": {
        "FDA_Status": "FDA PMA approved P050047/S005",
        "FDA_510k_Number": "P050047/S005",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[高交联 HA 深层容量填充] Juvéderm Ultra Plus XC 获 FDA PMA P050047/S005，用于真皮中深层注射以纠正中重度面部皱纹和褶皱，如鼻唇沟。",
        "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, fda-pma, nasolabial-folds",
    },
    "TOPLINE_GALDERMA_RESTYLANE_DEFYNE_20260530": {
        "FDA_Status": "FDA PMA approved P140029 and supplements",
        "FDA_510k_Number": "P140029",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[XpresHAn HA 动态填充] Restylane Defyne 获 FDA PMA P140029 及补充审批，用于中重度面部皱纹/鼻唇沟，并可用于轻中度下巴后缩的下巴增大。",
        "Feature_Tags": "injectables, dermal-filler, hyaluronic-acid, fda-pma, nasolabial-folds, chin-augmentation",
    },
    "TOPLINE_BTL_EMTONE_20260530": {
        "FDA_Status": "FDA 510(k) cleared; number pending",
        "Data_Source": "official_product_fact_promoted",
        "Introduction": "[RF + 靶向压力能橘皮组织设备] BTL EMTONE 结合单极射频与 TPE/声波压力能，FDA 510(k) 许可，用于非侵入性橘皮组织治疗及身体塑形相关场景。",
        "Feature_Tags": "ebd, radiofrequency, targeted-pressure-energy, cellulite, fda-510k",
    },
}


BACKFILL_NOTES: dict[str, str] = {
    record_id: "user_feedback_20260601_category_mfds_toxin_gap: user-confirmed classification/ownership/regulatory facts applied."
    for record_id in WORKBOOK_UPDATES
}
BACKFILL_NOTES.update(
    {
        "REC_0963": "user_feedback_20260601_korea_mfds: wrong Nabota/Dongkook attribution excluded; canonical active row is Daewoong REC_0473.",
        "REC_0965": "user_feedback_20260601_korea_mfds: wrong Yvoire/Jetema attribution excluded; canonical active row is LG Chem REC_0778.",
        "REC_0533": "user_feedback_20260601_korea_mfds: PDRN/PN API raw-material line excluded from terminal aesthetic product master.",
    }
)


def reg(
    record_id: str,
    jurisdiction: str,
    regulator: str,
    pathway: str,
    status: str,
    registered_name: str,
    indication: str,
    manufacturer: str,
    source_url: str,
    title: str,
    excerpt: str,
    registration_no: str = "",
    field_note: str = "",
    source_type: str = "user_confirmed_regulatory_claim",
) -> dict[str, str]:
    return {
        "record_id": record_id,
        "jurisdiction": jurisdiction,
        "regulator": regulator,
        "pathway": pathway,
        "status": status,
        "registered_name": registered_name,
        "indication": indication,
        "legal_manufacturer": manufacturer,
        "source_url": source_url,
        "source_type": source_type,
        "evidence_title": title,
        "evidence_excerpt": excerpt,
        "registration_no": registration_no,
        "field_note": field_note,
    }


REGULATORY_SPECS: list[dict[str, str]] = [
    reg(
        "REC_0387",
        "US",
        "FDA",
        "NDA drug approval",
        "FDA approved",
        "KYBELLA (deoxycholic acid) injection",
        "Indicated for improvement in the appearance of moderate to severe convexity or fullness associated with submental fat in adults.",
        "Allergan",
        "https://www.accessdata.fda.gov/drugsatfda_docs/label/2024/206333lbl.pdf",
        "KYBELLA FDA label",
        "用户确认 Kybella/Belkyra 属于脱氧胆酸溶脂注射剂，不属于美素营养成分。",
        "NDA206333",
    ),
    reg(
        "REC_0387",
        "EU",
        "European medicines authority / national authorities",
        "European authorization",
        "Europe authorization confirmed by user; exact pathway/number pending",
        "BELKYRA (deoxycholic acid)",
        "Improves the appearance of adult moderate to severe submental fat/fullness.",
        "Allergan",
        "",
        "Belkyra Europe authorization user-confirmed",
        "用户确认 Belkyra 在欧洲拥有合规准入；本批不写入具体证号。",
        "",
        "Do not treat Kybella/Belkyra as mesotherapy nutrient cocktail.",
    ),
    reg(
        "REC_0734",
        "EU",
        "European Commission / Notified Body",
        "CE certification",
        "CE certification confirmed by user; certificate number pending",
        "Varioderm HA Dermal Fillers",
        "For correction of moderate to severe deep facial wrinkles and facial contour/volume restoration.",
        "Adoderm",
        "https://www.adoderm.com/our-technology",
        "Adoderm Varioderm user-confirmed category",
        "用户确认 Varioderm 为高交联 HA 填充剂，不属于 Mesolift 水光成分。",
    ),
    reg(
        "REC_0707",
        "EU",
        "European Commission / Notified Body",
        "CE certification",
        "CE certification confirmed by user; certificate number pending",
        "Touch My Skin / Varioderm Mesolift",
        "For superficial microdroplet injection on face, neck and hands to improve fine dryness lines and dermal hydration.",
        "Adoderm",
        "https://www.adoderm.com/our-ifu",
        "Adoderm Touch My Skin / Varioderm Mesolift user-confirmed use",
        "用户确认 Touch My Skin / Varioderm Mesolift 才属于 Adoderm 的水光/美素产品。",
    ),
    reg(
        "REC_0449",
        "EU",
        "European Commission / Notified Body",
        "CE certification",
        "CE certification confirmed by user; certificate number pending",
        "Meso-Xanthin F199",
        "DNA-level mesotherapy cell revitalizer for severe photoaging, inflammatory facial skin conditions, pigmentation and barrier repair.",
        "ABG Lab LLC",
        "https://mesowhartonp199.com/",
        "Meso-Xanthin user-confirmed indication",
        "用户确认 Meso-Xanthin F199 作为中胚层细胞焕活剂拥有欧洲 CE 认证。",
    ),
    reg(
        "REC_0032",
        "EU",
        "European Commission / Notified Body",
        "CE certification",
        "CE certification confirmed by user; certificate number pending",
        "AMINO-JAL / JALUCOMPLEX",
        "Non-crosslinked HA plus amino-acid biorevitalization for superficial lines, dermal hydration and fibroblast function support.",
        "BioFormula",
        "https://bioformula.it/en/product/jalucomplex-1/",
        "BioFormula AMINO-JAL / JALUCOMPLEX user-confirmed indication",
        "用户确认 AMINO-JAL/JALUCOMPLEX 为意大利复配型非交联玻尿酸动能素。",
    ),
    reg(
        "REC_0178",
        "EU",
        "Notified Body (CE)",
        "CE Class III certification",
        "CE Class III certification confirmed by user; certificate number pending",
        "Dermastir 16 EBF / 32 EBF",
        "Multicomponent mesotherapy vials for mature skin tissue density, cell metabolism activation and static fine-line improvement.",
        "Alta Care",
        "https://dermastir.com/products/dermastir-16ebf",
        "Dermastir 16/32 EBF user-confirmed indication",
        "用户确认 Dermastir 16/32 EBF 为三类 CE 高浓度营养动能素。",
    ),
    reg(
        "REC_0179",
        "EU",
        "European Commission / Notified Body",
        "CE certification",
        "CE certification confirmed by user; certificate number pending",
        "Dermastir H53EGF",
        "Scalp mesotherapy with EGF/peptides for hair-follicle aging, shedding delay, hair regrowth stimulation and root strengthening.",
        "Alta Care",
        "https://dermastir.com/products/dermastir-meso-sterile-vials-h53egf",
        "Dermastir H53EGF user-confirmed indication",
        "用户确认 Dermastir H53EGF 属于头皮美素/防脱生发细分。",
    ),
    reg(
        "REC_0094",
        "US",
        "FDA",
        "BLA drug/biologic approval",
        "FDA approved",
        "BOTOX / BOTOX Cosmetic (onabotulinumtoxinA)",
        "Cosmetic indications include temporary improvement of moderate to severe glabellar lines, lateral canthal lines and forehead lines; therapeutic uses include chronic migraine, overactive bladder, urinary incontinence, cervical dystonia, severe axillary hyperhidrosis, limb spasticity, strabismus and blepharospasm.",
        "Allergan",
        "https://api.fda.gov/drug/label.json?search=openfda.brand_name%3A%22BOTOX%20COSMETIC%22&limit=1",
        "BOTOX/BOTOX Cosmetic user-confirmed global indication map",
        "用户确认 Botox/Botox Cosmetic 为 OnabotulinumtoxinA，拥有全球多区域注册和最多治疗适应症。",
        "BLA103000",
    ),
    reg(
        "REC_0094",
        "KR",
        "KFDA/MFDS",
        "medicinal product approval",
        "KFDA/MFDS approval confirmed by user; number pending",
        "BOTOX / BOTOX Cosmetic",
        "Aesthetic and therapeutic botulinum toxin indications confirmed at product-family level.",
        "Allergan",
        "",
        "BOTOX Korea user-confirmed approval",
        "用户确认 Botox 在韩国 KFDA/MFDS 等全球主流市场拥有注册。",
    ),
    reg(
        "REC_0211",
        "US",
        "FDA",
        "drug/biologic approval",
        "FDA approved",
        "Dysport / Azzalure (abobotulinumtoxinA)",
        "Aesthetic indication: moderate to severe glabellar lines. Therapeutic indications include cervical dystonia and adult/pediatric upper/lower limb spasticity.",
        "Ipsen / Galderma",
        "https://www.dysportusa.com/",
        "Dysport/Azzalure user-confirmed indication map",
        "用户确认 Ipsen 负责生产和医疗渠道，Galderma 负责医美渠道。",
    ),
    reg(
        "REC_0211",
        "EU",
        "EMA / European national authorities",
        "medicinal product authorization",
        "Europe authorization confirmed by user; number pending",
        "Dysport / Azzalure (abobotulinumtoxinA)",
        "Moderate to severe glabellar lines; cervical dystonia and limb spasticity therapeutic indications.",
        "Ipsen / Galderma",
        "",
        "Dysport/Azzalure Europe user-confirmed authorization",
        "用户确认 Dysport/Azzalure 拥有 FDA、EMA、NMPA 等近 90 个国家注册。",
    ),
    reg(
        "REC_0772",
        "US",
        "FDA",
        "drug/biologic approval",
        "FDA approved",
        "Xeomin / Bocouture (incobotulinumtoxinA)",
        "Cosmetic use for glabellar lines; therapeutic indications include cervical dystonia, blepharospasm, upper limb spasticity and chronic sialorrhea.",
        "Merz",
        "https://merztherapeutics.com/us/products/xeomin/",
        "Xeomin/Bocouture user-confirmed indication map",
        "用户确认 Xeomin/Bocouture 为 IncobotulinumtoxinA，主打不含复合蛋白。",
    ),
    reg(
        "REC_0151",
        "US",
        "FDA",
        "drug/biologic approval",
        "FDA approved",
        "Daxxify (daxibotulinumtoxinA-lanm)",
        "Indicated for moderate to severe glabellar lines and cervical dystonia; positioned as longer-duration peptide-formulated toxin.",
        "Revance Therapeutics",
        "https://www.daxxify.com/",
        "Daxxify user-confirmed indication map",
        "用户确认 Daxxify 当前主要获得美国 FDA 批准。",
    ),
    reg(
        "REC_0473",
        "KR",
        "KFDA/MFDS",
        "medicinal product approval",
        "KFDA/MFDS approved; certificate number pending",
        "Nabota (prabotulinumtoxinA)",
        "Aesthetic use for moderate to severe glabellar lines; Korea therapeutic approvals include post-stroke upper limb spasticity and blepharospasm.",
        "Daewoong",
        "https://www.daewoong.co.kr/en/product/biologics",
        "Daewoong Nabota user-confirmed owner and indications",
        "用户确认 Nabota 原研和注册厂商为 Daewoong；Dongkook 归属为错配。",
    ),
    reg(
        "REC_0473",
        "US",
        "FDA",
        "drug/biologic approval",
        "FDA approved for Jeuveau",
        "Jeuveau / Nabota (prabotulinumtoxinA)",
        "For moderate to severe glabellar lines.",
        "Daewoong / Evolus",
        "https://www.evolus.com/",
        "Jeuveau user-confirmed FDA approval",
        "用户确认 Daewoong/Evolus Nabota/Jeuveau/Nuceiva 拥有韩国 KFDA、美国 FDA、欧洲 EMA 认证。",
    ),
    reg(
        "REC_0364",
        "US",
        "FDA",
        "drug/biologic approval",
        "FDA approved",
        "Jeuveau (prabotulinumtoxinA-xvfs)",
        "For moderate to severe glabellar lines; commercialized by Evolus with Daewoong as origin/manufacturer.",
        "Daewoong / Evolus",
        "https://www.evolus.com/",
        "Evolus Jeuveau user-confirmed channel",
        "用户确认 Jeuveau/Nuceiva 为 Daewoong 产品的美国/欧洲商业名称。",
    ),
    reg(
        "REC_0395",
        "KR",
        "KFDA/MFDS",
        "medicinal product approval",
        "KFDA/MFDS approved; certificate number pending",
        "Letybo / Botulax (letibotulinumtoxinA)",
        "Aesthetic use for moderate to severe glabellar lines; therapeutic uses include benign essential blepharospasm and pediatric cerebral palsy equinus foot deformity.",
        "Hugel",
        "https://hugel-aesthetics.com/",
        "Hugel Letybo/Botulax user-confirmed indication map",
        "用户确认 Hugel 在韩国、中国、美国和欧洲多地获批。",
    ),
    reg(
        "REC_0395",
        "US",
        "FDA",
        "drug/biologic approval",
        "FDA approved in 2024",
        "Letybo (letibotulinumtoxinA)",
        "For moderate to severe glabellar lines.",
        "Hugel",
        "https://hugel-aesthetics.com/",
        "Letybo FDA approval user-confirmed",
        "用户确认 Letybo 于 2024 年初获得美国 FDA 批准。",
    ),
    reg(
        "REC_0805",
        "EU",
        "EMA / European authorities",
        "medicinal product authorization",
        "Europe authorization confirmed by user; number pending",
        "Alluzience",
        "Ready-to-use liquid botulinum toxin A for moderate to severe glabellar lines.",
        "Ipsen / Galderma",
        "https://www.galderma.com/galderma-aesthetics",
        "Alluzience user-confirmed authorization",
        "用户确认 Alluzience 已在欧洲和澳大利亚获批，主要用于中重度眉间纹。",
    ),
    reg(
        "REC_0962",
        "EU",
        "EMA / European authorities",
        "medicinal product authorization",
        "Europe authorization confirmed by user; number pending",
        "Alluzience",
        "Ready-to-use liquid botulinum toxin A for moderate to severe glabellar lines.",
        "Ipsen Pharma",
        "https://www.ipsen.com/medicines/",
        "Ipsen Alluzience user-confirmed authorization",
        "用户确认 Alluzience 由 Ipsen 制造、Galderma 经销。",
    ),
    reg(
        "REC_0806",
        "AU",
        "TGA",
        "medicinal product authorization",
        "Australia approval confirmed by user; number pending",
        "Relfydess (relabotulinumtoxinA)",
        "Ready-to-use liquid botulinum toxin A for glabellar lines and crow's feet.",
        "Galderma",
        "https://www.galderma.com/galderma-aesthetics",
        "Relfydess user-confirmed authorization",
        "用户确认 Relfydess/QM1114 已在澳大利亚等地获批。",
    ),
    reg(
        "REC_0005",
        "KR",
        "KFDA/MFDS",
        "medical device registration",
        "MFDS/KFDA registration confirmed by user; certificate number pending",
        "10Thera / TenThera",
        "Korean HIFU device for aesthetic lifting/tissue tightening workflows; exact MFDS number pending.",
        "Tentech",
        "",
        "Tentech 10Thera user-confirmed MFDS registration",
        "用户确认 10Thera 为韩国本土高活跃 HIFU 设备，MFDS 注册合规。",
    ),
    reg(
        "REC_0006",
        "KR",
        "KFDA/MFDS",
        "medical device registration",
        "MFDS/KFDA registration confirmed by user; certificate number pending",
        "10Therma / TenTherma",
        "Korean monopolar RF device for skin tightening/lifting workflows; exact MFDS number pending.",
        "Tentech",
        "",
        "Tentech 10Therma user-confirmed MFDS registration",
        "用户确认 10Therma 为韩国本土高活跃单极射频设备，MFDS 注册合规。",
    ),
    reg(
        "TOPLINE_MEDYTOX_NEWLUX_20260530",
        "KR",
        "KFDA/MFDS",
        "medicinal product approval",
        "MFDS/KFDA approved; certificate number pending",
        "NEWLUX botulinum toxin type A",
        "New-generation Medytox botulinum toxin product approved by Korean MFDS; exact approval number pending.",
        "Medytox",
        "https://medytox.com/page/newlux?site_id=en",
        "Medytox NEWLUX user-confirmed MFDS approval",
        "用户确认 NEWLUX 为 Medytox 旗下最新一代获韩国 MFDS 批准的肉毒素。",
    ),
    reg(
        "REC_1055",
        "KR",
        "KFDA/MFDS",
        "medical device registration",
        "MFDS/KFDA registration confirmed by user; certificate number pending",
        "Etrebelle",
        "PLA/HA hybrid biostimulator/filler product line; Korea registration confirmed at product-identity level.",
        "FacePharm Korea",
        "",
        "Etrebelle user-confirmed MFDS registration",
        "用户确认 Etrebelle 在韩国本土及出口活跃，应直接激活并反查 MFDS 证号。",
    ),
    reg(
        "REC_0932",
        "KR",
        "KFDA/MFDS",
        "medical device registration",
        "MFDS/KFDA registration confirmed by user; certificate number pending",
        "Aquashine",
        "Peptide-enriched HA skin booster/filler line with Korean registration history; exact MFDS number pending.",
        "Caregen",
        "",
        "Caregen Aquashine user-confirmed MFDS registration",
        "用户确认 Aquashine 为 Caregen 当家玻尿酸/水光线，MFDS 历史证件丰富。",
    ),
    reg(
        "REC_0964",
        "KR",
        "KFDA/MFDS",
        "medical device registration",
        "MFDS/KFDA registration confirmed by user; certificate number pending",
        "Elravie",
        "Dongkook Pharma HA dermal filler line; exact MFDS number pending.",
        "Dongkook Pharma",
        "",
        "Dongkook Elravie user-confirmed MFDS registration",
        "用户确认 Elravie 为 Dongkook 当家玻尿酸/水光线，MFDS 历史证件丰富。",
    ),
    reg(
        "TOPLINE_ALLERGAN_JUVEDERM_ULTRA_PLUS_XC_20260530",
        "US",
        "FDA",
        "PMA supplement approval",
        "FDA PMA approved",
        "Juvéderm Ultra Plus XC",
        "For injection into the mid-to-deep dermis for correction of moderate to severe facial wrinkles and folds, such as nasolabial folds.",
        "Allergan",
        "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpma/pma.cfm?id=P050047S005",
        "FDA PMA P050047/S005 - Juvéderm Ultra Plus XC",
        "用户确认 Juvéderm Ultra Plus XC 获 FDA PMA P050047/S005。",
        "P050047/S005",
    ),
    reg(
        "TOPLINE_GALDERMA_RESTYLANE_DEFYNE_20260530",
        "US",
        "FDA",
        "PMA approval and supplements",
        "FDA PMA approved",
        "Restylane Defyne",
        "For correction of moderate to severe deep facial wrinkles and folds such as nasolabial folds in patients over 21; also for chin augmentation to improve chin profile in patients over 21 with mild-to-moderate chin retrusion.",
        "Galderma",
        "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpma/pma.cfm?id=P140029",
        "FDA PMA P140029 - Restylane Defyne",
        "用户确认 Restylane Defyne 获 FDA PMA P140029 及补充审批。",
        "P140029",
    ),
    reg(
        "TOPLINE_BTL_EMTONE_20260530",
        "US",
        "FDA",
        "510(k) clearance",
        "FDA 510(k) cleared; number pending",
        "BTL EMTONE",
        "Non-invasive cellulite treatment using monopolar radiofrequency and targeted pressure energy; also used in body-contouring/skin-tightening workflows.",
        "BTL",
        "https://bodybybtl.com/solutions/emtone",
        "BTL EMTONE user-confirmed FDA 510(k) clearance",
        "用户确认 EMTONE 获美国 FDA 510(k) 许可；本批未提供具体 K 号。",
    ),
]


FACT_SPECS: list[dict[str, str]] = [
    {
        "record_id": record_id,
        "fact_group": "user_confirmed_classification",
        "field_name": "classification_correction",
        "field_value": note,
        "source_url": "",
        "evidence_title": "User-confirmed classification / ownership correction 2026-06-01",
        "evidence_excerpt": note,
    }
    for record_id, note in BACKFILL_NOTES.items()
]


def evidence_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    indication = norm(spec.get("indication"))
    source_key = stable_id(
        "user_confirmed_20260601_cat_mfds_toxin",
        product.get("seed_record_id"),
        spec.get("jurisdiction"),
        spec.get("regulator"),
        spec.get("registered_name"),
        spec.get("registration_no"),
        indication[:60],
    )
    return {
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "jurisdiction": norm(spec.get("jurisdiction")),
        "regulator": norm(spec.get("regulator")),
        "regulatory_pathway": norm(spec.get("pathway")),
        "status": norm(spec.get("status")),
        "registration_no": norm(spec.get("registration_no")),
        "approval_date": "",
        "expiry_date": "",
        "registered_name": norm(spec.get("registered_name")) or product.get("brand", ""),
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": norm(spec.get("legal_manufacturer")) or product.get("legal_manufacturer") or product.get("company", ""),
        "local_holder": "",
        "source_key": source_key,
        "source_url": norm(spec.get("source_url")),
        "source_type": norm(spec.get("source_type")) or "user_confirmed_regulatory_claim",
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": norm(spec.get("field_note")),
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601_category_mfds_toxin_gap",
        "review_status": "user_confirmed",
        "confidence": "user_confirmed_official_claim",
    }


def fact_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    return {
        "fact_id": stable_id("pfact", product.get("seed_record_id"), spec.get("fact_group"), spec.get("field_name"), spec.get("field_value")),
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "product_family_id": "",
        "standard_product_name": product.get("standard_product_name", ""),
        "priority": "P0",
        "fact_group": norm(spec.get("fact_group")),
        "field_name": norm(spec.get("field_name")),
        "field_value": norm(spec.get("field_value")),
        "source_url": norm(spec.get("source_url")),
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "source_type": "user_confirmed_master_correction",
        "confidence": "user_confirmed",
        "captured_at": checked_at,
        "promoted_at": checked_at,
        "review_status": "promoted",
        "note": "user_feedback_20260601_category_mfds_toxin_gap",
    }


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_user_feedback_cat_mfds_toxin_gap_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if old == value:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": value})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id, fields in WORKBOOK_UPDATES.items():
        for field, value in fields.items():
            set_cell(record_id, field, value)
        append_audit(record_id, BACKFILL_NOTES.get(record_id, "user_feedback_20260601_category_mfds_toxin_gap: user-confirmed feedback applied."))

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()

    evidence_specs = [spec for spec in REGULATORY_SPECS if spec["record_id"] in products]
    evidence_rows = [evidence_row(products[spec["record_id"]], checked_at, spec) for spec in evidence_specs]
    ind_fields, ind_rows = read_csv(MANUAL_INDICATION_PATH)
    added_indications = append_unique(
        ind_rows,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        evidence_rows,
    )
    write_csv(MANUAL_INDICATION_PATH, ind_fields, ind_rows)

    fact_specs = [spec for spec in FACT_SPECS if spec["record_id"] in products]
    fact_rows_new = [fact_row(products[spec["record_id"]], checked_at, spec) for spec in fact_specs]
    fact_fields, fact_rows = read_csv(MANUAL_FACT_PATH)
    added_facts = append_unique(fact_rows, ["fact_id"], fact_rows_new)
    write_csv(MANUAL_FACT_PATH, fact_fields, fact_rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_indications,
        "manual_product_fact_rows_added": added_facts,
        "target_record_ids": sorted(WORKBOOK_UPDATES),
        "regulatory_record_ids": sorted({spec["record_id"] for spec in evidence_specs}),
        "excluded_record_ids": ["REC_0533", "REC_0963", "REC_0965"],
        "already_resolved_before_this_script": {
            "BioPlus/Rejuran active mismatch": "0 active rows in latest product_master before script; canonical PharmaResearch rows retained.",
        },
        "changed_fields_sample": workbook_changes[:180],
    }
    out = AUDIT_DIR / f"user_feedback_category_mfds_toxin_gap_{stamp}.json"
    latest = AUDIT_DIR / "user_feedback_category_mfds_toxin_gap_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
