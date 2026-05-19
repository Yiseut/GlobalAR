#!/usr/bin/env python3
"""
Apply deterministic integrity cleanup back to the source Excel workbook.

The source workbook is treated as the editable seed database. This script only
performs mechanical fixes that are traceable from existing rows, and records
every changed cell in Cleanup_Log plus data/seed_cleanup_log.json.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from build_data import (
    COMPANY_BACKGROUND_EVIDENCE_PATH,
    COMPANY_CAPITAL_STRUCTURE_PATH,
    COMPANY_MEDIA_ASSET_INDEX_PATH,
    COMPANY_OFFICIAL_SOURCE_EVIDENCE_PATH,
    COMPANY_OFFICIAL_SOURCE_PLAN_PATH,
    COMPANY_OFFICIAL_WEBSITE_PATH,
    DATA_DIR,
    DATA_QUALITY_ISSUES_PATH,
    EVIDENCE_PROMOTION_LOG_PATH,
    LISTED_COMPANY_BATCH_PATH,
    MDR_CE_SEARCH_PLAN_PATH,
    OFFICIAL_WEBSITE_MASTER_PATH,
    OFFICIAL_INDICATION_EVIDENCE_PATH,
    POLICY_REGULATORY_SOURCE_PLAN_PATH,
    PRODUCT_MASTER_PATH,
    PRODUCT_FAMILY_MASTER_PATH,
    PRODUCT_SPECIFICATION_EVIDENCE_PATH,
    PRODUCT_SKU_MASTER_PATH,
    REGISTRATION_EVIDENCE_PATH,
    SOURCE_AUTHORITY_POLICY_PATH,
    SOURCE_DIR,
    compact_key,
    find_file,
    norm,
    stable_id,
)


LOG_PATH = DATA_DIR / "seed_cleanup_log.json"
SUMMARY_PATH = DATA_DIR / "seed_cleanup_summary.json"
SOURCE_PATTERN = "*v4.xlsx"
MARKET_VALUATION_RANK_PATH = DATA_DIR / "market_valuation_rank.csv"
MARKET_CONFLICT_PATH = DATA_DIR / "market_snapshot_conflicts.csv"
OFFICIAL_SOURCE_COVERAGE_PATH = DATA_DIR / "official_source_coverage.csv"
SOURCE_DIFF_REPORT_PATH = DATA_DIR / "source_diff_report.csv"
MDR_CE_EVIDENCE_PATH = DATA_DIR / "mdr_ce_evidence_candidates.jsonl"
XUEQIU_MARKET_CHECK_PATH = DATA_DIR / "xueqiu_market_check.csv"

TECH_TYPE_FIXES = {
    "REC_0216": "Electroporation",
    "REC_0780": "Hydrodermabrasion / Oxygen Infusion",
    "REC_0387": "Deoxycholic Acid",
    "REC_0030": "Probiotic Supplement",
    "REC_0453": "Microchannel Drug Delivery",
    "REC_0036": "Medical Aesthetics Training / Conference",
    "REC_0046": "Online Medical Education Platform",
    "REC_0303": "Silicone Body Implant",
    "REC_0475": "Needle Endoscopy",
    "REC_0087": "Digital Dermoscopy",
    "REC_0563": "Hair Nutraceutical Supplement",
    "REC_0500": "Needle-Free Injector",
    "REC_0501": "Needle-Free Injector Nozzle",
    "REC_0333": "Topical Booster Serum",
    "REC_0076": "Wet/Dry Microdermabrasion",
    "REC_0086": "Microcurrent",
    "REC_0010": "Pulsed Shortwave Therapy",
    "REC_0110": "PPC / Deoxycholate Lipolysis",
    "REC_0380": "Keratin Hair Supplement",
    "REC_0034": "Multispectral Skin Diagnosis",
    "REC_0035": "Cryotherapy Spray Tip",
    "REC_0350": "Total Body Photography",
    "REC_0612": "Portable Skin Imaging",
    "REC_0739": "3D Imaging / Simulation",
    "REC_0753": "Multispectral Skin Analysis",
    "REC_0756": "Deoxycholic Acid",
    "REC_0402": "PPC / Deoxycholate Lipolysis",
    "REC_0268": "Fat Transfer Cannula",
    "REC_0344": "Infiltration Pump",
    "REC_0557": "Power Assisted Liposuction",
    "REC_0788": "Injection Consumables",
    "REC_0792": "Vacuum Mesotherapy Injector",
    "REC_0793": "Electronic Injection Gun",
    "REC_0795": "Injection Needles",
    "REC_0796": "Biologics Platform",
}

CATEGORY_DEFINITION_FIXES = {
    ("Skincare", "Supplements"): ("Oral supplement products positioned around skin, hair, wellness or peri-procedure support.", "Probiotic Supplement"),
    ("Regenerative", "Biotech"): ("Biotechnology platforms, delivery systems or regenerative-enabling technologies not yet mapped to a narrower product class.", "Microchannel Drug Delivery"),
    ("Services", "Training-Education"): ("Education, training, conference or platform services for aesthetic clinicians.", "Medical Aesthetics Training / Conference"),
    ("Skincare", "Nutraceuticals"): ("Oral nutritional products for hair, skin or aesthetic wellness support.", "Hair Nutraceutical Supplement"),
    ("Pharma", "Pain Management"): ("Pharmaceutical or device-assisted pain management products used around aesthetic or surgical workflows.", "Pulsed Shortwave Therapy"),
    ("Skincare", "Nutricosmetics"): ("Ingestible beauty products marketed for skin, hair or anti-aging benefits.", "Keratin Hair Supplement"),
    ("EBD", "Skin Diagnosis"): ("Energy/device-enabled skin or scalp diagnosis and analysis systems.", "Multispectral Skin Diagnosis"),
    ("Consumables", "Injection Devices"): ("Reusable or consumable injection-assist devices and accessories for aesthetic delivery workflows.", "Vacuum Mesotherapy Injector, Electronic Injection Gun"),
}

DUPLICATE_FIXES = {
    "REC_0898": "REC_0668",
    "REC_0385": "REC_0384",
    "REC_0949": "REC_0543",
    "REC_0377": "REC_0376",
    "REC_0144": "REC_0143",
    "REC_0729": "REC_0727",
    "REC_0978": "REC_0415",
    "REC_0357": "REC_0355",
    "REC_0514": "REC_0515",
    "REC_0188": "REC_0187",
}


def locate_source_workbook() -> Path:
    return find_file(SOURCE_PATTERN)


def backup_workbook(source: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = source.with_name(f"{source.stem}_backup_before_integrity_cleanup_{timestamp}{source.suffix}")
    shutil.copy2(source, backup)
    return backup


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        norm(cell.value): index
        for index, cell in enumerate(ws[1], start=1)
        if norm(cell.value)
    }


def ensure_column(ws: openpyxl.worksheet.worksheet.Worksheet, name: str) -> int:
    colmap = headers(ws)
    if name in colmap:
        return colmap[name]
    col = ws.max_column + 1
    ws.cell(1, col).value = name
    return col


def cell_value(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, colmap: dict[str, int], field: str) -> Any:
    col = colmap.get(field)
    return ws.cell(row, col).value if col else None


def log_change(
    changes: list[dict[str, Any]],
    sheet: str,
    row_num: int | str,
    row_id: Any,
    field: str,
    old_value: Any,
    new_value: Any,
    action: str,
    note: str = "",
) -> None:
    if norm(old_value) == norm(new_value):
        return
    changes.append(
        {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "sheet": sheet,
            "row_num": row_num,
            "row_id": norm(row_id),
            "field": field,
            "old_value": old_value,
            "new_value": new_value,
            "action": action,
            "note": note,
        }
    )


def set_cell(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    colmap: dict[str, int],
    field: str,
    value: Any,
    changes: list[dict[str, Any]],
    action: str,
    row_id: Any | None = None,
    note: str = "",
) -> bool:
    col = colmap[field]
    old_value = ws.cell(row, col).value
    if norm(old_value) == norm(value):
        return False
    ws.cell(row, col).value = value
    log_change(changes, ws.title, row, row_id or cell_value(ws, row, colmap, "Record_ID") or row, field, old_value, value, action, note)
    return True


def canonical_company(value: Any) -> str:
    text = norm(value)
    if compact_key(text) == "deka":
        return "DEKA"
    return text


def cleanup_product_lines(wb: openpyxl.Workbook, changes: list[dict[str, Any]]) -> dict[str, Any]:
    ws = wb["Product_Lines"]
    ensure_column(ws, "Product_UUID")
    ensure_column(ws, "Is_Primary_Record")
    colmap = headers(ws)
    stats: Counter[str] = Counter()

    for row in range(2, ws.max_row + 1):
        record_id = norm(cell_value(ws, row, colmap, "Record_ID")) or f"row_{row}"
        company = canonical_company(cell_value(ws, row, colmap, "Company"))
        brand = norm(cell_value(ws, row, colmap, "Brand"))
        core_product = norm(cell_value(ws, row, colmap, "Core_Product"))
        category_l2 = norm(cell_value(ws, row, colmap, "Category_L2"))
        tech_type = norm(cell_value(ws, row, colmap, "Tech_Type_Std"))
        tech_original = norm(cell_value(ws, row, colmap, "Tech_Type_Original"))

        if company and company != norm(cell_value(ws, row, colmap, "Company")):
            if set_cell(ws, row, colmap, "Company", company, changes, "canonicalize_company_name", record_id):
                stats["company_names"] += 1

        if not core_product and brand:
            if set_cell(ws, row, colmap, "Core_Product", brand, changes, "fill_core_product_from_brand", record_id):
                core_product = brand
                stats["core_product"] += 1

        if not tech_type and tech_original:
            if set_cell(ws, row, colmap, "Tech_Type_Std", tech_original, changes, "fill_tech_type_from_original", record_id):
                tech_type = tech_original
                stats["tech_type"] += 1

        brand_type = norm(cell_value(ws, row, colmap, "Brand_Type"))
        if brand_type.lower() == "corporate":
            if set_cell(ws, row, colmap, "Is_Primary_Record", 0, changes, "mark_corporate_row_non_primary", record_id):
                stats["corporate_non_primary"] += 1

        product_uuid = norm(cell_value(ws, row, colmap, "Product_UUID"))
        if not product_uuid:
            new_uuid = stable_id("prod", record_id, company, brand, core_product or brand, category_l2, tech_type)
            if set_cell(ws, row, colmap, "Product_UUID", new_uuid, changes, "fill_stable_product_uuid", record_id):
                stats["product_uuid"] += 1

    return dict(stats)


def product_stats(wb: openpyxl.Workbook) -> dict[str, dict[str, Any]]:
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    stats: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "products": set(),
            "brands": set(),
            "countries": set(),
            "tracks": Counter(),
            "fda_products": set(),
            "nmpa_products": set(),
        }
    )
    for row in range(2, ws.max_row + 1):
        company = canonical_company(cell_value(ws, row, colmap, "Company"))
        if not company:
            continue
        record_id = norm(cell_value(ws, row, colmap, "Record_ID")) or stable_id(
            "row",
            company,
            cell_value(ws, row, colmap, "Brand"),
            cell_value(ws, row, colmap, "Core_Product"),
        )
        stats[company]["products"].add(record_id)
        brand = norm(cell_value(ws, row, colmap, "Brand"))
        country = norm(cell_value(ws, row, colmap, "Country"))
        category_l1 = norm(cell_value(ws, row, colmap, "Category_L1"))
        if brand:
            stats[company]["brands"].add(brand)
        if country:
            stats[company]["countries"].add(country)
        if category_l1:
            stats[company]["tracks"][category_l1] += 1
        if norm(cell_value(ws, row, colmap, "FDA_Status")) or norm(cell_value(ws, row, colmap, "FDA_510k_Number")):
            stats[company]["fda_products"].add(record_id)
        if norm(cell_value(ws, row, colmap, "NMPA_Status")) or norm(cell_value(ws, row, colmap, "NMPA_Reg_Number")):
            stats[company]["nmpa_products"].add(record_id)
    return stats


def merge_duplicate_companies(wb: openpyxl.Workbook, changes: list[dict[str, Any]]) -> int:
    ws = wb["Companies"]
    colmap = headers(ws)
    company_col = colmap["Company"]

    for row in range(2, ws.max_row + 1):
        old_company = ws.cell(row, company_col).value
        new_company = canonical_company(old_company)
        if new_company and norm(old_company) != new_company:
            set_cell(ws, row, colmap, "Company", new_company, changes, "canonicalize_company_name", new_company)

    colmap = headers(ws)
    rows_by_key: dict[str, list[int]] = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        key = compact_key(cell_value(ws, row, colmap, "Company"))
        if key:
            rows_by_key[key].append(row)

    duplicate_rows: list[int] = []
    for key, rows in rows_by_key.items():
        if len(rows) <= 1:
            continue
        target = min(rows)
        for duplicate in sorted((row for row in rows if row != target), reverse=True):
            for col in range(1, ws.max_column + 1):
                target_cell = ws.cell(target, col)
                duplicate_cell = ws.cell(duplicate, col)
                if not norm(target_cell.value) and norm(duplicate_cell.value):
                    target_cell.value = duplicate_cell.value
                    log_change(
                        changes,
                        "Companies",
                        target,
                        cell_value(ws, target, colmap, "Company"),
                        norm(ws.cell(1, col).value),
                        "",
                        duplicate_cell.value,
                        "merge_duplicate_company_row",
                        f"merged from row {duplicate}",
                    )
            duplicate_rows.append(duplicate)
            log_change(
                changes,
                "Companies",
                duplicate,
                cell_value(ws, duplicate, colmap, "Company"),
                "__row__",
                "duplicate row",
                "deleted after merge",
                "merge_duplicate_company_row",
                f"merged into row {target}",
            )

    for row in sorted(set(duplicate_rows), reverse=True):
        ws.delete_rows(row, 1)

    return len(set(duplicate_rows))


def refresh_company_stats(wb: openpyxl.Workbook, changes: list[dict[str, Any]]) -> int:
    ws = wb["Companies"]
    for field in [
        "Countries_All",
        "Country_Count",
        "Product_Count",
        "Brand_Count",
        "Primary_Track",
        "FDA_Products",
        "NMPA_Products",
    ]:
        ensure_column(ws, field)
    colmap = headers(ws)
    stats = product_stats(wb)
    changed = 0

    for row in range(2, ws.max_row + 1):
        company = canonical_company(cell_value(ws, row, colmap, "Company"))
        if not company or company not in stats:
            continue
        record = stats[company]
        track = record["tracks"].most_common(1)[0][0] if record["tracks"] else ""
        updates = {
            "Product_Count": len(record["products"]),
            "Brand_Count": len(record["brands"]),
            "Countries_All": ", ".join(sorted(record["countries"])),
            "Country_Count": len(record["countries"]),
            "Primary_Track": track,
            "FDA_Products": len(record["fda_products"]),
            "NMPA_Products": len(record["nmpa_products"]),
        }
        for field, value in updates.items():
            if set_cell(ws, row, colmap, field, value, changes, "refresh_company_stats_from_product_lines", company):
                changed += 1
    return changed


def rebuild_brand_portfolio(wb: openpyxl.Workbook, changes: list[dict[str, Any]]) -> dict[str, int]:
    product_ws = wb["Product_Lines"]
    product_cols = headers(product_ws)
    brand_ws = wb["Brand_Portfolio"]
    desired_headers = [
        "Company",
        "Brand",
        "Country",
        "Category_L1",
        "Category_L2",
        "Tech_Type",
        "Brand_Type",
        "Product_Count",
        "Products",
    ]
    old_rows = max(brand_ws.max_row - 1, 0)
    if brand_ws.max_row > 1:
        brand_ws.delete_rows(2, brand_ws.max_row - 1)
    for col, header in enumerate(desired_headers, start=1):
        brand_ws.cell(1, col).value = header

    groups: dict[tuple[str, ...], set[str]] = defaultdict(set)
    for row in range(2, product_ws.max_row + 1):
        company = canonical_company(cell_value(product_ws, row, product_cols, "Company"))
        brand = norm(cell_value(product_ws, row, product_cols, "Brand"))
        if not company or not brand:
            continue
        key = (
            company,
            brand,
            norm(cell_value(product_ws, row, product_cols, "Country")),
            norm(cell_value(product_ws, row, product_cols, "Category_L1")),
            norm(cell_value(product_ws, row, product_cols, "Category_L2")),
            norm(cell_value(product_ws, row, product_cols, "Tech_Type_Std")),
            norm(cell_value(product_ws, row, product_cols, "Brand_Type")),
        )
        product_name = norm(cell_value(product_ws, row, product_cols, "Core_Product")) or brand
        groups[key].add(product_name)

    for key in sorted(groups):
        products = sorted(groups[key])
        brand_ws.append([*key, len(products), "; ".join(products)])

    log_change(
        changes,
        "Brand_Portfolio",
        "aggregate",
        "Brand_Portfolio",
        "__sheet__",
        old_rows,
        len(groups),
        "rebuild_brand_portfolio_from_product_lines",
    )
    return {"old_rows": old_rows, "new_rows": len(groups)}


def extend_category_definitions(wb: openpyxl.Workbook, changes: list[dict[str, Any]]) -> int:
    ws = wb["Category_Definitions"]
    for field in ["Category_L1", "Category_L2", "Definition_EN", "Tech_Type_Examples"]:
        ensure_column(ws, field)
    colmap = headers(ws)
    existing = {
        (
            norm(cell_value(ws, row, colmap, "Category_L1")),
            norm(cell_value(ws, row, colmap, "Category_L2")),
        )
        for row in range(2, ws.max_row + 1)
    }

    product_ws = wb["Product_Lines"]
    product_cols = headers(product_ws)
    examples: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    for row in range(2, product_ws.max_row + 1):
        pair = (
            norm(cell_value(product_ws, row, product_cols, "Category_L1")),
            norm(cell_value(product_ws, row, product_cols, "Category_L2")),
        )
        if not all(pair) or pair in existing:
            continue
        tech = norm(cell_value(product_ws, row, product_cols, "Tech_Type_Std"))
        if tech:
            examples[pair][tech] += 1

    added = 0
    for pair in sorted(examples):
        tech_examples = ", ".join(tech for tech, _ in examples[pair].most_common(5))
        ws.append([pair[0], pair[1], "Needs definition", tech_examples])
        added += 1

    if added:
        log_change(
            changes,
            "Category_Definitions",
            "aggregate",
            "Category_Definitions",
            "__rows__",
            len(existing),
            len(existing) + added,
            "add_missing_category_paths_from_product_lines",
            "Definition_EN is a placeholder requiring domain review.",
        )
    return added


def style_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, min(ws.max_row, 200) + 1):
            max_len = max(max_len, len(norm(ws.cell(row, col).value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def replace_sheet(wb: openpyxl.Workbook, title: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def write_cleanup_log_sheet(wb: openpyxl.Workbook, changes: list[dict[str, Any]]) -> None:
    ws = replace_sheet(wb, "Cleanup_Log")
    headers_row = [
        "timestamp",
        "sheet",
        "row_num",
        "row_id",
        "field",
        "old_value",
        "new_value",
        "action",
        "note",
    ]
    ws.append(headers_row)
    for change in changes:
        ws.append([change.get(header, "") for header in headers_row])
    style_sheet(ws)


def write_seed_integrity_sheet(wb: openpyxl.Workbook) -> int:
    ws = replace_sheet(wb, "Seed_Integrity_Issues")
    if not DATA_QUALITY_ISSUES_PATH.exists():
        ws.append(["status", "message"])
        ws.append(["missing", str(DATA_QUALITY_ISSUES_PATH)])
        style_sheet(ws)
        return 0

    with DATA_QUALITY_ISSUES_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        ws.append(fieldnames)
        count = 0
        for row in reader:
            ws.append([row.get(field, "") for field in fieldnames])
            count += 1
    style_sheet(ws)
    return count


def write_csv_sheet(wb: openpyxl.Workbook, title: str, path: Path) -> int:
    ws = replace_sheet(wb, title)
    if not path.exists():
        ws.append(["status", "message"])
        ws.append(["missing", str(path)])
        style_sheet(ws)
        return 0

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        ws.append(fieldnames)
        count = 0
        for row in reader:
            ws.append([excel_safe(row.get(field, "")) for field in fieldnames])
            count += 1
    style_sheet(ws)
    return count


def write_jsonl_sheet(wb: openpyxl.Workbook, title: str, path: Path) -> int:
    ws = replace_sheet(wb, title)
    if not path.exists():
        ws.append(["status", "message"])
        ws.append(["missing", str(path)])
        style_sheet(ws)
        return 0

    rows: list[dict[str, Any]] = []
    fieldnames: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        flat = {
            key: json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
            for key, value in row.items()
        }
        rows.append(flat)
        for key in flat:
            if key not in fieldnames:
                fieldnames.append(key)
    ws.append(fieldnames)
    for row in rows:
        ws.append([excel_safe(row.get(field, "")) for field in fieldnames])
    style_sheet(ws)
    return len(rows)


def excel_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def save_change_files(changes: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    LOG_PATH.write_text(json.dumps(changes, ensure_ascii=False, indent=2), encoding="utf-8")
    SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def load_changes() -> list[dict[str, Any]]:
    if not LOG_PATH.exists():
        return []
    return json.loads(LOG_PATH.read_text(encoding="utf-8"))


def run_cleanup(no_backup: bool = False) -> dict[str, Any]:
    source = locate_source_workbook()
    backup = None if no_backup else backup_workbook(source)
    wb = openpyxl.load_workbook(source, data_only=False)
    changes: list[dict[str, Any]] = []

    summary = {
        "source_workbook": str(source),
        "backup_workbook": str(backup) if backup else "",
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "product_lines": cleanup_product_lines(wb, changes),
    }
    summary["duplicate_company_rows_removed"] = merge_duplicate_companies(wb, changes)
    summary["company_stat_cells_refreshed"] = refresh_company_stats(wb, changes)
    summary["brand_portfolio"] = rebuild_brand_portfolio(wb, changes)
    summary["category_paths_added"] = extend_category_definitions(wb, changes)
    summary["change_count"] = len(changes)
    summary["finished_at"] = datetime.now().isoformat(timespec="seconds")

    write_cleanup_log_sheet(wb, changes)
    wb.save(source)
    save_change_files(changes, summary)
    return summary


def sync_audit_sheet() -> dict[str, Any]:
    source = locate_source_workbook()
    wb = openpyxl.load_workbook(source, data_only=False)
    changes = load_changes()
    write_cleanup_log_sheet(wb, changes)
    issue_count = write_seed_integrity_sheet(wb)
    wb.save(source)
    return {
        "source_workbook": str(source),
        "cleanup_log_rows": len(changes),
        "seed_integrity_issue_rows": issue_count,
    }


def sync_hierarchy_sheets(no_backup: bool = False) -> dict[str, Any]:
    source = locate_source_workbook()
    backup = None if no_backup else backup_workbook(source)
    wb = openpyxl.load_workbook(source, data_only=False)
    family_rows = write_csv_sheet(wb, "Product_Family_Master", PRODUCT_FAMILY_MASTER_PATH)
    sku_rows = write_csv_sheet(wb, "Product_SKU_Master", PRODUCT_SKU_MASTER_PATH)
    wb.save(source)
    return {
        "source_workbook": str(source),
        "backup_workbook": str(backup) if backup else "",
        "product_family_rows": family_rows,
        "product_sku_rows": sku_rows,
    }


def sync_background_sheets(no_backup: bool = False) -> dict[str, Any]:
    source = locate_source_workbook()
    backup = None if no_backup else backup_workbook(source)
    wb = openpyxl.load_workbook(source, data_only=False)
    legacy_long_sheet = "Company_Official_Source_Evidence"
    if legacy_long_sheet in wb.sheetnames:
        del wb[legacy_long_sheet]
    evidence_rows = write_jsonl_sheet(wb, "Company_Background_Evidence", COMPANY_BACKGROUND_EVIDENCE_PATH)
    capital_rows = write_csv_sheet(wb, "Company_Capital_Structure", COMPANY_CAPITAL_STRUCTURE_PATH)
    listed_rows = write_csv_sheet(wb, "Listed_Company_Batch", LISTED_COMPANY_BATCH_PATH)
    policy_rows = write_csv_sheet(wb, "Source_Authority_Policy", SOURCE_AUTHORITY_POLICY_PATH)
    official_plan_rows = write_csv_sheet(wb, "Company_Official_Source_Plan", COMPANY_OFFICIAL_SOURCE_PLAN_PATH)
    official_evidence_rows = write_jsonl_sheet(wb, "Company_Official_Evidence", COMPANY_OFFICIAL_SOURCE_EVIDENCE_PATH)
    product_master_rows = write_csv_sheet(wb, "Product_Master", PRODUCT_MASTER_PATH)
    registration_evidence_rows = write_csv_sheet(wb, "Registration_Evidence", REGISTRATION_EVIDENCE_PATH)
    official_website_rows = write_csv_sheet(wb, "Official_Website_Master", OFFICIAL_WEBSITE_MASTER_PATH)
    company_website_rows = write_csv_sheet(wb, "Company_Official_Website", COMPANY_OFFICIAL_WEBSITE_PATH)
    media_asset_rows = write_csv_sheet(wb, "Media_Asset_Index", COMPANY_MEDIA_ASSET_INDEX_PATH)
    product_spec_rows = write_csv_sheet(wb, "Product_Spec_Evidence", PRODUCT_SPECIFICATION_EVIDENCE_PATH)
    regulatory_plan_rows = write_csv_sheet(wb, "Policy_Regulatory_Source_Plan", POLICY_REGULATORY_SOURCE_PLAN_PATH)
    valuation_rank_rows = write_csv_sheet(wb, "Market_Valuation_Rank", MARKET_VALUATION_RANK_PATH)
    market_conflict_rows = write_csv_sheet(wb, "Market_Conflicts", MARKET_CONFLICT_PATH)
    official_coverage_rows = write_csv_sheet(wb, "Official_Source_Coverage", OFFICIAL_SOURCE_COVERAGE_PATH)
    source_diff_rows = write_csv_sheet(wb, "Source_Diff_Report", SOURCE_DIFF_REPORT_PATH)
    mdr_ce_evidence_rows = write_jsonl_sheet(wb, "MDR_CE_Evidence", MDR_CE_EVIDENCE_PATH)
    evidence_promotion_rows = write_csv_sheet(wb, "Evidence_Promotion_Log", EVIDENCE_PROMOTION_LOG_PATH)
    official_indication_rows = write_csv_sheet(wb, "Official_Indication_Evidence", OFFICIAL_INDICATION_EVIDENCE_PATH)
    xueqiu_market_rows = write_csv_sheet(wb, "Xueqiu_Market_Check", XUEQIU_MARKET_CHECK_PATH)
    wb.save(source)
    return {
        "source_workbook": str(source),
        "backup_workbook": str(backup) if backup else "",
        "company_background_evidence_rows": evidence_rows,
        "company_capital_structure_rows": capital_rows,
        "listed_company_batch_rows": listed_rows,
        "source_authority_policy_rows": policy_rows,
        "company_official_source_plan_rows": official_plan_rows,
        "company_official_source_evidence_rows": official_evidence_rows,
        "product_master_rows": product_master_rows,
        "registration_evidence_rows": registration_evidence_rows,
        "official_website_master_rows": official_website_rows,
        "company_official_website_rows": company_website_rows,
        "media_asset_rows": media_asset_rows,
        "product_specification_rows": product_spec_rows,
        "policy_regulatory_source_plan_rows": regulatory_plan_rows,
        "market_valuation_rank_rows": valuation_rank_rows,
        "market_conflict_rows": market_conflict_rows,
        "official_source_coverage_rows": official_coverage_rows,
        "source_diff_report_rows": source_diff_rows,
        "mdr_ce_evidence_rows": mdr_ce_evidence_rows,
        "evidence_promotion_log_rows": evidence_promotion_rows,
        "official_indication_evidence_rows": official_indication_rows,
        "xueqiu_market_check_rows": xueqiu_market_rows,
    }


def sync_ce_plan_sheet(no_backup: bool = False) -> dict[str, Any]:
    source = locate_source_workbook()
    backup = None if no_backup else backup_workbook(source)
    wb = openpyxl.load_workbook(source, data_only=False)
    rows = write_csv_sheet(wb, "MDR_CE_Search_Plan", MDR_CE_SEARCH_PLAN_PATH)
    wb.save(source)
    return {
        "source_workbook": str(source),
        "backup_workbook": str(backup) if backup else "",
        "mdr_ce_search_plan_rows": rows,
    }


def resolve_remaining_issues(no_backup: bool = False) -> dict[str, Any]:
    source = locate_source_workbook()
    backup = None if no_backup else backup_workbook(source)
    wb = openpyxl.load_workbook(source, data_only=False)
    changes: list[dict[str, Any]] = []

    product_ws = wb["Product_Lines"]
    ensure_column(product_ws, "Duplicate_Note")
    ensure_column(product_ws, "Is_Primary_Record")
    product_cols = headers(product_ws)
    tech_filled = 0
    duplicates_marked = 0
    for row in range(2, product_ws.max_row + 1):
        record_id = norm(cell_value(product_ws, row, product_cols, "Record_ID"))
        if record_id in TECH_TYPE_FIXES and not norm(cell_value(product_ws, row, product_cols, "Tech_Type_Std")):
            if set_cell(
                product_ws,
                row,
                product_cols,
                "Tech_Type_Std",
                TECH_TYPE_FIXES[record_id],
                changes,
                "resolve_missing_tech_type",
                record_id,
                "Rule-based cleanup from product name, category and introduction.",
            ):
                tech_filled += 1
        if record_id in DUPLICATE_FIXES:
            target = DUPLICATE_FIXES[record_id]
            note = f"duplicate_of:{target}; marked non-primary during seed integrity cleanup"
            if set_cell(product_ws, row, product_cols, "Is_Primary_Record", 0, changes, "mark_reviewed_duplicate_non_primary", record_id, note):
                duplicates_marked += 1
            set_cell(product_ws, row, product_cols, "Duplicate_Note", note, changes, "mark_reviewed_duplicate_non_primary", record_id, note)

    category_ws = wb["Category_Definitions"]
    for field in ["Category_L1", "Category_L2", "Definition_EN", "Tech_Type_Examples"]:
        ensure_column(category_ws, field)
    category_cols = headers(category_ws)
    existing = {
        (
            norm(cell_value(category_ws, row, category_cols, "Category_L1")),
            norm(cell_value(category_ws, row, category_cols, "Category_L2")),
        )
        for row in range(2, category_ws.max_row + 1)
    }
    categories_added = 0
    for pair, (definition, examples) in sorted(CATEGORY_DEFINITION_FIXES.items()):
        if pair in existing:
            continue
        category_ws.append([pair[0], pair[1], definition, examples])
        categories_added += 1
        log_change(
            changes,
            "Category_Definitions",
            category_ws.max_row,
            " > ".join(pair),
            "__row__",
            "",
            "added",
            "resolve_missing_category_path",
            "Added dictionary path from reviewed seed issue.",
        )

    all_changes = load_changes() + changes
    write_cleanup_log_sheet(wb, all_changes)
    wb.save(source)

    summary = {
        "source_workbook": str(source),
        "backup_workbook": str(backup) if backup else "",
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "tech_types_filled": tech_filled,
        "duplicates_marked_non_primary": duplicates_marked,
        "category_paths_added": categories_added,
        "change_count": len(changes),
        "total_cleanup_log_rows": len(all_changes),
        "finished_at": datetime.now().isoformat(timespec="seconds"),
    }
    save_change_files(all_changes, summary)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply deterministic cleanup to the global aesthetics seed workbook.")
    parser.add_argument(
        "mode",
        choices=["cleanup", "resolve-remaining", "sync-audit", "sync-hierarchy", "sync-background", "sync-ce-plan"],
        help="Run cleanup, resolve deterministic issues, or sync generated audit/hierarchy sheets into the workbook.",
    )
    parser.add_argument("--no-backup", action="store_true", help="Skip creating a backup before cleanup.")
    args = parser.parse_args()

    if args.mode == "cleanup":
        result = run_cleanup(no_backup=args.no_backup)
    elif args.mode == "resolve-remaining":
        result = resolve_remaining_issues(no_backup=args.no_backup)
    elif args.mode == "sync-hierarchy":
        result = sync_hierarchy_sheets(no_backup=args.no_backup)
    elif args.mode == "sync-background":
        result = sync_background_sheets(no_backup=args.no_backup)
    elif args.mode == "sync-ce-plan":
        result = sync_ce_plan_sheet(no_backup=args.no_backup)
    else:
        result = sync_audit_sheet()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
