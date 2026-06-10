#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
PRODUCT_MASTER = DATA_DIR / "product_master.csv"

TARGET_SEED = "REC_0660"
TARGET_NOTE = (
    "excluded_by_scope_20260601: Bayer Skinoren is an azelaic-acid prescription dermatology product "
    "for acne/rosacea, not an upstream medical-aesthetics product line under the current scope rules."
)

SUMMARY_JSON = AUDIT_DIR / "skinoren_scope_exclusion_20260601_latest.json"


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def headers(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}


def append_note(existing: str, marker: str) -> str:
    existing = clean(existing)
    if marker in existing:
        return existing
    return f"{existing} | {marker}" if existing else marker


def read_product_id() -> str:
    if not PRODUCT_MASTER.exists():
        return ""
    with PRODUCT_MASTER.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if clean(row.get("seed_record_id")) == TARGET_SEED:
                return clean(row.get("product_id"))
    return ""


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_skinoren_scope_exclusion_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)

    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    record_col = colmap["Record_ID"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if clean(ws.cell(row=row_idx, column=record_col).value) == TARGET_SEED:
            target_row = row_idx
            break

    changes: list[dict[str, str]] = []
    if target_row is not None:
        updates = {
            "Inclusion_Status": "excluded",
            "Duplicate_Note": TARGET_NOTE,
            "V4_1_Registration_Review_Status": "excluded_scope",
        }
        for field, value in updates.items():
            col = colmap.get(field)
            if not col:
                continue
            old = clean(ws.cell(row=target_row, column=col).value)
            new = append_note(old, value) if field == "Duplicate_Note" else value
            if old == new:
                continue
            ws.cell(row=target_row, column=col).value = new
            changes.append({"field": field, "old": old, "new": new})
        wb.save(SOURCE_BOOK)

    summary = {
        "checked_at": datetime.now().astimezone().replace(microsecond=0).isoformat(),
        "seed_record_id": TARGET_SEED,
        "product_id_before_rebuild": read_product_id(),
        "workbook_row_found": target_row is not None,
        "changes": changes,
        "backup": str(backup),
        "note": TARGET_NOTE,
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
