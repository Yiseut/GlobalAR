#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_BOOK = Path(r"E:\shared\Documents\data\全球医美企业库_标准化版v4.xlsx")
AUDIT_DIR = Path(__file__).resolve().parents[1] / "data" / "audits"


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(norm(part).lower() for part in parts if norm(part))
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12] if raw else "0" * 12
    return f"{prefix}_{digest}"


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def row_dict(ws, row_idx: int, colmap: dict[str, int]) -> dict[str, Any]:
    return {field: ws.cell(row=row_idx, column=col).value for field, col in colmap.items()}


def set_values(ws, row_idx: int, colmap: dict[str, int], updates: dict[str, Any], changes: list[dict[str, Any]]) -> None:
    record_id = norm(ws.cell(row=row_idx, column=colmap["Record_ID"]).value)
    for field, value in updates.items():
        if field not in colmap:
            continue
        cell = ws.cell(row=row_idx, column=colmap[field])
        old_value = cell.value
        if old_value == value:
            continue
        cell.value = value
        changes.append(
            {
                "record_id": record_id,
                "field": field,
                "old": old_value,
                "new": value,
            }
        )


def taxonomy(l1: str, l2: str, l3: str, source: str, confidence: str, note: str) -> dict[str, str]:
    return {
        "Material_Taxonomy_L1_CN": l1,
        "Material_Taxonomy_L2_CN": l2,
        "Material_Taxonomy_L3_CN": l3,
        "Material_Taxonomy_Path_CN": " > ".join(part for part in [l1, l2, l3] if part),
        "Material_Taxonomy_Source": source,
        "Material_Taxonomy_Confidence": confidence,
        "Material_Taxonomy_Review_Status": "auto_applied" if confidence == "high" else "needs_review",
        "Material_Taxonomy_Note": note,
    }


CAHA_TAXONOMY = taxonomy(
    "注射类",
    "胶原刺激剂",
    "CaHA",
    "manual_decision:caha_boundary_20260527",
    "high",
    "人工复核确认：FACETEM 和 Neauvia Stimulate 按 CaHA/HA+CaHA 复合胶原刺激剂处理。",
)
HA_TAXONOMY = taxonomy(
    "注射类",
    "透明质酸 HA",
    "交联HA（填充剂形态）",
    "manual_decision:matex_lab_split_20260527",
    "high",
    "人工复核确认：纯 HA 与 PEG 交联 HA 仍归入 HA 赛道。",
)
PLASMA_TAXONOMY = taxonomy(
    "能量设备",
    "等离子",
    "等离子束 / 等离子笔",
    "manual_decision:neauvia_devices_20260527",
    "high",
    "该行是 Neauvia Devices/Sectum/Plasma 能量设备，不能因配合 Neauvia HA 填充剂使用而归入 HA。",
)


def next_record_id(ws, colmap: dict[str, int]) -> str:
    max_num = 0
    for row in range(2, ws.max_row + 1):
        value = norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
        match = re.fullmatch(r"REC_(\d+)", value)
        if match:
            max_num = max(max_num, int(match.group(1)))
    return f"REC_{max_num + 1:04d}"


def update_brand_portfolio(wb, changes: list[dict[str, Any]]) -> None:
    if "Brand_Portfolio" not in wb.sheetnames:
        return
    ws = wb["Brand_Portfolio"]
    colmap = headers(ws)
    required = {"Company", "Brand", "Product_Count", "Products", "Tech_Type", "Category_L1", "Category_L2"}
    if not required.issubset(colmap):
        return
    for row in range(2, ws.max_row + 1):
        if norm(ws.cell(row=row, column=colmap["Company"]).value) == "Matex Lab" and norm(ws.cell(row=row, column=colmap["Brand"]).value) == "Neauvia Organic":
            updates = {
                "Product_Count": 2,
                "Products": "Intense / PEG-crosslinked HA Filler; Stimulate (PEG-HA + 1% CaHA)",
                "Tech_Type": "Hyaluronic Acid; Calcium Hydroxylapatite",
                "Category_L1": "Injectables",
                "Category_L2": "Dermal Filler; Biostimulator",
            }
            for field, value in updates.items():
                cell = ws.cell(row=row, column=colmap[field])
                if cell.value != value:
                    changes.append({"record_id": "Brand_Portfolio:Matex Lab:Neauvia Organic", "field": field, "old": cell.value, "new": value})
                    cell.value = value
            return


def update_company_counts(wb, changes: list[dict[str, Any]]) -> None:
    if "Companies" not in wb.sheetnames:
        return
    ws = wb["Companies"]
    colmap = headers(ws)
    if not {"Company", "Product_Count", "Brand_Count"}.issubset(colmap):
        return
    for row in range(2, ws.max_row + 1):
        if norm(ws.cell(row=row, column=colmap["Company"]).value) == "Matex Lab":
            updates = {"Product_Count": 3, "Brand_Count": 2}
            for field, value in updates.items():
                cell = ws.cell(row=row, column=colmap[field])
                if cell.value != value:
                    changes.append({"record_id": "Companies:Matex Lab", "field": field, "old": cell.value, "new": value})
                    cell.value = value
            return


def append_manual_fact_evidence(wb, changes: list[dict[str, Any]]) -> None:
    if "Manual_Product_Fact_Evidence" not in wb.sheetnames:
        return
    ws = wb["Manual_Product_Fact_Evidence"]
    colmap = headers(ws)
    required = {
        "fact_id",
        "product_id",
        "seed_record_id",
        "company_id",
        "company",
        "brand",
        "product_family_id",
        "standard_product_name",
        "priority",
        "fact_group",
        "field_name",
        "field_value",
        "source_url",
        "evidence_title",
        "evidence_excerpt",
        "source_type",
        "confidence",
        "captured_at",
        "promoted_at",
        "review_status",
        "note",
    }
    if not required.issubset(colmap):
        return

    existing = {
        norm(ws.cell(row=row, column=colmap["fact_id"]).value)
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["fact_id"]).value)
    }
    captured_at = datetime.now().astimezone().isoformat(timespec="seconds")
    facts = [
        {
            "seed_record_id": "REC_0479",
            "product_id": "prod_1d51b3a7d560",
            "product_family_id": stable_id("pf", "Matex Lab", "Neauvia Organic", "Intense / PEG-crosslinked HA Filler", "Injectables", "Dermal Filler", "Hyaluronic Acid"),
            "company_id": stable_id("co", "Matex Lab"),
            "company": "Matex Lab",
            "brand": "Neauvia Organic",
            "standard_product_name": "Intense / PEG-crosslinked HA Filler",
            "priority": "P1",
            "fact_group": "official_product_document",
            "field_name": "composition",
            "field_value": "Biodegradable Hyaluronic Acid hydrogel crosslinked with PEG; HA concentration 28 mg/ml.",
            "source_url": "https://www.neauvia.com/wp-content/uploads/2025/07/Exe_ProductFactSheet_INTENSE_210x297_LR.pdf",
            "evidence_title": "Neauvia INTENSE Product Fact Sheet",
            "evidence_excerpt": "Intense is a biodegradable Hyaluronic Acid hydrogel crosslinked with PEG. It is intended to restore lost volume of soft tissue; the fact sheet lists crosslinking as PEG and HA concentration as 28 mg/ml.",
            "source_type": "official_product_document",
            "confidence": "product_official_domain_candidate",
            "review_status": "manual_cross_checked",
            "note": "Added during manual decision pass to separate HA/PEG-HA from Neauvia Stimulate.",
        },
        {
            "seed_record_id": "REC_1023",
            "product_id": stable_id("prod", "REC_1023", "Matex Lab", "Neauvia Organic", "Stimulate (PEG-HA + 1% CaHA)", "Biostimulator", "Calcium Hydroxylapatite"),
            "product_family_id": stable_id("pf", "Matex Lab", "Neauvia Organic", "Stimulate (PEG-HA + 1% CaHA)", "Injectables", "Biostimulator", "Calcium Hydroxylapatite"),
            "company_id": stable_id("co", "Matex Lab"),
            "company": "Matex Lab",
            "brand": "Neauvia Organic",
            "standard_product_name": "Stimulate (PEG-HA + 1% CaHA)",
            "priority": "P0",
            "fact_group": "official_product_document",
            "field_name": "composition",
            "field_value": "PEG-crosslinked HA enriched with Calcium Hydroxyapatite (CaHA); marketed as STIMULATE.",
            "source_url": "https://www.neauvia.com/wp-content/uploads/2024/04/NEAUVIA-PRESS-RELEASE-STIMULATE-AMWC-2024-AWARD-1.pdf",
            "evidence_title": "Neauvia STIMULATE AMWC 2024 Press Release",
            "evidence_excerpt": "Neauvia describes STIMULATE as a PEG-HA + CaHA filler made of Hyaluronic Acid cross-linked with PEG and enriched with Calcium Hydroxyapatite. The legal note identifies it as a class III device under EU MDR and lists MATEX LAB SPA as manufacturer.",
            "source_type": "official_product_document",
            "confidence": "product_official_domain_candidate",
            "review_status": "manual_cross_checked",
            "note": "Added during manual decision pass after splitting Neauvia Organic Stimulate into a standalone HA + CaHA line.",
        },
    ]
    for fact in facts:
        fact_id = stable_id("pfact", fact["seed_record_id"], fact["field_name"], fact["source_url"])
        if fact_id in existing:
            continue
        row_idx = ws.max_row + 1
        record = {
            "fact_id": fact_id,
            "captured_at": captured_at,
            "promoted_at": captured_at,
            **fact,
        }
        for field, col in colmap.items():
            ws.cell(row=row_idx, column=col, value=record.get(field, ""))
        existing.add(fact_id)
        changes.append({"record_id": fact["seed_record_id"], "field": "Manual_Product_Fact_Evidence", "old": "", "new": fact_id})


def main() -> None:
    if not SOURCE_BOOK.exists():
        raise FileNotFoundError(SOURCE_BOOK)

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_manual_decisions_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup_path)

    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, Any]] = []

    facetem_updates = {
        "Category_L1": "Injectables",
        "Category_L2": "Dermal Filler",
        "Tech_Type_Std": "Calcium Hydroxylapatite",
        "Tech_Type_Original": "Calcium hydroxylapatite filler (CaHA)",
        "Verified_Product_Type_CN": "CaHA 胶原刺激剂",
        **CAHA_TAXONOMY,
    }
    if "REC_0266" in row_by_id:
        set_values(
            ws,
            row_by_id["REC_0266"],
            colmap,
            {
                **facetem_updates,
                "Core_Product": "FACETEM / FACETEM S CaHA Filler",
                "Feature_Tags": "caha, calcium-hydroxylapatite, collagen-stimulator, dermal-filler",
                "Introduction": "CGBIO 官方产品页将 FACETEM/FACETEM S 标注为 Calcium Hydroxylapatite (CaHA) dermal filler；材料归属按 CaHA/胶原刺激剂处理。",
            },
            changes,
        )
    if "REC_0960" in row_by_id:
        set_values(
            ws,
            row_by_id["REC_0960"],
            colmap,
            {
                **facetem_updates,
                "Core_Product": "FACETEM CaHA Filler",
                "Feature_Tags": "caha, calcium-hydroxylapatite, collagen-stimulator, cgbio-daewoong",
                "Is_Primary_Record": False,
                "Duplicate_Note": "duplicate_of:REC_0266; FACETEM is a CGBIO/Daewoong-affiliate CaHA line; old HA label was corrected.",
                "Introduction": "FACETEM 属于 CGBIO/Daewoong 体系的 Calcium Hydroxylapatite (CaHA) 产品线；源行原 HA 描述已按人工复核改正，并作为 REC_0266 的非主重复线索保留。",
            },
            changes,
        )

    if "REC_0478" in row_by_id:
        set_values(
            ws,
            row_by_id["REC_0478"],
            colmap,
            {
                "Category_L1": "EBD",
                "Category_L2": "Plasma",
                "Tech_Type_Std": "Plasma",
                "Tech_Type_Original": "RF / Plasma energy device",
                "Verified_Product_Type_CN": "等离子/RF 能量设备",
                **PLASMA_TAXONOMY,
            },
            changes,
        )

    if "REC_0479" not in row_by_id:
        raise RuntimeError("REC_0479 not found; cannot split Neauvia Organic safely.")
    set_values(
        ws,
        row_by_id["REC_0479"],
        colmap,
        {
            "Core_Product": "Intense / PEG-crosslinked HA Filler",
            "Tech_Type_Std": "Hyaluronic Acid",
            "Tech_Type_Original": "PEG-crosslinked HA filler",
            "Feature_Tags": "ha, peg-crosslinked, dermal-filler",
            "Verified_Product_Type_CN": "PEG 交联透明质酸填充剂",
            "Introduction": "Neauvia Organic Intense 等纯 HA / PEG 交联 HA 产品归入 HA 赛道；含 1% CaHA 的 Stimulate 已拆分为独立产品线。",
            **HA_TAXONOMY,
        },
        changes,
    )

    existing_stimulate = False
    for row in range(2, ws.max_row + 1):
        company = norm(ws.cell(row=row, column=colmap["Company"]).value)
        brand = norm(ws.cell(row=row, column=colmap["Brand"]).value)
        core = norm(ws.cell(row=row, column=colmap["Core_Product"]).value).lower()
        tech = norm(ws.cell(row=row, column=colmap["Tech_Type_Std"]).value).lower()
        if company == "Matex Lab" and brand == "Neauvia Organic" and "stimulate" in core and "calcium" in tech:
            existing_stimulate = True
            break

    added_record_id = ""
    if not existing_stimulate:
        source_row = row_dict(ws, row_by_id["REC_0479"], colmap)
        added_record_id = next_record_id(ws, colmap)
        new_row_idx = ws.max_row + 1
        for field, col in colmap.items():
            ws.cell(row=new_row_idx, column=col, value=source_row.get(field))
        new_core = "Stimulate (PEG-HA + 1% CaHA)"
        new_updates = {
            "Record_ID": added_record_id,
            "Company": "Matex Lab",
            "Country": "Italy",
            "Region": "Europe",
            "Location_Full": "Brindisi, Italy",
            "Ownership": "Private",
            "Business_Role": "Manufacturer",
            "Status": "Active",
            "Category_L1": "Injectables",
            "Category_L2": "Biostimulator",
            "Tech_Type_Std": "Calcium Hydroxylapatite",
            "Brand": "Neauvia Organic",
            "Brand_Type": "Product",
            "Core_Product": new_core,
            "CE_Status": "EU MDR Class III medical device (official Neauvia statement)",
            "Feature_Tags": "ha-caha-hybrid, peg-crosslinked, biostimulator, skinbooster-use, skin-quality",
            "Is_Primary_Record": True,
            "Product_UUID": stable_id("prod", added_record_id, "Matex Lab", "Neauvia Organic", new_core, "Biostimulator", "Calcium Hydroxylapatite"),
            "Duplicate_Note": "",
            "Tech_Type_Original": "PEG-crosslinked HA + 1% Calcium Hydroxylapatite (CaHA)",
            "Introduction": "Neauvia/Matex Lab 官方资料将 STIMULATE 描述为 PEG 交联 HA 并加入 Calcium Hydroxyapatite (CaHA) 的复合注射剂；人工复核补充：该线按水光/肤质改善打法理解，不按传统填充剂统称。",
            "Verified_Product_Type_CN": "HA + CaHA 复合胶原刺激剂（水光/肤质改善打法）",
            **CAHA_TAXONOMY,
        }
        set_values(ws, new_row_idx, colmap, new_updates, changes)
        changes.append({"record_id": added_record_id, "field": "__row__", "old": "", "new": "added Neauvia Organic Stimulate split product line"})

    update_brand_portfolio(wb, changes)
    update_company_counts(wb, changes)
    append_manual_fact_evidence(wb, changes)

    wb.save(SOURCE_BOOK)
    wb.close()

    summary = {
        "source": str(SOURCE_BOOK),
        "backup": str(backup_path),
        "changed_cells": len(changes),
        "added_record_id": added_record_id,
        "changes": changes,
    }
    summary_path = AUDIT_DIR / f"manual_decision_apply_{stamp}.json"
    latest_path = AUDIT_DIR / "manual_decision_apply_latest.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
