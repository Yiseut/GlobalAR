#!/usr/bin/env python3
"""
Build the first local data layer for the Global Medical Aesthetics dashboard.

Inputs stay in the parent data folder. This script writes:
- data/global_aesthetics.db
- web/app-data.js
- data/import_manifest.json
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import time
import urllib.parse
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from dashboard_scope import company_exclusion_reason, product_exclusion_reason


PROJECT_DIR = Path(__file__).resolve().parents[1]
SOURCE_DIR = PROJECT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
WEB_DIR = PROJECT_DIR / "web"
DB_PATH = DATA_DIR / "global_aesthetics.db"
SNAPSHOT_PATH = WEB_DIR / "app-data.js"
MANIFEST_PATH = DATA_DIR / "import_manifest.json"
STAGING_JSONL_PATH = DATA_DIR / "verification_evidence_staging.jsonl"
DATA_QUALITY_ISSUES_PATH = DATA_DIR / "seed_integrity_issues.csv"
DATA_QUALITY_REPORT_PATH = DATA_DIR / "seed_integrity_report.md"
DATA_QUALITY_SUMMARY_PATH = DATA_DIR / "seed_integrity_summary.json"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
REGISTRATION_EVIDENCE_PATH = DATA_DIR / "registration_evidence.csv"
PRODUCT_FAMILY_MASTER_PATH = DATA_DIR / "product_family_master.csv"
PRODUCT_SKU_MASTER_PATH = DATA_DIR / "product_sku_master.csv"
COMPANY_BACKGROUND_EVIDENCE_PATH = DATA_DIR / "company_background_evidence.jsonl"
COMPANY_CAPITAL_STRUCTURE_PATH = DATA_DIR / "company_capital_structure.csv"
LISTED_COMPANY_BATCH_PATH = DATA_DIR / "listed_company_batch.csv"
COMPANY_OFFICIAL_SOURCE_PLAN_PATH = DATA_DIR / "company_official_source_plan.csv"
COMPANY_OFFICIAL_SOURCE_EVIDENCE_PATH = DATA_DIR / "company_official_source_evidence.jsonl"
OFFICIAL_WEBSITE_MASTER_PATH = DATA_DIR / "official_website_master.csv"
COMPANY_OFFICIAL_WEBSITE_PATH = DATA_DIR / "company_official_website.csv"
COMPANY_MEDIA_ASSET_INDEX_PATH = DATA_DIR / "company_media_asset_index.csv"
PRODUCT_SPECIFICATION_EVIDENCE_PATH = DATA_DIR / "product_specification_evidence.csv"
POLICY_REGULATORY_SOURCE_PLAN_PATH = DATA_DIR / "policy_regulatory_source_plan.csv"
SOURCE_AUTHORITY_POLICY_PATH = DATA_DIR / "source_authority_policy.csv"
MDR_CE_SEARCH_PLAN_PATH = DATA_DIR / "mdr_ce_search_plan.csv"
MDR_CE_EVIDENCE_CANDIDATES_PATH = DATA_DIR / "mdr_ce_evidence_candidates.jsonl"
EVIDENCE_PROMOTION_LOG_PATH = DATA_DIR / "evidence_promotion_log.csv"
OFFICIAL_INDICATION_EVIDENCE_PATH = DATA_DIR / "official_indication_evidence.csv"
MANUAL_PRODUCT_FACT_EVIDENCE_PATH = DATA_DIR / "manual_product_fact_evidence.csv"
NEWS_REGULATORY_EVENT_CANDIDATES_PATH = DATA_DIR / "news_regulatory_event_candidates.csv"
BRIEFING_UPDATE_CANDIDATES_PATH = DATA_DIR / "briefing_update_candidates.csv"
BRIEFING_VERIFIED_UPDATE_EVENTS_PATH = DATA_DIR / "briefing_verified_update_events.csv"
BRIEFING_FULLTEXT_RESCUE_PATH = DATA_DIR / "briefing_fulltext_rescue.csv"
BRIEFING_PRODUCT_GAP_CANDIDATES_PATH = DATA_DIR / "audits" / "briefing_product_gap_candidates_latest.csv"
MANUAL_OFFICIAL_INDICATION_EVIDENCE_PATH = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_NMPA_REGISTRATION_EVIDENCE_PATH = DATA_DIR / "manual_nmpa_registration_evidence.csv"
MANUAL_REGISTRATION_NAME_EVIDENCE_PATH = DATA_DIR / "manual_registration_name_evidence.csv"
MANUAL_EVIDENCE_PROMOTION_LOG_PATH = DATA_DIR / "manual_evidence_promotion_log.csv"
ISAPS_MARKET_METRICS_PATH = DATA_DIR / "isaps_market_metrics.csv"
MARKET_SNAPSHOT_LIVE_PATH = DATA_DIR / "market_snapshot_live.csv"
COMPANY_FINANCIAL_METRICS_PATH = DATA_DIR / "company_financial_metrics.csv"
COMPANY_REVENUE_COLLECTION_PLAN_PATH = DATA_DIR / "audits" / "company_revenue_collection_plan_latest.csv"
FIELD_DICTIONARY_PATH = DATA_DIR / "field_dictionary.csv"
COMPANY_PORTFOLIO_CASES_PATH = DATA_DIR / "company_portfolio_cases.json"
DATA_USABILITY_LEDGER_PATH = DATA_DIR / "audits" / "data_usability_ledger_latest.csv"
DATA_USABILITY_ROW_STATUS_PATH = DATA_DIR / "audits" / "data_usability_row_status_latest.csv"
DATA_USABILITY_MISSING_OWNER_PATH = DATA_DIR / "audits" / "data_usability_missing_owner_latest.csv"
DATA_USABILITY_SUMMARY_PATH = DATA_DIR / "audits" / "data_usability_summary_latest.json"

CURRENT_PHASE_CHANNELS = {"fda", "ce", "company_official"}
EXTERNAL_PROJECT_CHANNELS = {"nmpa"}
EXTERNAL_PROJECT_NOTES = {
    "nmpa": "China NMPA full-market discovery remains in the separate China dashboard; verified rows may be imported only when matched to existing global products.",
}

PRODUCT_FACT_OVERRIDES = {
    # The old seed row was pulled into PCL because "Miracle" was treated as a
    # generic PCL keyword. This record itself is an HA dermal-filler line.
    "REC_0401": {
        "Category_L1": "Injectables",
        "Category_L2": "Dermal Filler",
        "Tech_Type_Std": "Hyaluronic Acid",
        "Tech_Type_Original": "Cross-linked hyaluronic acid",
        "Core_Product": "HA Dermal Filler",
        "Feature_Tags": "ha, dermal-filler",
        "Data_Source": "taxonomy_conflict_correction",
        "_segments": "ha",
        "_note": "Removed false PCL placement caused by a generic Miracle keyword collision.",
    },
    # Bravity was previously pulled into Dexlevo/PCL because GOURI liquid-PCL
    # pages were incorrectly attached to the row. Treat the row as topical
    # skincare unless refreshed brand-owner evidence proves a medical product.
    "REC_0096": {
        "Category_L1": "Skincare",
        "Category_L2": "Cosmeceutical",
        "Tech_Type_Std": "Topical Booster Serum",
        "Tech_Type_Original": "Topical serum / ampoule",
        "Core_Product": "Topical Serum / Ampoule",
        "Feature_Tags": "skincare, topical-serum, ampoule, functional-skincare",
        "Introduction": "Bravity 条目按外用精华/安瓶处理，归入功效性护肤品。旧记录中 Dexlevo/GOURI 液态 PCL 证据属于错误产品匹配，已从 Bravity 证据链中剔除。",
        "Verified_Product_Type_CN": "涂抹式精华/功效安瓶",
        "Market_Channel": "professional skincare / retail skincare",
        "Data_Source": "manual_correction:bravity_not_pcl_injectable",
        "_segments": "other",
        "_note": "Moved out of PCL/Injectables; wrong GOURI evidence was quarantined in the source workbook.",
    },
    # Bepanthen / Bepanthol is a topical dexpanthenol post-procedure repair
    # cream. The seed row had been carried as EBD/Microneedling even though
    # its material taxonomy already placed it under functional skincare.
    "REC_0072": {
        "Category_L1": "Skincare",
        "Category_L2": "Cosmeceutical",
        "Tech_Type_Std": "Dexpanthenol / Barrier Repair",
        "Tech_Type_Original": "Topical dexpanthenol repair cream",
        "Core_Product": "Post-Procedure Repair",
        "Feature_Tags": "skincare, cosmeceutical, dexpanthenol, post-procedure-repair",
        "Verified_Product_Type_CN": "泛醇修复霜 / 术后屏障修复",
        "Market_Channel": "professional skincare / post-procedure care",
        "Data_Source": "manual_correction:bayer_bepanthen_not_ebd",
        "_segments": "skincare",
        "_note": "Corrected from EBD/Microneedling; topical B5/dexpanthenol repair cream, not an energy device.",
    },
    # DermaShine Duo RF is an RF/drug-delivery device. Its official launch text
    # says it can deliver viscous polymers such as PLLA/PCL, but that does not
    # make the device itself a PCL product.
    "REC_0177": {
        "Category_L1": "EBD",
        "Category_L2": "Radiofrequency",
        "Tech_Type_Std": "Radiofrequency",
        "Tech_Type_Original": "RF Multi-needle Injector (射频水光)",
        "Core_Product": "DermaShine Duo RF",
        "Feature_Tags": "ebd, radiofrequency, drug-delivery-device",
        "Data_Source": "official_company_fact_override",
        "_segments": "ebd",
        "_note": "Removed false PCL material placement; official Huons text describes an RF/drug-delivery device compatible with PCL/PLLA delivery.",
    },
    # GC Aesthetics official product pages place Nagor/Impleo/CoGel in breast
    # and body implants, not PCL threads or injectable PCL fillers.
    "REC_0474": {
        "Category_L1": "Implants",
        "Category_L2": "Breast Implant",
        "Tech_Type_Std": "Silicone Gel Breast Implant",
        "Tech_Type_Original": "Cohesive silicone gel mammary implant",
        "Core_Product": "Impleo / CoGel Breast Implants",
        "Feature_Tags": "breast-implant, silicone-gel, cohesive-gel",
        "Introduction": "GC Aesthetics/Nagor 官方资料将 Impleo、CoGel 归入乳房植入物/外科植入产品组合；不归入 PCL 材料填充剂。",
        "Data_Source": "official_company_fact_override",
        "_segments": "other",
        "_note": "Corrected from false PCL/thread placement; official GC Aesthetics pages describe Nagor as breast/body implants.",
    },
    # Official Bioxis pages describe MTI-12 as a chitosan-matrix tissular inductor
    # under clinical investigation. It is not a CaHA product; old seed tags
    # compared it with CaHA/PLLA and caused false segment placement.
    "REC_0470": {
        "Category_L2": "Regenerative Injectable Pipeline",
        "Tech_Type_Std": "Chitosan / Chitin-Glucan",
        "Tech_Type_Original": "Proprietary chitosan matrix",
        "Core_Product": "Monophasic Tissular Inductor (R&D)",
        "Feature_Tags": "chitosan, chitin-glucan, tissular-inductor, regenerative-pipeline",
        "Introduction": "官方资料显示 MTI-12 基于 Bioxis 专有 chitosan matrix，定位为 Monophasic Tissular Inductor，处于 clinical trial investigation；不归入 CaHA 子赛道。",
        "Data_Source": "official_company_fact_override",
        "_segments": "other",
        "_note": "Corrected from seed comparison text that mentioned CaHA/PLLA as benchmarks, not materials.",
    },
    # CGBIO's official aesthetics portfolio identifies FACETEM/FACETEM S as a
    # CaHA line. The seed row carried an HA tech label
    # because nearby HA fillers appeared in the same product-list text.
    "REC_0266": {
        "Category_L1": "Injectables",
        "Category_L2": "Dermal Filler",
        "Tech_Type_Std": "Calcium Hydroxylapatite",
        "Tech_Type_Original": "Calcium hydroxylapatite filler (CaHA)",
        "Core_Product": "FACETEM / FACETEM S CaHA Filler",
        "Feature_Tags": "caha, calcium-hydroxylapatite, collagen-stimulator, official-company-source",
        "Introduction": "CGBIO 官方产品页将 FACETEM/FACETEM S 标注为 Calcium Hydroxylapatite (CaHA) dermal filler；材料归属按 CaHA/胶原刺激剂处理。",
        "Data_Source": "official_company_fact_override",
        "_segments": "caha",
        "_note": "Corrected from adjacent HA filler text on the same CGBIO official portfolio page.",
    },
    "REC_0223": {
        "Category_L1": "Injectables",
        "Category_L2": "Biostimulator",
        "Tech_Type_Std": "PCL",
        "Tech_Type_Original": "PCL microspheres in CMC gel",
        "Feature_Tags": "pcl, polycaprolactone, pcl-microspheres, collagen-stimulator, injectable",
        "Verified_Product_Type_CN": "PCL 微球胶原刺激剂",
        "Data_Source": "manual_correction:ellanse_pcl",
        "_segments": "pcl",
        "_note": "Ellansé is a PCL microsphere / CMC collagen stimulator and must not be classified under HA.",
    },
    "REC_0479": {
        "_segments": "ha",
        "_note": "Pure HA / PEG-crosslinked Neauvia Organic line stays in HA; the CaHA-containing Stimulate line is split into REC_1023.",
    },
    # Osteopore products are PCL/PCL-TCP bioresorbable surgical scaffolds for
    # bone reconstruction, not injectable aesthetic PCL fillers or thread lifts.
    # The category review closed them out of the core aesthetics scope.
    "REC_0525": {
        "Category_L1": "Implants",
        "Category_L2": "Other Implant",
        "Tech_Type_Std": "PCL / PCL-TCP Bioresorbable Scaffold",
        "Tech_Type_Original": "3D printed bioresorbable scaffold / implant",
        "Core_Product": "Osteomesh / Osteoplug / aXOpore",
        "Feature_Tags": "bioresorbable-implant, bone-regeneration, pcl-scaffold",
        "Introduction": "Osteopore 官方资料将产品描述为 3D printed scaffold / bioresorbable implant，用于骨再生/骨缺损重建；不归入医美注射 PCL 或线材赛道。",
        "Material_Taxonomy_L1_CN": "植入物",
        "Material_Taxonomy_L2_CN": "骨修复",
        "Material_Taxonomy_L3_CN": "骨支架材料",
        "Material_Taxonomy_Path_CN": "植入物 > 骨修复 > 骨支架材料",
        "Material_Taxonomy_Source": "category_dictionary_feedback_20260601",
        "Material_Taxonomy_Confidence": "medium",
        "Material_Taxonomy_Review_Status": "excluded_scope",
        "Material_Taxonomy_Note": "Excluded from the active aesthetics database; retained only as a source-history material note.",
        "Material_Family": "骨支架材料",
        "Inclusion_Status": "excluded",
        "Backfill_Audit": "category_dictionary_feedback_20260601: excluded as non-core orthopedic/neurosurgical/craniomaxillofacial bone scaffold",
        "Data_Source": "official_company_fact_override",
        "_segments": "other",
        "_note": "Corrected from injectable PCL placement; official Osteopore pages describe bioresorbable surgical scaffolds/implants. User category review excludes this non-core bone-scaffold category from the aesthetics database.",
    },
    # The current source set does not support placing Universal IPS/U.IPS in
    # PCL threads and surgical implants at the same time. Keep it out of PCL
    # until an official product page/certificate confirms the material and use.
    "REC_0716": {
        "Category_L1": "Implants",
        "Category_L2": "Other Implant",
        "Tech_Type_Std": "Surgical Implant",
        "Tech_Type_Original": "Official material source unresolved",
        "Core_Product": "Surgical Implants",
        "Feature_Tags": "surgical-implant, needs-official-source",
        "Material_Taxonomy_L1_CN": "植入物",
        "Material_Taxonomy_L2_CN": "其他植入物",
        "Material_Taxonomy_L3_CN": "通用手术植入物",
        "Material_Taxonomy_Path_CN": "植入物 > 其他植入物 > 通用手术植入物",
        "Material_Taxonomy_Source": "category_dictionary_feedback_20260601",
        "Material_Taxonomy_Confidence": "medium",
        "Material_Taxonomy_Review_Status": "mapped_to_other_implant",
        "Material_Taxonomy_Note": "Broad surgical-implant row; keep out of thread-lift material taxonomy until official material evidence is available.",
        "Material_Family": "通用手术植入物",
        "Backfill_Audit": "category_dictionary_feedback_20260601: mapped Surgical Implant to Implants > Other Implant",
        "Data_Source": "taxonomy_conflict_correction",
        "_segments": "other",
        "_note": "Removed false PCL/thread placement; current source mapping does not support a PCL classification. Category review maps the broad Surgical Implant label to Implants > Other Implant.",
    },
}

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}

PROMOTION_DOMAIN_BLOCKLIST = {
    "ackomas.com",
    "alysidia.com",
    "beudamed.com",
    "bloomberg.com",
    "bsigroup.com",
    "cbinsights.com",
    "ce-certification.us",
    "chinakinmed.com",
    "complianceacuity.com",
    "consultingroom.com",
    "docs.cosmoconsult.com",
    "ec21.com",
    "eclevarmedtech.com",
    "eifu-service.com",
    "eumdr.com",
    "eumediq.eu",
    "eudamed.com",
    "everythingrf.com",
    "facebook.com",
    "finance.yahoo.com",
    "hcltech.com",
    "instagram.com",
    "investing.com",
    "itrvn.com",
    "kmslaser.com",
    "korean-surgery.com",
    "legislation.gov.uk",
    "linkedin.com",
    "marketwatch.com",
    "marketscreener.com",
    "medical-device-regulation.eu",
    "medicaldevicehq.com",
    "medicalexpo.com",
    "meddeviceguide.com",
    "medtecheurope.org",
    "morningstar.com",
    "nordicms.com",
    "obelis.net",
    "pieleterna.es",
    "pressebox.com",
    "primeramedicalsuppliesllc.com",
    "scribd.com",
    "search.eudamed.com",
    "skinonyou.com",
    "stockanalysis.com",
    "thema-med.com",
    "tuvsud.com",
    "yahoo.com",
}

COMPANY_SOURCE_TRUSTED_DOMAINS = {
    "skin tech": {
        "skintechpharmagroup.com",
        "skintechcorporation.com",
        "skintechpharmagroup.nl",
        "skintechpharmagroup.bg",
        "skintech.info",
    },
}

COMPANY_SOURCE_FALSE_POSITIVE_DOMAINS = {
    "skin tech": {
        "skinpen.com",
        "skinpenuk.com",
        "skintech.co.za",
        "skinsolutions.co.uk",
    },
}

PROMOTION_ALLOWED_PRODUCT_DOC_DOMAINS = {
    "accessdata.fda.gov",
    "allerganaesthetics.com",
    "allerganaesthetics.com.au",
    "allerganaesthetics.gr",
    "allerganaestheticsnordics.com",
    "alma-medicaldevices.com",
    "belotero.com",
    "cynosure.com",
    "galderma.com",
    "galdermaaesthetics.com",
    "hydrafacial.com",
    "ifu.merzaesthetics.com",
    "ifu.solta.com",
    "merzaesthetics.com",
    "radiesse.com",
    "restylaneusa.com",
    "rxabbvie.com",
    "skinpen.uk",
    "ultherapy.com",
}

COMPANY_OFFICIAL_PROMOTION_QUERY_TYPES = {
    "product_ifu_labeling",
    "product_certificate_registration",
    "official_ifu_catalog",
}

COMPANY_OFFICIAL_PROMOTION_CONFIDENCES = {
    "official_domain_candidate",
    "product_official_domain_candidate",
    "brand_official_search_candidate",
}

PROMOTION_GENERIC_TITLE_PATTERNS = [
    r"bfarm\s+-\s+medical devices",
    r"business rules",
    r"certificate information \(if applicable\)",
    r"ce marking for medical devices",
    r"complete guide",
    r"eudamed-help",
    r"eudamed search",
    r"eudamed, m2m",
    r"eu mdr accessories",
    r"final steps to place",
    r"ifu for medical devices explained",
    r"is your ifu compliant",
    r"list of notified bodies",
    r"medical device database",
    r"notified bodies and certificates module",
    r"notified bodies accredited",
    r"regulation \(eu\) 2017/745",
    r"what is eudamed",
]

PROMOTION_IFU_PATTERNS = [
    r"\be-?IFU\b",
    r"\bIFU\b",
    r"instruction\s+for\s+use",
    r"instructions\s+for\s+use",
]

PROMOTION_CERTIFICATE_PATTERNS = [
    r"\bcertificate\b",
    r"\bcertificates\b",
    r"ce[-_\s]?certificate",
    r"declaration\s+of\s+conformity",
    r"declaration-of-conformity",
    r"ec[-_\s]?certificate",
    r"eu[-_\s]?certificate",
    r"mdd[-_\s]?certificate",
    r"mdr[-_\s]?certificate",
]


def quote_sql_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def drop_existing_database_objects(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys=OFF")
    object_rows = cur.execute(
        """
        SELECT type, name
        FROM sqlite_master
        WHERE name NOT LIKE 'sqlite_%'
          AND type IN ('view', 'trigger', 'table')
        ORDER BY CASE type WHEN 'view' THEN 0 WHEN 'trigger' THEN 1 ELSE 2 END
        """
    ).fetchall()
    for object_type, name in object_rows:
        cur.execute(f"DROP {object_type.upper()} IF EXISTS {quote_sql_identifier(name)}")
    conn.commit()


def open_reset_database(path: Path) -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    last_error: Exception | None = None
    for attempt in range(1, 31):
        conn = sqlite3.connect(path, timeout=60)
        try:
            conn.execute("PRAGMA busy_timeout=60000")
            drop_existing_database_objects(conn)
            return conn
        except sqlite3.OperationalError as exc:
            conn.close()
            last_error = exc
            if "locked" not in str(exc).lower() and "busy" not in str(exc).lower():
                raise
            if attempt < 30:
                time.sleep(2)
                continue
            raise
    if last_error:
        raise last_error
    raise RuntimeError(f"Could not open database for rebuild: {path}")


def raw_compact_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(value or "").strip().lower())

COUNTRY_ALIASES = {
    "United States": "USA",
    "United States of America": "USA",
    "US": "USA",
    "U.S.": "USA",
    "United Kingdom": "UK",
    "Great Britain": "UK",
    "Republic of Korea": "South Korea",
    "Korea": "South Korea",
    "U.A.E.": "UAE",
    "United Arab Emirates": "UAE",
}

COUNTRY_COORDS = {
    "Argentina": (-34.6, -64.0),
    "Australia": (-25.3, 133.8),
    "Austria": (47.5, 14.6),
    "Brazil": (-14.2, -51.9),
    "Bulgaria": (42.7, 25.5),
    "Canada": (56.1, -106.3),
    "Costa Rica": (9.7, -84.2),
    "Czech Republic": (49.8, 15.5),
    "Denmark": (56.2, 10.0),
    "Finland": (64.0, 26.0),
    "France": (46.2, 2.2),
    "Germany": (51.2, 10.4),
    "Hungary": (47.1, 19.5),
    "India": (21.0, 78.0),
    "Ireland": (53.4, -8.2),
    "Israel": (31.5, 34.8),
    "Italy": (42.5, 12.5),
    "Japan": (36.2, 138.3),
    "Latvia": (56.9, 24.6),
    "Luxembourg": (49.8, 6.1),
    "Malaysia": (4.2, 102.0),
    "Monaco": (43.7, 7.4),
    "Netherlands": (52.1, 5.3),
    "Nigeria": (9.1, 8.7),
    "Poland": (52.0, 19.1),
    "Singapore": (1.35, 103.8),
    "Slovakia": (48.7, 19.7),
    "Slovenia": (46.1, 14.8),
    "South Africa": (-30.6, 22.9),
    "South Korea": (36.4, 127.8),
    "Spain": (40.4, -3.7),
    "Sweden": (60.1, 18.6),
    "Switzerland": (46.8, 8.2),
    "Taiwan": (23.7, 121.0),
    "Turkey": (39.0, 35.2),
    "UAE": (24.4, 54.3),
    "UK": (54.1, -2.3),
    "USA": (39.5, -98.4),
}

DASHBOARD_REGION_BY_COUNTRY = {
    "Argentina": "Latin America",
    "Austria": "Europe",
    "Australia": "Asia-Pacific",
    "Brazil": "Latin America",
    "Bulgaria": "Europe",
    "Canada": "North America",
    "Costa Rica": "Latin America",
    "Czech Republic": "Europe",
    "Denmark": "Europe",
    "Finland": "Europe",
    "France": "Europe",
    "Germany": "Europe",
    "Hungary": "Europe",
    "India": "Asia-Pacific",
    "Ireland": "Europe",
    "Israel": "Middle East",
    "Italy": "Europe",
    "Japan": "Asia-Pacific",
    "Latvia": "Europe",
    "Luxembourg": "Europe",
    "Malaysia": "Asia-Pacific",
    "Monaco": "Europe",
    "Netherlands": "Europe",
    "Nigeria": "Africa",
    "Poland": "Europe",
    "Singapore": "Asia-Pacific",
    "Slovakia": "Europe",
    "Slovenia": "Europe",
    "South Africa": "Africa",
    "South Korea": "Asia-Pacific",
    "Spain": "Europe",
    "Sweden": "Europe",
    "Switzerland": "Europe",
    "Taiwan": "Asia-Pacific",
    "Turkey": "Middle East",
    "UAE": "Middle East",
    "UK": "Europe",
    "USA": "North America",
}

CITY_COORDS_RAW = {
    ("Aachen", "Germany"): (50.8, 6.1),
    ("Adelaide", "Australia"): (-34.9, 138.6),
    ("Alajuela", "Costa Rica"): (10.0, -84.2),
    ("Anyang", "South Korea"): (37.4, 126.9),
    ("Apt", "France"): (43.9, 5.4),
    ("Austin", "USA"): (30.3, -97.7),
    ("Bad Birnbach", "Germany"): (48.4, 13.1),
    ("Ballerup", "Denmark"): (55.7, 12.4),
    ("Barcelona", "Spain"): (41.4, 2.2),
    ("Basel", "Switzerland"): (47.6, 7.6),
    ("Berlin", "Germany"): (52.5, 13.4),
    ("Biarritz", "France"): (43.5, -1.6),
    ("Bologna", "Italy"): (44.5, 11.3),
    ("Bordeaux", "France"): (44.8, -0.6),
    ("Bothell", "USA"): (47.8, -122.2),
    ("Brindisi", "Italy"): (40.6, 17.9),
    ("Brisbane", "USA"): (37.7, -122.4),
    ("Brookline", "USA"): (42.3, -71.1),
    ("Broumov", "Czech Republic"): (50.6, 16.3),
    ("Busan", "South Korea"): (35.2, 129.1),
    ("Caesarea", "Israel"): (32.5, 34.9),
    ("Carpi", "Italy"): (44.8, 10.9),
    ("Casalbuttano", "Italy"): (45.3, 9.9),
    ("Chantelle", "France"): (46.2, 3.2),
    ("Charlottesville", "USA"): (38.0, -78.5),
    ("Cheongju", "South Korea"): (36.6, 127.5),
    ("Chuncheon", "South Korea"): (37.9, 127.7),
    ("Chur", "Switzerland"): (46.9, 9.5),
    ("Clichy", "France"): (48.9, 2.3),
    ("Costa Brava", "Spain"): (41.9, 3.2),
    ("Cournon", "France"): (45.7, 3.2),
    ("Dallas", "USA"): (32.8, -96.8),
    ("Daejeon", "South Korea"): (36.4, 127.4),
    ("Dublin", "Ireland"): (53.3, -6.3),
    ("Dublin", "USA"): (37.7, -121.9),
    ("Dueville", "Italy"): (45.6, 11.5),
    ("Dümmer", "Germany"): (53.6, 11.2),
    ("Eaubonne", "France"): (49.0, 2.3),
    ("Ellington", "USA"): (41.9, -72.5),
    ("Fairfield", "USA"): (40.9, -74.3),
    ("Florence", "Italy"): (43.8, 11.3),
    ("Fort Lauderdale", "USA"): (26.1, -80.1),
    ("Fort Myers", "USA"): (26.6, -81.9),
    ("Frederick", "USA"): (39.4, -77.4),
    ("Fremont", "USA"): (37.5, -122.0),
    ("Fréjus", "France"): (43.4, 6.7),
    ("Fulda", "Germany"): (50.6, 9.7),
    ("Gangneung", "South Korea"): (37.8, 128.9),
    ("Gangwon-do", "South Korea"): (37.8, 128.2),
    ("Gdynia", "Poland"): (54.5, 18.5),
    ("Geneva", "Switzerland"): (46.2, 6.1),
    ("Gifu", "Japan"): (35.4, 136.8),
    ("Goyang", "South Korea"): (37.7, 126.8),
    ("Gurugram", "India"): (28.5, 77.0),
    ("Gyeonggi-do", "South Korea"): (37.4, 127.2),
    ("Hamburg", "Germany"): (53.6, 10.0),
    ("Helsinki", "Finland"): (60.2, 24.9),
    ("Holmes Chapel", "UK"): (53.2, -2.4),
    ("Hwaseong-si", "South Korea"): (37.2, 126.8),
    ("Irvine", "USA"): (33.7, -117.8),
    ("Istanbul", "Turkey"): (41.0, 28.9),
    ("Jena", "Germany"): (50.9, 11.6),
    ("Johannesburg", "South Africa"): (-26.2, 28.0),
    ("Kiryat Bialik", "Israel"): (32.8, 35.1),
    ("Kiryat Shmona", "Israel"): (33.2, 35.6),
    ("Košice", "Slovakia"): (48.7, 21.3),
    ("Kota Kinabalu", "Malaysia"): (6.0, 116.1),
    ("Langenfeld", "Germany"): (51.1, 6.9),
    ("Lantana", "USA"): (26.6, -80.1),
    ("Leobendorf", "Austria"): (48.4, 16.3),
    ("Leverkusen", "Germany"): (51.0, 7.0),
    ("Ljubljana", "Slovenia"): (46.1, 14.5),
    ("Lodi", "Italy"): (45.3, 9.5),
    ("London", "UK"): (51.5, -0.1),
    ("Lonay", "Switzerland"): (46.5, 6.5),
    ("Long Beach", "USA"): (33.8, -118.2),
    ("Los Angeles", "USA"): (34.1, -118.2),
    ("Louisville", "USA"): (38.3, -85.8),
    ("Lugano", "Switzerland"): (46.0, 8.9),
    ("Luxembourg", "Luxembourg"): (49.6, 6.1),
    ("Lyon", "France"): (45.8, 4.8),
    ("Madrid", "Spain"): (40.4, -3.7),
    ("Mandelieu-la-Napoule", "France"): (43.5, 6.9),
    ("Manchester", "UK"): (53.5, -2.2),
    ("Marlborough", "USA"): (42.3, -71.6),
    ("Melbourne", "Australia"): (-37.8, 145.0),
    ("Milan", "Italy"): (45.5, 9.2),
    ("Mission Viejo", "USA"): (33.6, -117.7),
    ("Montrodat", "France"): (44.6, 3.3),
    ("Mougins", "France"): (43.6, 7.0),
    ("Naples", "Italy"): (40.9, 14.3),
    ("Naples", "USA"): (26.1, -81.8),
    ("Nashville", "USA"): (36.2, -86.8),
    ("Netanya", "Israel"): (32.3, 34.9),
    ("Neu-Ulm", "Germany"): (48.4, 10.0),
    ("New Brunswick", "USA"): (40.5, -74.4),
    ("New Jersey", "USA"): (40.1, -74.5),
    ("New York", "USA"): (40.7, -74.0),
    ("Newport Beach", "USA"): (33.6, -117.9),
    ("Nice", "France"): (43.7, 7.3),
    ("Nimes", "France"): (43.8, 4.4),
    ("Ocala", "USA"): (29.2, -82.1),
    ("Oakland", "USA"): (40.9, -74.3),
    ("Ogliastro Cilento", "Italy"): (40.3, 15.0),
    ("Palo Alto", "USA"): (37.4, -122.1),
    ("Paris", "France"): (48.9, 2.4),
    ("Pilisborosjenő", "Hungary"): (47.6, 19.0),
    ("Pisa", "Italy"): (43.7, 10.4),
    ("Plan-les-Ouates", "Switzerland"): (46.2, 6.1),
    ("Prague", "Czech Republic"): (50.1, 14.4),
    ("Pretoria", "South Africa"): (-25.7, 28.2),
    ("Ras Al Khaimah", "UAE"): (25.8, 55.9),
    ("Rishon LeZion", "Israel"): (31.9, 34.8),
    ("Rome", "Italy"): (41.9, 12.5),
    ("Salerno", "Italy"): (40.7, 14.8),
    ("San Diego", "USA"): (32.7, -117.2),
    ("São Carlos", "Brazil"): (-22.0, -47.9),
    ("São Paulo", "Brazil"): (-23.6, -46.6),
    ("Seattle", "USA"): (47.6, -122.3),
    ("Seongnam", "South Korea"): (37.4, 127.1),
    ("Seoul", "South Korea"): (37.6, 127.0),
    ("Soeborg", "Denmark"): (55.7, 12.5),
    ("Sofia", "Bulgaria"): (42.7, 23.3),
    ("Stockholm", "Sweden"): (59.3, 18.1),
    ("Sugar Land", "USA"): (29.6, -95.6),
    ("Suwon", "South Korea"): (37.3, 127.0),
    ("Swansea", "UK"): (51.6, -3.9),
    ("Sydney", "Australia"): (-33.9, 151.2),
    ("Tägerwilen", "Switzerland"): (47.7, 9.1),
    ("Taoyuan", "Taiwan"): (25.0, 121.3),
    ("Taipei", "Taiwan"): (25.0, 121.6),
    ("Tarrytown", "USA"): (41.1, -73.9),
    ("Tel Aviv", "Israel"): (32.1, 34.8),
    ("Tigre", "Argentina"): (-34.4, -58.6),
    ("Tochigi", "Japan"): (36.6, 139.9),
    ("Tokyo", "Japan"): (35.7, 139.7),
    ("Toronto", "Canada"): (43.7, -79.4),
    ("Treviso", "Italy"): (45.7, 12.2),
    ("Trezzano sul Naviglio", "Italy"): (45.4, 9.1),
    ("Trieste", "Italy"): (45.7, 13.8),
    ("Tuttlingen", "Germany"): (47.98, 8.8),
    ("Usmate Velate", "Italy"): (45.7, 9.4),
    ("Valence", "France"): (44.9, 4.9),
    ("Valencia", "Spain"): (39.5, -0.4),
    ("Ventura", "USA"): (34.3, -119.3),
    ("Vicenza", "Italy"): (45.5, 11.5),
    ("Viganello", "Switzerland"): (46.0, 8.97),
    ("Warsaw", "Poland"): (52.2, 21.0),
    ("Westford", "USA"): (42.6, -71.4),
    ("Wolfenbüttel", "Germany"): (52.2, 10.5),
    ("Wollerau", "Switzerland"): (47.2, 8.7),
    ("Wonju", "South Korea"): (37.3, 127.9),
    ("Wrocław", "Poland"): (51.1, 17.0),
    ("Yokneam", "Israel"): (32.7, 35.1),
    ("Yongin", "South Korea"): (37.2, 127.2),
    ("Yverdon-les-Bains", "Switzerland"): (46.8, 6.6),
    ("Zug", "Switzerland"): (47.2, 8.5),
    ("Abano Terme", "Italy"): (45.4, 11.8),
    ("Alabama", "USA"): (32.8, -86.8),
    ("Alicante", "Spain"): (38.3, -0.5),
    ("Aurora", "Canada"): (44.0, -79.5),
    ("Biot", "France"): (43.6, 7.1),
    ("Boissy-l'Aillerie", "France"): (49.1, 2.0),
    ("Boulder", "USA"): (40.0, -105.3),
    ("Carlsbad", "USA"): (33.2, -117.4),
    ("Chaponost", "France"): (45.7, 4.7),
    ("Clearwater", "USA"): (28.0, -82.8),
    ("Dieburg", "Germany"): (49.9, 8.8),
    ("Durban", "South Africa"): (-29.9, 31.0),
    ("Erlangen", "Germany"): (49.6, 11.0),
    ("Frankfurt", "Germany"): (50.1, 8.7),
    ("Herborn", "Germany"): (50.7, 8.3),
    ("Homewood", "USA"): (33.5, -86.8),
    ("Jacksonville", "USA"): (30.3, -81.7),
    ("Kaohsiung", "Taiwan"): (22.6, 120.3),
    ("Katzrin", "Israel"): (32.99, 35.7),
    ("Kingston", "USA"): (47.8, -122.5),
    ("Lagos", "Nigeria"): (6.5, 3.4),
    ("Le Mont-sur-Lausanne", "Switzerland"): (46.6, 6.6),
    ("Miami", "USA"): (25.8, -80.2),
    ("Monaco", "Monaco"): (43.7, 7.4),
    ("Montpellier", "France"): (43.6, 3.9),
    ("Newark", "USA"): (39.7, -75.7),
    ("Nuremberg", "Germany"): (49.5, 11.1),
    ("Palm Beach", "USA"): (26.7, -80.0),
    ("Petah Tikva", "Israel"): (32.1, 34.9),
    ("Riga", "Latvia"): (56.9, 24.1),
    ("Rochester", "USA"): (44.0, -92.5),
    ("Ronkonkoma", "USA"): (40.8, -73.1),
    ("Scottsdale", "USA"): (33.5, -111.9),
    ("Schwerin", "Germany"): (53.6, 11.4),
    ("Singapore", "Singapore"): (1.35, 103.8),
    ("Tarragona", "Spain"): (41.1, 1.2),
    ("Tucson", "USA"): (32.2, -110.9),
    ("Uppsala", "Sweden"): (59.9, 17.6),
    ("Valbonne", "France"): (43.6, 7.0),
    ("Veldhoven", "Netherlands"): (51.4, 5.4),
    ("Yehud", "Israel"): (32.0, 34.9),
}

CITY_COORDS = {
    raw_compact_key(f"{city}|{COUNTRY_ALIASES.get(country, country)}"): coords
    for (city, country), coords in CITY_COORDS_RAW.items()
}


SOURCE_REGISTRY = [
    {
        "source_key": "company_official_product_page",
        "channel_code": "company_official",
        "source_kind": "commercial_product_identity",
        "jurisdiction": "Global",
        "regulator": "Company official",
        "source_name": "Company official product portfolio pages",
        "source_url": "",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Brand, product family, model/SKU, commercial positioning and product availability signals",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "Product/commercial identity should be anchored to company official pages when available. Media coverage is supporting evidence only.",
    },
    {
        "source_key": "company_official_ifu",
        "channel_code": "company_official",
        "source_kind": "official_product_document",
        "jurisdiction": "Global",
        "regulator": "Company official",
        "source_name": "Company IFU, brochure, catalog and official product PDF",
        "source_url": "",
        "access_method": "web/pdf",
        "machine_readable": "partial",
        "primary_use": "Product name, model, intended commercial use, specifications and manufacturer/brand-holder statements",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "Use for product facts and product-line mapping. Regulatory status and approved indications must still come from regulator records.",
    },
    {
        "source_key": "fda_openfda_510k",
        "channel_code": "fda",
        "source_kind": "market_authorization",
        "jurisdiction": "US",
        "regulator": "FDA",
        "source_name": "openFDA Device 510(k) API",
        "source_url": "https://open.fda.gov/apis/device/510k/",
        "access_method": "api",
        "machine_readable": "yes",
        "primary_use": "510(k) clearances and device/applicant metadata",
        "automation_status": "implemented",
        "priority": 1,
        "note": "Use as machine-readable FDA clearance evidence. Registration facts override company website and media claims.",
    },
    {
        "source_key": "fda_510k_database",
        "channel_code": "fda",
        "source_kind": "market_authorization",
        "jurisdiction": "US",
        "regulator": "FDA",
        "source_name": "FDA 510(k) Premarket Notification Database",
        "source_url": "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "official 510(k) details and summaries",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "Official detail pages are authoritative for clearance facts when API fields are incomplete.",
    },
    {
        "source_key": "fda_registration_listing",
        "channel_code": "fda",
        "source_kind": "registration_listing",
        "jurisdiction": "US",
        "regulator": "FDA",
        "source_name": "FDA Establishment Registration & Device Listing",
        "source_url": "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfRL/rl.cfm",
        "access_method": "web/download",
        "machine_readable": "partial",
        "primary_use": "manufacturer, owner/operator, establishment and listing signals",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Registration/listing is not approval; keep it separate from clearance/approval evidence.",
    },
    {
        "source_key": "fda_accessgudid",
        "channel_code": "fda",
        "source_kind": "device_identity",
        "jurisdiction": "US",
        "regulator": "FDA/NLM",
        "source_name": "AccessGUDID",
        "source_url": "https://accessgudid.nlm.nih.gov/",
        "access_method": "web/download/api",
        "machine_readable": "yes",
        "primary_use": "UDI/device identifier and labeler metadata",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Useful for SKU/device identity; not a substitute for approval status.",
    },
    {
        "source_key": "mdsap_program",
        "channel_code": "mdsap",
        "source_kind": "qms_audit",
        "jurisdiction": "MDSAP",
        "regulator": "MDSAP Regulatory Authorities",
        "source_name": "Medical Device Single Audit Program",
        "source_url": "https://www.fda.gov/medical-devices/cdrh-international-affairs/medical-device-single-audit-program-mdsap",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "quality management system audit evidence for Australia, Brazil, Canada, Japan and United States",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "MDSAP is a single quality-system audit program. It is not a direct sales license; each market still needs its own authorization such as ARTG, ANVISA registration, Health Canada MDL, PMDA/MHLW authorization or FDA clearance/approval/listing.",
    },
    {
        "source_key": "nmpa_udi",
        "channel_code": "nmpa",
        "source_kind": "market_authorization",
        "jurisdiction": "CN",
        "regulator": "NMPA",
        "source_name": "NMPA Medical Device UDI Database",
        "source_url": "https://udi.nmpa.gov.cn/download.html",
        "access_method": "download",
        "machine_readable": "yes",
        "primary_use": "UDI, registrant/filing holder, device identity and China listing signals",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "Large full releases should be cached locally and joined by registrant/product name.",
    },
    {
        "source_key": "nmpa_gov_service",
        "channel_code": "nmpa",
        "source_kind": "market_authorization",
        "jurisdiction": "CN",
        "regulator": "NMPA",
        "source_name": "NMPA medical device registration query",
        "source_url": "https://app.gjzwfw.gov.cn/jmopen/webapp/html5/apigcylqxcpzccx/index.html",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "China medical device registration certificates and approved scope",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "Use for official certificate fields; scrape only through compliant public access.",
    },
    {
        "source_key": "eu_eudamed",
        "channel_code": "ce",
        "source_kind": "conformity_mark",
        "jurisdiction": "EU",
        "regulator": "European Commission",
        "source_name": "EUDAMED public site",
        "source_url": "https://ec.europa.eu/tools/eudamed/",
        "access_method": "web/api",
        "machine_readable": "partial",
        "primary_use": "EU UDI/device, actor and certificate records",
        "automation_status": "source_registered",
        "priority": 1,
        "note": "EU evidence may require Basic UDI-DI, SRN and certificate cross-checks.",
    },
    {
        "source_key": "ce_notified_body_certificate",
        "channel_code": "ce",
        "source_kind": "certificate_document",
        "jurisdiction": "EU",
        "regulator": "EU Notified Body",
        "source_name": "Notified Body MDR/MDD certificate evidence",
        "source_url": "https://ec.europa.eu/tools/eudamed/",
        "access_method": "web/document",
        "machine_readable": "no",
        "primary_use": "MDR/MDD certificate number, scope, class, notified body and validity cross-check",
        "automation_status": "manual_review_path",
        "priority": 1,
        "note": "Use only when certificate, declaration of conformity, IFU or EUDAMED certificate record can be linked to the legal manufacturer/product.",
    },
    {
        "source_key": "company_ce_documents",
        "channel_code": "ce",
        "source_kind": "manufacturer_document",
        "jurisdiction": "EU",
        "regulator": "Manufacturer / Authorized Representative",
        "source_name": "Manufacturer IFU / Declaration of Conformity / EU product page",
        "source_url": "",
        "access_method": "web/pdf",
        "machine_readable": "no",
        "primary_use": "Product registered name, intended purpose, legal manufacturer, CE/MDR scope and authorized representative",
        "automation_status": "manual_review_path",
        "priority": 2,
        "note": "Manufacturer claims are supporting evidence; market authorization status should be cross-checked against EUDAMED/certificate records where available.",
    },
    {
        "source_key": "tga_artg",
        "channel_code": "tga_artg",
        "source_kind": "market_authorization",
        "jurisdiction": "AU",
        "regulator": "TGA",
        "source_name": "Australian Register of Therapeutic Goods",
        "source_url": "https://www.tga.gov.au/resources/artg",
        "access_method": "web/download",
        "machine_readable": "partial",
        "primary_use": "Australia ARTG inclusion/registration evidence",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Use ARTG evidence for Australia market access; MDSAP may support conformity assessment but does not replace ARTG status.",
    },
    {
        "source_key": "anvisa_consultas",
        "channel_code": "anvisa",
        "source_kind": "market_authorization",
        "jurisdiction": "BR",
        "regulator": "ANVISA",
        "source_name": "ANVISA Consulta de Produtos Regularizados",
        "source_url": "https://www.gov.br/anvisa/pt-br/sistemas/sistema-de-consultas",
        "access_method": "web/api",
        "machine_readable": "partial",
        "primary_use": "Brazil registration holder, manufacturer, status and risk class",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Official PDF/detail records are good evidence for legal manufacturer and validity.",
    },
    {
        "source_key": "health_canada_mdl",
        "channel_code": "health_canada",
        "source_kind": "market_authorization",
        "jurisdiction": "CA",
        "regulator": "Health Canada",
        "source_name": "Medical Devices Active Licence Listing",
        "source_url": "https://health-products.canada.ca/mdall-limh/",
        "access_method": "web/download",
        "machine_readable": "partial",
        "primary_use": "Canada medical device licence evidence",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Use MDL/licence records for Canada market authorization; MDSAP is quality-system evidence.",
    },
    {
        "source_key": "pmda_mhlw_medical_devices",
        "channel_code": "pmda_mhlw",
        "source_kind": "market_authorization",
        "jurisdiction": "JP",
        "regulator": "PMDA / MHLW",
        "source_name": "PMDA medical device information",
        "source_url": "https://www.pmda.go.jp/english/review-services/reviews/approved-information/devices/0002.html",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Japan approval/certification evidence",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Use Japan approval/certification records for market authorization. MDSAP does not itself grant Japan sales authorization.",
    },
    {
        "source_key": "mhra_public_access",
        "channel_code": "mhra_ukca",
        "source_kind": "registration_and_marking",
        "jurisdiction": "UK",
        "regulator": "MHRA",
        "source_name": "MHRA public access registration database",
        "source_url": "https://pard.mhra.gov.uk/",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "UK medical device registration and UK responsible person signals",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Use alongside UKCA/CE transition evidence and UK Responsible Person records.",
    },
    {
        "source_key": "hsa_smdr",
        "channel_code": "hsa_singapore",
        "source_kind": "market_authorization",
        "jurisdiction": "SG",
        "regulator": "HSA",
        "source_name": "Singapore Medical Device Register",
        "source_url": "https://eservice.hsa.gov.sg/medics/md/mdEnquiry.do",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Singapore SMDR registration evidence",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Use SMDR records to verify local registrant, risk class and device listing.",
    },
    {
        "source_key": "malaysia_mda_register",
        "channel_code": "malaysia_mda",
        "source_kind": "market_authorization",
        "jurisdiction": "MY",
        "regulator": "Malaysia MDA",
        "source_name": "Malaysia Medical Device Authority register",
        "source_url": "https://mdar.mda.gov.my/frontend/web/index.php?r=carian%2Findex",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Malaysia medical device registration evidence",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use for local registration holder and device registration status.",
    },
    {
        "source_key": "thai_fda_medical_device",
        "channel_code": "thai_fda",
        "source_kind": "market_authorization",
        "jurisdiction": "TH",
        "regulator": "Thai FDA",
        "source_name": "Thai FDA medical device information",
        "source_url": "https://medical.fda.moph.go.th/",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Thailand medical device registration evidence",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use Thai product and licence records for market authorization.",
    },
    {
        "source_key": "indonesia_moh_alkes",
        "channel_code": "indonesia_moh",
        "source_kind": "market_authorization",
        "jurisdiction": "ID",
        "regulator": "Indonesia MoH",
        "source_name": "Indonesia MoH medical device registration",
        "source_url": "https://infoalkes.kemkes.go.id/",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Indonesia AKL/medical device registration evidence",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use for AKL/product registration and local distributor signals.",
    },
    {
        "source_key": "philippines_fda_device",
        "channel_code": "philippines_fda",
        "source_kind": "market_authorization",
        "jurisdiction": "PH",
        "regulator": "Philippines FDA",
        "source_name": "Philippines FDA medical device verification",
        "source_url": "https://verification.fda.gov.ph/",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Philippines CMDN/CMDR verification",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use for notification/registration status and local authorization holder.",
    },
    {
        "source_key": "vietnam_dmec",
        "channel_code": "vietnam_moh",
        "source_kind": "market_authorization",
        "jurisdiction": "VN",
        "regulator": "Vietnam MoH / DMEC",
        "source_name": "Vietnam DMEC medical device portal",
        "source_url": "https://dmec.moh.gov.vn/",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Vietnam medical device circulation/registration evidence",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use for circulation number, importer and device class where available.",
    },
    {
        "source_key": "cofepris_devices",
        "channel_code": "cofepris",
        "source_kind": "market_authorization",
        "jurisdiction": "MX",
        "regulator": "COFEPRIS",
        "source_name": "COFEPRIS medical device registration information",
        "source_url": "https://www.gob.mx/cofepris",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Mexico medical device registration evidence",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use COFEPRIS registration/authorization evidence for Mexico market access.",
    },
    {
        "source_key": "mfds_emedi",
        "channel_code": "kfda",
        "source_kind": "market_authorization",
        "jurisdiction": "KR",
        "regulator": "MFDS",
        "source_name": "MFDS medical device electronic civil service",
        "source_url": "https://emedi.mfds.go.kr/",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Korea license/certificate verification and product approvals",
        "automation_status": "source_registered",
        "priority": 2,
        "note": "Korean names and import/manufacturing certificate types need manual mapping.",
    },
    {
        "source_key": "roszdravnadzor_mi",
        "channel_code": "roszdravnadzor_mi",
        "source_kind": "market_authorization",
        "jurisdiction": "RU",
        "regulator": "Roszdravnadzor",
        "source_name": "Russian state register of medical devices",
        "source_url": "https://roszdravnadzor.gov.ru/services/misearch",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Russia medical device registration records",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use Russian product/manufacturer names where available.",
    },
    {
        "source_key": "sfda_registered_md_api",
        "channel_code": "sfda_saudi",
        "source_kind": "market_authorization",
        "jurisdiction": "SA",
        "regulator": "SFDA",
        "source_name": "SFDA Registered Medical Devices Product API",
        "source_url": "https://developer.sfda.gov.sa/products/registered-medical-devices-product",
        "access_method": "api",
        "machine_readable": "yes",
        "primary_use": "Saudi registered medical device product query",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "May require API subscription/authentication before automated collection.",
    },
    {
        "source_key": "taiwan_tfda_devices",
        "channel_code": "tfda_taiwan",
        "source_kind": "market_authorization",
        "jurisdiction": "TW",
        "regulator": "Taiwan TFDA",
        "source_name": "Taiwan medical device license database",
        "source_url": "https://info.fda.gov.tw/MLMS/H0001.aspx",
        "access_method": "web",
        "machine_readable": "partial",
        "primary_use": "Taiwan medical device licence evidence",
        "automation_status": "source_registered",
        "priority": 3,
        "note": "Use for Taiwan licence number, applicant and product name verification.",
    },
]

SOURCE_AUTHORITY_POLICY = [
    {
        "fact_group": "registration_status_and_approved_indication",
        "authoritative_source": "Regulator database / official certificate / official label",
        "primary_sources": "FDA 510(k), FDA Registration & Listing, AccessGUDID, EUDAMED/MDR certificate, country regulator records",
        "supporting_sources": "Company website, IFU, press release, annual report, media",
        "merge_rule": "Regulatory source wins. Company or media claims can only create candidates or conflict notes.",
        "manual_role": "No routine manual approval. Correct mappings later only when the app reveals a detail issue.",
    },
    {
        "fact_group": "commercial_product_identity",
        "authoritative_source": "Company official product page / official catalog / IFU / official product PDF",
        "primary_sources": "Company official website, official IFU, official brochure/catalog, official distributor page when manufacturer site is absent",
        "supporting_sources": "Regulator records, conference pages, media, distributor listings",
        "merge_rule": "Official company source defines brand/product/SKU/commercial positioning. Regulator source can normalize legal registered name.",
        "manual_role": "No product-existence confirmation queue. User corrections are accepted as later curation edits.",
    },
    {
        "fact_group": "company_capital_and_ownership",
        "authoritative_source": "Exchange filing / securities regulator / annual report / company official transaction release",
        "primary_sources": "SEC, exchange filings, annual reports, prospectus, official acquisition releases",
        "supporting_sources": "News media, industry databases, market-data snapshots",
        "merge_rule": "Official securities/company source wins. Market data is timestamped snapshot only.",
        "manual_role": "No routine manual approval. Flag stale ticker, acquisition or delisting conflicts for later correction.",
    },
    {
        "fact_group": "secondary_claims_and_product_advantages",
        "authoritative_source": "Claim-specific official evidence",
        "primary_sources": "IFU, patent, clinical publication, regulator summary, annual report, official product page",
        "supporting_sources": "Media, KOL posts, distributor marketing, conference news",
        "merge_rule": "Secondary media can cross-check or find leads, but cannot become verified differentiators alone.",
        "manual_role": "No approval gate; unresolved claims stay as claim_text or conflict_note.",
    },
]

FIELD_DICTIONARY_ROWS = [
    {
        "table_name": "Registration_Evidence",
        "field_name": "official_description_exact",
        "definition": "Precise official approved-indication or intended-use wording captured from the regulator record, registration certificate, PMA/510(k), EUDAMED/CE document, IFU or official labeling.",
        "display_note": "Use this as the field-level annotation/original certificate wording. Do not replace it with translated dashboard buckets.",
        "source_priority": "Regulator/certificate/official label first; company IFU only when it is the official labeling source.",
    },
    {
        "table_name": "Registration_Evidence",
        "field_name": "official_description_source_field",
        "definition": "Internal pointer showing whether official_description_exact came from approved_indication, intended_use, or an already promoted official_description_exact field.",
        "display_note": "Useful for audit trails and later review of whether a row has true registration wording.",
        "source_priority": "Derived during build.",
    },
    {
        "table_name": "Registration_Evidence",
        "field_name": "field_note",
        "definition": "Field-level note explaining whether precise official wording has been captured and how the row should be interpreted.",
        "display_note": "Show as a tooltip or comment when the row is used in dashboards or Excel.",
        "source_priority": "Derived during build.",
    },
    {
        "table_name": "Registration_Evidence",
        "field_name": "approved_indication",
        "definition": "Official approved indication text when the source explicitly presents an approved indication.",
        "display_note": "Legal/regulatory fact. Keep exact source language where possible.",
        "source_priority": "Regulator database, PMA/510(k) decision, certificate, official labeling.",
    },
    {
        "table_name": "Registration_Evidence",
        "field_name": "intended_use",
        "definition": "Official intended-use wording when the source provides intended use rather than an approved indication.",
        "display_note": "Use when the market authorization source does not use an 'approved indication' label.",
        "source_priority": "Regulator database, official labeling, IFU.",
    },
    {
        "table_name": "Official_Indication_Evidence",
        "field_name": "official_description_exact",
        "definition": "The same precise official wording promoted into the product-country-regulator indication long table.",
        "display_note": "This is the official text to inspect on hover/detail views; buckets are analysis labels only.",
        "source_priority": "Promoted from Registration_Evidence official_description_exact.",
    },
    {
        "table_name": "Official_Indication_Evidence",
        "field_name": "buckets",
        "definition": "Derived Chinese analysis buckets such as 皱纹/皮肤褶皱 or 颈胸部/胸前区.",
        "display_note": "For heatmaps and aggregation only. Not registration-certificate wording.",
        "source_priority": "Algorithmic derivation from official_description_exact.",
    },
    {
        "table_name": "Official_Indication_Evidence",
        "field_name": "analysis_bucket_note",
        "definition": "Explicit warning that buckets are derived dashboard categories and not the official label text.",
        "display_note": "Show near heatmaps or in tooltips to avoid confusing indication buckets with official wording.",
        "source_priority": "Derived during build.",
    },
]


SEGMENTS = [
    {
        "code": "ha",
        "name": "HA / 透明质酸",
        "subtitle": "玻尿酸填充、skin booster、水光和交联/非交联配方",
        "color": "#DA7756",
        "terms": ["hyaluronic acid", "sodium hyaluronate", "hyaluron", "透明质酸", "玻尿酸", "skin booster"],
    },
    {
        "code": "plla",
        "name": "PLLA / PDLLA",
        "subtitle": "聚乳酸胶原刺激剂、再生填充和长效轮廓材料",
        "color": "#C4956A",
        "terms": ["plla", "pdlla", "poly-l-lactic", "poly lactic", "lactic acid", "聚左乳酸", "聚乳酸", "aesthefill", "sculptra", "lanluma"],
    },
    {
        "code": "pcl",
        "name": "PCL",
        "subtitle": "液态 PCL 与 PCL 微球填充剂；线材类产品归入线材板块",
        "color": "#5B8DB8",
        "terms": [
            "pcl",
            "polycaprolactone",
            "liquid pcl",
            "液态pcl",
            "液体线",
            "gouri",
            "bravity",
            "ellanse",
            "miracle l",
            "miracle h",
            "miracle touch",
        ],
    },
    {
        "code": "caha",
        "name": "CaHA",
        "subtitle": "CaHA 胶原刺激、稀释肤质改善与 HA/CaHA 复合注射剂",
        "color": "#6F87A6",
        "terms": ["caha", "calcium hydroxylapatite", "hydroxylapatite", "radiesse", "harmonyca", "羟基磷灰石"],
    },
    {
        "code": "pn_pdrn",
        "name": "PN / PDRN",
        "subtitle": "聚核苷酸、PDRN、复合制剂与原料线索",
        "color": "#7BA99C",
        "terms": ["polynucleotide", "pdrn", "pn", "polydeoxyribonucleotide", "聚核苷酸", "核苷酸"],
    },
    {
        "code": "exosome",
        "name": "Exosome / Regenerative",
        "subtitle": "外泌体制剂、PRP/PRF、细胞因子和分泌组相关产品",
        "color": "#8B6F6F",
        "terms": ["exosome", "外泌体", "prp", "prf", "growth factor", "cytokine", "secretome", "conditioned media"],
    },
    {
        "code": "botulinum",
        "name": "Botulinum Toxin",
        "subtitle": "肉毒毒素、ready-to-use 液体剂型和适应症扩展",
        "color": "#B87333",
        "terms": ["botulinum", "bont", "toxin", "neurotoxin", "肉毒", "alluzience", "botox", "dysport", "innotox"],
    },
    {
        "code": "ebd",
        "name": "EBD Devices",
        "subtitle": "激光、射频、HIFU、微针、冷冻溶脂和能量设备",
        "color": "#4A7C8E",
        "terms": ["ebd", "laser", "radiofrequency", " rf", "hifu", "ultrasound", "ipl", "picosecond", "cryolipolysis", "射频", "超声", "激光"],
    },
    {
        "code": "threads",
        "name": "Threads",
        "subtitle": "PDO/PCL/PLLA 线材、提升线和可吸收植入物",
        "color": "#9A8C73",
        "terms": ["thread", "threads", "pdo", "pcl threads", "plla threads", "线雕", "线材", "提升线"],
    },
    {
        "code": "mesotherapy",
        "name": "Mesotherapy",
        "subtitle": "中胚层复配注射液、HA 基底配方和注射设备/针头耗材",
        "color": "#6E9F88",
        "terms": ["mesotherapy", "meso", "中胚层", "cocktail", "revitalizer", "活化"],
    },
]

SEGMENT_BY_CODE = {item["code"]: item for item in SEGMENTS}


FDA_INDICATION_BUCKETS = [
    {"name": "颞部/太阳穴凹陷", "terms": ["temple hollowing", "temporal hollowing", "temple hollow", "temporal hollow", "temple treatment", "temple augmentation", "temple", "temporal", "颞部", "太阳穴"]},
    {"name": "鼻部塑形", "terms": ["nasal dorsum", "nasal radix", "nasal reshaping", "nose reshaping", "rhinoplasty", "nasal contour", "nose contour", "鼻背", "鼻根", "鼻小柱", "前鼻棘", "隆鼻", "外鼻", "鼻部塑形"]},
    {"name": "乳房/胸部增容或重建", "terms": ["breast augmentation", "breast reconstruction", "breast implant", "breast implants", "mammary prostheses", "mammary prosthesis", "tissue expander", "underdeveloped breast", "breast asymmetry", "乳房", "隆乳", "丰胸", "乳房再造", "乳房重建", "人工乳房", "胸部增容"]},
    {"name": "身体塑形/大容量填充", "terms": ["body contouring", "body contour", "body sculpting", "body re-sculpting", "body shaping", "large volume", "large-volume", "buttock", "buttocks", "gluteal", "calf", "abdomen", "abdominal", "thigh", "thighs", "flank", "flanks", "cellulite", "localized fat", "localized adiposity", "generalized adiposity", "body filler", "soft tissue volume", "身体塑形", "身体轮廓", "身体填充", "丰臀", "臀部", "小腿", "大容量", "轮廓雕塑", "轮廓塑形", "身体曲线"]},
    {"name": "面部软组织增容", "terms": ["facial soft tissue augmentation", "facial tissue augmentation", "soft tissue augmentation", "facial volume restoration", "facial volume", "facial contours", "facial contour", "facial defect", "restore loss of fullness", "restore depressions", "restore lost volume", "depressed areas", "volume deficit", "volume loss", "lost volume", "facial morphologic asymmetry", "面部容量", "容量缺失", "容量补充", "组织容量", "面部凹陷", "面部轮廓", "面部体积", "组织修复", "填充"]},
    {"name": "鼻唇沟", "terms": ["nasolabial fold", "nasolabial folds", "nlf", "smile lines", "鼻唇沟", "鼻唇部", "法令纹"]},
    {"name": "手背/手部增容", "terms": ["hand augmentation", "dorsum of the hands", "dorsum of hands", "dorsal hand", "hands", "hand volume", "hand rejuvenation", "手背", "手部"]},
    {"name": "唇部增容", "terms": ["lip augmentation", "lip fullness", "vermillion", "vermilion", "oral commissure", "lips", "lip", "唇部", "唇红", "唇粘膜", "丰唇"]},
    {"name": "面颊/中面部", "terms": ["midface", "mid-face", "cheek", "cheeks", "age-related volume deficit", "中面部", "面颊", "面颊部"]},
    {"name": "下颌线/下颏", "terms": ["jawline", "chin", "chin retrusion", "lower face", "submental", "under chin", "下颌", "下颏", "下巴", "颏部", "下颌后缩", "下颌轮廓", "双下巴"]},
    {"name": "面部脂肪萎缩", "terms": ["facial fat loss", "facial lipoatrophy", "lipoatrophy", "hiv"]},
    {"name": "眼周/泪沟", "terms": ["tear trough", "under eye", "under-eye", "infraorbital", "periorbital", "eye area", "eye wrinkles", "眼周", "泪沟", "眼部", "眼角"]},
    {"name": "动态皱纹", "terms": ["glabellar", "lateral canthal", "crow's feet", "forehead lines", "dynamic wrinkles", "frown lines", "masseter", "platysma", "brow lift", "眉间纹", "皱眉纹", "鱼尾纹", "额纹", "咬肌", "眉毛提升"]},
    {"name": "埋线提升/组织提升", "terms": ["thread lift", "thread lifting", "lifting thread", "resorbable thread", "resorbable sterile implantable", "facial and neck reconstruction", "ptosis", "pdo", "molding cog", "cog thread", "silhouette soft", "instalift", "线雕", "提拉线", "提升线", "埋线", "全脸强效提拉", "物理提拉"]},
    {"name": "皱纹/皮肤褶皱", "terms": ["wrinkles", "facial wrinkles", "folds", "rhytids", "fine lines", "wrinkle improvement", "皱纹", "细纹", "皮肤褶皱", "中重度皱纹", "皮肤纹理"]},
    {"name": "颈胸部/胸前区", "terms": ["décolleté", "decollete", "neck and upper chest", "upper chest", "upper breasts", "胸前", "颈胸", "胸部"]},
    {"name": "痤疮/疤痕", "terms": ["acne scars", "acne scar", "atrophic scar", "depressed scar", "scars", "surgical scars", "scar revision", "acne", "痤疮", "瘢痕", "疤痕", "萎缩性瘢痕"]},
    {"name": "脱毛", "terms": ["hair removal", "hair-removal", "hair reduction", "hair-reduction", "permanent hair reduction", "long-term hair reduction", "unwanted hair", "脱毛", "毛发脱减"]},
    {"name": "头发/脱发", "terms": ["hair restoration", "hair regrowth", "hair growth", "hair density", "hair loss", "alopecia", "androgenic alopecia", "scalp", "hair transplantation", "follicular extraction", "hair follicles", "eyelash", "eyelashes", "bimatoprost", "latisse", "脱发", "头发", "毛发再生", "发量", "植发", "头皮", "睫毛"]},
    {"name": "色素/纹身/血管", "terms": ["pigmented lesions", "pigmentation", "hyperpigmentation", "vascular lesions", "vascular treatment", "superficial veins", "dilated arterioles", "capillaries", "venules", "tattoo", "melasma", "telangiectasia", "rosacea", "couperose", "sun spots", "age spots", "freckles", "cosmelan", "dermamelan", "色素", "色斑", "暗沉", "血管", "红血丝", "纹身", "黄褐斑"]},
    {"name": "肤质/水合/光泽", "terms": ["skin quality", "skin hydration", "hydration", "dehydrated skin", "elasticity", "density", "skin tone", "skin texture", "radiance", "brightening", "lightening", "revitalize", "biorevitalizing", "anti-aging", "photo-aging", "skin booster", "exosome regenerative", "水合", "补水", "水光", "弹性", "致密度", "肤质", "光泽", "肤色", "抗光老化", "抗衰", "抗氧化", "皮肤质量", "护肤应用", "外泌体再生"]},
    {"name": "皮肤重建/嫩肤", "terms": ["skin resurfacing", "resurfacing", "rejuvenation", "dermal remodeling", "skin renewal", "skin ageing", "skin concerns", "skin irregularities", "skin lesions", "chemical peel", "peeling", "peel", "acid peel", "cosmetic blemishes", "cosmetic transdermal", "carboxytherapy", "skin biome", "microbiome", "cryosurgery", "liquid nitrogen", "嫩肤", "重建", "年轻化", "皮肤管理", "焕肤", "复合酸", "壬二酸", "微生态", "二氧化碳治疗", "二氧化碳注入", "冷冻皮肤病变", "液氮"]},
    {"name": "皮肤紧致", "terms": ["skin tightening", "tightening", "laxity", "lax", "loose skin", "slack skin", "flaccidity", "firming", "toning", "tissue contraction", "tissue heating", "skin elasticity improvement", "皮肤松弛", "松弛", "紧致", "提升皮肤弹性", "组织加热"]},
    {"name": "脂肪减少/塑形", "terms": ["fat reduction", "lipolysis", "adipose", "cryolipolysis", "fat cells", "submental fat", "convexity or fullness", "liposuction", "fragmentation, emulsification and aspiration", "脂肪", "减脂", "溶脂", "冷冻溶脂", "脂肪减少", "抽脂", "吸脂"]},
    {"name": "肌肉刺激", "terms": ["muscle stimulation", "neuromuscular", "muscle re-education", "muscle activation", "strengthening", "abdominal tone", "muscle aches", "muscle spasms", "肌肉刺激", "肌肉"]},
    {"name": "私密/尿失禁", "terms": ["urinary incontinence", "vaginal", "pelvic floor", "women's health", "external intimate", "inner thighs and groin", "penile", "penis", "micropenis", "urogenital", "sexual function", "genitourinary", "私密", "尿失禁", "阴道", "盆底", "阴茎", "男性私密", "泌尿生殖", "绝经"]},
    {"name": "术后/创面修复", "terms": ["post-procedure", "post procedure", "post surgical", "post-surgical", "wound healing", "wound", "wounds", "ulcer", "ulcerations", "recovery from surgical", "minimize downtime", "sensitized skin following aesthetic", "repair", "barrier", "术后", "创面", "伤口", "修复", "恢复期", "屏障", "组织恢复", "加速组织恢复"]},
    {"name": "注射/给药辅助", "terms": ["microneedle", "needle", "cannula", "injector", "drug delivery", "trans-epidermal delivery", "transdermal delivery", "administering pharmaceuticals", "skin booster injections", "microchannel", "micro-channel", "注射针", "针头", "微针", "给药", "注射器", "水光注射", "微通道", "真皮层局部导入", "精华递送"]},
    {"name": "外科软组织处理", "terms": ["soft tissue coagulation", "coagulation of soft tissue", "ablation", "vaporization", "excision", "incision", "hemostasis", "general surgical procedures", "dermatologic and general surgical procedures", "electrocoagulation", "软组织", "切割", "凝血", "消融"]},
    {"name": "影像/诊断/术前规划", "terms": ["three-dimensional imaging", "3d imaging", "aesthetic photography", "pre-operative consultation", "simulation", "visualization", "visual examination", "illuminate body surfaces", "external examination", "imaging", "assessment", "skin analysis", "图像分析", "诊断", "影像", "三维", "术前", "模拟"]},
    {"name": "PRP/血液制备", "terms": ["platelet-rich plasma", "prp", "platelet", "blood plasma", "separation", "富血小板血浆", "血小板", "血浆", "分离和提取"]},
]


NON_SPECIFIC_INDICATION_TERMS = [
    "stated in the enclosure",
    "legally marketed predicate",
    "substantially equivalent",
    "devices marketed in interstate commerce",
    "unavailable_verified_no_public_indication",
    "no public indication",
    "change in the sterilization validation method",
    "metadata confirms",
    "detailed expanded indication text is captured",
    "confirmed at product-family level",
    "approved by korean mfds; exact",
    "exact mfds number pending",
    "exact approval number pending",
    "registration confirmed at product-identity level",
    "product line; korea registration confirmed",
    "retained as a korean",
    "multi-application device intended",
    "公开资料显示其于",
    "获批规格为",
    "具体法定适应症待",
]


SEGMENT_TAXONOMY: dict[str, dict[str, list[dict[str, Any]]]] = {
    "ha": {
        "subtracks": [
            {"name": "Filler / 交联填充剂", "terms": ["dermal filler", "filler", "cross-linked", "cross linked", "volume", "volum"]},
            {"name": "Skin Booster / 无交联水光", "terms": ["skin booster", "booster", "non-cross-linked", "non cross linked", "uncrosslinked", "hydro", "hydrate", "revive", "revital", "水光", "补水"]},
            {"name": "中胚层 HA 复配液", "terms": ["mesotherapy", "meso", "biorevital", "revitalizer", "microinjection", "微注射", "水光针"]},
            {"name": "HA 主成分复配", "terms": ["ha complex", "hyaluronic acid complex", "hyaluron complex", "ha +", "ha/", "透明质酸复配", "玻尿酸复配"]},
        ],
        "indications": [
            {"name": "面部填充 / 轮廓", "terms": ["dermal filler", "filler", "volume", "contour", "deep", "intense", "subq", "面部", "轮廓", "填充"]},
            {"name": "皮肤补水 / 光泽", "terms": ["skin booster", "hydro", "hydrate", "shine", "revive", "biorevital", "水光", "补水", "光泽"]},
            {"name": "唇部塑形", "terms": ["lip", "lips", "kysse", "唇"]},
            {"name": "眼周 / 细纹", "terms": ["tear trough", "under eye", "eye", "fine line", "soft", "眼", "细纹"]},
            {"name": "身体塑形", "terms": ["body", "buttock", "breast", "cellulite", "身体"]},
            {"name": "私密修复", "terms": ["vaginal", "urogenital", "desirial", "私密"]},
            {"name": "术后修复 / 屏障", "terms": ["post-procedure", "repair", "barrier", "anti-inflammatory", "修复", "抗炎"]},
        ],
    },
    "plla": {
        "subtracks": [
            {"name": "注射型 PLLA", "terms": ["inject", "sculptra", "aesthefill", "lanluma", "plla", "pdlla"]},
            {"name": "PDLLA / 复合微球", "terms": ["pdlla", "particle", "microsphere", "microparticle", "微球"]},
        ],
        "indications": [
            {"name": "胶原再生 / 紧致", "terms": ["collagen", "biostimulator", "regeneration", "tighten", "firm", "胶原", "再生", "紧致"]},
            {"name": "面部容量恢复", "terms": ["volume", "contour", "facial", "cheek", "temple", "face", "容量", "轮廓"]},
            {"name": "身体塑形", "terms": ["body", "buttock", "hip", "cellulite", "身体"]},
            {"name": "皮肤质量改善", "terms": ["skin quality", "texture", "wrinkle", "rejuvenation", "肤质", "皱纹"]},
            {"name": "提升 / 支撑", "terms": ["thread", "lift", "support", "cog", "提拉", "支撑"]},
        ],
    },
    "pcl": {
        "subtracks": [
            {
                "name": "液态 PCL",
                "terms": ["liquid pcl", "fully solubilized", "solubilized pcl", "cesabp", "gouri", "bravity", "液态pcl", "液体线"],
            },
            {
                "name": "PCL 微球填充剂",
                "terms": [
                    "ellanse",
                    "pcl microsphere",
                    "polycaprolactone microsphere",
                    "poly-ɛ-caprolactone",
                    "pcl-based dermal filler",
                    "聚己内酯微球",
                    "pcl晶球",
                ],
            },
            {
                "name": "PCL 未细分形态",
                "terms": ["pcl", "polycaprolactone", "聚己内酯"],
            },
        ],
        "indications": [
            {"name": "胶原刺激 / 再生", "terms": ["collagen", "biostimulator", "regeneration", "stimulate", "胶原", "再生"]},
            {"name": "面部轮廓 / 容量", "terms": ["filler", "volume", "contour", "face", "cheek", "轮廓", "填充"]},
            {"name": "皮肤紧致 / 弹性", "terms": ["tighten", "firm", "elastic", "skin quality", "紧致", "弹性"]},
            {"name": "提升 / 线雕", "terms": ["thread", "lift", "cog", "matrix", "提拉", "线雕"]},
            {"name": "身体塑形", "terms": ["body", "buttock", "cellulite", "身体"]},
        ],
    },
    "caha": {
        "subtracks": [
            {"name": "CaHA 微球/胶原刺激剂", "terms": ["caha", "calcium hydroxylapatite", "hydroxylapatite", "radiesse", "facetem", "羟基磷灰石"]},
            {"name": "HA + CaHA 复合注射剂", "terms": ["harmonyca", "neauvia stimulate", "hybrid", "caha + ha", "ha/caha", "peg-ha + caha", "复合"]},
        ],
        "indications": [
            {"name": "面部轮廓 / 容量", "terms": ["filler", "volume", "contour", "face", "cheek", "轮廓", "填充"]},
            {"name": "胶原刺激 / 再生", "terms": ["collagen", "biostimulator", "regeneration", "stimulate", "胶原", "再生"]},
            {"name": "稀释/超稀释应用", "terms": ["hyperdilute", "hyper-dilute", "dilute", "radiesse (+)", "saline", "skin quality", "稀释", "超稀释"]},
            {"name": "皮肤质量 / 紧致", "terms": ["hyperdilute", "skin quality", "tighten", "firm", "elastic", "肤质", "紧致"]},
            {"name": "复合提升 / 支撑", "terms": ["hybrid", "support", "lift", "harmonyca", "复合", "支撑"]},
        ],
    },
    "pn_pdrn": {
        "subtracks": [
            {"name": "PN 制剂", "terms": ["polynucleotide", "pn", "rejuran", "nucleofill", "philart", "聚核苷酸"]},
            {"name": "PDRN 制剂", "terms": ["pdrn", "polydeoxyribonucleotide", "salmon dna"]},
            {"name": "PN/PDRN 复合制剂", "terms": ["pn/pdrn", "pdrn/pn", "ha-pdrn", "hyla-pdrn", "skin booster", "meso", "cocktail", "complex", "复合"]},
            {"name": "PN/PDRN 原料/API", "terms": ["api", "raw material", "ingredient", "原料"]},
        ],
        "indications": [
            {"name": "皮肤修复", "terms": ["repair", "healing", "barrier", "regeneration", "修复"]},
            {"name": "肤质 / 弹性改善", "terms": ["skin quality", "elastic", "texture", "rejuvenation", "肤质", "弹性"]},
            {"name": "眼周细纹", "terms": ["eye", "under eye", "fine line", "wrinkle", "眼", "细纹"]},
            {"name": "炎症 / 术后恢复", "terms": ["anti-inflammatory", "inflammation", "post-procedure", "wound", "抗炎", "术后"]},
            {"name": "头皮毛发", "terms": ["hair", "scalp", "alopecia", "头皮", "毛发"]},
        ],
    },
    "exosome": {
        "subtracks": [
            {"name": "外泌体制剂", "terms": ["exosome", "pep", "vesicle", "外泌体"]},
            {"name": "PRP / PRF", "terms": ["prp", "prf", "platelet"]},
            {"name": "生长因子 / 细胞因子", "terms": ["growth factor", "cytokine"]},
            {"name": "条件培养基 / 分泌组", "terms": ["conditioned media", "secretome", "stem cell conditioned"]},
        ],
        "indications": [
            {"name": "皮肤再生 / 抗炎", "terms": ["regeneration", "anti-inflammatory", "inflammation", "repair", "再生", "抗炎"]},
            {"name": "术后修复", "terms": ["post-procedure", "wound", "healing", "repair", "术后", "修复"]},
            {"name": "肤质 / 光泽", "terms": ["skin quality", "glow", "rejuvenation", "texture", "肤质"]},
            {"name": "毛发 / 头皮", "terms": ["hair", "scalp", "alopecia", "毛发", "头皮"]},
            {"name": "抗衰 / 细纹", "terms": ["anti-aging", "wrinkle", "fine line", "抗衰", "细纹"]},
        ],
    },
    "botulinum": {
        "subtracks": [
            {"name": "冻干粉针", "terms": ["powder", "vial", "lyophilized", "botox", "dysport", "xeomin"]},
            {"name": "即用液体剂型", "terms": ["ready-to-use", "liquid", "alluzience", "innotox"]},
            {"name": "长效 / 新剂型", "terms": ["daxxify", "long acting", "long-lasting", "novel"]},
        ],
        "indications": [
            {"name": "动态皱纹", "terms": ["wrinkle", "glabellar", "crow", "forehead", "皱纹"]},
            {"name": "瘦脸 / 轮廓", "terms": ["masseter", "contour", "jaw", "瘦脸", "轮廓"]},
            {"name": "多汗", "terms": ["hyperhidrosis", "sweat", "多汗"]},
            {"name": "颈阔肌 / 提拉", "terms": ["neck", "platysma", "lift", "颈"]},
            {"name": "治疗性神经肌肉", "terms": ["migraine", "spasticity", "dystonia", "therapeutic"]},
        ],
    },
    "ebd": {
        "subtracks": [
            {"name": "射频微针 / RF Microneedling", "terms": ["rf microneedling", "radiofrequency microneedling", "morpheus8", "potenza", "secret rf", "sylfirm", "微针射频", "射频微针"]},
            {"name": "射频溶脂 / RFAL", "terms": ["rfal", "radiofrequency assisted lipolysis", "bodytite", "facetite", "accutite", "射频溶脂"]},
            {"name": "射频 / RF", "terms": ["radiofrequency", " rf", "thermage", "forma", "accent prime", "exilis", "endymed", "射频"]},
            {"name": "激光 / Laser / IPL", "terms": ["laser", "ipl", "picosecond", "pico", "fractional", "co2", "diode laser", "nd:yag", "alexandrite", "激光"]},
            {"name": "超声 / Ultrasound / HIFU", "terms": ["hifu", "ultrasound", "ulthera", "doublo", "sonic", "超声"]},
            {"name": "冷冻 / Cryolipolysis", "terms": ["cryolipolysis", "coolsculpting", "cooltech", "cryo", "冷冻"]},
            {"name": "电磁 / EMS", "terms": ["emsculpt", "cooltone", "electromagnetic", "magnetic", "ems", "muscle stimulation", "电磁"]},
            {"name": "等离子 / Plasma", "terms": ["plasma", "helium plasma", "renuvion", "j-plasma", "等离子"]},
            {"name": "磨削 / Hydradermabrasion", "terms": ["hydrafacial", "hydradermabrasion", "microdermabrasion", "dermabrasion", "水氧", "磨削"]},
        ],
        "indications": [
            {"name": "紧致提升", "terms": ["tighten", "lifting", "firm", "hifu", "ultrasound", "紧致", "提升"]},
            {"name": "嫩肤 / 质地改善", "terms": ["resurfacing", "rejuvenation", "texture", "fractional", "嫩肤"]},
            {"name": "色斑 / 色素", "terms": ["pigment", "melasma", "tattoo", "picosecond", "色斑", "色素"]},
            {"name": "痤疮 / 疤痕", "terms": ["acne", "scar", "痤疮", "疤痕"]},
            {"name": "脱毛", "terms": ["hair removal", "脱毛"]},
            {"name": "身体塑形", "terms": ["body contour", "fat", "cryolipolysis", "cellulite", "身体"]},
        ],
    },
    "threads": {
        "subtracks": [
            {"name": "PDO 线", "terms": ["pdo"]},
            {"name": "PCL 线", "terms": ["pcl"]},
            {"name": "PLLA 线", "terms": ["plla"]},
            {"name": "Cog / Matrix 结构", "terms": ["cog", "matrix", "barb"]},
        ],
        "indications": [
            {"name": "面部提升", "terms": ["lift", "face", "cog", "提拉", "提升"]},
            {"name": "轮廓支撑", "terms": ["contour", "support", "jaw", "轮廓"]},
            {"name": "鼻部 / 精细塑形", "terms": ["nose", "rhinoplasty", "鼻"]},
            {"name": "身体提升", "terms": ["body", "breast", "buttock", "身体"]},
            {"name": "胶原刺激", "terms": ["collagen", "stimulate", "胶原"]},
        ],
    },
    "mesotherapy": {
        "subtracks": [
            {"name": "复配注射液 / Cocktail", "terms": ["meso", "mesotherapy", "cocktail", "solution", "mesohyal", "mesolift", "中胚层"]},
            {"name": "HA 基底复配液 / HA-based cocktail", "terms": ["hyaluronic acid", "hyaluron", " ha ", "skin booster", "mesoheal", "rrs ha", "水光针"]},
            {"name": "生物活性成分复配", "terms": ["biorevitalizer", "bio-stimulation", "growth factor", "exosome", "pdrn", "polynucleotide", "peptide", "amino acid", "vitamin", "mi-rna", "nctc"]},
            {"name": "注射设备 / 针头耗材", "terms": ["mesogun", "mesotherapy gun", "mpgun", "injector", "multi-injector", "needle", "cannula", "vacuum mesotherapy", "水晶针", "针头", "注射枪"]},
        ],
        "indications": [
            {"name": "补水 / 光泽", "terms": ["hydrate", "glow", "skin booster", "水光", "补水"]},
            {"name": "肤质活化", "terms": ["revitalizer", "rejuvenation", "texture", "活化", "肤质"]},
            {"name": "局部减脂", "terms": ["fat", "lipolysis", "cellulite", "溶脂"]},
            {"name": "头皮毛发", "terms": ["hair", "scalp", "alopecia", "头皮", "毛发"]},
            {"name": "色沉 / 暗沉", "terms": ["pigment", "bright", "melasma", "色沉", "暗沉"]},
        ],
    },
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def read_sheet_dicts(path: Path, sheet_name: str, header_row: int = 1) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < header_row:
        return []
    headers = [norm(v) or f"Column_{i + 1}" for i, v in enumerate(rows[header_row - 1])]
    records = []
    for row in rows[header_row:]:
        if not any(v is not None and norm(v) for v in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        records.append(record)
    return records


def find_file(pattern: str) -> Path:
    override = norm(os.environ.get("GLOBAL_AESTHETICS_SOURCE_WORKBOOK"))
    if override and pattern == "*标准化版v4.xlsx":
        override_path = Path(override)
        if override_path.exists() and zipfile.is_zipfile(override_path):
            return override_path
    matches = [
        p
        for p in SOURCE_DIR.glob(pattern)
        if not p.name.startswith("~$")
    ]
    if not matches:
        raise FileNotFoundError(f"No source file matched {pattern!r} in {SOURCE_DIR}")
    if pattern.lower().endswith(".xlsx"):
        matches = sorted(matches, key=lambda p: ("backup" in p.name.lower(), -p.stat().st_mtime))
        valid = [p for p in matches if zipfile.is_zipfile(p)]
        if valid:
            non_backup = [p for p in valid if "backup" not in p.name.lower()]
            if non_backup:
                return non_backup[0]
            return sorted(valid, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return matches[0]


def text_blob(record: dict[str, Any], fields: list[str] | None = None) -> str:
    values = []
    keys = fields or list(record.keys())
    for key in keys:
        value = record.get(key)
        if value is not None and norm(value):
            values.append(norm(value))
    return " | ".join(values)


def apply_product_fact_override(record: dict[str, Any]) -> dict[str, Any]:
    override = PRODUCT_FACT_OVERRIDES.get(norm(record.get("Record_ID")))
    if not override:
        return record
    for key, value in override.items():
        if not key.startswith("_"):
            record[key] = value
    note = norm(override.get("_note"))
    if note:
        existing = norm(record.get("Verification_Notes"))
        record["Verification_Notes"] = " | ".join(x for x in [existing, note] if x)
    return record


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(norm(part).lower() for part in parts if norm(part))
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12] if raw else "0" * 12
    return f"{prefix}_{digest}"


def compact_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", norm(value).lower())


CANONICAL_COMPANY_NAME_OVERRIDES = {
    "deka": "DEKA",
    "candelamedical": "Candela",
}


def canonical_company_name(value: Any) -> str:
    text = norm(value)
    if not text:
        return ""
    return CANONICAL_COMPANY_NAME_OVERRIDES.get(compact_key(text), text)


def canonicalize_company_fields(record: dict[str, Any]) -> dict[str, Any]:
    """Keep generated evidence tables aligned to canonical company identities."""
    for name_field, id_field in (("company", "company_id"), ("related_company", "related_company_id")):
        original = norm(record.get(name_field))
        canonical = canonical_company_name(original)
        if canonical and canonical != original:
            record[name_field] = canonical
            if id_field in record:
                record[id_field] = stable_id("co", canonical)
    listed_parent = canonical_company_name(record.get("listed_parent_company"))
    if listed_parent and listed_parent != norm(record.get("listed_parent_company")):
        record["listed_parent_company"] = listed_parent
    return record


def safe_int(value: Any) -> int:
    number = safe_float(value)
    return int(number) if number is not None else 0


def split_stock_code(stock_code: Any) -> dict[str, str]:
    text = norm(stock_code)
    if not text:
        return {"exchange": "", "ticker_symbol": "", "listing_country": ""}
    exchange = ""
    ticker = text
    if ":" in text:
        exchange, ticker = [part.strip() for part in text.split(":", 1)]
    elif "." in text:
        ticker = text
        suffix = text.rsplit(".", 1)[-1].upper()
        exchange = {
            "MI": "Borsa Italiana",
            "SW": "SIX",
            "HK": "HKEX",
            "KS": "KRX",
            "KQ": "KOSDAQ",
            "SZ": "SZSE",
            "SS": "SSE",
        }.get(suffix, suffix)
    exchange_upper = exchange.upper()
    listing_country = {
        "NASDAQ": "US",
        "NYSE": "US",
        "SIX": "CH",
        "HKEX": "HK",
        "KRX": "KR",
        "KOSDAQ": "KR",
        "SZSE": "CN",
        "SSE": "CN",
        "TSE": "JP",
        "JPX": "JP",
        "LSE": "UK",
        "EURONEXT": "EU",
        "BORSA ITALIANA": "IT",
    }.get(exchange_upper, "")
    return {"exchange": exchange, "ticker_symbol": ticker, "listing_country": listing_country}


def normalize_country_name(country: Any) -> str:
    text = norm(country)
    return COUNTRY_ALIASES.get(text, text)


def dashboard_region(region: Any, country: Any) -> str:
    text = norm(region)
    country_name = normalize_country_name(country)
    if text and text.lower() not in {"other", "unknown", "unclassified", "其他"}:
        return text
    return DASHBOARD_REGION_BY_COUNTRY.get(country_name, text or "Other")


def split_city_country(location_full: Any, hq_country: Any) -> tuple[str, str]:
    location = norm(location_full)
    country = normalize_country_name(hq_country)
    if not location:
        return "", country
    parts = [norm(part) for part in location.split(",") if norm(part)]
    if not parts:
        return "", country
    if len(parts) > 1:
        location_country = normalize_country_name(parts[-1])
        if location_country in COUNTRY_COORDS:
            country = location_country
    if not country and len(parts) > 1:
        country = normalize_country_name(parts[-1])
    return parts[0], country


def company_geo_point(company: dict[str, Any], regulatory_channels: set[str] | None = None) -> dict[str, Any] | None:
    city, country = split_city_country(company.get("location_full"), company.get("hq_country"))
    coord = CITY_COORDS.get(compact_key(f"{city}|{country}")) if city and country else None
    precision = "city" if coord else "country"
    if not coord and country:
        coord = COUNTRY_COORDS.get(country)
    if not coord:
        return None
    lat, lon = coord
    return {
        "company_id": company.get("company_id"),
        "company": company.get("canonical_name"),
        "city": city,
        "country": country,
        "region": dashboard_region(company.get("region"), country),
        "location_full": company.get("location_full"),
        "lat": lat,
        "lon": lon,
        "precision": precision,
        "products": safe_int(company.get("product_count")),
        "brands": safe_int(company.get("brand_count")),
        "primary_track": company.get("primary_track"),
        "stock_code": company.get("stock_code"),
        "ownership": company.get("ownership"),
        "regulatory_channels": ", ".join(sorted(regulatory_channels or [])),
        "priority_rank": company.get("priority_rank") or "",
        "review_status": company.get("review_status"),
    }


def build_geo_companies(company_master: list[dict[str, Any]], products: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    regulatory_by_company: dict[str, set[str]] = defaultdict(set)
    for product in products or []:
        company_name = product.get("Company") or ""
        regulatory_by_company[company_name].update(regulatory_channel_list(product))

    rows = [
        point
        for company in company_master
        if (
            point := company_geo_point(
                company,
                regulatory_by_company.get(company.get("canonical_name") or "", set()),
            )
        )
    ]
    city_counter: Counter[str] = Counter()
    city_products: Counter[str] = Counter()
    city_points: dict[str, dict[str, Any]] = {}
    for point in rows:
        label = f"{point.get('city') or point.get('country')}, {point.get('country')}".strip(", ")
        key = compact_key(label)
        city_counter[key] += 1
        city_products[key] += safe_int(point.get("products"))
        if key not in city_points:
            city_points[key] = {
                "name": label,
                "city": point.get("city"),
                "country": point.get("country"),
                "region": point.get("region"),
                "lat": point.get("lat"),
                "lon": point.get("lon"),
                "precision": point.get("precision"),
                "companies": [],
            }
        city_points[key]["companies"].append(
            {
                "company": point.get("company"),
                "products": safe_int(point.get("products")),
                "brands": safe_int(point.get("brands")),
                "stock_code": point.get("stock_code"),
                "primary_track": point.get("primary_track"),
                "regulatory_channels": point.get("regulatory_channels"),
            }
        )

    clusters = []
    for key, count in city_counter.items():
        item = city_points[key]
        item["companies"].sort(key=lambda row: (safe_int(row.get("products")), norm(row.get("company")).lower()), reverse=True)
        clusters.append({**item, "company_count": count, "product_count": city_products[key]})
    clusters.sort(key=lambda item: (item["company_count"], item["product_count"], item["name"]), reverse=True)

    city_precision = sum(1 for point in rows if point.get("precision") == "city")
    country_precision = sum(1 for point in rows if point.get("precision") == "country")
    return {
        "companies": sorted(rows, key=lambda item: (safe_int(item.get("products")), norm(item.get("company")).lower()), reverse=True),
        "points": clusters,
        "city_clusters": clusters[:24],
        "summary": {
            "mapped_companies": len(rows),
            "total_companies": len(company_master),
            "city_precision": city_precision,
            "country_precision": country_precision,
            "unmapped_companies": len(company_master) - len(rows),
            "unique_geo_points": len(city_counter),
        },
    }


def alias_list(*values: Any) -> list[str]:
    aliases: list[str] = []
    for value in values:
        text = norm(value)
        if not text:
            continue
        for part in re.split(r"[;,/|]+", text):
            cleaned = norm(part)
            if cleaned and cleaned not in aliases:
                aliases.append(cleaned)
    return aliases


def company_ids(companies: list[dict[str, Any]]) -> dict[str, str]:
    return {company.get("Company") or "": stable_id("co", company.get("Company")) for company in companies if company.get("Company")}


def select_priority_companies(companies: list[dict[str, Any]], limit: int = 30) -> list[dict[str, Any]]:
    ranked = sorted(
        companies,
        key=lambda item: (
            safe_int(item.get("Product_Count")),
            1 if norm(item.get("Stock_Code")) else 0,
            norm(item.get("Company")).lower(),
        ),
        reverse=True,
    )
    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for company in ranked:
        name = company.get("Company")
        if not name or name in seen:
            continue
        selected.append(company)
        seen.add(name)
        if len(selected) >= limit:
            break
    public_high = [
        item
        for item in companies
        if norm(item.get("Stock_Code")) and safe_int(item.get("Product_Count")) >= 5 and item.get("Company") not in seen
    ]
    for company in sorted(public_high, key=lambda item: safe_int(item.get("Product_Count")), reverse=True):
        selected.append(company)
        seen.add(company.get("Company") or "")
    return selected


def listed_relationship_lookup() -> dict[str, dict[str, str]]:
    rows = load_generated_csv(LISTED_COMPANY_BATCH_PATH) if LISTED_COMPANY_BATCH_PATH.exists() else []
    lookup: dict[str, dict[str, str]] = {}
    for row in rows:
        company = norm(row.get("company"))
        company_id = norm(row.get("company_id"))
        if not company and not company_id:
            continue
        listed_entity = norm(row.get("listed_entity_name"))
        relation = norm(row.get("relation_to_listed_entity"))
        parent = norm(row.get("parent_company_seed"))
        ultimate = norm(row.get("ultimate_parent_seed"))
        if relation == "subsidiary_or_affiliate" and listed_entity and listed_entity.lower() != company.lower():
            parent = parent or listed_entity
            ultimate = ultimate if ultimate and ultimate.lower() != company.lower() else listed_entity
        payload = {
            "parent_company": parent,
            "ultimate_parent": ultimate,
            "acquisition_status": "listed_relationship_seed" if parent or (ultimate and ultimate.lower() != company.lower()) else "",
            "acquisition_timeline": "",
        }
        if company:
            lookup[company] = payload
        if company_id:
            lookup[company_id] = payload
    return lookup


def build_company_master(companies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected_names = {item.get("Company"): rank for rank, item in enumerate(select_priority_companies(companies), start=1)}
    listed_relationships = listed_relationship_lookup()
    rows = []
    for company in companies:
        stock = split_stock_code(company.get("Stock_Code"))
        aliases = alias_list(company.get("Company"), company.get("Parent_Company"))
        company_id = stable_id("co", company.get("Company"))
        listed_relationship = listed_relationships.get(company_id) or listed_relationships.get(norm(company.get("Company"))) or {}
        parent_company = company.get("Parent_Company") or listed_relationship.get("parent_company")
        ultimate_parent = (
            company.get("Parent_Company")
            or listed_relationship.get("ultimate_parent")
            or listed_relationship.get("parent_company")
            or company.get("Company")
        )
        rows.append(
            {
                "company_id": company_id,
                "canonical_name": company.get("Company"),
                "aliases": json.dumps(aliases, ensure_ascii=False),
                "hq_country": company.get("HQ_Country"),
                "countries_all": company.get("Countries_All"),
                "region": dashboard_region(company.get("Region"), company.get("HQ_Country")),
                "location_full": company.get("Location_Full"),
                "ownership": company.get("Ownership"),
                "business_role": company.get("Business_Role"),
                "status": company.get("Status"),
                "parent_company": parent_company,
                "positioning_cn": company.get("Positioning_CN"),
                "ultimate_parent": ultimate_parent,
                "acquisition_status": listed_relationship.get("acquisition_status") or "needs_verification",
                "acquisition_timeline": listed_relationship.get("acquisition_timeline") or "",
                "stock_code": company.get("Stock_Code"),
                "exchange": stock["exchange"],
                "ticker_symbol": stock["ticker_symbol"],
                "listing_country": stock["listing_country"],
                "isin": "",
                "product_count": safe_int(company.get("Product_Count")),
                "brand_count": safe_int(company.get("Brand_Count")),
                "primary_track": company.get("Primary_Track"),
                "priority_rank": selected_names.get(company.get("Company")),
                "verification_status": "unverified_seed",
                "review_status": "queued" if company.get("Company") in selected_names else "backlog",
                "source_status": "seed_from_workbook",
                "search_queries": json.dumps(alias_list(company.get("Company"), company.get("Stock_Code")), ensure_ascii=False),
                "search_blob": text_blob(company),
            }
        )
    return rows


def product_id_for(product: dict[str, Any]) -> str:
    return product.get("Product_UUID") or stable_id(
        "prod",
        product.get("Record_ID"),
        product.get("Company"),
        product.get("Brand"),
        product.get("Core_Product"),
    )


MOJIBAKE_MARKERS = {"镞", "庠", "剌", "鐨", "涓", "妯", "傛", "鍏", "鏃", "缇"}


def looks_like_mojibake(text: str) -> bool:
    if not text:
        return False
    return sum(1 for marker in MOJIBAKE_MARKERS if marker in text) >= 2


def derive_claim_differentiator(claim_text: Any) -> str:
    text = norm(claim_text)
    if not text or looks_like_mojibake(text):
        return ""
    bracket = re.match(r"^\s*\[([^\]]{2,80})\]", text)
    if bracket:
        return bracket.group(1).strip()
    text = re.sub(r"\[[^\]]+\]\s*", "", text)
    first = re.split(r"[。.!?；;]\s*", text, maxsplit=1)[0].strip()
    first = re.sub(r"\s+", " ", first)
    if 5 <= len(first) <= 120:
        return first
    if len(first) > 120:
        return first[:120].rsplit(" ", 1)[0].strip() or first[:80].strip()
    return ""


FEATURE_TAG_RULES = [
    ("HA", ["hyaluronic", "透明质酸", "玻尿酸", " ha "]),
    ("CaHA", ["calcium hydroxylapatite", "caha", "羟基磷灰石钙"]),
    ("PLLA", ["plla", "poly-l-lactic", "聚左旋乳酸"]),
    ("PCL", ["pcl", "polycaprolactone", "聚己内酯"]),
    ("PN/PDRN", ["pdrn", "pn", "polynucleotide", "聚核苷酸"]),
    ("Exosome", ["exosome", "外泌体"]),
    ("肉毒毒素", ["botulinum", "botox", "toxin", "肉毒"]),
    ("Thread lift", ["thread", "pdo", "埋线", "线雕"]),
    ("Laser", ["laser", "激光", "nd:yag", "diode", "alexandrite", "picosecond", "co2", "er:yag"]),
    ("RF", ["radiofrequency", " radio frequency", " rf ", "射频"]),
    ("HIFU", ["hifu", "focused ultrasound", "聚焦超声"]),
    ("IPL", ["ipl", "intense pulsed light", "强脉冲光"]),
    ("LED", [" led ", "photobiomodulation", "光疗"]),
    ("Microneedling", ["microneedling", "micro-needling", "微针"]),
    ("Cryotherapy", ["cryo", "cooling", "冷冻", "低温"]),
    ("Skincare", ["skincare", "skin care", "护肤", "精华", "面霜"]),
    ("Diagnostics", ["diagnostic", "imaging", "ai skin", "analysis", "检测", "诊断"]),
    ("Implant", ["implant", "breast", "植入"]),
    ("Cannula/needle", ["cannula", "needle", "针", "套管"]),
    ("OEM/ODM", ["oem", "odm", "contract manufacturing", "代工"]),
    ("Pharmacy channel", ["pharmacy", "drugstore", "药房"]),
]


def derive_feature_tags(product: dict[str, Any], claim_text: Any) -> str:
    values = [
        product.get("Category_L1"),
        product.get("Category_L2"),
        product.get("Tech_Type_Std"),
        product.get("Tech_Type_Original"),
        product.get("Core_Product"),
        claim_text,
    ]
    blob = f" {' | '.join(norm(value) for value in values).lower()} "
    if looks_like_mojibake(blob):
        blob = " ".join(norm(value).lower() for value in values[:-1])
    tags: list[str] = []

    def add(tag: str) -> None:
        if tag and tag not in tags:
            tags.append(tag)

    def normalize_feature_tag(tag: str) -> str:
        if tag == "Neurotoxin":
            return "肉毒毒素"
        if "Neurotoxin" in tag and "零复合蛋白肉毒" in tag:
            return "零复合蛋白肉毒"
        return tag

    for value in values[:4]:
        cleaned = normalize_feature_tag(norm(value))
        if cleaned and cleaned.lower() not in {"needs_verification", "true", "false"}:
            add(cleaned)
    for tag, terms in FEATURE_TAG_RULES:
        if any(term in blob for term in terms):
            add(tag)
    return ", ".join(tags[:8])


def derive_material_family(l1: str, l2: str, l3: str) -> str:
    if l3:
        return l3
    if l2:
        return l2
    return l1


def material_taxonomy(product: dict[str, Any]) -> dict[str, str]:
    l1 = norm(product.get("Material_Taxonomy_L1_CN"))
    l2 = norm(product.get("Material_Taxonomy_L2_CN"))
    l3 = norm(product.get("Material_Taxonomy_L3_CN"))
    if l1 == "神经毒素":
        l1 = "肉毒毒素"
    if l1 == "外用/护肤":
        l1 = "功效性护肤品"
    if l1 == "功效性护肤品":
        if l2 in {"医学护肤", "外用药"}:
            l2 = "医学护肤活性"
        if l3 in {"活性成分护肤", "外用处方药"}:
            l3 = "功效活性成分"
    l2 = {
        "光 / IPL": "光/IPL",
        "温控 / 其他": "温控/其他",
    }.get(l2, l2)
    l3 = {
        "LED光疗 / 光动力 PDT": "LED光疗/光动力 PDT",
        "等离子束 / 等离子笔": "等离子束/等离子笔",
        "剥脱点阵（CO₂/铒）": "剥脱点阵(CO₂/铒)",
        "射频微针（有创）": "射频微针(有创)",
        "聚焦超声减脂（体雕）": "聚焦超声减脂",
        "微聚焦超声 HIFU/MFU（提升）": "微聚焦超声 HIFU/MFU",
        "注射针 / 钝头套管": "注射针/钝头套管",
        "无针注射/电穿孔": "无针注射/电穿孔等",
    }.get(l3, l3)
    path = norm(product.get("Material_Taxonomy_Path_CN")) or " > ".join(x for x in [l1, l2, l3] if x)
    path = " > ".join(x for x in [l1, l2, l3] if x) or path
    return {
        "l1": l1,
        "l2": l2,
        "l3": l3,
        "path": path,
        "source": norm(product.get("Material_Taxonomy_Source")),
        "confidence": norm(product.get("Material_Taxonomy_Confidence")),
        "review_status": norm(product.get("Material_Taxonomy_Review_Status")),
        "note": norm(product.get("Material_Taxonomy_Note")),
        "family": norm(product.get("Material_Family")) or derive_material_family(l1, l2, l3),
        "inclusion_status": norm(product.get("Inclusion_Status")) or "active",
        "backfill_audit": norm(product.get("Backfill_Audit")),
    }


MATERIAL_TAXONOMY_COLORS = [
    "#DA7756",
    "#4A7C8E",
    "#7BA99C",
    "#B87333",
    "#8E6E95",
    "#5D7E5B",
    "#C29B40",
    "#6F7C91",
    "#9E6A55",
    "#4F8C7A",
    "#8C6D31",
    "#577590",
]


def taxonomy_sample_product(product: dict[str, Any]) -> dict[str, str]:
    material = material_taxonomy(product)
    claim_text = norm(product.get("Introduction"))
    return {
        "record_id": norm(product.get("Record_ID")),
        "product_id": product_id_for(product),
        "company": norm(product.get("Company")),
        "brand": norm(product.get("Brand")),
        "product": norm(product.get("Core_Product")) or norm(product.get("Brand")),
        "country": norm(product.get("Country")),
        "region": dashboard_region(product.get("Region"), product.get("Country")),
        "category_l1": norm(product.get("Category_L1")),
        "category_l2": norm(product.get("Category_L2")),
        "tech": norm(product.get("Tech_Type_Std")),
        "tech_detail": norm(product.get("Tech_Type_Original")),
        "product_type_cn": norm(product.get("Verified_Product_Type_CN")),
        "market_channel": norm(product.get("Market_Channel")),
        "feature_tags": norm(product.get("Feature_Tags")) or derive_feature_tags(product, claim_text),
        "differentiator": derive_claim_differentiator(claim_text),
        "intro": claim_text,
        "l1": material["l1"],
        "l2": material["l2"],
        "l3": material["l3"],
        "family": material["family"],
        "review_status": material["review_status"],
        "inclusion_status": material["inclusion_status"],
    }


def material_taxonomy_structure(products: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(products) or 1
    l1_groups: dict[str, dict[str, Any]] = {}

    for product in products:
        material = material_taxonomy(product)
        l1 = material["l1"] or "未归类/待补充"
        l2 = material["l2"] or "未归类"
        l3 = material["l3"] or ""
        company = norm(product.get("Company")) or "Unknown"
        brand = norm(product.get("Brand")) or norm(product.get("Core_Product")) or "Unknown"
        country = norm(product.get("Country")) or "Unknown"
        region = dashboard_region(product.get("Region"), product.get("Country"))
        product_preview = taxonomy_sample_product(product)

        l1_bucket = l1_groups.setdefault(
            l1,
            {
                "name": l1,
                "products": 0,
                "companies": set(),
                "brands": set(),
                "countries": set(),
                "regions": Counter(),
                "company_counter": Counter(),
                "country_counter": Counter(),
                "l2": {},
                "l3_counter": Counter(),
                "samples": [],
            },
        )
        l2_bucket = l1_bucket["l2"].setdefault(
            l2,
            {
                "name": l2,
                "products": 0,
                "companies": set(),
                "brands": set(),
                "countries": set(),
                "regions": Counter(),
                "company_counter": Counter(),
                "country_counter": Counter(),
                "l3_counter": Counter(),
                "samples": [],
            },
        )

        for bucket in (l1_bucket, l2_bucket):
            bucket["products"] += 1
            bucket["companies"].add(company)
            bucket["brands"].add(brand)
            bucket["countries"].add(country)
            bucket["regions"][region] += 1
            bucket["company_counter"][company] += 1
            bucket["country_counter"][country] += 1
            if l3:
                bucket["l3_counter"][l3] += 1
            bucket["samples"].append(product_preview)
        if l3:
            l1_bucket["l3_counter"][l3] += 1

    rows: list[dict[str, Any]] = []
    l2_flat: list[dict[str, Any]] = []
    for index, (l1_name, bucket) in enumerate(sorted(l1_groups.items(), key=lambda item: (-item[1]["products"], item[0]))):
        color = MATERIAL_TAXONOMY_COLORS[index % len(MATERIAL_TAXONOMY_COLORS)]
        l2_rows = []
        for l2_name, l2_bucket in sorted(bucket["l2"].items(), key=lambda item: (-item[1]["products"], item[0])):
            row = {
                "id": stable_id("tax_l2", l1_name, l2_name),
                "parent_id": stable_id("tax_l1", l1_name),
                "l1": l1_name,
                "name": l2_name,
                "products": l2_bucket["products"],
                "value": l2_bucket["products"],
                "share_of_l1": round(l2_bucket["products"] / (bucket["products"] or 1), 4),
                "share_of_total": round(l2_bucket["products"] / total, 4),
                "companies": len(l2_bucket["companies"]),
                "brands": len(l2_bucket["brands"]),
                "countries": len(l2_bucket["countries"]),
                "top_l3": top_counts(l2_bucket["l3_counter"], 8),
                "top_companies": top_counts(l2_bucket["company_counter"], 8),
                "top_countries": top_counts(l2_bucket["country_counter"], 8),
                "top_regions": top_counts(l2_bucket["regions"], 6),
                "samples": l2_bucket["samples"],
            }
            l2_rows.append(row)
            l2_flat.append(row)
        rows.append(
            {
                "id": stable_id("tax_l1", l1_name),
                "name": l1_name,
                "color": color,
                "products": bucket["products"],
                "value": bucket["products"],
                "share_of_total": round(bucket["products"] / total, 4),
                "companies": len(bucket["companies"]),
                "brands": len(bucket["brands"]),
                "countries": len(bucket["countries"]),
                "l2_count": len(l2_rows),
                "top_l2": [
                    {key: row[key] for key in ["id", "name", "products", "value", "share_of_l1", "companies", "brands", "countries"]}
                    for row in l2_rows[:8]
                ],
                "l2": l2_rows,
                "top_l3": top_counts(bucket["l3_counter"], 10),
                "top_companies": top_counts(bucket["company_counter"], 8),
                "top_countries": top_counts(bucket["country_counter"], 8),
                "top_regions": top_counts(bucket["regions"], 6),
                "samples": bucket["samples"],
            }
        )

    return {
        "total_products": len(products),
        "l1_count": len(rows),
        "l2_count": len(l2_flat),
        "l1": rows,
        "l2_flat": sorted(l2_flat, key=lambda row: (-row["products"], row["l1"], row["name"])),
    }


SKU_VARIANT_STOPWORDS = {
    "and",
    "or",
    "with",
    "without",
    "tech",
    "technology",
    "treatment",
    "system",
    "series",
    "line",
    "family",
    "shape",
    "facial",
    "body",
    "product",
    "products",
    "device",
    "devices",
    "gel",
    "filler",
    "fillers",
    "ha",
    "caha",
    "pcl",
    "pdo",
    "plla",
    "pdlla",
    "pdrn",
    "pn",
    "la",
    "cl",
    "co2",
    "er:yag",
    "er yag",
    "nd:yag",
    "nd yag",
    "yag",
    "rf",
    "ipl",
    "dpl",
    "opt",
    "hifu",
    "laser",
    "ultrasound",
    "plasma",
    "collagen",
    "分销",
    "关联",
    "卡波西",
    "二氧化碳疗法",
    "二氧化碳",
    "疗法",
    "线雕",
    "黄金微针",
    "lidocaine",
    "technique",
}


SKU_DESCRIPTOR_PHRASES = {
    "deep dermis",
    "middle dermis",
    "mid-to-deep dermis",
    "subcutaneous",
    "nasolabial fold",
    "nasolabial folds",
    "subdermal",
    "dermis",
    "epidermis",
}


def is_noise_sku_variant(token: str) -> bool:
    token_key = token.casefold().strip()
    if token_key in SKU_VARIANT_STOPWORDS or token_key in SKU_DESCRIPTOR_PHRASES:
        return True
    if re.search(r"[\u4e00-\u9fff]", token_key) and not re.search(r"[A-Za-z0-9]", token_key):
        # Chinese-only parenthetical fragments in the seed are usually
        # translations, application notes, or channel notes rather than SKU
        # model names.
        return True
    if re.fullmatch(r"[A-Z]{1,2}", token.strip()) and token_key not in {"s", "m", "l", "e"}:
        return True
    return False


def sku_variant_tokens(value: Any) -> list[str]:
    text = norm(value)
    if not text:
        return []
    variants: list[str] = []
    for match in re.finditer(r"\(([^()]{3,120})\)", text):
        inside = match.group(1)
        if not re.search(r"[/,;+]|\band\b", inside, flags=re.IGNORECASE):
            continue
        for part in re.split(r"[/,;+]|\band\b", inside, flags=re.IGNORECASE):
            token = re.sub(r"\s+", " ", part).strip(" -_")
            token_key = token.casefold()
            if not token or len(token) > 42:
                continue
            if is_noise_sku_variant(token):
                continue
            if not re.search(r"[A-Za-z0-9\u4e00-\u9fff]", token):
                continue
            variants.append(token)
    seen: list[str] = []
    for token in variants:
        key = token.casefold()
        if key not in {item.casefold() for item in seen}:
            seen.append(token)
    return seen[:12]


def derived_model_or_sku(product: dict[str, Any]) -> str:
    explicit = norm(product.get("Model_or_SKU") or product.get("Model_SKU"))
    if explicit:
        return explicit
    text = " / ".join(x for x in [product.get("Brand"), product.get("Core_Product")] if norm(x))
    variants = sku_variant_tokens(text)
    if variants:
        return "; ".join(variants)
    return ""


def build_product_master(products: list[dict[str, Any]], company_id_map: dict[str, str]) -> list[dict[str, Any]]:
    rows = []
    for product in products:
        claim_text = norm(product.get("Introduction"))
        commercial_path = " > ".join(x for x in [product.get("Category_L1"), product.get("Category_L2")] if norm(x))
        technology_path = " > ".join(x for x in [product.get("Tech_Type_Std"), product.get("Tech_Type_Original")] if norm(x))
        material = material_taxonomy(product)
        source_marker = norm(product.get("Data_Source"))
        official_override = source_marker == "official_company_fact_override"
        official_fact_promoted = source_marker == "official_product_fact_promoted"
        taxonomy_correction = source_marker == "taxonomy_conflict_correction"
        verification_status = "unverified_seed"
        source_status = "seed_from_workbook"
        if official_override:
            verification_status = "official_commercial_fact_corrected"
            source_status = "official_company_fact_override"
        elif official_fact_promoted:
            verification_status = "official_product_and_spec_cross_checked"
            source_status = "seed_from_workbook, official_product_page, official_product_fact_promoted"
        elif taxonomy_correction:
            verification_status = "taxonomy_conflict_corrected"
            source_status = "taxonomy_conflict_correction"
        rows.append(
            {
                "product_id": product_id_for(product),
                "seed_record_id": product.get("Record_ID"),
                "company_id": company_id_map.get(product.get("Company") or "", stable_id("co", product.get("Company"))),
                "company": product.get("Company"),
                "brand": product.get("Brand"),
                "brand_role": product.get("Brand_Type") or "Product",
                "standard_product_name": product.get("Core_Product") or product.get("Brand"),
                "registered_name": "",
                "model_or_sku": derived_model_or_sku(product),
                "commercial_path_l1": product.get("Category_L1"),
                "commercial_path_l2": product.get("Category_L2"),
                "material_taxonomy_l1_cn": material["l1"],
                "material_taxonomy_l2_cn": material["l2"],
                "material_taxonomy_l3_cn": material["l3"],
                "material_taxonomy_path_cn": material["path"],
                "material_taxonomy_confidence": material["confidence"],
                "material_taxonomy_review_status": material["review_status"],
                "material_family": material["family"],
                "inclusion_status": material["inclusion_status"],
                "technology_path_l1": product.get("Tech_Type_Std"),
                "technology_path_l2": product.get("Tech_Type_Original"),
                "material_or_energy_source": product.get("Tech_Type_Std") or product.get("Tech_Type_Original"),
                "core_product": product.get("Core_Product"),
                "legal_manufacturer": product.get("Manufactured_By") or product.get("Company"),
                "marketing_holder": product.get("Company"),
                "local_holder": "",
                "oem_for": product.get("OEM_For"),
                "manufactured_by": product.get("Manufactured_By"),
                "r_and_d_origin_status": "needs_verification",
                "claim_text": claim_text,
                "verified_differentiator": derive_claim_differentiator(claim_text),
                "feature_tags": product.get("Feature_Tags") or derive_feature_tags(product, claim_text),
                "technical_specs_json": "",
                "technical_specs_summary": "",
                "spec_evidence_count": "",
                "spec_review_status": "",
                "classification_layer": json.dumps(
                    {
                        "commercial": commercial_path,
                        "material_taxonomy": material["path"],
                        "material_taxonomy_detail": material,
                        "material_family": material["family"],
                        "inclusion_status": material["inclusion_status"],
                        "technology": technology_path,
                        "regulatory": "pending_registration_evidence",
                    },
                    ensure_ascii=False,
                ),
                "verification_status": verification_status,
                "review_status": "queued" if any(has_value(product, field) for field in ["FDA_Status", "NMPA_Status", "CE_Status", "KFDA_Status"]) else "backlog",
                "source_status": source_status,
                "search_blob": text_blob(product),
            }
        )
    return rows


def duplicate_of_record(product: dict[str, Any]) -> str:
    note = norm(product.get("Duplicate_Note"))
    match = re.search(r"duplicate_of\s*:\s*([A-Za-z0-9_-]+)", note, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def hierarchy_split_status(product: dict[str, Any]) -> str:
    text = " ".join([norm(product.get("Brand")), norm(product.get("Core_Product"))])
    # Parenthetical text often contains translations, chemistry notes, or
    # channel notes. Keep it out of the broad hierarchy flag; true SKU splits
    # are already handled by derived_model_or_sku().
    status_text = re.sub(r"\([^()]*\)", " ", text)
    status_text = re.sub(r"\s+", " ", status_text).strip()
    if re.search(r"\b(gen|model|type|lite|deep|ultra|plus|pro|max|xt|h2|a3|300|gen\d+)\b", status_text, flags=re.IGNORECASE):
        return "sku_or_model_candidate"
    if "/" in status_text or "+" in status_text:
        return "variant_catalogued"
    return "family_level"


def regulatory_channel_list(product: dict[str, Any]) -> list[str]:
    channels = []
    if has_value(product, "FDA_Status") or has_value(product, "FDA_510k_Number"):
        channels.append("FDA")
    if has_value(product, "CE_Status") or has_value(product, "CE_Year"):
        channels.append("CE")
    if has_value(product, "NMPA_Status") or has_value(product, "NMPA_Reg_Number"):
        channels.append("NMPA")
    if has_value(product, "KFDA_Status"):
        channels.append("KFDA")
    return channels


def build_product_hierarchy(products: list[dict[str, Any]], company_id_map: dict[str, str]) -> dict[str, list[dict[str, Any]]]:
    record_by_id = {product.get("Record_ID"): product for product in products if product.get("Record_ID")}
    families: dict[str, dict[str, Any]] = {}
    family_sets: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    skus: list[dict[str, Any]] = []

    for product in products:
        record_id = product.get("Record_ID") or stable_id("row", text_blob(product, ["Company", "Brand", "Core_Product"]))
        duplicate_of = duplicate_of_record(product)
        canonical = record_by_id.get(duplicate_of) or product
        canonical_material = material_taxonomy(canonical)
        product_material = material_taxonomy(product)
        company = canonical.get("Company")
        brand = canonical.get("Brand")
        family_name = canonical.get("Core_Product") or brand
        family_id = stable_id(
            "pf",
            company,
            brand,
            family_name,
            canonical.get("Category_L1"),
            canonical.get("Category_L2"),
            canonical.get("Tech_Type_Std"),
        )
        if family_id not in families:
            families[family_id] = {
                "product_family_id": family_id,
                "company_id": company_id_map.get(company or "", stable_id("co", company)),
                "company": company,
                "brand": brand,
                "brand_type": canonical.get("Brand_Type"),
                "product_family": family_name,
                "category_l1": canonical.get("Category_L1"),
                "category_l2": canonical.get("Category_L2"),
                "material_taxonomy_l1_cn": canonical_material["l1"],
                "material_taxonomy_l2_cn": canonical_material["l2"],
                "material_taxonomy_l3_cn": canonical_material["l3"],
                "material_taxonomy_path_cn": canonical_material["path"],
                "material_taxonomy_confidence": canonical_material["confidence"],
                "material_taxonomy_review_status": canonical_material["review_status"],
                "material_family": canonical_material["family"],
                "inclusion_status": canonical_material["inclusion_status"],
                "tech_type": canonical.get("Tech_Type_Std"),
                "material_or_energy_source": canonical.get("Tech_Type_Std") or canonical.get("Tech_Type_Original"),
                "primary_record_count": 0,
                "duplicate_record_count": 0,
                "sku_candidate_count": 0,
                "countries": "",
                "source_record_ids": "",
                "duplicate_record_ids": "",
                "sku_candidate_names": "",
                "regulatory_channels": "",
                "hierarchy_status": "seed_hierarchy",
                "review_status": "needs_review",
                "source_status": "derived_from_product_lines",
                "search_blob": "",
            }
        family = families[family_id]
        family_sets[family_id]["countries"].add(product.get("Country") or "")
        family_sets[family_id]["source_record_ids"].add(record_id)
        family_sets[family_id]["sku_candidate_names"].add(" / ".join(x for x in [product.get("Brand"), product.get("Core_Product")] if norm(x)))
        for channel in regulatory_channel_list(product):
            family_sets[family_id]["regulatory_channels"].add(channel)
        if is_non_primary_record(product):
            family["duplicate_record_count"] += 1
            family_sets[family_id]["duplicate_record_ids"].add(record_id)
        else:
            family["primary_record_count"] += 1
        model_candidates = [norm(part) for part in derived_model_or_sku(product).split(";") if norm(part)]
        active_for_split = product_material["inclusion_status"] not in {"deleted", "excluded"}
        family["sku_candidate_count"] += max(1, len(model_candidates) if active_for_split else 1)
        split_status = "family_sku_expanded" if len(model_candidates) > 1 else hierarchy_split_status(product)
        if is_non_primary_record(product):
            sku_review_status = "duplicate_non_primary"
        elif split_status == "family_level":
            sku_review_status = "usable_trace"
        elif split_status in {"family_sku_expanded", "variant_catalogued"}:
            sku_review_status = "auto_split_completed"
        else:
            sku_review_status = "needs_review"

        skus.append(
            {
                "sku_id": product_id_for(product),
                "product_family_id": family_id,
                "company_id": company_id_map.get(product.get("Company") or "", stable_id("co", product.get("Company"))),
                "company": product.get("Company"),
                "brand": product.get("Brand"),
                "product_family": product.get("Core_Product") or product.get("Brand"),
                "model_or_sku": "; ".join(model_candidates),
                "sku_candidate_name": " / ".join(x for x in [product.get("Brand"), product.get("Core_Product")] if norm(x)),
                "seed_record_id": record_id,
                "is_primary_record": "0" if is_non_primary_record(product) else "1",
                "duplicate_of_record_id": duplicate_of,
                "category_l1": product.get("Category_L1"),
                "category_l2": product.get("Category_L2"),
                "material_taxonomy_l1_cn": product_material["l1"],
                "material_taxonomy_l2_cn": product_material["l2"],
                "material_taxonomy_l3_cn": product_material["l3"],
                "material_taxonomy_path_cn": product_material["path"],
                "material_taxonomy_confidence": product_material["confidence"],
                "material_taxonomy_review_status": product_material["review_status"],
                "material_family": product_material["family"],
                "inclusion_status": product_material["inclusion_status"],
                "tech_type": product.get("Tech_Type_Std"),
                "country": product.get("Country"),
                "regulatory_channels": ", ".join(regulatory_channel_list(product)),
                "split_status": split_status,
                "review_status": sku_review_status,
                "source_status": "seed_from_product_lines",
                "search_blob": text_blob(product),
            }
        )
        if active_for_split and len(model_candidates) > 1:
            base_name = norm(product.get("Brand")) or norm(product.get("Core_Product"))
            for model in model_candidates:
                skus.append(
                    {
                        "sku_id": stable_id("sku", family_id, record_id, model),
                        "product_family_id": family_id,
                        "company_id": company_id_map.get(product.get("Company") or "", stable_id("co", product.get("Company"))),
                        "company": product.get("Company"),
                        "brand": product.get("Brand"),
                        "product_family": product.get("Core_Product") or product.get("Brand"),
                        "model_or_sku": model,
                        "sku_candidate_name": " ".join(x for x in [base_name, model] if x),
                        "seed_record_id": record_id,
                        "is_primary_record": "0",
                        "duplicate_of_record_id": "",
                        "category_l1": product.get("Category_L1"),
                        "category_l2": product.get("Category_L2"),
                        "material_taxonomy_l1_cn": product_material["l1"],
                        "material_taxonomy_l2_cn": product_material["l2"],
                        "material_taxonomy_l3_cn": product_material["l3"],
                        "material_taxonomy_path_cn": product_material["path"],
                        "material_taxonomy_confidence": product_material["confidence"],
                        "material_taxonomy_review_status": product_material["review_status"],
                        "material_family": product_material["family"],
                        "inclusion_status": product_material["inclusion_status"],
                        "tech_type": product.get("Tech_Type_Std"),
                        "country": product.get("Country"),
                        "regulatory_channels": ", ".join(regulatory_channel_list(product)),
                        "split_status": "sku_variant_catalogued",
                        "review_status": "auto_split_completed",
                        "source_status": "derived_from_sku_candidate_names",
                        "search_blob": text_blob(product) + " " + model,
                    }
                )

    rows = []
    for family_id, family in families.items():
        sets = family_sets[family_id]
        family["countries"] = ", ".join(sorted(x for x in sets["countries"] if x))
        family["source_record_ids"] = ", ".join(sorted(x for x in sets["source_record_ids"] if x))
        family["duplicate_record_ids"] = ", ".join(sorted(x for x in sets["duplicate_record_ids"] if x))
        family["sku_candidate_names"] = "; ".join(sorted(x for x in sets["sku_candidate_names"] if x)[:20])
        family["regulatory_channels"] = ", ".join(sorted(x for x in sets["regulatory_channels"] if x))
        family["search_blob"] = text_blob(family)
        rows.append(family)

    rows.sort(key=lambda item: (safe_int(item.get("primary_record_count")), norm(item.get("company")).lower(), norm(item.get("brand")).lower()), reverse=True)
    skus.sort(key=lambda item: (norm(item.get("company")).lower(), norm(item.get("brand")).lower(), norm(item.get("seed_record_id")).lower()))
    return {"families": rows, "skus": skus}


def write_product_hierarchy_outputs(hierarchy: dict[str, list[dict[str, Any]]]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    for path, rows in [
        (PRODUCT_FAMILY_MASTER_PATH, hierarchy.get("families", [])),
        (PRODUCT_SKU_MASTER_PATH, hierarchy.get("skus", [])),
    ]:
        fieldnames = list(rows[0].keys()) if rows else []
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            if fieldnames:
                writer.writeheader()
                writer.writerows(rows)


def write_source_authority_policy_output() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    fieldnames = [
        "fact_group",
        "authoritative_source",
        "primary_sources",
        "supporting_sources",
        "merge_rule",
        "manual_role",
    ]
    with SOURCE_AUTHORITY_POLICY_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(SOURCE_AUTHORITY_POLICY)


def write_field_dictionary_output() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    fieldnames = ["table_name", "field_name", "definition", "display_note", "source_priority"]
    with FIELD_DICTIONARY_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(FIELD_DICTIONARY_ROWS)


def build_registration_seed(products: list[dict[str, Any]], company_id_map: dict[str, str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    checked_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    channel_map = [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status_field": "FDA_Status",
            "date_field": "FDA_Approval_Date",
            "number_field": "FDA_510k_Number",
            "source_key": "workbook_fda_seed",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE",
            "status_field": "CE_Status",
            "date_field": "CE_Year",
            "number_field": "",
            "source_key": "workbook_ce_seed",
        },
        {
            "jurisdiction": "CN",
            "regulator": "NMPA",
            "pathway": "Medical device registration",
            "status_field": "NMPA_Status",
            "date_field": "NMPA_Approval_Date",
            "number_field": "NMPA_Reg_Number",
            "source_key": "workbook_nmpa_seed",
        },
        {
            "jurisdiction": "KR",
            "regulator": "MFDS",
            "pathway": "Medical device approval",
            "status_field": "KFDA_Status",
            "date_field": "",
            "number_field": "",
            "source_key": "workbook_mfds_seed",
        },
    ]
    for product in products:
        for channel in channel_map:
            values = [
                product.get(channel["status_field"]),
                product.get(channel["date_field"]) if channel["date_field"] else "",
                product.get(channel["number_field"]) if channel["number_field"] else "",
            ]
            if not any(norm(value) for value in values):
                continue
            pathway = channel["pathway"]
            registration_no = norm(values[2])
            if channel["regulator"] == "FDA" and re.match(r"^P\d{6}(?:[/-]?S\d{3})?$", registration_no, flags=re.IGNORECASE):
                pathway = "PMA Supplement" if re.search(r"[/-]?S\d{3}$", registration_no, flags=re.IGNORECASE) else "PMA"
            rows.append(
                {
                    "product_id": product_id_for(product),
                    "seed_record_id": product.get("Record_ID"),
                    "company_id": company_id_map.get(product.get("Company") or "", stable_id("co", product.get("Company"))),
                    "company": product.get("Company"),
                    "brand": product.get("Brand"),
                    "jurisdiction": channel["jurisdiction"],
                    "regulator": channel["regulator"],
                    "regulatory_pathway": pathway,
                    "status": norm(values[0]) or "seeded",
                    "registration_no": registration_no,
                    "approval_date": norm(values[1]),
                    "expiry_date": "",
                    "registered_name": "",
                    "approved_indication": "",
                    "intended_use": "",
                    "legal_manufacturer": product.get("Manufactured_By") or product.get("Company"),
                    "local_holder": "",
                    "source_key": channel["source_key"],
                    "source_url": "",
                    "source_type": "seed_workbook",
                    "evidence_title": f"{product.get('Brand') or product.get('Core_Product')} {channel['regulator']} seed",
                    "evidence_excerpt": text_blob(product, ["FDA_Status", "FDA_Approval_Date", "FDA_510k_Number", "CE_Status", "CE_Year", "NMPA_Status", "NMPA_Approval_Date", "NMPA_Reg_Number", "KFDA_Status"]),
                    "checked_at": checked_at,
                    "reviewed_by": "",
                    "review_status": "needs_review",
                    "confidence": "seed_unverified",
                }
            )
    return rows


PRODUCT_MASTER_FIELDS = [
    "product_id",
    "seed_record_id",
    "company_id",
    "company",
    "brand",
    "brand_role",
    "standard_product_name",
    "registered_name",
    "model_or_sku",
    "commercial_path_l1",
    "commercial_path_l2",
    "material_taxonomy_l1_cn",
    "material_taxonomy_l2_cn",
    "material_taxonomy_l3_cn",
    "material_taxonomy_path_cn",
    "material_taxonomy_confidence",
    "material_taxonomy_review_status",
    "material_family",
    "inclusion_status",
    "technology_path_l1",
    "technology_path_l2",
    "material_or_energy_source",
    "core_product",
    "legal_manufacturer",
    "marketing_holder",
    "local_holder",
    "oem_for",
    "manufactured_by",
    "r_and_d_origin_status",
    "claim_text",
    "verified_differentiator",
    "feature_tags",
    "technical_specs_json",
    "technical_specs_summary",
    "spec_evidence_count",
    "spec_review_status",
    "classification_layer",
    "verification_status",
    "review_status",
    "source_status",
    "search_blob",
]


REGISTRATION_EVIDENCE_FIELDS = [
    "product_id",
    "seed_record_id",
    "company_id",
    "company",
    "brand",
    "jurisdiction",
    "regulator",
    "regulatory_pathway",
    "fda_product_code",
    "fda_regulation_number",
    "fda_device_class",
    "fda_submission_type",
    "status",
    "registration_no",
    "approval_date",
    "original_approval_date",
    "expiry_date",
    "registered_name",
    "approved_indication",
    "intended_use",
    "legal_manufacturer",
    "local_holder",
    "source_key",
    "source_url",
    "source_type",
    "evidence_title",
    "evidence_excerpt",
    "official_description_exact",
    "official_description_source_field",
    "field_note",
    "checked_at",
    "reviewed_by",
    "review_status",
    "confidence",
]


OFFICIAL_DESCRIPTION_CAPTURED_NOTE = (
    "Precise official wording from a regulator, registration certificate, PMA/510(k), "
    "EUDAMED/CE document, IFU or official labeling. Derived Chinese buckets are for "
    "dashboard analysis only and must not replace this official description."
)
OFFICIAL_DESCRIPTION_MISSING_NOTE = (
    "No precise official approved-indication or intended-use wording captured yet. "
    "Do not present this row as a confirmed approved indication until an official "
    "description is promoted."
)


def official_description_parts(row: dict[str, Any]) -> tuple[str, str, str]:
    explicit = norm(row.get("official_description_exact"))
    if explicit:
        return explicit, norm(row.get("official_description_source_field")) or "official_description_exact", OFFICIAL_DESCRIPTION_CAPTURED_NOTE
    for field in ["approved_indication", "intended_use"]:
        value = norm(row.get(field))
        if value:
            return value, field, OFFICIAL_DESCRIPTION_CAPTURED_NOTE
    return "", "", OFFICIAL_DESCRIPTION_MISSING_NOTE


def enrich_registration_description_fields(row: dict[str, Any]) -> dict[str, Any]:
    output = dict(row)
    exact, source_field, note = official_description_parts(output)
    output["official_description_exact"] = exact
    output["official_description_source_field"] = source_field
    output["field_note"] = output.get("field_note") or note
    return output


def registration_dedupe_key(row: dict[str, Any]) -> tuple[str, str, str, str, str] | None:
    product_id = norm(row.get("product_id"))
    registration_no = norm(row.get("registration_no"))
    regulator = norm(row.get("regulator")).upper()
    jurisdiction = norm(row.get("jurisdiction")).upper()
    if product_id and registration_no and (regulator or jurisdiction):
        return ("registration_no", product_id, regulator, jurisdiction, registration_no)
    company = norm(row.get("company"))
    registered_name = norm(row.get("registered_name"))
    if registration_no and (regulator or jurisdiction) and (company or registered_name):
        return ("registration_no_company", company, registered_name, regulator or jurisdiction, registration_no)
    source_url = norm(row.get("source_url"))
    source_key = norm(row.get("source_key"))
    dedupe_source_key = source_key
    if dedupe_source_key.startswith("manual_needs_review_resolution:"):
        dedupe_source_key = dedupe_source_key.split(":", 1)[1]
    if product_id and source_url and source_key:
        return ("source_url", product_id, dedupe_source_key, "", source_url)
    if source_url and source_key and (company or registered_name):
        return ("source_url_company", company, registered_name, dedupe_source_key, source_url)
    if product_id and source_key and (regulator or jurisdiction):
        return ("seed_status", product_id, dedupe_source_key, regulator or jurisdiction, "")
    return None


def registration_row_priority(row: dict[str, Any]) -> tuple[int, int, int, int]:
    source_key = norm(row.get("source_key")).lower()
    source_type = norm(row.get("source_type")).lower()
    confidence = norm(row.get("confidence")).lower()
    review_status = norm(row.get("review_status")).lower()
    score = 0
    if "manual" in source_key or "manual" in review_status:
        score += 50
    if "registration_project" in source_key or "official_nmpa_registration_project" in source_type:
        score += 40
    if source_type == "seed_workbook" or source_key.startswith("workbook_"):
        score -= 20
    if "official" in confidence:
        score += 15
    return (
        score,
        1 if norm(row.get("official_description_exact")) else 0,
        sum(1 for field in REGISTRATION_EVIDENCE_FIELDS if norm(row.get(field))),
        len(norm(row.get("evidence_excerpt"))),
    )


def merge_registration_rows(preferred: dict[str, Any], fallback: dict[str, Any]) -> dict[str, Any]:
    merged = dict(preferred)
    for field in REGISTRATION_EVIDENCE_FIELDS:
        if not norm(merged.get(field)) and norm(fallback.get(field)):
            merged[field] = fallback.get(field)
    return merged


def dedupe_registration_evidence_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    passthrough: list[dict[str, Any]] = []
    order: list[tuple[str, str, str, str, str]] = []
    for row in rows:
        key = registration_dedupe_key(row)
        if key is None:
            passthrough.append(row)
            continue
        existing = deduped.get(key)
        if existing is None:
            deduped[key] = row
            order.append(key)
            continue
        if registration_row_priority(row) >= registration_row_priority(existing):
            deduped[key] = merge_registration_rows(row, existing)
        else:
            deduped[key] = merge_registration_rows(existing, row)
    return [deduped[key] for key in order] + passthrough


def drop_superseded_seed_registration_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    confirmed_keys: set[tuple[str, str, str]] = set()
    for row in rows:
        product_id = norm(row.get("product_id"))
        jurisdiction = norm(row.get("jurisdiction")).upper()
        regulator = norm(row.get("regulator")).upper()
        if not product_id or not (jurisdiction or regulator):
            continue
        source_type = norm(row.get("source_type")).lower()
        source_key = norm(row.get("source_key")).lower()
        has_registered_identity = bool(norm(row.get("registration_no")) or norm(row.get("registered_name")))
        if has_registered_identity and source_type != "seed_workbook" and not source_key.startswith("workbook_"):
            confirmed_keys.add((product_id, jurisdiction, regulator))

    filtered: list[dict[str, Any]] = []
    for row in rows:
        product_id = norm(row.get("product_id"))
        jurisdiction = norm(row.get("jurisdiction")).upper()
        regulator = norm(row.get("regulator")).upper()
        source_type = norm(row.get("source_type")).lower()
        source_key = norm(row.get("source_key")).lower()
        is_seed = source_type == "seed_workbook" or source_key.startswith("workbook_") or source_key.startswith("manual_needs_review_resolution:workbook_")
        is_empty_identity = not norm(row.get("registration_no")) and not norm(row.get("registered_name"))
        if is_seed and is_empty_identity and (product_id, jurisdiction, regulator) in confirmed_keys:
            continue
        filtered.append(row)
    return filtered


def write_rows_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def write_product_master_output(product_master: list[dict[str, Any]]) -> None:
    write_rows_csv(PRODUCT_MASTER_PATH, PRODUCT_MASTER_FIELDS, product_master)


def build_registration_evidence_output(
    registration_seed: list[dict[str, Any]],
    staging_records: list[dict[str, Any]],
    promoted_registration_rows: list[dict[str, Any]],
    allowed_product_ids: set[str] | None = None,
) -> list[dict[str, Any]]:
    rows = [
        {field: enrich_registration_description_fields(item).get(field, "") for field in REGISTRATION_EVIDENCE_FIELDS}
        for item in registration_seed
        if allowed_product_ids is None or norm(item.get("product_id")) in allowed_product_ids
    ]
    for item in staging_records:
        if item.get("merge_target") != "registration_evidence":
            continue
        product_id = norm(item.get("product_id"))
        if allowed_product_ids is not None and product_id and product_id not in allowed_product_ids:
            continue
        candidates = item.get("field_candidates") or {}
        if isinstance(candidates, str):
            try:
                candidates = json.loads(candidates)
            except json.JSONDecodeError:
                candidates = {}
        rows.append(
            enrich_registration_description_fields(
                {
                "product_id": item.get("product_id"),
                "seed_record_id": "",
                "company_id": item.get("company_id"),
                "company": item.get("company"),
                "brand": item.get("brand"),
                "jurisdiction": item.get("jurisdiction"),
                "regulator": "FDA" if item.get("source_key") == "fda_openfda_510k" else "",
                "regulatory_pathway": candidates.get("regulatory_pathway"),
                "status": candidates.get("status"),
                "registration_no": candidates.get("registration_no"),
                "approval_date": candidates.get("approval_date"),
                "expiry_date": "",
                "registered_name": candidates.get("registered_name"),
                "approved_indication": candidates.get("approved_indication"),
                "intended_use": candidates.get("intended_use"),
                "legal_manufacturer": candidates.get("legal_manufacturer"),
                "local_holder": "",
                "source_key": item.get("source_key"),
                "source_url": item.get("url"),
                "source_type": "official_api",
                "evidence_title": item.get("title"),
                "evidence_excerpt": item.get("excerpt"),
                "field_note": item.get("field_note") or item.get("review_note"),
                "checked_at": item.get("reviewed_at") or item.get("captured_at"),
                "reviewed_by": item.get("reviewed_by") or "",
                "review_status": item.get("review_status") or "needs_review",
                "confidence": item.get("confidence") or "official_api_unreviewed",
                }
            )
        )
    rows.extend(
        {field: enrich_registration_description_fields(item).get(field, "") for field in REGISTRATION_EVIDENCE_FIELDS}
        for item in promoted_registration_rows
        if allowed_product_ids is None or norm(item.get("product_id")) in allowed_product_ids
    )
    return drop_superseded_seed_registration_rows(dedupe_registration_evidence_rows(rows))


def derived_registration_counts(registration_rows: list[dict[str, Any]]) -> dict[str, Counter[str]]:
    counters: dict[str, Counter[str]] = {"FDA": Counter(), "NMPA": Counter()}
    seen: set[tuple[str, str, str]] = set()
    for row in registration_rows:
        product_id = norm(row.get("product_id"))
        company = norm(row.get("company"))
        if not product_id or not company:
            continue
        regulator_blob = " ".join(
            norm(row.get(field)).lower()
            for field in ["regulator", "jurisdiction", "source_key", "source_type", "regulatory_pathway"]
        )
        if "fda" in regulator_blob and "kfda" not in regulator_blob and "mfds" not in regulator_blob:
            key = ("FDA", company, product_id)
            if key not in seen:
                counters["FDA"][company] += 1
                seen.add(key)
        if "nmpa" in regulator_blob or "jurisdiction cn" in regulator_blob or re.search(r"\bcn\b", regulator_blob):
            key = ("NMPA", company, product_id)
            if key not in seen:
                counters["NMPA"][company] += 1
                seen.add(key)
    return counters


def apply_registration_product_backfill(
    product_master: list[dict[str, Any]],
    registration_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    product_by_id = {row.get("product_id"): row for row in product_master if row.get("product_id")}
    updated = 0
    updates_by_source = Counter()
    for row in registration_rows:
        product_id = norm(row.get("product_id"))
        product = product_by_id.get(product_id)
        if not product:
            continue
        registered_name = norm(row.get("registered_name"))
        legal_manufacturer = norm(row.get("legal_manufacturer"))
        local_holder = norm(row.get("local_holder"))
        changed = False
        if registered_name and not norm(product.get("registered_name")):
            product["registered_name"] = registered_name
            changed = True
        if legal_manufacturer and not norm(product.get("legal_manufacturer")):
            product["legal_manufacturer"] = legal_manufacturer
            changed = True
        if local_holder and not norm(product.get("local_holder")):
            product["local_holder"] = local_holder
            changed = True
        if changed:
            updated += 1
            source = norm(row.get("source_type")) or norm(row.get("source_key")) or "registration_evidence"
            updates_by_source[source] += 1
            if norm(product.get("verification_status")) == "unverified_seed":
                product["verification_status"] = "official_evidence_promoted"
            if norm(product.get("review_status")) in {"backlog", "queued", ""}:
                product["review_status"] = "auto_cross_checked"
            if "official_evidence_promoted" not in norm(product.get("source_status")):
                product["source_status"] = ", ".join(
                    part
                    for part in [norm(product.get("source_status")), "official_evidence_promoted"]
                    if part
                )
    return {"product_registered_name_backfilled": updated, "by_source": top_counts(updates_by_source, 12)}


REGISTRATION_INDICATION_UNAVAILABLE = "unavailable_verified_no_public_indication"


def has_unavailable_public_indication_marker(rows: list[dict[str, Any]]) -> bool:
    marker = REGISTRATION_INDICATION_UNAVAILABLE.casefold()
    return any(marker in norm(row.get("review_status")).casefold() for row in rows)


def apply_v4_registration_indication_closure(registration_rows: list[dict[str, Any]]) -> dict[str, Any]:
    status_updates = 0
    approved_filled = 0
    intended_filled = 0
    exact_promoted = 0
    unavailable_marked = 0
    for row in registration_rows:
        review_status = norm(row.get("review_status")).casefold()
        if review_status in {"needs_source_followup", "pdf_indication_not_found"}:
            row["review_status"] = REGISTRATION_INDICATION_UNAVAILABLE
            status_updates += 1

        official_text = norm(row.get("official_description_exact"))
        approved = norm(row.get("approved_indication"))
        intended = norm(row.get("intended_use"))
        if official_text:
            if not approved:
                row["approved_indication"] = official_text
                approved_filled += 1
                exact_promoted += 1
            if not intended:
                row["intended_use"] = official_text
                intended_filled += 1
                exact_promoted += 1
            continue

        if approved and not intended:
            row["intended_use"] = approved
            intended_filled += 1
            continue
        if intended and not approved:
            row["approved_indication"] = intended
            approved_filled += 1
            continue

        if not approved:
            row["approved_indication"] = REGISTRATION_INDICATION_UNAVAILABLE
            approved_filled += 1
            unavailable_marked += 1
        if not intended:
            row["intended_use"] = REGISTRATION_INDICATION_UNAVAILABLE
            intended_filled += 1
            unavailable_marked += 1
        if not norm(row.get("field_note")):
            row["field_note"] = "No public approved-indication/intended-use text was captured in the current official-source pass."

    return {
        "registration_followup_status_closed": status_updates,
        "registration_approved_indication_filled": approved_filled,
        "registration_intended_use_filled": intended_filled,
        "registration_official_exact_text_promoted": exact_promoted,
        "registration_unavailable_markers_added": unavailable_marked,
    }


def write_registration_evidence_output(rows: list[dict[str, Any]]) -> None:
    write_rows_csv(REGISTRATION_EVIDENCE_PATH, REGISTRATION_EVIDENCE_FIELDS, rows)


def build_market_snapshots(company_master: list[dict[str, Any]]) -> list[dict[str, Any]]:
    as_of = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    live = load_market_snapshot_live()
    financial = load_company_financial_metrics()
    rows = []
    for company in company_master:
        if not company.get("stock_code"):
            continue
        live_row = live.get(company["company_id"]) or live.get(company["stock_code"]) or live.get(company["ticker_symbol"]) or {}
        financial_row = (
            financial.get(company["company_id"])
            or financial.get(company["stock_code"])
            or financial.get(company["ticker_symbol"])
            or financial.get(company["canonical_name"])
            or {}
        )
        rows.append(
            {
                "company_id": company["company_id"],
                "company": company["canonical_name"],
                "stock_code": company["stock_code"],
                "exchange": company["exchange"],
                "ticker_symbol": company["ticker_symbol"],
                "listing_country": company["listing_country"],
                "as_of": live_row.get("as_of") or as_of,
                "price": live_row.get("price") or "",
                "currency": live_row.get("currency") or "",
                "market_cap_usd_m": live_row.get("market_cap_usd_m") or "",
                "pe_ratio": live_row.get("pe_ratio") or financial_row.get("pe_ratio") or "",
                "pb_ratio": live_row.get("pb_ratio") or financial_row.get("pb_ratio") or "",
                "eps_ttm": live_row.get("eps_ttm") or financial_row.get("eps_ttm") or financial_row.get("eps_diluted") or financial_row.get("eps_basic") or "",
                "net_profit_growth_yoy_pct": financial_row.get("net_income_growth_yoy_pct") or live_row.get("net_profit_growth_yoy_pct") or "",
                "revenue_usd_m": financial_row.get("revenue_usd_m") or "",
                "revenue_year": financial_row.get("fiscal_year") or "",
                "gross_margin_pct": financial_row.get("gross_margin_pct") or "",
                "ps_ratio": financial_row.get("ps_ratio") or "",
                "financial_period": financial_row.get("financial_period") or (f"FY{financial_row.get('fiscal_year')}" if financial_row.get("fiscal_year") else ""),
                "filing_date": financial_row.get("filing_date") or financial_row.get("revenue_filed") or live_row.get("filing_date") or "",
                "metric_basis": financial_row.get("metric_basis") or live_row.get("metric_basis") or "",
                "market_refreshed_at": live_row.get("market_refreshed_at") or live_row.get("as_of") or "",
                "financial_refreshed_at": financial_row.get("captured_at") or live_row.get("financial_refreshed_at") or "",
                "financial_source_url": financial_row.get("source_url") or "",
                "financial_review_status": financial_row.get("review_status") or "",
                "day_change_pct": live_row.get("day_change_pct") or "",
                "source": live_row.get("source") or "",
                "source_url": live_row.get("source_url") or "",
                "snapshot_status": live_row.get("snapshot_status") or "pending_live_fetch",
                "note": live_row.get("note") or "Dynamic market data is intentionally separate from the static master and requires a live data collector.",
            }
        )
    def market_cap_sort_key(row: dict[str, Any]) -> tuple[int, float, str]:
        try:
            market_cap = float(row.get("market_cap_usd_m") or 0)
        except (TypeError, ValueError):
            market_cap = 0
        return (0 if market_cap else 1, -market_cap, row.get("company") or "")

    return sorted(rows, key=market_cap_sort_key)


def source_scope_status(source: dict[str, Any]) -> str:
    channel = source.get("channel_code")
    if channel in EXTERNAL_PROJECT_CHANNELS:
        return "external_project"
    if channel in CURRENT_PHASE_CHANNELS:
        return "active_phase"
    return "roadmap"


def build_verification_queue(company_master: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    created_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    lanes = [
        ("company_background", "公司背景/母公司/并购", "corporate", "company official site + annual report + acquisition releases"),
        ("product_map", "品牌-产品映射", "corporate", "company official product portfolio / IFU"),
        ("registration_us", "美国 FDA 注册/获批", "regulatory", "FDA 510(k), Registration & Listing, AccessGUDID"),
        ("registration_eu", "欧盟 MDR / CE", "regulatory", "EUDAMED device/certificate records + CE/MDR certificate evidence"),
        ("registration_other_markets", "其他国家/地区获批适应症", "regulatory", "TGA/ARTG, ANVISA, Health Canada MDL, PMDA/MHLW, MHRA/UKCA, HSA/SMDR, COFEPRIS, Saudi SFDA, Taiwan TFDA and other official market authorization sources"),
    ]
    for company in sorted(
        [item for item in company_master if item.get("priority_rank")],
        key=lambda item: int(item.get("priority_rank") or 999),
    ):
        aliases = json.loads(company.get("aliases") or "[]")
        query_base = " OR ".join(aliases[:3]) or company.get("canonical_name")
        for fact_group, label, lane, expected_source in lanes:
            rows.append(
                {
                    "priority_rank": company["priority_rank"],
                    "company_id": company["company_id"],
                    "company": company["canonical_name"],
                    "fact_group": fact_group,
                    "target_label": label,
                    "source_lane": lane,
                    "query": query_base,
                    "expected_source": expected_source,
                    "status": "queued",
                    "created_at": created_at,
                    "evidence_count": 0,
                    "reviewer_note": "",
                }
            )
    return rows


def load_staging_records() -> list[dict[str, Any]]:
    if not STAGING_JSONL_PATH.exists():
        return []
    records = []
    for line in STAGING_JSONL_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
            if isinstance(payload, dict):
                records.append(canonicalize_company_fields(payload))
        except json.JSONDecodeError:
            continue
    return records


def load_company_background_evidence() -> list[dict[str, Any]]:
    if not COMPANY_BACKGROUND_EVIDENCE_PATH.exists():
        return []
    records = []
    for line in COMPANY_BACKGROUND_EVIDENCE_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return records


def load_company_capital_structure() -> list[dict[str, Any]]:
    if not COMPANY_CAPITAL_STRUCTURE_PATH.exists():
        return []
    with COMPANY_CAPITAL_STRUCTURE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_listed_company_batch() -> list[dict[str, Any]]:
    if not LISTED_COMPANY_BATCH_PATH.exists():
        return []
    with LISTED_COMPANY_BATCH_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_company_official_source_plan() -> list[dict[str, Any]]:
    if not COMPANY_OFFICIAL_SOURCE_PLAN_PATH.exists():
        return []
    with COMPANY_OFFICIAL_SOURCE_PLAN_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            row
            for row in csv.DictReader(handle)
            if company_exclusion_reason(row) != "no_medical_aesthetic_product"
        ]


def load_policy_regulatory_source_plan() -> list[dict[str, Any]]:
    if not POLICY_REGULATORY_SOURCE_PLAN_PATH.exists():
        return []
    with POLICY_REGULATORY_SOURCE_PLAN_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_company_official_source_evidence() -> list[dict[str, Any]]:
    if not COMPANY_OFFICIAL_SOURCE_EVIDENCE_PATH.exists():
        return []
    records: dict[str, dict[str, Any]] = {}
    unkeyed: list[dict[str, Any]] = []
    for line in COMPANY_OFFICIAL_SOURCE_EVIDENCE_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(payload, dict):
            continue
        record = canonicalize_company_fields(payload)
        if company_exclusion_reason(record) == "no_medical_aesthetic_product":
            continue
        record_id = norm(record.get("evidence_id")) or norm(record.get("id"))
        if record_id:
            records[record_id] = record
        else:
            unkeyed.append(record)
    return list(records.values()) + unkeyed


def load_mdr_ce_evidence_candidates() -> list[dict[str, Any]]:
    if not MDR_CE_EVIDENCE_CANDIDATES_PATH.exists():
        return []
    records = []
    for line in MDR_CE_EVIDENCE_CANDIDATES_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(record, dict) and company_exclusion_reason(record) == "no_medical_aesthetic_product":
            continue
        records.append(record)
    return records


def normalized_host(url: Any) -> str:
    text = norm(url)
    if not text:
        return ""
    if "://" not in text:
        text = "https://" + text
    try:
        host = re.sub(r"^www\.", "", urllib.parse.urlparse(text).netloc.lower())
    except Exception:
        return ""
    return host.split(":")[0]


def host_matches_domain(host: str, domain: str) -> bool:
    host = re.sub(r"^www\.", "", norm(host).lower())
    domain = re.sub(r"^www\.", "", norm(domain).lower())
    return bool(host and domain and (host == domain or host.endswith("." + domain)))


def company_source_role(company: str, url: Any) -> str:
    company_key = norm(company).lower()
    host = normalized_host(url)
    if not host:
        return "unknown"
    if any(host_matches_domain(host, domain) for domain in COMPANY_SOURCE_FALSE_POSITIVE_DOMAINS.get(company_key, set())):
        return "excluded_similar_name"
    if any(host_matches_domain(host, domain) for domain in COMPANY_SOURCE_TRUSTED_DOMAINS.get(company_key, set())):
        return "trusted_official"
    return "crosscheck"


def promotion_domain_blocked(host: str) -> bool:
    return any(host_matches_domain(host, domain) for domain in PROMOTION_DOMAIN_BLOCKLIST)


def promotion_pattern_match(patterns: list[str], text: str) -> bool:
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in patterns)


def promotion_doc_signal(text: str, url: str, host: str) -> str:
    if host_matches_domain(host, "accessdata.fda.gov") and ("cdrh_docs" in url.lower() or "cfpma" in url.lower()):
        return "fda_official_document"
    if promotion_pattern_match(PROMOTION_IFU_PATTERNS, text) or re.search(r"/e?ifu(?:/|\.|$)", url, flags=re.IGNORECASE) or host.startswith("ifu."):
        return "ifu"
    if promotion_pattern_match(PROMOTION_CERTIFICATE_PATTERNS, text):
        return "certificate"
    return ""


def official_domains_by_company(
    official_website_master: list[dict[str, Any]], company_official_website: list[dict[str, Any]]
) -> dict[str, set[str]]:
    domains: dict[str, set[str]] = defaultdict(set)
    for row in official_website_master:
        company = norm(row.get("company"))
        domain = normalized_host(row.get("official_domain") or row.get("official_website_url"))
        if company and domain and not promotion_domain_blocked(domain):
            domains[company].add(domain)
    for row in company_official_website:
        company = norm(row.get("company"))
        for field in [
            "listed_parent_domain",
            "operating_company_domain",
            "primary_official_domain",
            "brand_website_urls",
            "product_line_page_urls",
        ]:
            for value in re.split(r"[,;\s]+", norm(row.get(field))):
                domain = normalized_host(value)
                if company and domain and not promotion_domain_blocked(domain):
                    domains[company].add(domain)
    return domains


def host_is_promotable_official(host: str, company: str, official_domains: dict[str, set[str]]) -> bool:
    if not host or promotion_domain_blocked(host):
        return False
    if any(host_matches_domain(host, domain) for domain in PROMOTION_ALLOWED_PRODUCT_DOC_DOMAINS):
        return True
    for domain in official_domains.get(company, set()):
        if host_matches_domain(host, domain):
            return True
    company_tokens = [token for token in re.split(r"[^a-z0-9]+", company.lower()) if len(token) >= 5]
    return any(token in host for token in company_tokens)


def is_generic_promotion_candidate(item: dict[str, Any]) -> bool:
    text = " ".join(norm(item.get(field)) for field in ["title", "url", "evidence_excerpt"])
    if promotion_pattern_match(PROMOTION_GENERIC_TITLE_PATTERNS, text):
        return True
    host = normalized_host(item.get("url"))
    if host_matches_domain(host, "health.ec.europa.eu") and "document/download" not in norm(item.get("url")).lower():
        return True
    if "eudamed-help" in norm(item.get("url")).lower():
        return True
    return False


def classify_mdr_ce_promotion(
    item: dict[str, Any], official_domains: dict[str, set[str]]
) -> dict[str, str] | None:
    url = norm(item.get("url"))
    host = normalized_host(url)
    title = norm(item.get("title"))
    text = " ".join(norm(item.get(field)) for field in ["title", "url", "evidence_excerpt", "raw_text"])
    short_text = " ".join(norm(item.get(field)) for field in ["title", "url"])
    if not url or promotion_domain_blocked(host) or is_generic_promotion_candidate(item):
        return None

    signal = promotion_doc_signal(short_text, url, host)
    company = norm(item.get("company"))

    if signal == "fda_official_document":
        return {
            "jurisdiction": "US",
            "regulator": "FDA",
            "regulatory_pathway": "FDA label / summary / IFU",
            "status": "official_fda_document_found",
            "source_type": "official_fda_document",
            "confidence": "official_fda_document_promoted",
        }

    if host_matches_domain(host, "ec.europa.eu") or host_matches_domain(host, "webgate.ec.europa.eu"):
        if "devices" in url.lower() and not is_generic_promotion_candidate(item):
            return {
                "jurisdiction": "EU",
                "regulator": "EUDAMED / European Commission",
                "regulatory_pathway": "EUDAMED",
                "status": "official_eudamed_source_found",
                "source_type": "official_eudamed",
                "confidence": "official_eudamed_promoted",
            }
        return None

    if signal and host_is_promotable_official(host, company, official_domains):
        if signal == "ifu":
            return {
                "jurisdiction": "EU / Global",
                "regulator": "Manufacturer IFU",
                "regulatory_pathway": "IFU / official labeling",
                "status": "official_ifu_source_found",
                "source_type": "official_company_ifu",
                "confidence": "official_ifu_promoted",
            }
        if signal == "certificate":
            return {
                "jurisdiction": "EU",
                "regulator": "Notified Body / Manufacturer",
                "regulatory_pathway": "CE/MDR certificate or declaration",
                "status": "official_certificate_source_found",
                "source_type": "official_company_certificate",
                "confidence": "official_certificate_promoted",
            }

    return None


def classify_company_official_promotion(
    item: dict[str, Any], official_domains: dict[str, set[str]]
) -> dict[str, str] | None:
    query_type = norm(item.get("query_type"))
    family_id = norm(item.get("product_family_id"))
    confidence = norm(item.get("confidence"))
    official_candidate = norm(item.get("official_candidate")).lower()
    if query_type not in COMPANY_OFFICIAL_PROMOTION_QUERY_TYPES or not family_id:
        return None
    if confidence not in COMPANY_OFFICIAL_PROMOTION_CONFIDENCES and official_candidate != "likely":
        return None

    url = norm(item.get("url"))
    host = normalized_host(url)
    company = norm(item.get("company"))
    if not url or promotion_domain_blocked(host) or is_generic_promotion_candidate(item):
        return None
    if not host_is_promotable_official(host, company, official_domains):
        return None

    full_text = " ".join(norm(item.get(field)) for field in ["title", "url", "evidence_excerpt", "raw_text"])
    signal = promotion_doc_signal(full_text, url, host)
    has_ce_signal = bool(re.search(r"\b(CE\s*(?:Class|2409|mark|marked|certificate)|MDR|EU\s*MDR|EUDAMED)\b", full_text, re.I))
    has_ifu_signal = signal == "ifu" or "instructions for use" in full_text.lower()
    if not (has_ifu_signal or has_ce_signal or signal == "certificate"):
        return None

    if has_ifu_signal:
        return {
            "jurisdiction": "EU / Global" if has_ce_signal or ".gr" in host or "nordics" in host else "Global",
            "regulator": "Manufacturer IFU",
            "regulatory_pathway": "IFU / official labeling",
            "status": "official_company_ifu_found",
            "source_type": "official_company_ifu",
            "confidence": "official_company_ifu_promoted",
        }
    return {
        "jurisdiction": "EU" if has_ce_signal else "Global",
        "regulator": "Notified Body / Manufacturer",
        "regulatory_pathway": "CE/MDR certificate or declaration",
        "status": "official_company_certificate_found",
        "source_type": "official_company_certificate",
        "confidence": "official_company_certificate_promoted",
    }


def family_product_refs(
    product_hierarchy: dict[str, list[dict[str, Any]]],
    product_id_by_seed_record_id: dict[str, str] | None = None,
) -> dict[str, list[dict[str, str]]]:
    refs: dict[str, list[dict[str, str]]] = defaultdict(list)
    product_id_by_seed_record_id = product_id_by_seed_record_id or {}
    for sku in product_hierarchy.get("skus", []):
        family_id = norm(sku.get("product_family_id"))
        seed_record_id = norm(sku.get("seed_record_id"))
        product_id = product_id_by_seed_record_id.get(seed_record_id) or norm(sku.get("product_id"))
        if not family_id or not product_id:
            continue
        refs[family_id].append(
            {
                "product_id": product_id,
                "seed_record_id": seed_record_id,
                "company_id": norm(sku.get("company_id")),
                "company": norm(sku.get("company")),
                "brand": norm(sku.get("brand")),
                "product_family": norm(sku.get("product_family")),
                "is_primary": norm(sku.get("is_primary_record")) == "1",
            }
        )
    for family_id, rows in refs.items():
        rows.sort(key=lambda row: (0 if row["is_primary"] else 1, row["seed_record_id"], row["product_id"]))
    return refs


def clean_evidence_text(value: Any) -> str:
    text = re.sub(r"\s+", " ", norm(value)).strip()
    if not text or text.count("\ufffd") > 8:
        return ""
    return text


INDICATION_SECTION_BREAKS = [
    "CONTRAINDICATIONS",
    "WARNINGS",
    "PRECAUTIONS",
    "ADVERSE EVENTS",
    "ADVERSE EXPERIENCES",
    "ADVERSE REACTIONS",
    "PRE-MARKET CLINICAL TRIAL",
    "PREMARKET CLINICAL TRIAL",
    "CLINICAL TRIAL",
    "CLINICAL TRIALS",
    "CLINICAL STUDIES",
    "CLINICAL EVALUATION",
    "SUMMARY OF SAFETY",
    "DIRECTIONS FOR USE",
    "INSTRUCTIONS FOR USE",
    "HOW SUPPLIED",
    "DEVICE DESCRIPTION",
    "DESCRIPTION",
    "LIMITATIONS",
    "DATE SUMMARY PREPARED",
    "SUBSTANTIAL EQUIVALENCE",
    "PREDICATE DEVICE",
    "PERFORMANCE DATA",
    "MEDDRA",
    "TEAE",
]


NOISY_INDICATION_MARKERS = [
    "adverse experiences",
    "adverse events",
    "adverse reactions",
    "clinical trial",
    "clinical trials",
    "clinical studies",
    "date summary prepared",
    "meddra",
    "summary of safety",
    "substantial equivalence",
    "teae",
    "predicate device",
    "performance data",
    "classification number",
    "product code",
    "510(k) summary",
    "premarket approval",
]


INDICATION_SENTENCE_TERMS = [
    "approved for",
    "correction of",
    "indicated for",
    "intended for",
    "soft tissue augmentation",
    "hand augmentation",
    "dorsum of the hands",
    "volume loss",
    "wrinkles",
    "folds",
    "décolleté",
    "decollete",
    "chest",
    "jawline",
    "facial",
]

INDICATION_SENTENCE_STOP_TERMS = [
    "adverse",
    "anesthetic",
    "anaesthetic",
    "clinical trial",
    "discomfort",
    "erythema",
    "infection",
    "massage",
    "meddra",
    "nodule",
    "palpable",
    "swelling",
    "teae",
]


def trim_to_indication_sentences(value: str) -> str:
    sentences = re.split(r"(?<=[.!?])\s+", value)
    kept: list[str] = []
    for sentence in sentences:
        sentence = sentence.strip()
        if not sentence:
            continue
        lower = sentence.lower()
        if kept and any(term in lower for term in INDICATION_SENTENCE_STOP_TERMS):
            break
        if not kept or any(term in lower for term in INDICATION_SENTENCE_TERMS):
            kept.append(sentence)
            continue
        break
    return " ".join(kept).strip() or value


def trim_indication_excerpt(value: str) -> str:
    excerpt = re.sub(r"\s+", " ", norm(value)).strip()
    stop_at = len(excerpt)
    upper_excerpt = excerpt.upper()
    for marker in INDICATION_SECTION_BREAKS:
        pos = upper_excerpt.find(marker)
        if pos > 20:
            stop_at = min(stop_at, pos)
    heading = re.search(
        r"\s(?:I{1,3}|IV|V|VI{0,3}|IX|X)\.\s+[A-Z][A-Z /&,\-]{4,}",
        excerpt,
    )
    if heading and heading.start() > 20:
        stop_at = min(stop_at, heading.start())
    all_caps = re.search(r"\s[A-Z][A-Z /&,\-]{10,}:\s", excerpt)
    if all_caps and all_caps.start() > 20:
        stop_at = min(stop_at, all_caps.start())
    ellipsis = excerpt.find("[...]")
    if ellipsis > 60:
        stop_at = min(stop_at, ellipsis)
    excerpt = excerpt[:stop_at].strip(" .;:-")
    excerpt = re.sub(r"\s+", " ", excerpt)
    excerpt = trim_to_indication_sentences(excerpt)
    return excerpt[:700]


def is_promotable_indication_text(value: Any) -> bool:
    text = clean_evidence_text(value)
    if len(text) < 30:
        return False
    lower = text.lower()
    if any(marker in lower for marker in NOISY_INDICATION_MARKERS):
        return False
    return bool(
        re.search(r"\b(indicated for|indications? for use|intended use|intended purpose|intended for)\b", lower)
        and any(
            bucket in lower
            for bucket in [
                "augmentation",
                "dermal filler",
                "facial",
                "soft tissue",
                "nasolabial",
                "décolleté",
                "decollete",
                "jawline",
                "midface",
                "wrinkles",
                "folds",
            ]
        )
    )


def extract_indication_text(value: Any) -> str:
    text = clean_evidence_text(value)
    if not text:
        return ""
    patterns = [
        r"(?:INTENDED USE\s*/\s*INDICATIONS|INDICATIONS\s+FOR\s+USE|INTENDED PURPOSE|INTENDED USE|INDICATION STATEMENT|INDICATIONS?)[:\s.\-–]+(.{30,1400})",
        r"([A-Z0-9®™+\-\s]{1,100}\s+(?:is|are)\s+indicated for .{30,1000})",
        r"(\b(?:is|are)\s+indicated for .{30,1000})",
        r"([A-Z0-9®™+\-\s]{1,100}\s+(?:is|are)\s+(?:a|an\s+)?[^.]{0,120}\s+intended for .{30,1000})",
        r"(\b(?:is|are)\s+(?:a|an\s+)?[^.]{0,120}\s+intended for .{30,1000})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        excerpt = trim_indication_excerpt(match.group(1))
        if is_promotable_indication_text(excerpt):
            return excerpt
    return ""


def official_indication_buckets(text: Any) -> list[str]:
    blob = f" {norm(text).lower()} "
    buckets = []
    for item in FDA_INDICATION_BUCKETS:
        if any(matches_term(blob, term) for term in item["terms"]):
            buckets.append(item["name"])
    if not buckets and any(term in blob for term in NON_SPECIFIC_INDICATION_TERMS):
        buckets.append("待补具体适应症")
    if not buckets and norm(text):
        buckets.append("其他官方适应症")
    return buckets[:6]


def regulatory_family(item: dict[str, Any]) -> str:
    regulator = norm(item.get("regulator")).lower()
    jurisdiction = norm(item.get("jurisdiction")).lower()
    source_type = norm(item.get("source_type")).lower()
    source_key = norm(item.get("source_key")).lower()
    pathway = norm(item.get("regulatory_pathway")).lower()
    if "mfds" in regulator or "kfda" in regulator or "mfds" in source_key or "kfda" in source_key or "korea" in jurisdiction or jurisdiction == "kr":
        return "MFDS"
    if "tfda" in regulator or "tfda" in source_key or "taiwan" in regulator or "taiwan" in jurisdiction:
        return "Taiwan TFDA"
    if "thai" in regulator or "thai" in source_key or "thailand" in jurisdiction:
        return "Thai FDA"
    if "philippines" in regulator or "philippines" in source_key or "philippines" in jurisdiction or "cmdn" in regulator or "cmdr" in regulator:
        return "Philippines CMDN/CMDR"
    if "sfda" in regulator or "sfda" in source_key or "saudi" in regulator or "saudi" in jurisdiction:
        return "Saudi SFDA"
    if "fda" in regulator or "fda" in source_type or "fda" in source_key or jurisdiction == "us":
        return "FDA"
    if (
        "eudamed" in regulator
        or "eudamed" in pathway
        or "notified body" in regulator
        or "notified body" in pathway
        or re.search(r"\bce\b|ce/mdr|\bmdr\b|ce mark|ce-mark", pathway)
        or jurisdiction.startswith("eu")
    ):
        return "CE/MDR"
    if "nmpa" in regulator or jurisdiction in {"cn", "china", "mainland china"}:
        return "NMPA"
    if "tga" in regulator or "artg" in regulator or "australia" in jurisdiction:
        return "TGA/ARTG"
    if "anvisa" in regulator or "brazil" in jurisdiction:
        return "ANVISA"
    if "health canada" in regulator or "canada" in jurisdiction:
        return "Health Canada"
    if "pmda" in regulator or "mhlw" in regulator or "japan" in jurisdiction:
        return "PMDA/MHLW"
    if "mhra" in regulator or "ukca" in regulator:
        return "MHRA/UKCA"
    if "hsa" in regulator or "smdr" in regulator or "singapore" in jurisdiction:
        return "HSA/SMDR"
    if "malaysia" in jurisdiction or re.search(r"\bmda\b", regulator):
        return "Malaysia MDA"
    if "indonesia" in jurisdiction or "akl" in regulator:
        return "Indonesia MoH/AKL"
    if "vietnam" in jurisdiction or "dmec" in regulator:
        return "Vietnam MoH/DMEC"
    if "cofepris" in regulator or "mexico" in jurisdiction:
        return "COFEPRIS"
    return norm(item.get("regulator")) or norm(item.get("jurisdiction")) or "Other"


def build_official_indication_analysis(
    registration_seed: list[dict[str, Any]],
    staging_records: list[dict[str, Any]],
    promoted_registration_rows: list[dict[str, Any]],
    product_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []

    def append_row(row: dict[str, Any], source_label: str) -> None:
        enriched = enrich_registration_description_fields(row)
        product_id = norm(enriched.get("product_id"))
        if product_id and product_id not in product_lookup:
            return
        indication = norm(enriched.get("official_description_exact"))
        if not indication:
            return
        product = product_lookup.get(product_id, {})
        buckets = official_indication_buckets(indication)
        rows.append(
            {
                "product_id": enriched.get("product_id"),
                "seed_record_id": enriched.get("seed_record_id"),
                "company": enriched.get("company") or product.get("Company"),
                "brand": enriched.get("brand") or product.get("Brand"),
                "product": enriched.get("registered_name") or product.get("Core_Product") or product.get("Brand"),
                "country": enriched.get("jurisdiction") or "Unknown",
                "regulator": regulatory_family(enriched),
                "pathway": enriched.get("regulatory_pathway"),
                "registration_no": enriched.get("registration_no"),
                "approval_date": enriched.get("approval_date"),
                "year": extract_year(enriched.get("approval_date")),
                "indication": indication,
                "official_description_exact": indication,
                "official_description_source_field": enriched.get("official_description_source_field"),
                "field_note": enriched.get("field_note"),
                "analysis_bucket_note": "中文标签仅用于图表聚合；正式获批原文见 official_description_exact。",
                "buckets": buckets,
                "source_url": enriched.get("source_url"),
                "source_type": enriched.get("source_type"),
                "confidence": enriched.get("confidence"),
                "source_label": source_label,
            }
        )

    for item in staging_records:
        if item.get("merge_target") != "registration_evidence":
            continue
        candidates = item.get("field_candidates") or {}
        if isinstance(candidates, str):
            try:
                candidates = json.loads(candidates)
            except json.JSONDecodeError:
                candidates = {}
        append_row(
            {
                "product_id": item.get("product_id"),
                "seed_record_id": item.get("source_record_id"),
                "company": item.get("company"),
                "brand": item.get("brand"),
                "jurisdiction": item.get("jurisdiction"),
                "regulator": "FDA" if item.get("source_key") == "fda_openfda_510k" else "",
                "regulatory_pathway": candidates.get("regulatory_pathway"),
                "registration_no": candidates.get("registration_no"),
                "approval_date": candidates.get("approval_date"),
                "registered_name": candidates.get("registered_name"),
                "approved_indication": candidates.get("approved_indication"),
                "intended_use": candidates.get("intended_use"),
                "source_url": item.get("url"),
                "source_type": "official_api",
                "confidence": item.get("confidence"),
                "source_key": item.get("source_key"),
            },
            "FDA 公共记录",
        )
    for row in promoted_registration_rows:
        append_row(row, "官方文件")
    for row in registration_seed:
        append_row(row, "原表线索")

    bucket_rows: list[dict[str, Any]] = []
    regulator_bucket = defaultdict(Counter)
    country_bucket = defaultdict(Counter)
    company_bucket = defaultdict(Counter)
    timeline = defaultdict(Counter)
    for row in rows:
        for bucket in row["buckets"]:
            bucket_rows.append({**row, "bucket": bucket})
            regulator_bucket[bucket][row["regulator"]] += 1
            country_bucket[bucket][row["country"]] += 1
            company_bucket[bucket][row["company"] or "Unknown"] += 1
            if row["year"]:
                timeline[int(row["year"])][bucket] += 1

    top_buckets = top_counts(Counter(row["bucket"] for row in bucket_rows), 16)
    top_regulators = [name for name, _ in Counter(row["regulator"] for row in bucket_rows).most_common(7)] or ["FDA"]
    top_countries = [name for name, _ in Counter(row["country"] for row in bucket_rows).most_common(7)] or ["US"]
    top_companies = [name for name, _ in Counter(row["company"] or "Unknown" for row in bucket_rows).most_common(10)] or ["Unknown"]

    def matrix(counter_by_bucket: dict[str, Counter], columns: list[str], column_field: str, limit: int = 12) -> dict[str, Any]:
        rows_out = []
        for item in top_buckets[:limit]:
            bucket = item["name"]
            counter = counter_by_bucket.get(bucket, Counter())
            examples: dict[str, list[dict[str, Any]]] = {column: [] for column in columns}
            for source_row in bucket_rows:
                if source_row.get("bucket") != bucket:
                    continue
                column_value = source_row.get(column_field) or "Unknown"
                if column_value not in examples or len(examples[column_value]) >= 5:
                    continue
                examples[column_value].append(
                    {
                        "company": source_row.get("company"),
                        "brand": source_row.get("brand"),
                        "product": source_row.get("product"),
                        "registration_no": source_row.get("registration_no"),
                        "approval_date": source_row.get("approval_date"),
                        "official_description_exact": source_row.get("official_description_exact") or source_row.get("indication"),
                        "source_url": source_row.get("source_url"),
                        "field_note": source_row.get("field_note"),
                    }
                )
            rows_out.append(
                {
                    "name": bucket,
                    "total": sum(counter.values()),
                    "values": {column: counter.get(column, 0) for column in columns},
                    "examples": examples,
                }
            )
        return {"columns": columns, "rows": rows_out}

    return {
        "rows": len(rows),
        "records": rows,
        "bucket_rows": len(bucket_rows),
        "top_buckets": top_buckets,
        "top_regulators": top_counts(Counter(row["regulator"] for row in bucket_rows), 10),
        "top_countries": top_counts(Counter(row["country"] for row in bucket_rows), 10),
        "by_regulator_heatmap": matrix(regulator_bucket, top_regulators, "regulator"),
        "by_country_heatmap": matrix(country_bucket, top_countries, "country"),
        "by_company_heatmap": matrix(company_bucket, top_companies, "company"),
        "timeline": [
            {
                "year": str(year),
                "total": sum(counter.values()),
                "fda": sum(value for bucket, value in counter.items()),
                "ce": 0,
                "nmpa": 0,
                "launch": 0,
            }
            for year, counter in sorted(timeline.items())[-10:]
        ],
        "preview": rows[:40],
        "source_note": "官方适应症按产品、国家/地区、监管机构、注册号、批准时间和来源链接记录；图表只显示聚合标签，原文保存在明细字段。",
    }


def write_official_indication_evidence(rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "product_id",
        "seed_record_id",
        "company",
        "brand",
        "product",
        "country",
        "regulator",
        "pathway",
        "registration_no",
        "approval_date",
        "year",
        "indication",
        "official_description_exact",
        "official_description_source_field",
        "field_note",
        "analysis_bucket_note",
        "buckets",
        "source_url",
        "source_type",
        "confidence",
        "source_label",
    ]
    with OFFICIAL_INDICATION_EVIDENCE_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            output = dict(row)
            output["buckets"] = ", ".join(row.get("buckets") or [])
            writer.writerow({field: output.get(field, "") for field in fieldnames})


def official_indication_heatmap(records: list[dict[str, Any]], column_key: str = "regulator", limit: int = 12) -> dict[str, Any]:
    bucket_counter: Counter = Counter()
    matrix: dict[str, Counter] = defaultdict(Counter)
    column_counter: Counter = Counter()
    examples: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for row in records:
        columns_value = norm(row.get(column_key)) or "Unknown"
        column_counter[columns_value] += 1
        for bucket in row.get("buckets") or []:
            bucket_counter[bucket] += 1
            matrix[bucket][columns_value] += 1
            if len(examples[bucket][columns_value]) < 5:
                examples[bucket][columns_value].append(
                    {
                        "company": row.get("company"),
                        "brand": row.get("brand"),
                        "product": row.get("product"),
                        "registration_no": row.get("registration_no"),
                        "approval_date": row.get("approval_date"),
                        "official_description_exact": row.get("official_description_exact") or row.get("indication"),
                        "source_url": row.get("source_url"),
                        "field_note": row.get("field_note"),
                    }
                )
    columns = [name for name, _ in column_counter.most_common(7)] or ["FDA"]
    rows = []
    for bucket, total in bucket_counter.most_common(limit):
        rows.append(
            {
                "name": bucket,
                "total": total,
                "values": {column: matrix[bucket].get(column, 0) for column in columns},
                "examples": {column: examples[bucket].get(column, []) for column in columns},
            }
        )
    return {"columns": columns, "rows": rows}


def extract_registration_number_from_url_or_title(item: dict[str, Any]) -> str:
    text = " ".join(norm(item.get(field)) for field in ["title", "url", "evidence_excerpt"])
    match = re.search(r"\b(K\d{6}|P\d{6}(?:S\d{3})?)\b", text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def update_classification_layer(row: dict[str, Any], regulatory_status: str) -> None:
    payload: dict[str, Any]
    try:
        payload = json.loads(row.get("classification_layer") or "{}")
        if not isinstance(payload, dict):
            payload = {}
    except json.JSONDecodeError:
        payload = {}
    payload["regulatory"] = regulatory_status
    row["classification_layer"] = json.dumps(payload, ensure_ascii=False)


def build_evidence_promotions(
    product_master: list[dict[str, Any]],
    product_hierarchy: dict[str, list[dict[str, Any]]],
    staging_records: list[dict[str, Any]],
    mdr_ce_evidence_candidates: list[dict[str, Any]],
    company_official_source_evidence: list[dict[str, Any]],
    official_website_master: list[dict[str, Any]],
    company_official_website: list[dict[str, Any]],
) -> dict[str, Any]:
    promoted_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    product_by_id = {row.get("product_id"): row for row in product_master if row.get("product_id")}
    product_id_by_seed_record_id = {
        norm(row.get("seed_record_id")): norm(row.get("product_id"))
        for row in product_master
        if norm(row.get("seed_record_id")) and norm(row.get("product_id"))
    }
    refs_by_family = family_product_refs(product_hierarchy, product_id_by_seed_record_id)
    official_domains = official_domains_by_company(official_website_master, company_official_website)
    product_updates: dict[str, dict[str, Any]] = {}
    registration_rows: list[dict[str, Any]] = []
    log_rows: list[dict[str, Any]] = []
    seen_registration_keys: set[tuple[str, str, str]] = set()

    def remember_product_update(product_id: str, source_type: str, registered_name: str = "", legal_manufacturer: str = "") -> None:
        if not product_id or product_id not in product_by_id:
            return
        update = product_updates.setdefault(
            product_id,
            {
                "product_id": product_id,
                "sources": set(),
                "registered_name": "",
                "legal_manufacturer": "",
            },
        )
        update["sources"].add(source_type)
        if registered_name and not update["registered_name"]:
            update["registered_name"] = registered_name
        if legal_manufacturer and not update["legal_manufacturer"]:
            update["legal_manufacturer"] = legal_manufacturer

    def append_log(row: dict[str, Any], field_name: str, promoted_value: str, note: str = "") -> None:
        log_rows.append(
            {
                "promotion_id": stable_id("promo", row.get("product_id"), field_name, row.get("source_url"), promoted_value),
                "product_id": row.get("product_id"),
                "seed_record_id": row.get("seed_record_id"),
                "company_id": row.get("company_id"),
                "company": row.get("company"),
                "brand": row.get("brand"),
                "product_family_id": row.get("product_family_id", ""),
                "source_key": row.get("source_key"),
                "source_type": row.get("source_type"),
                "field_name": field_name,
                "promoted_value": promoted_value,
                "source_url": row.get("source_url"),
                "evidence_title": row.get("evidence_title"),
                "confidence": row.get("confidence"),
                "promoted_at": promoted_at,
                "note": note,
            }
        )

    for item in staging_records:
        if item.get("merge_target") != "registration_evidence":
            continue
        candidates = item.get("field_candidates") or {}
        if isinstance(candidates, str):
            try:
                candidates = json.loads(candidates)
            except json.JSONDecodeError:
                candidates = {}
        product_id = norm(item.get("product_id"))
        if product_id:
            remember_product_update(
                product_id,
                "official_fda_api",
                norm(candidates.get("registered_name")),
                norm(candidates.get("legal_manufacturer")),
            )

    for item in mdr_ce_evidence_candidates:
        authority = classify_mdr_ce_promotion(item, official_domains)
        if not authority:
            continue
        family_id = norm(item.get("product_family_id"))
        refs = refs_by_family.get(family_id) or [
            {
                "product_id": "",
                "seed_record_id": "",
                "company_id": norm(item.get("company_id")),
                "company": norm(item.get("company")),
                "brand": norm(item.get("brand")),
                "product_family": norm(item.get("product_family")),
                "is_primary": True,
            }
        ]
        indication = extract_indication_text(item.get("evidence_excerpt") or item.get("raw_text"))
        registration_no = extract_registration_number_from_url_or_title(item)
        for ref in refs:
            key = (ref.get("product_id", ""), norm(item.get("url")), authority["source_type"])
            if key in seen_registration_keys:
                continue
            seen_registration_keys.add(key)
            row = {
                "product_id": ref.get("product_id", ""),
                "seed_record_id": ref.get("seed_record_id", ""),
                "company_id": ref.get("company_id") or norm(item.get("company_id")),
                "company": ref.get("company") or norm(item.get("company")),
                "brand": ref.get("brand") or norm(item.get("brand")),
                "product_family_id": family_id,
                "jurisdiction": authority["jurisdiction"],
                "regulator": authority["regulator"],
                "regulatory_pathway": authority["regulatory_pathway"],
                "status": authority["status"],
                "registration_no": registration_no,
                "approval_date": "",
                "expiry_date": "",
                "registered_name": ref.get("product_family") or norm(item.get("product_family")),
                "approved_indication": indication if authority["source_type"] == "official_fda_document" else "",
                "intended_use": indication,
                "legal_manufacturer": norm(item.get("company")),
                "local_holder": "",
                "source_key": norm(item.get("source_key")),
                "source_url": norm(item.get("url")),
                "source_type": authority["source_type"],
                "evidence_title": norm(item.get("title")) or norm(item.get("product_family")),
                "evidence_excerpt": clean_evidence_text(item.get("evidence_excerpt"))[:1200],
                "checked_at": norm(item.get("captured_at")) or promoted_at,
                "reviewed_by": "auto_cross_check",
                "review_status": "auto_cross_checked",
                "confidence": authority["confidence"],
            }
            registration_rows.append(row)
            remember_product_update(row["product_id"], authority["source_type"], row["registered_name"], row["legal_manufacturer"])
            append_log(row, "registration_evidence", row["status"], "Promoted from conservative official document rules.")
            if indication:
                append_log(row, "approved_indication" if authority["source_type"] == "official_fda_document" else "intended_use", indication)

    for item in company_official_source_evidence:
        authority = classify_company_official_promotion(item, official_domains)
        if not authority:
            continue
        indication = extract_indication_text(item.get("evidence_excerpt") or item.get("raw_text"))
        if not indication:
            continue
        family_id = norm(item.get("product_family_id"))
        refs = refs_by_family.get(family_id)
        if not refs:
            continue
        registration_no = extract_registration_number_from_url_or_title(item)
        for ref in refs:
            key = (ref.get("product_id", ""), norm(item.get("url")), authority["source_type"])
            if key in seen_registration_keys:
                continue
            seen_registration_keys.add(key)
            row = {
                "product_id": ref.get("product_id", ""),
                "seed_record_id": ref.get("seed_record_id", ""),
                "company_id": ref.get("company_id") or norm(item.get("company_id")),
                "company": ref.get("company") or norm(item.get("company")),
                "brand": ref.get("brand") or norm(item.get("brand")),
                "product_family_id": family_id,
                "jurisdiction": authority["jurisdiction"],
                "regulator": authority["regulator"],
                "regulatory_pathway": authority["regulatory_pathway"],
                "status": authority["status"],
                "registration_no": registration_no,
                "approval_date": "",
                "expiry_date": "",
                "registered_name": ref.get("product_family") or norm(item.get("product_family")),
                "approved_indication": "",
                "intended_use": indication,
                "legal_manufacturer": norm(item.get("company")),
                "local_holder": "",
                "source_key": "company_official_ifu",
                "source_url": norm(item.get("url")),
                "source_type": authority["source_type"],
                "evidence_title": norm(item.get("title")) or norm(item.get("product_family")),
                "evidence_excerpt": clean_evidence_text(item.get("evidence_excerpt"))[:1200],
                "checked_at": norm(item.get("captured_at")) or promoted_at,
                "reviewed_by": "auto_cross_check",
                "review_status": "auto_cross_checked",
                "confidence": authority["confidence"],
            }
            registration_rows.append(row)
            remember_product_update(row["product_id"], authority["source_type"], row["registered_name"], row["legal_manufacturer"])
            append_log(row, "registration_evidence", row["status"], "Promoted from official company IFU/certificate rules.")
            append_log(row, "intended_use", indication, "Exact official IFU/labeling wording; not a marketing bucket.")

    for product_id, update in product_updates.items():
        row = product_by_id.get(product_id)
        if not row:
            continue
        if update.get("registered_name") and not norm(row.get("registered_name")):
            row["registered_name"] = update["registered_name"]
        if update.get("legal_manufacturer") and not norm(row.get("legal_manufacturer")):
            row["legal_manufacturer"] = update["legal_manufacturer"]
        source_names = sorted(update.get("sources") or [])
        row["verification_status"] = "official_evidence_promoted"
        row["review_status"] = "auto_cross_checked"
        row["source_status"] = ", ".join(source_names)
        update_classification_layer(row, "official_evidence_promoted")
        append_log(
            {
                "product_id": product_id,
                "seed_record_id": row.get("seed_record_id"),
                "company_id": row.get("company_id"),
                "company": row.get("company"),
                "brand": row.get("brand"),
                "source_key": "product_master_promotion",
                "source_type": "product_master_update",
                "source_url": "",
                "evidence_title": row.get("standard_product_name"),
                "confidence": "official_evidence_promoted",
            },
            "Product_Master",
            ", ".join(source_names),
            "Official specs intentionally left blank until search completion.",
        )

    summary = {
        "product_master_promoted": len(product_updates),
        "registration_rows_promoted": len(registration_rows),
        "log_rows": len(log_rows),
        "by_source_type": top_counts(Counter(row.get("source_type") for row in registration_rows), 10),
        "by_jurisdiction": top_counts(Counter(row.get("jurisdiction") for row in registration_rows), 10),
        "indication_rows": sum(1 for row in registration_rows if row.get("approved_indication") or row.get("intended_use")),
        "fda_rows": sum(1 for row in registration_rows if row.get("regulator") == "FDA"),
        "mdr_ce_rows": sum(1 for row in registration_rows if row.get("jurisdiction") in {"EU", "EU / Global"}),
        "eudamed_rows": sum(1 for row in registration_rows if row.get("source_type") == "official_eudamed"),
        "spec_fields_promoted": 0,
        "spec_note": "Official specifications are held as candidates and intentionally not written into Product_Master in this promotion pass.",
        "promoted_at": promoted_at,
    }
    return {"registration_rows": registration_rows, "product_updates": product_updates, "log_rows": log_rows, "summary": summary}


def write_evidence_promotion_log(rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "promotion_id",
        "product_id",
        "seed_record_id",
        "company_id",
        "company",
        "brand",
        "product_family_id",
        "source_key",
        "source_type",
        "field_name",
        "promoted_value",
        "source_url",
        "evidence_title",
        "confidence",
        "promoted_at",
        "note",
    ]
    rows = merge_manual_evidence_promotion_log(rows)
    with EVIDENCE_PROMOTION_LOG_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def merge_manual_evidence_promotion_log(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged_rows = {norm(row.get("promotion_id")): row for row in rows if norm(row.get("promotion_id"))}
    for row in load_generated_csv(MANUAL_EVIDENCE_PROMOTION_LOG_PATH):
        promotion_id = norm(row.get("promotion_id"))
        if promotion_id:
            merged_rows[promotion_id] = row
    return list(merged_rows.values())


PRODUCT_FACT_EVIDENCE_FIELDS = [
    "fact_id",
    "product_id",
    "seed_record_id",
    "company_id",
    "company",
    "brand",
    "product_family_id",
    "standard_product_name",
    "priority",
    "fact_group",
    "field_name",
    "field_value",
    "source_url",
    "evidence_title",
    "evidence_excerpt",
    "source_type",
    "confidence",
    "captured_at",
    "promoted_at",
    "review_status",
    "note",
]


def append_unique_status(existing: Any, values: list[str]) -> str:
    seen: list[str] = []
    for part in re.split(r"[,;|]+", norm(existing)):
        value = norm(part)
        if value and value not in seen:
            seen.append(value)
    for value in values:
        value = norm(value)
        if value and value not in seen:
            seen.append(value)
    return ", ".join(seen)


def set_classification_source_review(row: dict[str, Any], source_review_status: str) -> None:
    try:
        payload = json.loads(row.get("classification_layer") or "{}")
        if not isinstance(payload, dict):
            payload = {}
    except json.JSONDecodeError:
        payload = {}
    payload["source_review"] = source_review_status
    row["classification_layer"] = json.dumps(payload, ensure_ascii=False)


def apply_manual_product_fact_evidence(
    product_master: list[dict[str, Any]],
    fact_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    product_by_id = {norm(row.get("product_id")): row for row in product_master if norm(row.get("product_id"))}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in fact_rows:
        product_id = norm(row.get("product_id"))
        if product_id in product_by_id:
            grouped[product_id].append(row)

    updated_products = 0
    by_status: Counter = Counter()
    source_status_values = {
        "official_product_page": "official_product_page",
        "official_product_document": "official_product_document",
        "official_specification_candidate": "official_spec_candidate",
    }
    for product_id, rows in grouped.items():
        product = product_by_id[product_id]
        fact_groups = {norm(row.get("fact_group")) for row in rows}
        spec_rows = [
            row
            for row in rows
            if norm(row.get("fact_group")) == "official_specification_candidate"
            and norm(row.get("field_value"))
        ]
        if spec_rows:
            specs: list[dict[str, str]] = []
            seen_specs: set[tuple[str, str, str]] = set()
            for row in sorted(
                spec_rows,
                key=lambda item: (
                    norm(item.get("field_name")),
                    norm(item.get("field_value")),
                    norm(item.get("source_url")),
                ),
            ):
                category = norm(row.get("field_name")) or "official_specification"
                value = norm(row.get("field_value"))
                source_url = norm(row.get("source_url"))
                key = (category.casefold(), value.casefold(), source_url.casefold())
                if key in seen_specs:
                    continue
                seen_specs.add(key)
                specs.append(
                    {
                        "category": category,
                        "value": value,
                        "source_url": source_url,
                        "source_type": norm(row.get("source_type")),
                        "confidence": norm(row.get("confidence")),
                        "review_status": norm(row.get("review_status")),
                    }
                )
            product["technical_specs_json"] = json.dumps(specs[:18], ensure_ascii=False)
            product["technical_specs_summary"] = "; ".join(
                f"{item['category']}: {item['value']}" for item in specs[:6]
            )
            product["spec_evidence_count"] = str(len(specs))
            if any(norm(row.get("review_status")) == "auto_cross_checked" for row in spec_rows):
                product["spec_review_status"] = "auto_promoted_specs"
            else:
                product["spec_review_status"] = "specs_need_review"
        new_sources = sorted({source_status_values.get(group, group) for group in fact_groups if group})
        if new_sources:
            product["source_status"] = append_unique_status(product.get("source_status"), new_sources)
        current_status = norm(product.get("verification_status"))
        if current_status != "official_evidence_promoted":
            if "official_product_page" in fact_groups and "official_specification_candidate" in fact_groups:
                next_status = "official_product_and_spec_cross_checked"
            elif "official_product_document" in fact_groups:
                next_status = "official_document_cross_checked"
            elif "official_product_page" in fact_groups:
                next_status = "official_product_page_cross_checked"
            elif "official_specification_candidate" in fact_groups:
                next_status = "official_spec_candidate_cross_checked"
            else:
                next_status = current_status or "needs_verification"
            product["verification_status"] = next_status
            set_classification_source_review(product, next_status)
        if norm(product.get("review_status")) in {"", "backlog", "needs_review", "unreviewed"}:
            product["review_status"] = "auto_cross_checked"
        updated_products += 1
        by_status[product.get("verification_status") or "unknown"] += 1

    return {
        "manual_product_fact_rows": len(fact_rows),
        "manual_product_fact_products": len(grouped),
        "manual_product_fact_updates": updated_products,
        "manual_product_fact_by_status": top_counts(by_status, 10),
    }


def append_manual_product_fact_trace_logs(
    log_rows: list[dict[str, Any]],
    fact_rows: list[dict[str, Any]],
    product_master: list[dict[str, Any]],
    promoted_at: str,
) -> int:
    product_by_id = {norm(row.get("product_id")): row for row in product_master if norm(row.get("product_id"))}
    existing_keys = {
        (
            norm(row.get("product_id")),
            norm(row.get("field_name")).casefold(),
            norm(row.get("promoted_value")).casefold(),
        )
        for row in log_rows + load_generated_csv(MANUAL_EVIDENCE_PROMOTION_LOG_PATH)
        if norm(row.get("product_id")) and norm(row.get("field_name")) and norm(row.get("promoted_value"))
    }

    added = 0
    for row in fact_rows:
        product_id = norm(row.get("product_id"))
        field_name = norm(row.get("field_name"))
        field_value = norm(row.get("field_value"))
        review_status = norm(row.get("review_status")).casefold()
        if not product_id or product_id not in product_by_id or not field_name or not field_value:
            continue
        if "orphan" in review_status or "unlinked" in review_status:
            continue
        key = (product_id, field_name.casefold(), field_value.casefold())
        if key in existing_keys:
            continue
        product = product_by_id[product_id]
        log_rows.append(
            {
                "promotion_id": stable_id(
                    "promo_manual_fact",
                    row.get("fact_id"),
                    product_id,
                    field_name,
                    row.get("source_url"),
                    field_value,
                ),
                "product_id": product_id,
                "seed_record_id": norm(row.get("seed_record_id")) or norm(product.get("seed_record_id")),
                "company_id": norm(row.get("company_id")) or norm(product.get("company_id")),
                "company": norm(row.get("company")) or norm(product.get("company")),
                "brand": norm(row.get("brand")) or norm(product.get("brand")),
                "product_family_id": norm(row.get("product_family_id")),
                "source_key": norm(row.get("fact_group")) or "manual_product_fact_evidence",
                "source_type": norm(row.get("source_type")),
                "field_name": field_name,
                "promoted_value": field_value,
                "source_url": norm(row.get("source_url")),
                "evidence_title": norm(row.get("evidence_title")),
                "confidence": norm(row.get("confidence")),
                "promoted_at": norm(row.get("promoted_at")) or promoted_at,
                "note": f"Trace log backfilled from manual_product_fact_evidence fact_id={norm(row.get('fact_id'))}.",
            }
        )
        existing_keys.add(key)
        added += 1
    return added


def apply_product_specification_evidence_to_master(
    product_master: list[dict[str, Any]],
    spec_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    product_by_id = {norm(row.get("product_id")): row for row in product_master if norm(row.get("product_id"))}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in spec_rows:
        product_id = norm(row.get("product_id"))
        review_status = norm(row.get("review_status")).casefold()
        if not product_id or product_id not in product_by_id:
            continue
        if "promoted" not in review_status and "cross_checked" not in review_status:
            continue
        if "filtered" in review_status or "out_of_scope" in review_status or "pending" in review_status:
            continue
        if not norm(row.get("spec_value")):
            continue
        grouped[product_id].append(row)

    updated = 0
    for product_id, rows in grouped.items():
        product = product_by_id[product_id]
        specs: list[dict[str, str]] = []
        try:
            existing = json.loads(product.get("technical_specs_json") or "[]")
            if isinstance(existing, list):
                specs.extend(item for item in existing if isinstance(item, dict))
        except json.JSONDecodeError:
            specs = []

        seen: set[tuple[str, str, str]] = set()
        normalized_specs: list[dict[str, str]] = []
        for item in specs:
            category = norm(item.get("category")) or norm(item.get("spec_category")) or "official_specification"
            value = norm(item.get("value")) or norm(item.get("spec_value"))
            source_url = norm(item.get("source_url"))
            if not value:
                continue
            key = (category.casefold(), value.casefold(), source_url.casefold())
            if key in seen:
                continue
            seen.add(key)
            normalized_specs.append(
                {
                    "category": category,
                    "value": value,
                    "source_url": source_url,
                    "source_type": norm(item.get("source_type")),
                    "confidence": norm(item.get("confidence")),
                    "review_status": norm(item.get("review_status")),
                }
            )

        by_category_count: Counter[str] = Counter(item["category"] for item in normalized_specs)
        for row in sorted(
            rows,
            key=lambda item: (
                0 if "promoted" in norm(item.get("review_status")).casefold() else 1,
                norm(item.get("spec_category")),
                norm(item.get("spec_name")),
                norm(item.get("spec_value")),
            ),
        ):
            category = norm(row.get("spec_category")) or norm(row.get("spec_name")) or "official_specification"
            if by_category_count[category] >= 4:
                continue
            value_parts = [norm(row.get("spec_name")), norm(row.get("spec_value"))]
            unit = norm(row.get("spec_unit"))
            if unit and unit.casefold() not in value_parts[-1].casefold():
                value_parts[-1] = f"{value_parts[-1]} {unit}".strip()
            value = ": ".join(part for part in value_parts if part)
            source_url = norm(row.get("source_page_url"))
            key = (category.casefold(), value.casefold(), source_url.casefold())
            if not value or key in seen:
                continue
            seen.add(key)
            by_category_count[category] += 1
            normalized_specs.append(
                {
                    "category": category,
                    "value": value,
                    "source_url": source_url,
                    "source_type": norm(row.get("source_query_type")),
                    "confidence": norm(row.get("confidence")),
                    "review_status": norm(row.get("review_status")),
                }
            )
            if len(normalized_specs) >= 18:
                break

        if not normalized_specs:
            continue
        product["technical_specs_json"] = json.dumps(normalized_specs[:18], ensure_ascii=False)
        product["technical_specs_summary"] = "; ".join(
            f"{item['category']}: {item['value']}" for item in normalized_specs[:6]
        )
        product["spec_evidence_count"] = str(len(normalized_specs))
        if any("promoted" in norm(row.get("review_status")).casefold() for row in rows):
            product["spec_review_status"] = "auto_promoted_specs"
        elif not norm(product.get("spec_review_status")):
            product["spec_review_status"] = "cross_checked_specs"
        updated += 1

    return {
        "product_spec_evidence_products": len(grouped),
        "product_spec_evidence_master_updates": updated,
    }


GENERATED_CSV_PRIMARY_KEYS = {
    COMPANY_OFFICIAL_WEBSITE_PATH.name: ("company_id",),
    COMPANY_MEDIA_ASSET_INDEX_PATH.name: ("asset_id",),
    PRODUCT_SPECIFICATION_EVIDENCE_PATH.name: ("spec_id",),
    OFFICIAL_WEBSITE_MASTER_PATH.name: ("website_id",),
    MANUAL_PRODUCT_FACT_EVIDENCE_PATH.name: ("fact_id",),
}


def generated_csv_record_key(path: Path, record: dict[str, Any]) -> str:
    for field in GENERATED_CSV_PRIMARY_KEYS.get(path.name, ()):
        value = norm(record.get(field))
        if value:
            return value
    for field in (
        "id",
        "evidence_id",
        "record_id",
        "source_id",
        "plan_id",
        "candidate_id",
        "event_id",
        "promotion_id",
        "spec_id",
        "asset_id",
        "website_id",
    ):
        value = norm(record.get(field))
        if value:
            return value
    return ""


def load_generated_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        keyed: dict[str, dict[str, Any]] = {}
        unkeyed: list[dict[str, Any]] = []
        for row in csv.DictReader(handle):
            record = canonicalize_company_fields(row)
            if company_exclusion_reason(record) == "no_medical_aesthetic_product":
                continue
            record_id = generated_csv_record_key(path, record)
            if record_id:
                keyed[record_id] = record
            else:
                unkeyed.append(record)
        return list(keyed.values()) + unkeyed


def load_audit_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_data_usability_summary() -> dict[str, Any]:
    if not DATA_USABILITY_SUMMARY_PATH.exists():
        return {}
    try:
        return json.loads(DATA_USABILITY_SUMMARY_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def load_company_portfolio_cases() -> list[dict[str, Any]]:
    if not COMPANY_PORTFOLIO_CASES_PATH.exists():
        return []
    try:
        payload = json.loads(COMPANY_PORTFOLIO_CASES_PATH.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError:
        return []
    if isinstance(payload, dict):
        payload = payload.get("cases") or []
    return [item for item in payload if isinstance(item, dict)]


def enrich_company_portfolio_cases(
    cases: list[dict[str, Any]],
    product_master: list[dict[str, Any]],
    official_website_master: list[dict[str, Any]],
    product_specification_evidence: list[dict[str, Any]],
    company_official_source_evidence: list[dict[str, Any]],
    official_indication_analysis: dict[str, Any],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    official_rows = official_indication_analysis.get("records") or []
    for case in cases:
        company_id = norm(case.get("company_id"))
        company = norm(case.get("company"))
        company_blob = company.lower()

        def is_case_row(row: dict[str, Any]) -> bool:
            row_company_id = norm(row.get("company_id"))
            row_company = norm(row.get("company")).lower()
            return bool(
                (company_id and row_company_id == company_id)
                or (company_blob and row_company == company_blob)
                or (company_blob and company_blob in row_company)
            )

        product_rows = [row for row in product_master if is_case_row(row)]
        website_rows = [row for row in official_website_master if is_case_row(row)]
        spec_rows = [row for row in product_specification_evidence if is_case_row(row)]
        source_rows_all = [row for row in company_official_source_evidence if is_case_row(row)]
        source_rows = [row for row in source_rows_all if company_source_role(company, row.get("url")) != "excluded_similar_name"]
        trusted_source_rows = [row for row in source_rows if company_source_role(company, row.get("url")) == "trusted_official"]
        crosscheck_source_rows = [row for row in source_rows if company_source_role(company, row.get("url")) == "crosscheck"]
        excluded_source_rows = [row for row in source_rows_all if company_source_role(company, row.get("url")) == "excluded_similar_name"]
        trusted_website_rows = [
            row
            for row in website_rows
            if company_source_role(company, row.get("official_website_url") or row.get("source_url") or row.get("url")) == "trusted_official"
        ]
        trusted_spec_rows = [row for row in spec_rows if company_source_role(company, row.get("source_page_url")) == "trusted_official"]
        indication_rows = [row for row in official_rows if is_case_row(row)]

        def query_count(rows: list[dict[str, Any]], *query_types: str) -> int:
            wanted = {norm(item) for item in query_types if norm(item)}
            return sum(
                1
                for row in rows
                if norm(row.get("query_type") or row.get("source_query_type")) in wanted
            )

        trusted_domains = sorted(
            {
                normalized_host(url_value)
                for row in trusted_source_rows + trusted_website_rows + trusted_spec_rows
                for url_value in [row.get("url") or row.get("source_page_url") or row.get("official_website_url") or row.get("source_url")]
                if normalized_host(url_value)
            }
        )
        source_depth = [
            {
                "label": "官网/产品页",
                "label_en": "Official pages",
                "value": query_count(trusted_source_rows, "official_product_portfolio", "product_official_page"),
                "note": "官网与产品页",
            },
            {
                "label": "IFU/证书",
                "label_en": "IFU / certificates",
                "value": query_count(trusted_source_rows, "official_ifu_catalog", "product_ifu_labeling", "product_certificate_registration"),
                "note": "说明书、证书、注册线索",
            },
            {
                "label": "规格字段",
                "label_en": "Spec fields",
                "value": len(trusted_spec_rows),
                "note": "包装、容量、成分等字段",
            },
            {
                "label": "应用热力图",
                "label_en": "Application map",
                "value": len((case.get("portfolio_indication_heatmap") or {}).get("columns") or []),
                "note": "来自官网资料图口径",
            },
            {
                "label": "官方适应症",
                "label_en": "Approved indications",
                "value": len(indication_rows),
                "note": "监管证据长表",
            },
            {
                "label": "交叉验证",
                "label_en": "Cross-check",
                "value": len(crosscheck_source_rows),
                "note": "经销商或第三方线索",
            },
        ]
        brand_names = {
            norm(row.get("brand") or row.get("standard_product_name") or row.get("product_family"))
            for row in product_rows
            if norm(row.get("brand") or row.get("standard_product_name") or row.get("product_family"))
        }
        reference_brands = [
            row.get("name")
            for row in (case.get("brand_portfolio_heatmap") or {}).get("rows", [])
            if row.get("name")
        ]
        matched_brands = [
            brand
            for brand in reference_brands
            if any(brand.lower() in existing.lower() or existing.lower() in brand.lower() for existing in brand_names)
        ]
        missing_reference_brands = [brand for brand in reference_brands if brand not in matched_brands]
        evidence_links = []
        seen_urls: set[str] = set()
        for row in trusted_source_rows + crosscheck_source_rows:
            url = norm(row.get("url"))
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            evidence_links.append(
                {
                    "title": row.get("title") or row.get("query_type") or "Official source",
                    "url": url,
                    "query_type": row.get("query_type"),
                    "confidence": row.get("confidence"),
                    "official_candidate": row.get("official_candidate"),
                    "source_role": company_source_role(company, url),
                    "domain": normalized_host(url),
                    "evidence_id": row.get("evidence_id"),
                    "excerpt": norm(row.get("evidence_excerpt"))[:420],
                }
            )
            if len(evidence_links) >= 8:
                break

        merged = dict(case)
        merged["database_coverage"] = {
            "product_master_rows": len(product_rows),
            "official_website_rows": len(website_rows),
            "trusted_website_rows": len(trusted_website_rows),
            "product_spec_rows": len(spec_rows),
            "trusted_product_spec_rows": len(trusted_spec_rows),
            "company_official_source_rows": len(source_rows),
            "trusted_official_source_rows": len(trusted_source_rows),
            "crosscheck_source_rows": len(crosscheck_source_rows),
            "excluded_similar_name_rows": len(excluded_source_rows),
            "trusted_domains": trusted_domains,
            "source_depth": source_depth,
            "official_indication_rows": len(indication_rows),
            "reference_brand_rows": len(reference_brands),
            "matched_reference_brands": len(matched_brands),
            "missing_reference_brands": missing_reference_brands,
            "gap_note": "If reference brands exceed Product_Master rows, promote official website product-family evidence before using the company page as a complete master record.",
        }
        merged["evidence_links"] = evidence_links
        enriched.append(merged)
    return enriched


def load_manual_official_indication_rows() -> list[dict[str, Any]]:
    rows = load_generated_csv(MANUAL_OFFICIAL_INDICATION_EVIDENCE_PATH)
    checked_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    out: list[dict[str, Any]] = []
    for row in rows:
        normalized = {field: row.get(field, "") for field in REGISTRATION_EVIDENCE_FIELDS}
        normalized["checked_at"] = normalized.get("checked_at") or checked_at
        normalized["reviewed_by"] = normalized.get("reviewed_by") or "official_source_sync"
        normalized["review_status"] = normalized.get("review_status") or "auto_cross_checked"
        normalized["confidence"] = normalized.get("confidence") or "official_regulator_record"
        out.append(normalized)
    return out


def load_manual_nmpa_registration_rows() -> list[dict[str, Any]]:
    rows = load_generated_csv(MANUAL_NMPA_REGISTRATION_EVIDENCE_PATH)
    checked_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    out: list[dict[str, Any]] = []
    for row in rows:
        normalized = {field: row.get(field, "") for field in REGISTRATION_EVIDENCE_FIELDS}
        normalized["jurisdiction"] = normalized.get("jurisdiction") or "CN"
        normalized["regulator"] = normalized.get("regulator") or "NMPA"
        normalized["source_key"] = normalized.get("source_key") or "registration_project_nmpa_master"
        normalized["source_type"] = normalized.get("source_type") or "official_nmpa_registration_project"
        normalized["checked_at"] = normalized.get("checked_at") or checked_at
        normalized["reviewed_by"] = normalized.get("reviewed_by") or "registration_project_sync"
        normalized["review_status"] = normalized.get("review_status") or "auto_cross_checked"
        normalized["confidence"] = normalized.get("confidence") or "official_regulator_record"
        out.append(normalized)
    return out


def load_manual_registration_name_rows() -> list[dict[str, Any]]:
    rows = load_generated_csv(MANUAL_REGISTRATION_NAME_EVIDENCE_PATH)
    checked_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    out: list[dict[str, Any]] = []
    for row in rows:
        normalized = {field: row.get(field, "") for field in REGISTRATION_EVIDENCE_FIELDS}
        normalized["source_key"] = normalized.get("source_key") or f"manual_registered_name_review:{normalized.get('seed_record_id')}"
        normalized["source_type"] = normalized.get("source_type") or "manual_regulatory_registered_name_lookup"
        normalized["checked_at"] = normalized.get("checked_at") or checked_at
        normalized["reviewed_by"] = normalized.get("reviewed_by") or "user_supplied_v4_1_registered_name_file"
        normalized["review_status"] = normalized.get("review_status") or "manual_verified"
        normalized["confidence"] = normalized.get("confidence") or "manual_verified_regulatory_name"
        out.append(normalized)
    return out


def load_mdr_ce_search_plan() -> list[dict[str, Any]]:
    if not MDR_CE_SEARCH_PLAN_PATH.exists():
        return []
    with MDR_CE_SEARCH_PLAN_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_market_snapshot_live() -> dict[str, dict[str, Any]]:
    if not MARKET_SNAPSHOT_LIVE_PATH.exists():
        return {}
    with MARKET_SNAPSHOT_LIVE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    live: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = row.get("company_id") or row.get("stock_code") or row.get("ticker_symbol") or ""
        if key:
            live[key] = row
    return live


def load_company_financial_metrics() -> dict[str, dict[str, Any]]:
    if not COMPANY_FINANCIAL_METRICS_PATH.exists():
        return {}
    with COMPANY_FINANCIAL_METRICS_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    metrics: dict[str, dict[str, Any]] = {}
    for row in rows:
        for key in [
            row.get("company_id"),
            row.get("company"),
            row.get("stock_code"),
            row.get("ticker_symbol"),
            row.get("sec_cik"),
        ]:
            value = norm(key)
            if value:
                metrics[value] = row
    return metrics


def load_news_regulatory_event_candidates() -> list[dict[str, Any]]:
    return load_generated_csv(NEWS_REGULATORY_EVENT_CANDIDATES_PATH)


def load_briefing_update_candidates() -> list[dict[str, Any]]:
    return load_generated_csv(BRIEFING_UPDATE_CANDIDATES_PATH)


def load_briefing_verified_update_events() -> list[dict[str, Any]]:
    return load_generated_csv(BRIEFING_VERIFIED_UPDATE_EVENTS_PATH)


def load_briefing_fulltext_rescue() -> list[dict[str, Any]]:
    return load_generated_csv(BRIEFING_FULLTEXT_RESCUE_PATH)


def load_briefing_product_gap_candidates() -> list[dict[str, Any]]:
    return load_generated_csv(BRIEFING_PRODUCT_GAP_CANDIDATES_PATH)


def load_category_pairs() -> set[tuple[str, str]]:
    try:
        company_book = find_file("*标准化版v4.xlsx")
        rows = read_sheet_dicts(company_book, "Category_Definitions")
    except Exception:
        return set()
    return {(norm(row.get("Category_L1")), norm(row.get("Category_L2"))) for row in rows if norm(row.get("Category_L1")) and norm(row.get("Category_L2"))}


def add_quality_issue(
    issues: list[dict[str, Any]],
    *,
    source_table: str,
    source_row_id: str,
    entity_type: str,
    entity_name: str,
    issue_type: str,
    severity: str,
    field_name: str = "",
    current_value: Any = "",
    suggested_value: Any = "",
    description: str,
    action: str,
) -> None:
    issues.append(
        {
            "issue_id": stable_id("dq", source_table, source_row_id, issue_type, field_name, entity_name),
            "source_table": source_table,
            "source_row_id": source_row_id,
            "entity_type": entity_type,
            "entity_name": entity_name,
            "issue_type": issue_type,
            "severity": severity,
            "field_name": field_name,
            "current_value": norm(current_value),
            "suggested_value": norm(suggested_value),
            "description": description,
            "action": action,
            "status": "open",
        }
    )


def data_quality_summary(issues: list[dict[str, Any]]) -> dict[str, Any]:
    by_severity = Counter(issue["severity"] for issue in issues)
    by_type = Counter(issue["issue_type"] for issue in issues)
    by_table = Counter(issue["source_table"] for issue in issues)
    sorted_issues = sorted(
        issues,
        key=lambda issue: (
            SEVERITY_ORDER.get(issue["severity"], 99),
            issue["source_table"],
            issue["issue_type"],
            issue["entity_name"],
        ),
    )
    return {
        "total": len(issues),
        "critical": by_severity.get("critical", 0),
        "high": by_severity.get("high", 0),
        "medium": by_severity.get("medium", 0),
        "low": by_severity.get("low", 0),
        "info": by_severity.get("info", 0),
        "by_severity": top_counts(by_severity, 10),
        "by_type": top_counts(by_type, 14),
        "by_table": top_counts(by_table, 10),
        "top_issues": sorted_issues[:16],
        "report_path": str(DATA_QUALITY_REPORT_PATH),
        "issues_path": str(DATA_QUALITY_ISSUES_PATH),
    }


def build_data_quality(
    products: list[dict[str, Any]],
    companies: list[dict[str, Any]],
    brands: list[dict[str, Any]],
) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    valid_category_pairs = load_category_pairs()
    company_by_name = {company.get("Company"): company for company in companies if company.get("Company")}
    company_key_to_names: dict[str, list[str]] = defaultdict(list)
    for company in companies:
        key = compact_key(company.get("Company"))
        if key and company.get("Company") not in company_key_to_names[key]:
            company_key_to_names[key].append(company.get("Company"))

    product_required_fields = ["Record_ID", "Company", "Country", "Region", "Category_L1", "Category_L2", "Brand", "Core_Product"]
    for product in products:
        row_id = product.get("Record_ID") or stable_id("row", text_blob(product, ["Company", "Brand", "Core_Product"]))
        entity_name = " / ".join(x for x in [product.get("Company"), product.get("Brand"), product.get("Core_Product")] if norm(x))
        for field in product_required_fields:
            if not has_value(product, field):
                add_quality_issue(
                    issues,
                    source_table="Product_Lines",
                    source_row_id=row_id,
                    entity_type="product",
                    entity_name=entity_name,
                    issue_type="required_field_missing",
                    severity="high" if field in {"Company", "Brand", "Core_Product"} else "medium",
                    field_name=field,
                    description=f"Product seed row is missing required field {field}.",
                    action="补齐字段或确认该行应移出产品主数据。",
                )
        if not has_value(product, "Tech_Type_Std"):
            add_quality_issue(
                issues,
                source_table="Product_Lines",
                source_row_id=row_id,
                entity_type="product",
                entity_name=entity_name,
                issue_type="tech_type_missing",
                severity="medium",
                field_name="Tech_Type_Std",
                description="Product has no standardized technical type, which weakens taxonomy and dashboard grouping.",
                action="从产品说明、注册名或官网资料补标准技术路线。",
            )
        pair = (product.get("Category_L1") or "", product.get("Category_L2") or "")
        if valid_category_pairs and all(pair) and pair not in valid_category_pairs:
            add_quality_issue(
                issues,
                source_table="Product_Lines",
                source_row_id=row_id,
                entity_type="product",
                entity_name=entity_name,
                issue_type="category_path_not_in_dictionary",
                severity="medium",
                field_name="Category_L1/Category_L2",
                current_value=" > ".join(pair),
                description="Product category path is not defined in Category_Definitions.",
                action="确认是新增分类，还是应映射到已有标准分类。",
            )
        company = product.get("Company")
        if company and company not in company_by_name:
            candidates = company_key_to_names.get(compact_key(company), [])
            add_quality_issue(
                issues,
                source_table="Product_Lines",
                source_row_id=row_id,
                entity_type="company",
                entity_name=company,
                issue_type="company_not_exactly_in_company_master",
                severity="high" if not candidates else "medium",
                field_name="Company",
                current_value=company,
                suggested_value=", ".join(candidates),
                description="Product row company does not exactly match a Companies row.",
                action="统一公司标准名；若是大小写/别名问题，映射到 canonical_name。",
            )

    record_counts = Counter(product.get("Record_ID") for product in products if product.get("Record_ID"))
    for record_id, count in record_counts.items():
        if count > 1:
            add_quality_issue(
                issues,
                source_table="Product_Lines",
                source_row_id=record_id,
                entity_type="product",
                entity_name=record_id,
                issue_type="duplicate_record_id",
                severity="critical",
                field_name="Record_ID",
                current_value=record_id,
                description=f"Record_ID appears {count} times.",
                action="Record_ID 必须唯一；保留一条并重编号重复行。",
            )

    duplicate_fingerprints: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        if is_non_primary_record(product):
            continue
        fingerprint = "|".join(
            compact_key(product.get(field))
            for field in ["Company", "Brand", "Core_Product", "Category_L2", "Tech_Type_Std"]
        )
        if fingerprint.strip("|"):
            duplicate_fingerprints[fingerprint].append(product)
    for rows in duplicate_fingerprints.values():
        if len(rows) <= 1:
            continue
        sample = rows[0]
        add_quality_issue(
            issues,
            source_table="Product_Lines",
            source_row_id=", ".join(row.get("Record_ID") or "" for row in rows[:8]),
            entity_type="product",
            entity_name=" / ".join(x for x in [sample.get("Company"), sample.get("Brand"), sample.get("Core_Product")] if norm(x)),
            issue_type="possible_duplicate_product",
            severity="medium",
            current_value=str(len(rows)),
            description="Multiple product rows share the same company, brand, core product, category and technology fingerprint.",
            action="确认是否是同一产品重复、不同型号/SKU，或应拆成品牌-产品-SKU 层级。",
        )

    company_product_counts = Counter(product.get("Company") for product in products if product.get("Company"))
    brands_by_company: dict[str, set[str]] = defaultdict(set)
    countries_by_company: dict[str, set[str]] = defaultdict(set)
    for product in products:
        if product.get("Company") and product.get("Brand"):
            brands_by_company[product["Company"]].add(product["Brand"])
        if product.get("Company") and product.get("Country"):
            countries_by_company[product["Company"]].add(product["Country"])
    for company in companies:
        name = company.get("Company")
        if not name:
            continue
        expected_products = safe_int(company.get("Product_Count"))
        actual_products = company_product_counts.get(name, 0)
        if expected_products and expected_products != actual_products:
            add_quality_issue(
                issues,
                source_table="Companies",
                source_row_id=name,
                entity_type="company",
                entity_name=name,
                issue_type="derived_product_count_mismatch",
                severity="medium",
                field_name="Product_Count",
                current_value=expected_products,
                suggested_value=actual_products,
                description="Companies.Product_Count does not match Product_Lines count.",
                action="用 Product_Lines 重新派生公司统计，或确认是否排除了非主记录。",
            )
        expected_brands = safe_int(company.get("Brand_Count"))
        actual_brands = len(brands_by_company.get(name, set()))
        if expected_brands and expected_brands != actual_brands:
            add_quality_issue(
                issues,
                source_table="Companies",
                source_row_id=name,
                entity_type="company",
                entity_name=name,
                issue_type="derived_brand_count_mismatch",
                severity="low",
                field_name="Brand_Count",
                current_value=expected_brands,
                suggested_value=actual_brands,
                description="Companies.Brand_Count does not match unique brands in Product_Lines.",
                action="确认 Brand_Portfolio 与 Product_Lines 的品牌聚合口径。",
            )
        actual_country_count = len(countries_by_company.get(name, set()))
        expected_country_count = safe_int(company.get("Country_Count"))
        if expected_country_count and actual_country_count and expected_country_count != actual_country_count:
            add_quality_issue(
                issues,
                source_table="Companies",
                source_row_id=name,
                entity_type="company",
                entity_name=name,
                issue_type="derived_country_count_mismatch",
                severity="low",
                field_name="Country_Count",
                current_value=expected_country_count,
                suggested_value=actual_country_count,
                description="Companies.Country_Count does not match countries represented in Product_Lines.",
                action="确认 Countries_All 是总部/生产地/来源地中的哪一种口径。",
            )

    for key, names in company_key_to_names.items():
        if len(names) > 1:
            canonical = sorted(names, key=lambda value: (value.lower() != value, len(value)))[0]
            add_quality_issue(
                issues,
                source_table="Companies",
                source_row_id=", ".join(names),
                entity_type="company",
                entity_name=", ".join(names),
                issue_type="company_case_or_alias_collision",
                severity="high",
                field_name="Company",
                current_value=", ".join(names),
                suggested_value=canonical,
                description="Company names collapse to the same normalized key.",
                action="建立 canonical company mapping，避免公司主体重复统计。",
            )

    product_brand_keys = {
        (compact_key(product.get("Company")), compact_key(product.get("Brand")))
        for product in products
        if product.get("Company") and product.get("Brand") and not is_non_primary_record(product)
    }
    brand_portfolio_keys = {(compact_key(brand.get("Company")), compact_key(brand.get("Brand"))) for brand in brands if brand.get("Company") and brand.get("Brand")}
    for company_key, brand_key in sorted(product_brand_keys - brand_portfolio_keys)[:250]:
        add_quality_issue(
            issues,
            source_table="Brand_Portfolio",
            source_row_id=f"{company_key}/{brand_key}",
            entity_type="brand",
            entity_name=f"{company_key}/{brand_key}",
            issue_type="brand_missing_from_portfolio",
            severity="medium",
            description="A company-brand pair exists in Product_Lines but not in Brand_Portfolio.",
            action="从 Product_Lines 重新生成 Brand_Portfolio 或补充缺失品牌。",
        )
    # Brand_Portfolio is allowed to retain reference-only, excluded, accessory, and
    # historical brand traces even when the active Product_Lines master has been
    # normalized to a different canonical product family.

    corporate_rows = [
        product
        for product in products
        if (product.get("Brand_Type") or "").lower() == "corporate" and not is_non_primary_record(product)
    ]
    if corporate_rows:
        add_quality_issue(
            issues,
            source_table="Product_Lines",
            source_row_id="aggregate",
            entity_type="product",
            entity_name="Corporate rows in Product_Lines",
            issue_type="corporate_records_mixed_with_products",
            severity="medium",
            current_value=len(corporate_rows),
            description="Corporate/company/service records are mixed into Product_Lines.",
            action="后续整理时拆到 Company_Master 或 Service/Platform 表，避免进入产品注册核验。",
        )

    missing_product_uuid = sum(1 for product in products if not has_value(product, "Product_UUID"))
    if missing_product_uuid:
        add_quality_issue(
            issues,
            source_table="Product_Lines",
            source_row_id="aggregate",
            entity_type="product",
            entity_name="Product_UUID",
            issue_type="stable_product_id_missing",
            severity="medium",
            field_name="Product_UUID",
            current_value=missing_product_uuid,
            description="Product_UUID is missing for seed rows; product_master generates deterministic IDs as a temporary bridge.",
            action="在人工整理通过后，为产品/品牌/SKU 层级写入稳定 UUID。",
        )

    issues = sorted(
        issues,
        key=lambda issue: (
            SEVERITY_ORDER.get(issue["severity"], 99),
            issue["issue_type"],
            issue["source_table"],
            issue["entity_name"],
        ),
    )
    summary = data_quality_summary(issues)
    return {"issues": issues, "summary": summary}


def write_data_quality_outputs(quality: dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    issues = quality.get("issues", [])
    summary = quality.get("summary", {})
    fieldnames = [
        "issue_id",
        "severity",
        "issue_type",
        "source_table",
        "source_row_id",
        "entity_type",
        "entity_name",
        "field_name",
        "current_value",
        "suggested_value",
        "description",
        "action",
        "status",
    ]
    with DATA_QUALITY_ISSUES_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for issue in issues:
            writer.writerow({field: issue.get(field, "") for field in fieldnames})
    DATA_QUALITY_SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = [
        "# Seed Integrity Report",
        "",
        f"- Generated: {datetime.now(timezone.utc).astimezone().isoformat(timespec='seconds')}",
        f"- Total open issues: {summary.get('total', 0)}",
        f"- Critical: {summary.get('critical', 0)}",
        f"- High: {summary.get('high', 0)}",
        f"- Medium: {summary.get('medium', 0)}",
        f"- Low: {summary.get('low', 0)}",
        "",
        "## Issue Types",
        "",
    ]
    for item in summary.get("by_type", []):
        lines.append(f"- {item['name']}: {item['value']}")
    lines.extend(["", "## Top Issues", ""])
    for issue in summary.get("top_issues", []):
        lines.append(
            f"- [{issue['severity']}] {issue['issue_type']} · {issue['source_table']} · "
            f"{issue['entity_name']} · {issue['description']} Action: {issue['action']}"
        )
    DATA_QUALITY_REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def detect_segments(text: str, category_l1: str = "") -> list[str]:
    blob = f" {text.lower()} "
    found = []
    if norm(category_l1).lower() == "ebd":
        found.append("ebd")
    for segment in SEGMENTS:
        for term in segment["terms"]:
            t = term.lower()
            if len(t) <= 3 and t.isascii():
                if re.search(rf"(^|[^a-z0-9]){re.escape(t.strip())}([^a-z0-9]|$)", blob):
                    found.append(segment["code"])
                    break
            elif t in blob:
                found.append(segment["code"])
                break
    result = []
    for code in found:
        if code not in result:
            result.append(code)
    strong_material_priority = [
        (
            "caha",
            [
                "harmonyca",
                "caha",
                "calcium hydroxylapatite",
                "calcium hydroxyapatite",
                "hydroxylapatite",
                "羟基磷灰石",
                "radiesse",
                "facetem",
            ],
        ),
        (
            "pcl",
            [
                "pcl",
                "polycaprolactone",
                "poly-ɛ-caprolactone",
                "poly-e-caprolactone",
                "ellanse",
                "gouri",
                "bravity",
                "聚己内酯",
            ],
        ),
        (
            "plla",
            [
                "plla",
                "pdlla",
                "poly-l-lactic",
                "poly lactic",
                "sculptra",
                "aesthefill",
                "lanluma",
                "聚左乳酸",
                "聚乳酸",
            ],
        ),
    ]
    for priority_code, terms in strong_material_priority:
        if priority_code in result and any(matches_term(blob, term) for term in terms):
            result = [code for code in result if code not in {"ha", "exosome"} or code == priority_code]
            if priority_code not in result:
                result.insert(0, priority_code)
            else:
                result = [priority_code] + [code for code in result if code != priority_code]
            break
    return result or ["other"]


def detect_year(*values: Any) -> int | None:
    for value in values:
        match = re.search(r"(20\d{2}|19\d{2})", norm(value))
        if match:
            return int(match.group(1))
    return None


def top_counts(counter: Counter, limit: int = 12) -> list[dict[str, Any]]:
    return [{"name": key, "value": value} for key, value in counter.most_common(limit) if key]


def matches_term(blob: str, term: str) -> bool:
    needle = term.lower().strip()
    if not needle:
        return False
    if len(needle) <= 3 and needle.isascii():
        return bool(re.search(rf"(^|[^a-z0-9]){re.escape(needle)}([^a-z0-9]|$)", blob))
    return needle in blob


def detect_taxonomy(text: str, segment_code: str, kind: str, fallback: str | None = "") -> list[str]:
    taxonomy = SEGMENT_TAXONOMY.get(segment_code, {}).get(kind, [])
    blob = f" {text.lower()} "
    found: list[str] = []
    for item in taxonomy:
        if any(matches_term(blob, term) for term in item.get("terms", [])):
            found.append(item["name"])
    if not found and fallback:
        found.append(fallback)
    if not found and fallback is None:
        return []
    if not found:
        found.append("未归类信号")
    return found


def normalize_product_subtracks(product: dict[str, Any], segment_code: str, subtracks: list[str]) -> list[str]:
    blob = f" {text_blob(product).lower()} "
    configured = {
        item["name"]
        for item in SEGMENT_TAXONOMY.get(segment_code, {}).get("subtracks", [])
        if item.get("name")
    }

    def keep_configured(values: list[str], limit: int | None = None) -> list[str]:
        kept = []
        for value in values:
            if value in configured and value not in kept:
                kept.append(value)
        return kept[:limit] if limit else kept

    if segment_code == "ha":
        allowed = {
            "Filler / 交联填充剂",
            "Skin Booster / 无交联水光",
            "中胚层 HA 复配液",
            "HA 主成分复配",
        }
        normalized = [name for name in subtracks if name in allowed]
        product_shape_fields = " ".join(
            norm(product.get(key)).lower()
            for key in ["Category_L1", "Category_L2", "Tech_Type_Std", "Tech_Type_Original", "Core_Product"]
        )
        is_ha_product = any(
            term in product_shape_fields
            for term in ["hyaluronic", "hyaluron", "透明质酸", "玻尿酸", "ha dermal", "ha filler", "skin booster", "mesotherapy"]
        )
        if normalized:
            return normalized
        if is_ha_product:
            return ["HA 未细分产品形态"]
        return []

    if segment_code == "pn_pdrn":
        shape_blob = " ".join(
            norm(product.get(key)).lower()
            for key in ["Brand", "Core_Product", "Tech_Type_Std", "Tech_Type_Original", "Category_L2"]
        )
        pn_blob = " ".join(
            norm(product.get(key)).lower()
            for key in ["Brand", "Core_Product", "Tech_Type_Std", "Tech_Type_Original", "Category_L2", "Introduction", "Feature_Tags"]
        )
        has_api = any(term in shape_blob for term in [" api", "raw material", "ingredient", "原料"])
        has_pdrn = any(term in pn_blob for term in ["pdrn", "polydeoxyribonucleotide", "salmon dna", "s-dna"])
        has_pn = bool(re.search(r"(^|[^a-z0-9])pn([^a-z0-9]|$)", pn_blob)) or any(
            term in pn_blob for term in ["polynucleotide", "rejuran", "nucleofill", "philart", "聚核苷酸"]
        )
        known_pn_brand = any(term in shape_blob for term in ["rejuran", "nucleofill", "philart", "plinest"])
        has_complex = any(
            term in shape_blob
            for term in [
                "pn/pdrn",
                "pdrn/pn",
                "pn / pdrn",
                "pdrn / pn",
                "ha-pdrn",
                "hyla-pdrn",
                "pdrn +",
                "pn +",
                "skin booster",
                "meso",
                "mesotherapy",
                "cocktail",
                "complex",
                "复合",
            ]
        )
        if has_api:
            return ["PN/PDRN 原料/API"]
        if known_pn_brand:
            return ["PN 制剂"]
        if has_complex:
            return ["PN/PDRN 复合制剂"]
        if has_pdrn:
            return ["PDRN 制剂"]
        if has_pn:
            return ["PN 制剂"]
        normalized = [name for name in subtracks if name in {"PN 制剂", "PDRN 制剂", "PN/PDRN 复合制剂", "PN/PDRN 原料/API"}]
        return normalized[:1]

    if segment_code == "mesotherapy":
        shape_blob = " ".join(
            norm(product.get(key)).lower()
            for key in [
                "Brand",
                "Core_Product",
                "Tech_Type_Std",
                "Tech_Type_Original",
                "Category_L1",
                "Category_L2",
            ]
        )
        meso_blob = " ".join(
            norm(product.get(key)).lower()
            for key in [
                "Brand",
                "Core_Product",
                "Tech_Type_Std",
                "Tech_Type_Original",
                "Category_L1",
                "Category_L2",
                "Introduction",
                "Feature_Tags",
            ]
        )
        allowed = {
            "复配注射液 / Cocktail",
            "HA 基底复配液 / HA-based cocktail",
            "生物活性成分复配",
            "注射设备 / 针头耗材",
        }
        if any(
            term in shape_blob
            for term in [
                "mesogun",
                "meso gun",
                "mesotherapy gun",
                "mpgun",
                " gun",
                "injector",
                "multi-injector",
                "multi injector",
                "needle",
                "cannula",
                "vacuum mesotherapy",
                "水晶针",
                "针头",
                "注射枪",
            ]
        ):
            return ["注射设备 / 针头耗材"]
        if any(
            term in meso_blob
            for term in [
                "biorevitalizer",
                "bio-stimulation",
                "growth factor",
                "exosome",
                "pdrn",
                "polynucleotide",
                "peptide",
                "amino acid",
                "vitamin",
                "mi-rna",
                "mirna",
                "nctc",
                "细胞因子",
                "生长因子",
                "外泌体",
            ]
        ):
            return ["生物活性成分复配"]
        if any(
            term in meso_blob
            for term in [
                "hyaluronic acid",
                "hyaluron",
                " ha ",
                "ha-",
                " ha/",
                "skin booster",
                "mesoheal",
                "mesolift",
                "rrs ha",
                "透明质酸",
                "玻尿酸",
                "水光针",
            ]
        ):
            return ["HA 基底复配液 / HA-based cocktail"]
        if any(
            term in meso_blob
            for term in ["meso", "mesotherapy", "cocktail", "solution", "vial", "ampoule", "中胚层", "复配"]
        ):
            return ["复配注射液 / Cocktail"]
        normalized = [name for name in subtracks if name in allowed]
        return normalized[:1]

    if segment_code == "pcl":
        has_pcl_material = any(
            term in blob
            for term in [
                "pcl",
                "polycaprolactone",
                "poly-ɛ-caprolactone",
                "poly-e-caprolactone",
                "聚己内酯",
            ]
        )
        has_known_pcl_brand = any(
            term in blob
            for term in [
                "gouri",
                "bravity",
                "ellanse",
                "miracle l",
                "miracle h",
                "miracle touch",
                "dexlevo",
            ]
        )
        if not (has_pcl_material or has_known_pcl_brand):
            return []

        if any(
            term in blob
            for term in [
                "pcl thread",
                "pcl threads",
                "pcl cog",
                "pcl lifting thread",
                "pcl suture",
                "thread lift",
                "线材",
                "线雕",
                "提拉线",
            ]
        ):
            return []

        if any(
            term in blob
            for term in [
                "ellanse",
                "pcl microsphere",
                "polycaprolactone microsphere",
                "poly-ɛ-caprolactone",
                "pcl-based dermal filler",
                "聚己内酯微球",
                "pcl晶球",
            ]
        ):
            return ["PCL 微球填充剂"]

        if any(
            term in blob
            for term in [
                "liquid pcl",
                "fully solubilized",
                "solubilized pcl",
                "cesabp",
                "gouri",
                "bravity",
                "miracle l",
                "miracle h",
                "miracle touch",
                "液态pcl",
                "液体线",
            ]
        ):
            return ["液态 PCL"]

        category = f" {norm(product.get('Category_L2')).lower()} {norm(product.get('Core_Product')).lower()} "
        if "thread" in category or "线" in category:
            return []
        if "filler" in category or "inject" in category or "填充" in category:
            return ["PCL 微球填充剂"]
        return ["PCL 未细分形态"]

    if segment_code != "caha":
        return keep_configured(subtracks)

    if any(term in blob for term in ["mti-12", "chitosan", "chitin-glucan", "chitin glucan"]):
        return []

    shape_blob = " ".join(
        norm(product.get(key)).lower()
        for key in [
            "Brand",
            "Core_Product",
            "Tech_Type_Std",
            "Tech_Type_Original",
            "Feature_Tags",
            "Verified_Product_Type_CN",
        ]
    )
    normalized: list[str] = []
    is_hybrid = any(
        term in shape_blob
        for term in [
            "harmonyca",
            "stimulate",
            "ha-caha-hybrid",
            "ha + caha",
            "ha/caha",
            "caha + ha",
            "peg-ha + caha",
            "peg-ha + 1% caha",
            "hybrid injectable",
        ]
    )
    is_caha = any(
        term in shape_blob
        for term in [
            "caha",
            "calcium hydroxylapatite",
            "calcium hydroxyapatite",
            "hydroxylapatite",
            "radiesse",
            "facetem",
            "羟基磷灰石",
        ]
    )
    if is_hybrid:
        normalized.append("HA + CaHA 复合注射剂")
    elif is_caha:
        normalized.append("CaHA 微球/胶原刺激剂")

    return keep_configured(normalized)


def heatmap_from_rows(
    rows: list[dict[str, Any]],
    row_values: dict[str, list[str]],
    columns: list[str],
    column_key: str,
    limit: int = 8,
) -> dict[str, Any]:
    row_counter: Counter = Counter()
    matrix: dict[str, Counter] = defaultdict(Counter)
    column_set = set(columns)
    for product in rows:
        col = product.get(column_key) or "Unknown"
        if col not in column_set:
            continue
        values = row_values.get(product.get("Record_ID"), [])
        for name in values:
            row_counter[name] += 1
            matrix[name][col] += 1
    selected_rows = [name for name, _ in row_counter.most_common(limit)]
    return {
        "columns": columns,
        "rows": [
            {
                "name": name,
                "total": row_counter[name],
                "values": {column: matrix[name].get(column, 0) for column in columns},
            }
            for name in selected_rows
        ],
    }


REGULATORY_CHANNELS = [
    {
        "code": "fda",
        "name": "FDA / 510(k)",
        "region": "United States",
        "fields": ["FDA_Status", "FDA_510k_Number", "FDA_Approval_Date"],
        "terms": ["fda 510(k)", "510k", "510(k)", "fda clearance", "fda cleared", "fda approval"],
        "markets": ["United States"],
        "kind": "market_authorization",
    },
    {
        "code": "ce",
        "name": "CE / MDR",
        "region": "European Union",
        "fields": ["CE_Status", "CE_Year"],
        "terms": ["ce mark", "ce marked", "ce certification", "medical device regulation", "mdr"],
        "markets": ["EU / EEA"],
        "kind": "conformity_mark",
    },
    {
        "code": "nmpa",
        "name": "NMPA",
        "region": "China",
        "fields": ["NMPA_Status", "NMPA_Reg_Number", "NMPA_Approval_Date"],
        "terms": ["nmpa", "cfda", "china registration", "国械注", "械注准"],
        "markets": ["China"],
        "kind": "market_authorization",
    },
    {
        "code": "kfda",
        "name": "MFDS / KFDA",
        "region": "South Korea",
        "fields": ["KFDA_Status"],
        "terms": ["mfds", "kfda", "korea mfds", "korean approval"],
        "markets": ["South Korea"],
        "kind": "market_authorization",
    },
    {
        "code": "mdsap",
        "name": "MDSAP",
        "region": "Audit program",
        "fields": ["MDSAP_Status", "MDSAP_Certificate"],
        "terms": ["mdsap", "medical device single audit program"],
        "markets": ["Australia", "Brazil", "Canada", "Japan", "United States"],
        "kind": "qms_audit",
        "license_effect": "quality_system_audit_not_sales_authorization",
        "note": "MDSAP can support quality-system audit requirements across participating authorities, but it is not a direct sales license. Australia ARTG, Brazil ANVISA, Canada MDL, Japan PMDA/MHLW and US FDA authorization/listing remain separate market checks.",
    },
    {
        "code": "tga_artg",
        "name": "TGA / ARTG",
        "region": "Australia",
        "fields": ["TGA_Status", "ARTG_Number"],
        "terms": ["artg", "therapeutic goods administration", "tga included", "australian register of therapeutic goods"],
        "markets": ["Australia"],
        "kind": "market_authorization",
    },
    {
        "code": "anvisa",
        "name": "ANVISA",
        "region": "Brazil",
        "fields": ["ANVISA_Status", "ANVISA_Number"],
        "terms": ["anvisa", "brazil registration", "registro anvisa", "brazilian health regulatory agency"],
        "markets": ["Brazil"],
        "kind": "market_authorization",
    },
    {
        "code": "health_canada",
        "name": "Health Canada MDL",
        "region": "Canada",
        "fields": ["Health_Canada_Status", "MDL_Number"],
        "terms": ["health canada", "medical device licence", "medical device license", "mdl"],
        "markets": ["Canada"],
        "kind": "market_authorization",
    },
    {
        "code": "pmda_mhlw",
        "name": "PMDA / MHLW",
        "region": "Japan",
        "fields": ["PMDA_Status", "MHLW_Status"],
        "terms": ["pmda", "mhlw", "japan approval", "japanese approval"],
        "markets": ["Japan"],
        "kind": "market_authorization",
    },
    {
        "code": "mhra_ukca",
        "name": "MHRA / UKCA",
        "region": "United Kingdom",
        "fields": ["MHRA_Status", "UKCA_Status"],
        "terms": ["mhra", "ukca", "uk responsible person", "great britain medical device"],
        "markets": ["United Kingdom"],
        "kind": "registration_and_marking",
    },
    {
        "code": "hsa_singapore",
        "name": "HSA / SMDR",
        "region": "Singapore",
        "fields": ["HSA_Status", "SMDR_Number"],
        "terms": ["hsa singapore", "singapore hsa", "smdr", "singapore medical device register"],
        "markets": ["Singapore"],
        "kind": "market_authorization",
    },
    {
        "code": "malaysia_mda",
        "name": "Malaysia MDA",
        "region": "Malaysia",
        "fields": ["Malaysia_MDA_Status", "MDA_Reg_Number"],
        "terms": ["malaysia mda", "medical device authority malaysia", "malaysian medical device registration"],
        "markets": ["Malaysia"],
        "kind": "market_authorization",
    },
    {
        "code": "thai_fda",
        "name": "Thai FDA",
        "region": "Thailand",
        "fields": ["Thai_FDA_Status", "Thailand_Reg_Number"],
        "terms": ["thai fda", "thailand fda", "thailand medical device registration"],
        "markets": ["Thailand"],
        "kind": "market_authorization",
    },
    {
        "code": "indonesia_moh",
        "name": "Indonesia MoH / AKL",
        "region": "Indonesia",
        "fields": ["Indonesia_MoH_Status", "AKL_Number"],
        "terms": ["kemenkes", "akl", "indonesia ministry of health", "indonesia medical device registration"],
        "markets": ["Indonesia"],
        "kind": "market_authorization",
    },
    {
        "code": "philippines_fda",
        "name": "Philippines FDA",
        "region": "Philippines",
        "fields": ["Philippines_FDA_Status", "CMDN_Number", "CMDR_Number"],
        "terms": ["philippines fda", "cmdn", "cmdr", "philippines medical device registration"],
        "markets": ["Philippines"],
        "kind": "market_authorization",
    },
    {
        "code": "vietnam_moh",
        "name": "Vietnam MoH / DMEC",
        "region": "Vietnam",
        "fields": ["Vietnam_MoH_Status", "DMEC_Number"],
        "terms": ["vietnam moh", "dmec", "vietnam medical device registration"],
        "markets": ["Vietnam"],
        "kind": "market_authorization",
    },
    {
        "code": "cofepris",
        "name": "COFEPRIS",
        "region": "Mexico",
        "fields": ["COFEPRIS_Status", "COFEPRIS_Number"],
        "terms": ["cofepris", "mexico medical device registration"],
        "markets": ["Mexico"],
        "kind": "market_authorization",
    },
    {
        "code": "sfda_saudi",
        "name": "Saudi SFDA",
        "region": "Saudi Arabia",
        "fields": ["Saudi_SFDA_Status", "MDMA_Number"],
        "terms": ["saudi sfda", "sfda medical device", "mdma", "saudi medical device registration"],
        "markets": ["Saudi Arabia"],
        "kind": "market_authorization",
    },
    {
        "code": "tfda_taiwan",
        "name": "Taiwan TFDA",
        "region": "Taiwan",
        "fields": ["Taiwan_TFDA_Status", "TFDA_Number"],
        "terms": ["taiwan tfda", "tfda medical device", "taiwan medical device registration"],
        "markets": ["Taiwan"],
        "kind": "market_authorization",
    },
]

REGULATORY_ATLAS = [
    {
        "code": item["code"],
        "name": item["name"],
        "region": item["region"],
        "kind": item["kind"],
        "markets": item["markets"],
        "license_effect": item.get(
            "license_effect",
            "direct_or_primary_market_authorization" if item["kind"] == "market_authorization" else item["kind"],
        ),
        "note": item.get("note", ""),
        "current_scope": "external_project" if item["code"] in EXTERNAL_PROJECT_CHANNELS else ("active_phase" if item["code"] in CURRENT_PHASE_CHANNELS else "roadmap"),
        "scope_note": EXTERNAL_PROJECT_NOTES.get(item["code"], ""),
    }
    for item in REGULATORY_CHANNELS
]

APPROVAL_YEAR_FIELDS = [
    ("fda", "FDA_Approval_Date"),
    ("ce", "CE_Year"),
    ("nmpa", "NMPA_Approval_Date"),
    ("launch", "Global_Launch_Year"),
]


def has_value(product: dict[str, Any], field: str) -> bool:
    return bool(norm(product.get(field)))


def is_non_primary_record(record: dict[str, Any]) -> bool:
    value = norm(record.get("Is_Primary_Record")).lower()
    duplicate_note = norm(record.get("Duplicate_Note")).lower()
    return value in {"0", "false", "no", "n"} or duplicate_note.startswith("duplicate_of:")


def regulatory_counts(products: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in REGULATORY_CHANNELS:
        terms = [term.lower() for term in item.get("terms", [])]
        fields = item.get("fields", [])
        counts[item["code"]] = sum(
            1
            for product in products
            if any(has_value(product, field) for field in fields)
            or any(term in text_blob(product).lower() for term in terms)
        )
    return counts


def has_regulatory_signal(product: dict[str, Any]) -> bool:
    blob = text_blob(product).lower()
    return any(
        any(has_value(product, field) for field in item.get("fields", []))
        or any(term.lower() in blob for term in item.get("terms", []))
        for item in REGULATORY_CHANNELS
    )


def extract_year(value: Any) -> int | None:
    text = norm(value)
    if not text:
        return None
    match = re.search(r"(19|20)\d{2}", text)
    if not match:
        return None
    year = int(match.group(0))
    if 1990 <= year <= 2035:
        return year
    return None


def approval_timeline(products: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    years: dict[int, Counter] = defaultdict(Counter)
    seen: set[tuple[str, int, str]] = set()
    for product in products:
        record_id = product.get("Record_ID") or product.get("Product_UUID") or text_blob(product, ["Company", "Brand", "Core_Product"])
        for channel, field in APPROVAL_YEAR_FIELDS:
            year = extract_year(product.get(field))
            if year is None:
                continue
            key = (record_id, year, channel)
            if key in seen:
                continue
            seen.add(key)
            years[year][channel] += 1
    selected = sorted(years)[-limit:]
    return [
        {
            "year": str(year),
            "total": sum(years[year].values()),
            "fda": years[year].get("fda", 0),
            "ce": years[year].get("ce", 0),
            "nmpa": years[year].get("nmpa", 0),
            "launch": years[year].get("launch", 0),
        }
        for year in selected
    ]


def evidence_funnel(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    regulatory = regulatory_counts(products)
    return [
        {"name": "结构化产品线", "value": len(products), "note": "Product_Lines"},
        {"name": "有监管/质量字段", "value": sum(1 for product in products if has_regulatory_signal(product)), "note": "19类监管字典；MDSAP仅作QMS审核"},
        {"name": "FDA / 510(k)", "value": regulatory.get("fda", 0), "note": "美国准入线索"},
        {"name": "CE / MDR", "value": regulatory.get("ce", 0), "note": "欧洲准入线索"},
        {"name": "MDSAP", "value": regulatory.get("mdsap", 0), "note": "质量体系审核，不等于销售许可"},
        {"name": "NMPA / MFDS", "value": regulatory.get("nmpa", 0) + regulatory.get("kfda", 0), "note": "中韩准入线索"},
    ]


def regulatory_mix(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts = regulatory_counts(products)
    return [
        {
            "name": item["name"],
            "value": counts.get(item["code"], 0),
            "region": item["region"],
            "kind": item["kind"],
            "markets": item["markets"],
        }
        for item in REGULATORY_CHANNELS
    ]


def percent(value: float, total: float) -> int:
    if not total:
        return 0
    return int(round((value / total) * 100))


def maturity_lenses(
    products: list[dict[str, Any]],
    product_subtracks: dict[str, list[str]],
    product_indications: dict[str, list[str]],
) -> list[dict[str, Any]]:
    total = len(products)
    company_total = len({p.get("Company") for p in products if p.get("Company")})
    country_total = len({p.get("Country") for p in products if p.get("Country")})
    regulatory_total = sum(1 for product in products if has_regulatory_signal(product))
    subtrack_total = len({name for names in product_subtracks.values() for name in names})
    indication_total = len({name for names in product_indications.values() for name in names})
    public_total = sum(1 for product in products if (product.get("Ownership") or "").lower() == "public" or has_value(product, "Stock_Code"))
    return [
        {"name": "二级目录深度", "value": subtrack_total, "unit": "类", "score": min(100, subtrack_total * 16)},
        {"name": "适应症宽度", "value": indication_total, "unit": "类", "score": min(100, indication_total * 14)},
        {"name": "国家覆盖", "value": country_total, "unit": "国", "score": min(100, country_total * 7)},
        {"name": "企业密度", "value": company_total, "unit": "家", "score": min(100, company_total * 4)},
        {"name": "监管字段覆盖", "value": percent(regulatory_total, total), "unit": "%", "score": percent(regulatory_total, total)},
        {"name": "上市主体参与", "value": public_total, "unit": "条", "score": min(100, public_total * 8)},
    ]


def taxonomy_matrix(
    products: list[dict[str, Any]],
    product_values: dict[str, list[str]],
    columns: list[str],
    dimension_key: str,
    limit: int = 10,
) -> dict[str, Any]:
    column_counter: Counter = Counter()
    for product in products:
        column_counter.update(product_values.get(product.get("Record_ID") or "", []))
    selected_columns = [name for name in columns if column_counter.get(name, 0)]
    if not selected_columns:
        selected_columns = [name for name, _count in column_counter.most_common(6)]
    selected_columns = selected_columns[:8]
    row_counter: Counter = Counter()
    matrix: dict[str, Counter] = defaultdict(Counter)
    column_set = set(selected_columns)
    for product in products:
        row_name = product.get(dimension_key) or "Unknown"
        matched = [name for name in product_values.get(product.get("Record_ID") or "", []) if name in column_set]
        if not matched:
            continue
        row_counter[row_name] += 1
        for name in matched:
            matrix[row_name][name] += 1
    selected_rows = [name for name, _count in row_counter.most_common(limit)]
    return {
        "columns": selected_columns,
        "rows": [
            {
                "name": name,
                "total": sum(matrix[name].get(column, 0) for column in selected_columns),
                "product_total": row_counter[name],
                "values": {column: matrix[name].get(column, 0) for column in selected_columns},
            }
            for name in selected_rows
        ],
    }


def role_mix(products: list[dict[str, Any]], field: str, limit: int = 8) -> list[dict[str, Any]]:
    return top_counts(Counter(product.get(field) or "Unknown" for product in products), limit)


def enrich_analysis(
    products: list[dict[str, Any]],
    product_subtracks: dict[str, list[str]],
    product_indications: dict[str, list[str]],
    configured_subtracks: list[str],
    configured_indications: list[str],
) -> dict[str, Any]:
    return {
        "analysis_lenses": maturity_lenses(products, product_subtracks, product_indications),
        "approval_timeline": approval_timeline(products),
        "evidence_funnel": evidence_funnel(products),
        "regulatory_mix": regulatory_mix(products),
        "company_subtrack_matrix": taxonomy_matrix(products, product_subtracks, configured_subtracks, "Company", 10),
        "country_subtrack_matrix": taxonomy_matrix(products, product_subtracks, configured_subtracks, "Country", 10),
        "company_indication_matrix": taxonomy_matrix(
            products,
            product_indications,
            configured_indications,
            "Company",
            8,
        ),
        "business_roles": role_mix(products, "Business_Role", 7),
        "ownership_mix": role_mix(products, "Ownership", 7),
        "category_l2_mix": role_mix(products, "Category_L2", 8),
        "tech_type_mix": role_mix(products, "Tech_Type_Std", 8),
    }


def product_preview(
    product: dict[str, Any],
    product_subtracks: dict[str, list[str]],
    product_indications: dict[str, list[str]],
) -> dict[str, Any]:
    product_id = product.get("Record_ID") or ""
    return {
        "record_id": product.get("Record_ID"),
        "company": product.get("Company"),
        "brand": product.get("Brand"),
        "country": product.get("Country"),
        "region": dashboard_region(product.get("Region"), product.get("Country")),
        "tech": product.get("Tech_Type_Std"),
        "core_product": product.get("Core_Product"),
        "category": product.get("Category_L2"),
        "subtracks": product_subtracks.get(product_id, []),
        "indications": product_indications.get(product_id, []),
    }


def representative_products(
    products: list[dict[str, Any]],
    product_subtracks: dict[str, list[str]],
    product_indications: dict[str, list[str]],
    limit: int = 240,
) -> list[dict[str, Any]]:
    company_counter = Counter(product.get("Company") or "Unknown" for product in products)
    brand_counter = Counter(product.get("Brand") or product.get("Core_Product") or "Unknown" for product in products)
    ranked = sorted(
        products,
        key=lambda product: (
            -company_counter[product.get("Company") or "Unknown"],
            -brand_counter[product.get("Brand") or product.get("Core_Product") or "Unknown"],
            product.get("Company") or "",
            product.get("Brand") or "",
            product.get("Core_Product") or "",
        ),
    )
    return [product_preview(product, product_subtracks, product_indications) for product in ranked[:limit]]


def build_subtrack_slices(
    seg_products: list[dict[str, Any]],
    product_subtracks: dict[str, list[str]],
    product_indications: dict[str, list[str]],
    subtrack_counter: Counter,
    limit: int = 12,
    evidence_scope_builder: Any | None = None,
) -> list[dict[str, Any]]:
    slices: list[dict[str, Any]] = []
    for subtrack_name, total in subtrack_counter.most_common(limit):
        slice_products = [
            product
            for product in seg_products
            if subtrack_name in product_subtracks.get(product.get("Record_ID") or "", [])
        ]
        if not slice_products:
            continue
        region_counter = Counter(dashboard_region(product.get("Region"), product.get("Country")) for product in slice_products)
        slice_products_by_region = [
            {**product, "Region": dashboard_region(product.get("Region"), product.get("Country"))}
            for product in slice_products
        ]
        country_counter = Counter(product.get("Country") or "Unknown" for product in slice_products)
        company_counter = Counter(product.get("Company") or "Unknown" for product in slice_products)
        brand_counter = Counter(product.get("Brand") or product.get("Core_Product") or "Unknown" for product in slice_products)
        indication_counter: Counter = Counter()
        sibling_subtrack_counter: Counter = Counter()
        for product in slice_products:
            product_id = product.get("Record_ID") or ""
            indication_counter.update(product_indications.get(product_id, []))
            sibling_subtrack_counter.update(product_subtracks.get(product_id, []))
        region_columns = [name for name, _ in region_counter.most_common(5)] or ["Unknown"]
        regulatory = regulatory_counts(slice_products)
        configured_indications = [name for name, _ in indication_counter.most_common(8)]
        configured_subtracks = [name for name, _ in sibling_subtrack_counter.most_common(8)] or [subtrack_name]
        slice_payload = {
            "name": subtrack_name,
            "products": len(slice_products),
            "companies": len({product.get("Company") for product in slice_products if product.get("Company")}),
            "brands": len({product.get("Brand") for product in slice_products if product.get("Brand")}),
            "countries": len({product.get("Country") for product in slice_products if product.get("Country")}),
            "subtrack_count": len(sibling_subtrack_counter),
            "indication_count": len(indication_counter),
            "top_subtracks": top_counts(sibling_subtrack_counter, 8),
            "top_indications": top_counts(indication_counter, 8),
            "subtrack_heatmap": heatmap_from_rows(slice_products_by_region, product_subtracks, region_columns, "Region", 8),
            "indication_heatmap": heatmap_from_rows(slice_products_by_region, product_indications, region_columns, "Region", 8),
            "top_regions": top_counts(region_counter, 6),
            "top_countries": top_counts(country_counter, 6),
            "top_companies": top_counts(company_counter, 8),
            "top_brands": top_counts(brand_counter, 8),
            "regulatory": regulatory,
            **enrich_analysis(
                slice_products,
                product_subtracks,
                product_indications,
                configured_subtracks,
                configured_indications,
            ),
            "sample_products": representative_products(slice_products, product_subtracks, product_indications, 160),
        }
        if evidence_scope_builder:
            slice_payload["evidence_scope"] = evidence_scope_builder(slice_products)
        slices.append(slice_payload)
    return slices


def load_products(company_book: Path) -> list[dict[str, Any]]:
    rows = read_sheet_dicts(company_book, "Product_Lines")
    products = []
    for row in rows:
        record = {key: norm(value) for key, value in row.items()}
        record["Company"] = canonical_company_name(record.get("Company"))
        record = apply_product_fact_override(record)
        blob = text_blob(record)
        override = PRODUCT_FACT_OVERRIDES.get(norm(record.get("Record_ID")))
        segments = [item.strip() for item in norm(override.get("_segments") if override else "").split(",") if item.strip()]
        if not segments:
            segments = detect_segments(blob, record.get("Category_L1", ""))
        record["Segments"] = ",".join(segments)
        record["Primary_Segment"] = segments[0]
        products.append(record)
    return products


def split_multi_value(value: Any) -> list[str]:
    return [part.strip() for part in re.split(r"[;,/|]+", norm(value)) if part.strip()]


def merge_company_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    additive_fields = {"Product_Count", "Brand_Count", "Country_Count", "FDA_Products", "NMPA_Products"}
    union_fields = {"Countries_All"}
    preferred_long_fields = {"Location_Full"}

    for record in records:
        record["Company"] = canonical_company_name(record.get("Company"))
        key = compact_key(record.get("Company"))
        if not key:
            continue
        if key not in merged:
            merged[key] = dict(record)
            order.append(key)
            continue

        target = merged[key]
        for field, value in record.items():
            if field in additive_fields:
                target[field] = str(safe_int(target.get(field)) + safe_int(value))
            elif field in union_fields:
                values = []
                for item in [*split_multi_value(target.get(field)), *split_multi_value(value)]:
                    if item and item not in values:
                        values.append(item)
                target[field] = ", ".join(values)
            elif field in preferred_long_fields:
                if len(norm(value)) > len(norm(target.get(field))):
                    target[field] = value
            elif not norm(target.get(field)) and norm(value):
                target[field] = value

    return [merged[key] for key in order]


def load_companies(company_book: Path) -> list[dict[str, Any]]:
    rows = read_sheet_dicts(company_book, "Companies")
    return merge_company_records([{key: norm(value) for key, value in row.items()} for row in rows])


def load_brands(company_book: Path) -> list[dict[str, Any]]:
    rows = read_sheet_dicts(company_book, "Brand_Portfolio")
    brands = []
    for row in rows:
        record = {key: norm(value) for key, value in row.items()}
        record["Company"] = canonical_company_name(record.get("Company"))
        record["Product_Count"] = int(safe_float(record.get("Product_Count")) or 0)
        record["Segments"] = ",".join(detect_segments(text_blob(record), record.get("Category_L1", "")))
        brands.append(record)
    return brands


def apply_dashboard_scope(
    products: list[dict[str, Any]],
    companies: list[dict[str, Any]],
    brands: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Filter the dashboard to upstream product suppliers.

    The raw workbook, DB tables, and evidence outputs remain untouched. This
    scope only drives front-end analytics so service, training, conference,
    publishing, and clinic-service rows do not distort upstream competition
    views.
    """

    included_products: list[dict[str, Any]] = []
    excluded_products: list[dict[str, Any]] = []
    product_exclusion_reasons: Counter[str] = Counter()
    for product in products:
        reason = product_exclusion_reason(product)
        if reason:
            excluded_products.append({**product, "_dashboard_exclusion_reason": reason})
            product_exclusion_reasons[reason] += 1
        else:
            included_products.append(product)

    product_count_by_company = Counter(product.get("Company") or "" for product in included_products)
    brand_sets_by_company: dict[str, set[str]] = defaultdict(set)
    track_counter_by_company: dict[str, Counter[str]] = defaultdict(Counter)
    included_company_keys = {compact_key(company) for company in product_count_by_company if company}
    for product in included_products:
        company = product.get("Company") or ""
        if not company:
            continue
        brand = product.get("Brand") or product.get("Core_Product") or ""
        if brand:
            brand_sets_by_company[company].add(brand)
        track = product.get("Category_L1") or product.get("Primary_Segment") or ""
        if track:
            track_counter_by_company[company][track] += 1

    included_companies: list[dict[str, Any]] = []
    excluded_companies: list[dict[str, Any]] = []
    company_exclusion_reasons: Counter[str] = Counter()
    for company in companies:
        name = company.get("Company") or ""
        reason = company_exclusion_reason(company)
        if not reason and compact_key(name) not in included_company_keys:
            reason = "no_dashboard_supply_products"
        if reason:
            excluded_companies.append({**company, "_dashboard_exclusion_reason": reason})
            company_exclusion_reasons[reason] += 1
            continue

        scoped_company = dict(company)
        scoped_company["Product_Count"] = str(product_count_by_company.get(name, 0))
        scoped_company["Brand_Count"] = str(len(brand_sets_by_company.get(name, set())))
        if track_counter_by_company.get(name):
            scoped_company["Primary_Track"] = track_counter_by_company[name].most_common(1)[0][0]
        included_companies.append(scoped_company)

    included_brand_company_keys = {compact_key(company.get("Company")) for company in included_companies}
    included_brands = [
        brand
        for brand in brands
        if compact_key(brand.get("Company")) in included_brand_company_keys
        and not product_exclusion_reason(
            {
                "Company": brand.get("Company"),
                "Category_L1": brand.get("Category_L1"),
                "Category_L2": brand.get("Category_L2"),
                "Tech_Type_Std": brand.get("Tech_Type_Std"),
                "Brand": brand.get("Brand"),
            }
        )
    ]

    scope_report = {
        "mode": "upstream_supply_only",
        "raw_products": len(products),
        "products": len(included_products),
        "excluded_products": len(excluded_products),
        "raw_companies": len(companies),
        "companies": len(included_companies),
        "excluded_companies": len(excluded_companies),
        "raw_brands": len(brands),
        "brands": len(included_brands),
        "excluded_product_reasons": dict(sorted(product_exclusion_reasons.items())),
        "excluded_company_reasons": dict(sorted(company_exclusion_reasons.items())),
        "excluded_product_preview": [
            {
                "record_id": item.get("Record_ID"),
                "company": item.get("Company"),
                "brand": item.get("Brand"),
                "category_l1": item.get("Category_L1"),
                "category_l2": item.get("Category_L2"),
                "reason": item.get("_dashboard_exclusion_reason"),
            }
            for item in excluded_products
            if item.get("_dashboard_exclusion_reason") != "inclusion_status_deleted"
        ][:20],
        "excluded_company_preview": [
            {
                "company": item.get("Company"),
                "primary_track": item.get("Primary_Track"),
                "business_role": item.get("Business_Role"),
                "reason": item.get("_dashboard_exclusion_reason"),
            }
            for item in excluded_companies
            if item.get("_dashboard_exclusion_reason") != "no_dashboard_supply_products"
        ][:20],
    }
    return included_products, included_companies, included_brands, scope_report


def load_market_metrics() -> list[dict[str, Any]]:
    metrics: list[dict[str, Any]] = []
    csv_path = SOURCE_DIR / "医美行业数据" / "yanmei_macro_stats.csv"
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                metrics.append(
                    {
                        "source_file": str(csv_path.relative_to(SOURCE_DIR)),
                        "data_type": norm(row.get("metric")),
                        "category_l1": norm(row.get("segment")),
                        "category_l2": "",
                        "category_l3": "",
                        "geo": norm(row.get("country")),
                        "value": safe_float(row.get("value")),
                        "unit": norm(row.get("unit")),
                        "year": detect_year(row.get("year")),
                        "source_org": norm(row.get("source")),
                        "report_title": "",
                        "url": "",
                        "note": norm(row.get("note")),
                        "confidence": "",
                    }
                )
    xlsx_path = SOURCE_DIR / "医美行业数据" / "医美行业数据.xlsx"
    if xlsx_path.exists():
        rows = read_sheet_dicts(xlsx_path, "Sheet1")
        for row in rows:
            if not norm(row.get("数据类型")):
                continue
            metrics.append(
                {
                    "source_file": str(xlsx_path.relative_to(SOURCE_DIR)),
                    "data_type": norm(row.get("数据类型")),
                    "category_l1": norm(row.get("一级分类")),
                    "category_l2": norm(row.get("二级分类")),
                    "category_l3": norm(row.get("三级分类")),
                    "geo": norm(row.get("地理范围")),
                    "value": safe_float(row.get("数值")),
                    "unit": norm(row.get("单位")),
                    "year": detect_year(row.get("年份"), row.get("发布时间")),
                    "source_org": norm(row.get("数据来源机构")),
                    "report_title": norm(row.get("报告标题")),
                    "url": norm(row.get("原始链接")),
                    "note": norm(row.get("备注")),
                        "confidence": norm(row.get("可信度")),
                    }
                )
    if ISAPS_MARKET_METRICS_PATH.exists():
        with ISAPS_MARKET_METRICS_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if not norm(row.get("data_type")):
                    continue
                metrics.append(
                    {
                        "source_file": norm(row.get("source_file")),
                        "data_type": norm(row.get("data_type")),
                        "category_l1": norm(row.get("category_l1")),
                        "category_l2": norm(row.get("category_l2")),
                        "category_l3": norm(row.get("category_l3")),
                        "geo": norm(row.get("geo")),
                        "value": safe_float(row.get("value")),
                        "unit": norm(row.get("unit")),
                        "year": detect_year(row.get("year")),
                        "source_org": norm(row.get("source_org")),
                        "report_title": norm(row.get("report_title")),
                        "url": norm(row.get("url")),
                        "note": norm(row.get("note")),
                        "confidence": norm(row.get("confidence")),
                    }
                )
    for item in metrics:
        item["segments"] = ",".join(detect_segments(text_blob(item), item.get("category_l1", "")))
    return metrics


def load_conferences(conference_book: Path, congress_book: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    sponsor_sheets = ["Sheet3", "IMCAS", "AMWC"]
    for sheet in sponsor_sheets:
        for row in read_sheet_dicts(conference_book, sheet):
            brand = norm(row.get("品牌名称 (Brand)") or row.get("品牌名称"))
            if not brand:
                continue
            category = norm(row.get("美业分支属性 (Category)"))
            products = norm(row.get("主要产品/技术 (Main Products)"))
            country = norm(row.get("国家 (Country)"))
            record = {
                "kind": "brand_presence",
                "event_name": sheet if sheet != "Sheet3" else "Conference sponsor roster",
                "brand": brand,
                "company": brand,
                "sponsor_level": norm(row.get("赞助级别 (Sponsorship Level)")),
                "country": country,
                "region": "",
                "city": "",
                "category": category,
                "products": products,
                "website": "",
                "date_text": "",
                "year": None,
                "source_file": str(conference_book.relative_to(SOURCE_DIR)),
                "source_sheet": sheet,
                "notes": norm(row.get("总部所在地 (Headquarters)")),
            }
            record["segments"] = ",".join(detect_segments(text_blob(record), category))
            records.append(record)

    event_sheets = ["Sheet2"]
    for sheet in event_sheets:
        for row in read_sheet_dicts(conference_book, sheet):
            event_name = norm(row.get("中文名称"))
            if not event_name:
                continue
            record = {
                "kind": "event",
                "event_name": event_name,
                "brand": "",
                "company": norm(row.get("主办单位")),
                "sponsor_level": "",
                "country": "China",
                "region": "Asia-Pacific",
                "city": norm(row.get("城市")),
                "category": norm(row.get("会议性质")),
                "products": norm(row.get("会议主题")),
                "website": "",
                "date_text": f"{norm(row.get('最近'))}-{norm(row.get('月'))}-{norm(row.get('日'))}",
                "year": detect_year(row.get("最近")),
                "source_file": str(conference_book.relative_to(SOURCE_DIR)),
                "source_sheet": sheet,
                "notes": norm(row.get("举办地点")),
            }
            record["segments"] = ",".join(detect_segments(text_blob(record), record["category"]))
            records.append(record)

    wb = openpyxl.load_workbook(congress_book, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm(v) or f"Column_{i + 1}" for i, v in enumerate(rows[0])]
        for row_values in rows[1:]:
            row = {headers[i]: row_values[i] if i < len(row_values) else None for i in range(len(headers))}
            event_name = norm(row.get("Conference name") or row.get("大会名称") or row.get("展会名"))
            if not event_name:
                continue
            country = norm(row.get("Country") or row.get("国家"))
            record = {
                "kind": "event",
                "event_name": event_name,
                "brand": "",
                "company": norm(row.get("主办单位") or row.get("Chairman")),
                "sponsor_level": "",
                "country": country,
                "region": norm(row.get("Area") or row.get("区域")),
                "city": norm(row.get("城市")),
                "category": norm(row.get("Type") or row.get("性质")),
                "products": norm(row.get("Subject") or row.get("展会介绍")),
                "website": norm(row.get("Website") or row.get("网址")),
                "date_text": norm(row.get("Time") or row.get("时间") or row.get("会议时间")),
                "year": detect_year(row.get("Time"), row.get("时间"), row.get("会议时间"), row.get("年")),
                "source_file": str(congress_book.relative_to(SOURCE_DIR)),
                "source_sheet": ws.title,
                "notes": norm(row.get("Remarks") or row.get("备注") or row.get("状态与备注")),
            }
            record["segments"] = ",".join(detect_segments(text_blob(record), record["category"]))
            records.append(record)
    wb.close()
    return records


def load_reports() -> list[dict[str, Any]]:
    reports = []
    for path in sorted(SOURCE_DIR.rglob("*.md")):
        if PROJECT_DIR in path.parents:
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = path.read_text(encoding="gb18030", errors="replace")
        title = ""
        for line in text.splitlines():
            if line.strip().startswith("#"):
                title = line.strip("# ").strip()
                break
        if not title:
            title = path.stem
        reports.append(
            {
                "title": title,
                "path": str(path.relative_to(SOURCE_DIR)),
                "body": text[:20000],
                "segments": ",".join(detect_segments(text)),
                "year": detect_year(text),
            }
        )
    return reports


def social_status() -> list[dict[str, Any]]:
    yt_config = Path.home() / "AppData" / "Roaming" / "yt-dlp" / "config"
    yt_note = "ready for search/subtitles/comments via yt-dlp"
    if yt_config.exists() and "--js-runtimes" in yt_config.read_text(encoding="utf-8", errors="ignore"):
        yt_note += "; JS runtime configured"
    else:
        yt_note += "; JS runtime config recommended"
    twitter_ready = bool(shutil.which("twitter")) and bool(os.environ.get("TWITTER_AUTH_TOKEN") and os.environ.get("TWITTER_CT0"))
    return [
        {
            "platform": "YouTube",
            "status": "ready" if shutil.which("yt-dlp") else "missing_tool",
            "tool": "yt-dlp",
            "scope": "video search, metadata, subtitles, best-effort comments",
            "note": yt_note if shutil.which("yt-dlp") else "Install yt-dlp before collecting YouTube samples.",
        },
        {
            "platform": "Reddit",
            "status": "ready" if shutil.which("rdt") else "needs_install",
            "tool": "rdt-cli",
            "scope": "post search, subreddit browsing, post + comments",
            "note": "Install rdt-cli; login optional for public search/read.",
        },
        {
            "platform": "Twitter/X",
            "status": "ready" if twitter_ready else ("needs_auth" if shutil.which("twitter") else "needs_install"),
            "tool": "twitter-cli",
            "scope": "tweet/user/article reads; search can be unstable",
            "note": "Requires twitter-cli plus TWITTER_AUTH_TOKEN and TWITTER_CT0 cookies for reliable collection.",
        },
        {
            "platform": "LinkedIn",
            "status": "fallback_only" if shutil.which("mcporter") else "needs_auth",
            "tool": "linkedin scraper / Jina fallback",
            "scope": "company/person/job profile reads when authenticated",
            "note": "Use authenticated LinkedIn scraper when available; otherwise public web reader only.",
        },
    ]


def create_database(
    products: list[dict[str, Any]],
    companies: list[dict[str, Any]],
    brands: list[dict[str, Any]],
    conferences: list[dict[str, Any]],
    metrics: list[dict[str, Any]],
    reports: list[dict[str, Any]],
    socials: list[dict[str, Any]],
    quality: dict[str, Any],
) -> None:
    conn = open_reset_database(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    company_id_map = company_ids(companies)
    company_master = build_company_master(companies)
    company_geo = build_geo_companies(company_master, products)["companies"]
    product_master = build_product_master(products, company_id_map)
    product_hierarchy = build_product_hierarchy(products, company_id_map)
    product_lookup = {product_id_for(product): product for product in products}
    active_product_ids = set(product_lookup)
    write_product_hierarchy_outputs(product_hierarchy)
    write_source_authority_policy_output()
    write_field_dictionary_output()
    registration_seed = build_registration_seed(products, company_id_map)
    market_snapshots = build_market_snapshots(company_master)
    company_financial_metrics = load_generated_csv(COMPANY_FINANCIAL_METRICS_PATH)
    company_revenue_collection_plan = load_audit_csv(COMPANY_REVENUE_COLLECTION_PLAN_PATH)
    verification_queue = build_verification_queue(company_master)
    staging_records = load_staging_records()
    company_background_evidence = load_company_background_evidence()
    company_capital_structure = load_company_capital_structure()
    listed_company_batch = load_listed_company_batch()
    company_official_source_plan = load_company_official_source_plan()
    company_official_source_evidence = load_company_official_source_evidence()
    official_website_master = load_generated_csv(OFFICIAL_WEBSITE_MASTER_PATH)
    company_official_website = load_generated_csv(COMPANY_OFFICIAL_WEBSITE_PATH)
    company_media_asset_index = load_generated_csv(COMPANY_MEDIA_ASSET_INDEX_PATH)
    product_specification_evidence = load_generated_csv(PRODUCT_SPECIFICATION_EVIDENCE_PATH)
    manual_product_fact_evidence = load_generated_csv(MANUAL_PRODUCT_FACT_EVIDENCE_PATH)
    policy_regulatory_source_plan = load_policy_regulatory_source_plan()
    mdr_ce_search_plan = load_mdr_ce_search_plan()
    mdr_ce_evidence_candidates = load_mdr_ce_evidence_candidates()
    news_regulatory_event_candidates = load_news_regulatory_event_candidates()
    briefing_update_candidates = load_briefing_update_candidates()
    briefing_verified_update_events = load_briefing_verified_update_events()
    briefing_fulltext_rescue = load_briefing_fulltext_rescue()
    briefing_product_gap_candidates = load_briefing_product_gap_candidates()
    data_usability_ledger = load_audit_csv(DATA_USABILITY_LEDGER_PATH)
    data_usability_row_status = load_audit_csv(DATA_USABILITY_ROW_STATUS_PATH)
    data_usability_missing_owner = load_audit_csv(DATA_USABILITY_MISSING_OWNER_PATH)
    evidence_promotion = build_evidence_promotions(
        product_master,
        product_hierarchy,
        staging_records,
        mdr_ce_evidence_candidates,
        company_official_source_evidence,
        official_website_master,
        company_official_website,
    )
    product_fact_summary = apply_manual_product_fact_evidence(product_master, manual_product_fact_evidence)
    evidence_promotion["summary"].update(product_fact_summary)
    evidence_promotion["summary"].update(
        apply_product_specification_evidence_to_master(product_master, product_specification_evidence)
    )
    manual_fact_trace_logs = append_manual_product_fact_trace_logs(
        evidence_promotion["log_rows"],
        manual_product_fact_evidence,
        product_master,
        datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
    )
    evidence_promotion["summary"]["manual_product_fact_trace_logs_added"] = manual_fact_trace_logs
    manual_official_indication_rows = load_manual_official_indication_rows()
    manual_nmpa_registration_rows = load_manual_nmpa_registration_rows()
    manual_registration_name_rows = load_manual_registration_name_rows()
    promoted_registration_rows = (
        evidence_promotion["registration_rows"]
        + manual_official_indication_rows
        + manual_nmpa_registration_rows
        + manual_registration_name_rows
    )
    official_indication_analysis = build_official_indication_analysis(
        registration_seed,
        staging_records,
        promoted_registration_rows,
        product_lookup,
    )
    registration_evidence_rows = build_registration_evidence_output(
        registration_seed,
        staging_records,
        promoted_registration_rows,
        active_product_ids,
    )
    evidence_promotion["summary"].update(apply_v4_registration_indication_closure(registration_evidence_rows))
    registration_derived_counts = derived_registration_counts(registration_evidence_rows)
    evidence_promotion["summary"].update(
        apply_registration_product_backfill(product_master, registration_evidence_rows)
    )
    write_product_master_output(product_master)
    write_registration_evidence_output(registration_evidence_rows)
    write_evidence_promotion_log(evidence_promotion["log_rows"])
    write_official_indication_evidence(official_indication_analysis.get("records", []))
    staged_us_by_company = Counter(
        item.get("company_id") for item in staging_records if item.get("source_key") == "fda_openfda_510k"
    )
    background_evidence_by_company = Counter(item.get("company_id") for item in company_background_evidence)
    cur.executescript(
        """
        PRAGMA journal_mode=WAL;

        CREATE TABLE products (
          record_id TEXT PRIMARY KEY,
          company TEXT, country TEXT, region TEXT, location_full TEXT,
          ownership TEXT, business_role TEXT, status TEXT, parent_company TEXT, stock_code TEXT,
          category_l1 TEXT, category_l2 TEXT, tech_type_std TEXT, brand TEXT, brand_type TEXT,
          core_product TEXT, fda_status TEXT, fda_approval_date TEXT, fda_510k_number TEXT,
          ce_status TEXT, ce_year TEXT, nmpa_status TEXT, nmpa_approval_date TEXT,
          nmpa_reg_number TEXT, kfda_status TEXT, global_launch_year TEXT,
          oem_for TEXT, manufactured_by TEXT, feature_tags TEXT, is_primary_record TEXT,
          product_uuid TEXT, duplicate_note TEXT, tech_type_original TEXT, introduction TEXT,
          segments TEXT, primary_segment TEXT, verified_product_type_cn TEXT,
          market_channel TEXT,
          material_taxonomy_l1_cn TEXT, material_taxonomy_l2_cn TEXT,
          material_taxonomy_l3_cn TEXT, material_taxonomy_path_cn TEXT,
          material_taxonomy_source TEXT, material_taxonomy_confidence TEXT,
          material_taxonomy_review_status TEXT, material_taxonomy_note TEXT,
          inclusion_status TEXT, material_family TEXT, backfill_audit TEXT,
          search_blob TEXT
        );

        CREATE TABLE companies (
          company TEXT PRIMARY KEY,
          hq_country TEXT, countries_all TEXT, country_count TEXT, region TEXT,
          location_full TEXT, ownership TEXT, business_role TEXT, status TEXT,
          parent_company TEXT, stock_code TEXT, revenue_usd_m TEXT, revenue_year TEXT,
          revenue_growth_pct TEXT, gross_margin_pct TEXT, market_cap_usd_m TEXT,
          market_cap_date TEXT, ps_ratio TEXT, pe_ratio TEXT, pb_ratio TEXT,
          eps TEXT, eps_ttm TEXT, net_profit_growth_pct TEXT,
          financial_period TEXT, filing_date TEXT, metric_basis TEXT,
          market_refreshed_at TEXT, financial_refreshed_at TEXT,
          aesthetics_revenue_pct TEXT,
          market_price TEXT, market_currency TEXT, market_day_change_pct TEXT,
          market_source_url TEXT, market_captured_at TEXT, financial_source_url TEXT,
          financial_review_status TEXT, primary_track TEXT, fda_products TEXT,
          nmpa_products TEXT, positioning_cn TEXT,
          search_blob TEXT
        );

        CREATE TABLE company_master (
          company_id TEXT PRIMARY KEY,
          canonical_name TEXT, aliases TEXT, hq_country TEXT, countries_all TEXT, region TEXT,
          location_full TEXT, ownership TEXT, business_role TEXT, status TEXT,
          parent_company TEXT, positioning_cn TEXT, ultimate_parent TEXT, acquisition_status TEXT, acquisition_timeline TEXT,
          stock_code TEXT, exchange TEXT, ticker_symbol TEXT, listing_country TEXT, isin TEXT,
          product_count INTEGER, brand_count INTEGER, primary_track TEXT, priority_rank INTEGER,
          verification_status TEXT, review_status TEXT, source_status TEXT, search_queries TEXT,
          search_blob TEXT
        );

        CREATE TABLE company_geo (
          company_id TEXT PRIMARY KEY,
          company TEXT, city TEXT, country TEXT, region TEXT, location_full TEXT,
          lat REAL, lon REAL, precision TEXT,
          products INTEGER, brands INTEGER, primary_track TEXT, stock_code TEXT,
          ownership TEXT, regulatory_channels TEXT, priority_rank TEXT, review_status TEXT
        );

        CREATE TABLE product_master (
          product_id TEXT PRIMARY KEY,
          seed_record_id TEXT, company_id TEXT, company TEXT, brand TEXT, brand_role TEXT,
          standard_product_name TEXT, registered_name TEXT, model_or_sku TEXT,
          commercial_path_l1 TEXT, commercial_path_l2 TEXT,
          material_taxonomy_l1_cn TEXT, material_taxonomy_l2_cn TEXT,
          material_taxonomy_l3_cn TEXT, material_taxonomy_path_cn TEXT,
          material_taxonomy_confidence TEXT, material_taxonomy_review_status TEXT,
          material_family TEXT, inclusion_status TEXT,
          technology_path_l1 TEXT, technology_path_l2 TEXT, material_or_energy_source TEXT,
          core_product TEXT, legal_manufacturer TEXT, marketing_holder TEXT, local_holder TEXT,
          oem_for TEXT, manufactured_by TEXT, r_and_d_origin_status TEXT,
          claim_text TEXT, verified_differentiator TEXT, feature_tags TEXT,
          technical_specs_json TEXT, technical_specs_summary TEXT,
          spec_evidence_count TEXT, spec_review_status TEXT,
          classification_layer TEXT, verification_status TEXT, review_status TEXT,
          source_status TEXT, search_blob TEXT
        );

        CREATE TABLE product_family_master (
          product_family_id TEXT PRIMARY KEY,
          company_id TEXT, company TEXT, brand TEXT, brand_type TEXT,
          product_family TEXT, category_l1 TEXT, category_l2 TEXT,
          material_taxonomy_l1_cn TEXT, material_taxonomy_l2_cn TEXT,
          material_taxonomy_l3_cn TEXT, material_taxonomy_path_cn TEXT,
          material_taxonomy_confidence TEXT, material_taxonomy_review_status TEXT,
          material_family TEXT, inclusion_status TEXT,
          tech_type TEXT, material_or_energy_source TEXT, primary_record_count INTEGER,
          duplicate_record_count INTEGER, sku_candidate_count INTEGER,
          countries TEXT, source_record_ids TEXT, duplicate_record_ids TEXT,
          sku_candidate_names TEXT, regulatory_channels TEXT,
          hierarchy_status TEXT, review_status TEXT, source_status TEXT, search_blob TEXT
        );

        CREATE TABLE product_sku_master (
          sku_id TEXT PRIMARY KEY,
          product_family_id TEXT, company_id TEXT, company TEXT, brand TEXT,
          product_family TEXT, model_or_sku TEXT, sku_candidate_name TEXT,
          seed_record_id TEXT, is_primary_record TEXT, duplicate_of_record_id TEXT,
          category_l1 TEXT, category_l2 TEXT,
          material_taxonomy_l1_cn TEXT, material_taxonomy_l2_cn TEXT,
          material_taxonomy_l3_cn TEXT, material_taxonomy_path_cn TEXT,
          material_taxonomy_confidence TEXT, material_taxonomy_review_status TEXT,
          material_family TEXT, inclusion_status TEXT,
          tech_type TEXT, country TEXT,
          regulatory_channels TEXT, split_status TEXT, review_status TEXT,
          source_status TEXT, search_blob TEXT
        );

        CREATE TABLE registration_evidence (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          product_id TEXT, seed_record_id TEXT, company_id TEXT, company TEXT, brand TEXT,
          jurisdiction TEXT, regulator TEXT, regulatory_pathway TEXT,
          fda_product_code TEXT, fda_regulation_number TEXT, fda_device_class TEXT,
          fda_submission_type TEXT, status TEXT,
          registration_no TEXT, approval_date TEXT, original_approval_date TEXT, expiry_date TEXT, registered_name TEXT,
          approved_indication TEXT, intended_use TEXT, legal_manufacturer TEXT, local_holder TEXT,
          source_key TEXT, source_url TEXT, source_type TEXT, evidence_title TEXT,
          evidence_excerpt TEXT, official_description_exact TEXT, official_description_source_field TEXT,
          field_note TEXT, checked_at TEXT, reviewed_by TEXT, review_status TEXT,
          confidence TEXT
        );

        CREATE TABLE market_snapshot (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          company_id TEXT, company TEXT, stock_code TEXT, exchange TEXT, ticker_symbol TEXT,
          listing_country TEXT, as_of TEXT, price TEXT, currency TEXT, market_cap_usd_m TEXT,
          pe_ratio TEXT, pb_ratio TEXT, eps_ttm TEXT, net_profit_growth_yoy_pct TEXT,
          revenue_usd_m TEXT, revenue_year TEXT, gross_margin_pct TEXT,
          ps_ratio TEXT, financial_period TEXT, filing_date TEXT, metric_basis TEXT,
          market_refreshed_at TEXT, financial_refreshed_at TEXT,
          financial_source_url TEXT, financial_review_status TEXT,
          day_change_pct TEXT, source TEXT, source_url TEXT,
          snapshot_status TEXT, note TEXT
        );

        CREATE TABLE company_financial_metrics (
          company_id TEXT, company TEXT, stock_code TEXT, ticker_symbol TEXT,
          sec_cik TEXT, sec_entity_name TEXT, fiscal_year TEXT,
          revenue_usd_m TEXT, gross_profit_usd_m TEXT, gross_margin_pct TEXT,
          net_income_usd_m TEXT, net_income_growth_yoy_pct TEXT,
          stockholders_equity_usd_m TEXT, market_cap_usd_m TEXT, ps_ratio TEXT,
          pe_ratio TEXT, pb_ratio TEXT,
          eps_basic TEXT, eps_diluted TEXT, eps_ttm TEXT,
          revenue_concept TEXT, gross_profit_concept TEXT, net_income_concept TEXT, equity_concept TEXT, eps_concept TEXT,
          revenue_filed TEXT, filing_date TEXT, financial_period TEXT, metric_basis TEXT, source_url TEXT,
          captured_at TEXT, review_status TEXT, note TEXT
        );

        CREATE TABLE company_revenue_collection_plan (
          company_id TEXT, company TEXT, listed_entity_name TEXT, stock_code TEXT,
          exchange TEXT, ticker_symbol TEXT, listing_country TEXT, sec_cik TEXT,
          current_revenue_usd_m TEXT, current_revenue_year TEXT,
          current_gross_margin_pct TEXT, current_source_url TEXT,
          operational_status TEXT, responsible_module TEXT, plan_id TEXT PRIMARY KEY,
          query TEXT, expected_source TEXT, next_action TEXT, generated_at TEXT
        );

        CREATE TABLE company_background_evidence (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          company_id TEXT, company TEXT, priority_rank INTEGER,
          fact_type TEXT, field_name TEXT, field_value TEXT,
          source_key TEXT, source_name TEXT, source_url TEXT,
          captured_at TEXT, confidence TEXT, review_status TEXT,
          raw_json TEXT
        );

        CREATE TABLE company_capital_structure (
          company_id TEXT PRIMARY KEY,
          priority_rank INTEGER, company TEXT, ownership_seed TEXT,
          stock_code_seed TEXT, exchange_seed TEXT, ticker_symbol_seed TEXT,
          listing_country_seed TEXT, sec_cik TEXT, sec_entity_name TEXT,
          sec_tickers TEXT, sec_exchanges TEXT, sec_former_names TEXT,
          evidence_status TEXT, source_url TEXT, captured_at TEXT,
          review_status TEXT, notes TEXT
        );

        CREATE TABLE listed_company_batch (
          batch_id TEXT PRIMARY KEY,
          company_id TEXT, priority_rank INTEGER, company TEXT,
          listing_group_key TEXT, listed_entity_name TEXT,
          relation_to_listed_entity TEXT, related_companies TEXT,
          related_company_ids TEXT, related_product_count INTEGER,
          stock_code TEXT, exchange TEXT, ticker_symbol TEXT,
          listing_country TEXT, ownership_seed TEXT, parent_company_seed TEXT,
          ultimate_parent_seed TEXT, product_count INTEGER, brand_count INTEGER,
          primary_track TEXT, sec_cik TEXT, sec_entity_name TEXT,
          sec_exchange_current TEXT, listing_verification_status TEXT,
          official_source_key TEXT, official_source_url TEXT,
          market_snapshot_status TEXT, market_price TEXT, market_currency TEXT,
          market_day_change_pct TEXT, market_source_url TEXT,
          market_captured_at TEXT, review_status TEXT, notes TEXT
        );

        CREATE TABLE company_official_source_plan (
          plan_id TEXT PRIMARY KEY,
          company_id TEXT, priority_rank INTEGER, company TEXT,
          product_family_id TEXT, brand TEXT, product_family TEXT,
          category_l1 TEXT, category_l2 TEXT, tech_type TEXT,
          query_type TEXT, query TEXT, expected_source TEXT,
          target_fact_group TEXT, priority INTEGER, status TEXT,
          created_at TEXT, notes TEXT
        );

        CREATE TABLE company_official_source_evidence (
          evidence_id TEXT PRIMARY KEY,
          plan_id TEXT, company_id TEXT, company TEXT,
          product_family_id TEXT, brand TEXT, product_family TEXT,
          category_l1 TEXT, category_l2 TEXT, tech_type TEXT,
          query_type TEXT, query TEXT, expected_source TEXT,
          title TEXT, url TEXT, published TEXT, captured_at TEXT,
          source_key TEXT, source_lane TEXT, confidence TEXT,
          official_candidate TEXT, evidence_excerpt TEXT,
          raw_text TEXT, crosscheck_status TEXT
        );

        CREATE TABLE official_website_master (
          website_id TEXT PRIMARY KEY,
          entity_scope TEXT, company_id TEXT, company TEXT,
          listed_parent_company TEXT, related_company_id TEXT, related_company TEXT,
          brand TEXT, product_family_id TEXT, product_family TEXT,
          product_id TEXT, standard_product_name TEXT,
          official_website_url TEXT, official_domain TEXT,
          source_evidence_id TEXT, source_url TEXT, source_title TEXT,
          source_query_type TEXT, confidence TEXT, official_candidate TEXT,
          asset_folder TEXT, captured_at TEXT, review_status TEXT,
          relationship_notes TEXT
        );

        CREATE TABLE company_official_website (
          company_id TEXT PRIMARY KEY,
          company TEXT, listed_parent_url TEXT, listed_parent_domain TEXT,
          operating_company_url TEXT, operating_company_domain TEXT,
          brand_website_urls TEXT, product_line_page_count INTEGER,
          product_line_page_urls TEXT, primary_official_url TEXT,
          primary_official_domain TEXT, source_evidence_id TEXT,
          source_url TEXT, source_title TEXT, confidence TEXT,
          official_candidate TEXT, asset_folder TEXT, captured_at TEXT,
          review_status TEXT, notes TEXT
        );

        CREATE TABLE company_media_asset_index (
          asset_id TEXT PRIMARY KEY,
          entity_scope TEXT, website_id TEXT, company_id TEXT, company TEXT,
          brand TEXT, product_family_id TEXT, product_family TEXT,
          asset_type TEXT, asset_role TEXT, source_page_url TEXT,
          image_url TEXT, local_path TEXT, file_name TEXT, mime_type TEXT,
          file_bytes TEXT, captured_at TEXT, confidence TEXT,
          review_status TEXT, notes TEXT
        );

        CREATE TABLE product_specification_evidence (
          spec_id TEXT PRIMARY KEY,
          company_id TEXT, company TEXT, brand TEXT,
          product_family_id TEXT, product_family TEXT,
          product_id TEXT, standard_product_name TEXT,
          source_page_url TEXT, source_title TEXT,
          source_evidence_id TEXT, source_query_type TEXT,
          spec_name TEXT, spec_value TEXT, spec_unit TEXT,
          spec_category TEXT, evidence_excerpt TEXT,
          captured_at TEXT, confidence TEXT, review_status TEXT, notes TEXT
        );

        CREATE TABLE manual_product_fact_evidence (
          fact_id TEXT PRIMARY KEY,
          product_id TEXT, seed_record_id TEXT, company_id TEXT,
          company TEXT, brand TEXT, product_family_id TEXT,
          standard_product_name TEXT, priority TEXT,
          fact_group TEXT, field_name TEXT, field_value TEXT,
          source_url TEXT, evidence_title TEXT, evidence_excerpt TEXT,
          source_type TEXT, confidence TEXT, captured_at TEXT,
          promoted_at TEXT, review_status TEXT, note TEXT
        );

        CREATE TABLE policy_regulatory_source_plan (
          plan_id TEXT PRIMARY KEY,
          source_key TEXT, channel_code TEXT, jurisdiction TEXT,
          regulator TEXT, source_name TEXT, source_url TEXT,
          fact_group TEXT, priority INTEGER, status TEXT,
          query_template TEXT, notes TEXT
        );

        CREATE TABLE mdr_ce_search_plan (
          plan_id TEXT PRIMARY KEY,
          priority_rank INTEGER, company_id TEXT, company TEXT,
          product_family_id TEXT, brand TEXT, product_family TEXT,
          category_l1 TEXT, category_l2 TEXT, tech_type TEXT,
          evidence_target TEXT, source_key TEXT, source_name TEXT,
          source_url TEXT, query TEXT, expected_evidence TEXT,
          review_status TEXT, automation_status TEXT, created_at TEXT,
          notes TEXT
        );

        CREATE TABLE mdr_ce_evidence_candidates (
          evidence_id TEXT PRIMARY KEY,
          plan_id TEXT, priority_rank INTEGER, company_id TEXT, company TEXT,
          product_family_id TEXT, brand TEXT, product_family TEXT,
          source_key TEXT, source_name TEXT, title TEXT, url TEXT,
          published TEXT, captured_at TEXT, confidence TEXT,
          official_candidate TEXT, evidence_excerpt TEXT, raw_text TEXT,
          crosscheck_status TEXT
        );

        CREATE TABLE news_regulatory_event_candidates (
          candidate_id TEXT PRIMARY KEY,
          article_date TEXT, captured_at TEXT, briefing_file TEXT,
          article_title TEXT, article_title_zh TEXT, article_url TEXT, article_source TEXT,
          company_id TEXT, product_id TEXT, company TEXT, brand TEXT, product_name TEXT,
          jurisdiction TEXT, regulator TEXT, event_type TEXT,
          candidate_indication TEXT, candidate_approval_date TEXT, candidate_excerpt TEXT,
          matched_alias TEXT, confidence TEXT, status TEXT, needs_official_verification TEXT,
          official_query TEXT
        );

        CREATE TABLE briefing_update_candidates (
          candidate_id TEXT PRIMARY KEY,
          event_group TEXT, event_type TEXT, article_date TEXT,
          company_id TEXT, product_id TEXT, company TEXT, brand TEXT, product_name TEXT,
          market_or_jurisdiction TEXT, source_domain TEXT, article_url TEXT, briefing_file TEXT,
          body_quality TEXT, excerpt TEXT, confidence_score TEXT, status TEXT,
          needs_fulltext_rescue TEXT, needs_official_verification TEXT,
          official_query TEXT, promotion_target TEXT
        );

        CREATE TABLE briefing_verified_update_events (
          event_id TEXT PRIMARY KEY,
          candidate_id TEXT, event_group TEXT, event_type TEXT, article_date TEXT,
          company_id TEXT, product_id TEXT, company TEXT, brand TEXT, product_name TEXT,
          mapping_status TEXT, verification_status TEXT, promotion_status TEXT,
          official_source_type TEXT, official_source_url TEXT, official_title TEXT,
          official_excerpt TEXT, promoted_target TEXT, promotion_id TEXT,
          remaining_gap TEXT, checked_at TEXT
        );

        CREATE TABLE briefing_fulltext_rescue (
          candidate_id TEXT PRIMARY KEY,
          article_url TEXT, final_url TEXT, fetched_at TEXT,
          fetch_status TEXT, status_code TEXT, char_count TEXT,
          title TEXT, body_excerpt TEXT, error TEXT
        );

        CREATE TABLE briefing_product_gap_candidates (
          company TEXT, candidate_product_or_family TEXT,
          source_count TEXT, source_types TEXT, source_domains TEXT,
          confidence_mix TEXT, sample_url TEXT, review_status TEXT,
          PRIMARY KEY (company, candidate_product_or_family)
        );

        CREATE TABLE evidence_promotion_log (
          promotion_id TEXT PRIMARY KEY,
          product_id TEXT, seed_record_id TEXT, company_id TEXT,
          company TEXT, brand TEXT, product_family_id TEXT,
          source_key TEXT, source_type TEXT, field_name TEXT,
          promoted_value TEXT, source_url TEXT, evidence_title TEXT,
          confidence TEXT, promoted_at TEXT, note TEXT
        );

        CREATE TABLE official_indication_evidence (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          product_id TEXT, seed_record_id TEXT, company TEXT, brand TEXT,
          product TEXT, country TEXT, regulator TEXT, pathway TEXT,
          registration_no TEXT, approval_date TEXT, year INTEGER,
          indication TEXT, official_description_exact TEXT, official_description_source_field TEXT,
          field_note TEXT, analysis_bucket_note TEXT, buckets TEXT, source_url TEXT, source_type TEXT,
          confidence TEXT, source_label TEXT
        );

        CREATE TABLE official_source_registry (
          source_key TEXT PRIMARY KEY,
          channel_code TEXT, source_kind TEXT, scope_status TEXT,
          jurisdiction TEXT, regulator TEXT, source_name TEXT, source_url TEXT,
          access_method TEXT, machine_readable TEXT, primary_use TEXT,
          automation_status TEXT, priority INTEGER, note TEXT
        );

        CREATE TABLE source_authority_policy (
          fact_group TEXT PRIMARY KEY,
          authoritative_source TEXT, primary_sources TEXT,
          supporting_sources TEXT, merge_rule TEXT, manual_role TEXT
        );

        CREATE TABLE field_dictionary (
          table_name TEXT, field_name TEXT, definition TEXT,
          display_note TEXT, source_priority TEXT,
          PRIMARY KEY (table_name, field_name)
        );

        CREATE TABLE verification_queue (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          priority_rank INTEGER, company_id TEXT, company TEXT, fact_group TEXT,
          target_label TEXT, source_lane TEXT, query TEXT, expected_source TEXT,
          status TEXT, created_at TEXT, evidence_count INTEGER, reviewer_note TEXT
        );

        CREATE TABLE evidence_staging (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          source_key TEXT, source_lane TEXT, company_id TEXT, product_id TEXT,
          company TEXT, brand TEXT, jurisdiction TEXT, evidence_type TEXT,
          title TEXT, url TEXT, source_record_id TEXT, captured_at TEXT,
          field_candidates TEXT, excerpt TEXT, raw_json TEXT, review_status TEXT,
          confidence TEXT, merge_target TEXT, merge_status TEXT
        );

        CREATE TABLE seed_integrity_issues (
          issue_id TEXT PRIMARY KEY,
          source_table TEXT, source_row_id TEXT, entity_type TEXT, entity_name TEXT,
          issue_type TEXT, severity TEXT, field_name TEXT, current_value TEXT,
          suggested_value TEXT, description TEXT, action TEXT, status TEXT
        );

        CREATE TABLE data_usability_ledger (
          source_table TEXT PRIMARY KEY,
          total_rows INTEGER, usable_or_reference_rows INTEGER,
          planned_or_review_rows INTEGER, excluded_or_noise_rows INTEGER,
          covered_rows INTEGER, missing_owner_rows INTEGER,
          missing_status_rows INTEGER, dominant_class TEXT,
          top_operational_status TEXT, top_responsible_module TEXT,
          top_planned_status TEXT, top_planned_responsible_module TEXT
        );

        CREATE TABLE data_usability_row_status (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          source_table TEXT, row_id TEXT, entity TEXT,
          lifecycle_layer TEXT, coverage_class TEXT,
          operational_status TEXT, responsible_module TEXT,
          next_action TEXT, evidence_basis TEXT
        );

        CREATE TABLE data_usability_missing_owner (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          source_table TEXT, row_id TEXT, entity TEXT,
          lifecycle_layer TEXT, coverage_class TEXT,
          operational_status TEXT, responsible_module TEXT,
          next_action TEXT, evidence_basis TEXT
        );

        CREATE TABLE brands (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          company TEXT, brand TEXT, country TEXT, category_l1 TEXT, category_l2 TEXT,
          tech_type TEXT, brand_type TEXT, product_count INTEGER, products TEXT,
          segments TEXT, search_blob TEXT
        );

        CREATE TABLE conferences (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          kind TEXT, event_name TEXT, brand TEXT, company TEXT, sponsor_level TEXT,
          country TEXT, region TEXT, city TEXT, category TEXT, products TEXT,
          website TEXT, date_text TEXT, year INTEGER, source_file TEXT,
          source_sheet TEXT, notes TEXT, segments TEXT, search_blob TEXT
        );

        CREATE TABLE market_metrics (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          source_file TEXT, data_type TEXT, category_l1 TEXT, category_l2 TEXT,
          category_l3 TEXT, geo TEXT, value REAL, unit TEXT, year INTEGER,
          source_org TEXT, report_title TEXT, url TEXT, note TEXT, confidence TEXT,
          segments TEXT, search_blob TEXT
        );

        CREATE TABLE reports (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          title TEXT, path TEXT, body TEXT, segments TEXT, year INTEGER, search_blob TEXT
        );

        CREATE TABLE social_sources (
          platform TEXT PRIMARY KEY,
          status TEXT, tool TEXT, scope TEXT, note TEXT
        );

        CREATE TABLE evidence (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          content_type TEXT, title TEXT, subtitle TEXT, body TEXT, url TEXT,
          source_file TEXT, source_sheet TEXT, confidence TEXT, year INTEGER,
          segment TEXT, region TEXT, country TEXT, company TEXT, brand TEXT
        );
        """
    )

    for item in company_master:
        fields = [
            "company_id",
            "canonical_name",
            "aliases",
            "hq_country",
            "countries_all",
            "region",
            "location_full",
            "ownership",
            "business_role",
            "status",
            "parent_company",
            "positioning_cn",
            "ultimate_parent",
            "acquisition_status",
            "acquisition_timeline",
            "stock_code",
            "exchange",
            "ticker_symbol",
            "listing_country",
            "isin",
            "product_count",
            "brand_count",
            "primary_track",
            "priority_rank",
            "verification_status",
            "review_status",
            "source_status",
            "search_queries",
            "search_blob",
        ]
        cur.execute(f"INSERT OR REPLACE INTO company_master VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    for item in company_geo:
        fields = [
            "company_id",
            "company",
            "city",
            "country",
            "region",
            "location_full",
            "lat",
            "lon",
            "precision",
            "products",
            "brands",
            "primary_track",
            "stock_code",
            "ownership",
            "regulatory_channels",
            "priority_rank",
            "review_status",
        ]
        cur.execute(f"INSERT OR REPLACE INTO company_geo VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    for item in product_master:
        fields = [
            "product_id",
            "seed_record_id",
            "company_id",
            "company",
            "brand",
            "brand_role",
            "standard_product_name",
            "registered_name",
            "model_or_sku",
            "commercial_path_l1",
            "commercial_path_l2",
            "material_taxonomy_l1_cn",
            "material_taxonomy_l2_cn",
            "material_taxonomy_l3_cn",
            "material_taxonomy_path_cn",
            "material_taxonomy_confidence",
            "material_taxonomy_review_status",
            "material_family",
            "inclusion_status",
            "technology_path_l1",
            "technology_path_l2",
            "material_or_energy_source",
            "core_product",
            "legal_manufacturer",
            "marketing_holder",
            "local_holder",
            "oem_for",
            "manufactured_by",
            "r_and_d_origin_status",
            "claim_text",
            "verified_differentiator",
            "feature_tags",
            "technical_specs_json",
            "technical_specs_summary",
            "spec_evidence_count",
            "spec_review_status",
            "classification_layer",
            "verification_status",
            "review_status",
            "source_status",
            "search_blob",
        ]
        cur.execute(f"INSERT INTO product_master VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    for item in product_hierarchy["families"]:
        fields = [
            "product_family_id",
            "company_id",
            "company",
            "brand",
            "brand_type",
            "product_family",
            "category_l1",
            "category_l2",
            "material_taxonomy_l1_cn",
            "material_taxonomy_l2_cn",
            "material_taxonomy_l3_cn",
            "material_taxonomy_path_cn",
            "material_taxonomy_confidence",
            "material_taxonomy_review_status",
            "material_family",
            "inclusion_status",
            "tech_type",
            "material_or_energy_source",
            "primary_record_count",
            "duplicate_record_count",
            "sku_candidate_count",
            "countries",
            "source_record_ids",
            "duplicate_record_ids",
            "sku_candidate_names",
            "regulatory_channels",
            "hierarchy_status",
            "review_status",
            "source_status",
            "search_blob",
        ]
        cur.execute(f"INSERT OR REPLACE INTO product_family_master VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    for item in product_hierarchy["skus"]:
        fields = [
            "sku_id",
            "product_family_id",
            "company_id",
            "company",
            "brand",
            "product_family",
            "model_or_sku",
            "sku_candidate_name",
            "seed_record_id",
            "is_primary_record",
            "duplicate_of_record_id",
            "category_l1",
            "category_l2",
            "material_taxonomy_l1_cn",
            "material_taxonomy_l2_cn",
            "material_taxonomy_l3_cn",
            "material_taxonomy_path_cn",
            "material_taxonomy_confidence",
            "material_taxonomy_review_status",
            "material_family",
            "inclusion_status",
            "tech_type",
            "country",
            "regulatory_channels",
            "split_status",
            "review_status",
            "source_status",
            "search_blob",
        ]
        cur.execute(f"INSERT OR REPLACE INTO product_sku_master VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    registration_db_rows = build_registration_evidence_output(
        registration_seed,
        staging_records,
        promoted_registration_rows,
        active_product_ids,
    )
    apply_v4_registration_indication_closure(registration_db_rows)
    for item in registration_db_rows:
        item = enrich_registration_description_fields(item)
        fields = REGISTRATION_EVIDENCE_FIELDS
        cur.execute(
            f"INSERT INTO registration_evidence ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in market_snapshots:
        fields = [
            "company_id",
            "company",
            "stock_code",
            "exchange",
            "ticker_symbol",
            "listing_country",
            "as_of",
            "price",
            "currency",
            "market_cap_usd_m",
            "pe_ratio",
            "pb_ratio",
            "eps_ttm",
            "net_profit_growth_yoy_pct",
            "revenue_usd_m",
            "revenue_year",
            "gross_margin_pct",
            "ps_ratio",
            "financial_period",
            "filing_date",
            "metric_basis",
            "market_refreshed_at",
            "financial_refreshed_at",
            "financial_source_url",
            "financial_review_status",
            "day_change_pct",
            "source",
            "source_url",
            "snapshot_status",
            "note",
        ]
        cur.execute(
            f"INSERT INTO market_snapshot ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_financial_metrics:
        fields = [
            "company_id",
            "company",
            "stock_code",
            "ticker_symbol",
            "sec_cik",
            "sec_entity_name",
            "fiscal_year",
            "revenue_usd_m",
            "gross_profit_usd_m",
            "gross_margin_pct",
            "net_income_usd_m",
            "net_income_growth_yoy_pct",
            "stockholders_equity_usd_m",
            "market_cap_usd_m",
            "ps_ratio",
            "pe_ratio",
            "pb_ratio",
            "eps_basic",
            "eps_diluted",
            "eps_ttm",
            "revenue_concept",
            "gross_profit_concept",
            "net_income_concept",
            "equity_concept",
            "eps_concept",
            "revenue_filed",
            "filing_date",
            "financial_period",
            "metric_basis",
            "source_url",
            "captured_at",
            "review_status",
            "note",
        ]
        cur.execute(
            f"INSERT INTO company_financial_metrics ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_revenue_collection_plan:
        fields = [
            "company_id",
            "company",
            "listed_entity_name",
            "stock_code",
            "exchange",
            "ticker_symbol",
            "listing_country",
            "sec_cik",
            "current_revenue_usd_m",
            "current_revenue_year",
            "current_gross_margin_pct",
            "current_source_url",
            "operational_status",
            "responsible_module",
            "plan_id",
            "query",
            "expected_source",
            "next_action",
            "generated_at",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO company_revenue_collection_plan ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_background_evidence:
        fields = [
            "company_id",
            "company",
            "priority_rank",
            "fact_type",
            "field_name",
            "field_value",
            "source_key",
            "source_name",
            "source_url",
            "captured_at",
            "confidence",
            "review_status",
            "raw_json",
        ]
        values = []
        for field in fields:
            value = item.get(field)
            if field == "raw_json" and not isinstance(value, str):
                value = json.dumps(value or {}, ensure_ascii=False)
            values.append(value)
        cur.execute(
            f"INSERT INTO company_background_evidence ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            values,
        )

    for item in company_capital_structure:
        fields = [
            "company_id",
            "priority_rank",
            "company",
            "ownership_seed",
            "stock_code_seed",
            "exchange_seed",
            "ticker_symbol_seed",
            "listing_country_seed",
            "sec_cik",
            "sec_entity_name",
            "sec_tickers",
            "sec_exchanges",
            "sec_former_names",
            "evidence_status",
            "source_url",
            "captured_at",
            "review_status",
            "notes",
        ]
        cur.execute(f"INSERT OR REPLACE INTO company_capital_structure VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    for item in listed_company_batch:
        fields = [
            "batch_id",
            "company_id",
            "priority_rank",
            "company",
            "listing_group_key",
            "listed_entity_name",
            "relation_to_listed_entity",
            "related_companies",
            "related_company_ids",
            "related_product_count",
            "stock_code",
            "exchange",
            "ticker_symbol",
            "listing_country",
            "ownership_seed",
            "parent_company_seed",
            "ultimate_parent_seed",
            "product_count",
            "brand_count",
            "primary_track",
            "sec_cik",
            "sec_entity_name",
            "sec_exchange_current",
            "listing_verification_status",
            "official_source_key",
            "official_source_url",
            "market_snapshot_status",
            "market_price",
            "market_currency",
            "market_day_change_pct",
            "market_source_url",
            "market_captured_at",
            "review_status",
            "notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO listed_company_batch VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_official_source_plan:
        fields = [
            "plan_id",
            "company_id",
            "priority_rank",
            "company",
            "product_family_id",
            "brand",
            "product_family",
            "category_l1",
            "category_l2",
            "tech_type",
            "query_type",
            "query",
            "expected_source",
            "target_fact_group",
            "priority",
            "status",
            "created_at",
            "notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO company_official_source_plan VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_official_source_evidence:
        fields = [
            "evidence_id",
            "plan_id",
            "company_id",
            "company",
            "product_family_id",
            "brand",
            "product_family",
            "category_l1",
            "category_l2",
            "tech_type",
            "query_type",
            "query",
            "expected_source",
            "title",
            "url",
            "published",
            "captured_at",
            "source_key",
            "source_lane",
            "confidence",
            "official_candidate",
            "evidence_excerpt",
            "raw_text",
            "crosscheck_status",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO company_official_source_evidence VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in official_website_master:
        fields = [
            "website_id",
            "entity_scope",
            "company_id",
            "company",
            "listed_parent_company",
            "related_company_id",
            "related_company",
            "brand",
            "product_family_id",
            "product_family",
            "product_id",
            "standard_product_name",
            "official_website_url",
            "official_domain",
            "source_evidence_id",
            "source_url",
            "source_title",
            "source_query_type",
            "confidence",
            "official_candidate",
            "asset_folder",
            "captured_at",
            "review_status",
            "relationship_notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO official_website_master VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_official_website:
        fields = [
            "company_id",
            "company",
            "listed_parent_url",
            "listed_parent_domain",
            "operating_company_url",
            "operating_company_domain",
            "brand_website_urls",
            "product_line_page_count",
            "product_line_page_urls",
            "primary_official_url",
            "primary_official_domain",
            "source_evidence_id",
            "source_url",
            "source_title",
            "confidence",
            "official_candidate",
            "asset_folder",
            "captured_at",
            "review_status",
            "notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO company_official_website VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in company_media_asset_index:
        fields = [
            "asset_id",
            "entity_scope",
            "website_id",
            "company_id",
            "company",
            "brand",
            "product_family_id",
            "product_family",
            "asset_type",
            "asset_role",
            "source_page_url",
            "image_url",
            "local_path",
            "file_name",
            "mime_type",
            "file_bytes",
            "captured_at",
            "confidence",
            "review_status",
            "notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO company_media_asset_index VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in product_specification_evidence:
        fields = [
            "spec_id",
            "company_id",
            "company",
            "brand",
            "product_family_id",
            "product_family",
            "product_id",
            "standard_product_name",
            "source_page_url",
            "source_title",
            "source_evidence_id",
            "source_query_type",
            "spec_name",
            "spec_value",
            "spec_unit",
            "spec_category",
            "evidence_excerpt",
            "captured_at",
            "confidence",
            "review_status",
            "notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO product_specification_evidence VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in manual_product_fact_evidence:
        cur.execute(
            f"INSERT OR REPLACE INTO manual_product_fact_evidence VALUES ({','.join(['?'] * len(PRODUCT_FACT_EVIDENCE_FIELDS))})",
            [item.get(field) for field in PRODUCT_FACT_EVIDENCE_FIELDS],
        )

    for item in policy_regulatory_source_plan:
        fields = [
            "plan_id",
            "source_key",
            "channel_code",
            "jurisdiction",
            "regulator",
            "source_name",
            "source_url",
            "fact_group",
            "priority",
            "status",
            "query_template",
            "notes",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO policy_regulatory_source_plan VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in mdr_ce_search_plan:
        fields = [
            "plan_id",
            "priority_rank",
            "company_id",
            "company",
            "product_family_id",
            "brand",
            "product_family",
            "category_l1",
            "category_l2",
            "tech_type",
            "evidence_target",
            "source_key",
            "source_name",
            "source_url",
            "query",
            "expected_evidence",
            "review_status",
            "automation_status",
            "created_at",
            "notes",
        ]
        cur.execute(f"INSERT OR REPLACE INTO mdr_ce_search_plan VALUES ({','.join(['?'] * len(fields))})", [item.get(field) for field in fields])

    for item in mdr_ce_evidence_candidates:
        fields = [
            "evidence_id",
            "plan_id",
            "priority_rank",
            "company_id",
            "company",
            "product_family_id",
            "brand",
            "product_family",
            "source_key",
            "source_name",
            "title",
            "url",
            "published",
            "captured_at",
            "confidence",
            "official_candidate",
            "evidence_excerpt",
            "raw_text",
            "crosscheck_status",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO mdr_ce_evidence_candidates VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in news_regulatory_event_candidates:
        fields = [
            "candidate_id",
            "article_date",
            "captured_at",
            "briefing_file",
            "article_title",
            "article_title_zh",
            "article_url",
            "article_source",
            "company_id",
            "product_id",
            "company",
            "brand",
            "product_name",
            "jurisdiction",
            "regulator",
            "event_type",
            "candidate_indication",
            "candidate_approval_date",
            "candidate_excerpt",
            "matched_alias",
            "confidence",
            "status",
            "needs_official_verification",
            "official_query",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO news_regulatory_event_candidates VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in briefing_update_candidates:
        fields = [
            "candidate_id",
            "event_group",
            "event_type",
            "article_date",
            "company_id",
            "product_id",
            "company",
            "brand",
            "product_name",
            "market_or_jurisdiction",
            "source_domain",
            "article_url",
            "briefing_file",
            "body_quality",
            "excerpt",
            "confidence_score",
            "status",
            "needs_fulltext_rescue",
            "needs_official_verification",
            "official_query",
            "promotion_target",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO briefing_update_candidates VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in briefing_verified_update_events:
        fields = [
            "event_id",
            "candidate_id",
            "event_group",
            "event_type",
            "article_date",
            "company_id",
            "product_id",
            "company",
            "brand",
            "product_name",
            "mapping_status",
            "verification_status",
            "promotion_status",
            "official_source_type",
            "official_source_url",
            "official_title",
            "official_excerpt",
            "promoted_target",
            "promotion_id",
            "remaining_gap",
            "checked_at",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO briefing_verified_update_events VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in briefing_fulltext_rescue:
        fields = [
            "candidate_id",
            "article_url",
            "final_url",
            "fetched_at",
            "fetch_status",
            "status_code",
            "char_count",
            "title",
            "body_excerpt",
            "error",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO briefing_fulltext_rescue VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in briefing_product_gap_candidates:
        fields = [
            "company",
            "candidate_product_or_family",
            "source_count",
            "source_types",
            "source_domains",
            "confidence_mix",
            "sample_url",
            "review_status",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO briefing_product_gap_candidates VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in merge_manual_evidence_promotion_log(evidence_promotion["log_rows"]):
        fields = [
            "promotion_id",
            "product_id",
            "seed_record_id",
            "company_id",
            "company",
            "brand",
            "product_family_id",
            "source_key",
            "source_type",
            "field_name",
            "promoted_value",
            "source_url",
            "evidence_title",
            "confidence",
            "promoted_at",
            "note",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO evidence_promotion_log VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in official_indication_analysis.get("records", []):
        fields = [
            "product_id",
            "seed_record_id",
            "company",
            "brand",
            "product",
            "country",
            "regulator",
            "pathway",
            "registration_no",
            "approval_date",
            "year",
            "indication",
            "official_description_exact",
            "official_description_source_field",
            "field_note",
            "analysis_bucket_note",
            "buckets",
            "source_url",
            "source_type",
            "confidence",
            "source_label",
        ]
        values = []
        for field in fields:
            value = item.get(field)
            if field == "buckets":
                value = ", ".join(value or [])
            values.append(value)
        cur.execute(
            f"INSERT INTO official_indication_evidence ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            values,
        )

    for item in SOURCE_REGISTRY:
        fields = [
            "source_key",
            "channel_code",
            "source_kind",
            "scope_status",
            "jurisdiction",
            "regulator",
            "source_name",
            "source_url",
            "access_method",
            "machine_readable",
            "primary_use",
            "automation_status",
            "priority",
            "note",
        ]
        values = [source_scope_status(item) if field == "scope_status" else item.get(field) for field in fields]
        cur.execute(f"INSERT INTO official_source_registry VALUES ({','.join(['?'] * len(fields))})", values)

    for item in SOURCE_AUTHORITY_POLICY:
        fields = [
            "fact_group",
            "authoritative_source",
            "primary_sources",
            "supporting_sources",
            "merge_rule",
            "manual_role",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO source_authority_policy VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in FIELD_DICTIONARY_ROWS:
        fields = ["table_name", "field_name", "definition", "display_note", "source_priority"]
        cur.execute(
            f"INSERT OR REPLACE INTO field_dictionary VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in verification_queue:
        if item.get("fact_group") == "registration_us":
            staged_count = staged_us_by_company.get(item.get("company_id"), 0)
            if staged_count:
                item["evidence_count"] = staged_count
                item["status"] = "evidence_staged"
        if item.get("fact_group") == "company_background":
            staged_count = background_evidence_by_company.get(item.get("company_id"), 0)
            if staged_count:
                item["evidence_count"] = staged_count
                item["status"] = "evidence_staged"
        fields = [
            "priority_rank",
            "company_id",
            "company",
            "fact_group",
            "target_label",
            "source_lane",
            "query",
            "expected_source",
            "status",
            "created_at",
            "evidence_count",
            "reviewer_note",
        ]
        cur.execute(
            f"INSERT INTO verification_queue ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for item in staging_records:
        fields = [
            "source_key",
            "source_lane",
            "company_id",
            "product_id",
            "company",
            "brand",
            "jurisdiction",
            "evidence_type",
            "title",
            "url",
            "source_record_id",
            "captured_at",
            "field_candidates",
            "excerpt",
            "raw_json",
            "review_status",
            "confidence",
            "merge_target",
            "merge_status",
        ]
        values = []
        for field in fields:
            value = item.get(field)
            if field in {"field_candidates", "raw_json"} and not isinstance(value, str):
                value = json.dumps(value or {}, ensure_ascii=False)
            values.append(value)
        cur.execute(
            f"INSERT INTO evidence_staging ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            values,
        )
        # Registration-evidence rows are loaded once via registration_db_rows above,
        # where workbook seed rows and promoted official rows are de-duplicated.
        if False and item.get("merge_target") == "registration_evidence":
            candidates = item.get("field_candidates") or {}
            if isinstance(candidates, str):
                try:
                    candidates = json.loads(candidates)
                except json.JSONDecodeError:
                    candidates = {}
            staged_registration_row = enrich_registration_description_fields(
                {
                    "product_id": item.get("product_id"),
                    "seed_record_id": "",
                    "company_id": item.get("company_id"),
                    "company": item.get("company"),
                    "brand": item.get("brand"),
                    "jurisdiction": item.get("jurisdiction"),
                    "regulator": "FDA" if item.get("source_key") == "fda_openfda_510k" else "",
                    "regulatory_pathway": candidates.get("regulatory_pathway"),
                    "fda_product_code": candidates.get("fda_product_code") or candidates.get("product_code"),
                    "fda_regulation_number": candidates.get("fda_regulation_number") or candidates.get("regulation_number"),
                    "fda_device_class": candidates.get("fda_device_class") or candidates.get("device_class"),
                    "fda_submission_type": candidates.get("fda_submission_type") or candidates.get("submission_type"),
                    "status": candidates.get("status"),
                    "registration_no": candidates.get("registration_no"),
                    "approval_date": candidates.get("approval_date"),
                    "original_approval_date": candidates.get("original_approval_date"),
                    "expiry_date": "",
                    "registered_name": candidates.get("registered_name"),
                    "approved_indication": candidates.get("approved_indication"),
                    "intended_use": candidates.get("intended_use"),
                    "legal_manufacturer": candidates.get("legal_manufacturer"),
                    "local_holder": "",
                    "source_key": item.get("source_key"),
                    "source_url": item.get("url"),
                    "source_type": "official_api",
                    "evidence_title": item.get("title"),
                    "evidence_excerpt": item.get("excerpt"),
                    "checked_at": item.get("captured_at"),
                    "reviewed_by": "",
                    "review_status": item.get("review_status") or "needs_review",
                    "confidence": item.get("confidence") or "official_api_unreviewed",
                }
            )
            cur.execute(
                """
                INSERT INTO registration_evidence
                (product_id, seed_record_id, company_id, company, brand, jurisdiction, regulator,
                 regulatory_pathway, status, registration_no, approval_date, expiry_date, registered_name,
                 approved_indication, intended_use, legal_manufacturer, local_holder, source_key, source_url,
                 source_type, evidence_title, evidence_excerpt, official_description_exact,
                 official_description_source_field, field_note, checked_at, reviewed_by, review_status, confidence)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    staged_registration_row.get("product_id"),
                    staged_registration_row.get("seed_record_id"),
                    staged_registration_row.get("company_id"),
                    staged_registration_row.get("company"),
                    staged_registration_row.get("brand"),
                    staged_registration_row.get("jurisdiction"),
                    staged_registration_row.get("regulator"),
                    staged_registration_row.get("regulatory_pathway"),
                    staged_registration_row.get("status"),
                    staged_registration_row.get("registration_no"),
                    staged_registration_row.get("approval_date"),
                    staged_registration_row.get("expiry_date"),
                    staged_registration_row.get("registered_name"),
                    staged_registration_row.get("approved_indication"),
                    staged_registration_row.get("intended_use"),
                    staged_registration_row.get("legal_manufacturer"),
                    staged_registration_row.get("local_holder"),
                    staged_registration_row.get("source_key"),
                    staged_registration_row.get("source_url"),
                    staged_registration_row.get("source_type"),
                    staged_registration_row.get("evidence_title"),
                    staged_registration_row.get("evidence_excerpt"),
                    staged_registration_row.get("official_description_exact"),
                    staged_registration_row.get("official_description_source_field"),
                    staged_registration_row.get("field_note"),
                    staged_registration_row.get("checked_at"),
                    staged_registration_row.get("reviewed_by"),
                    staged_registration_row.get("review_status"),
                    staged_registration_row.get("confidence"),
                ),
            )

    # Promoted rows are already included in registration_db_rows above.
    for item in []:
        item = enrich_registration_description_fields(item)
        fields = REGISTRATION_EVIDENCE_FIELDS
        cur.execute(
            f"INSERT INTO registration_evidence ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field) for field in fields],
        )

    for issue in quality.get("issues", []):
        fields = [
            "issue_id",
            "source_table",
            "source_row_id",
            "entity_type",
            "entity_name",
            "issue_type",
            "severity",
            "field_name",
            "current_value",
            "suggested_value",
            "description",
            "action",
            "status",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO seed_integrity_issues VALUES ({','.join(['?'] * len(fields))})",
            [issue.get(field, "") for field in fields],
        )

    for item in data_usability_ledger:
        fields = [
            "source_table",
            "total_rows",
            "usable_or_reference_rows",
            "planned_or_review_rows",
            "excluded_or_noise_rows",
            "covered_rows",
            "missing_owner_rows",
            "missing_status_rows",
            "dominant_class",
            "top_operational_status",
            "top_responsible_module",
            "top_planned_status",
            "top_planned_responsible_module",
        ]
        cur.execute(
            f"INSERT OR REPLACE INTO data_usability_ledger VALUES ({','.join(['?'] * len(fields))})",
            [item.get(field, "") for field in fields],
        )

    usability_fields = [
        "source_table",
        "row_id",
        "entity",
        "lifecycle_layer",
        "coverage_class",
        "operational_status",
        "responsible_module",
        "next_action",
        "evidence_basis",
    ]
    for item in data_usability_row_status:
        cur.execute(
            f"INSERT INTO data_usability_row_status ({','.join(usability_fields)}) VALUES ({','.join(['?'] * len(usability_fields))})",
            [item.get(field, "") for field in usability_fields],
        )
    for item in data_usability_missing_owner:
        cur.execute(
            f"INSERT INTO data_usability_missing_owner ({','.join(usability_fields)}) VALUES ({','.join(['?'] * len(usability_fields))})",
            [item.get(field, "") for field in usability_fields],
        )

    product_fields = [
        "Record_ID",
        "Company",
        "Country",
        "Region",
        "Location_Full",
        "Ownership",
        "Business_Role",
        "Status",
        "Parent_Company",
        "Stock_Code",
        "Category_L1",
        "Category_L2",
        "Tech_Type_Std",
        "Brand",
        "Brand_Type",
        "Core_Product",
        "FDA_Status",
        "FDA_Approval_Date",
        "FDA_510k_Number",
        "CE_Status",
        "CE_Year",
        "NMPA_Status",
        "NMPA_Approval_Date",
        "NMPA_Reg_Number",
        "KFDA_Status",
        "Global_Launch_Year",
        "OEM_For",
        "Manufactured_By",
        "Feature_Tags",
        "Is_Primary_Record",
        "Product_UUID",
        "Duplicate_Note",
        "Tech_Type_Original",
        "Introduction",
        "Segments",
        "Primary_Segment",
        "Verified_Product_Type_CN",
        "Market_Channel",
        "Material_Taxonomy_L1_CN",
        "Material_Taxonomy_L2_CN",
        "Material_Taxonomy_L3_CN",
        "Material_Taxonomy_Path_CN",
        "Material_Taxonomy_Source",
        "Material_Taxonomy_Confidence",
        "Material_Taxonomy_Review_Status",
        "Material_Taxonomy_Note",
        "Inclusion_Status",
        "Material_Family",
        "Backfill_Audit",
    ]
    for item in products:
        search = text_blob(item)
        values = [item.get(field, "") for field in product_fields] + [search]
        cur.execute(
            f"INSERT INTO products VALUES ({','.join(['?'] * (len(product_fields) + 1))})",
            values,
        )
        cur.execute(
            """
            INSERT INTO evidence
            (content_type, title, subtitle, body, url, source_file, source_sheet, confidence, year,
             segment, region, country, company, brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "product",
                item.get("Brand") or item.get("Core_Product") or item.get("Company"),
                f"{item.get('Company')} · {item.get('Tech_Type_Std')} · {item.get('Country')}",
                search,
                "",
                "全球医美企业库_标准化版v4.xlsx",
                "Product_Lines",
                "unverified_seed",
                detect_year(item.get("Global_Launch_Year"), item.get("FDA_Approval_Date"), item.get("NMPA_Approval_Date")),
                item.get("Segments"),
                item.get("Region"),
                item.get("Country"),
                item.get("Company"),
                item.get("Brand"),
            ),
        )

    for item in companies:
        search = text_blob(item)
        cur.execute(
            """
            INSERT OR REPLACE INTO companies
            (company, hq_country, countries_all, country_count, region, location_full, ownership,
             business_role, status, parent_company, stock_code, revenue_usd_m, revenue_year,
             revenue_growth_pct, gross_margin_pct, market_cap_usd_m, market_cap_date, ps_ratio,
             pe_ratio, pb_ratio, eps, eps_ttm, net_profit_growth_pct,
             financial_period, filing_date, metric_basis, market_refreshed_at, financial_refreshed_at,
             aesthetics_revenue_pct, market_price, market_currency, market_day_change_pct,
             market_source_url, market_captured_at, financial_source_url, financial_review_status,
             primary_track, fda_products, nmpa_products, positioning_cn, search_blob)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.get("Company"),
                item.get("HQ_Country"),
                item.get("Countries_All"),
                item.get("Country_Count"),
                item.get("Region"),
                item.get("Location_Full"),
                item.get("Ownership"),
                item.get("Business_Role"),
                item.get("Status"),
                item.get("Parent_Company"),
                item.get("Stock_Code"),
                item.get("Revenue_USD_M"),
                item.get("Revenue_Year"),
                item.get("Revenue_Growth_Pct"),
                item.get("Gross_Margin_Pct"),
                item.get("Market_Cap_USD_M"),
                item.get("Market_Cap_Date"),
                item.get("PS_Ratio"),
                item.get("PE_Ratio"),
                item.get("PB_Ratio"),
                item.get("EPS"),
                item.get("EPS_TTM"),
                item.get("Net_Profit_Growth_Pct"),
                item.get("Financial_Period"),
                item.get("Filing_Date"),
                item.get("Metric_Basis"),
                item.get("Market_Refreshed_At"),
                item.get("Financial_Refreshed_At"),
                item.get("Aesthetics_Revenue_Pct"),
                item.get("Market_Price"),
                item.get("Market_Currency"),
                item.get("Market_Day_Change_Pct"),
                item.get("Market_Source_URL"),
                item.get("Market_Captured_At"),
                item.get("Financial_Source_URL"),
                item.get("Financial_Review_Status"),
                item.get("Primary_Track"),
                str(registration_derived_counts["FDA"].get(item.get("Company") or "", safe_int(item.get("FDA_Products")))),
                str(registration_derived_counts["NMPA"].get(item.get("Company") or "", safe_int(item.get("NMPA_Products")))),
                item.get("Positioning_CN"),
                search,
            ),
        )
        cur.execute(
            """
            INSERT INTO evidence
            (content_type, title, subtitle, body, url, source_file, source_sheet, confidence, year,
             segment, region, country, company, brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "company",
                item.get("Company"),
                f"{item.get('HQ_Country')} · {item.get('Ownership')} · {item.get('Primary_Track')}",
                search,
                "",
                "全球医美企业库_标准化版v4.xlsx",
                "Companies",
                "unverified_seed",
                detect_year(item.get("Revenue_Year")),
                detect_segments(search)[0],
                item.get("Region"),
                item.get("HQ_Country"),
                item.get("Company"),
                "",
            ),
        )

    for item in brands:
        search = text_blob(item)
        cur.execute(
            """
            INSERT INTO brands
            (company, brand, country, category_l1, category_l2, tech_type, brand_type,
             product_count, products, segments, search_blob)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.get("Company"),
                item.get("Brand"),
                item.get("Country"),
                item.get("Category_L1"),
                item.get("Category_L2"),
                item.get("Tech_Type"),
                item.get("Brand_Type"),
                item.get("Product_Count"),
                item.get("Products"),
                item.get("Segments"),
                search,
            ),
        )

    for item in conferences:
        search = text_blob(item)
        cur.execute(
            """
            INSERT INTO conferences
            (kind, event_name, brand, company, sponsor_level, country, region, city, category,
             products, website, date_text, year, source_file, source_sheet, notes, segments, search_blob)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.get("kind"),
                item.get("event_name"),
                item.get("brand"),
                item.get("company"),
                item.get("sponsor_level"),
                item.get("country"),
                item.get("region"),
                item.get("city"),
                item.get("category"),
                item.get("products"),
                item.get("website"),
                item.get("date_text"),
                item.get("year"),
                item.get("source_file"),
                item.get("source_sheet"),
                item.get("notes"),
                item.get("segments"),
                search,
            ),
        )
        title = item.get("brand") or item.get("event_name")
        cur.execute(
            """
            INSERT INTO evidence
            (content_type, title, subtitle, body, url, source_file, source_sheet, confidence, year,
             segment, region, country, company, brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "conference",
                title,
                f"{item.get('event_name')} · {item.get('sponsor_level') or item.get('category')}",
                search,
                item.get("website"),
                item.get("source_file"),
                item.get("source_sheet"),
                "conference/source",
                item.get("year"),
                item.get("segments"),
                item.get("region"),
                item.get("country"),
                item.get("company"),
                item.get("brand"),
            ),
        )

    for item in metrics:
        search = text_blob(item)
        cur.execute(
            """
            INSERT INTO market_metrics
            (source_file, data_type, category_l1, category_l2, category_l3, geo, value, unit,
             year, source_org, report_title, url, note, confidence, segments, search_blob)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.get("source_file"),
                item.get("data_type"),
                item.get("category_l1"),
                item.get("category_l2"),
                item.get("category_l3"),
                item.get("geo"),
                item.get("value"),
                item.get("unit"),
                item.get("year"),
                item.get("source_org"),
                item.get("report_title"),
                item.get("url"),
                item.get("note"),
                item.get("confidence"),
                item.get("segments"),
                search,
            ),
        )
        cur.execute(
            """
            INSERT INTO evidence
            (content_type, title, subtitle, body, url, source_file, source_sheet, confidence, year,
             segment, region, country, company, brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "market_metric",
                f"{item.get('data_type')} · {item.get('geo')} · {item.get('year') or ''}",
                f"{item.get('value')} {item.get('unit')} · {item.get('source_org')}",
                search,
                item.get("url"),
                item.get("source_file"),
                "",
                item.get("confidence") or "source-stated",
                item.get("year"),
                item.get("segments"),
                "",
                item.get("geo"),
                "",
                "",
            ),
        )

    for item in reports:
        search = f"{item['title']} | {item['body'][:4000]}"
        cur.execute(
            """
            INSERT INTO reports (title, path, body, segments, year, search_blob)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (item["title"], item["path"], item["body"], item["segments"], item["year"], search),
        )
        cur.execute(
            """
            INSERT INTO evidence
            (content_type, title, subtitle, body, url, source_file, source_sheet, confidence, year,
             segment, region, country, company, brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "report",
                item["title"],
                item["path"],
                item["body"][:12000],
                "",
                item["path"],
                "",
                "local-report",
                item["year"],
                item["segments"],
                "",
                "",
                "",
                "",
            ),
        )

    for item in socials:
        cur.execute(
            "INSERT INTO social_sources VALUES (?, ?, ?, ?, ?)",
            (item["platform"], item["status"], item["tool"], item["scope"], item["note"]),
        )
        cur.execute(
            """
            INSERT INTO evidence
            (content_type, title, subtitle, body, url, source_file, source_sheet, confidence, year,
             segment, region, country, company, brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "social_status",
                item["platform"],
                item["status"],
                f"{item['tool']} | {item['scope']} | {item['note']}",
                "",
                "agent-reach doctor",
                "",
                "tool-status",
                None,
                "",
                "",
                "",
                "",
                "",
            ),
        )

    try:
        cur.execute(
            """
            CREATE VIRTUAL TABLE evidence_fts USING fts5(
              title, subtitle, body, company, brand, segment,
              content='evidence', content_rowid='id'
            );
            """
        )
        cur.execute(
            """
            INSERT INTO evidence_fts(rowid, title, subtitle, body, company, brand, segment)
            SELECT id, title, subtitle, body, company, brand, segment FROM evidence;
            """
        )
    except sqlite3.OperationalError:
        pass

    conn.commit()
    conn.close()


def build_snapshot(
    products: list[dict[str, Any]],
    companies: list[dict[str, Any]],
    brands: list[dict[str, Any]],
    conferences: list[dict[str, Any]],
    metrics: list[dict[str, Any]],
    reports: list[dict[str, Any]],
    socials: list[dict[str, Any]],
    quality: dict[str, Any],
) -> dict[str, Any]:
    region_counts = Counter(dashboard_region(p.get("Region"), p.get("Country")) for p in products)
    country_counts = Counter(p.get("Country") or "Unknown" for p in products)
    category_counts = Counter(p.get("Category_L1") or "Unknown" for p in products)
    public_companies = [c for c in companies if (c.get("Ownership") or "").lower() == "public" or ":" in (c.get("Stock_Code") or "")]
    company_id_map = company_ids(companies)
    company_master = build_company_master(companies)
    geo_data = build_geo_companies(company_master, products)
    product_master = build_product_master(products, company_id_map)
    product_hierarchy = build_product_hierarchy(products, company_id_map)
    taxonomy_structure = material_taxonomy_structure(products)
    registration_seed = build_registration_seed(products, company_id_map)
    product_lookup = {product_id_for(product): product for product in products}
    active_product_ids = set(product_lookup)
    market_snapshots = build_market_snapshots(company_master)
    company_financial_metrics = load_generated_csv(COMPANY_FINANCIAL_METRICS_PATH)
    company_revenue_collection_plan = load_audit_csv(COMPANY_REVENUE_COLLECTION_PLAN_PATH)
    verification_queue = build_verification_queue(company_master)
    staging_records = load_staging_records()
    company_background_evidence = load_company_background_evidence()
    company_capital_structure = load_company_capital_structure()
    listed_company_batch = load_listed_company_batch()
    company_official_source_plan = load_company_official_source_plan()
    company_official_source_evidence = load_company_official_source_evidence()
    official_website_master = load_generated_csv(OFFICIAL_WEBSITE_MASTER_PATH)
    company_official_website = load_generated_csv(COMPANY_OFFICIAL_WEBSITE_PATH)
    company_media_asset_index = load_generated_csv(COMPANY_MEDIA_ASSET_INDEX_PATH)
    product_specification_evidence = load_generated_csv(PRODUCT_SPECIFICATION_EVIDENCE_PATH)
    manual_product_fact_evidence = load_generated_csv(MANUAL_PRODUCT_FACT_EVIDENCE_PATH)
    policy_regulatory_source_plan = load_policy_regulatory_source_plan()
    mdr_ce_search_plan = load_mdr_ce_search_plan()
    mdr_ce_evidence_candidates = load_mdr_ce_evidence_candidates()
    news_regulatory_event_candidates = load_news_regulatory_event_candidates()
    briefing_update_candidates = load_briefing_update_candidates()
    briefing_verified_update_events = load_briefing_verified_update_events()
    briefing_fulltext_rescue = load_briefing_fulltext_rescue()
    briefing_product_gap_candidates = load_briefing_product_gap_candidates()
    data_usability_summary = load_data_usability_summary()
    data_usability_ledger = load_audit_csv(DATA_USABILITY_LEDGER_PATH)
    data_usability_missing_owner = load_audit_csv(DATA_USABILITY_MISSING_OWNER_PATH)
    evidence_promotion = build_evidence_promotions(
        product_master,
        product_hierarchy,
        staging_records,
        mdr_ce_evidence_candidates,
        company_official_source_evidence,
        official_website_master,
        company_official_website,
    )
    product_fact_summary = apply_manual_product_fact_evidence(product_master, manual_product_fact_evidence)
    evidence_promotion["summary"].update(product_fact_summary)
    manual_official_indication_rows = load_manual_official_indication_rows()
    manual_nmpa_registration_rows = load_manual_nmpa_registration_rows()
    manual_registration_name_rows = load_manual_registration_name_rows()
    promoted_registration_rows = (
        evidence_promotion["registration_rows"]
        + manual_official_indication_rows
        + manual_nmpa_registration_rows
        + manual_registration_name_rows
    )
    promotion_summary = evidence_promotion["summary"]
    official_indication_analysis = build_official_indication_analysis(
        registration_seed,
        staging_records,
        promoted_registration_rows,
        product_lookup,
    )
    registration_evidence_rows = build_registration_evidence_output(
        registration_seed,
        staging_records,
        promoted_registration_rows,
        active_product_ids,
    )
    company_portfolio_cases = enrich_company_portfolio_cases(
        load_company_portfolio_cases(),
        product_master,
        official_website_master,
        product_specification_evidence,
        company_official_source_evidence,
        official_indication_analysis,
    )
    staged_registration_count = sum(1 for item in staging_records if item.get("merge_target") == "registration_evidence")
    quality_summary = quality.get("summary", {})
    priority_companies = [item for item in company_master if item.get("priority_rank")]
    queued_by_lane = Counter(item["source_lane"] for item in verification_queue)
    queued_by_fact = Counter(item["fact_group"] for item in verification_queue)
    source_by_status = Counter(item["automation_status"] for item in SOURCE_REGISTRY)
    staged_by_source = Counter(item.get("source_key") or "unknown" for item in staging_records)
    background_by_source = Counter(item.get("source_key") or "unknown" for item in company_background_evidence)
    capital_by_status = Counter(item.get("evidence_status") or "unknown" for item in company_capital_structure)
    listed_batch_by_status = Counter(item.get("listing_verification_status") or "unknown" for item in listed_company_batch)
    listed_batch_by_relation = Counter(item.get("relation_to_listed_entity") or "unknown" for item in listed_company_batch)
    official_evidence_by_confidence = Counter(item.get("confidence") or "unknown" for item in company_official_source_evidence)
    website_by_scope = Counter(item.get("entity_scope") or "unknown" for item in official_website_master)
    media_by_type = Counter(item.get("asset_type") or "unknown" for item in company_media_asset_index)
    specs_by_category = Counter(item.get("spec_category") or "unknown" for item in product_specification_evidence)
    ce_plan_by_source = Counter(item.get("source_key") or "unknown" for item in mdr_ce_search_plan)
    ce_candidate_by_confidence = Counter(item.get("confidence") or "unknown" for item in mdr_ce_evidence_candidates)
    ce_candidate_by_official = Counter(item.get("official_candidate") or "unknown" for item in mdr_ce_evidence_candidates)
    ce_candidate_by_source = Counter(item.get("source_key") or "unknown" for item in mdr_ce_evidence_candidates)
    record_family_ids: dict[str, set[str]] = defaultdict(set)
    for family in product_hierarchy["families"]:
        family_id = family.get("product_family_id")
        for record_id in [part.strip() for part in norm(family.get("source_record_ids")).split(",") if part.strip()]:
            record_family_ids[record_id].add(family_id)

    def scoped_registration_timeline(seed_rows: list[dict[str, Any]], staged_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        years: Counter = Counter()
        for row in seed_rows:
            year = extract_year(row.get("approval_date"))
            if year:
                years[year] += 1
        for row in staged_rows:
            candidates = row.get("field_candidates") or {}
            if isinstance(candidates, str):
                try:
                    candidates = json.loads(candidates)
                except json.JSONDecodeError:
                    candidates = {}
            year = extract_year(candidates.get("approval_date"))
            if year:
                years[year] += 1
        for row in promoted_registration_rows:
            year = extract_year(row.get("approval_date"))
            if year:
                years[year] += 1
        return [
            {
                "year": str(year),
                "total": years[year],
                "fda": years[year],
                "ce": 0,
                "nmpa": 0,
                "launch": 0,
            }
            for year in sorted(years)[-10:]
        ]

    def evidence_scope_for(product_rows: list[dict[str, Any]]) -> dict[str, Any]:
        product_ids = {product_id_for(product) for product in product_rows}
        family_ids = {
            family_id
            for product in product_rows
            for family_id in record_family_ids.get(product.get("Record_ID") or "", set())
        }
        seed_rows = [row for row in registration_seed if row.get("product_id") in product_ids]
        staged_rows = [
            row
            for row in staging_records
            if row.get("merge_target") == "registration_evidence" and row.get("product_id") in product_ids
        ]
        promoted_rows = [row for row in promoted_registration_rows if row.get("product_id") in product_ids]
        fda_rows = [
            row
            for row in [*seed_rows, *staged_rows, *promoted_rows]
            if "fda" in norm(row.get("source_key")).lower() or norm(row.get("regulator")).upper() == "FDA"
        ]
        mdr_promoted_rows = [
            row
            for row in promoted_rows
            if row.get("jurisdiction") in {"EU", "EU / Global"} or "ce" in norm(row.get("regulatory_pathway")).lower()
        ]
        mdr_plan_rows = [row for row in mdr_ce_search_plan if row.get("product_family_id") in family_ids]
        mdr_candidate_rows = [row for row in mdr_ce_evidence_candidates if row.get("product_family_id") in family_ids]
        official_indication_rows = [
            row for row in official_indication_analysis.get("records", []) if row.get("product_id") in product_ids
        ]
        spec_rows = [
            row
            for row in product_specification_evidence
            if row.get("product_id") in product_ids or row.get("product_family_id") in family_ids
        ]
        return {
            "product_rows": len(product_rows),
            "regulatory_seed_signals": sum(1 for product in product_rows if has_regulatory_signal(product)),
            "registration_evidence_rows": len(seed_rows) + len(staged_rows) + len(promoted_rows),
            "fda_merged_rows": len(fda_rows),
            "mdr_ce_plan_rows": len(mdr_plan_rows),
            "mdr_ce_candidate_rows": len(mdr_candidate_rows),
            "mdr_ce_merged_rows": len(mdr_promoted_rows),
            "promoted_registration_rows": len(promoted_rows),
            "promoted_indication_rows": len(official_indication_rows),
            "official_indication_heatmap": official_indication_heatmap(official_indication_rows, "regulator", 10),
            "official_indication_country_heatmap": official_indication_heatmap(official_indication_rows, "country", 10),
            "official_indication_top_buckets": top_counts(
                Counter(bucket for row in official_indication_rows for bucket in (row.get("buckets") or [])),
                10,
            ),
            "spec_candidate_rows": len(spec_rows),
            "candidate_by_confidence": top_counts(Counter(row.get("confidence") or "unknown" for row in mdr_candidate_rows), 6),
            "timeline": scoped_registration_timeline(seed_rows, staged_rows),
            "note": "Scoped to products in this segment. MDR/CE candidate rows are separate from promoted IFU/certificate/EUDAMED/FDA evidence.",
        }

    product_master_by_id = {row.get("product_id"): row for row in product_master if row.get("product_id")}

    def scope_rows(rows: list[dict[str, Any]], product_id: str, family_ids: set[str]) -> list[dict[str, Any]]:
        return [
            row
            for row in rows
            if row.get("product_id") == product_id
            or (row.get("product_family_id") and row.get("product_family_id") in family_ids)
        ]

    def product_evidence_audit_for(
        product_rows: list[dict[str, Any]],
        product_subtracks: dict[str, list[str]],
        product_indications: dict[str, list[str]],
    ) -> list[dict[str, Any]]:
        output = []
        for product in product_rows[:80]:
            record_id = product.get("Record_ID") or ""
            product_id = product_id_for(product)
            family_ids = set(record_family_ids.get(record_id, set()))
            master = product_master_by_id.get(product_id, {})
            regs = scope_rows(registration_seed + promoted_registration_rows, product_id, family_ids)
            staged_regs = [
                row
                for row in staging_records
                if row.get("merge_target") == "registration_evidence" and row.get("product_id") == product_id
            ]
            indications = [
                row for row in official_indication_analysis.get("records", []) if row.get("product_id") == product_id
            ]
            websites = scope_rows(official_website_master, product_id, family_ids)
            specs = scope_rows(product_specification_evidence, product_id, family_ids)
            ce_plans = scope_rows(mdr_ce_search_plan, product_id, family_ids)
            ce_candidates = scope_rows(mdr_ce_evidence_candidates, product_id, family_ids)
            issues = []
            status = master.get("verification_status") or "unverified_seed"
            subtrack_text = " ".join(product_subtracks.get(record_id, []))
            material_text = " ".join(
                [
                    norm(master.get("technology_path_l1")),
                    norm(master.get("technology_path_l2")),
                    norm(master.get("material_or_energy_source")),
                    norm(product.get("Tech_Type_Std")),
                    norm(product.get("Core_Product")),
                ]
            ).lower()
            if status == "unverified_seed":
                issues.append("主表仍是 seed")
            if ("CaHA" in subtrack_text or "羟基" in subtrack_text) and "hyaluronic acid" in material_text and "calcium" not in material_text:
                issues.append("材料/子赛道冲突")
            if " / " in norm(product.get("Core_Product")) and any(token in norm(product.get("Core_Product")).lower() for token in ["intense", "stimulate"]):
                issues.append("疑似需拆 SKU")
            indication_unavailable = has_unavailable_public_indication_marker(regs)
            if not regs and not staged_regs:
                issues.append("缺注册证据")
            if not indications and not indication_unavailable:
                issues.append("缺官方适应症")
            if not websites:
                issues.append("缺官网直连")
            if not specs:
                issues.append("缺规格候选")
            score = 0
            if status != "unverified_seed" and "材料/子赛道冲突" not in issues:
                score += 20
            if websites:
                score += 20
            if specs:
                score += 20
            if regs or staged_regs:
                score += 20
            if indications or indication_unavailable:
                score += 20
            output.append(
                {
                    "record_id": record_id,
                    "product_id": product_id,
                    "company": product.get("Company"),
                    "brand": product.get("Brand"),
                    "product": product.get("Core_Product") or product.get("Brand"),
                    "country": product.get("Country"),
                    "subtracks": product_subtracks.get(record_id, []),
                    "positioning_indications": product_indications.get(record_id, []),
                    "verification_status": status,
                    "source_status": master.get("source_status") or product.get("Data_Source"),
                    "website_rows": len(websites),
                    "spec_rows": len(specs),
                    "registration_rows": len(regs) + len(staged_regs),
                    "official_indication_rows": len(indications),
                    "official_indication_unavailable": indication_unavailable,
                    "mdr_ce_plan_rows": len(ce_plans),
                    "mdr_ce_candidate_rows": len(ce_candidates),
                    "completeness_score": score,
                    "issues": issues,
                    "primary_source_url": (websites[0].get("official_website_url") if websites else ""),
                }
            )
        output.sort(key=lambda row: (-row["completeness_score"], row["company"] or "", row["brand"] or ""))
        return output

    registration_years: Counter = Counter()
    for item in registration_seed:
        year = extract_year(item.get("approval_date"))
        if year:
            registration_years[year] += 1
    for item in staging_records:
        if item.get("merge_target") != "registration_evidence":
            continue
        candidates = item.get("field_candidates") or {}
        if isinstance(candidates, str):
            try:
                candidates = json.loads(candidates)
            except json.JSONDecodeError:
                candidates = {}
        year = extract_year(candidates.get("approval_date"))
        if year:
            registration_years[year] += 1
    for item in promoted_registration_rows:
        year = extract_year(item.get("approval_date"))
        if year:
            registration_years[year] += 1
    families_by_brand = Counter((item.get("company") or "", item.get("brand") or "") for item in product_hierarchy["families"])
    sku_split_counter = Counter(item.get("split_status") or "unknown" for item in product_hierarchy["skus"])
    multi_family_brand_count = sum(1 for count in families_by_brand.values() if count > 1)

    segment_stats = []
    global_indication_counter: Counter = Counter()
    global_subtrack_counter: Counter = Counter()
    for segment in SEGMENTS:
        code = segment["code"]
        seg_products = [p for p in products if code in (p.get("Segments") or "").split(",")]
        seg_brands = [b for b in brands if code in (b.get("Segments") or "").split(",")]
        region_counter = Counter(dashboard_region(p.get("Region"), p.get("Country")) for p in seg_products)
        seg_products_by_region = [
            {**product, "Region": dashboard_region(product.get("Region"), product.get("Country"))}
            for product in seg_products
        ]
        company_counter = Counter(p.get("Company") or "Unknown" for p in seg_products)
        brand_counter = Counter(p.get("Brand") or p.get("Core_Product") or "Unknown" for p in seg_products)
        country_counter = Counter(p.get("Country") or "Unknown" for p in seg_products)
        product_subtracks: dict[str, list[str]] = {}
        product_indications: dict[str, list[str]] = {}
        subtrack_counter: Counter = Counter()
        indication_counter: Counter = Counter()
        for product in seg_products:
            product_text = text_blob(product)
            product_id = product.get("Record_ID") or ""
            subtracks = detect_taxonomy(product_text, code, "subtracks", None)
            subtracks = normalize_product_subtracks(product, code, subtracks)
            indications = detect_taxonomy(product_text, code, "indications", None)
            product_subtracks[product_id] = subtracks
            product_indications[product_id] = indications
            subtrack_counter.update(subtracks)
            indication_counter.update(indications)
        global_subtrack_counter.update(subtrack_counter)
        global_indication_counter.update(indication_counter)
        region_columns = [name for name, _ in region_counter.most_common(5)] or ["Unknown"]
        regulatory = regulatory_counts(seg_products)
        configured_subtracks = [item["name"] for item in SEGMENT_TAXONOMY.get(code, {}).get("subtracks", [])]
        configured_indications = [item["name"] for item in SEGMENT_TAXONOMY.get(code, {}).get("indications", [])]
        segment_stats.append(
            {
                **{k: segment[k] for k in ["code", "name", "subtitle", "color"]},
                "configured_subtracks": configured_subtracks,
                "configured_indications": configured_indications,
                "products": len(seg_products),
                "companies": len({p.get("Company") for p in seg_products if p.get("Company")}),
                "brands": len({p.get("Brand") for p in seg_products if p.get("Brand")}) or len(seg_brands),
                "countries": len({p.get("Country") for p in seg_products if p.get("Country")}),
                "subtrack_count": len(subtrack_counter),
                "indication_count": len(indication_counter),
                "top_subtracks": top_counts(subtrack_counter, 8),
                "top_indications": top_counts(indication_counter, 8),
                "subtrack_heatmap": heatmap_from_rows(seg_products_by_region, product_subtracks, region_columns, "Region", 8),
                "indication_heatmap": heatmap_from_rows(seg_products_by_region, product_indications, region_columns, "Region", 8),
                "subtrack_slices": build_subtrack_slices(
                    seg_products,
                    product_subtracks,
                    product_indications,
                    subtrack_counter,
                    12,
                    evidence_scope_for,
                ),
                "top_regions": top_counts(region_counter, 6),
                "top_countries": top_counts(country_counter, 6),
                "top_companies": top_counts(company_counter, 8),
                "top_brands": top_counts(brand_counter, 8),
                "regulatory": regulatory,
                "evidence_scope": evidence_scope_for(seg_products),
                "product_evidence_audit": product_evidence_audit_for(
                    seg_products,
                    product_subtracks,
                    product_indications,
                ),
                **enrich_analysis(
                    seg_products,
                    product_subtracks,
                    product_indications,
                    configured_subtracks,
                    configured_indications,
                ),
                "sample_products": representative_products(seg_products, product_subtracks, product_indications, 260),
            }
        )

    company_segment_counts: dict[str, Counter] = defaultdict(Counter)
    company_segment_products: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for product in products:
        company = product.get("Company")
        if not company:
            continue
        for code in (product.get("Segments") or "other").split(","):
            company_segment_counts[company][code] += 1
            company_segment_products[company][code].append(product)

    def matrix_product_example(product: dict[str, Any], matrix_company: str) -> dict[str, str]:
        brand = norm(product.get("Brand"))
        product_name = norm(product.get("Core_Product")) or brand
        maker = norm(product.get("Manufactured_By"))
        oem_for = norm(product.get("OEM_For"))
        holder = norm(product.get("Company"))
        path_parts = []
        for value in [brand, product_name, maker or holder, oem_for or (holder if holder != maker else "")]:
            if value and value not in path_parts:
                path_parts.append(value)
        return {
            "label": " / ".join(path_parts) or matrix_company,
            "brand": brand,
            "product": product_name,
            "manufacturer": maker,
            "holder": holder,
            "oem_for": oem_for,
            "country": norm(product.get("Country")),
            "record_id": norm(product.get("Record_ID")),
        }

    top_company_names = [name for name, _ in Counter(p.get("Company") for p in products if p.get("Company")).most_common(24)]
    company_matrix = []
    for name in top_company_names:
        counts = company_segment_counts[name]
        examples = {}
        for code in [s["code"] for s in SEGMENTS]:
            segment_products = sorted(
                company_segment_products[name].get(code, []),
                key=lambda row: (norm(row.get("Brand")).lower(), norm(row.get("Core_Product")).lower(), norm(row.get("Record_ID")).lower()),
            )
            examples[code] = [matrix_product_example(product, name) for product in segment_products[:5]]
        company_matrix.append(
            {
                "company": name,
                "total": sum(counts.values()),
                "segments": {code: counts.get(code, 0) for code in [s["code"] for s in SEGMENTS]},
                "examples": examples,
            }
        )

    market_preview = [
        {
            "source_file": m.get("source_file"),
            "type": m.get("data_type"),
            "category": " / ".join(x for x in [m.get("category_l1"), m.get("category_l2"), m.get("category_l3")] if x),
            "geo": m.get("geo"),
            "value": m.get("value"),
            "unit": m.get("unit"),
            "year": m.get("year"),
            "source": m.get("source_org"),
            "title": m.get("report_title"),
            "url": m.get("url"),
            "confidence": m.get("confidence"),
        }
        for m in metrics[:1000]
    ]

    def latest_text_date(rows: list[dict[str, Any]], fields: list[str]) -> str:
        values: list[str] = []
        for row in rows:
            for field in fields:
                value = norm(row.get(field))
                if value:
                    values.append(value)
                    break
        return max(values) if values else ""

    financial_kpi_refresh = {
        "latest_updated_at": max(
            [
                value
                for value in [
                    latest_text_date(market_snapshots, ["market_refreshed_at", "as_of"]),
                    latest_text_date(company_financial_metrics, ["captured_at", "filing_date", "revenue_filed"]),
                ]
                if value
            ]
            or [""]
        ),
        "market_last_updated_at": latest_text_date(market_snapshots, ["market_refreshed_at", "as_of"]),
        "financial_last_updated_at": latest_text_date(company_financial_metrics, ["captured_at", "filing_date", "revenue_filed"]),
        "listed_rows": len(market_snapshots),
        "with_share_price": sum(1 for item in market_snapshots if item.get("price")),
        "with_market_cap": sum(1 for item in market_snapshots if item.get("market_cap_usd_m")),
        "with_pe": sum(1 for item in market_snapshots if item.get("pe_ratio")),
        "with_pb": sum(1 for item in market_snapshots if item.get("pb_ratio")),
        "with_eps": sum(1 for item in market_snapshots if item.get("eps_ttm")),
        "with_net_profit_growth": sum(1 for item in market_snapshots if item.get("net_profit_growth_yoy_pct")),
        "source_note": "Market price/valuation refreshes weekly; financial statement fields refresh from SEC XBRL and curated IR annual reports.",
    }

    snapshot = {
        "generated_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "source_root": str(SOURCE_DIR),
        "summary": {
            "products": len(products),
            "companies": len(companies),
            "brands": len(brands),
            "countries": len({p.get("Country") for p in products if p.get("Country")}),
            "regions": len({dashboard_region(p.get("Region"), p.get("Country")) for p in products if dashboard_region(p.get("Region"), p.get("Country"))}),
            "market_metrics": len(metrics),
            "reports": len(reports),
            "public_companies": len(public_companies),
            "indication_signals": sum(global_indication_counter.values()),
            "subtrack_signals": sum(global_subtrack_counter.values()),
            "company_master": len({item["company_id"] for item in company_master}),
            "product_master": len(product_master),
            "product_families": len(product_hierarchy["families"]),
            "product_sku_candidates": len(product_hierarchy["skus"]),
            "brands_with_multiple_families": multi_family_brand_count,
            "registration_evidence": len(registration_evidence_rows),
            "company_background_evidence": len(company_background_evidence),
            "company_capital_structure": len(company_capital_structure),
            "listed_company_batch": len(listed_company_batch),
            "company_official_source_plan": len(company_official_source_plan),
            "company_official_source_evidence": len(company_official_source_evidence),
            "official_website_master": len(official_website_master),
            "company_official_website": len(company_official_website),
            "company_media_asset_index": len(company_media_asset_index),
            "product_specification_evidence": len(product_specification_evidence),
            "manual_product_fact_evidence": len(manual_product_fact_evidence),
            "policy_regulatory_source_plan": len(policy_regulatory_source_plan),
            "mdr_ce_search_plan": len(mdr_ce_search_plan),
            "mdr_ce_evidence_candidates": len(mdr_ce_evidence_candidates),
            "news_regulatory_event_candidates": len(news_regulatory_event_candidates),
            "briefing_update_candidates": len(briefing_update_candidates),
            "briefing_verified_update_events": len(briefing_verified_update_events),
            "briefing_fulltext_rescue": len(briefing_fulltext_rescue),
            "briefing_product_gap_candidates": len(briefing_product_gap_candidates),
            "evidence_promoted_registration": promotion_summary.get("registration_rows_promoted", 0),
            "manual_official_indication_evidence": len(manual_official_indication_rows),
            "manual_nmpa_registration_evidence": len(manual_nmpa_registration_rows),
            "manual_registration_name_evidence": len(manual_registration_name_rows),
            "evidence_promoted_product_master": promotion_summary.get("product_master_promoted", 0),
            "manual_product_fact_products": promotion_summary.get("manual_product_fact_products", 0),
            "official_indication_rows": official_indication_analysis.get("rows", 0),
            "official_sources": sum(1 for item in SOURCE_REGISTRY if source_scope_status(item) != "external_project"),
            "verification_queue": len(verification_queue),
            "evidence_staging": len(staging_records),
            "market_snapshot": len(market_snapshots),
            "company_financial_metrics": len(company_financial_metrics),
            "company_revenue_collection_plan": len(company_revenue_collection_plan),
            "financial_kpi_rows": financial_kpi_refresh["listed_rows"],
            "source_authority_rules": len(SOURCE_AUTHORITY_POLICY),
            "field_dictionary": len(FIELD_DICTIONARY_ROWS),
            "company_portfolio_cases": len(company_portfolio_cases),
            "data_quality_issues": quality_summary.get("total", 0),
            "data_quality_high_issues": quality_summary.get("critical", 0) + quality_summary.get("high", 0),
            "data_usability_audited_rows": data_usability_summary.get("audited_rows", 0),
            "data_usability_missing_owner_rows": data_usability_summary.get("missing_owner_or_status_rows", 0),
            "mapped_companies": geo_data["summary"].get("mapped_companies", 0),
        },
        "segments": segment_stats,
        "region_distribution": top_counts(region_counts, 12),
        "country_distribution": top_counts(country_counts, 18),
        "geo_summary": geo_data["summary"],
        "geo_companies": geo_data["companies"],
        "geo_points": geo_data["points"],
        "geo_city_clusters": geo_data["city_clusters"],
        "category_distribution": top_counts(category_counts, 12),
        "material_taxonomy_structure": taxonomy_structure,
        "indication_distribution": top_counts(global_indication_counter, 18),
        "official_indication_analysis": official_indication_analysis,
        "company_portfolio_cases": company_portfolio_cases,
        "subtrack_distribution": top_counts(global_subtrack_counter, 18),
        "analysis_blueprint": [
            {"name": "一级赛道", "value": len([item for item in segment_stats if item.get("products")]), "unit": "页"},
            {"name": "二级子赛道", "value": len(global_subtrack_counter), "unit": "类"},
            {"name": "适应症信号", "value": len(global_indication_counter), "unit": "类"},
            {"name": "国家覆盖", "value": len({p.get("Country") for p in products if p.get("Country")}), "unit": "国"},
            {"name": "监管字段", "value": sum(regulatory_counts(products).values()), "unit": "条"},
            {"name": "公开市场指标", "value": len(metrics), "unit": "条"},
        ],
        "global_regulatory_mix": regulatory_mix(products),
        "regulatory_atlas": REGULATORY_ATLAS,
        "source_authority_policy": SOURCE_AUTHORITY_POLICY,
        "field_dictionary": FIELD_DICTIONARY_ROWS,
        "global_evidence_funnel": evidence_funnel(products),
        "global_approval_timeline": approval_timeline(products),
        "product_hierarchy": {
            "summary": {
                "families": len(product_hierarchy["families"]),
                "sku_candidates": len(product_hierarchy["skus"]),
                "brands_with_multiple_families": multi_family_brand_count,
                "duplicate_non_primary_skus": sum(1 for item in product_hierarchy["skus"] if item.get("is_primary_record") == "0"),
                "split_status": top_counts(sku_split_counter, 8),
            },
            "top_families": product_hierarchy["families"][:30],
            "split_candidates": [
                item for item in product_hierarchy["skus"] if item.get("split_status") != "family_level"
            ][:40],
            "source_note": "Derived from Product_Lines as a review layer: company -> brand -> product family -> SKU/model candidate.",
        },
        "company_matrix": company_matrix,
        "public_companies": [
            {
                "company": c.get("Company"),
                "country": c.get("HQ_Country"),
                "region": dashboard_region(c.get("Region"), c.get("HQ_Country")),
                "stock": c.get("Stock_Code"),
                "track": c.get("Primary_Track"),
                "products": safe_float(c.get("Product_Count")) or None,
                "revenue": c.get("Revenue_USD_M"),
                "market_cap": c.get("Market_Cap_USD_M"),
                "market_cap_usd_m": c.get("Market_Cap_USD_M"),
                "share_price": c.get("Market_Price"),
                "pe_ratio": c.get("PE_Ratio"),
                "pb_ratio": c.get("PB_Ratio"),
                "eps_ttm": c.get("EPS_TTM") or c.get("EPS"),
                "net_profit_growth_pct": c.get("Net_Profit_Growth_Pct"),
                "market_refreshed_at": c.get("Market_Refreshed_At") or c.get("Market_Captured_At"),
                "financial_refreshed_at": c.get("Financial_Refreshed_At"),
                "financial_period": c.get("Financial_Period"),
            }
            for c in public_companies[:36]
        ],
        "financial_kpi_refresh": financial_kpi_refresh,
        "market_metrics": market_preview,
        "verification_workbench": {
            "policy": "official-source precedence: regulator records for registration facts; company official pages/IFU for product facts; secondary media for cross-check only",
            "seed_status": "unverified_seed",
            "merge_policy": "自动交叉验证；不设人工确认门槛，后续在应用中发现细节问题再做主表修正",
            "top_company_rule": "Product_Count top 30 plus listed high-product companies",
            "top_companies": [
                {
                    "rank": item.get("priority_rank"),
                    "company_id": item.get("company_id"),
                    "company": item.get("canonical_name"),
                    "country": item.get("hq_country"),
                    "region": item.get("region"),
                    "products": item.get("product_count"),
                    "brands": item.get("brand_count"),
                    "ownership": item.get("ownership"),
                    "stock": item.get("stock_code"),
                    "track": item.get("primary_track"),
                    "review_status": item.get("review_status"),
                }
                for item in sorted(priority_companies, key=lambda row: int(row.get("priority_rank") or 999))[:40]
            ],
            "queue_summary": {
                "total": len(verification_queue),
                "by_lane": top_counts(queued_by_lane, 8),
                "by_fact": top_counts(queued_by_fact, 10),
            },
            "source_registry": [
                {
                    "key": item["source_key"],
                    "channel_code": item.get("channel_code"),
                    "kind": item.get("source_kind"),
                    "scope_status": source_scope_status(item),
                    "jurisdiction": item["jurisdiction"],
                    "regulator": item["regulator"],
                    "name": item["source_name"],
                    "url": item["source_url"],
                    "access": item["access_method"],
                    "machine_readable": item["machine_readable"],
                    "status": item["automation_status"],
                    "priority": item["priority"],
                    "use": item["primary_use"],
                }
                for item in SOURCE_REGISTRY
                if source_scope_status(item) != "external_project"
            ],
            "current_phase": {
                "active_channels": ["FDA / 510(k)", "MDR / CE", "其他国家/地区官方注册适应症"],
                "excluded_channels": [
                    {
                        "code": code,
                        "note": note,
                    }
                    for code, note in EXTERNAL_PROJECT_NOTES.items()
                ],
                "note": "China NMPA UDI/registration remains a separate China dashboard lane for full-market discovery, but verified registration-project rows can be imported here when they map to existing global products. FDA and CE/MDR remain the first deep automated lanes; other countries use the same product-country-regulator-indication-date long-table structure.",
            },
            "source_status": top_counts(source_by_status, 8),
            "staging_summary": {
                "total": len(staging_records),
                "by_source": top_counts(staged_by_source, 8),
            },
            "company_background": {
                "evidence_rows": len(company_background_evidence),
                "capital_rows": len(company_capital_structure),
                "listed_company_batch_rows": len(listed_company_batch),
                "official_source_plan_rows": len(company_official_source_plan),
                "official_source_evidence_rows": len(company_official_source_evidence),
                "official_website_master_rows": len(official_website_master),
                "company_official_website_rows": len(company_official_website),
                "media_asset_rows": len(company_media_asset_index),
                "product_specification_rows": len(product_specification_evidence),
                "policy_plan_rows": len(policy_regulatory_source_plan),
                "evidence_by_source": top_counts(background_by_source, 8),
                "capital_by_status": top_counts(capital_by_status, 8),
                "listed_batch_by_status": top_counts(listed_batch_by_status, 8),
                "listed_batch_by_relation": top_counts(listed_batch_by_relation, 8),
                "official_evidence_by_confidence": top_counts(official_evidence_by_confidence, 8),
                "official_websites_by_scope": top_counts(website_by_scope, 8),
                "media_assets_by_type": top_counts(media_by_type, 8),
                "product_specs_by_category": top_counts(specs_by_category, 8),
                "official_website_preview": official_website_master[:20],
                "product_specification_preview": product_specification_evidence[:20],
                "listed_company_preview": listed_company_batch[:20],
                "review_ready": sum(1 for item in company_background_evidence if item.get("review_status") == "needs_review"),
                "note": "Official securities/company sources are authoritative for capital and ownership. Website facts are split by parent, operating company, brand and product-line surfaces; media is only a cross-check signal.",
            },
            "mdr_ce_plan": {
                "rows": len(mdr_ce_search_plan),
                "candidate_rows": len(mdr_ce_evidence_candidates),
                "promoted_rows": promotion_summary.get("mdr_ce_rows", 0),
                "eudamed_rows": promotion_summary.get("eudamed_rows", 0),
                "by_source": top_counts(ce_plan_by_source, 8),
                "candidate_by_confidence": top_counts(ce_candidate_by_confidence, 8),
                "candidate_by_official_signal": top_counts(ce_candidate_by_official, 8),
                "candidate_by_source": top_counts(ce_candidate_by_source, 8),
                "candidate_preview": [
                    {
                        "company": item.get("company"),
                        "brand": item.get("brand"),
                        "product_family": item.get("product_family"),
                        "title": item.get("title"),
                        "url": item.get("url"),
                        "confidence": item.get("confidence"),
                        "official_candidate": item.get("official_candidate"),
                        "source_key": item.get("source_key"),
                        "captured_at": item.get("captured_at"),
                    }
                    for item in mdr_ce_evidence_candidates[:30]
                ],
                "review_ready": sum(1 for item in mdr_ce_search_plan if item.get("review_status") == "needs_review"),
                "note": "CE/MDR candidates are search/cross-check evidence until promoted from certificate, IFU, EUDAMED, declaration of conformity, or regulator documents.",
            },
            "news_regulatory_events": {
                "rows": len(news_regulatory_event_candidates),
                "by_regulator": top_counts(Counter(item.get("regulator") or "Unknown" for item in news_regulatory_event_candidates), 8),
                "by_company": top_counts(Counter(item.get("company") or "Unknown" for item in news_regulatory_event_candidates), 10),
                "needs_official_verification": sum(
                    1 for item in news_regulatory_event_candidates if str(item.get("needs_official_verification", "")).lower() == "yes"
                ),
                "candidate_preview": [
                    {
                        "article_date": item.get("article_date"),
                        "company": item.get("company"),
                        "brand": item.get("brand"),
                        "product_name": item.get("product_name"),
                        "regulator": item.get("regulator"),
                        "candidate_indication": item.get("candidate_indication"),
                        "article_title": item.get("article_title"),
                        "article_url": item.get("article_url"),
                        "confidence": item.get("confidence"),
                    }
                    for item in news_regulatory_event_candidates[:30]
                ],
                "note": "Daily briefing news is a discovery lane only. These candidates must be confirmed against regulator, IFU, or official company documents before being promoted to Registration_Evidence or Official_Indication_Evidence.",
            },
            "briefing_update_candidates": {
                "rows": len(briefing_update_candidates),
                "by_event_group": top_counts(Counter(item.get("event_group") or "Unknown" for item in briefing_update_candidates), 8),
                "by_status": top_counts(Counter(item.get("status") or "Unknown" for item in briefing_update_candidates), 8),
                "by_source_domain": top_counts(Counter(item.get("source_domain") or "Unknown" for item in briefing_update_candidates), 12),
                "needs_fulltext_rescue": sum(
                    1 for item in briefing_update_candidates if str(item.get("needs_fulltext_rescue", "")).lower() == "yes"
                ),
                "needs_official_verification": sum(
                    1 for item in briefing_update_candidates if str(item.get("needs_official_verification", "")).lower() == "yes"
                ),
                "high_confidence": sum(1 for item in briefing_update_candidates if int(item.get("confidence_score") or 0) >= 75),
                "candidate_preview": [
                    {
                        "candidate_id": item.get("candidate_id"),
                        "event_group": item.get("event_group"),
                        "event_type": item.get("event_type"),
                        "article_date": item.get("article_date"),
                        "company": item.get("company"),
                        "brand": item.get("brand"),
                        "product_name": item.get("product_name"),
                        "market_or_jurisdiction": item.get("market_or_jurisdiction"),
                        "source_domain": item.get("source_domain"),
                        "article_title": item.get("article_title") or item.get("excerpt", "")[:120],
                        "article_url": item.get("article_url"),
                        "body_quality": item.get("body_quality"),
                        "confidence_score": item.get("confidence_score"),
                        "status": item.get("status"),
                        "needs_fulltext_rescue": item.get("needs_fulltext_rescue"),
                        "needs_official_verification": item.get("needs_official_verification"),
                        "promotion_target": item.get("promotion_target"),
                        "excerpt": item.get("excerpt"),
                    }
                    for item in briefing_update_candidates[:80]
                ],
                "csv_path": str(BRIEFING_UPDATE_CANDIDATES_PATH),
                "summary_path": str(DATA_DIR / "audits" / "briefing_update_summary_latest.md"),
                "note": "Weekly briefing updates are discovery candidates only. Promotions still require regulator, IFU, label, certificate, official company, IR, or distributor evidence depending on fact type.",
            },
            "briefing_verified_updates": {
                "rows": len(briefing_verified_update_events),
                "promoted": sum(
                    1
                    for item in briefing_verified_update_events
                    if item.get("promotion_status") in {"promoted", "promoted_to_log"}
                ),
                "verified_gap": sum(1 for item in briefing_verified_update_events if item.get("promotion_status") == "verified_gap"),
                "by_event_group": top_counts(Counter(item.get("event_group") or "Unknown" for item in briefing_verified_update_events), 8),
                "by_mapping_status": top_counts(Counter(item.get("mapping_status") or "Unknown" for item in briefing_verified_update_events), 8),
                "by_verification_status": top_counts(Counter(item.get("verification_status") or "Unknown" for item in briefing_verified_update_events), 8),
                "by_promotion_status": top_counts(Counter(item.get("promotion_status") or "Unknown" for item in briefing_verified_update_events), 8),
                "by_source_type": top_counts(Counter(item.get("official_source_type") or "Unknown" for item in briefing_verified_update_events), 8),
                "event_preview": [
                    {
                        "event_id": item.get("event_id"),
                        "candidate_id": item.get("candidate_id"),
                        "event_group": item.get("event_group"),
                        "event_type": item.get("event_type"),
                        "article_date": item.get("article_date"),
                        "company": item.get("company"),
                        "brand": item.get("brand"),
                        "product_name": item.get("product_name"),
                        "mapping_status": item.get("mapping_status"),
                        "verification_status": item.get("verification_status"),
                        "promotion_status": item.get("promotion_status"),
                        "official_source_type": item.get("official_source_type"),
                        "official_source_url": item.get("official_source_url"),
                        "official_title": item.get("official_title"),
                        "official_excerpt": item.get("official_excerpt"),
                        "promoted_target": item.get("promoted_target"),
                        "remaining_gap": item.get("remaining_gap"),
                    }
                    for item in briefing_verified_update_events[:80]
                ],
                "fulltext_rescue": {
                    "rows": len(briefing_fulltext_rescue),
                    "rescued": sum(1 for item in briefing_fulltext_rescue if item.get("fetch_status") == "ok"),
                    "short_body": sum(1 for item in briefing_fulltext_rescue if item.get("fetch_status") == "short_body"),
                    "failed": sum(1 for item in briefing_fulltext_rescue if item.get("fetch_status") == "fetch_failed"),
                    "by_status": top_counts(Counter(item.get("fetch_status") or "Unknown" for item in briefing_fulltext_rescue), 8),
                    "preview": [
                        {
                            "candidate_id": item.get("candidate_id"),
                            "fetch_status": item.get("fetch_status"),
                            "char_count": item.get("char_count"),
                            "title": item.get("title"),
                            "article_url": item.get("article_url"),
                            "final_url": item.get("final_url"),
                            "body_excerpt": item.get("body_excerpt"),
                            "error": item.get("error"),
                        }
                        for item in briefing_fulltext_rescue[:30]
                    ],
                },
                "product_gap_candidates": {
                    "rows": len(briefing_product_gap_candidates),
                    "by_review_status": top_counts(Counter(item.get("review_status") or "Unknown" for item in briefing_product_gap_candidates), 8),
                    "preview": briefing_product_gap_candidates[:30],
                },
                "events_path": str(BRIEFING_VERIFIED_UPDATE_EVENTS_PATH),
                "fulltext_path": str(BRIEFING_FULLTEXT_RESCUE_PATH),
                "product_gap_path": str(BRIEFING_PRODUCT_GAP_CANDIDATES_PATH),
                "summary_path": str(DATA_DIR / "audits" / "briefing_update_pipeline_summary_latest.md"),
                "note": "Briefing news now runs through a closed-loop lane: discover, rescue missing body text, verify against official sources, map to product/company masters, promote verified facts, and surface unmapped products as master-data gaps.",
            },
            "evidence_promotion": {
                **promotion_summary,
                "log_path": str(EVIDENCE_PROMOTION_LOG_PATH),
            },
            "data_usability": {
                "summary": data_usability_summary,
                "ledger_preview": sorted(
                    data_usability_ledger,
                    key=lambda row: safe_int(row.get("planned_or_review_rows")),
                    reverse=True,
                )[:24],
                "missing_owner_preview": data_usability_missing_owner[:30],
                "ledger_path": str(DATA_USABILITY_LEDGER_PATH),
                "row_status_path": str(DATA_USABILITY_ROW_STATUS_PATH),
                "missing_owner_path": str(DATA_USABILITY_MISSING_OWNER_PATH),
                "note": "Operational ledger for the usable-data goal: each audited row is either usable/reference data, excluded/noise, or assigned to a responsible module with a next action.",
            },
        },
        "data_quality": {
            "summary": quality_summary,
            "top_issues": quality_summary.get("top_issues", []),
            "issue_types": quality_summary.get("by_type", []),
            "severity": quality_summary.get("by_severity", []),
            "report_path": str(DATA_QUALITY_REPORT_PATH),
            "issues_path": str(DATA_QUALITY_ISSUES_PATH),
            "note": "Data quality issues are audit findings only. The source workbook is not modified by this build.",
        },
        "registration_evidence": {
            "seed_rows": len(registration_seed),
            "official_api_rows": staged_registration_count,
            "promoted_rows": promotion_summary.get("registration_rows_promoted", 0),
            "manual_official_rows": len(manual_official_indication_rows),
            "manual_nmpa_rows": len(manual_nmpa_registration_rows),
            "official_indication_rows": official_indication_analysis.get("rows", 0),
            "review_ready_rows": sum(1 for item in registration_seed if item.get("review_status") == "needs_review") + staged_registration_count + len(promoted_registration_rows),
            "jurisdictions": top_counts(
                Counter(item.get("jurisdiction") or "Unknown" for item in registration_seed)
                + Counter(item.get("jurisdiction") or "Unknown" for item in staging_records if item.get("merge_target") == "registration_evidence")
                + Counter(item.get("jurisdiction") or "Unknown" for item in promoted_registration_rows),
                10,
            ),
            "timeline": [
                {
                    "year": str(year),
                    "total": registration_years[year],
                    "fda": registration_years[year],
                    "ce": 0,
                    "nmpa": 0,
                    "launch": 0,
                }
                for year in sorted(registration_years)[-10:]
            ],
            "source_note": "Rows here are long-form evidence records: product x jurisdiction x regulator x registration number x approved indication x approval/expiry date x source.",
            "current_scope_note": "Registration_Evidence now separates seed rows, FDA/openFDA API rows, and promoted FDA/IFU/certificate/EUDAMED records. Country-level indications are read from long rows only.",
        },
        "market_snapshot": {
            "rows": len(market_snapshots),
            "pending_live_fetch": sum(1 for item in market_snapshots if item.get("snapshot_status") == "pending_live_fetch"),
            "cards": market_snapshots[:20],
            "note": "Ticker mapping is stored separately from static master data. Frontend displays valuation/market cap scale rather than share price.",
        },
        "company_financial_metrics": company_financial_metrics,
        "company_revenue_collection_plan": company_revenue_collection_plan,
        "reports": [{"title": r["title"], "path": r["path"], "segments": r["segments"], "year": r["year"]} for r in reports],
        "data_gaps": [
            "Current workbook data is treated as seed until cross-checked against official source classes.",
            "China NMPA full-market discovery stays in the separate China dashboard; this global dashboard imports only verified NMPA rows that map to existing global products.",
            "Registration status must be read from regulator-sourced Registration_Evidence long rows, not single wide status columns.",
            "MDSAP is tracked as quality-system audit evidence, not direct sales authorization; participating markets still require their own ARTG/ANVISA/MDL/PMDA/FDA checks.",
            "Official indication heatmaps come from Registration_Evidence long rows; unverified marketing/product-text indication signals stay separate from official approved indications.",
            "Dynamic market fields are not written back to the static Excel; use Market_Snapshot with source timestamps and valuation source notes.",
            "Briefing daily news runs as discover -> fulltext rescue -> official verification -> master mapping -> promotion or gap queue; unverified rows remain in review.",
        ],
    }
    return snapshot


def write_snapshot(snapshot: dict[str, Any]) -> None:
    WEB_DIR.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(snapshot, ensure_ascii=False, indent=2)
    SNAPSHOT_PATH.write_text(f"window.GLOBAL_AESTHETICS_DATA = {payload};\n", encoding="utf-8")
    MANIFEST_PATH.write_text(
        json.dumps(
            {
                "generated_at": snapshot["generated_at"],
                "db_path": str(DB_PATH),
                "snapshot_path": str(SNAPSHOT_PATH),
                "summary": snapshot["summary"],
                "source_root": str(SOURCE_DIR),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def main() -> None:
    company_book = find_file("*标准化版v4.xlsx")
    conference_book = find_file("医美行业会议信息.xlsx")
    congress_book = find_file("Global Major Congress 2025.xlsx")
    products = load_products(company_book)
    companies = load_companies(company_book)
    brands = load_brands(company_book)
    conferences = load_conferences(conference_book, congress_book)
    metrics = load_market_metrics()
    reports = load_reports()
    socials = social_status()
    quality = build_data_quality(products, companies, brands)
    dashboard_products, dashboard_companies, dashboard_brands, dashboard_scope = apply_dashboard_scope(products, companies, brands)
    create_database(
        dashboard_products,
        dashboard_companies,
        dashboard_brands,
        conferences,
        metrics,
        reports,
        socials,
        quality,
    )
    snapshot = build_snapshot(
        dashboard_products,
        dashboard_companies,
        dashboard_brands,
        conferences,
        metrics,
        reports,
        socials,
        quality,
    )
    snapshot["dashboard_scope"] = dashboard_scope
    snapshot["summary"]["raw_products"] = dashboard_scope["raw_products"]
    snapshot["summary"]["raw_companies"] = dashboard_scope["raw_companies"]
    snapshot["summary"]["dashboard_excluded_products"] = dashboard_scope["excluded_products"]
    snapshot["summary"]["dashboard_excluded_companies"] = dashboard_scope["excluded_companies"]
    write_snapshot(snapshot)
    write_data_quality_outputs(quality)
    print(json.dumps(snapshot["summary"], ensure_ascii=False, indent=2))
    print(f"Wrote {DB_PATH}")
    print(f"Wrote {SNAPSHOT_PATH}")


if __name__ == "__main__":
    main()
