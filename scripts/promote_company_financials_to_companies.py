#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = PROJECT_DIR.parent / "全球医美企业库_标准化版v4.xlsx"
DATA_DIR = PROJECT_DIR / "data"
AUDIT_DIR = DATA_DIR / "audits"

MARKET_VALUATION_RANK_PATH = DATA_DIR / "market_valuation_rank.csv"
MARKET_SNAPSHOT_PATH = DATA_DIR / "market_snapshot_live.csv"
LISTED_COMPANY_BATCH_PATH = DATA_DIR / "listed_company_batch.csv"
SEC_FINANCIAL_METRICS_PATH = DATA_DIR / "company_financial_metrics.csv"
MANUAL_NON_US_FINANCIAL_METRICS_PATH = DATA_DIR / "manual_non_us_financial_metrics.csv"

AUDIT_CSV = AUDIT_DIR / "company_financial_promote_latest.csv"
AUDIT_MD = AUDIT_DIR / "company_financial_promote_latest.md"


PROMOTED_FIELDS = [
    "Revenue_USD_M",
    "Revenue_Year",
    "Revenue_Growth_Pct",
    "Gross_Margin_Pct",
    "Market_Cap_USD_M",
    "Market_Cap_Date",
    "PS_Ratio",
    "PE_Ratio",
    "PB_Ratio",
    "EPS",
    "EPS_TTM",
    "Net_Profit_Growth_Pct",
    "Aesthetics_Revenue_Pct",
    "Market_Price",
    "Market_Currency",
    "Market_Day_Change_Pct",
    "Market_Source_URL",
    "Market_Captured_At",
    "Market_Refreshed_At",
    "Financial_Period",
    "Filing_Date",
    "Metric_Basis",
    "Financial_Refreshed_At",
    "Financial_Source_URL",
    "Financial_Review_Status",
]


FIELD_LABELS = {
    "Revenue_USD_M": "revenue_usd_m",
    "Revenue_Year": "revenue_year",
    "Gross_Margin_Pct": "gross_margin_pct",
    "Market_Cap_USD_M": "market_cap_usd_m",
    "Market_Cap_Date": "market_cap_date",
    "PS_Ratio": "ps_ratio",
    "PE_Ratio": "pe_ratio",
    "PB_Ratio": "pb_ratio",
    "EPS": "eps",
    "EPS_TTM": "eps_ttm",
    "Net_Profit_Growth_Pct": "net_profit_growth_pct",
    "Market_Price": "market_price",
    "Market_Currency": "market_currency",
    "Market_Day_Change_Pct": "market_day_change_pct",
    "Market_Source_URL": "market_source_url",
    "Market_Captured_At": "market_captured_at",
    "Market_Refreshed_At": "market_refreshed_at",
    "Financial_Period": "financial_period",
    "Filing_Date": "filing_date",
    "Metric_Basis": "metric_basis",
    "Financial_Refreshed_At": "financial_refreshed_at",
    "Financial_Source_URL": "financial_source_url",
    "Financial_Review_Status": "financial_review_status",
}

ALWAYS_REFRESH_FIELDS = {
    "Revenue_USD_M",
    "Revenue_Year",
    "Gross_Margin_Pct",
    "Market_Cap_USD_M",
    "Market_Cap_Date",
    "PS_Ratio",
    "PE_Ratio",
    "PB_Ratio",
    "EPS",
    "EPS_TTM",
    "Net_Profit_Growth_Pct",
    "Market_Price",
    "Market_Currency",
    "Market_Day_Change_Pct",
    "Market_Source_URL",
    "Market_Captured_At",
    "Market_Refreshed_At",
    "Financial_Period",
    "Filing_Date",
    "Metric_Basis",
    "Financial_Refreshed_At",
    "Financial_Source_URL",
    "Financial_Review_Status",
}


SOURCE_PRIORITY = {
    "sec_financial_metrics": 90,
    "manual_non_us_financial_metrics": 80,
    "market_valuation_rank": 60,
    "market_snapshot_live": 55,
    "listed_company_batch": 50,
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def compact_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", norm(value).casefold())


def stock_keys(*values: Any) -> set[str]:
    keys: set[str] = set()
    for value in values:
        raw = norm(value).upper()
        if not raw:
            continue
        normalized = re.sub(r"\s+", "", raw)
        keys.add(normalized)
        if ":" in normalized:
            keys.add(normalized.split(":", 1)[1])
        if "(" in normalized:
            keys.add(normalized.split("(", 1)[0])
        if "." in normalized:
            keys.add(normalized.split(".", 1)[0])
    return {key for key in keys if key}


def is_blank(value: Any) -> bool:
    text = norm(value)
    return text in {"", "0", "0.0", "0.00", "None", "nan"}


def clean_number(value: Any) -> str:
    text = norm(value)
    if not text:
        return ""
    text = text.replace(",", "")
    return text


def as_float(value: Any) -> float | None:
    text = clean_number(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def money_m(value: Any) -> str:
    number = as_float(value)
    if number is None:
        return ""
    return f"{number:.2f}".rstrip("0").rstrip(".")


def pct_value(value: Any) -> str:
    number = as_float(value)
    if number is None:
        return ""
    return f"{number:.2f}".rstrip("0").rstrip(".")


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {norm(cell.value): index for index, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def ensure_columns(ws: openpyxl.worksheet.worksheet.Worksheet, fields: list[str]) -> dict[str, int]:
    colmap = headers(ws)
    for field in fields:
        if field not in colmap:
            col = ws.max_column + 1
            ws.cell(1, col).value = field
            colmap[field] = col
    return colmap


def cell(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, colmap: dict[str, int], field: str) -> Any:
    col = colmap.get(field)
    return ws.cell(row, col).value if col else None


def set_cell(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, colmap: dict[str, int], field: str, value: Any) -> None:
    ws.cell(row, colmap[field]).value = value


def candidate_key(row: dict[str, str], companies_by_name: dict[str, int], companies_by_stock: dict[str, int]) -> int | None:
    name = compact_key(row.get("company"))
    if name in companies_by_name:
        return companies_by_name[name]
    for key in stock_keys(row.get("stock_code"), row.get("ticker_symbol"), row.get("yahoo_symbol")):
        if key in companies_by_stock:
            return companies_by_stock[key]
    return None


def add_candidate(
    buckets: dict[int, list[dict[str, Any]]],
    row_index: int | None,
    source: str,
    row: dict[str, str],
    values: dict[str, str],
) -> None:
    if row_index is None:
        return
    cleaned = {field: norm(value) for field, value in values.items() if not is_blank(value)}
    if not cleaned:
        return
    buckets[row_index].append(
        {
            "source": source,
            "priority": SOURCE_PRIORITY[source],
            "company": norm(row.get("company")),
            "stock_code": norm(row.get("stock_code")),
            "ticker_symbol": norm(row.get("ticker_symbol")),
            "values": cleaned,
        }
    )


def collect_candidates(companies_by_name: dict[str, int], companies_by_stock: dict[str, int]) -> dict[int, list[dict[str, Any]]]:
    buckets: dict[int, list[dict[str, Any]]] = defaultdict(list)

    for row in read_csv(MARKET_VALUATION_RANK_PATH):
        add_candidate(
            buckets,
            candidate_key(row, companies_by_name, companies_by_stock),
            "market_valuation_rank",
            row,
            {
                "Market_Cap_USD_M": money_m(row.get("market_cap_usd_m")),
                "Market_Cap_Date": row.get("as_of"),
                "Financial_Source_URL": row.get("source_url"),
                "Financial_Review_Status": row.get("snapshot_status"),
            },
        )

    for row in read_csv(MARKET_SNAPSHOT_PATH):
        add_candidate(
            buckets,
            candidate_key(row, companies_by_name, companies_by_stock),
            "market_snapshot_live",
            row,
            {
                "Market_Cap_USD_M": money_m(row.get("market_cap_usd_m")),
                "Market_Cap_Date": row.get("as_of"),
                "Market_Price": clean_number(row.get("price")),
                "Market_Currency": row.get("currency"),
                "Market_Day_Change_Pct": pct_value(row.get("day_change_pct")),
                "PB_Ratio": pct_value(row.get("pb_ratio")),
                "PE_Ratio": pct_value(row.get("pe_ratio")),
                "EPS": pct_value(row.get("eps_ttm")),
                "EPS_TTM": pct_value(row.get("eps_ttm")),
                "Net_Profit_Growth_Pct": pct_value(row.get("net_profit_growth_yoy_pct")),
                "Market_Source_URL": row.get("source_url"),
                "Market_Captured_At": row.get("as_of"),
                "Market_Refreshed_At": row.get("market_refreshed_at") or row.get("as_of"),
                "Financial_Period": row.get("financial_period"),
                "Filing_Date": row.get("filing_date"),
                "Metric_Basis": row.get("metric_basis"),
                "Financial_Refreshed_At": row.get("financial_refreshed_at"),
                "Financial_Source_URL": row.get("source_url"),
                "Financial_Review_Status": row.get("snapshot_status"),
            },
        )

    for row in read_csv(LISTED_COMPANY_BATCH_PATH):
        add_candidate(
            buckets,
            candidate_key(row, companies_by_name, companies_by_stock),
            "listed_company_batch",
            row,
            {
                "Market_Price": clean_number(row.get("market_price")),
                "Market_Currency": row.get("market_currency"),
                "Market_Day_Change_Pct": pct_value(row.get("market_day_change_pct")),
                "Market_Source_URL": row.get("market_source_url"),
                "Market_Captured_At": row.get("market_captured_at"),
                "Financial_Review_Status": row.get("market_snapshot_status") or row.get("review_status"),
            },
        )

    for row in read_csv(SEC_FINANCIAL_METRICS_PATH):
        add_candidate(
            buckets,
            candidate_key(row, companies_by_name, companies_by_stock),
            "sec_financial_metrics",
            row,
            {
                "Revenue_USD_M": money_m(row.get("revenue_usd_m")),
                "Revenue_Year": row.get("fiscal_year"),
                "Gross_Margin_Pct": pct_value(row.get("gross_margin_pct")),
                "Market_Cap_USD_M": money_m(row.get("market_cap_usd_m")),
                "Market_Cap_Date": row.get("captured_at"),
                "PS_Ratio": pct_value(row.get("ps_ratio")),
                "PE_Ratio": pct_value(row.get("pe_ratio")),
                "PB_Ratio": pct_value(row.get("pb_ratio")),
                "EPS": pct_value(row.get("eps_ttm") or row.get("eps_diluted") or row.get("eps_basic")),
                "EPS_TTM": pct_value(row.get("eps_ttm") or row.get("eps_diluted") or row.get("eps_basic")),
                "Net_Profit_Growth_Pct": pct_value(row.get("net_income_growth_yoy_pct")),
                "Financial_Period": row.get("financial_period") or (f"FY{row.get('fiscal_year')}" if row.get("fiscal_year") else ""),
                "Filing_Date": row.get("filing_date") or row.get("revenue_filed"),
                "Metric_Basis": row.get("metric_basis"),
                "Financial_Refreshed_At": row.get("captured_at"),
                "Financial_Source_URL": row.get("source_url"),
                "Financial_Review_Status": row.get("review_status"),
            },
        )

    for row in read_csv(MANUAL_NON_US_FINANCIAL_METRICS_PATH):
        add_candidate(
            buckets,
            candidate_key(row, companies_by_name, companies_by_stock),
            "manual_non_us_financial_metrics",
            row,
            {
                "Revenue_USD_M": money_m(row.get("revenue_usd_m")),
                "Revenue_Year": row.get("fiscal_year"),
                "Gross_Margin_Pct": pct_value(row.get("gross_margin_pct")),
                "Market_Cap_USD_M": money_m(row.get("market_cap_usd_m")),
                "Market_Cap_Date": row.get("captured_at") or row.get("revenue_filed"),
                "PS_Ratio": pct_value(row.get("ps_ratio")),
                "PE_Ratio": pct_value(row.get("pe_ratio")),
                "PB_Ratio": pct_value(row.get("pb_ratio")),
                "EPS": pct_value(row.get("eps_ttm") or row.get("eps_diluted") or row.get("eps_basic")),
                "EPS_TTM": pct_value(row.get("eps_ttm") or row.get("eps_diluted") or row.get("eps_basic")),
                "Net_Profit_Growth_Pct": pct_value(row.get("net_income_growth_yoy_pct")),
                "Financial_Period": row.get("financial_period") or (f"FY{row.get('fiscal_year')}" if row.get("fiscal_year") else ""),
                "Filing_Date": row.get("filing_date") or row.get("revenue_filed"),
                "Metric_Basis": row.get("metric_basis") or row.get("revenue_basis"),
                "Financial_Refreshed_At": row.get("captured_at") or row.get("revenue_filed"),
                "Financial_Source_URL": row.get("source_url"),
                "Financial_Review_Status": row.get("review_status"),
            },
        )

    return buckets


def best_value(candidates: list[dict[str, Any]], field: str) -> tuple[str, str]:
    for candidate in sorted(candidates, key=lambda item: item["priority"], reverse=True):
        value = candidate["values"].get(field)
        if not is_blank(value):
            return norm(value), candidate["source"]
    return "", ""


def derive_ps_ratio(market_cap: Any, revenue: Any) -> str:
    cap = as_float(market_cap)
    rev = as_float(revenue)
    if cap is None or rev is None or rev == 0:
        return ""
    return f"{cap / rev:.2f}".rstrip("0").rstrip(".")


def backup_source() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = SOURCE_WORKBOOK.with_name(
        f"{SOURCE_WORKBOOK.stem}_backup_before_company_financials_{timestamp}{SOURCE_WORKBOOK.suffix}"
    )
    shutil.copy2(SOURCE_WORKBOOK, backup)
    return backup


def run() -> int:
    checked_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    backup = backup_source()
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK, data_only=False)
    ws = wb["Companies"]
    colmap = ensure_columns(ws, PROMOTED_FIELDS)

    companies_by_name: dict[str, int] = {}
    companies_by_stock: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        company = norm(cell(ws, row, colmap, "Company"))
        if not company:
            continue
        companies_by_name[compact_key(company)] = row
        for key in stock_keys(cell(ws, row, colmap, "Stock_Code")):
            companies_by_stock[key] = row

    candidates = collect_candidates(companies_by_name, companies_by_stock)
    audit_rows: list[dict[str, Any]] = []
    changed_fields = Counter()
    row_status = Counter()

    for row in range(2, ws.max_row + 1):
        company = norm(cell(ws, row, colmap, "Company"))
        if not company:
            continue
        row_candidates = candidates.get(row, [])
        if not row_candidates:
            continue
        changed: list[str] = []
        sources: dict[str, str] = {}
        for field in PROMOTED_FIELDS:
            if field in {"Revenue_Growth_Pct", "Aesthetics_Revenue_Pct"}:
                continue
            value, source = best_value(row_candidates, field)
            if not value:
                continue
            current = cell(ws, row, colmap, field)
            if is_blank(current) or (field in ALWAYS_REFRESH_FIELDS and norm(current) != value):
                set_cell(ws, row, colmap, field, value)
                changed.append(field)
                sources[field] = source
                changed_fields[field] += 1

        if is_blank(cell(ws, row, colmap, "PS_Ratio")):
            derived = derive_ps_ratio(cell(ws, row, colmap, "Market_Cap_USD_M"), cell(ws, row, colmap, "Revenue_USD_M"))
            if derived:
                set_cell(ws, row, colmap, "PS_Ratio", derived)
                changed.append("PS_Ratio")
                sources["PS_Ratio"] = "derived_market_cap_revenue"
                changed_fields["PS_Ratio"] += 1

        if changed:
            row_status["updated"] += 1
            audit_rows.append(
                {
                    "company": company,
                    "stock_code": norm(cell(ws, row, colmap, "Stock_Code")),
                    "changed_fields": ";".join(changed),
                    "field_sources": json.dumps(sources, ensure_ascii=False, sort_keys=True),
                    "candidate_sources": ";".join(sorted({item["source"] for item in row_candidates})),
                    "checked_at": checked_at,
                }
            )
        else:
            row_status["matched_no_blank_fields"] += 1

    coverage = {
        field: sum(
            1
            for row in range(2, ws.max_row + 1)
            if not is_blank(cell(ws, row, colmap, field))
        )
        for field in [
            "Revenue_USD_M",
            "Gross_Margin_Pct",
            "Market_Cap_USD_M",
            "PS_Ratio",
            "PE_Ratio",
            "PB_Ratio",
            "EPS_TTM",
            "Net_Profit_Growth_Pct",
            "Market_Price",
            "Market_Refreshed_At",
            "Financial_Refreshed_At",
        ]
    }
    wb.save(SOURCE_WORKBOOK)
    wb.close()

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    write_csv(
        AUDIT_CSV,
        ["company", "stock_code", "changed_fields", "field_sources", "candidate_sources", "checked_at"],
        audit_rows,
    )
    lines = [
        "# Company financial promotion",
        "",
        f"- Generated: {checked_at}",
        f"- Source workbook: `{SOURCE_WORKBOOK}`",
        f"- Backup: `{backup}`",
        f"- Updated companies: {row_status.get('updated', 0)}",
        f"- Matched companies with no blank target fields: {row_status.get('matched_no_blank_fields', 0)}",
        "",
        "## Field Fill Counts",
        "",
    ]
    for field, count in changed_fields.most_common():
        lines.append(f"- {field}: {count}")
    lines.extend(["", "## Coverage After Promotion", ""])
    for field, count in coverage.items():
        lines.append(f"- {field}: {count}")
    lines.extend(["", "## Files", "", f"- Audit CSV: `{AUDIT_CSV}`"])
    AUDIT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "source_workbook": str(SOURCE_WORKBOOK),
                "backup": str(backup),
                "updated_companies": row_status.get("updated", 0),
                "field_fill_counts": dict(changed_fields),
                "coverage_after": coverage,
                "audit_csv": str(AUDIT_CSV),
                "audit_md": str(AUDIT_MD),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
