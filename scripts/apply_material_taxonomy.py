#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_DIR / "data"
AUDIT_DIR = DATA_DIR / "audits"
DEFAULT_TAXONOMY_BOOK = Path(r"E:\shared\Downloads\医美产品材料分类手册_双语.xlsx")
DEFAULT_SOURCE_BOOK = Path(r"E:\shared\Documents\data\全球医美企业库_标准化版v4.xlsx")

TAXONOMY_FIELDS = [
    "Material_Taxonomy_L1_CN",
    "Material_Taxonomy_L2_CN",
    "Material_Taxonomy_L3_CN",
    "Material_Taxonomy_Path_CN",
    "Material_Taxonomy_Source",
    "Material_Taxonomy_Confidence",
    "Material_Taxonomy_Review_Status",
    "Material_Taxonomy_Note",
]

TAXONOMY_EXPORT_FIELDS = [
    "record_id",
    "company",
    "brand",
    "core_product",
    "old_category_l1",
    "old_category_l2",
    "old_technology_l1",
    "old_technology_l2",
    *TAXONOMY_FIELDS,
    "Inclusion_Status",
    "Material_Family",
    "Backfill_Audit",
]

MANUAL_HEADER = [
    "L1 一级类",
    "L2 二级类",
    "L3 三级 / 具体材料",
    "代表产品 / 品牌(示例)",
    "给药术式 / 形态",
    "使用场景 Setting",
    "US-FDA",
    "CN-NMPA",
    "EU-CE",
    "作用机制 / 备注",
    "source_sheet",
]


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def lower(value: Any) -> str:
    return norm(value).casefold()


def compact(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", lower(value))


def manual_value(value: Any) -> str:
    """Keep the Chinese taxonomy label from bilingual handbook cells."""
    text = norm(value)
    return text.splitlines()[0].strip() if text else ""


def has_any(text: str, terms: list[str]) -> bool:
    return any(term.casefold() in text for term in terms)


def has_word(text: str, terms: list[str]) -> bool:
    for term in terms:
        term_l = term.casefold()
        if re.search(rf"(?<![a-z0-9]){re.escape(term_l)}(?![a-z0-9])", text):
            return True
    return False


def load_manual_taxonomy(path: Path) -> tuple[list[dict[str, str]], set[tuple[str, str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    paths: set[tuple[str, str, str]] = set()
    for ws in wb.worksheets:
        if ws.title.startswith("说明") or ws.title.startswith("RF"):
            continue
        for raw in ws.iter_rows(min_row=4, values_only=True):
            if not raw or not norm(raw[0]):
                continue
            values = [manual_value(x) for x in raw[:10]]
            item = {MANUAL_HEADER[i]: values[i] if i < len(values) else "" for i in range(10)}
            item["source_sheet"] = ws.title
            rows.append(item)
            paths.add((item["L1 一级类"], item["L2 二级类"], item["L3 三级 / 具体材料"]))
    return rows, paths


def product_text(row: dict[str, str]) -> str:
    fields = [
        "Brand",
        "Core_Product",
        "Category_L1",
        "Category_L2",
        "Tech_Type_Std",
        "Tech_Type_Original",
        "Feature_Tags",
        "Introduction",
        "Verified_Product_Type_CN",
        "Market_Channel",
    ]
    return " | ".join(lower(row.get(field)) for field in fields)


def product_display(row: dict[str, str]) -> str:
    return " / ".join(x for x in [norm(row.get("Company")), norm(row.get("Brand")), norm(row.get("Core_Product"))] if x)


def result(
    l1: str,
    l2: str,
    l3: str,
    source: str,
    confidence: str,
    note: str,
    valid_paths: set[tuple[str, str, str]],
) -> dict[str, str]:
    path = (l1, l2, l3)
    in_manual = path in valid_paths
    review_status = "auto_applied" if confidence == "high" and in_manual else "needs_review"
    if not in_manual:
        review_status = "needs_review"
        note = f"{note}; classification path is outside current manual and may require taxonomy extension"
    return {
        "Material_Taxonomy_L1_CN": l1,
        "Material_Taxonomy_L2_CN": l2,
        "Material_Taxonomy_L3_CN": l3,
        "Material_Taxonomy_Path_CN": " > ".join(x for x in [l1, l2, l3] if x),
        "Material_Taxonomy_Source": source,
        "Material_Taxonomy_Confidence": confidence,
        "Material_Taxonomy_Review_Status": review_status,
        "Material_Taxonomy_Note": note,
    }


def classify(row: dict[str, str], valid_paths: set[tuple[str, str, str]]) -> dict[str, str]:
    text = product_text(row)
    category_l1 = lower(row.get("Category_L1"))
    category_l2 = lower(row.get("Category_L2"))
    tech = lower(row.get("Tech_Type_Std"))
    tech2 = lower(row.get("Tech_Type_Original"))
    brand = lower(row.get("Brand"))
    core = lower(row.get("Core_Product"))
    intro = lower(row.get("Introduction"))
    combined = " ".join([text, category_l1, category_l2, tech, tech2, brand, core, intro])
    axis = " ".join([category_l1, category_l2, tech, tech2, brand, core, lower(row.get("Verified_Product_Type_CN"))])
    material_axis = " ".join([category_l2, tech, tech2, brand, core, lower(row.get("Verified_Product_Type_CN"))])
    inject_context = "inject" in category_l1 or "mesotherapy" in category_l2 or "skin booster" in category_l2 or "dermal filler" in category_l2

    topical_context = has_any(
        combined,
        [
            "skincare",
            "skin care",
            "cosmeceutical",
            "cream",
            "serum",
            "gel",
            "mask",
            "ampoule",
            "concentrate",
            "topical",
            "post-procedure",
            "护肤",
            "面霜",
            "精华",
            "凝胶",
            "面膜",
            "外用",
            "术后修复",
        ],
    )

    if (
        "services" in category_l1
        or "training" in category_l2
        or "education" in category_l2
        or (has_any(axis, ["genetic data", "genomics", "research service", "数据服务", "基因"]) and not has_any(axis, ["exosome", "外泌体"]))
        or (has_any(axis, ["r&d", "pipeline"]) and not has_any(axis, ["exosome", "外泌体", "pdrn", "pn ", "polynucleotide"]))
    ):
        return result("未归类/待补充", "非产品服务", "科研/数据服务", "rule:out_of_scope_service", "low", "service or R&D row is outside the material taxonomy", valid_paths)

    if has_any(combined, ["visia", "skin analysis", "skin analyzer", "analysis imaging", "skin diagnostic", "皮肤分析", "肤质量化"]):
        return result("诊断/影像", "皮肤分析", "皮肤分析仪/VISIA类", "rule:diagnostics_skin_analysis", "high", "skin analysis / VISIA term", valid_paths)
    if has_any(combined, ["ai imaging", "ai skin", "face assessment", "facial assessment", "ai影像", "面部评估", "术前规划"]):
        return result("诊断/影像", "AI影像", "AI影像/面部评估", "rule:diagnostics_ai", "high", "AI imaging or face assessment term", valid_paths)
    if has_any(combined, ["vascular locator", "high-frequency ultrasound", "高频超声", "血管定位", "术中影像"]):
        return result("诊断/影像", "术中影像", "高频超声/血管定位", "rule:diagnostics_intraoperative", "high", "intraoperative ultrasound or vascular location term", valid_paths)
    if has_any(combined, ["total body photography", "3d simulation", "dermagraphix", "vectra", "intellistudio", "photography", "simulation", "影像", "三维", "3d"]):
        return result("诊断/影像", "AI影像", "AI影像/面部评估", "rule:diagnostics_imaging_simulation", "medium", "photography/3D imaging term", valid_paths)

    if has_any(axis, ["needle", "microneedle", "micro-needle", "cannula", "针头", "微针", "微通道", "套管"]):
        return result("耗材/器械", "针具", "注射针/钝头套管", "rule:consumable_needle_cannula", "high", "needle/cannula term", valid_paths)
    if has_any(axis, ["injector", "injection device", "注射枪", "给药设备"]):
        return result("耗材/器械", "给药设备", "电子注射枪", "rule:consumable_injection_device", "high", "injection-assist device term", valid_paths)
    if has_any(axis, ["radiofrequency", " radio frequency", " rf ", "射频"]):
        return result("能量设备", "射频 RF", "聚焦射频", "rule:ebd_radiofrequency", "high", "radiofrequency term", valid_paths)
    if has_any(axis, ["cooling", "cryotherapy", "cryo", "冷疗", "冷却"]):
        return result("能量设备", "温控/其他", "其他能量设备（待补充）", "rule:ebd_cooling", "medium", "precision cooling term", valid_paths)
    if has_any(axis, ["peel", "gaba", "nana", "chemical peel", "焕肤", "刷酸"]):
        return result("功效性护肤品", "化学焕肤", "果酸/水杨酸/TCA等", "rule:skincare_peel", "high", "chemical peel or neurocosmetic peel term", valid_paths)
    if has_any(axis, ["exosome", "外泌体"]):
        return result("注射类", "再生材料", "外泌体", "rule:injectable_exosome", "high", "exosome injectable term", valid_paths)
    if topical_context and has_any(combined, ["botox-like", "botox like", "类肉毒", "类似肉毒", "涂抹式肉毒", "不含真正的肉毒", "hair therapy"]):
        return result("功效性护肤品", "医学护肤活性", "功效活性成分", "rule:skincare_botox_like_not_toxin", "high", "botox-like marketing claim is not botulinum toxin", valid_paths)

    if has_any(combined, ["botulinum toxin b", "rimabotulinum", "myobloc", "b型肉毒"]):
        return result("肉毒毒素", "B型肉毒毒素", "B型", "rule:botulinum_b", "high", "type B botulinum toxin term", valid_paths)
    if has_any(
        combined,
        [
            "botulinum",
            "botox",
            "dysport",
            "xeomin",
            "jeuveau",
            "daxxify",
            "nabota",
            "letibotulinum",
            "botulax",
            "innotox",
            "relatox",
            "rentox",
            "肉毒",
        ],
    ):
        l3 = "液态制剂" if has_any(combined, ["liquid", "ready-to-use", "ready to use", "innotox", "液态", "即用"]) else "冻干粉制剂"
        return result("肉毒毒素", "A型肉毒毒素", l3, "rule:botulinum_a", "high", "type A botulinum toxin term", valid_paths)

    if has_word(axis, ["cannula", "cannulas", "needle", "needles", "syringe", "acupuncture"]) or has_any(axis, ["套管", "针具", "注射针", "钝头针"]):
        return result("耗材/器械", "针具", "注射针/钝头套管", "rule:consumable_needle_cannula", "high", "needle/cannula term", valid_paths)

    if has_any(combined, ["thread", "thread lift", "pdo", "埋线", "线雕", "线材", "silhouette instalift", "aptos"]):
        if has_any(combined, ["silicone", "pet", "polypropylene", "non absorb", "non-absorb", "永久", "不可吸收", "spring thread"]):
            return result("埋线提升", "不可吸收线", "硅胶/聚丙烯等", "rule:thread_nonabsorbable", "high", "thread-lift row with non-absorbable material", valid_paths)
        if has_any(combined, ["pcl", "polycaprolactone", "聚己内酯"]):
            return result("埋线提升", "可吸收线", "PCL 线", "rule:thread_pcl", "high", "thread-lift row with PCL material", valid_paths)
        if has_any(combined, ["plla", "pla", "聚乳酸", "左旋"]):
            return result("埋线提升", "可吸收线", "PLLA/PLA 线", "rule:thread_plla", "high", "thread-lift row with PLA/PLLA material", valid_paths)
        return result("埋线提升", "可吸收线", "PDO 线", "rule:thread_default_absorbable", "medium", "thread-lift row without explicit material; PDO used as provisional default", valid_paths)

    if has_any(combined, ["breast implant", "mammary implant", "乳房假体", "乳房植入"]):
        if has_any(combined, ["saline", "盐水"]):
            return result("植入物", "乳房假体", "盐水假体", "rule:implant_breast_saline", "high", "saline breast implant term", valid_paths)
        return result("植入物", "乳房假体", "硅胶假体", "rule:implant_breast_silicone", "high", "breast implant term", valid_paths)
    if has_any(combined, ["facial implant", "chin implant", "鼻假体", "下颌假体", "面部假体"]):
        if has_any(combined, ["ptfe", "gore-tex", "膨体"]):
            return result("植入物", "面部假体", "膨体 PTFE", "rule:implant_face_ptfe", "high", "PTFE facial implant term", valid_paths)
        return result("植入物", "面部假体", "硅胶假体", "rule:implant_face_silicone", "medium", "facial implant term without explicit material", valid_paths)
    if has_word(combined, ["adm"]) or has_any(combined, ["acellular dermal matrix", "dermal matrix", "tissue matrix", "soft tissue scaffold", "soft tissue reinforcement", "脱细胞", "组织基质", "异体真皮", "软组织补片"]):
        return result("植入物", "组织基质/补片", "ADM/异体真皮等", "rule:implant_adm_matrix", "high", "ADM or soft-tissue matrix implant term", valid_paths)
    if has_any(combined, ["bone scaffold", "bone filler", "bone graft", "bone substitute", "骨修复", "骨填充"]):
        return result("植入物", "骨修复", "骨填充材料", "rule:implant_bone", "high", "bone repair or filler material term", valid_paths)
    if has_any(combined, ["cartilage", "软骨"]):
        return result("植入物", "软骨", "自体/异体/合成软骨", "rule:implant_cartilage", "medium", "cartilage term", valid_paths)

    if has_any(combined, ["dressing", "hydrocolloid", "silicone sheet", "anesthetic", "lidocaine cream", "表麻", "敷料", "术后修复", "疤痕贴"]) and not has_any(axis, ["led", "phototherapy", "laser", "rf", "radiofrequency", "hifu", "激光", "射频"]):
        return result("耗材/器械", "辅助", "表麻/敷料/术后修复", "rule:consumable_auxiliary", "high", "dressing/anesthetic/post-care consumable term", valid_paths)
    if has_any(axis, ["laser arm", "guide arm", "light guide", "handpiece", "accessory", "component", "导光臂", "配件", "组件"]):
        return result("耗材/器械", "设备配件/组件", "导光臂/配件等", "rule:consumable_device_component", "high", "device component or handpiece accessory term", valid_paths)
    if (
        has_any(axis, ["liposuction", "liposculpture", "lipoaspiration", "infiltration pump", "surgical instrument", "surgical scissors", "vibrasat", "power assisted liposuction", "吸脂", "浸润泵", "剪刀", "手术器械", "震动吸脂"])
        or has_word(axis, ["pal"])
    ):
        return result("耗材/器械", "手术器械/吸脂器械", "吸脂/手术器械", "rule:consumable_surgical_lipo", "high", "liposuction or surgical instrument term", valid_paths)
    if has_any(axis, ["mesogun", "electroporation", "needle-free", "needle free", "drug delivery", "carboxytherapy", "无针", "电穿孔", "给药", "注射笔", "co2 therapy"]):
        return result("耗材/器械", "给药设备", "无针注射/电穿孔", "rule:consumable_delivery_device", "high", "delivery-device or mesogun term", valid_paths)
    skin_cleansing_signal = has_any(
        combined,
        [
            "skin management",
            "hydrafacial",
            "hydradermabrasion",
            "hydrodermabrasion",
            "oxygen infusion",
            "aqua peel",
            "hydro peel",
            "hydropeel",
            "小气泡",
            "水氧",
            "水飞梭",
            "皮肤清洁",
        ],
    )
    if skin_cleansing_signal:
        companion_signal = has_any(
            axis,
            [
                "booster",
                "boosters",
                "serum",
                "essence",
                "solution",
                "solutions",
                "ampoule",
                "concentrate",
                "精华",
                "精华液",
                "配套液",
                "导入液",
            ],
        )
        consumable_signal = has_any(axis, ["tip", "tips", "consumable", "consumables", "耗材头", "耗材"])
        device_signal = has_any(
            axis,
            [
                "device",
                "system",
                "platform",
                "machine",
                "apparatus",
                "hydradermabrasion",
                "hydrodermabrasion",
                "aqua peel",
                "hydro peel",
                "hydropeel",
                "oxygen infusion",
                "设备",
                "仪",
                "系统",
                "平台",
                "小气泡",
                "水氧",
                "水飞梭",
            ],
        )
        if consumable_signal:
            return result(
                "耗材/器械",
                "皮肤清洁/护理设备",
                "清洁耗材头/精华液",
                "rule:skin_cleansing_consumable",
                "high",
                "skin-cleansing tips or solution consumable under the new consumables/devices class",
                valid_paths,
            )
        if companion_signal and not device_signal:
            return result(
                "功效性护肤品",
                "医学护肤活性",
                "功效活性成分",
                "rule:skin_cleansing_companion_essence",
                "medium",
                "small-bubble companion essence/booster; user split as cosmeceutical companion product",
                valid_paths,
            )
        return result(
            "耗材/器械",
            "皮肤清洁/护理设备",
            "气泡/水氧",
            "rule:skin_cleansing_device",
            "high",
            "small-bubble/HydraFacial-type cleansing device under consumables/devices",
            valid_paths,
        )
    if has_any(
        axis,
        [
            "multi-platform",
            "multi platform",
            "multi-technology platform",
            "multi technology platform",
            "multi-modality platform",
            "modular laser",
            "aesthetic workstation",
            "lase-station",
            "laser platform",
            "body contouring platform",
            "robotic platform",
            "oxygeneo",
            "geneo x",
            "glo2facial",
            "etherea-mx",
            "joule x",
            "v-series",
            "多功能平台",
            "多技术",
            "集成平台",
            "平台机",
            "全能激光平台",
        ],
    ):
        return result("能量设备", "多功能平台", "集成式平台", "rule:ebd_integrated_platform", "high", "integrated or multi-technology aesthetic platform", valid_paths)
    if has_any(combined, ["chemical peel", "peel", "tca", "salicylic", "glycolic", "mandelic", "prx-t33", "果酸", "水杨酸", "焕肤", "换肤"]) and not has_any(axis, ["laser", "激光"]):
        return result("功效性护肤品", "化学焕肤", "果酸/水杨酸/TCA等", "rule:topical_chemical_peel", "high", "chemical peel/acids term", valid_paths)
    if has_any(combined, ["bimatoprost", "hydroquinone", "tretinoin", "azelaic", "外用处方", "处方药"]):
        return result("功效性护肤品", "医学护肤活性", "功效活性成分", "rule:topical_prescription", "medium", "topical prescription/active skincare term", valid_paths)
    if topical_context and not inject_context and not has_any(axis, ["device", "system", "platform", "led", "phototherapy", "light therapy", "仪", "设备", "系统", "光疗"]):
        return result("功效性护肤品", "医学护肤活性", "功效活性成分", "rule:topical_cosmeceutical", "medium", "topical/skincare commercial context", valid_paths)

    if "ebd" in category_l1 and has_any(combined, ["renuvion", "j-plasma", "rf plasma", "plasma", "射频等离子", "等离子"]):
        if has_any(combined, ["renuvion", "j-plasma", "rf plasma", "射频等离子"]):
            return result("能量设备", "等离子", "射频等离子(Renuvion类)", "rule:ebd_rf_plasma", "high", "RF plasma term", valid_paths)
        return result("能量设备", "等离子", "等离子束/等离子笔", "rule:ebd_plasma", "high", "plasma EBD row; avoid HA context leakage from paired-injection wording", valid_paths)
    if has_any(combined, ["pmma", "polymethylmethacrylate", "bellafill", "聚甲基丙烯酸"]):
        return result("注射类", "胶原刺激剂", "PMMA", "rule:biostimulator_pmma", "high", "PMMA term", valid_paths)
    caha_terms = ["caha", "calcium hydroxylapatite", "calcium hydroxyapatite", "微晶瓷", "羟基磷灰石"]
    caha_specific_text = [
        "calcium hydroxylapatite filler",
        "calcium hydroxyapatite filler",
        "1% caha",
        "1 % caha",
        "1% calcium hydroxyapatite",
        "1% calcium hydroxylapatite",
        "ha + caha",
        "peg-ha + caha",
        "ha/caha",
    ]
    if has_any(material_axis, caha_terms) or has_any(combined, caha_specific_text):
        return result("注射类", "胶原刺激剂", "CaHA 微晶瓷", "rule:biostimulator_caha", "high", "CaHA term", valid_paths)
    if has_any(combined, ["pdlla", "poly-d,l-lactic", "poly d,l lactic", "消旋"]):
        return result("注射类", "胶原刺激剂", "PLA→PDLLA(消旋)", "rule:biostimulator_pdlla", "high", "PDLLA term", valid_paths)
    if has_any(combined, ["plla", "poly-l-lactic", "sculptra", "lanluma", "juvelook", "聚左旋乳酸", "左旋"]):
        return result("注射类", "胶原刺激剂", "PLA→PLLA(左旋)", "rule:biostimulator_plla", "high", "PLLA term", valid_paths)
    if has_any(combined, ["bravity", "topical serum", "topical booster", "ampoule", "涂抹", "精华", "安瓶"]):
        return result(
            "功效性护肤品",
            "医学护肤活性",
            "功效活性成分",
            "rule:topical_skincare_before_pcl",
            "high",
            "topical serum/ampoule form takes precedence over PCL material mention",
            valid_paths,
        )
    if has_any(combined, ["pcl", "polycaprolactone", "ellanse", "ellansé", "聚己内酯"]):
        return result("注射类", "胶原刺激剂", "PCL 聚己内酯", "rule:biostimulator_pcl", "high", "PCL term", valid_paths)

    if has_any(combined, ["hyaluronidase", "透明质酸酶"]):
        return result("注射类", "透明质酸 HA", "透明质酸酶", "rule:ha_hyaluronidase", "high", "hyaluronidase term", valid_paths)
    if has_any(combined, ["hyaluronic", " hyal ", "ha filler", "dermalax", "juvederm", "jvederm", "restylane", "teosyal", "belotero", "yvoire", "revolax", "玻尿酸", "透明质酸"]):
        booster_context = has_any(
            combined,
            ["skin booster", "booster", "mesolift", "mesotherapy", "profhilo", "bioremodel", "hydrobooster", "水光", "肤质", "补水", "非交联", "微交联"],
        )
        filler_context = has_any(
            combined,
            ["dermal filler", "filler", "volum", "sub-q", "sub q", "implant plus", "deep", "lips", "body contour", "buttock", "cross-linked", "cross linked", "交联", "填充", "塑形"],
        )
        if booster_context and not filler_context:
            return result("注射类", "透明质酸 HA", "非交联HA(水光形态)", "rule:ha_skin_booster", "high", "HA with skin-booster or mesotherapy context", valid_paths)
        return result("注射类", "透明质酸 HA", "交联HA(填充剂形态)", "rule:ha_filler", "high", "HA with filler or default HA commercial context", valid_paths)

    if has_any(combined, ["exosome", "外泌体"]):
        if topical_context:
            return result("功效性护肤品", "再生外用", "外用生长因子/外泌体凝胶", "rule:topical_exosome", "high", "exosome with topical/skincare context", valid_paths)
        return result("注射类", "再生材料", "外泌体", "rule:injectable_exosome", "medium", "exosome without explicit topical context", valid_paths)
    if has_any(combined, ["pdrn", "polynucleotide", "pn ", " pn/", "多聚核苷酸", "聚核苷酸"]):
        return result("注射类", "再生材料", "PDRN/PN 多聚核苷酸", "rule:regenerative_pn_pdrn", "high", "PDRN/PN term", valid_paths)
    if has_any(combined, ["ecm", "extracellular matrix", "细胞外基质"]):
        return result("注射类", "再生材料", "ECM(注射用)", "rule:regenerative_ecm", "high", "injectable ECM term", valid_paths)
    if has_any(combined, ["silk fibroin", "fibroin", "丝素"]):
        return result("注射类", "再生材料", "丝素蛋白等新材料", "rule:regenerative_silk", "high", "silk fibroin term", valid_paths)

    if (
        has_any(material_axis, ["collagen", "胶原"])
        and "ebd" not in category_l1
        and not has_any(material_axis, ["laser", "radiofrequency", "rf", "led", "hifu", "ultrasound", "microneed", "激光", "射频", "超声", "光疗"])
    ):
        if has_any(material_axis, ["recombinant", "human", "rhcollagen", "重组", "人源"]):
            return result("注射类", "胶原蛋白", "重组人源胶原蛋白", "rule:collagen_recombinant", "high", "recombinant/human collagen term", valid_paths)
        if has_any(material_axis, ["bovine", "porcine", "fish", "animal", "牛", "猪", "鱼", "动物源"]):
            return result("注射类", "胶原蛋白", "动物源胶原蛋白", "rule:collagen_animal", "high", "animal-origin collagen term", valid_paths)
        return result("注射类", "胶原蛋白", "重组人源胶原蛋白", "rule:collagen_unspecified", "medium", "collagen term in material/product fields without clear origin", valid_paths)

    if has_any(combined, ["prp", "platelet-rich plasma", "富血小板血浆"]):
        return result("注射类", "自体来源", "PRP 富血小板血浆", "rule:autologous_prp", "high", "PRP term", valid_paths)
    if has_any(combined, ["prf", "platelet-rich fibrin", "富血小板纤维蛋白"]):
        return result("注射类", "自体来源", "PRF 富血小板纤维蛋白", "rule:autologous_prf", "high", "PRF term", valid_paths)
    if has_any(combined, ["svf", "nanofat", "nano fat", "autologous fat", "fat graft", "自体脂肪", "纳米脂肪"]):
        return result("注射类", "自体来源", "自体脂肪/SVF/纳米脂肪", "rule:autologous_fat", "high", "autologous fat/SVF term", valid_paths)

    if has_any(combined, ["deoxycholic", "kybella", "lipolytic", "fat dissolv", "lipolysis injection", "溶脂", "脂解"]) and "ebd" not in category_l1:
        return result("注射类", "美塑成分", "脂解剂", "rule:mesotherapy_lipolytic", "high", "lipolytic injection term", valid_paths)
    if has_any(combined, ["vitamin", "amino acid", "mineral", "trace element", "维生素", "氨基酸", "微量元素"]):
        if topical_context and not inject_context:
            return result("功效性护肤品", "医学护肤活性", "功效活性成分", "rule:topical_vitamin", "medium", "vitamin/amino-acid term with topical/skincare context", valid_paths)
        return result("注射类", "美塑成分", "维生素/氨基酸/微量元素", "rule:mesotherapy_vitamin", "medium", "vitamin/amino-acid mesotherapy term", valid_paths)
    if has_any(combined, ["growth factor", "peptide", "wharton", "cocktail", "生长因子", "多肽", "鸡尾酒"]):
        if topical_context:
            return result("功效性护肤品", "再生外用", "外用生长因子/外泌体凝胶", "rule:topical_growth_factor", "medium", "growth-factor term with topical/skincare context", valid_paths)
        return result("注射类", "美塑成分", "生长因子/多肽鸡尾酒", "rule:mesotherapy_peptide", "medium", "peptide/growth-factor mesotherapy term", valid_paths)

    if has_any(combined, ["rf microneed", "radiofrequency microneed", "morpheus8", "potenza", "vivace", "secret rf", "intracel", "射频微针"]):
        return result("能量设备", "射频 RF", "射频微针(有创)", "rule:ebd_rf_microneedling", "high", "RF microneedling term", valid_paths)
    if has_any(combined, ["focused rf", "聚焦射频"]):
        return result("能量设备", "射频 RF", "聚焦射频", "rule:ebd_focused_rf", "high", "focused RF term", valid_paths)
    if has_any(combined, ["monopolar rf", "bipolar rf", "multipolar rf", "radiofrequency", "thermage", "venus legacy", "endymed", "单极射频", "双极射频", "多极射频", "射频"]):
        return result("能量设备", "射频 RF", "单极/双极/多极射频(无创)", "rule:ebd_rf_noninvasive", "high", "RF term without microneedling context", valid_paths)

    if has_any(combined, ["hifu", "mfu-v", "microfocused ultrasound", "micro-focused ultrasound", "ultherapy", "微聚焦超声", "超声刀"]):
        if has_any(combined, ["fat", "body", "contour", "lipolysis", "减脂", "体雕"]):
            return result("能量设备", "超声 US", "聚焦超声减脂", "rule:ebd_ultrasound_body", "high", "focused ultrasound with body-contouring context", valid_paths)
        return result("能量设备", "超声 US", "微聚焦超声 HIFU/MFU", "rule:ebd_hifu_lift", "high", "HIFU/MFU lifting term", valid_paths)
    if has_any(combined, ["lipus", "ultrasound", "sonophoresis", "ultrasonic", "超声"]):
        return result("能量设备", "超声 US", "面部/体表超声护理", "rule:ebd_ultrasound_surface", "medium", "ultrasound term without HIFU/body contouring context", valid_paths)

    if has_any(combined, ["co2 laser", "co₂", "er:yag", "erbium", "ablative", "剥脱", "点阵 co2", "二氧化碳激光"]):
        return result("能量设备", "激光 Laser", "剥脱点阵(CO₂/铒)", "rule:laser_ablative", "high", "CO2/Er:YAG or ablative laser term", valid_paths)
    if has_any(combined, ["non-ablative", "non ablative", "fraxel", "thulium", "1550", "1927", "非剥脱"]):
        return result("能量设备", "激光 Laser", "非剥脱点阵", "rule:laser_nonablative", "high", "non-ablative fractional laser term", valid_paths)
    if has_any(axis, ["diode laser", "alexandrite", "soprano", "lightsheer", "hair removal", "脱毛"]):
        return result("能量设备", "激光 Laser", "脱毛激光", "rule:laser_hair_removal", "high", "hair-removal laser term", valid_paths)
    if has_any(axis, ["picosecond", "pico", "q-switched", "q switched", "nd:yag", "pdl", "vascular laser", "pigment", "tattoo", "皮秒", "调q", "色素", "血管", "纹身"]):
        return result("能量设备", "激光 Laser", "色素/血管激光(调Q/皮秒)", "rule:laser_pigment_vascular", "high", "pigment/vascular/tattoo laser term", valid_paths)
    if has_any(axis, ["laser", "激光"]):
        return result("能量设备", "激光 Laser", "色素/血管激光(调Q/皮秒)", "rule:laser_generic", "medium", "generic laser term; assigned to broad pigment/vascular bucket for review", valid_paths)

    if has_any(combined, ["led", "photobiomodulation", "pdt", "photodynamic", "light therapy", "光疗", "光动力", "红光", "蓝光"]):
        return result("能量设备", "光/IPL", "LED光疗/光动力 PDT", "rule:ebd_led_pdt", "high", "LED/PDT/light-therapy term", valid_paths)
    if has_word(axis, ["ipl", "dpl", "opt"]) or has_any(combined, ["intense pulsed light", "强脉冲光", "光子"]):
        return result("能量设备", "光/IPL", "强脉冲光 IPL/DPL/OPT", "rule:ebd_ipl", "high", "IPL/DPL/OPT term", valid_paths)

    if has_any(combined, ["renuvion", "j-plasma", "rf plasma", "射频等离子"]):
        return result("能量设备", "等离子", "射频等离子(Renuvion类)", "rule:ebd_rf_plasma", "high", "RF plasma term", valid_paths)
    if has_any(combined, ["plasma", "等离子"]):
        return result("能量设备", "等离子", "等离子束/等离子笔", "rule:ebd_plasma", "high", "plasma term", valid_paths)

    if has_any(combined, ["cryolipolysis", "coolsculpt", "cooltech", "fat freezing", "冷冻溶脂"]):
        return result("能量设备", "温控/其他", "冷冻溶脂", "rule:ebd_cryolipolysis", "high", "cryolipolysis term", valid_paths)
    if has_any(combined, ["emsculpt", "hifem", "ems", "electromagnetic", "电磁", "肌肉刺激"]):
        return result("能量设备", "温控/其他", "电磁场肌肉刺激(EMSculpt类)", "rule:ebd_ems", "high", "electromagnetic muscle stimulation term", valid_paths)
    if has_any(combined, ["shockwave", "radial shockwave", "冲击波"]):
        return result("能量设备", "温控/其他", "冲击波", "rule:ebd_shockwave", "high", "shockwave term", valid_paths)
    if has_any(combined, ["skin cooling", "cooling", "cryo 6", "cryo 7", "冷风", "冷却"]):
        return result("能量设备", "温控/其他", "术中冷却/皮肤冷却（待补充）", "rule:ebd_skin_cooling_extension", "medium", "skin cooling is not an exact L3 in current manual", valid_paths)

    if has_any(combined, ["microneedle", "microneedling", "micro-needle", "滚针", "纳米针", "微针"]):
        return result("耗材/器械", "微针类", "微针/滚针/纳米针", "rule:consumable_microneedle", "high", "non-RF microneedling term", valid_paths)
    if topical_context or "skincare" in category_l1 or "cosmeceutical" in category_l2:
        return result("功效性护肤品", "医学护肤活性", "功效活性成分", "rule:topical_cosmeceutical", "medium", "topical/skincare commercial context", valid_paths)

    if "inject" in category_l1 or "mesotherapy" in category_l2 or "mesotherapy" in tech:
        return result("注射类", "美塑成分", "生长因子/多肽鸡尾酒", "rule:injectable_mesotherapy_fallback", "low", "injectable/mesotherapy row without precise material signal", valid_paths)
    if "ebd" in category_l1:
        return result("能量设备", "温控/其他", "其他能量设备（待补充）", "rule:ebd_fallback_extension", "low", "EBD row without precise material/energy signal", valid_paths)
    if "implant" in category_l1 or "surgical" in category_l1:
        return result("植入物", "面部假体", "硅胶假体", "rule:implant_fallback", "low", "implant/surgical row without precise material signal", valid_paths)

    return result("未归类/待补充", "待人工确认", "待人工确认", "rule:unclassified", "low", "no reliable taxonomy signal found", valid_paths)


def worksheet_headers(ws) -> dict[str, int]:
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def ensure_headers(ws, fields: list[str]) -> dict[str, int]:
    headers = worksheet_headers(ws)
    next_col = ws.max_column + 1
    for field in fields:
        if field not in headers:
            ws.cell(row=1, column=next_col, value=field)
            headers[field] = next_col
            next_col += 1
    return headers


def row_to_dict(ws, row_idx: int, headers: dict[str, int]) -> dict[str, str]:
    return {field: norm(ws.cell(row=row_idx, column=col).value) for field, col in headers.items()}


def update_sheet_by_record_id(ws, assignments: dict[str, dict[str, str]], record_field: str) -> int:
    headers = ensure_headers(ws, TAXONOMY_FIELDS)
    record_col = headers.get(record_field)
    if not record_col:
        return 0
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        record_id = norm(ws.cell(row=row_idx, column=record_col).value)
        assignment = assignments.get(record_id)
        if not assignment:
            continue
        for field in TAXONOMY_FIELDS:
            ws.cell(row=row_idx, column=headers[field], value=assignment.get(field, ""))
        updated += 1
    return updated


def preserved_assignment(row: dict[str, str]) -> dict[str, str] | None:
    review_status = norm(row.get("Material_Taxonomy_Review_Status"))
    inclusion_status = norm(row.get("Inclusion_Status"))
    terminal_or_manual_statuses = {
        "accepted",
        "confirmed",
        "manual_verified",
        "pending_review",
        "pending_subclass",
    }
    should_preserve = review_status in terminal_or_manual_statuses or inclusion_status in {"excluded", "deleted"}
    if not should_preserve:
        return None
    assignment = {field: norm(row.get(field)) for field in TAXONOMY_FIELDS}
    if not assignment.get("Material_Taxonomy_Path_CN"):
        parts = [
            assignment.get("Material_Taxonomy_L1_CN"),
            assignment.get("Material_Taxonomy_L2_CN"),
            assignment.get("Material_Taxonomy_L3_CN"),
        ]
        assignment["Material_Taxonomy_Path_CN"] = " > ".join(part for part in parts if part)
    assignment["Material_Taxonomy_Source"] = assignment.get("Material_Taxonomy_Source") or "manual_backfill"
    assignment["Material_Taxonomy_Confidence"] = assignment.get("Material_Taxonomy_Confidence") or "manual"
    assignment["Material_Taxonomy_Review_Status"] = review_status or "confirmed"
    return assignment


def replace_taxonomy_definition_sheet(wb, manual_rows: list[dict[str, str]]) -> None:
    sheet_name = "Material_Taxonomy_Definitions"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
    headers = MANUAL_HEADER
    ws.append(headers)
    for item in manual_rows:
        ws.append([item.get(header, "") for header in headers])
    header_fill = PatternFill("solid", fgColor="E7EEF8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [16, 20, 28, 26, 20, 18, 14, 14, 14, 42, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_taxonomy_columns(ws) -> None:
    headers = worksheet_headers(ws)
    fill = PatternFill("solid", fgColor="FFF4CC")
    for field in TAXONOMY_FIELDS:
        col = headers.get(field)
        if not col:
            continue
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col)].width = min(38, max(18, len(field) + 2))
    for field in ["Material_Taxonomy_Path_CN", "Material_Taxonomy_Note"]:
        col = headers.get(field)
        if col:
            ws.column_dimensions[get_column_letter(col)].width = 42 if "Path" in field else 64


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_html_report(rows: list[dict[str, str]], summary: dict[str, Any], path: Path) -> None:
    review_rows = [
        row
        for row in rows
        if row.get("Material_Taxonomy_Review_Status") == "needs_review"
        and row.get("Inclusion_Status", "active") not in {"excluded", "deleted"}
    ]
    bucket_map: dict[str, list[dict[str, str]]] = {}
    for row in review_rows:
        key = row.get("Material_Taxonomy_Path_CN") or "未归类/待补充 > 待人工确认 > 待人工确认"
        bucket_map.setdefault(key, []).append(row)

    buckets: list[dict[str, Any]] = []
    for index, (bucket_path, bucket_rows) in enumerate(
        sorted(bucket_map.items(), key=lambda item: (-len(item[1]), item[0])), start=1
    ):
        first = bucket_rows[0]
        buckets.append(
            {
                "bucket_id": f"B{index:03d}",
                "path": bucket_path,
                "l1": first.get("Material_Taxonomy_L1_CN", ""),
                "l2": first.get("Material_Taxonomy_L2_CN", ""),
                "l3": first.get("Material_Taxonomy_L3_CN", ""),
                "count": len(bucket_rows),
                "sources": dict(Counter(row.get("Material_Taxonomy_Source", "") for row in bucket_rows)),
                "confidence": dict(Counter(row.get("Material_Taxonomy_Confidence", "") for row in bucket_rows)),
                "old_l1": dict(Counter(row.get("old_category_l1", "") for row in bucket_rows)),
                "rows": bucket_rows,
                "record_ids": [row.get("record_id", "") for row in bucket_rows],
            }
        )

    auto_summary = [
        {"path": key, "count": value}
        for key, value in Counter(
            row.get("Material_Taxonomy_Path_CN", "未填写路径")
            for row in rows
            if row.get("Material_Taxonomy_Review_Status") == "auto_applied"
        ).most_common(25)
    ]
    paths = sorted({row.get("Material_Taxonomy_Path_CN", "") for row in rows if row.get("Material_Taxonomy_Path_CN")})
    paths.extend(
        [
            "耗材/器械 > 皮肤清洁/护理设备 > 气泡/水氧",
            "耗材/器械 > 皮肤清洁/护理设备 > 清洁耗材头/精华液",
            "功效性护肤品 > 医学护肤活性 > 功效活性成分",
            "功效性护肤品 > 化学焕肤 > 果酸/水杨酸/TCA等",
        ]
    )
    paths = sorted(dict.fromkeys(paths))
    payload = {
        "summary": {
            **summary,
            "bucket_count": len(buckets),
            "needs_review_rows": len(review_rows),
            "auto_rows": len(rows) - len(review_rows),
        },
        "buckets": buckets,
        "auto_summary": auto_summary,
        "paths": paths,
    }
    payload_json = json.dumps(payload, ensure_ascii=False).replace("<", "\\u003c")
    doc = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>材料主轴分类复核</title>
  <style>
    :root { --bg:#eef2f6; --panel:#fff; --text:#172033; --muted:#5f6b7a; --line:#d9e0e8; --accent:#176b87; --soft:#f5f7fa; --warn:#996d12; --ok:#166534; --hold:#5947a3; --danger:#9f2f2f; }
    * { box-sizing:border-box; }
    body { margin:0; background:var(--bg); color:var(--text); font:14px/1.6 "Microsoft YaHei","PingFang SC","Segoe UI",Arial,sans-serif; }
    header { padding:22px 28px 16px; background:#fff; border-bottom:1px solid var(--line); }
    h1 { margin:0 0 8px; font-size:28px; letter-spacing:0; }
    p { margin:0; color:var(--muted); }
    .stats { display:flex; flex-wrap:wrap; gap:10px; margin-top:16px; }
    .stat { min-width:130px; padding:10px 12px; border:1px solid var(--line); border-radius:8px; background:var(--soft); }
    .stat b { display:block; font-size:22px; line-height:1.1; }
    .stat span { color:var(--muted); font-size:12px; }
    .notice { margin:14px 28px 0; padding:12px 14px; border:1px solid #f0d38b; background:#fff7df; border-radius:8px; color:#513a05; }
    .toolbar { position:sticky; top:0; z-index:4; display:grid; grid-template-columns:1.2fr .7fr .7fr auto auto; gap:10px; padding:14px 28px; background:#fff; border-bottom:1px solid var(--line); }
    input, select, textarea, button { min-height:38px; border:1px solid var(--line); border-radius:7px; background:#fff; color:var(--text); font:inherit; }
    input, select, textarea { width:100%; padding:8px 10px; }
    textarea { min-height:84px; resize:vertical; }
    button { padding:8px 12px; cursor:pointer; font-weight:700; white-space:nowrap; }
    button.primary { background:var(--accent); border-color:var(--accent); color:#fff; }
    .layout { display:grid; grid-template-columns:380px minmax(0,1fr); gap:16px; padding:16px 28px 28px; }
    .bucket-list { display:grid; gap:10px; }
    .card, .panel { background:var(--panel); border:1px solid var(--line); border-radius:8px; }
    .card { padding:12px; cursor:pointer; }
    .card.active { border-color:var(--accent); box-shadow:0 0 0 2px #e7f3f6; }
    .card.done { border-color:#9bd3aa; background:#fbfffc; }
    .path { font-weight:800; line-height:1.45; }
    .meta { margin-top:7px; color:var(--muted); font-size:12px; line-height:1.5; }
    .badge { display:inline-flex; border-radius:999px; padding:3px 8px; font-size:12px; font-weight:700; }
    .review { color:var(--warn); background:#fff5d9; }
    .done { color:var(--ok); background:#e9f7ef; }
    .hold { color:var(--hold); background:#efedff; }
    .exclude { color:var(--danger); background:#fff0f0; }
    .panel { overflow:hidden; }
    .panel-head, .section { padding:16px; border-bottom:1px solid var(--line); }
    .panel-head h2 { margin:0 0 8px; font-size:22px; letter-spacing:0; }
    .decision { display:grid; grid-template-columns:1fr 2fr 1fr; gap:10px; padding:16px; background:var(--soft); border-bottom:1px solid var(--line); }
    .full { grid-column:1 / -1; }
    label { display:block; margin-bottom:6px; color:var(--muted); font-size:12px; font-weight:700; }
    .chips { display:flex; flex-wrap:wrap; gap:8px; }
    .chip { border:1px solid var(--line); border-radius:999px; padding:5px 9px; color:var(--muted); font-size:12px; background:#fff; }
    table { width:100%; border-collapse:collapse; font-size:13px; }
    th, td { padding:9px 8px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; }
    th { color:var(--muted); background:#fafbfd; font-size:12px; }
    td strong { display:block; }
    .empty { padding:24px; border:1px dashed var(--line); border-radius:8px; background:#fff; text-align:center; color:var(--muted); }
    details { margin-top:16px; padding:14px; }
    details summary { cursor:pointer; font-weight:800; }
    @media (max-width:1100px) { .toolbar,.layout,.decision { grid-template-columns:1fr; } .full { grid-column:auto; } }
  </style>
</head>
<body>
  <header>
    <h1>材料主轴分类：按分类口径批量复核</h1>
    <p>自动应用的项目已经收起。你只需要先判断左侧这些分类口径；本组有例外时再单独标注。</p>
    <div class="stats">
      <div class="stat"><b id="bucketCount">0</b><span>待判断问题组</span></div>
      <div class="stat"><b id="reviewCount">0</b><span>需复核产品行</span></div>
      <div class="stat"><b id="autoCount">0</b><span>已自动应用</span></div>
      <div class="stat"><b id="doneCount">0</b><span>已处理问题组</span></div>
    </div>
  </header>
  <div class="notice"><b>你要做什么：</b>不用逐条过。先判断每个问题组的分类口径；小气泡设备本体选“耗材/器械 > 皮肤清洁/护理设备 > 气泡/水氧”，精华、Booster、配套液这类产品选“功效性护肤品 > 医学护肤活性 > 功效活性成分”。只有少数产品例外时，才在下方逐行说明。</div>
  <div class="toolbar">
    <input id="search" placeholder="搜索公司、品牌、产品、旧分类、建议路径">
    <select id="decisionFilter"><option value="">全部处理状态</option><option value="undecided">未处理</option><option value="accept">已采纳</option><option value="change">改路径</option><option value="hold">暂缓</option><option value="exclude">排除</option></select>
    <select id="l1Filter"><option value="">全部一级</option></select>
    <button class="primary" id="exportJson">导出复核意见 JSON</button>
    <button id="exportCsv">导出 CSV</button>
  </div>
  <main class="layout">
    <aside><div id="bucketList" class="bucket-list"></div></aside>
    <section>
      <div id="detail" class="panel"></div>
      <details class="panel">
        <summary>自动应用结果抽查摘要</summary>
        <div id="autoSummary" style="margin-top:12px"></div>
      </details>
    </section>
  </main>
  <datalist id="pathOptions"></datalist>
  <script>
    const DATA = __PAYLOAD__;
    const SUMMARY = DATA.summary;
    const BUCKETS = DATA.buckets;
    const STORAGE_KEY = "materialTaxonomyReviewOneTable.v1";
    let state = JSON.parse(localStorage.getItem(STORAGE_KEY) || '{"buckets":{},"rows":{}}');
    let active = BUCKETS[0]?.bucket_id || "";
    const $ = id => document.getElementById(id);
    function esc(v) { return String(v || "").replace(/[&<>"']/g, s => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[s])); }
    function save() { localStorage.setItem(STORAGE_KEY, JSON.stringify(state)); updateStats(); }
    function mix(obj) { return Object.entries(obj || {}).filter(([k]) => k).map(([k,v]) => `${k} (${v})`).join(" / "); }
    function displayPath(v) {
      const s = String(v || "");
      if (s.startsWith("功效性护肤品 > 皮肤清洁类设备")) return "耗材/器械 > 皮肤清洁/护理设备 > 气泡/水氧";
      if (s.startsWith("外用/护肤 > 化学焕肤")) return s.replace("外用/护肤 > 化学焕肤", "功效性护肤品 > 化学焕肤");
      if (s === "外用/护肤 > 医学护肤 > 活性成分护肤") return "功效性护肤品 > 医学护肤活性 > 功效活性成分";
      return s.replaceAll("外用/护肤", "功效性护肤品");
    }
    function displayL1(v) { return v === "外用/护肤" ? "功效性护肤品" : (v || "未标注"); }
    function decision(id) { return state.buckets[id] || {}; }
    function label(v) { return ({accept:"已采纳",change:"改路径",hold:"暂缓",exclude:"排除"}[v] || "待判断"); }
    function cls(v) { return v === "accept" || v === "change" ? "done" : v === "hold" ? "hold" : v === "exclude" ? "exclude" : "review"; }
    function hay(bucket) {
      return [bucket.path,bucket.l1,bucket.l2,bucket.l3,mix(bucket.sources),mix(bucket.old_l1),...bucket.rows.flatMap(r => [r.company,r.brand,r.core_product,r.old_category_l1,r.old_category_l2,r.old_technology_l1,r.old_technology_l2])].join(" ").toLowerCase();
    }
    function filtered() {
      const q = $("search").value.trim().toLowerCase();
      const df = $("decisionFilter").value;
      const l1 = $("l1Filter").value;
      return BUCKETS.filter(bucket => {
        const d = decision(bucket.bucket_id).decision || "undecided";
        if (df && d !== df) return false;
        if (l1 && bucket.l1 !== l1) return false;
        if (q && !hay(bucket).includes(q)) return false;
        return true;
      });
    }
    function renderList() {
      const items = filtered();
      if (!items.some(bucket => bucket.bucket_id === active)) active = items[0]?.bucket_id || "";
      $("bucketList").innerHTML = items.map(bucket => {
        const d = decision(bucket.bucket_id).decision || "";
        return `<article class="card ${bucket.bucket_id === active ? "active" : ""} ${d ? "done" : ""}" data-id="${esc(bucket.bucket_id)}">
          <div style="display:flex;justify-content:space-between;gap:8px"><div class="path">${esc(displayPath(bucket.path))}</div><span class="badge ${cls(d)}">${label(d)}</span></div>
          <div class="meta">${bucket.count} 条产品行 · ${esc(mix(bucket.confidence) || "未标置信度")}</div>
          <div class="meta">规则来源：${esc(mix(bucket.sources) || "未标注")}</div>
          <div class="meta">旧一级：${esc(mix(bucket.old_l1) || "未标注")}</div>
        </article>`;
      }).join("") || '<div class="empty">没有匹配的问题组</div>';
      document.querySelectorAll(".card[data-id]").forEach(card => card.onclick = () => { active = card.dataset.id; render(); });
    }
    function renderDetail() {
      const bucket = BUCKETS.find(item => item.bucket_id === active);
      if (!bucket) { $("detail").innerHTML = '<div class="empty">请选择一个问题组</div>'; return; }
      const saved = decision(bucket.bucket_id);
      $("detail").innerHTML = `<div class="panel-head"><h2>${esc(displayPath(bucket.path))}</h2><p>${bucket.count} 条产品行。先判断这一组应该归到哪里；下面表格只处理少数例外。</p></div>
        <div class="decision">
          <div><label>我的判断</label><select id="bucketDecision"><option value="">待判断</option><option value="accept">同意当前建议</option><option value="change">我认为应改为右侧路径</option><option value="split">这一组需要拆分</option><option value="hold">暂时拿不准</option><option value="exclude">不是医美产品/不入主盘</option></select></div>
          <div><label>我认为应归为</label><input id="changePath" list="pathOptions" value="${esc(saved.change_to_path || "")}" placeholder="例如：耗材/器械 > 皮肤清洁/护理设备 > 气泡/水氧"></div>
          <div><label>本组记录</label><button id="copyIds" type="button">复制 record_id</button></div>
          <div class="full"><label>人工意见 / 判断口径</label><textarea id="note" placeholder="例如：小气泡设备本体走耗材/器械；精华、Booster、配套液走功效性护肤品。">${esc(saved.note || "")}</textarea></div>
        </div>
        <div class="section"><div class="chips">
          <span class="chip">当前建议一级：${esc(displayL1(bucket.l1))}</span><span class="chip">当前建议二级：${esc(bucket.l2)}</span><span class="chip">当前建议三级：${esc(bucket.l3)}</span>
          <span class="chip">规则来源：${esc(mix(bucket.sources) || "未标注")}</span><span class="chip">旧一级：${esc(mix(bucket.old_l1) || "未标注")}</span>
        </div></div>
        <div class="section"><h3>本组产品，只有例外才需要填写</h3><div style="overflow:auto"><table><thead><tr><th>产品</th><th>旧分类</th><th>技术线索</th><th>行级处理</th><th>我认为应归为</th><th>例外备注</th></tr></thead><tbody>
          ${bucket.rows.map(row => rowHtml(row)).join("")}
        </tbody></table></div></div>`;
      $("bucketDecision").value = saved.decision || "";
      $("bucketDecision").onchange = saveBucket;
      $("changePath").oninput = saveBucket;
      $("note").oninput = saveBucket;
      $("copyIds").onclick = async () => { await navigator.clipboard.writeText(bucket.record_ids.join(", ")); $("copyIds").textContent = "已复制"; setTimeout(() => $("copyIds").textContent = "复制 record_id", 1000); };
      document.querySelectorAll("[data-row-action],[data-row-path],[data-row-note]").forEach(el => el.oninput = saveRow);
      function saveBucket() {
        state.buckets[bucket.bucket_id] = { decision:$("bucketDecision").value, change_to_path:$("changePath").value.trim(), note:$("note").value.trim(), updated_at:new Date().toISOString() };
        save(); renderList();
      }
    }
    function rowHtml(row) {
      const saved = state.rows[row.record_id] || {};
      return `<tr><td><strong>${esc(row.brand || row.core_product || row.record_id)}</strong><div class="meta">${esc(row.company)}</div><div class="meta">${esc(row.core_product)}</div><div class="meta">${esc(row.record_id)}</div></td>
        <td>${esc([row.old_category_l1,row.old_category_l2].filter(Boolean).join(" > ") || "未标注")}</td>
        <td>${esc([row.old_technology_l1,row.old_technology_l2].filter(Boolean).join(" > ") || "未标注")}</td>
        <td><select data-row-action="${esc(row.record_id)}"><option value="" ${!saved.action ? "selected" : ""}>随本组处理</option><option value="change" ${saved.action === "change" ? "selected" : ""}>此行改路径</option><option value="hold" ${saved.action === "hold" ? "selected" : ""}>此行暂缓</option><option value="exclude" ${saved.action === "exclude" ? "selected" : ""}>此行排除</option></select></td>
        <td><input data-row-path="${esc(row.record_id)}" list="pathOptions" value="${esc(saved.change_to_path || "")}" placeholder="只在例外时填写"></td>
        <td><input data-row-note="${esc(row.record_id)}" value="${esc(saved.note || "")}" placeholder="只写例外说明"></td></tr>`;
    }
    function saveRow(e) {
      const id = e.target.dataset.rowAction || e.target.dataset.rowPath || e.target.dataset.rowNote;
      const action = document.querySelector(`[data-row-action="${CSS.escape(id)}"]`)?.value || "";
      const change_to_path = document.querySelector(`[data-row-path="${CSS.escape(id)}"]`)?.value.trim() || "";
      const note = document.querySelector(`[data-row-note="${CSS.escape(id)}"]`)?.value.trim() || "";
      if (action || change_to_path || note) state.rows[id] = { action, change_to_path, note, updated_at:new Date().toISOString() }; else delete state.rows[id];
      save();
    }
    function updateStats() {
      $("bucketCount").textContent = SUMMARY.bucket_count || BUCKETS.length;
      $("reviewCount").textContent = SUMMARY.needs_review_rows || 0;
      $("autoCount").textContent = SUMMARY.auto_rows || 0;
      $("doneCount").textContent = BUCKETS.filter(bucket => decision(bucket.bucket_id).decision).length;
    }
    function renderAuto() {
      $("autoSummary").innerHTML = `<table><thead><tr><th>路径</th><th>自动应用行数</th></tr></thead><tbody>${DATA.auto_summary.map(row => `<tr><td>${esc(displayPath(row.path))}</td><td>${row.count}</td></tr>`).join("")}</tbody></table>`;
    }
    function exportPayload() { return { exported_at:new Date().toISOString(), summary:SUMMARY, bucket_decisions:state.buckets, row_exceptions:state.rows, buckets:BUCKETS.map(b => ({ bucket_id:b.bucket_id, path:b.path, count:b.count, record_ids:b.record_ids })) }; }
    function download(name, content, type) { const blob = new Blob([content], { type }); const url = URL.createObjectURL(blob); const a = document.createElement("a"); a.href = url; a.download = name; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(url); }
    function csvCell(v) { return `"${String(v || "").replace(/"/g, '""')}"`; }
    $("exportJson").onclick = () => download("material_taxonomy_review_decisions.json", JSON.stringify(exportPayload(), null, 2), "application/json;charset=utf-8");
    $("exportCsv").onclick = () => {
      const lines = [["level","bucket_id","record_id","current_path","decision","change_to_path","note","count"].join(",")];
      BUCKETS.forEach(bucket => { const saved = decision(bucket.bucket_id); lines.push(["group",bucket.bucket_id,"",displayPath(bucket.path),saved.decision || "",saved.change_to_path || "",saved.note || "",bucket.count].map(csvCell).join(",")); });
      Object.entries(state.rows).forEach(([recordId,saved]) => { const bucket = BUCKETS.find(b => b.record_ids.includes(recordId)); lines.push(["row",bucket?.bucket_id || "",recordId,displayPath(bucket?.path || ""),saved.action || "",saved.change_to_path || "",saved.note || "",""].map(csvCell).join(",")); });
      download("material_taxonomy_review_decisions.csv", "\\ufeff" + lines.join("\\n"), "text/csv;charset=utf-8");
    };
    function render() { renderList(); renderDetail(); updateStats(); }
    [...new Set(BUCKETS.map(bucket => bucket.l1).filter(Boolean))].sort().forEach(value => { const opt = document.createElement("option"); opt.value = value; opt.textContent = displayL1(value); $("l1Filter").appendChild(opt); });
    $("pathOptions").innerHTML = DATA.paths.map(path => `<option value="${esc(displayPath(path))}"></option>`).join("");
    ["search","decisionFilter","l1Filter"].forEach(id => $(id).oninput = render);
    renderAuto();
    render();
  </script>
</body>
</html>
""".replace("__PAYLOAD__", payload_json)
    path.write_text(doc, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply material-axis taxonomy to the global aesthetics source workbook.")
    parser.add_argument("--taxonomy-book", type=Path, default=DEFAULT_TAXONOMY_BOOK)
    parser.add_argument("--source-book", type=Path, default=DEFAULT_SOURCE_BOOK)
    parser.add_argument("--no-save-excel", action="store_true", help="Only write audit files; do not update the source workbook.")
    args = parser.parse_args()

    manual_rows, valid_paths = load_manual_taxonomy(args.taxonomy_book)
    wb = load_workbook(args.source_book)
    if "Product_Lines" not in wb.sheetnames:
        raise SystemExit("Product_Lines sheet not found.")
    product_ws = wb["Product_Lines"]
    headers = ensure_headers(product_ws, TAXONOMY_FIELDS)
    product_records: dict[str, dict[str, str]] = {}
    audit_rows: list[dict[str, str]] = []
    for row_idx in range(2, product_ws.max_row + 1):
        row = row_to_dict(product_ws, row_idx, headers)
        record_id = norm(row.get("Record_ID"))
        if not record_id:
            continue
        assignment = preserved_assignment(row) or classify(row, valid_paths)
        product_records[record_id] = assignment
        for field in TAXONOMY_FIELDS:
            product_ws.cell(row=row_idx, column=headers[field], value=assignment.get(field, ""))
        audit_rows.append(
            {
                "record_id": record_id,
                "company": norm(row.get("Company")),
                "brand": norm(row.get("Brand")),
                "core_product": norm(row.get("Core_Product")),
                "old_category_l1": norm(row.get("Category_L1")),
                "old_category_l2": norm(row.get("Category_L2")),
                "old_technology_l1": norm(row.get("Tech_Type_Std")),
                "old_technology_l2": norm(row.get("Tech_Type_Original")),
                **assignment,
                "Inclusion_Status": norm(row.get("Inclusion_Status")) or "active",
                "Material_Family": norm(row.get("Material_Family")),
                "Backfill_Audit": norm(row.get("Backfill_Audit")),
            }
        )
    style_taxonomy_columns(product_ws)

    sheet_updates = {"Product_Lines": len(audit_rows)}
    for sheet_name, record_field in [
        ("Product_Master", "seed_record_id"),
        ("Product_SKU_Master", "seed_record_id"),
    ]:
        if sheet_name in wb.sheetnames:
            sheet_updates[sheet_name] = update_sheet_by_record_id(wb[sheet_name], product_records, record_field)
            style_taxonomy_columns(wb[sheet_name])

    replace_taxonomy_definition_sheet(wb, manual_rows)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    latest_csv = AUDIT_DIR / "material_taxonomy_assignment_latest.csv"
    write_csv(latest_csv, TAXONOMY_EXPORT_FIELDS, audit_rows)

    review_rows = [
        row
        for row in audit_rows
        if row.get("Material_Taxonomy_Review_Status") == "needs_review"
        and row.get("Inclusion_Status", "active") not in {"excluded", "deleted"}
    ]
    latest_review_csv = AUDIT_DIR / "material_taxonomy_review_queue_latest.csv"
    write_csv(latest_review_csv, TAXONOMY_EXPORT_FIELDS, review_rows)

    summary = {
        "generated_at": generated_at,
        "source_book": str(args.source_book),
        "taxonomy_book": str(args.taxonomy_book),
        "rows": len(audit_rows),
        "manual_paths": len(valid_paths),
        "by_l1": dict(Counter(row["Material_Taxonomy_L1_CN"] for row in audit_rows)),
        "by_review_status": dict(Counter(row["Material_Taxonomy_Review_Status"] for row in audit_rows)),
        "by_confidence": dict(Counter(row["Material_Taxonomy_Confidence"] for row in audit_rows)),
        "sheet_updates": sheet_updates,
        "audit_csv": str(latest_csv),
        "review_queue_csv": str(latest_review_csv),
    }
    latest_html = AUDIT_DIR / "material_taxonomy_review_latest.html"
    write_html_report(audit_rows, summary, latest_html)
    summary["review_html"] = str(latest_html)

    summary_path = AUDIT_DIR / "material_taxonomy_summary_latest.md"
    summary_lines = [
        "# Material taxonomy assignment",
        "",
        f"- Generated: {generated_at}",
        f"- Source workbook: `{args.source_book}`",
        f"- Taxonomy manual: `{args.taxonomy_book}`",
        f"- Product rows classified: {len(audit_rows)}",
        f"- Manual taxonomy paths: {len(valid_paths)}",
        f"- Auto applied: {summary['by_review_status'].get('auto_applied', 0)}",
        f"- Needs review: {summary['by_review_status'].get('needs_review', 0)}",
        "",
        "## L1 distribution",
        "",
    ]
    for key, value in Counter(row["Material_Taxonomy_L1_CN"] for row in audit_rows).most_common():
        summary_lines.append(f"- {key}: {value}")
    summary_lines.extend(
        [
            "",
            "## Artifacts",
            "",
            f"- Assignment CSV: `{latest_csv}`",
            f"- Review queue CSV: `{latest_review_csv}`",
            f"- Review HTML: `{latest_html}`",
        ]
    )
    summary_path.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")

    if not args.no_save_excel:
        backup = args.source_book.with_name(f"{args.source_book.stem}.backup_before_material_taxonomy_{timestamp}{args.source_book.suffix}")
        shutil.copy2(args.source_book, backup)
        try:
            wb.save(args.source_book)
            summary["excel_saved"] = str(args.source_book)
            summary["backup"] = str(backup)
        except PermissionError:
            fallback = args.source_book.with_name(f"{args.source_book.stem}.material_taxonomy_{timestamp}{args.source_book.suffix}")
            wb.save(fallback)
            summary["excel_saved"] = str(fallback)
            summary["backup"] = str(backup)
            summary["warning"] = "Source workbook was locked; wrote a timestamped copy instead."

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
