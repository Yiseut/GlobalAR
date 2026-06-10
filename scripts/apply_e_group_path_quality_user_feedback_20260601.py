#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"

PRODUCT_MASTER = DATA_DIR / "product_master.csv"
MANUAL_INDICATION = DATA_DIR / "manual_official_indication_evidence.csv"
REVIEW_PACK = AUDIT_DIR / "e_group_remaining_review_pack_latest.csv"

SUMMARY_JSON = AUDIT_DIR / "e_group_path_quality_user_feedback_latest.json"
APPLIED_CSV = AUDIT_DIR / "e_group_path_quality_user_feedback_applied_latest.csv"
EXCLUDED_CSV = AUDIT_DIR / "e_group_path_quality_user_feedback_excluded_latest.csv"
DEFERRED_CSV = AUDIT_DIR / "e_group_path_quality_user_feedback_deferred_latest.csv"

EXCLUDE_SEEDS = {
    "REC_0773": "excluded_by_user_e_group_path_20260601: Biovico Xerthra/AOORA points to ophthalmology, orthopedics, wound regeneration and veterinary/product paths, outside core aesthetics upstream scope.",
    "REC_0045": "excluded_by_user_e_group_path_20260601: Contura Aquamid reconstruction evidence points to orthopaedics, animal health and urinary incontinence, outside current aesthetics scope.",
    "REC_0567": "excluded_by_user_e_group_path_20260601: Crown ProGen PRP IFU is for bone-defect graft handling, not terminal aesthetic PRP use.",
    "REC_0215": "excluded_by_user_e_group_path_20260601: Eclipse PRP Gold / ProGen PRP IFU is for bone-defect graft handling, not terminal aesthetic PRP use.",
    "REC_0310": "excluded_by_user_e_group_path_20260601: Lameditech HandyRay is a personal laser lancing/blood-sampling device, not aesthetic treatment equipment.",
    "REC_0526": "excluded_by_user_e_group_path_20260601: Mesotech Overage text indicates ophthalmic surgical aid, outside core aesthetics scope.",
}

DEFERRED_SEEDS = {
    "REC_0355": "Jalupro Super Hydro source text is truncated at indicated for; needs fresh official IFU/product page.",
    "REC_0738": "Promoitalia V-Carbon source text is truncated at indicated for improving the appearance of; needs fresh official IFU/product page.",
    "REC_0309": "Revitacare HairCare source text is truncated at intended for; needs fresh official IFU/product page.",
}

INDICATIONS = [
    {
        "seed": "REC_0849",
        "label": "Alma Lasers - Alma Duo",
        "indication": "Intended for sexual wellness applications using focused low-intensity shock wave therapy (LI-ESWT) to stimulate blood flow, support sexual function, and restore spontaneity and pleasure.",
        "status": "user_confirmed_path_keep_official_use",
    },
    {
        "seed": "REC_0859",
        "label": "Candela - CO2RE Intima",
        "indication": "Indicated for the treatment of genitourinary syndrome of menopause (GSM) and stress urinary incontinence (SUI) in adult women.",
        "status": "user_confirmed_path_keep_official_indication",
    },
    {
        "seed": "REC_0336",
        "label": "Halozyme Therapeutics - Hylenex",
        "indication": "Indicated as an adjuvant to increase the dispersion and absorption of other injected drugs.",
        "status": "user_confirmed_path_keep_official_indication",
    },
    {
        "seed": "REC_0818",
        "label": "Solta Medical - VASERlipo",
        "indication": "Indicated for the fragmentation, emulsification and aspiration of soft tissues in surgical specialties including plastic and reconstructive surgery.",
        "status": "user_confirmed_path_keep_official_indication",
    },
    {
        "seed": "REC_0931",
        "label": "Caregen - DR. CYJ Hair Filler",
        "indication": "Indicated for the treatment of scalp and hair lesions caused by external aggressions.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0201",
        "label": "Caregen - Dr. CYJ Hair Filler",
        "indication": "Indicated for the treatment of scalp and hair lesions caused by external aggressions.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0561",
        "label": "Cocoon Medical - Primelase HR excellence",
        "indication": "Intended for the permanent reduction in hair regrowth.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0834",
        "label": "Cutera - truSculpt flex+",
        "indication": "Indicated for the improvement of abdominal tone, strengthening of abdominal muscles, and development of a firmer abdomen.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0892",
        "label": "IBSA Derma - Aliaxin EV",
        "indication": "Indicated for the correction of deep skin damage in the face and to increase facial volume.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0890",
        "label": "IBSA Derma - Aliaxin SR",
        "indication": "Indicated for the correction of medium and deep skin damages of the face, and to increase the volume and contour of the lips.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0893",
        "label": "IBSA Derma - Aliaxin SV",
        "indication": "Indicated for increasing cheek volume, remodeling facial contours, and filling very deep wrinkles.",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0791",
        "label": "Needle Concept - Mesotherapy Needle",
        "indication": "适用于注射肉毒杆菌毒素和透明质酸等美容活性成分，用于除皱、美白、祛斑、脱发治疗和妊娠纹淡化。",
        "status": "user_confirmed_quality_repair",
    },
    {
        "seed": "REC_0316",
        "label": "APS - Hifu Top",
        "indication": "Indicated for tightening loose skin around the eyes and neck, lifting downturned mouth and jowls, enhancing flattened cheekbones, and reducing perioral wrinkles.",
        "status": "user_confirmed_not_explicit_rewrite",
    },
    {
        "seed": "REC_0853",
        "label": "Alma Lasers - Alma PrimeX",
        "indication": "Indicated for comprehensive body and facial contouring using a combination of guided ultrasound and deep radiofrequency heating.",
        "status": "user_confirmed_not_explicit_rewrite",
    },
    {
        "seed": "REC_0869",
        "label": "BTL - EMFEMME 360",
        "indication": "适用于治疗阴道松弛、性交疼痛、更年期泌尿生殖综合征 (GSM) 及改善尿渗漏问题。",
        "status": "user_confirmed_not_explicit_rewrite",
    },
    {
        "seed": "REC_0833",
        "label": "Cutera - xeo+",
        "indication": "The 1064 nm wavelength is indicated for the coagulation and hemostasis of benign vascular lesions, including port wine stains, hemangiomas, telangiectasias and rosacea, and for the treatment of benign cutaneous lesions including warts, scars and striae.",
        "status": "user_confirmed_not_explicit_rewrite",
    },
    {
        "seed": "REC_0191",
        "label": "Dives Med - Dives Filler",
        "indication": "Indicated for the correction of deep wrinkles and furrows, facial modeling including cheeks, jawline, nose and chin, and restoration of moderate to advanced volume loss.",
        "status": "user_confirmed_not_explicit_rewrite",
    },
]

EVIDENCE_FILES = [
    {"path": DATA_DIR / "manual_product_fact_evidence.csv", "note_field": "note", "status_field": "review_status"},
    {"path": DATA_DIR / "manual_evidence_promotion_log.csv", "note_field": "note", "status_field": None},
    {"path": DATA_DIR / "product_specification_evidence.csv", "note_field": "notes", "status_field": "review_status"},
]

INDICATION_FIELDS = [
    "product_id",
    "seed_record_id",
    "company_id",
    "company",
    "brand",
    "jurisdiction",
    "regulator",
    "regulatory_pathway",
    "status",
    "registration_no",
    "approval_date",
    "expiry_date",
    "registered_name",
    "approved_indication",
    "intended_use",
    "legal_manufacturer",
    "local_holder",
    "source_key",
    "source_url",
    "source_type",
    "evidence_title",
    "evidence_excerpt",
    "official_description_exact",
    "official_description_source_field",
    "field_note",
    "checked_at",
    "reviewed_by",
    "review_status",
    "confidence",
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def stable_id(prefix: str, *parts: Any) -> str:
    blob = "||".join(clean(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists():
        return [], []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def headers(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}


def append_note(existing: str, marker: str) -> str:
    existing = clean(existing)
    if marker in existing:
        return existing
    if existing:
        return f"{existing} | {marker}"
    return marker


def apply_workbook_exclusions(stamp: str) -> tuple[str, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_e_group_path_exclusions_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        clean(ws.cell(row=row_idx, column=colmap["Record_ID"]).value): row_idx
        for row_idx in range(2, ws.max_row + 1)
        if "Record_ID" in colmap
    }
    changes: list[dict[str, str]] = []
    for record_id, note in EXCLUDE_SEEDS.items():
        row_idx = row_by_id.get(record_id)
        if not row_idx:
            changes.append({"seed_record_id": record_id, "field": "row", "old": "", "new": "not_found"})
            continue
        for field, value in {
            "Inclusion_Status": "excluded",
            "Duplicate_Note": note,
            "V4_1_Registration_Review_Status": "excluded_scope",
        }.items():
            col = colmap.get(field)
            if not col:
                continue
            old = clean(ws.cell(row=row_idx, column=col).value)
            if old == value:
                continue
            ws.cell(row=row_idx, column=col).value = value
            changes.append({"seed_record_id": record_id, "field": field, "old": old, "new": value})
    if changes:
        wb.save(SOURCE_BOOK)
    return str(backup), changes


def unlink_excluded_product_evidence(product_ids: dict[str, str], stamp: str) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for item in EVIDENCE_FILES:
        path = item["path"]
        fields, rows = read_csv(path)
        if not rows:
            continue
        backup = path.with_name(f"{path.stem}_backup_before_e_group_path_excluded_unlink_{stamp}{path.suffix}")
        shutil.copy2(path, backup)
        by_id: Counter[str] = Counter()
        for row in rows:
            original_product_id = clean(row.get("product_id"))
            if original_product_id not in product_ids:
                continue
            marker = (
                "excluded_product_unlink_e_group_path_20260601: "
                f"{product_ids[original_product_id]} Retained as company/source-history evidence only, not active product evidence."
            )
            row["product_id"] = ""
            if "product_family_id" in row:
                row["product_family_id"] = ""
            status_field = item["status_field"]
            if status_field and status_field in row:
                row[status_field] = "excluded_scope_unlinked"
            note_field = item["note_field"]
            if note_field in row:
                row[note_field] = append_note(row.get(note_field, ""), marker)
            by_id[original_product_id] += 1
        write_csv(path, fields, rows)
        changes.append(
            {
                "file": str(path),
                "backup": str(backup),
                "rows_unlinked": sum(by_id.values()),
                "by_original_product_id": dict(by_id),
            }
        )
    return changes


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    checked_at = datetime.now().astimezone().replace(microsecond=0).isoformat()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    _, product_rows = read_csv(PRODUCT_MASTER)
    manual_fields, manual_rows = read_csv(MANUAL_INDICATION)
    _, review_rows = read_csv(REVIEW_PACK)

    product_by_seed = {clean(row.get("seed_record_id")): row for row in product_rows if clean(row.get("seed_record_id"))}
    review_by_seed = {clean(row.get("seed_record_id")): row for row in review_rows if clean(row.get("seed_record_id"))}

    existing_keys = {
        (
            clean(row.get("product_id")),
            clean(row.get("source_url")),
            clean(row.get("official_description_exact") or row.get("approved_indication")),
        )
        for row in manual_rows
    }

    new_manual_rows: list[dict[str, str]] = []
    applied_rows: list[dict[str, str]] = []
    for item in INDICATIONS:
        seed = item["seed"]
        product = product_by_seed.get(seed)
        source = review_by_seed.get(seed, {})
        if not product:
            applied_rows.append({"seed_record_id": seed, "label": item["label"], "action": "not_found", "product_id": "", "source_url": ""})
            continue
        product_id = clean(product.get("product_id"))
        source_url = clean(source.get("source_url"))
        indication = clean(item["indication"])
        key = (product_id, source_url, indication)
        if key in existing_keys:
            applied_rows.append({"seed_record_id": seed, "label": item["label"], "action": "already_present", "product_id": product_id, "source_url": source_url})
            continue
        row = {
            "product_id": product_id,
            "seed_record_id": seed,
            "company_id": clean(product.get("company_id")),
            "company": clean(product.get("company")),
            "brand": clean(product.get("brand")),
            "jurisdiction": "Global",
            "regulator": "Official source / user-confirmed extraction",
            "regulatory_pathway": "User-confirmed E-group indication normalization",
            "status": "User-confirmed official indication/use wording",
            "registration_no": "",
            "approval_date": "",
            "expiry_date": "",
            "registered_name": clean(product.get("registered_name") or product.get("standard_product_name")),
            "approved_indication": indication,
            "intended_use": indication,
            "legal_manufacturer": clean(product.get("legal_manufacturer") or product.get("company")),
            "local_holder": clean(product.get("local_holder")),
            "source_key": stable_id("e_group_path_quality", product_id, source_url, indication),
            "source_url": source_url,
            "source_type": clean(source.get("source_type")) or "existing_candidate_source",
            "evidence_title": clean(source.get("evidence_title")) or item["label"],
            "evidence_excerpt": clean(source.get("candidate_text")),
            "official_description_exact": indication,
            "official_description_source_field": "user_confirmed_normalized_from_existing_candidate_text",
            "field_note": (
                "User feedback 2026-06-01 normalized this E-group item from clinical-path/quality/not-explicit review. "
                "For nonofficial or distributor source URLs, treat as a user-confirmed interim wording and replace with stronger official IFU when available."
            ),
            "checked_at": checked_at,
            "reviewed_by": "user_feedback_20260601",
            "review_status": item["status"],
            "confidence": "user_confirmed_normalized_indication",
        }
        new_manual_rows.append(row)
        existing_keys.add(key)
        applied_rows.append({"seed_record_id": seed, "label": item["label"], "action": "added", "product_id": product_id, "source_url": source_url})

    manual_backup = ""
    if new_manual_rows:
        backup = AUDIT_DIR / f"manual_official_indication_evidence_backup_before_e_group_path_quality_{stamp}.csv"
        shutil.copy2(MANUAL_INDICATION, backup)
        manual_backup = str(backup)
        write_csv(MANUAL_INDICATION, manual_fields or INDICATION_FIELDS, [*manual_rows, *new_manual_rows])

    workbook_backup, workbook_changes = apply_workbook_exclusions(stamp)

    excluded_product_ids: dict[str, str] = {}
    excluded_rows: list[dict[str, str]] = []
    for seed, reason in EXCLUDE_SEEDS.items():
        product = product_by_seed.get(seed, {})
        product_id = clean(product.get("product_id"))
        if product_id:
            excluded_product_ids[product_id] = reason
        excluded_rows.append(
            {
                "seed_record_id": seed,
                "product_id": product_id,
                "company": clean(product.get("company")),
                "brand": clean(product.get("brand")),
                "standard_product_name": clean(product.get("standard_product_name")),
                "decision": "exclude",
                "reason": reason,
            }
        )
    unlink_changes = unlink_excluded_product_evidence(excluded_product_ids, stamp)

    deferred_rows = []
    for seed, reason in DEFERRED_SEEDS.items():
        product = product_by_seed.get(seed, {})
        source = review_by_seed.get(seed, {})
        deferred_rows.append(
            {
                "seed_record_id": seed,
                "product_id": clean(product.get("product_id")),
                "company": clean(product.get("company") or source.get("company")),
                "brand": clean(product.get("brand") or source.get("brand")),
                "standard_product_name": clean(product.get("standard_product_name") or source.get("standard_product_name")),
                "decision": "defer_to_reacquire_official_source",
                "reason": reason,
                "source_url": clean(source.get("source_url")),
            }
        )

    write_csv(APPLIED_CSV, ["seed_record_id", "label", "action", "product_id", "source_url"], applied_rows)
    write_csv(EXCLUDED_CSV, ["seed_record_id", "product_id", "company", "brand", "standard_product_name", "decision", "reason"], excluded_rows)
    write_csv(DEFERRED_CSV, ["seed_record_id", "product_id", "company", "brand", "standard_product_name", "decision", "reason", "source_url"], deferred_rows)

    summary = {
        "checked_at": checked_at,
        "new_manual_indication_rows_added": len(new_manual_rows),
        "indication_decisions": len(INDICATIONS),
        "excluded_seed_rows": len(EXCLUDE_SEEDS),
        "deferred_seed_rows": len(DEFERRED_SEEDS),
        "manual_backup": manual_backup,
        "workbook_backup": workbook_backup,
        "workbook_changes": len(workbook_changes),
        "evidence_unlink_total_rows": sum(change["rows_unlinked"] for change in unlink_changes),
        "evidence_unlink_changes": unlink_changes,
        "outputs": {
            "applied_csv": str(APPLIED_CSV),
            "excluded_csv": str(EXCLUDED_CSV),
            "deferred_csv": str(DEFERRED_CSV),
        },
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
