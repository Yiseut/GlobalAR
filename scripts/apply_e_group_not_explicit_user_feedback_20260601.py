#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"

PRODUCT_MASTER = DATA_DIR / "product_master.csv"
MANUAL_INDICATION = DATA_DIR / "manual_official_indication_evidence.csv"
CANDIDATES = AUDIT_DIR / "e_group_indication_extraction_candidates_latest.csv"
UNCERTAIN = AUDIT_DIR / "e_group_indication_extraction_uncertain_latest.csv"
MANUAL_FACT = DATA_DIR / "manual_product_fact_evidence.csv"

SUMMARY_JSON = AUDIT_DIR / "e_group_not_explicit_user_feedback_apply_latest.json"
APPLIED_CSV = AUDIT_DIR / "e_group_not_explicit_user_feedback_applied_latest.csv"
UNMATCHED_CSV = AUDIT_DIR / "e_group_not_explicit_user_feedback_unmatched_latest.csv"


FIELDS = [
    "product_id",
    "seed_record_id",
    "company_id",
    "company",
    "brand",
    "jurisdiction",
    "regulator",
    "regulatory_pathway",
    "status",
    "registration_no",
    "approval_date",
    "expiry_date",
    "registered_name",
    "approved_indication",
    "intended_use",
    "legal_manufacturer",
    "local_holder",
    "source_key",
    "source_url",
    "source_type",
    "evidence_title",
    "evidence_excerpt",
    "official_description_exact",
    "official_description_source_field",
    "field_note",
    "checked_at",
    "reviewed_by",
    "review_status",
    "confidence",
]


USER_INDICATIONS = [
    # EBD
    (["REC_0905"], "Asclepion - Juliet", "适用于作为激素疗法或手术的替代方案，治疗女性绝经期泌尿生殖综合征等隐秘疾病。"),
    (["REC_0904", "REC_0545"], "Asclepion - PicoStar", "适用于纹身及色素病变去除。"),
    (["REC_0580"], "Asclepion Laser - QuadroStar Pro", "适用于治疗毛细血管扩张、小血管瘤、红血丝或良性色素病变。"),
    (["REC_0398", "REC_0399"], "Asterasys - Liftera", "聚焦超声设备，适用于面部及颈部的组织提拉与紧致。"),
    (["REC_0190"], "Aurora Medical - Diode 808nm", "808nm 半导体激光设备，适用于破坏毛乳头以达到永久性脱毛的目的。"),
    (["REC_0366"], "BRERA Medical - Jovena", "适用于美容医学和皮肤科中的面部及身体皮下组织治疗，结合射频产生热效应并伴随肌肉刺激。"),
    (["REC_0229"], "BTL - Emsella", "适用于通过非侵入性 HIFEM 电磁刺激，对盆底肌肉进行康复并恢复神经肌肉控制，以治疗男性和女性的尿失禁。"),
    (["REC_0857", "REC_0547"], "Candela - PicoWay", "适用于治疗黄褐斑、雀斑、咖啡斑、太田痣、皱纹、良性色素病变、痤疮疤痕以及纹身去除。"),
    (["REC_0116", "REC_0913"], "Classys - Clatuu Alpha", "适用于通过冷冻溶脂技术减少身体多余脂肪。"),
    (["REC_0916"], "Classys - FORSHAPE", "射频治疗系统，适用于促进新陈代谢、改善血液循环并加速组织恢复。"),
    (["REC_0633"], "Classys - Scizer", "适用于减少腹部、侧腰等部位的顽固脂肪。"),
    (["REC_0757", "REC_0912"], "Classys - Volnewmer", "采用单极射频及表皮冷却系统，适用于组织加热治疗。"),
    (["REC_0238"], "Cutera - enlighten III", "采用点阵技术，适用于全肤质的痤疮疤痕、纹理问题和皮肤老化治疗。"),
    (["REC_0256"], "Cutera - excel V+", "适用于良性血管和皮肤病变的凝固与止血，包括玫瑰痤疮、弥漫性红斑、面部/腿部/眶周静脉、血管瘤、鲜红斑痣、皮肤异色症、皱纹、静脉畸形、炎症性痤疮、良性色素病变及疤痕等。"),
    (["REC_0556"], "Cynosure - Potenza", "射频微针设备，适用于软组织电凝和止血等皮肤科常规手术。"),
    (["REC_0678"], "Storz Medical - Storz", "体外冲击波疗法，适用于橘皮组织、皱纹改善、身体塑形及皮肤紧致。"),
    # Injectables
    (["REC_0735"], "Adoderm - Variofill", "透明质酸凝胶填充剂，专用于臀部增大或塑形。"),
    (["REC_0973"], "BNC Korea - BONIVA", "适用于暂时性改善 19-65 岁成人因皱眉肌和/或降眉肌活动引起的中重度眉间纹。"),
    (["REC_0417"], "BNC Korea - Lumifil", "具备高粘度特性的透明质酸填充剂，适用于面部轮廓重塑及 3D 容量填充。"),
    (["REC_0249"], "BioFormula - Evanthia", "Class III 医疗器械，适用于专业医疗面部注射填充。"),
    (["REC_0148"], "Bioxis Pharmaceuticals - CYTOSIAL", "透明质酸真皮填充剂，用于面部皱纹纠正与容量恢复。"),
    (["REC_0194"], "Biovico - DKK DERM", "医疗器械，适用于分离和提取具有增强治疗特性的富血小板血浆。"),
    (["REC_0136"], "Croma Pharma - Croma PDO Threads", "适用于面部及身体多部位的埋线提拉与皮肤改善，包括前额、脸颊、眉毛、鼻子、下颌线、颈部、大腿内侧和腹部等。"),
    # Skincare / regenerative
    (["REC_0180"], "Alta Care - Dermastir Luxury", "亮肤面部爽肤产品，旨在均匀肤色并恢复皮肤光泽。"),
    (["REC_0044", "REC_0453"], "Aquavit Pharmaceuticals - Aquagold", "微通道递送系统，适用于真皮层局部导入与精华递送。"),
    (["REC_0060"], "Arkana Cosmetics - AzAc Solution", "复合酸焕肤液，适用于配合 15% 壬二酸面霜使用的治疗前准备与强化。"),
    (["REC_0258"], "Arkana Cosmetics - Exo Complex", "高级外泌体疗法，适用于增强皮肤结构、促进皮肤再生及治疗疤痕。"),
    (["REC_0072"], "Bayer - Bepanthen / Bepanthol", "适用于保持皲裂或干燥皮肤的完整性、哺乳期乳房护理，以及支持轻微损伤皮肤的再生。"),
    (["REC_0334", "REC_0956"], "Beauty Health - HydraFacial", "适用于改善皱纹、轻中度痤疮外观，镇静偶发性红斑，以及清理油性或充血皮肤和粉刺。"),
    (["REC_0302"], "BeautyBio - GloPRO Scalp / Body", "居家微针工具，适用于抚平皱纹外观、改善大腿后侧、臀部和腋下皮肤的紧致度。"),
    (["REC_0033"], "BenQ - Anscare", "硅胶贴/棒，适用于减少表现为隆起或变色的肥厚性疤痕和瘢痕疙瘩。"),
    (["REC_0165"], "BenQ - DermaAngel", "含有水杨酸的痤疮贴，适用于深层疏通毛孔和镇静红肿的密集护理。"),
    (["REC_0070"], "Benev - Benev Exosome", "冻干外泌体复合物，适用于皮肤和头皮再生的局部外用涂抹。"),
    (["REC_0110"], "BioFormula - Cellulysis", "含有液态磷脂酰胆碱的活性成分混合物，适用于局部脂肪减少。"),
    (["REC_0968"], "BioPlus - Kiara Reju", "旨在支持皮肤平滑并改善不均匀皮肤纹理的皮肤动能素。"),
    (["REC_0443"], "Biopark Medical - Mesopotamia", "适用于护肤应用的外泌体再生产品。"),
    (["REC_0621"], "Christina Cosmeceuticals - Rose de Mer", "100% 纯天然海洋焕肤产品，适用于无需化学剥脱的专业皮肤更新。"),
    (["REC_0078"], "Crown - BIOJUVE", "局部活体微生态精华与喷雾，适用于恢复皮肤微生物群平衡。"),
    # Other
    (["REC_0103"], "A.A.M.S. - Carboxy-Pen", "二氧化碳注入设备，适用于深度可调及剂量可控的皮下 CO2 治疗。"),
    (["REC_0303"], "Arion - Gluteal / Calf Implants", "硅胶假体，适用于通过放置在现有腓肠肌上方来增强小腿肌肉的形状和大小。"),
    (["REC_0189"], "Bomtech Electronics - Digital Pop", "半永久化妆机器，适用于将色素吸入一次性针头并标记到皮肤上。"),
    (["REC_0139"], "Brymill - Cry-Baby", "手持式冷冻外科设备，适用于通过开放式喷雾或接触探头技术受控分配液氮以冷冻皮肤病变。"),
    (["REC_0753"], "Canfield Scientific - VISIA", "面部图像分析系统，适用于利用标准光、交叉偏振光和 UV 光记录并测量表面及皮肤底层状况。"),
]

EXCLUDE_SCOPE_SEEDS = {
    "REC_0475": "excluded_by_user_e_group_scope_20260601: Arthrex NanoScope is a generic endoscopic surgical device outside core aesthetics upstream scope.",
    "REC_0101": "excluded_by_user_e_group_scope_20260601: Contura Bulkamid is a urogynecology bulking agent for stress urinary incontinence, outside core aesthetics upstream scope.",
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def stable_id(prefix: str, *parts: Any) -> str:
    blob = "||".join(clean(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)


def headers(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}


def source_index() -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for path in [UNCERTAIN, CANDIDATES]:
        if not path.exists():
            continue
        _, rows = read_csv(path)
        for row in rows:
            product_id = clean(row.get("product_id"))
            if product_id and clean(row.get("source_url")) and product_id not in out:
                out[product_id] = row
    if MANUAL_FACT.exists():
        _, rows = read_csv(MANUAL_FACT)
        for row in rows:
            product_id = clean(row.get("product_id"))
            if product_id and clean(row.get("source_url")) and product_id not in out:
                out[product_id] = {
                    "source_url": row.get("source_url", ""),
                    "source_type": row.get("source_type", ""),
                    "evidence_title": row.get("evidence_title", ""),
                    "source_evidence_excerpt": row.get("evidence_excerpt", ""),
                }
    return out


def manual_row(product: dict[str, str], indication: str, label: str, source: dict[str, str], checked_at: str) -> dict[str, str]:
    product_id = clean(product.get("product_id"))
    source_url = clean(source.get("source_url"))
    return {
        "product_id": product_id,
        "seed_record_id": clean(product.get("seed_record_id")),
        "company_id": clean(product.get("company_id")),
        "company": clean(product.get("company")),
        "brand": clean(product.get("brand")),
        "jurisdiction": "Global",
        "regulator": "Official product/IFU/source text; user-normalized",
        "regulatory_pathway": "user-confirmed E-group indication/use backfill",
        "status": "User confirmed official indication/intended-use wording from E-group not_explicit review",
        "registration_no": "",
        "approval_date": "",
        "expiry_date": "",
        "registered_name": clean(product.get("registered_name") or product.get("standard_product_name") or product.get("brand")),
        "approved_indication": "",
        "intended_use": indication,
        "legal_manufacturer": clean(product.get("legal_manufacturer") or product.get("manufactured_by") or product.get("company")),
        "local_holder": clean(product.get("local_holder")),
        "source_key": stable_id("egroup_user_ind", product_id, label, indication),
        "source_url": source_url,
        "source_type": "user_confirmed_e_group_indication",
        "evidence_title": clean(source.get("evidence_title")) or f"{label} user-confirmed official indication",
        "evidence_excerpt": clean(source.get("source_evidence_excerpt") or source.get("extracted_text"))[:1200],
        "official_description_exact": indication,
        "official_description_source_field": "user_confirmed_normalized_official_use",
        "field_note": "User provided normalized Chinese official indication/intended-use wording after reviewing the E-group not_explicit candidates. Original source text should remain traceable through source_url/evidence_excerpt.",
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_e_group_not_explicit_20260601",
        "review_status": "user_confirmed_official_indication",
        "confidence": "user_confirmed_normalized_official_use",
    }


def apply_workbook_exclusions(stamp: str) -> tuple[str, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_e_group_scope_exclusions_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        clean(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if clean(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []
    for record_id, note in EXCLUDE_SCOPE_SEEDS.items():
        row_idx = row_by_id.get(record_id)
        if not row_idx:
            continue
        for field, value in {
            "Inclusion_Status": "excluded",
            "Duplicate_Note": note,
            "V4_1_Registration_Review_Status": "excluded_scope",
        }.items():
            col = colmap.get(field)
            if not col:
                continue
            old = clean(ws.cell(row=row_idx, column=col).value)
            if old == value:
                continue
            ws.cell(row=row_idx, column=col).value = value
            changes.append({"record_id": record_id, "field": field, "old": old, "new": value})
    if changes:
        wb.save(SOURCE_BOOK)
    return str(backup), changes


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")

    _, products = read_csv(PRODUCT_MASTER)
    product_by_seed = {clean(row.get("seed_record_id")): row for row in products if clean(row.get("seed_record_id"))}
    product_by_id = {clean(row.get("product_id")): row for row in products if clean(row.get("product_id"))}
    source_by_product = source_index()
    manual_fields, manual_rows = read_csv(MANUAL_INDICATION)

    target_product_ids = {
        clean(product_by_seed.get(seed, {}).get("product_id"))
        for seeds, _, _ in USER_INDICATIONS
        for seed in seeds
        if clean(product_by_seed.get(seed, {}).get("product_id"))
    }
    retained_manual = [
        row
        for row in manual_rows
        if not (
            clean(row.get("product_id")) in target_product_ids
            and clean(row.get("reviewed_by")) == "auto_e_group_indication_extractor_20260601"
        )
    ]
    existing_keys = {
        (clean(row.get("product_id")), clean(row.get("official_description_exact") or row.get("intended_use")))
        for row in retained_manual
    }

    applied: list[dict[str, str]] = []
    unmatched: list[dict[str, str]] = []
    new_rows: list[dict[str, str]] = []
    for seeds, label, indication in USER_INDICATIONS:
        for seed in seeds:
            product = product_by_seed.get(seed)
            if not product:
                unmatched.append({"label": label, "seed_record_id": seed, "reason": "seed_not_found"})
                continue
            product_id = clean(product.get("product_id"))
            if not product_id:
                unmatched.append({"label": label, "seed_record_id": seed, "reason": "product_id_missing"})
                continue
            source = source_by_product.get(product_id, {})
            key = (product_id, indication)
            if key not in existing_keys:
                new_rows.append(manual_row(product, indication, label, source, checked_at))
                existing_keys.add(key)
            applied.append(
                {
                    "label": label,
                    "seed_record_id": seed,
                    "product_id": product_id,
                    "company": clean(product.get("company")),
                    "brand": clean(product.get("brand")),
                    "indication": indication,
                    "source_url": clean(source.get("source_url")),
                }
            )

    backup_manual = AUDIT_DIR / f"manual_official_indication_evidence_backup_before_e_group_user_feedback_{stamp}.csv"
    shutil.copy2(MANUAL_INDICATION, backup_manual)
    write_csv(MANUAL_INDICATION, manual_fields or FIELDS, [*retained_manual, *new_rows])

    workbook_backup, workbook_changes = apply_workbook_exclusions(stamp)

    write_csv(APPLIED_CSV, ["label", "seed_record_id", "product_id", "company", "brand", "indication", "source_url"], applied)
    write_csv(UNMATCHED_CSV, ["label", "seed_record_id", "reason"], unmatched)

    summary = {
        "checked_at": checked_at,
        "user_labels": len(USER_INDICATIONS),
        "target_product_ids": len(target_product_ids),
        "applied_product_rows": len(applied),
        "new_manual_indication_rows_added": len(new_rows),
        "auto_e_group_rows_replaced": len(manual_rows) - len(retained_manual),
        "unmatched": len(unmatched),
        "workbook_scope_exclusion_changes": len(workbook_changes),
        "workbook_scope_exclusion_records": sorted(EXCLUDE_SCOPE_SEEDS),
        "by_company": dict(Counter(row["company"] for row in applied).most_common()),
        "manual_backup": str(backup_manual),
        "workbook_backup": workbook_backup,
        "outputs": {
            "applied_csv": str(APPLIED_CSV),
            "unmatched_csv": str(UNMATCHED_CSV),
        },
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
