"""Mark source workbook rows as verified when promoted official product facts exist."""

from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER = DATA_DIR / "product_master.csv"
MANUAL_FACTS = DATA_DIR / "manual_product_fact_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def promoted_official_seed_ids() -> set[str]:
    seeds: set[str] = set()
    for row in read_csv(MANUAL_FACTS):
        fact_group = norm(row.get("fact_group")).lower()
        source_type = norm(row.get("source_type")).lower()
        review_status = norm(row.get("review_status")).lower()
        if fact_group not in {"top_company_productline_official_verification", "upstream_report_reference_official_verification"}:
            continue
        if source_type != "official_product_page" or review_status != "promoted":
            continue
        seed_id = norm(row.get("seed_record_id"))
        if seed_id:
            seeds.add(seed_id)
    return seeds


def current_unverified_products() -> set[str]:
    out: set[str] = set()
    for row in read_csv(PRODUCT_MASTER):
        if norm(row.get("verification_status")) == "unverified_seed":
            out.add(norm(row.get("seed_record_id")))
    return out


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    target_ids = promoted_official_seed_ids() & current_unverified_products()

    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_promoted_fact_status_sync_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if old == value:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": value})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id in sorted(target_ids):
        set_cell(record_id, "Data_Source", "official_product_fact_promoted")
        append_audit(record_id, "promoted_fact_status_sync_20260601: promoted official product facts treated as source-verified product identity.")

    wb.save(SOURCE_BOOK)
    wb.close()

    summary = {
        "backup": str(backup),
        "target_record_ids": sorted(target_ids),
        "workbook_changes": len(changes),
        "changed_fields_sample": changes[:120],
    }
    out = AUDIT_DIR / f"promoted_product_fact_status_sync_{stamp}.json"
    latest = AUDIT_DIR / "promoted_product_fact_status_sync_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
