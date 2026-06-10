"""Apply user-confirmed EBD regulatory/indication notes from 2026-06-01."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_FACT_PATH = DATA_DIR / "manual_product_fact_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_fields: list[str], new_rows: list[dict[str, str]]) -> int:
    existing = {tuple(norm(row.get(field)) for field in key_fields) for row in rows}
    added = 0
    for row in new_rows:
        key = tuple(norm(row.get(field)) for field in key_fields)
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def evidence_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    indication = norm(spec.get("indication"))
    source_key = norm(spec.get("source_key")) or stable_id(
        "user_confirmed_reg_ind_20260601_b4",
        product.get("seed_record_id"),
        spec.get("jurisdiction"),
        spec.get("regulator"),
        spec.get("registered_name"),
        spec.get("registration_no"),
    )
    return {
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "jurisdiction": norm(spec.get("jurisdiction")),
        "regulator": norm(spec.get("regulator")),
        "regulatory_pathway": norm(spec.get("pathway")),
        "status": norm(spec.get("status")),
        "registration_no": norm(spec.get("registration_no")),
        "approval_date": norm(spec.get("approval_date")),
        "expiry_date": norm(spec.get("expiry_date")),
        "registered_name": norm(spec.get("registered_name")) or product.get("brand", ""),
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": norm(spec.get("legal_manufacturer")) or product.get("legal_manufacturer") or product.get("company", ""),
        "local_holder": norm(spec.get("local_holder")),
        "source_key": source_key,
        "source_url": norm(spec.get("source_url")),
        "source_type": norm(spec.get("source_type")) or "user_confirmed_official_claim",
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": norm(spec.get("field_note")),
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601_batch4",
        "review_status": norm(spec.get("review_status")) or "user_confirmed",
        "confidence": norm(spec.get("confidence")) or "user_confirmed_official_claim",
    }


def fact_row(product: dict[str, str], checked_at: str, source_url: str, title: str, excerpt: str) -> dict[str, str]:
    return {
        "fact_id": stable_id("pfact", product.get("seed_record_id"), "official_product_page", source_url),
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "product_family_id": "",
        "standard_product_name": product.get("standard_product_name", ""),
        "priority": "P0",
        "fact_group": "official_product_page",
        "field_name": "official_product_page",
        "field_value": source_url,
        "source_url": source_url,
        "evidence_title": title,
        "evidence_excerpt": excerpt,
        "source_type": "official_product_page",
        "confidence": "official_site_user_confirmed",
        "captured_at": checked_at,
        "promoted_at": checked_at,
        "review_status": "auto_cross_checked",
        "note": "user_confirmed_regulatory_indications_20260601_batch4",
    }


TARGET_SPECS: dict[str, list[dict[str, str]]] = {
    "REC_0009": [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing FDA record and user feedback",
            "registration_no": "K072699",
            "registered_name": "Accent Prime / Accent RF System",
            "indication": "Accent Prime 结合单极射频与超声技术，用于面部及身体皮肤深层紧致、轮廓提升、局部非侵入性脂肪消融和身体塑形。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://almalasers.com/product/accent-prime/",
            "source_type": "official_product_page_user_confirmed_fda_claim",
            "evidence_title": "Alma Accent Prime official page / user-confirmed FDA status",
            "evidence_excerpt": "用户确认 Accent Prime 拥有 FDA、CE 及中国 NMPA 注册；官方页面列示 RF + ultrasound、body contouring、skin tightening 等定位。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Accent Prime",
            "indication": "Accent Prime 用于身体和面部轮廓塑形、皮肤紧致、皱纹/皮肤松弛和脂肪组织处理等医美能量治疗。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://almalasers.com/product/accent-prime/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Alma Accent Prime official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Accent Prime 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
        {
            "jurisdiction": "CN",
            "regulator": "NMPA",
            "pathway": "Class III medical device registration",
            "status": "NMPA Class III registration confirmed by user; certificate number not captured",
            "registered_name": "Accent Prime / ThermoLift",
            "indication": "Accent Prime/热拉提用于皮肤紧致、轮廓提升和身体塑形；具体中国注册证号待证号级核验。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://almalasers.com/product/accent-prime/",
            "source_type": "user_confirmed_nmpa_claim_with_official_product_page",
            "evidence_title": "Alma Accent Prime official page / user-confirmed NMPA status",
            "evidence_excerpt": "用户确认 Accent Prime 在中国拥有 NMPA 三类医疗器械注册；本批未提供具体注册证号。",
            "field_note": "Keep China status as confirmed presence until exact 国械注进 number is captured.",
        },
    ],
    "REC_0522": [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing FDA record and user feedback",
            "registration_no": "K201520",
            "registered_name": "Opus / Opus Plasma",
            "indication": "Opus Plasma 采用高频单极射频激发的微剥脱等离子体/分段 RF 技术，用于皮肤重建、改善面部细纹、皱纹、痤疮疤痕以及肤色和肤质不均。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://almalasers.com/product/opus/",
            "source_type": "official_product_page_user_confirmed_fda_claim",
            "evidence_title": "Alma Opus official page / user-confirmed FDA status",
            "evidence_excerpt": "用户确认 Opus Plasma 拥有 FDA 与 CE；官方页面列示 fractional RF、Opus Plasma、skin resurfacing、skin tightening、scars 等定位。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Opus Plasma",
            "indication": "Opus Plasma 用于皮肤重建、紧致、面部轮廓、疤痕和肤质改善等分段射频/等离子体治疗。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://almalasers.com/product/opus/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Alma Opus official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Opus Plasma 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
    ],
    "REC_0675": [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing FDA records and user feedback",
            "registration_no": "K230371; K222064",
            "registered_name": "Soprano Titanium / Soprano ICE Platinum",
            "indication": "Soprano Titanium / ICE Platinum 将 755nm、810nm、1064nm 三种波长整合在同一治疗头中，用于所有皮肤类型的安全、快速毛发脱减。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://www.alma-soprano.com/soprano-ice-platinum/",
            "source_type": "official_product_page_user_confirmed_fda_claim",
            "evidence_title": "Soprano ICE Platinum official page / user-confirmed FDA status",
            "evidence_excerpt": "用户确认 Soprano Titanium / ICE Platinum 拥有 FDA、CE 与中国 NMPA；官方页面列示 755/810/1064nm 三波长、all skin types、hair removal。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Soprano Titanium / Soprano ICE Platinum",
            "indication": "Soprano Titanium / ICE Platinum 用于三波长半导体激光脱毛，适用于不同毛发和皮肤类型。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://www.alma-soprano.com/soprano-ice-platinum/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Soprano ICE Platinum official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Soprano Titanium / ICE Platinum 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
        {
            "jurisdiction": "CN",
            "regulator": "NMPA",
            "pathway": "Medical device registration",
            "status": "NMPA registration confirmed by user; certificate number not captured",
            "registered_name": "Soprano Titanium / ICE Platinum",
            "indication": "Soprano Titanium / ICE Platinum 在中国注册状态由用户确认；作为半导体脱毛激光用于毛发脱减，具体型号和注册证号待证号级核验。",
            "legal_manufacturer": "Alma Lasers",
            "source_url": "https://www.alma-soprano.com/soprano-ice-platinum/",
            "source_type": "user_confirmed_nmpa_claim_with_official_product_page",
            "evidence_title": "Soprano ICE Platinum official page / user-confirmed NMPA status",
            "evidence_excerpt": "用户确认 Soprano Titanium / ICE Platinum 拥有中国 NMPA 注册；本批未提供具体注册证号。",
        },
    ],
    "REC_0937": [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing FDA records and user feedback",
            "registration_no": "K202880",
            "registered_name": "J-Plasma Precise FLEX Handpiece",
            "indication": "J-Plasma 是氦气等离子加射频技术平台，用于外科手术中对软组织进行精准切割、凝固和消融，并降低周围组织热扩散。",
            "legal_manufacturer": "Apyx Medical",
            "source_url": "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm?ID=K202880",
            "source_type": "official_fda_record_user_confirmed",
            "evidence_title": "FDA 510(k) K202880 - J-Plasma Precise FLEX Handpiece",
            "evidence_excerpt": "用户确认 J-Plasma 拥有美国 FDA 与欧洲 CE 认证；FDA 510(k) 数据库已有 J-Plasma handpiece 记录。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "J-Plasma",
            "indication": "J-Plasma 用于软组织切割、凝固和消融的氦气等离子/RF 外科能量平台。",
            "legal_manufacturer": "Apyx Medical",
            "source_url": "https://eifu.apyxmedical.com/apyxmedical/en/handpieces?keycode=apyxmedical0000033",
            "source_type": "user_confirmed_ce_claim_with_manufacturer_ifu",
            "evidence_title": "Apyx eIFU / user-confirmed J-Plasma CE status",
            "evidence_excerpt": "用户确认 J-Plasma 拥有欧洲 CE 认证；未提供具体证书编号，使用厂家 eIFU 承接产品身份。",
        },
    ],
    "REC_0936": [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing FDA record and user feedback",
            "registration_no": "K230272",
            "registered_name": "Renuvion APR Handpiece",
            "indication": "Renuvion APR 通过皮下微创导入氦气等离子和射频能量，用于软组织凝固/收缩、吸脂后皮下软组织凝固，以及改善颈部和颌下区域松弛皮肤外观。",
            "legal_manufacturer": "Apyx Medical",
            "source_url": "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm?ID=K230272",
            "source_type": "official_fda_record_user_confirmed",
            "evidence_title": "FDA 510(k) K230272 - Renuvion APR Handpiece",
            "evidence_excerpt": "用户确认 Renuvion/Renuvion APR 拥有 FDA 与 CE；FDA 记录和安全沟通列示 Renuvion APR 的当前适用范围。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Renuvion APR",
            "indication": "Renuvion APR 是面向整形美容市场的氦气等离子/RF 皮下收紧方案，常与吸脂联合用于身体及颈部/下颌缘等部位软组织收缩。",
            "legal_manufacturer": "Apyx Medical",
            "source_url": "https://eifu.apyxmedical.com/apyxmedical/en/handpieces?keycode=apyxmedical0000033",
            "source_type": "user_confirmed_ce_claim_with_manufacturer_ifu",
            "evidence_title": "Apyx eIFU / user-confirmed Renuvion CE status",
            "evidence_excerpt": "用户确认 Renuvion/Renuvion APR 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
    ],
    "REC_0903": [
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing FDA record and user feedback",
            "registration_no": "K210634",
            "registered_name": "MCL 31 Dermablate System / Dermablate",
            "indication": "Dermablate 是 2940nm Er:YAG 铒激光设备，用于高精度表皮良性病变消融、微剥脱/全剥脱性皮肤重建、重度疤痕改善和面部深层除皱。",
            "legal_manufacturer": "Asclepion Laser Technologies",
            "source_url": "https://asclepion.com/dermablate/",
            "source_type": "official_product_page_user_confirmed_fda_claim",
            "evidence_title": "Asclepion Dermablate official page / user-confirmed FDA status",
            "evidence_excerpt": "用户确认 Dermablate 拥有 FDA、CE 及中国 NMPA；官方页面列示 Er:YAG、全剥脱/分段剥脱、疤痕、皮肤年轻化等定位。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Dermablate",
            "indication": "Dermablate 作为 Er:YAG 铒激光用于皮肤消融、分段治疗、疤痕治疗、皮肤年轻化及相关皮肤重建操作。",
            "legal_manufacturer": "Asclepion Laser Technologies",
            "source_url": "https://asclepion.com/dermablate/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Asclepion Dermablate official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Dermablate 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
        {
            "jurisdiction": "CN",
            "regulator": "NMPA",
            "pathway": "Medical device registration",
            "status": "NMPA registration confirmed by user; certificate number not captured",
            "registered_name": "Dermablate / MCL31 Dermablate",
            "indication": "Dermablate 在中国注册状态由用户确认；作为 2940nm Er:YAG 铒激光用于皮肤重建和消融相关医美治疗，具体注册证号待证号级核验。",
            "legal_manufacturer": "Asclepion Laser Technologies",
            "source_url": "https://asclepion.com/dermablate/",
            "source_type": "user_confirmed_nmpa_claim_with_official_product_page",
            "evidence_title": "Asclepion Dermablate official page / user-confirmed NMPA status",
            "evidence_excerpt": "用户确认 Dermablate 拥有中国 NMPA 注册；本批未提供具体注册证号。",
        },
    ],
}


WORKBOOK_UPDATES: dict[str, dict[str, str]] = {
    "REC_0009": {
        "FDA_Status": "FDA 510(k), CE and NMPA confirmed by user; CE/NMPA certificate numbers pending.",
        "FDA_510k_Number": "K072699",
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "NMPA_Status": "NMPA Class III registration confirmed by user; certificate number pending.",
        "Category_L1": "EBD",
        "Category_L2": "Radiofrequency",
        "Tech_Type_Std": "Unipolar RF + Ultrasound",
        "Manufactured_By": "Alma Lasers",
        "Marketing_Holder": "Alma Lasers",
    },
    "REC_0522": {
        "FDA_Status": "FDA 510(k) and CE confirmed by user; CE certificate number pending.",
        "FDA_510k_Number": "K201520",
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "EBD",
        "Category_L2": "Plasma",
        "Tech_Type_Std": "Fractional RF Plasma",
        "Manufactured_By": "Alma Lasers",
        "Marketing_Holder": "Alma Lasers",
    },
    "REC_0675": {
        "FDA_Status": "FDA 510(k), CE and NMPA confirmed by user; CE/NMPA certificate numbers pending.",
        "FDA_510k_Number": "K230371; K222064",
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "NMPA_Status": "NMPA registration confirmed by user; certificate number pending.",
        "Category_L1": "EBD",
        "Category_L2": "Laser",
        "Tech_Type_Std": "755/810/1064nm Diode Laser",
        "Manufactured_By": "Alma Lasers",
        "Marketing_Holder": "Alma Lasers",
    },
    "REC_0937": {
        "FDA_Status": "FDA 510(k) and CE confirmed by user; certificate numbers pending where not already captured.",
        "FDA_510k_Number": "K202880",
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "EBD",
        "Category_L2": "Plasma",
        "Tech_Type_Std": "Helium Plasma + RF",
        "Manufactured_By": "Apyx Medical",
        "Marketing_Holder": "Apyx Medical",
    },
    "REC_0936": {
        "FDA_Status": "FDA 510(k) and CE confirmed by user; CE certificate number pending.",
        "FDA_510k_Number": "K230272",
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "EBD",
        "Category_L2": "Plasma",
        "Tech_Type_Std": "Helium Plasma + RF",
        "Manufactured_By": "Apyx Medical",
        "Marketing_Holder": "Apyx Medical",
    },
    "REC_0903": {
        "FDA_Status": "FDA 510(k), CE and NMPA confirmed by user; CE/NMPA certificate numbers pending.",
        "FDA_510k_Number": "K210634",
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "NMPA_Status": "NMPA registration confirmed by user; certificate number pending.",
        "Category_L1": "EBD",
        "Category_L2": "Laser",
        "Tech_Type_Std": "Er:YAG Laser 2940nm",
        "Manufactured_By": "Asclepion Laser Technologies",
        "Marketing_Holder": "Asclepion Laser Technologies",
    },
}


DIRECT_FACT_SOURCES: dict[str, tuple[str, str, str]] = {
    "REC_0009": ("https://almalasers.com/product/accent-prime/", "Alma Accent Prime official product page", "Official page supports Accent Prime RF + ultrasound body and face contouring / skin tightening positioning."),
    "REC_0522": ("https://almalasers.com/product/opus/", "Alma Opus official product page", "Official page supports Opus fractional RF / plasma skin resurfacing and tightening positioning."),
    "REC_0675": ("https://www.alma-soprano.com/soprano-ice-platinum/", "Soprano ICE Platinum official product page", "Official page supports triple-wavelength diode laser hair removal positioning."),
    "REC_0937": ("https://eifu.apyxmedical.com/apyxmedical/en/handpieces?keycode=apyxmedical0000033", "Apyx eIFU handpieces page", "Manufacturer eIFU page retained as product identity support for J-Plasma handpieces."),
    "REC_0936": ("https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm?ID=K230272", "FDA 510(k) K230272 - Renuvion APR Handpiece", "FDA page supports Renuvion APR product identity and clearance record."),
    "REC_0903": ("https://asclepion.com/dermablate/", "Asclepion Dermablate official product page", "Official page supports Dermablate Er:YAG laser positioning."),
}


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_user_reg_indications_b4_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: object) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        new = norm(value)
        if old == new:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id, fields in WORKBOOK_UPDATES.items():
        for field, value in fields.items():
            set_cell(record_id, field, value)
        append_audit(record_id, "user_reg_indications_20260601_batch4: user-confirmed EBD indication/regulatory status applied; precise CE/NMPA certificate numbers remain blank unless supplied.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_evidence_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, specs in TARGET_SPECS.items():
        product = products.get(record_id)
        if not product:
            continue
        for spec in specs:
            rows.append(evidence_row(product, checked_at, spec))
    return rows


def build_fact_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, (url, title, excerpt) in DIRECT_FACT_SOURCES.items():
        product = products.get(record_id)
        if product:
            rows.append(fact_row(product, checked_at, url, title, excerpt))
    return rows


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()

    ind_fields, ind_rows = read_csv(MANUAL_INDICATION_PATH)
    added_indications = append_unique(
        ind_rows,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        build_evidence_rows(products, checked_at),
    )
    write_csv(MANUAL_INDICATION_PATH, ind_fields, ind_rows)

    fact_fields, fact_rows = read_csv(MANUAL_FACT_PATH)
    added_facts = append_unique(fact_rows, ["fact_id"], build_fact_rows(products, checked_at))
    write_csv(MANUAL_FACT_PATH, fact_fields, fact_rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_indications,
        "manual_product_fact_rows_added": added_facts,
        "target_record_ids": sorted(TARGET_SPECS),
        "fda_confirmed_record_ids": sorted(
            record_id for record_id, specs in TARGET_SPECS.items() if any(spec.get("jurisdiction") == "US" for spec in specs)
        ),
        "ce_confirmed_record_ids": sorted(
            record_id for record_id, specs in TARGET_SPECS.items() if any(spec.get("jurisdiction") == "EU" for spec in specs)
        ),
        "nmpa_confirmed_record_ids": sorted(
            record_id for record_id, specs in TARGET_SPECS.items() if any(spec.get("jurisdiction") == "CN" for spec in specs)
        ),
        "known_residual_notes": [
            "CE and NMPA certificate numbers remain blank where the user confirmed presence but did not supply precise numbers.",
            "Asclepion parent/brand rows were not merged; this batch updates the main Asclepion Dermablate product line REC_0903.",
            "Apyx legacy combined Renuvion (J-Plasma) row REC_0606 was not modified because separate J-Plasma and Renuvion APR rows already carry the granular claims.",
        ],
        "changed_fields_sample": workbook_changes[:80],
    }
    out_path = AUDIT_DIR / f"user_confirmed_regulatory_indications_batch4_{stamp}.json"
    latest_path = AUDIT_DIR / "user_confirmed_regulatory_indications_batch4_latest.json"
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
