from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


TZ = timezone(timedelta(hours=8))
RUN_TS = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = REPO_ROOT.parent
WORKBOOK_PATH = DATA_ROOT / "全球医美企业库_标准化版v4.xlsx"
AUDIT_DIR = REPO_ROOT / "data" / "audits"

CORRECTION_TAG = "cosmeceutical_scope_cleanup_20260602"


def norm(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(norm(part).lower() for part in parts if norm(part))
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12] if raw else "0" * 12
    return f"{prefix}_{digest}"


def backup_file(path: Path) -> Path:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".xlsx":
        backup = path.with_name(f"{path.stem}.backup_before_{CORRECTION_TAG}_{RUN_TS}{path.suffix}")
    else:
        backup = AUDIT_DIR / f"{path.stem}.backup_before_{CORRECTION_TAG}_{RUN_TS}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def sheet_header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = norm(cell.value)
        if key:
            headers[key] = idx
    return headers


def append_note(existing: Any, message: str) -> str:
    before = norm(existing)
    entry = f"{CORRECTION_TAG}: {message}"
    if entry in before:
        return before
    return " | ".join(part for part in [before, entry] if part)


EXCLUDE_RECORDS = {
    "REC_0113": "excluded: mass-market CeraVe moisturizer is supermarket/drugstore skincare, outside medical-aesthetic professional product scope.",
    "REC_0114": "excluded: Cetaphil cleanser is mass dermatological skincare / pharmacy retail line, outside medical-aesthetic professional product scope.",
    "REC_0236": "excluded: Enhel hydrogen water generator is wellness / health appliance, not medical-aesthetic upstream product line.",
    "REC_0237": "excluded: Enhel hydrogen mask is hydrogen wellness skincare, not medical-aesthetic professional product line.",
    "REC_0375": "excluded: Kamedis eczema/acne kit is OTC dermatology/pharmacy consumer care, not medical-aesthetic professional channel.",
    "REC_0511": "excluded: mixed Waldencast holding-company row combines Obagi professional skincare with Milk Makeup cosmetics; replaced by dedicated Obagi rows.",
}


PRODUCT_LINE_UPDATES: dict[str, dict[str, str]] = {
    **{
        record_id: {
            "Inclusion_Status": "excluded",
            "V4_1_Registration_Review_Status": "excluded_scope",
            "Backfill_Audit": reason,
            "Duplicate_Note": reason,
        }
        for record_id, reason in EXCLUDE_RECORDS.items()
    },
    "REC_0801": {
        "Category_L1": "Pharma",
        "Category_L2": "Topical Rx",
        "Tech_Type_Std": "Prostaglandin / Bimatoprost",
        "Verified_Product_Type_CN": "处方药 / 前列腺素睫毛增长",
        "Market_Channel": "prescription dermatology / aesthetic-adjacent",
        "Material_Taxonomy_L1_CN": "药品/皮肤科药物",
        "Material_Taxonomy_L2_CN": "外用处方药",
        "Material_Taxonomy_L3_CN": "前列腺素类",
        "Material_Taxonomy_Path_CN": "药品/皮肤科药物 > 外用处方药 > 前列腺素类",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": "LATISSE is bimatoprost ophthalmic solution, a prescription prostaglandin product; moved from skincare display bucket to Pharma.",
        "Material_Family": "前列腺素类处方药",
        "Backfill_Audit": "moved from Skincare/Cosmeceutical to Pharma/Topical Rx; prescription bimatoprost/prostaglandin product.",
    },
    "REC_0593": {
        "Category_L1": "EBD",
        "Category_L2": "Other EBD",
        "Tech_Type_Std": "Pulsed Shortwave Therapy (PSWT)",
        "Verified_Product_Type_CN": "术后恢复能量设备 / 可穿戴脉冲短波治疗",
        "Market_Channel": "post-procedure recovery / medical device",
        "Material_Taxonomy_L1_CN": "能量设备",
        "Material_Taxonomy_L2_CN": "温控/其他",
        "Material_Taxonomy_L3_CN": "脉冲短波治疗 PSWT",
        "Material_Taxonomy_Path_CN": "能量设备 > 温控/其他 > 脉冲短波治疗 PSWT",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": "RecoveryRx is a wearable pulsed shortwave therapy device for recovery/pain applications; it is an energy-based medical device, not topical skincare.",
        "Material_Family": "脉冲短波治疗 PSWT",
        "Backfill_Audit": "moved from Skincare/Cosmeceutical to EBD/Other EBD; PSWT is pulsed shortwave energy device.",
    },
    "REC_0679": {
        "Category_L1": "Skincare",
        "Category_L2": "Cosmeceutical",
        "Tech_Type_Std": "Scar Management / Silicone Gel",
        "Verified_Product_Type_CN": "术后疤痕管理 / 硅凝胶敷膜",
        "Market_Channel": "plastic surgery / post-procedure scar management",
        "Material_Taxonomy_L1_CN": "功效性护肤品",
        "Material_Taxonomy_L2_CN": "术后修复/疤痕管理",
        "Material_Taxonomy_L3_CN": "硅凝胶敷膜",
        "Material_Taxonomy_Path_CN": "功效性护肤品 > 术后修复/疤痕管理 > 硅凝胶敷膜",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": "Strataderm/Stratamed are silicone gel scar management films for post-procedure care; corrected from breast implant taxonomy.",
        "Material_Family": "硅凝胶敷膜",
        "Backfill_Audit": "kept under Skincare/Cosmeceutical but corrected tech/material from Breast Implant to Scar Management / Silicone Gel.",
    },
    "REC_0509": {
        "Category_L1": "Skincare",
        "Category_L2": "Cosmeceutical",
        "Tech_Type_Std": "Hydroquinone / Rx Cosmeceutical",
        "Verified_Product_Type_CN": "医生渠道功效护肤 / Nu-Derm 系统",
        "Market_Channel": "physician-dispensed / professional cosmeceutical",
        "Material_Taxonomy_L1_CN": "功效性护肤品",
        "Material_Taxonomy_L2_CN": "医学护肤活性",
        "Material_Taxonomy_L3_CN": "功效活性成分",
        "Material_Taxonomy_Path_CN": "功效性护肤品 > 医学护肤活性 > 功效活性成分",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": "Dedicated Obagi Nu-Derm row restored as physician-dispensed/professional cosmeceutical; mixed Waldencast/Milk row is excluded separately.",
        "Material_Family": "医生渠道功效护肤",
        "Inclusion_Status": "active",
        "Backfill_Audit": "restored dedicated Obagi Nu-Derm row as professional cosmeceutical after excluding mixed Waldencast/Milk row.",
    },
    "REC_0510": {
        "Category_L1": "Skincare",
        "Category_L2": "Cosmeceutical",
        "Tech_Type_Std": "Vitamin C / Elastin Support Cosmeceutical",
        "Verified_Product_Type_CN": "医生渠道功效护肤 / 抗氧化与弹性支持",
        "Market_Channel": "physician-dispensed / professional cosmeceutical",
        "Material_Taxonomy_L1_CN": "功效性护肤品",
        "Material_Taxonomy_L2_CN": "医学护肤活性",
        "Material_Taxonomy_L3_CN": "功效活性成分",
        "Material_Taxonomy_Path_CN": "功效性护肤品 > 医学护肤活性 > 功效活性成分",
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": "Obagi Professional-C/ELASTIderm is professional cosmeceutical skincare, not Regenerative/Cell Therapy.",
        "Material_Family": "医生渠道功效护肤",
        "Backfill_Audit": "moved from Regenerative/Cell Therapy to Skincare/Cosmeceutical; professional Obagi skincare line.",
    },
    "REC_0339": {
        "Verified_Product_Type_CN": "专业化学焕肤 / 诊所刷酸",
        "Market_Channel": "medspa / professional peel",
        "Backfill_Audit": "C-case retained: professional peel line used in medspa/clinic channel.",
    },
    "REC_0180": {
        "Verified_Product_Type_CN": "医生/美容院线功效护肤 / 真空保鲜护肤",
        "Market_Channel": "professional cosmeceutical / physician or beauty-clinic channel",
        "Backfill_Audit": "C-case retained: Dermastir is treated as professional/doctor-channel cosmeceutical rather than mass retail skincare.",
    },
    "REC_0755": {
        "Verified_Product_Type_CN": "专业美容院功效护理 / Professional Range",
        "Market_Channel": "professional salon / aesthetic-adjacent",
        "Backfill_Audit": "C-case retained with boundary note: keep only as professional salon/aesthetic-adjacent range, not mass consumer skincare.",
    },
    "REC_SKINTECH_PEEL2GLOW": {
        "Verified_Product_Type_CN": "专业品牌居家延展 / 焕肤护理",
        "Market_Channel": "professional brand home-care extension / physician-dispensed adjacency",
        "Backfill_Audit": "C-case retained: home-care extension of Skin Tech professional aesthetic portfolio, not generic retail skincare.",
    },
    "REC_SKINTECH_HAPPY_INTIM": {
        "Verified_Product_Type_CN": "专业私密护理 / 医美私密年轻化",
        "Market_Channel": "medical aesthetics / gyne-aesthetic professional channel",
        "Backfill_Audit": "C-case retained: intimate beauty line is kept only as professional aesthetic/gyne-aesthetic category.",
    },
    "REC_0080": {
        "Verified_Product_Type_CN": "重组蛋白功效护肤 / 术后修复霜",
        "Market_Channel": "professional cosmeceutical / post-procedure skincare",
        "Backfill_Audit": "C-case retained: recombinant protein topical skincare remains in cosmeceutical scope, not toxin/injectable.",
    },
}


FAMILY_UPDATES = {
    "REC_0801": {
        "product_id": "prod_d40850afd216",
        "old_family_id": "pf_6c699cec5074",
        "company": "Allergan",
        "brand": "LATISSE",
        "product_family": "Bimatoprost Eyelash Growth",
        "category_l1": "Pharma",
        "category_l2": "Topical Rx",
        "tech_type": "Prostaglandin / Bimatoprost",
    },
    "REC_0593": {
        "product_id": "prod_96179facd1ca",
        "old_family_id": "pf_32294d13bd46",
        "company": "BioElectronics",
        "brand": "RecoveryRx",
        "product_family": "Wearable PSWT Device",
        "category_l1": "EBD",
        "category_l2": "Other EBD",
        "tech_type": "Pulsed Shortwave Therapy (PSWT)",
    },
    "REC_0679": {
        "product_id": "prod_6df40f5cb83f",
        "old_family_id": "pf_b083e8f719bf",
        "company": "Stratpharma",
        "brand": "Strataderm",
        "product_family": "Strataderm / Stratamed",
        "category_l1": "Skincare",
        "category_l2": "Cosmeceutical",
        "tech_type": "Scar Management / Silicone Gel",
    },
    "REC_0509": {
        "product_id": "prod_1328390d9da3",
        "old_family_id": "pf_9105a9247ff3",
        "company": "Obagi",
        "brand": "Obagi",
        "product_family": "Nu-Derm System",
        "category_l1": "Skincare",
        "category_l2": "Cosmeceutical",
        "tech_type": "Hydroquinone / Rx Cosmeceutical",
    },
    "REC_0510": {
        "product_id": "prod_8818ec371431",
        "old_family_id": "pf_fd710695e595",
        "company": "Obagi",
        "brand": "Obagi",
        "product_family": "Professional-C / ELASTIderm",
        "category_l1": "Skincare",
        "category_l2": "Cosmeceutical",
        "tech_type": "Vitamin C / Elastin Support Cosmeceutical",
    },
}

for item in FAMILY_UPDATES.values():
    item["new_family_id"] = stable_id(
        "pf",
        item["company"],
        item["brand"],
        item["product_family"],
        item["category_l1"],
        item["category_l2"],
        item["tech_type"],
    )


COSMECEUTICAL_DEFINITION = (
    "Professional functional skincare / cosmeceuticals used in medical-aesthetic clinics, "
    "physician-dispensed channels, peri-procedure repair, scar management, professional peel "
    "or treatment-cycle support. Excludes mass supermarket/drugstore consumer skincare unless "
    "the specific product line is physician-dispensed or professional-channel."
)

CATEGORY_DEFINITION_UPDATES = {
    ("Skincare", "Cosmeceutical"): {
        "Definition_EN": COSMECEUTICAL_DEFINITION,
        "Tech_Type_Examples": "Professional functional skincare; peri-procedure repair; scar management; professional peel support; physician-dispensed cosmeceutical",
    },
    ("Skincare", "Professional"): {
        "Definition_EN": "Professional clinic, medspa, physician-dispensed or aesthetic-adjacent skincare lines; excludes purely mass-market beauty/cosmetics.",
        "Tech_Type_Examples": "Professional skincare line; clinic channel skincare; aesthetic-adjacent treatment support",
    },
    ("Skincare", "Chemical Peel"): {
        "Definition_EN": "Professional chemical peel / exfoliation products used by clinics, medspas or trained professionals.",
        "Tech_Type_Examples": "AHA peel; TCA peel; salicylic peel; professional exfoliation protocol",
    },
    ("Skincare", "Wellness"): {
        "Definition_EN": "Legacy wellness skincare bucket. Products are out of active dashboard scope unless clearly tied to medical-aesthetic professional treatment cycles.",
        "Tech_Type_Examples": "Wellness-only skincare is normally excluded_scope",
    },
}


def update_workbook() -> dict[str, Any]:
    backup = backup_file(WORKBOOK_PATH)
    wb = load_workbook(WORKBOOK_PATH)
    if "Product_Lines" not in wb.sheetnames:
        raise RuntimeError("Workbook has no Product_Lines sheet")
    ws = wb["Product_Lines"]
    headers = sheet_header_map(ws)
    if "Record_ID" not in headers:
        raise RuntimeError("Product_Lines has no Record_ID column")

    changed: dict[str, dict[str, dict[str, str]]] = {}
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        record_id = norm(row[headers["Record_ID"] - 1].value)
        updates = PRODUCT_LINE_UPDATES.get(record_id)
        if not updates:
            continue
        seen.add(record_id)
        row_changes: dict[str, dict[str, str]] = {}
        for field, value in updates.items():
            if field not in headers:
                if field == "V4_1_Registration_Review_Status":
                    continue
                raise KeyError(f"Missing Product_Lines column: {field}")
            cell = row[headers[field] - 1]
            before = "" if cell.value is None else str(cell.value)
            if field in {"Backfill_Audit", "Duplicate_Note"}:
                after = append_note(before, value)
            else:
                after = value
            cell.value = after
            after_text = "" if after is None else str(after)
            if before != after_text:
                row_changes[field] = {"before": before, "after": after_text}
        if row_changes:
            changed[record_id] = row_changes

    missing = sorted(set(PRODUCT_LINE_UPDATES) - seen)
    if missing:
        raise RuntimeError(f"Missing target rows in Product_Lines: {', '.join(missing)}")

    category_changes: dict[str, dict[str, dict[str, str]]] = {}
    if "Category_Definitions" in wb.sheetnames:
        cws = wb["Category_Definitions"]
        cheaders = sheet_header_map(cws)
        for row in cws.iter_rows(min_row=2):
            key = (
                norm(row[cheaders["Category_L1"] - 1].value) if "Category_L1" in cheaders else "",
                norm(row[cheaders["Category_L2"] - 1].value) if "Category_L2" in cheaders else "",
            )
            updates = CATEGORY_DEFINITION_UPDATES.get(key)
            if not updates:
                continue
            row_key = " / ".join(key)
            row_changes: dict[str, dict[str, str]] = {}
            for field, value in updates.items():
                if field not in cheaders:
                    continue
                cell = row[cheaders[field] - 1]
                before = "" if cell.value is None else str(cell.value)
                cell.value = value
                if before != value:
                    row_changes[field] = {"before": before, "after": value}
            if row_changes:
                category_changes[row_key] = row_changes

    wb.save(WORKBOOK_PATH)
    return {
        "path": str(WORKBOOK_PATH),
        "backup": str(backup),
        "changed_records": changed,
        "category_definition_changes": category_changes,
    }


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise RuntimeError(f"{path} has no header")
        return list(reader.fieldnames), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def matching_update(row: dict[str, str]) -> dict[str, str] | None:
    company = norm(row.get("company"))
    brand = norm(row.get("brand"))
    seed_id = norm(row.get("seed_record_id") or row.get("record_id"))
    product_id = norm(row.get("product_id") or row.get("sku_id"))
    family_id = norm(row.get("product_family_id"))
    for record_id, item in FAMILY_UPDATES.items():
        if (
            seed_id == record_id
            or product_id == item["product_id"]
            or family_id in {item["old_family_id"], item["new_family_id"]}
            or (company == item["company"] and brand == item["brand"] and family_id in {item["old_family_id"], item["new_family_id"]})
        ):
            return item
    return None


def official_source_query(row: dict[str, str], item: dict[str, str]) -> str:
    base = (
        f"\"{item['company']}\" \"{item['brand']}\" {item['product_family']} "
        f"{item['category_l1']} {item['category_l2']} {item['tech_type']}"
    )
    query_type = norm(row.get("query_type"))
    if query_type == "product_ifu_labeling":
        return f"{base} IFU instructions for use intended purpose official PDF"
    if query_type == "product_certificate_registration":
        return f"{base} certificate declaration of conformity registration official"
    return f"{base} official product page professional aesthetic source"


def mdr_ce_query(row: dict[str, str], item: dict[str, str]) -> str:
    base = f"{item['company']} {item['brand']} {item['product_family']}"
    source_key = norm(row.get("source_key"))
    evidence_target = norm(row.get("evidence_target"))
    if source_key == "eu_eudamed" or "EUDAMED" in evidence_target:
        return f"{base} EUDAMED Basic UDI-DI SRN certificate device"
    if source_key == "ce_notified_body_certificate" or "Notified-body" in evidence_target:
        return f"{base} CE MDR certificate notified body declaration conformity scope"
    if source_key == "company_ce_documents" or "IFU" in evidence_target:
        return f"{base} IFU instructions for use intended purpose declaration of conformity official PDF"
    return f"{base} CE MDR official evidence"


def update_classification_fields(row: dict[str, str], item: dict[str, str]) -> None:
    if "product_family_id" in row:
        row["product_family_id"] = item["new_family_id"]
    for field, value in [
        ("category_l1", item["category_l1"]),
        ("category_l2", item["category_l2"]),
        ("tech_type", item["tech_type"]),
    ]:
        if field in row:
            row[field] = value


def update_plan_row(row: dict[str, str], query_builder: Callable[[dict[str, str], dict[str, str]], str]) -> bool:
    item = matching_update(row)
    if not item:
        return False
    before = dict(row)
    update_classification_fields(row, item)
    if "query" in row:
        row["query"] = query_builder(row, item)
    return row != before


def update_family_only(row: dict[str, str]) -> bool:
    item = matching_update(row)
    if not item:
        return False
    before = dict(row)
    update_classification_fields(row, item)
    return row != before


def update_csv(path: Path, updater: Callable[[dict[str, str]], bool]) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path), "backup": "", "changed_rows": 0, "missing": True}
    fieldnames, rows = read_csv(path)
    changed = 0
    for row in rows:
        if updater(row):
            changed += 1
    backup = None
    if changed:
        backup = backup_file(path)
        write_csv(path, fieldnames, rows)
    return {"path": str(path), "backup": str(backup) if backup else "", "changed_rows": changed}


def main() -> None:
    expected_family_ids = {
        "REC_0801": "pf_ef9b01774c9d",
        "REC_0593": "pf_f13dd9ad5cfb",
        "REC_0679": "pf_061822ea670e",
        "REC_0509": "pf_351c7e1240da",
        "REC_0510": "pf_5191e48337f4",
    }
    for record_id, expected in expected_family_ids.items():
        actual = FAMILY_UPDATES[record_id]["new_family_id"]
        if actual != expected:
            raise RuntimeError(f"Unexpected family id for {record_id}: {actual}")

    workbook_result = update_workbook()
    csv_results = [
        update_csv(REPO_ROOT / "data" / "company_official_source_plan.csv", lambda row: update_plan_row(row, official_source_query)),
        update_csv(REPO_ROOT / "data" / "mdr_ce_search_plan.csv", lambda row: update_plan_row(row, mdr_ce_query)),
        update_csv(REPO_ROOT / "data" / "manual_product_fact_evidence.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "product_specification_evidence.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "company_media_asset_index.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "manual_evidence_promotion_log.csv", update_family_only),
    ]

    audit = {
        "run_ts": RUN_TS,
        "correction": CORRECTION_TAG,
        "scope_policy": {
            "display_label": "Cosmeceutical / 功能性护肤品",
            "schema_key": "Skincare",
            "definition": COSMECEUTICAL_DEFINITION,
        },
        "excluded_records": EXCLUDE_RECORDS,
        "changed_family_ids": {
            record_id: {
                "product_id": item["product_id"],
                "old_family_id": item["old_family_id"],
                "new_family_id": item["new_family_id"],
                "category_l1": item["category_l1"],
                "category_l2": item["category_l2"],
                "tech_type": item["tech_type"],
            }
            for record_id, item in FAMILY_UPDATES.items()
        },
        "c_case_decisions": {
            "REC_0339": "retain: professional peel / medspa channel",
            "REC_0180": "retain: doctor/professional cosmeceutical line",
            "REC_0755": "retain as boundary professional salon/aesthetic-adjacent range only",
            "REC_SKINTECH_PEEL2GLOW": "retain: home-care extension of professional Skin Tech portfolio",
            "REC_SKINTECH_HAPPY_INTIM": "retain: professional intimate/gyne-aesthetic line",
            "REC_0080": "retain: recombinant protein topical cosmeceutical, not injectable toxin",
            "REC_0375": "exclude: OTC dermatology/pharmacy consumer care",
        },
        "workbook": workbook_result,
        "csv_updates": csv_results,
    }

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    audit_path = AUDIT_DIR / f"{CORRECTION_TAG}_{RUN_TS}.json"
    latest_path = AUDIT_DIR / f"{CORRECTION_TAG}_latest.json"
    audit_json = json.dumps(audit, ensure_ascii=False, indent=2)
    audit_path.write_text(audit_json, encoding="utf-8")
    latest_path.write_text(audit_json, encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
