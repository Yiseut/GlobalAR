"""Attach user-confirmed Bimini product-line evidence to the source workbook."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER = DATA_DIR / "product_master.csv"
MANUAL_FACTS = DATA_DIR / "manual_product_fact_evidence.csv"


SOURCE_URL = "https://www.prnewswire.com/news-releases/bimini-health-tech-achieves-eu-mdr-certification-302770477.html"


SPECS = {
    "REC_1056": {
        "brand": "Dermapose",
        "title": "Bimini Health Tech Dermapose company announcement",
        "excerpt": "User confirmed Dermapose as an in-scope upstream closed microfat/autologous fat transfer suite; company announcement supports EU MDR certification and product-line identity.",
    },
    "REC_1057": {
        "brand": "Puregraft",
        "title": "Bimini Health Tech Puregraft company announcement",
        "excerpt": "User confirmed Puregraft as an in-scope upstream fat purification/filtration product suite; company announcement supports EU MDR certification and product-line identity.",
    },
}


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER)
    return {norm(row.get("seed_record_id")): row for row in rows}


def append_unique(rows: list[dict[str, str]], key_field: str, new_rows: list[dict[str, str]]) -> int:
    existing = {norm(row.get(key_field)) for row in rows}
    added = 0
    for row in new_rows:
        key = norm(row.get(key_field))
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def fact_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    return {
        "fact_id": stable_id("pfact", product.get("seed_record_id"), "bimini_user_confirmed_company_claim", SOURCE_URL),
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "product_family_id": "",
        "standard_product_name": product.get("standard_product_name", ""),
        "priority": "P0",
        "fact_group": "official_company_claim",
        "field_name": "official_company_claim",
        "field_value": SOURCE_URL,
        "source_url": SOURCE_URL,
        "evidence_title": spec["title"],
        "evidence_excerpt": spec["excerpt"],
        "source_type": "official_company_claim",
        "confidence": "user_confirmed_company_announcement",
        "captured_at": checked_at,
        "promoted_at": checked_at,
        "review_status": "user_confirmed",
        "note": "bimini_missing_product_family_feedback_20260601",
    }


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_bimini_user_confirmed_product_evidence_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if old == value:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": value})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id in SPECS:
        set_cell(record_id, "Data_Source", "official_product_fact_promoted")
        append_audit(record_id, "bimini_user_confirmed_product_evidence_20260601: user-confirmed missing product-family feedback and company announcement attached as product identity evidence.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    products = product_lookup()
    backup, changes = update_workbook(stamp)

    fields, rows = read_csv(MANUAL_FACTS)
    new_rows = [fact_row(products[record_id], checked_at, spec) for record_id, spec in SPECS.items() if record_id in products]
    added = append_unique(rows, "fact_id", new_rows)
    write_csv(MANUAL_FACTS, fields, rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(changes),
        "manual_product_fact_rows_added": added,
        "target_record_ids": sorted(SPECS),
        "source_url": SOURCE_URL,
        "changed_fields_sample": changes[:80],
    }
    out = AUDIT_DIR / f"bimini_user_confirmed_product_evidence_{stamp}.json"
    latest = AUDIT_DIR / "bimini_user_confirmed_product_evidence_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
