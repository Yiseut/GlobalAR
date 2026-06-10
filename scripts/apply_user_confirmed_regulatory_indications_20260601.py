"""Apply user-confirmed regulatory and indication notes from the 2026-06-01 QA loop."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(*parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return hashlib.sha1(blob.encode("utf-8")).hexdigest()[:12]


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def append_unique(rows: list[dict[str, str]], new_rows: list[dict[str, str]]) -> int:
    existing = {
        (
            norm(row.get("seed_record_id")),
            norm(row.get("jurisdiction")),
            norm(row.get("regulator")),
            norm(row.get("registered_name")),
            norm(row.get("source_key")),
        )
        for row in rows
    }
    added = 0
    for row in new_rows:
        key = (
            norm(row.get("seed_record_id")),
            norm(row.get("jurisdiction")),
            norm(row.get("regulator")),
            norm(row.get("registered_name")),
            norm(row.get("source_key")),
        )
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def row_for(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    indication = norm(spec.get("indication"))
    source_key = norm(spec.get("source_key")) or f"user_confirmed_regulatory_indication_20260601:{product.get('seed_record_id')}:{stable_id(spec.get('jurisdiction'), spec.get('regulator'), spec.get('registered_name'))}"
    return {
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "jurisdiction": norm(spec.get("jurisdiction")),
        "regulator": norm(spec.get("regulator")),
        "regulatory_pathway": norm(spec.get("pathway")),
        "fda_product_code": "",
        "fda_regulation_number": "",
        "fda_device_class": "",
        "fda_submission_type": "",
        "status": norm(spec.get("status")),
        "registration_no": norm(spec.get("registration_no")),
        "approval_date": norm(spec.get("approval_date")),
        "original_approval_date": "",
        "expiry_date": norm(spec.get("expiry_date")),
        "registered_name": norm(spec.get("registered_name")) or product.get("brand", ""),
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": norm(spec.get("legal_manufacturer")) or product.get("legal_manufacturer") or product.get("company", ""),
        "local_holder": norm(spec.get("local_holder")),
        "source_key": source_key,
        "source_url": norm(spec.get("source_url")),
        "source_type": norm(spec.get("source_type")) or "user_confirmed_official_claim",
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": norm(spec.get("field_note")),
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601",
        "review_status": norm(spec.get("review_status")) or "user_confirmed",
        "confidence": norm(spec.get("confidence")) or "user_confirmed_official_claim",
    }


TARGET_SPECS: dict[str, list[dict[str, str]]] = {
    "REC_0537": [
        {
            "jurisdiction": "CN",
            "regulator": "NMPA",
            "pathway": "Class III medical device registration",
            "status": "NMPA Class III registration confirmed by user; certificate number not captured",
            "registered_name": "Persnica",
            "indication": "用于面部真皮组织注射，以纠正皱纹及面部容量缺失。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub02_02.php",
            "source_type": "user_confirmed_regulatory_claim_with_official_product_page",
            "evidence_title": "Across Persnica official product page / user-confirmed NMPA status",
            "evidence_excerpt": "用户确认 Persnica 已获得中国 NMPA 三类医疗器械注册证；证书编号未捕获。Across 官方 HA filler 页面用于产品身份承接。",
            "field_note": "User confirmed NMPA registration presence; no certificate number was supplied in this batch.",
        }
    ],
    "REC_0693": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device manufacturing/product authorization",
            "status": "MFDS/KFDA authorization confirmed by official Across history and user feedback",
            "registered_name": "The Chaeum",
            "indication": "The Chaeum 系列用于面部皱纹的临时改善；不同 Pure/Premium 型号覆盖浅层真皮、中深层真皮至皮下等注射层次。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub01_03.php",
            "source_type": "official_company_history",
            "evidence_title": "Across company history - MFDS authorization for The Chaeum",
            "evidence_excerpt": "Across 官方沿革记录 The Chaeum 获 MFDS 医疗器械制造/产品授权；用户同步确认 The Chaeum 拥有韩国 KFDA 注册。",
            "field_note": "No product-specific certificate number was supplied; retain as registration-presence evidence.",
            "confidence": "official_company_history_user_confirmed",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; product-specific certificate number not captured",
            "registered_name": "The Chaeum",
            "indication": "The Chaeum Premium 3/4 等型号用于中重度至深层面部皱纹填充，并可按型号用于唇部容量和面部深层结构塑形。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub02_05.php",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Across The Chaeum official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 The Chaeum 具备欧洲 CE 认证；Across 官方产品页列明 The Chaeum Pure/Premium/Shape 系列与注射层次。",
            "field_note": "User confirmed CE presence; no product-specific CE certificate number was supplied in this batch.",
        },
    ],
    "REC_0615": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device export manufacturing approval",
            "status": "MFDS/KFDA export manufacturing approval confirmed by official Across history and user feedback",
            "registered_name": "Revolax",
            "indication": "Revolax 是透明质酸填充剂，用于皱纹改善和组织修复；Fine、Deep、Sub-Q 型号按适用层次覆盖浅层真皮至深层皮下。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub01_03.php",
            "source_type": "official_company_history",
            "evidence_title": "Across company history - MFDS approval for Dermalax/Revolax",
            "evidence_excerpt": "Across 官方沿革记录 Dermalax/Revolax 获 MFDS 医疗器械出口制造批准；用户同步确认 Revolax 具备韩国 KFDA 认证。",
            "field_note": "No product-specific certificate number was supplied; retain as registration-presence evidence.",
            "confidence": "official_company_history_user_confirmed",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; product-specific certificate number not captured",
            "registered_name": "Revolax",
            "indication": "适用于面部中重度皱纹填充、面部轮廓塑形及唇部丰盈。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub02_03.php",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Across Revolax official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Revolax 具备欧洲 CE 认证；Across 官方产品页列明 Revolax Fine/Deep/Sub-Q 产品及适用层次。",
            "field_note": "User confirmed CE presence; no product-specific CE certificate number was supplied in this batch.",
        },
    ],
    "REC_0758": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device export manufacturing approval",
            "status": "MFDS export manufacturing approval confirmed by official Across history",
            "registered_name": "Volus 10",
            "indication": "Volus 10 是 10 mL 大容量透明质酸填充剂；Across 官方页面当前列示的应用部位为 penile enlargement。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub01_03.php",
            "source_type": "official_company_history",
            "evidence_title": "Across company history - MFDS approval for Volus",
            "evidence_excerpt": "Across 官方沿革记录 Volus/Volus with lidocaine 获 MFDS 医疗器械出口制造批准。",
            "field_note": "User described broader body-contouring uses; the stronger official Across page currently lists penile enlargement, so this row keeps the narrower official wording.",
            "confidence": "official_company_history_cross_checked",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; product-specific certificate number not captured",
            "registered_name": "Volus 10",
            "indication": "专为大容量身体填充设计，用于大容量软组织容量补充和轮廓修饰；具体部位需以当地说明书和医生适应证判断为准。",
            "legal_manufacturer": "Across Co., Ltd.",
            "source_url": "https://www.across.kr/eng/sub/sub02_04.php",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Across Volus 10 official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Volus 10 获欧洲 CE 认证；Across 官方产品页列明 Volus 10 为 10 mL 大容量 HA filler。",
            "field_note": "User confirmed CE presence; no product-specific CE certificate number was supplied in this batch.",
        },
    ],
    "REC_0163": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; product-specific certificate number not captured",
            "registered_name": "DeneB Classic-S / Classic-H",
            "indication": "DeneB Classic-S 偏向面部深层皱纹矫正及组织容量扩充；Classic-H 主要用于大容量身体塑形，如丰臀、非手术丰胸及身体轮廓雕塑。",
            "legal_manufacturer": "BioPlus Co., Ltd.",
            "source_url": "https://bioplus-deneb.com/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "DeneB official product site / user-confirmed CE status",
            "evidence_excerpt": "用户确认 DeneB Classic-S/Classic-H 获欧洲 CE 认证，并补充面部与身体大容量填充的产品适应症定位。",
            "field_note": "User confirmed CE presence; no product-specific CE certificate number was supplied in this batch.",
        }
    ],
    "REC_0934": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; product-specific certificate number not captured",
            "registered_name": "Revofil / Aquashine",
            "indication": "Revofil Ultra 等型号适用于深层皱纹修复及面颊、下巴等轮廓提升；Aquashine 细分系列作为轻度填充/水光产品，用于改善细纹、暗沉并提升皮肤水合度与弹性。",
            "legal_manufacturer": "Caregen Co., Ltd.",
            "source_url": "http://caregen.co.kr/eng/news/?vid=11",
            "source_type": "user_confirmed_ce_claim_with_official_caregen_reference",
            "evidence_title": "Caregen Revofil / Aquashine user-confirmed CE status",
            "evidence_excerpt": "用户确认 Revofil/Aquashine 全系产品获得欧洲 CE 认证，并补充 Revofil Ultra 与 Aquashine 系列的适应症定位。",
            "field_note": "User confirmed CE presence; no product-specific CE certificate number was supplied in this batch.",
        }
    ],
    "REC_0449": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; product-specific certificate number not captured",
            "registered_name": "Meso-Xanthin F199",
            "indication": "偏向再生医学的中胚层疗法/水光产品，用于抗光老化、改善色素沉着、面部细纹及整体皮肤抗氧化修复。",
            "legal_manufacturer": "ABG Lab LLC",
            "source_url": "https://mesowhartonp199.com/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "ABG Lab Meso-Xanthin F199 user-confirmed CE status",
            "evidence_excerpt": "用户确认 Meso-Xanthin F199 获欧洲 CE 认证，并补充其再生医学/中胚层疗法定位。",
            "field_note": "User confirmed CE presence; no product-specific CE certificate number was supplied in this batch.",
        }
    ],
    "REC_0178": [
        {
            "jurisdiction": "EU",
            "regulator": "Notified Body (CE 0123)",
            "pathway": "CE Class III medical device",
            "status": "CE0123 Class III medical device confirmed by official Dermastir product pages",
            "registration_no": "CE 0123 (Class III; no product-specific certificate number captured)",
            "registered_name": "Dermastir 16 EBF / 32 EBF",
            "indication": "Dermastir 16 EBF / 32 EBF 注射瓶用于深层真皮补水、改善细纹和初生皱纹、维持水合、致密度和光泽，并可改善皮肤弹性、肤色和整体肤质。",
            "legal_manufacturer": "Alta Care",
            "source_url": "https://dermastir.com/products/dermastir-16ebf",
            "source_type": "official_product_page",
            "evidence_title": "Dermastir 16EBF / 32EBF official product page",
            "evidence_excerpt": "Dermastir 官方产品页列明 16EBF/32EBF 为 CE0123 医疗器械，并给出补水、改善细纹、弹性和肤质等用途。",
            "field_note": "Official product pages confirm CE0123 and intended use; no product-specific certificate number was supplied.",
            "confidence": "official_product_page_cross_checked",
        }
    ],
}


WORKBOOK_STATUS = {
    "REC_0537": "NMPA Class III registration confirmed by user; certificate number pending.",
    "REC_0693": "KFDA/MFDS registered and CE certified; certificate numbers pending.",
    "REC_0615": "KFDA/MFDS registered and CE certified; certificate numbers pending.",
    "REC_0758": "CE certified; MFDS export manufacturing approval confirmed from official Across history.",
    "REC_0163": "CE certified; certificate number pending.",
    "REC_0934": "CE certified; certificate number pending.",
    "REC_0449": "CE certified; certificate number pending.",
    "REC_0178": "CE 0123 Class III medical device; official intended use captured.",
}

CHAEUM_DUPLICATE_ROWS = {
    "REC_0694": "duplicate_of:REC_0693; The Chaeum active product family is kept under Across as legal/manufacturing subject; Hugel retained as parent/portfolio trace.",
    "REC_0877": "duplicate_of:REC_0693; The Chaeum Premium No.1 should be treated as SKU/variant under Across The Chaeum, not a separate Hugel product line.",
    "REC_0878": "duplicate_of:REC_0693; The Chaeum Premium No.2 should be treated as SKU/variant under Across The Chaeum, not a separate Hugel product line.",
    "REC_0879": "duplicate_of:REC_0693; The Chaeum Premium No.3 should be treated as SKU/variant under Across The Chaeum, not a separate Hugel product line.",
    "REC_0880": "duplicate_of:REC_0693; The Chaeum Premium No.4 should be treated as SKU/variant under Across The Chaeum, not a separate Hugel product line.",
}


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_user_reg_indications_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: object) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        new = norm(value)
        if old == new:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id, status in WORKBOOK_STATUS.items():
        set_cell(record_id, "CE_Status", status)
        append_audit(record_id, f"user_reg_indications_20260601: user/official confirmation applied; precise certificate number remains blank unless supplied.")

    for record_id in ["REC_0537", "REC_0693", "REC_0615", "REC_0758"]:
        set_cell(record_id, "Manufactured_By", "Across Co., Ltd.")
    for record_id in ["REC_0537", "REC_0693", "REC_0615", "REC_0758"]:
        set_cell(record_id, "Marketing_Holder", "Across Co., Ltd.")
    set_cell("REC_0163", "Manufactured_By", "BioPlus Co., Ltd.")
    set_cell("REC_0934", "Manufactured_By", "Caregen Co., Ltd.")
    set_cell("REC_0449", "Manufactured_By", "ABG Lab LLC")
    set_cell("REC_0178", "Manufactured_By", "Alta Care")

    for record_id, note in CHAEUM_DUPLICATE_ROWS.items():
        set_cell(record_id, "Is_Primary_Record", 0)
        set_cell(record_id, "Duplicate_Note", note)
        append_audit(record_id, "user_reg_indications_20260601: collapsed into Across The Chaeum product family for active dashboard/product-gap scope.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, specs in TARGET_SPECS.items():
        product = products.get(record_id)
        if not product:
            continue
        for spec in specs:
            rows.append(row_for(product, checked_at, spec))
    return rows


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()
    fields, rows = read_csv(MANUAL_INDICATION_PATH)
    new_rows = build_rows(products, checked_at)
    added_rows = append_unique(rows, new_rows)
    write_csv(MANUAL_INDICATION_PATH, fields, rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_rows,
        "target_record_ids": sorted(TARGET_SPECS),
        "collapsed_chaeum_duplicate_record_ids": sorted(CHAEUM_DUPLICATE_ROWS),
        "changed_fields_sample": workbook_changes[:40],
    }
    out_path = AUDIT_DIR / f"user_confirmed_regulatory_indications_{stamp}.json"
    latest_path = AUDIT_DIR / "user_confirmed_regulatory_indications_latest.json"
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
