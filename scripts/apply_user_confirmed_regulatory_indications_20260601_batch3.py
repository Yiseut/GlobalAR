"""Apply user-confirmed liquid PCL and thread-lift regulatory notes from 2026-06-01."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_FACT_PATH = DATA_DIR / "manual_product_fact_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_fields: list[str], new_rows: list[dict[str, str]]) -> int:
    existing = {tuple(norm(row.get(field)) for field in key_fields) for row in rows}
    added = 0
    for row in new_rows:
        key = tuple(norm(row.get(field)) for field in key_fields)
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def registration_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    indication = norm(spec.get("indication"))
    source_key = norm(spec.get("source_key")) or stable_id(
        "user_confirmed_reg_ind_20260601_b3",
        product.get("seed_record_id"),
        spec.get("jurisdiction"),
        spec.get("regulator"),
        spec.get("registered_name"),
    )
    return {
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "jurisdiction": norm(spec.get("jurisdiction")),
        "regulator": norm(spec.get("regulator")),
        "regulatory_pathway": norm(spec.get("pathway")),
        "fda_product_code": norm(spec.get("fda_product_code")),
        "fda_regulation_number": norm(spec.get("fda_regulation_number")),
        "fda_device_class": norm(spec.get("fda_device_class")),
        "fda_submission_type": norm(spec.get("fda_submission_type")),
        "status": norm(spec.get("status")),
        "registration_no": norm(spec.get("registration_no")),
        "approval_date": norm(spec.get("approval_date")),
        "expiry_date": norm(spec.get("expiry_date")),
        "registered_name": norm(spec.get("registered_name")) or product.get("brand", ""),
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": norm(spec.get("legal_manufacturer")) or product.get("legal_manufacturer") or product.get("company", ""),
        "local_holder": norm(spec.get("local_holder")),
        "source_key": source_key,
        "source_url": norm(spec.get("source_url")),
        "source_type": norm(spec.get("source_type")) or "user_confirmed_official_claim",
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": norm(spec.get("field_note")),
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601_batch3",
        "review_status": norm(spec.get("review_status")) or "user_confirmed",
        "confidence": norm(spec.get("confidence")) or "user_confirmed_official_claim",
    }


def fact_row(product: dict[str, str], checked_at: str, source_url: str, title: str, excerpt: str) -> dict[str, str]:
    return {
        "fact_id": stable_id("pfact", product.get("seed_record_id"), "official_product_page", source_url),
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "product_family_id": "",
        "standard_product_name": product.get("standard_product_name", ""),
        "priority": "P0",
        "fact_group": "official_product_page",
        "field_name": "official_product_page",
        "field_value": source_url,
        "source_url": source_url,
        "evidence_title": title,
        "evidence_excerpt": excerpt,
        "source_type": "official_product_page",
        "confidence": "official_site_user_confirmed",
        "captured_at": checked_at,
        "promoted_at": checked_at,
        "review_status": "auto_cross_checked",
        "note": "user_confirmed_regulatory_indications_20260601_batch3",
    }


TARGET_SPECS: dict[str, list[dict[str, str]]] = {
    "REC_0307": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "GOURI",
            "indication": "GOURI 是全液态 PCL 生物刺激剂，不含微粒，注射后可在皮下自然扩散，用于全面部胶原蛋白新生、提升皮肤弹性、全面部抗衰和肤质改善。",
            "legal_manufacturer": "Dexlevo",
            "source_url": "https://gorgeousgouri.com/why_gouri",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "GOURI official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 GOURI 获得欧洲 CE 认证，并补充其全液态 PCL 胶原生物刺激剂定位。",
            "field_note": "User confirmed CE presence only; Korea MFDS/KFDA status still requires separate confirmation.",
        }
    ],
    "REC_0459": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Miracle L / Miracle H",
            "indication": "Miracle L/H 同属液态 PCL 技术；Miracle L 为纯液态 PCL，主打胶原再生与紧致，Miracle H 结合 HA，在刺激胶原的同时兼具深层补水和水光效果。",
            "legal_manufacturer": "Dexlevo",
            "source_url": "https://dexlevoaesthetic.com/27",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Dexlevo Aesthetic Miracle page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Miracle L/H 获得欧洲 CE 认证，并补充液态 PCL 与 HA 组合定位。",
            "field_note": "User confirmed CE presence only; Korea MFDS/KFDA status still requires separate confirmation.",
        }
    ],
    "REC_0039": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Aptos Nano",
            "indication": "Aptos Nano 是用于精细部位的细微提拉线，面向眼周、唇周等皮肤较薄区域的紧致、平滑和细纹改善。",
            "legal_manufacturer": "Aptos",
            "source_url": "https://aptos.global/index.php/product/34/nano-excellence-method",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Aptos Nano Excellence Method official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Aptos Nano 拥有欧洲 CE 认证；官方页面用于承接产品身份。",
        }
    ],
    "REC_0040": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE and multi-country medical device registrations",
            "status": "CE and multi-country registrations confirmed by user; certificate numbers not captured",
            "registered_name": "Aptos Wire / Needle",
            "indication": "Aptos Wire/Needle 为聚丙烯不可吸收外科级线材，提供强组织锚定力，主要用于中重度面部下垂、法令纹、下颌缘松弛和颈部提升的持久悬吊复位。",
            "legal_manufacturer": "Aptos",
            "source_url": "https://aptos.cz/en/product/light-lift-needle-method",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Aptos Light Lift Needle Method official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Aptos Wire/Needle 拥有欧洲 CE 及多国医疗器械注册；官方页面用于承接产品身份。",
        }
    ],
    "REC_0692": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "Tesslift / Tesslift Soft",
            "indication": "Tesslift/Tesslift Soft 采用 3D 网状结构加中心倒刺设计，植入后促进组织长入网孔以形成组织融合和固定，主要用于中下面部提升、下颌线精雕及眉眼提升。",
            "legal_manufacturer": "Tesslift",
            "source_url": "https://tesslift.com/",
            "source_type": "user_confirmed_kfda_claim_with_official_product_page",
            "evidence_title": "Tesslift official page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 Tesslift 拥有韩国 KFDA 与欧洲 CE 认证；官方页面用于承接产品身份。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Tesslift / Tesslift Soft",
            "indication": "Tesslift/Tesslift Soft 是具有 3D 网状结构和中心倒刺设计的高阶提拉线，用于中下面部、下颌线及眉眼区域微创提升。",
            "legal_manufacturer": "Tesslift",
            "source_url": "https://tesslift.com/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Tesslift official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Tesslift 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
    ],
    "REC_0676": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Spider Web Aesthetics",
            "indication": "Spider Web 是线雕技术/耗材，通过多根平滑线在皮下交织成网状支架，刺激胶原蛋白和弹性蛋白生成，用于面部、颈部及身体的非手术紧致与回春。",
            "legal_manufacturer": "Nexgen Biopharma",
            "source_url": "https://nexgenbiopharma.com/",
            "source_type": "user_confirmed_ce_claim_with_corporate_page",
            "evidence_title": "Nexgen Biopharma corporate page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Spider Web 拥有欧洲 CE 认证；当前本地证据为公司官网承接，仍需后续补强产品直链。",
            "field_note": "Corporate URL used because no stronger official product URL was captured in the local evidence set.",
        }
    ],
    "REC_0721": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "Ultra V Lift",
            "indication": "Ultra V Lift 覆盖 Hiko 与 Cog 等线型；Hiko 用于非手术鼻部塑形，Cog 用于中下面部深层物理提拉。",
            "legal_manufacturer": "Ultra V",
            "source_url": "https://www.ultravmed.com/",
            "source_type": "user_confirmed_kfda_claim_with_official_product_page",
            "evidence_title": "Ultra V Medical official page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 Ultra V Lift 拥有韩国 KFDA、欧洲 CE，且部分型号获 NMPA 注册。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Ultra V Lift",
            "indication": "Ultra V Lift 的 Hiko 用于非手术鼻部塑形，Cog 用于中下面部深层物理提拉。",
            "legal_manufacturer": "Ultra V",
            "source_url": "https://www.ultravmed.com/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Ultra V Medical official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Ultra V Lift 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
        {
            "jurisdiction": "CN",
            "regulator": "NMPA",
            "pathway": "Imported Class III medical device registration",
            "status": "Partial NMPA registration confirmed by user; exact model/certificate number not captured",
            "registered_name": "Ultra V Lift",
            "indication": "Ultra V Lift 部分型号在中国获 NMPA 注册；产品用于线雕提拉、鼻部塑形和面部轮廓改善，具体获批型号待证号级核验。",
            "legal_manufacturer": "Ultra V",
            "source_url": "https://www.ultravmed.com/",
            "source_type": "user_confirmed_nmpa_claim_with_official_product_page",
            "evidence_title": "Ultra V Medical official page / user-confirmed partial NMPA status",
            "evidence_excerpt": "用户确认 Ultra V Lift 部分型号获 NMPA 注册；本批未提供具体 NMPA 注册证号。",
            "field_note": "Treat as partial China-registration tag until the exact imported certificate number is captured.",
        },
    ],
    "REC_0353": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "i-THREAD Matrix / Cog",
            "indication": "i-THREAD 提供多种 PDO 线型；Matrix 平滑/网状线用于胶原再生和紧致，Cog 用于面部轮廓复位。",
            "legal_manufacturer": "Healux",
            "source_url": "https://www.healuxgroup.com/ithread",
            "source_type": "user_confirmed_kfda_claim_with_official_product_page",
            "evidence_title": "Healux i-THREAD official page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 i-THREAD 拥有韩国 KFDA、欧洲 CE 及美国 FDA 510(k) 许可。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "i-THREAD Matrix / Cog",
            "indication": "i-THREAD Matrix 用于胶原再生和紧致，Cog 用于面部轮廓复位。",
            "legal_manufacturer": "Healux",
            "source_url": "https://www.healuxgroup.com/ithread",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Healux i-THREAD official page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 i-THREAD 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA 510(k) clearance confirmed by user; 510(k) number not captured",
            "registered_name": "i-THREAD Matrix / Cog",
            "indication": "i-THREAD 作为 PDO 提拉线用于面部和身体线雕操作；本行仅保留用户确认的 FDA 510(k) 存在状态，具体法定适应症待 510(k) 号核验。",
            "legal_manufacturer": "Healux",
            "source_url": "https://www.healuxgroup.com/ithread",
            "source_type": "user_confirmed_fda510k_claim_with_official_product_page",
            "evidence_title": "Healux i-THREAD official page / user-confirmed FDA 510(k) status",
            "evidence_excerpt": "用户确认 i-THREAD 拥有美国 FDA 510(k) 许可；本批未提供具体 510(k) 编号。",
            "field_note": "FDA row intentionally keeps the number blank until a regulator record is captured.",
        },
    ],
    "REC_0775": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "X-Lift / X-Line",
            "indication": "X-Lift/X-Line 覆盖 PDO、PLLA、PCL 材质，以压印成型 Molding Cog 高拉力锯齿线为特征，用于全脸强效提拉。",
            "legal_manufacturer": "Grand Aespio",
            "source_url": "https://aespio.com/products/",
            "source_type": "user_confirmed_kfda_claim_with_official_product_page",
            "evidence_title": "Grand Aespio product page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 X-Lift/X-Line 拥有韩国 KFDA 与欧洲 CE 认证；官方产品页用于承接产品身份。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "X-Lift / X-Line",
            "indication": "X-Lift/X-Line 为多材质 Molding Cog 提拉线，用于全脸强效提拉。",
            "legal_manufacturer": "Grand Aespio",
            "source_url": "https://aespio.com/products/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Grand Aespio product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 X-Lift/X-Line 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
    ],
    "REC_0487": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "Nesfill / Reanzen Threads",
            "indication": "Nesfill/Reanzen Threads 是综合性 PDO 埋线产品，用于面颊下垂提升、额头提拉、改善法令纹及颈纹。",
            "legal_manufacturer": "Reanzen",
            "source_url": "https://www.reanzen.com",
            "source_type": "user_confirmed_kfda_claim",
            "evidence_title": "User-confirmed Reanzen thread-lift registration status",
            "evidence_excerpt": "用户确认 Reanzen Nesfill/Reanzen Threads 拥有韩国 KFDA 与欧洲 CE 认证；本批未捕获可用产品直链或证书编号。",
            "field_note": "Keep as user-confirmed regulatory presence; official product URL remains a follow-up gap.",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Nesfill / Reanzen Threads",
            "indication": "Nesfill/Reanzen Threads 是综合性 PDO 埋线产品，用于面部和颈部线雕提升与皱纹改善。",
            "legal_manufacturer": "Reanzen",
            "source_url": "https://www.reanzen.com",
            "source_type": "user_confirmed_ce_claim",
            "evidence_title": "User-confirmed Reanzen thread-lift CE status",
            "evidence_excerpt": "用户确认 Reanzen Nesfill/Reanzen Threads 拥有欧洲 CE 认证；本批未捕获具体证书编号。",
            "field_note": "Official product URL remains a follow-up gap.",
        },
    ],
    "REC_0532": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "PDO / PCL Threads",
            "indication": "Dermakor PDO/PCL Threads 包含 PDO 短线和 PCL 长效线；PDO 短线用于肤质改善，PCL 长效线用于深层提升、组织松垂和轮廓模糊管理。",
            "legal_manufacturer": "Dermakor",
            "source_url": "https://dermakor.com/",
            "source_type": "user_confirmed_kfda_claim_with_official_company_page",
            "evidence_title": "Dermakor official company page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 Dermakor PDO/PCL Threads 拥有韩国 KFDA 与欧洲 CE 认证；Dermakor 官网列示 MESOTHREADS 及 PDO/PCL/PLCL/PLLA 相关产品资料。",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "PDO / PCL Threads",
            "indication": "Dermakor PDO/PCL Threads 用于肤质改善、深层提升及轮廓改善，覆盖 PDO 短线与 PCL 长效线。",
            "legal_manufacturer": "Dermakor",
            "source_url": "https://dermakor.com/",
            "source_type": "user_confirmed_ce_claim_with_official_company_page",
            "evidence_title": "Dermakor official company page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Dermakor PDO/PCL Threads 拥有欧洲 CE 认证；未提供具体证书编号。",
        },
    ],
    "REC_0294": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "The Lift / PLLA Threads",
            "indication": "GCS The Lift/PLLA Threads 采用 PLLA 材质，除物理提拉外，着重于 18-24 个月周期内的自身胶原大量再生。",
            "legal_manufacturer": "GCS",
            "source_url": "https://www.gcsbio.com/eng/brands/brands.asp?seq=17",
            "source_type": "user_confirmed_kfda_claim_with_official_product_page",
            "evidence_title": "GCS The Lift official product page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 GCS The Lift/PLLA Threads 拥有韩国 KFDA 认证；官方产品页用于承接产品身份。",
        }
    ],
}


WORKBOOK_UPDATES: dict[str, dict[str, str]] = {
    "REC_0307": {
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Biostimulator",
        "Tech_Type_Std": "Solubilized PCL",
        "Manufactured_By": "Dexlevo",
        "Marketing_Holder": "Dexlevo",
    },
    "REC_0459": {
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Biostimulator",
        "Tech_Type_Std": "Solubilized PCL / HA",
        "Manufactured_By": "Dexlevo",
        "Marketing_Holder": "Dexlevo",
    },
    "REC_0039": {
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "P(LA/CL) Fine Thread",
        "Manufactured_By": "Aptos",
        "Marketing_Holder": "Aptos",
    },
    "REC_0040": {
        "CE_Status": "CE and multi-country registrations confirmed by user; certificate numbers pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "Polypropylene Thread",
        "Manufactured_By": "Aptos",
        "Marketing_Holder": "Aptos",
    },
    "REC_0692": {
        "CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.",
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "3D Mesh Thread",
        "Manufactured_By": "Tesslift",
        "Marketing_Holder": "Tesslift",
    },
    "REC_0676": {
        "CE_Status": "CE confirmed by user; certificate number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "Spider Web Smooth Thread",
        "Manufactured_By": "Nexgen Biopharma",
        "Marketing_Holder": "Nexgen Biopharma",
    },
    "REC_0721": {
        "CE_Status": "KFDA/MFDS and CE confirmed by user; some NMPA models confirmed, numbers pending.",
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "NMPA_Status": "Partial NMPA registration confirmed by user; exact model/certificate number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "PDO / Cog Thread",
        "Manufactured_By": "Ultra V",
        "Marketing_Holder": "Ultra V",
    },
    "REC_0353": {
        "FDA_Status": "FDA 510(k) clearance confirmed by user; 510(k) number pending.",
        "CE_Status": "KFDA/MFDS, CE and FDA 510(k) confirmed by user; certificate numbers pending.",
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "PDO Thread",
        "Manufactured_By": "Healux",
        "Marketing_Holder": "Healux",
    },
    "REC_0775": {
        "CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.",
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "PDO / PLLA / PCL Molding Cog Thread",
        "Manufactured_By": "Grand Aespio",
        "Marketing_Holder": "Grand Aespio",
    },
    "REC_0487": {
        "CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.",
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "PDO Thread",
        "Core_Product": "Nesfill / Reanzen Threads",
        "Manufactured_By": "Reanzen",
        "Marketing_Holder": "Reanzen",
    },
    "REC_0532": {
        "CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.",
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "PDO / PCL Thread",
        "Core_Product": "PDO / PCL Threads",
        "Manufactured_By": "Dermakor",
        "Marketing_Holder": "Dermakor",
    },
    "REC_0294": {
        "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.",
        "Category_L1": "Injectables",
        "Category_L2": "Thread Lift",
        "Tech_Type_Std": "PLLA Thread",
        "Manufactured_By": "GCS",
        "Marketing_Holder": "GCS",
    },
}


DIRECT_FACT_SOURCES: dict[str, tuple[str, str, str]] = {
    "REC_0307": ("https://gorgeousgouri.com/why_gouri", "GOURI official product page", "Official GOURI page used as product identity support for the user-confirmed liquid PCL positioning and CE presence."),
    "REC_0459": ("https://dexlevoaesthetic.com/27", "Dexlevo Aesthetic Miracle page", "Official Dexlevo Aesthetic page used as product identity support for the user-confirmed Miracle L/H liquid PCL positioning."),
    "REC_0039": ("https://aptos.global/index.php/product/34/nano-excellence-method", "Aptos Nano Excellence Method official page", "Official Aptos page used as product identity support for the user-confirmed Nano thread positioning."),
    "REC_0040": ("https://aptos.cz/en/product/light-lift-needle-method", "Aptos Light Lift Needle Method official page", "Official Aptos page used as product identity support for the user-confirmed polypropylene thread positioning."),
    "REC_0692": ("https://tesslift.com/", "Tesslift official page", "Official Tesslift page used as product identity support for user-confirmed KFDA/CE status."),
    "REC_0676": ("https://nexgenbiopharma.com/", "Nexgen Biopharma official company page", "Corporate page retained for Spider Web until a stronger official product URL is captured."),
    "REC_0721": ("https://www.ultravmed.com/", "Ultra V Medical official page", "Official Ultra V page used as product identity support for user-confirmed KFDA/CE/NMPA status."),
    "REC_0353": ("https://www.healuxgroup.com/ithread", "Healux i-THREAD official page", "Official Healux page used as product identity support for user-confirmed KFDA/CE/FDA status."),
    "REC_0775": ("https://aespio.com/products/", "Grand Aespio products page", "Official Grand Aespio page used as product identity support for user-confirmed KFDA/CE status."),
    "REC_0532": ("https://dermakor.com/", "Dermakor official company page", "Official Dermakor page lists MESOTHREADS and supports company identity for user-confirmed PDO/PCL thread line."),
    "REC_0294": ("https://www.gcsbio.com/eng/brands/brands.asp?seq=17", "GCS Art Lift / suture official product page", "Official GCS suture page used as product identity support for user-confirmed KFDA status."),
}


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_user_reg_indications_b3_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: object) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        new = norm(value)
        if old == new:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id, fields in WORKBOOK_UPDATES.items():
        for field, value in fields.items():
            set_cell(record_id, field, value)
        append_audit(record_id, "user_reg_indications_20260601_batch3: user-confirmed indication/regulatory status applied; precise certificate numbers remain blank unless supplied.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_registration_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, specs in TARGET_SPECS.items():
        product = products.get(record_id)
        if not product:
            continue
        for spec in specs:
            rows.append(registration_row(product, checked_at, spec))
    return rows


def build_fact_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, (url, title, excerpt) in DIRECT_FACT_SOURCES.items():
        product = products.get(record_id)
        if product:
            rows.append(fact_row(product, checked_at, url, title, excerpt))
    return rows


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()

    ind_fields, ind_rows = read_csv(MANUAL_INDICATION_PATH)
    added_indications = append_unique(
        ind_rows,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        build_registration_rows(products, checked_at),
    )
    write_csv(MANUAL_INDICATION_PATH, ind_fields, ind_rows)

    fact_fields, fact_rows = read_csv(MANUAL_FACT_PATH)
    added_facts = append_unique(fact_rows, ["fact_id"], build_fact_rows(products, checked_at))
    write_csv(MANUAL_FACT_PATH, fact_fields, fact_rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_indications,
        "manual_product_fact_rows_added": added_facts,
        "target_record_ids": sorted(TARGET_SPECS),
        "korea_confirmed_record_ids": sorted(
            record_id
            for record_id, specs in TARGET_SPECS.items()
            if any(spec.get("jurisdiction") == "KR" for spec in specs)
        ),
        "ce_confirmed_record_ids": sorted(
            record_id
            for record_id, specs in TARGET_SPECS.items()
            if any(spec.get("jurisdiction") == "EU" for spec in specs)
        ),
        "classification_corrected_record_ids": sorted(
            record_id
            for record_id, fields in WORKBOOK_UPDATES.items()
            if "Category_L2" in fields or "Tech_Type_Std" in fields
        ),
        "known_residual_notes": [
            "Dexlevo GOURI and Miracle have user-confirmed CE rows only; Korea MFDS/KFDA remains unconfirmed unless supplied separately.",
            "Reanzen has user-confirmed KFDA/CE rows, but no strong official product URL was captured in local/web evidence during this batch.",
            "Nexgen Spider Web currently uses a corporate URL rather than a direct product URL.",
        ],
        "changed_fields_sample": workbook_changes[:80],
    }
    out_path = AUDIT_DIR / f"user_confirmed_regulatory_indications_batch3_{stamp}.json"
    latest_path = AUDIT_DIR / "user_confirmed_regulatory_indications_batch3_latest.json"
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
