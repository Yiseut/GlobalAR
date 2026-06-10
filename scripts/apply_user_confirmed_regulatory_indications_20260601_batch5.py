"""Apply user-confirmed PN/PDRN, exosome, HA filler, and thread notes from 2026-06-01."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_FACT_PATH = DATA_DIR / "manual_product_fact_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_fields: list[str], new_rows: list[dict[str, str]]) -> int:
    existing = {tuple(norm(row.get(field)) for field in key_fields) for row in rows}
    added = 0
    for row in new_rows:
        key = tuple(norm(row.get(field)) for field in key_fields)
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def evidence_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    indication = norm(spec.get("indication"))
    source_key = norm(spec.get("source_key")) or stable_id(
        "user_confirmed_reg_ind_20260601_b5",
        product.get("seed_record_id"),
        spec.get("jurisdiction"),
        spec.get("regulator"),
        spec.get("registered_name"),
        spec.get("registration_no"),
    )
    return {
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "jurisdiction": norm(spec.get("jurisdiction")),
        "regulator": norm(spec.get("regulator")),
        "regulatory_pathway": norm(spec.get("pathway")),
        "status": norm(spec.get("status")),
        "registration_no": norm(spec.get("registration_no")),
        "approval_date": norm(spec.get("approval_date")),
        "expiry_date": norm(spec.get("expiry_date")),
        "registered_name": norm(spec.get("registered_name")) or product.get("brand", ""),
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": norm(spec.get("legal_manufacturer")) or product.get("legal_manufacturer") or product.get("company", ""),
        "local_holder": norm(spec.get("local_holder")),
        "source_key": source_key,
        "source_url": norm(spec.get("source_url")),
        "source_type": norm(spec.get("source_type")) or "user_confirmed_official_claim",
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": norm(spec.get("field_note")),
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601_batch5",
        "review_status": norm(spec.get("review_status")) or "user_confirmed",
        "confidence": norm(spec.get("confidence")) or "user_confirmed_official_claim",
    }


def fact_row(product: dict[str, str], checked_at: str, source_url: str, title: str, excerpt: str) -> dict[str, str]:
    return {
        "fact_id": stable_id("pfact", product.get("seed_record_id"), "official_product_page", source_url),
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "product_family_id": "",
        "standard_product_name": product.get("standard_product_name", ""),
        "priority": "P0",
        "fact_group": "official_product_page",
        "field_name": "official_product_page",
        "field_value": source_url,
        "source_url": source_url,
        "evidence_title": title,
        "evidence_excerpt": excerpt,
        "source_type": "official_product_page",
        "confidence": "official_site_user_confirmed",
        "captured_at": checked_at,
        "promoted_at": checked_at,
        "review_status": "auto_cross_checked",
        "note": "user_confirmed_regulatory_indications_20260601_batch5",
    }


def reg(jurisdiction: str, regulator: str, pathway: str, status: str, registered_name: str, indication: str, manufacturer: str, source_url: str, title: str, excerpt: str, field_note: str = "") -> dict[str, str]:
    return {
        "jurisdiction": jurisdiction,
        "regulator": regulator,
        "pathway": pathway,
        "status": status,
        "registered_name": registered_name,
        "indication": indication,
        "legal_manufacturer": manufacturer,
        "source_url": source_url,
        "source_type": "user_confirmed_regulatory_claim_with_official_product_page",
        "evidence_title": title,
        "evidence_excerpt": excerpt,
        "field_note": field_note,
    }


REJURAN_SOURCE = "https://pharmaresearch.co.kr/en/product/list.html?code=2"
REJURAN_INDICATION_MAIN = "Rejuran 为 PharmaResearch / PR Bio 的高浓度 PN 再生注射剂；Healer 用于全面部真皮层环境修复、改善细纹与红血丝，HB Plus 结合 PN、HA 与利多卡因，兼顾再生、深层补水和减痛。"
REJURAN_INDICATION_EYE = "Rejuran i / Eyes 为 PharmaResearch / PR Bio 的 PN 眼周再生注射剂，面向眼周极薄皮肤，用于改善黑眼圈、眼周初老、细纹和局部肤质。"
REJURAN_INDICATION_SHINE = "Rejuran Shine 属于 Rejuran 家族的 PN/HA 复合皮肤动能素，用于真皮层修复、补水、细纹改善和肤质提升。"


TARGET_SPECS: dict[str, list[dict[str, str]]] = {
    "REC_0601": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "Rejuran Healer / HB Plus", REJURAN_INDICATION_MAIN, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 原研与注册主体为 PharmaResearch / PR Bio，拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "Rejuran Healer / HB Plus", REJURAN_INDICATION_MAIN, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 拥有欧洲 CE 认证；未提供具体证书编号。"),
        reg("CN", "NMPA", "imported medical device registration", "NMPA registration confirmed for some Rejuran models by user; exact certificate/model mapping pending", "Rejuran family", REJURAN_INDICATION_MAIN, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 部分型号获中国 NMPA 及多国批准；本批未提供逐型号证号。", "Do not assign a China certificate number until exact model-level NMPA evidence is captured."),
    ],
    "REC_0602": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "Rejuran i / Eyes", REJURAN_INDICATION_EYE, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran i / Eyes 归属 PharmaResearch / PR Bio 并拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "Rejuran i / Eyes", REJURAN_INDICATION_EYE, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 家族拥有欧洲 CE 认证；未提供具体证书编号。"),
    ],
    "REC_0970": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "Rejuran Healer", REJURAN_INDICATION_MAIN, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran Healer 归属 PharmaResearch / PR Bio 并拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "Rejuran Healer", REJURAN_INDICATION_MAIN, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 家族拥有欧洲 CE 认证；未提供具体证书编号。"),
    ],
    "REC_0971": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user at Rejuran family level; certificate number not captured", "Rejuran Shine", REJURAN_INDICATION_SHINE, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 家族归属 PharmaResearch / PR Bio 并拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user at Rejuran family level; certificate number not captured", "Rejuran Shine", REJURAN_INDICATION_SHINE, "PharmaResearch / PR Bio", REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "用户确认 Rejuran 家族拥有欧洲 CE 认证；未提供具体证书编号。"),
    ],
    "REC_0540": [
        reg("EU", "European Commission / Notified Body", "CE Class III certification", "CE Class III certification confirmed by user; certificate number not captured", "PhilArt / PhilArt Eye / PhilArt Hair", "PhilArt 为 Croma Pharma 的 PN 产品线，包含基础抗衰、眼周和头皮型号，用于恢复皮肤弹性、深层补水和细胞级修复。", "Croma Pharma", "https://www.croma-polyphil.co.uk/about-polynucleotidesss/polyphil-product-range", "Croma PolyPhil / PhilArt product range", "用户确认 PhilArt 拥有欧洲 CE 三类医疗器械认证。"),
    ],
    "REC_0319": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "HP Cell Vitaran", "HP Cell Vitaran 为 BR Pharm 的纯 PN/PDRN 注射剂，用于面部抗衰、组织修复，也可用于骨科/关节炎症修复场景。", "BR Pharm", "https://brpharm.com/product/medical_device?category_code=1911&code=335&tpf=product%2Fview", "BR Pharm HP Cell Vitaran product page", "用户确认 Vitaran 拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "HP Cell Vitaran", "HP Cell Vitaran 为 BR Pharm 的 PN/PDRN 再生注射剂，用于皮肤抗衰和组织修复。", "BR Pharm", "https://brpharm.com/product/medical_device?category_code=1911&code=335&tpf=product%2Fview", "BR Pharm HP Cell Vitaran product page", "用户确认 Vitaran 拥有欧洲 CE 认证。"),
    ],
    "REC_0619": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "Richesse / Richesse Eyes", "Richesse / Richesse Eyes 为 JDBIO 的高端水光/复配产品，眼周型号使用 PDRN、HA 和多肽，面向眼周暗沉、泪沟凹陷和眼部细纹。", "JDBIO", "https://en.jdbio.com/pages/richesse", "JDBIO Richesse official page", "用户确认 Richesse 拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "Richesse / Richesse Eyes", "Richesse / Richesse Eyes 用于眼周修复、补水、肤质改善和局部容量/细纹管理。", "JDBIO", "https://en.jdbio.com/pages/richesse", "JDBIO Richesse official page", "用户确认 Richesse 拥有欧洲 CE 认证。"),
    ],
    "REC_0636": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS family-level claim captured from user feedback; certificate number not captured", "S-DNA / PDRN skin booster", "S-DNA 属于 PDRN + 非交联 HA 等复配水光/皮肤动能素，用于提亮肤色、深层补水和光老化修复。", "DFK Biolab", "https://sbodyline.com/shop/sardenya-nucleo-plus-with-lidocaine-2/", "DFK Biolab S-DNA / skin booster product evidence", "用户将 DFK Biolab S-DNA 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE family-level claim captured from user feedback; certificate number not captured", "S-DNA / PDRN skin booster", "S-DNA 为 PDRN 复合皮肤动能素，用于补水、提亮和光老化修复。", "DFK Biolab", "https://sbodyline.com/shop/sardenya-nucleo-plus-with-lidocaine-2/", "DFK Biolab S-DNA / skin booster product evidence", "用户将 DFK Biolab S-DNA 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
    ],
    "REC_0731": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS family-level claim captured from user feedback; certificate number not captured", "Chaeum V7 / V-Seven", "Chaeum V7 为 PDRN 复合皮肤动能素，通过微针或手针定点注射，用于提亮肤色、深层锁水和光老化修复。", "Chaeum Pharma", "https://www.chaeumpharma.com/aesthetics.html", "Chaeum Pharma aesthetics product page", "用户将 Chaeum V7 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE family-level claim captured from user feedback; certificate number not captured", "Chaeum V7 / V-Seven", "Chaeum V7 为 PDRN 复合皮肤动能素，用于补水、提亮和肤质改善。", "Chaeum Pharma", "https://www.chaeumpharma.com/aesthetics.html", "Chaeum Pharma aesthetics product page", "用户将 Chaeum V7 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
    ],
    "REC_0293": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS family-level claim captured from user feedback; certificate number not captured", "GANA PNV / PNV+", "GANA PNV / PNV+ 为 PDRN + HA 复合水光/皮肤动能素，用于提亮肤色、深层补水和光老化修复。", "GANA", "https://ganafillers.com/product/gana-pn-pdrn-2-x-2-5ml/", "GANA PN/PDRN product page", "用户将 GANA PNV 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE family-level claim captured from user feedback; certificate number not captured", "GANA PNV / PNV+", "GANA PNV / PNV+ 为 PDRN + HA 复合水光/皮肤动能素，用于补水、提亮和肤质改善。", "GANA", "https://ganafillers.com/product/gana-pn-pdrn-2-x-2-5ml/", "GANA PN/PDRN product page", "用户将 GANA PNV 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
    ],
    "REC_0202": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS family-level claim captured from user feedback; certificate number not captured", "Dr. DMA / PDRN Injection", "Dr. DMA 为 PDRN 复合水光/中胚层疗法耗材，用于提亮肤色、深层锁水、光老化修复和整体肤质改善。", "Daejoo", "", "Daejoo Dr. DMA user-confirmed product note", "用户将 Daejoo Dr. DMA 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE family-level claim captured from user feedback; certificate number not captured", "Dr. DMA / PDRN Injection", "Dr. DMA 为 PDRN 复合水光/中胚层疗法耗材，用于补水、提亮和肤质改善。", "Daejoo", "", "Daejoo Dr. DMA user-confirmed product note", "用户将 Daejoo Dr. DMA 纳入其他 PDRN 复合水光组，说明多在韩国及欧洲以中胚层疗法耗材获得 KFDA 或 CE。", "Group-level KFDA/CE claim; exact product certificate still requires model-level capture."),
    ],
    "REC_0648": [
        reg("EU", "European Commission / Notified Body", "CE certification", "CE family-level claim captured from user feedback; certificate number not captured", "HYLA-PDRN / LIPAX", "HYLA-PDRN 属于 PDRN + 非交联 HA 复合水光/皮肤动能素，用于提亮肤色、深层补水和光老化修复；LIPAX 关联局部脂肪管理。", "Laboratoire Skin France", "https://www.globalskinfrance.com/", "Laboratoire GlobalSkin France product page", "用户将 Laboratoire Skin France 纳入其他 PDRN 复合水光组，说明欧洲侧 CE 注册/备案需作为家族级线索保留。", "Group-level CE claim; exact product certificate still requires model-level capture."),
    ],
    "REC_0260": [
        reg("EU", "European Commission / CPNP or national device authority", "CPNP cosmetic notification / selected Class II device registration", "CPNP or selected Class II registration captured from user feedback; injection registration not assumed", "ExoComplex / Purasomes exosome complex", "ExoComplex 为外泌体再生抗衰产品，主要与微针/导入联合使用，用于抗炎、敏感肌和玫瑰痤疮修复、屏障重塑、创面愈合加速和毛发再生。", "Dermoaroma", "https://dermoaroma.com/purasomes/", "Dermoaroma Purasomes / ExoComplex official page", "用户确认外泌体产品多数不作直接注射，Dermoaroma ExoComplex 通常作为高端无菌化妆品在欧洲完成 CPNP 备案，或部分国家二类器械注册。", "Do not treat this as HA filler or direct-injection CE evidence."),
    ],
    "REC_0554": [
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "Pluryal Classic / Volume", "Pluryal Classic 用于中度皱纹和唇部塑形；Pluryal Volume 用于颧骨、下巴、下颌线等面部轮廓重塑和深层容量补充。", "MD Skin Solutions", "https://mdskin-solutions.com/", "MD Skin Solutions Pluryal official page", "用户确认 Pluryal Classic / Volume 拥有欧洲 CE 认证。"),
    ],
    "REC_0143": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "Crystal / Avalon", "Koru Pharma Crystal Lite/Deep/Ultra 分别用于浅层细纹、中度鼻唇沟和深层骨膜上容量支撑；Avalon 为 HA 填充/抗衰复配系列。", "Koru Pharma", "https://korupharma.com/", "Koru Pharma Crystal / Avalon official page", "用户确认 Crystal / Avalon 拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "Crystal / Avalon", "Koru Pharma Crystal / Avalon 用于面部皱纹填充、容量支撑、轮廓塑形和部分抗衰复配应用。", "Koru Pharma", "https://korupharma.com/", "Koru Pharma Crystal / Avalon official page", "用户确认 Crystal / Avalon 拥有欧洲 CE 认证。"),
    ],
    "REC_0389": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "La Pomme / Caratfill", "La Pomme 为 HA 面部容量填充剂；Caratfill 为 HA、PDRN 和营养复合物皮肤动能素，用于水光焕亮、补水和肤质改善。", "Jaysean / J Syeon", "https://www.jay-sean.com/hafiller-lapomme", "Jaysean La Pomme official page", "用户确认 La Pomme / Caratfill 拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "La Pomme / Caratfill", "La Pomme / Caratfill 覆盖 HA 填充和复配水光动能素，用于容量补充、轮廓塑形和肤质提升。", "Jaysean / J Syeon", "https://www.jay-sean.com/skinbooster-caratfill", "Jaysean Caratfill official page", "用户确认 La Pomme / Caratfill 拥有欧洲 CE 认证。"),
    ],
    "REC_0592": [
        reg("KR", "KFDA/MFDS", "medical device registration", "KFDA/MFDS registration confirmed by user; certificate number not captured", "RE: Thread / RE: N-COG", "RE: N-COG 为 PDO 锯齿线，通过多向倒刺提供软组织咬合力，用于中下面部机械提拉；RE: Thread 可结合 PDO 线与 PDRN 涂层/技术以促进胶原再生。", "N-Finders", "https://nfinders.com/eng/product_thread_n-cog/", "N-Finders RE:N-COG official page", "用户确认 RE: Thread / RE: N-COG 拥有韩国 KFDA/MFDS 注册。"),
        reg("EU", "European Commission / Notified Body", "CE certification", "CE certification confirmed by user; certificate number not captured", "RE: Thread / RE: N-COG", "N-Finders RE: Thread / RE: N-COG 用于中下面部机械提拉、紧致和局部胶原再生。", "N-Finders", "https://nfinders.com/eng/product_thread_n-cog/", "N-Finders RE:N-COG official page", "用户确认 RE: Thread / RE: N-COG 拥有欧洲 CE 认证。"),
    ],
}


WORKBOOK_UPDATES: dict[str, dict[str, str]] = {
    "REC_0968": {
        "Brand": "Kiara Reju",
        "Core_Product": "Kiara Reju",
        "Category_L1": "Injectables",
        "Category_L2": "Skin Booster",
        "Tech_Type_Std": "PDRN + Hyaluronic Acid",
        "Tech_Type_Original": "PDRN / HA skin booster",
        "Manufactured_By": "BioPlus Co., Ltd.",
        "Marketing_Holder": "BioPlus",
        "Data_Source": "official_company_fact_override",
        "Duplicate_Note": "corrected_from:REJURAN; not PharmaResearch Rejuran. BioPlus official product evidence points to Kiara Reju.",
    },
    "REC_0969": {
        "Inclusion_Status": "excluded",
        "Data_Source": "official_company_fact_override",
        "Duplicate_Note": "wrong_attribution: REJURAN HB belongs to PharmaResearch / PR Bio; BioPlus mapping excluded.",
    },
    "REC_0601": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending.", "NMPA_Status": "NMPA confirmed for selected Rejuran models by user; exact certificate/model mapping pending.", "Manufactured_By": "PharmaResearch / PR Bio", "Marketing_Holder": "PharmaResearch / PR Bio"},
    "REC_0602": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending.", "Manufactured_By": "PharmaResearch / PR Bio", "Marketing_Holder": "PharmaResearch / PR Bio"},
    "REC_0970": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending.", "Manufactured_By": "PharmaResearch / PR Bio", "Marketing_Holder": "PharmaResearch / PR Bio"},
    "REC_0971": {"KFDA_Status": "KFDA/MFDS confirmed at Rejuran family level by user; certificate number pending.", "CE_Status": "CE confirmed at Rejuran family level by user; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Skin Booster", "Tech_Type_Std": "PN / HA Skin Booster", "Manufactured_By": "PharmaResearch / PR Bio", "Marketing_Holder": "PharmaResearch / PR Bio"},
    "REC_0540": {"CE_Status": "CE Class III confirmed by user; certificate number pending."},
    "REC_0319": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending."},
    "REC_0619": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Skin Booster", "Tech_Type_Std": "PDRN + Hyaluronic Acid + Peptides"},
    "REC_0636": {"KFDA_Status": "KFDA/MFDS family-level claim captured from user feedback; certificate number pending.", "CE_Status": "CE family-level claim captured from user feedback; certificate number pending.", "Tech_Type_Std": "PDRN + HA Skin Booster"},
    "REC_0731": {"KFDA_Status": "KFDA/MFDS family-level claim captured from user feedback; certificate number pending.", "CE_Status": "CE family-level claim captured from user feedback; certificate number pending.", "Tech_Type_Std": "PDRN / PN Skin Booster"},
    "REC_0293": {"KFDA_Status": "KFDA/MFDS family-level claim captured from user feedback; certificate number pending.", "CE_Status": "CE family-level claim captured from user feedback; certificate number pending.", "Tech_Type_Std": "PDRN + HA Skin Booster"},
    "REC_0202": {"KFDA_Status": "KFDA/MFDS family-level claim captured from user feedback; certificate number pending.", "CE_Status": "CE family-level claim captured from user feedback; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Skin Booster", "Tech_Type_Std": "PDRN / PN Skin Booster", "Tech_Type_Original": "PDRN mesotherapy injection"},
    "REC_0648": {"CE_Status": "CE family-level claim captured from user feedback; certificate number pending.", "Tech_Type_Std": "PDRN + HA Mesotherapy"},
    "REC_0260": {"Category_L1": "Regenerative", "Category_L2": "Exosome", "Tech_Type_Std": "Exosome / Topical Microneedling", "Tech_Type_Original": "Exosome complex for topical or microneedling-assisted application", "CE_Status": "EU CPNP or selected Class II registration captured from user feedback; direct-injection CE not assumed."},
    "REC_0396": {"Category_L1": "Regenerative", "Category_L2": "Exosome", "Tech_Type_Std": "PDRN / Exosome Skin Booster", "Tech_Type_Original": "Microneedling-compatible exosome/PDRN skin booster"},
    "REC_0554": {"CE_Status": "CE confirmed by user; certificate number pending."},
    "REC_0143": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending."},
    "REC_0389": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending.", "Tech_Type_Std": "HA Filler / PDRN Skin Booster"},
    "REC_0592": {"KFDA_Status": "KFDA/MFDS confirmed by user; certificate number pending.", "CE_Status": "CE confirmed by user; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Thread Lift", "Tech_Type_Std": "PDO / PDRN-coated Thread", "Tech_Type_Original": "PDO cog threads / PDRN coating"},
}


DIRECT_FACT_SOURCES: dict[str, tuple[str, str, str]] = {
    "REC_0601": (REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "Official source retained for Rejuran product-family identity; user corrected ownership away from BioPlus."),
    "REC_0602": (REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "Official source retained for Rejuran i / Eyes product-family identity."),
    "REC_0970": (REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "Official source retained for Rejuran Healer product identity."),
    "REC_0971": (REJURAN_SOURCE, "PharmaResearch Rejuran official product information", "Official source retained for Rejuran family product identity."),
    "REC_0540": ("https://www.croma-polyphil.co.uk/about-polynucleotidesss/polyphil-product-range", "Croma PolyPhil / PhilArt product range", "Official page supports PhilArt/PolyPhil PN product range and skin quality positioning."),
    "REC_0319": ("https://brpharm.com/product/medical_device?category_code=1911&code=335&tpf=product%2Fview", "BR Pharm HP Cell Vitaran product page", "Official page retained for HP Cell Vitaran product identity."),
    "REC_0619": ("https://en.jdbio.com/pages/richesse", "JDBIO Richesse official page", "Official page retained for Richesse product-family identity."),
    "REC_0731": ("https://www.chaeumpharma.com/aesthetics.html", "Chaeum Pharma aesthetics product page", "Official page retained for Chaeum V7 / dls PN family identity."),
    "REC_0293": ("https://ganafillers.com/product/gana-pn-pdrn-2-x-2-5ml/", "GANA PN/PDRN product page", "Official page retained for GANA PNV / PDRN product identity."),
    "REC_0260": ("https://dermoaroma.com/purasomes/", "Dermoaroma Purasomes / ExoComplex official page", "Official page retained for exosome product identity; not treated as HA filler or direct injection."),
    "REC_0554": ("https://mdskin-solutions.com/", "MD Skin Solutions Pluryal official page", "Official page retained for Pluryal product-family identity."),
    "REC_0389": ("https://www.jay-sean.com/hafiller-lapomme", "Jaysean La Pomme official page", "Official page retained for La Pomme HA filler identity."),
    "REC_0592": ("https://nfinders.com/eng/product_thread_n-cog/", "N-Finders RE:N-COG official page", "Official page retained for N-Finders thread product identity."),
}


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_user_reg_indications_b5_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: object) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        new = norm(value)
        if old == new:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id, fields in WORKBOOK_UPDATES.items():
        for field, value in fields.items():
            set_cell(record_id, field, value)
        if record_id in {"REC_0968", "REC_0969"}:
            append_audit(record_id, "user_reg_indications_20260601_batch5: corrected BioPlus/Rejuran attribution; PharmaResearch / PR Bio retained as Rejuran owner.")
        else:
            append_audit(record_id, "user_reg_indications_20260601_batch5: user-confirmed PN/PDRN/exosome/HA/thread indication and regulatory status applied; certificate numbers remain blank unless supplied.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_evidence_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, specs in TARGET_SPECS.items():
        product = products.get(record_id)
        if not product:
            continue
        for spec in specs:
            rows.append(evidence_row(product, checked_at, spec))
    return rows


def build_fact_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, (url, title, excerpt) in DIRECT_FACT_SOURCES.items():
        product = products.get(record_id)
        if product:
            rows.append(fact_row(product, checked_at, url, title, excerpt))
    return rows


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()

    ind_fields, ind_rows = read_csv(MANUAL_INDICATION_PATH)
    added_indications = append_unique(
        ind_rows,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        build_evidence_rows(products, checked_at),
    )
    write_csv(MANUAL_INDICATION_PATH, ind_fields, ind_rows)

    fact_fields, fact_rows = read_csv(MANUAL_FACT_PATH)
    added_facts = append_unique(fact_rows, ["fact_id"], build_fact_rows(products, checked_at))
    write_csv(MANUAL_FACT_PATH, fact_fields, fact_rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_indications,
        "manual_product_fact_rows_added": added_facts,
        "target_record_ids": sorted(TARGET_SPECS),
        "korea_confirmed_record_ids": sorted(
            record_id for record_id, specs in TARGET_SPECS.items() if any(spec.get("jurisdiction") == "KR" for spec in specs)
        ),
        "ce_confirmed_record_ids": sorted(
            record_id for record_id, specs in TARGET_SPECS.items() if any(spec.get("jurisdiction") == "EU" for spec in specs)
        ),
        "nmpa_confirmed_record_ids": sorted(
            record_id for record_id, specs in TARGET_SPECS.items() if any(spec.get("jurisdiction") == "CN" for spec in specs)
        ),
        "bioPlus_rejuran_correction": {
            "REC_0968": "renamed to BioPlus Kiara Reju based on existing official BioPlus Kiara Reju evidence; not treated as PharmaResearch Rejuran.",
            "REC_0969": "excluded as wrong attribution for Rejuran HB.",
        },
        "known_residual_notes": [
            "Certificate numbers remain blank where the user confirmed registration presence but did not supply precise numbers.",
            "DFK/Chaeum/GANA/Daejoo/Laboratoire Skin France PDRN rows are captured as family-level KFDA/CE claims, not exact certificate-level records.",
            "Haim Exovair/LeVair received category/technology correction only; no registration status was supplied in this user batch.",
            "BioPlus Kiara Reju remains an active BioPlus product identity without a KFDA row until product-specific registration is supplied.",
        ],
        "changed_fields_sample": workbook_changes[:120],
    }
    out_path = AUDIT_DIR / f"user_confirmed_regulatory_indications_batch5_{stamp}.json"
    latest_path = AUDIT_DIR / "user_confirmed_regulatory_indications_batch5_latest.json"
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
