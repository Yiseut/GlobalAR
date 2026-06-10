#!/usr/bin/env python3
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
AUDIT_DIR = ROOT / "data" / "audits"


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def set_value(ws, row_idx: int, colmap: dict[str, int], field: str, value: Any, changes: list[dict[str, Any]]) -> None:
    if field not in colmap:
        return
    cell = ws.cell(row=row_idx, column=colmap[field])
    old = cell.value
    if old == value:
        return
    cell.value = value
    changes.append({"sheet": ws.title, "row": row_idx, "field": field, "old": old, "new": value})


def close_alma_ted_duplicate(wb, changes: list[dict[str, Any]]) -> None:
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    for row_idx in range(2, ws.max_row + 1):
        if norm(ws.cell(row=row_idx, column=colmap["Record_ID"]).value) != "REC_0848":
            continue
        set_value(ws, row_idx, colmap, "Is_Primary_Record", False, changes)
        set_value(
            ws,
            row_idx,
            colmap,
            "Duplicate_Note",
            "duplicate_of:REC_0029; duplicate Alma TED / Trans-Epidermal Drug Delivery row closed by seed integrity audit 20260601.",
            changes,
        )
        return
    raise RuntimeError("REC_0848 not found in Product_Lines")


def close_bioplus_rejuranhb_duplicate(wb, changes: list[dict[str, Any]]) -> None:
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    for row_idx in range(2, ws.max_row + 1):
        record_id = norm(ws.cell(row=row_idx, column=colmap["Record_ID"]).value)
        company = norm(ws.cell(row=row_idx, column=colmap["Company"]).value).casefold()
        brand = norm(ws.cell(row=row_idx, column=colmap["Brand"]).value).casefold()
        core = norm(ws.cell(row=row_idx, column=colmap["Core_Product"]).value).casefold()
        if record_id == "REC_0969" or (company == "bioplus" and brand == "rejuran hb" and core == "rejuran hb"):
            set_value(ws, row_idx, colmap, "Is_Primary_Record", False, changes)
            set_value(
                ws,
                row_idx,
                colmap,
                "Duplicate_Note",
                "duplicate_of:REC_0601; wrong_attribution: REJURAN HB belongs to PharmaResearch / PR Bio, not BioPlus.",
                changes,
            )
            return
    raise RuntimeError("BioPlus / REJURAN HB row not found in Product_Lines")


def close_bioplus_brand_portfolio(wb, changes: list[dict[str, Any]]) -> None:
    ws = wb["Brand_Portfolio"]
    colmap = headers(ws)
    delete_rows: list[int] = []
    has_kiara = False
    for row_idx in range(2, ws.max_row + 1):
        company = norm(ws.cell(row=row_idx, column=colmap["Company"]).value).casefold()
        brand = norm(ws.cell(row=row_idx, column=colmap["Brand"]).value).casefold()
        if company == "bioplus" and brand in {"rejuran", "rejuran hb"}:
            delete_rows.append(row_idx)
        if company == "bioplus" and brand == "kiara reju":
            has_kiara = True
            updates = {
                "Country": "South Korea",
                "Category_L1": "Injectables",
                "Category_L2": "Skin Booster",
                "Tech_Type": "PDRN + Hyaluronic Acid",
                "Brand_Type": "Product",
                "Product_Count": 1,
                "Products": "Kiara Reju",
            }
            for field, value in updates.items():
                set_value(ws, row_idx, colmap, field, value, changes)
    for row_idx in sorted(delete_rows, reverse=True):
        changes.append(
            {
                "sheet": ws.title,
                "row": row_idx,
                "field": "delete_row",
                "old": f"{ws.cell(row=row_idx, column=colmap['Company']).value} / {ws.cell(row=row_idx, column=colmap['Brand']).value}",
                "new": "",
            }
        )
        ws.delete_rows(row_idx, 1)
    if not has_kiara:
        row_idx = ws.max_row + 1
        row = {
            "Company": "BioPlus",
            "Brand": "Kiara Reju",
            "Country": "South Korea",
            "Category_L1": "Injectables",
            "Category_L2": "Skin Booster",
            "Tech_Type": "PDRN + Hyaluronic Acid",
            "Brand_Type": "Product",
            "Product_Count": 1,
            "Products": "Kiara Reju",
        }
        for field, value in row.items():
            if field in colmap:
                ws.cell(row=row_idx, column=colmap[field]).value = value
        changes.append({"sheet": ws.title, "row": row_idx, "field": "insert_row", "old": "", "new": "BioPlus / Kiara Reju"})


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = AUDIT_DIR / f"source_workbook_backup_before_seed_integrity_close_{stamp}.xlsx"
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    changes: list[dict[str, Any]] = []
    close_alma_ted_duplicate(wb, changes)
    close_bioplus_rejuranhb_duplicate(wb, changes)
    close_bioplus_brand_portfolio(wb, changes)
    wb.save(SOURCE_BOOK)
    print({"backup": str(backup), "changes": len(changes)})


if __name__ == "__main__":
    main()
