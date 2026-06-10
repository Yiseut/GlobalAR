"""Apply user-confirmed regenerative/thread/EBD indication notes from 2026-06-01."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"
MANUAL_FACT_PATH = DATA_DIR / "manual_product_fact_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_fields: list[str], new_rows: list[dict[str, str]]) -> int:
    existing = {tuple(norm(row.get(field)) for field in key_fields) for row in rows}
    added = 0
    for row in new_rows:
        key = tuple(norm(row.get(field)) for field in key_fields)
        if key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def registration_row(product: dict[str, str], checked_at: str, spec: dict[str, str]) -> dict[str, str]:
    indication = norm(spec.get("indication"))
    source_key = norm(spec.get("source_key")) or stable_id(
        "user_confirmed_reg_ind_20260601_b2",
        product.get("seed_record_id"),
        spec.get("jurisdiction"),
        spec.get("regulator"),
        spec.get("registered_name"),
    )
    return {
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "jurisdiction": norm(spec.get("jurisdiction")),
        "regulator": norm(spec.get("regulator")),
        "regulatory_pathway": norm(spec.get("pathway")),
        "fda_product_code": norm(spec.get("fda_product_code")),
        "fda_regulation_number": norm(spec.get("fda_regulation_number")),
        "fda_device_class": norm(spec.get("fda_device_class")),
        "fda_submission_type": norm(spec.get("fda_submission_type")),
        "status": norm(spec.get("status")),
        "registration_no": norm(spec.get("registration_no")),
        "approval_date": norm(spec.get("approval_date")),
        "original_approval_date": norm(spec.get("original_approval_date")),
        "expiry_date": norm(spec.get("expiry_date")),
        "registered_name": norm(spec.get("registered_name")) or product.get("brand", ""),
        "approved_indication": indication,
        "intended_use": indication,
        "legal_manufacturer": norm(spec.get("legal_manufacturer")) or product.get("legal_manufacturer") or product.get("company", ""),
        "local_holder": norm(spec.get("local_holder")),
        "source_key": source_key,
        "source_url": norm(spec.get("source_url")),
        "source_type": norm(spec.get("source_type")) or "user_confirmed_official_claim",
        "evidence_title": norm(spec.get("evidence_title")),
        "evidence_excerpt": norm(spec.get("evidence_excerpt")),
        "official_description_exact": indication,
        "official_description_source_field": "approved_indication",
        "field_note": norm(spec.get("field_note")),
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260601_batch2",
        "review_status": norm(spec.get("review_status")) or "user_confirmed",
        "confidence": norm(spec.get("confidence")) or "user_confirmed_official_claim",
    }


def fact_row(product: dict[str, str], checked_at: str, source_url: str, title: str, excerpt: str) -> dict[str, str]:
    return {
        "fact_id": stable_id("pfact", product.get("seed_record_id"), "official_product_page", source_url),
        "product_id": product.get("product_id", ""),
        "seed_record_id": product.get("seed_record_id", ""),
        "company_id": product.get("company_id", ""),
        "company": product.get("company", ""),
        "brand": product.get("brand", ""),
        "product_family_id": "",
        "standard_product_name": product.get("standard_product_name", ""),
        "priority": "P0",
        "fact_group": "official_product_page",
        "field_name": "official_product_page",
        "field_value": source_url,
        "source_url": source_url,
        "evidence_title": title,
        "evidence_excerpt": excerpt,
        "source_type": "official_product_page",
        "confidence": "official_site_user_confirmed",
        "captured_at": checked_at,
        "promoted_at": checked_at,
        "review_status": "auto_cross_checked",
        "note": "user_confirmed_regulatory_indications_20260601_batch2",
    }


TARGET_SPECS: dict[str, list[dict[str, str]]] = {
    "REC_0367": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE Class III medical device",
            "status": "CE Class III certification confirmed by user; certificate number not captured",
            "registered_name": "JULÄINE",
            "indication": "JULÄINE 是 PDLLA 生物刺激剂，主要用于面部深层组织容量补充，通过刺激胶原蛋白新生改善深层皱纹、面部凹陷并提升整体肤质。",
            "legal_manufacturer": "Nordberg Medical",
            "source_url": "https://nordbergmedical.com/products/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Nordberg Medical JULÄINE official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 JULÄINE 拥有欧洲 CE 认证（三类医疗器械），并补充 PDLLA 胶原生物刺激剂适应症定位。",
            "field_note": "User confirmed CE Class III presence; no product-specific certificate number was supplied in this batch.",
        }
    ],
    "REC_1049": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device manufacture license / approval",
            "status": "KFDA/MFDS approval confirmed by user and official company materials; certificate number not captured",
            "registered_name": "PowerFill",
            "indication": "PowerFill 是面向身体大容量填充的 PDLLA 生物刺激剂，用于身体轮廓塑形和组织容量改善，包括非手术丰臀、身体曲线改善和组织松弛管理。",
            "legal_manufacturer": "Regen Biotech, Inc.",
            "source_url": "https://regenbioglobal.com/PRODUCTS",
            "source_type": "official_company_product_page",
            "evidence_title": "RegenBio Global PowerFill product page / user-confirmed KFDA status",
            "evidence_excerpt": "官方页面说明 PowerFill 为 PDLLA body filler，并列示 AestheFill/PowerFill 的 KFDA manufacture license 历史；用户确认 KFDA 与 CE。",
            "field_note": "Existing local row retained narrower Korea claim; this user-confirmed row captures the broader body-contouring positioning.",
            "confidence": "official_company_page_user_confirmed",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user and official company materials; certificate number not captured",
            "registered_name": "PowerFill",
            "indication": "PowerFill 是 PDLLA 身体填充/胶原刺激产品，用于身体轮廓塑形和大容量软组织容量补充。",
            "legal_manufacturer": "Regen Biotech, Inc.",
            "source_url": "https://regenbioglobal.com/PRODUCTS",
            "source_type": "official_company_product_page",
            "evidence_title": "RegenBio Global PowerFill CE status",
            "evidence_excerpt": "官方页面列示 2021 年 CE approval；用户确认 PowerFill 拥有欧洲 CE 认证。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        },
    ],
    "REC_0196": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "DLS Face / Body",
            "indication": "DLS Face/Body 是 PDLLA 生物刺激剂，用于深层真皮或皮下注射，以恢复组织容量并持续刺激胶原再生。",
            "legal_manufacturer": "Chaeum Pharma",
            "source_url": "https://www.chaeumpharma.com/aesthetics.html",
            "source_type": "user_confirmed_kfda_claim_with_official_product_page",
            "evidence_title": "Chaeum Pharma DLS PDLLA official page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 DLS Face/Body 拥有韩国 KFDA 与欧洲 CE；Chaeum Pharma 官方页面列示 dlS PDLLA aesthetics product.",
            "field_note": "User confirmed KFDA presence; no product-specific certificate number was supplied in this batch.",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "DLS Face / Body",
            "indication": "DLS Face/Body 是 PDLLA 生物刺激剂，用于组织容量恢复和胶原再生刺激，覆盖面部与身体型号。",
            "legal_manufacturer": "Chaeum Pharma",
            "source_url": "https://www.chaeumpharma.com/aesthetics.html",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Chaeum Pharma DLS PDLLA user-confirmed CE status",
            "evidence_excerpt": "用户确认 DLS Face/Body 拥有欧洲 CE 认证。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        },
    ],
    "REC_0291": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "GANA Fill / GANA V",
            "indication": "GANA Fill / GANA V 是 PLLA 生物刺激剂，用于面部重度皱纹纠正及容量补充；部分大规格型号用于身体轮廓雕塑。",
            "legal_manufacturer": "GANA R&D",
            "source_url": "http://ganarnd.co.kr/kwa-790695",
            "source_type": "official_product_page",
            "evidence_title": "GANA FILL official product page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 GANA Fill/GANA V 拥有韩国 KFDA 与欧洲 CE；GANA 官方页面列示 GANA FILL 为 CE approved PLLA filler.",
            "field_note": "User confirmed KFDA presence; no product-specific certificate number was supplied in this batch.",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user and official product page; certificate number not captured",
            "registered_name": "GANA Fill / GANA V",
            "indication": "GANA Fill / GANA V 是 PLLA 胶原刺激型填充产品，用于重度皱纹、容量补充及部分身体轮廓塑形场景。",
            "legal_manufacturer": "GANA R&D",
            "source_url": "http://ganarnd.co.kr/kwa-790695",
            "source_type": "official_product_page",
            "evidence_title": "GANA FILL CE approved PLLA filler official page",
            "evidence_excerpt": "GANA 官方页面标题列示 GANA FILL (CE approved PLLA filler)；用户同步确认 CE。",
            "field_note": "No product-specific certificate number was supplied in this batch.",
            "confidence": "official_product_page_user_confirmed",
        },
    ],
    "REC_0515": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS authorization confirmed by official Olidia page and user feedback",
            "registered_name": "Olidia",
            "indication": "Olidia 是 PLLA 生物刺激剂，主要用于注射入真皮深层或皮下组织，以纠正中重度鼻唇沟等深层皱纹及面部老化导致的严重容量缺失。",
            "legal_manufacturer": "PRP Science / Ubio",
            "source_url": "https://olidia.co.kr/",
            "source_type": "official_product_page",
            "evidence_title": "Olidia official page - CE/KFDA/NMPA status",
            "evidence_excerpt": "Olidia 官方页面列示欧洲 CE 认证、韩国食药处许可、中国 NMPA 注册等状态；用户确认 Olidia 获 CE 与 KFDA。",
            "field_note": "No Korea product-specific certificate number was supplied; NMPA number remains in the existing NMPA row.",
            "confidence": "official_product_page_user_confirmed",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by official Olidia page and user feedback",
            "registered_name": "Olidia",
            "indication": "Olidia 是 PLLA 生物刺激剂，用于鼻唇沟等面部深层皱纹和容量缺失的组织修复/胶原刺激治疗。",
            "legal_manufacturer": "PRP Science / Ubio",
            "source_url": "https://olidia.co.kr/",
            "source_type": "official_product_page",
            "evidence_title": "Olidia official page - CE certification",
            "evidence_excerpt": "Olidia 官方页面列示欧洲 CE 认证；用户确认 CE。",
            "field_note": "No product-specific CE certificate number was supplied in this batch.",
            "confidence": "official_product_page_user_confirmed",
        },
    ],
    "REC_0037": [
        {
            "jurisdiction": "EU / Global",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification / multi-country registrations",
            "status": "CE certification and broad global registration confirmed by user; certificate number not captured",
            "registered_name": "Aptos Excellence",
            "indication": "Aptos Excellence 是 P(LA/CL) 材质带刺可吸收线材，用于面部、下颌缘和颈部软组织的微创提拉、复位，并在材料降解过程中刺激胶原再生。",
            "legal_manufacturer": "Aptos",
            "source_url": "https://aptos.global/index.php/product/7/visage-excellence-method",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Aptos Excellence official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Aptos Excellence 拥有欧洲 CE 认证，并在全球 70 多个国家具有合规注册。",
            "field_note": "User confirmed CE/global registration presence; no product-specific certificate number was supplied in this batch.",
        }
    ],
    "REC_0156": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Definisse Threads",
            "indication": "Definisse Threads 是 p-LA/CL 材质双向倒刺可吸收线，主要用于中面部下垂及下颌缘轮廓的提升复位。",
            "legal_manufacturer": "Relife / Menarini",
            "source_url": "https://www.relife-aesthetics.com/global/en/products/definisse-threads.html",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Relife Definisse Threads official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Definisse Threads 拥有欧洲 CE 认证。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        }
    ],
    "REC_0677": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Spring Thread",
            "indication": "Spring Thread 是不可吸收的永久性弹性线材，通常由医用级硅胶和聚酯制成，用于中重度面部松弛的长期提拉固定。",
            "legal_manufacturer": "1st SurgiConcept",
            "source_url": "https://springthread.com/en/product/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Spring Thread official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Spring Thread 拥有欧洲 CE 认证。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        }
    ],
    "REC_0591": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "Rainbow Threads",
            "indication": "Rainbow Threads 涵盖 PDO、PLLA、PCL 多种材质和平滑线、螺旋线、锯齿线等形态，用于全脸、颈部及身体的紧致、提拉与肤质改善。",
            "legal_manufacturer": "Oreon Life Science",
            "source_url": "http://oreonglobal.com/sub/Thread_RAINBOW_1.php",
            "source_type": "official_product_page",
            "evidence_title": "Oreon Rainbow Threads official page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 Rainbow Threads 拥有韩国 KFDA 与欧洲 CE；Oreon 官方页面列示 Rainbow thread 产品并出现 CE mark 线索。",
            "field_note": "User confirmed KFDA presence; no product-specific certificate number was supplied in this batch.",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE mark confirmed by official product evidence and user feedback; certificate number not captured",
            "registered_name": "Rainbow Threads",
            "indication": "Rainbow Threads 是综合线雕耗材品牌，用于软组织提拉、紧致和肤质改善。",
            "legal_manufacturer": "Oreon Life Science",
            "source_url": "http://oreonglobal.com/sub/Thread_RAINBOW_1.php",
            "source_type": "official_product_page",
            "evidence_title": "Oreon Rainbow Threads CE mark evidence",
            "evidence_excerpt": "Oreon 官方产品页/本地规格证据含 CE mark 线索；用户确认 CE。",
            "field_note": "No product-specific CE certificate number was supplied in this batch.",
            "confidence": "official_product_page_user_confirmed",
        },
    ],
    "REC_0548": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "V Line / Omega",
            "indication": "V Line / Omega 是韩国可吸收蛋白线/锯齿线品牌，主要用于面部松弛组织的物理提拉及 V 脸轮廓重塑。",
            "legal_manufacturer": "ITC",
            "source_url": "https://xlinethreads.com/xline-r-placl",
            "source_type": "user_confirmed_kfda_claim_with_product_page",
            "evidence_title": "ITC V Line / Omega user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 V Line / Omega 拥有韩国 KFDA 与欧洲 CE。",
            "field_note": "User confirmed KFDA presence; no product-specific certificate number was supplied in this batch.",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "V Line / Omega",
            "indication": "V Line / Omega 是可吸收线雕耗材，用于面部轮廓提拉和 V 脸塑形。",
            "legal_manufacturer": "ITC",
            "source_url": "https://xlinethreads.com/xline-r-placl",
            "source_type": "user_confirmed_ce_claim_with_product_page",
            "evidence_title": "ITC V Line / Omega user-confirmed CE status",
            "evidence_excerpt": "用户确认 V Line / Omega 拥有欧洲 CE。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        },
    ],
    "REC_0359": [
        {
            "jurisdiction": "KR",
            "regulator": "MFDS / KFDA",
            "pathway": "Medical device registration",
            "status": "KFDA/MFDS registration confirmed by user; certificate number not captured",
            "registered_name": "Potenza",
            "indication": "Potenza 是射频微针设备，通过单极、双极射频与绝缘/非绝缘微针组合，用于改善痤疮疤痕、毛孔粗大、皮肤紧致及面部细纹。",
            "legal_manufacturer": "Jeisys Medical",
            "source_url": "https://jeisys-inc.com/products/potenza/",
            "source_type": "official_product_page",
            "evidence_title": "Jeisys Potenza official product page / user-confirmed KFDA status",
            "evidence_excerpt": "用户确认 Potenza 拥有美国 FDA、欧洲 CE 与韩国 KFDA；Jeisys 为韩国原研制造商。",
            "field_note": "User confirmed KFDA presence; no product-specific certificate number was supplied in this batch.",
        },
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Potenza",
            "indication": "Potenza 是多模式射频微针设备，用于痤疮疤痕、毛孔、紧肤和面部细纹改善。",
            "legal_manufacturer": "Jeisys Medical",
            "source_url": "https://jeisyseurope.com/",
            "source_type": "official_product_page",
            "evidence_title": "Jeisys Europe Potenza official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Potenza 拥有欧洲 CE 认证。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        },
        {
            "jurisdiction": "US",
            "regulator": "FDA",
            "pathway": "510(k)",
            "status": "FDA clearance confirmed by existing K190678 evidence and user feedback",
            "registration_no": "K190678",
            "registered_name": "Potenza / TempSure platform",
            "indication": "Potenza 是 RF microneedling 平台；本库按医美用途补充痤疮疤痕、毛孔、紧肤和面部细纹改善场景，具体美国法定适应证以 510(k) 文件为准。",
            "legal_manufacturer": "Jeisys Medical",
            "source_url": "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm",
            "source_type": "official_fda_510k_reference",
            "evidence_title": "FDA 510(k) K190678 / Potenza user-confirmed product mapping",
            "evidence_excerpt": "用户确认 Potenza 拥有美国 FDA 许可；现库已有 K190678 证据行。",
            "field_note": "Clinical/aesthetic use wording is user supplied; legal FDA indication should be read from the original 510(k) record.",
            "confidence": "official_regulator_record_user_mapped",
        },
    ],
    "REC_0823": [
        {
            "jurisdiction": "EU / Global",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Potenza",
            "indication": "Potenza 是 Jeisys 原研、Cynosure 全球推广的射频微针设备，用于痤疮疤痕、毛孔、紧肤和面部细纹改善。",
            "legal_manufacturer": "Jeisys Medical",
            "local_holder": "Cynosure",
            "source_url": "https://www.cynosurelutronicanz.com/products/potenza",
            "source_type": "user_confirmed_ce_claim_with_distributor_product_page",
            "evidence_title": "Cynosure Potenza product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Potenza 拥有美国 FDA、欧洲 CE 与韩国 KFDA；Cynosure 为全球主要商业分销商。",
            "field_note": "Cynosure row represents commercial distribution; Jeisys row carries Korea manufacturer context.",
        }
    ],
    "REC_0402": [
        {
            "jurisdiction": "EU",
            "regulator": "European Commission / Notified Body",
            "pathway": "CE certification",
            "status": "CE certification confirmed by user; certificate number not captured",
            "registered_name": "Lipo Shoock+",
            "indication": "Lipo Shoock+ 是专业局部减脂中胚层注射产品，通常结合 PPC、脱氧胆酸钠和代谢促进成分，用于双下巴、腹部、大腿等局部顽固脂肪沉积的溶脂和局部紧致管理。",
            "legal_manufacturer": "Dives Med",
            "source_url": "https://pro.divesmed.com/liposhoock/",
            "source_type": "user_confirmed_ce_claim_with_official_product_page",
            "evidence_title": "Dives Med LipoShoock+ official product page / user-confirmed CE status",
            "evidence_excerpt": "用户确认 Lipo Shoock+ 获欧洲 CE 认证，并补充其局部减脂中胚层疗法耗材定位。",
            "field_note": "User confirmed CE presence; no product-specific certificate number was supplied in this batch.",
        }
    ],
}


WORKBOOK_UPDATES: dict[str, dict[str, str]] = {
    "REC_0367": {"CE_Status": "CE Class III medical device confirmed by user; certificate number pending.", "Manufactured_By": "Nordberg Medical", "Marketing_Holder": "Nordberg Medical"},
    "REC_1049": {"CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.", "KFDA_Status": "KFDA/MFDS approval confirmed by user; number pending.", "Manufactured_By": "Regen Biotech, Inc.", "Marketing_Holder": "RegenBio Global / Regen Biotech"},
    "REC_0196": {"CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.", "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.", "Manufactured_By": "Chaeum Pharma", "Marketing_Holder": "Chaeum Pharma"},
    "REC_0291": {"CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.", "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.", "Manufactured_By": "GANA R&D", "Marketing_Holder": "GANA"},
    "REC_0515": {"CE_Status": "KFDA/MFDS and CE confirmed by official Olidia page/user; NMPA row already present.", "KFDA_Status": "KFDA/MFDS authorization confirmed; number pending.", "Manufactured_By": "PRP Science / Ubio", "Marketing_Holder": "PRP Science / Ubio"},
    "REC_0037": {"CE_Status": "CE and multi-country registrations confirmed by user; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Thread Lift", "Tech_Type_Std": "P(LA/CL) Thread", "Manufactured_By": "Aptos", "Marketing_Holder": "Aptos"},
    "REC_0156": {"CE_Status": "CE confirmed by user; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Thread Lift", "Tech_Type_Std": "P(LA/CL) Thread", "Manufactured_By": "Relife / Menarini", "Marketing_Holder": "Relife / Menarini"},
    "REC_0677": {"CE_Status": "CE confirmed by user; certificate number pending.", "Category_L1": "Injectables", "Category_L2": "Thread Lift", "Tech_Type_Std": "Permanent Elastic Thread", "Manufactured_By": "1st SurgiConcept", "Marketing_Holder": "1st SurgiConcept"},
    "REC_0591": {"CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.", "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.", "Category_L1": "Injectables", "Category_L2": "Thread Lift", "Tech_Type_Std": "PDO / PLLA / PCL Thread", "Manufactured_By": "Oreon Life Science", "Marketing_Holder": "Oreon Life Science"},
    "REC_0548": {"CE_Status": "KFDA/MFDS and CE confirmed by user; certificate numbers pending.", "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.", "Category_L1": "Injectables", "Category_L2": "Thread Lift", "Tech_Type_Std": "PDO / PCL Thread", "Manufactured_By": "ITC", "Marketing_Holder": "ITC"},
    "REC_0359": {"FDA_Status": "FDA clearance confirmed via Potenza/TempSure 510(k) mapping; verify legal indication in source record.", "FDA_510k_Number": "K190678", "CE_Status": "FDA / CE / KFDA-MFDS confirmed by user; CE/KFDA numbers pending.", "KFDA_Status": "KFDA/MFDS registration confirmed by user; number pending.", "Category_L1": "EBD", "Category_L2": "Microneedling", "Tech_Type_Std": "RF Microneedling", "Manufactured_By": "Jeisys Medical", "Marketing_Holder": "Jeisys Medical"},
    "REC_0823": {"FDA_Status": "FDA cleared; existing K190678 evidence present.", "FDA_510k_Number": "K190678", "CE_Status": "FDA / CE / KFDA-MFDS confirmed by user; CE/KFDA numbers pending.", "Manufactured_By": "Jeisys Medical", "Marketing_Holder": "Cynosure"},
    "REC_0402": {"CE_Status": "CE confirmed by user; certificate number pending.", "Manufactured_By": "Dives Med", "Marketing_Holder": "Dives Med"},
}


DIRECT_FACT_SOURCES: dict[str, tuple[str, str, str]] = {
    "REC_0515": ("https://olidia.co.kr/", "Olidia official product page", "Official Olidia page states CE certification, Korean authorization, China NMPA registration, TFDA import authorization, composition and injection layers."),
    "REC_1049": ("https://regenbioglobal.com/PRODUCTS", "RegenBio Global PowerFill official product page", "Official product page describes PowerFill as a PDLLA body filler and lists KFDA/CE-related history for AestheFill and PowerFill."),
}


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_user_reg_indications_b2_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: object) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        new = norm(value)
        if old == new:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=row, column=col, value=new)
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": new})

    for record_id, fields in WORKBOOK_UPDATES.items():
        for field, value in fields.items():
            set_cell(record_id, field, value)
        append_audit(record_id, "user_reg_indications_20260601_batch2: user-confirmed indication/regulatory status applied; precise certificate numbers remain blank unless supplied.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_registration_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, specs in TARGET_SPECS.items():
        product = products.get(record_id)
        if not product:
            continue
        for spec in specs:
            rows.append(registration_row(product, checked_at, spec))
    return rows


def build_fact_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record_id, (url, title, excerpt) in DIRECT_FACT_SOURCES.items():
        product = products.get(record_id)
        if product:
            rows.append(fact_row(product, checked_at, url, title, excerpt))
    return rows


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()

    ind_fields, ind_rows = read_csv(MANUAL_INDICATION_PATH)
    added_indications = append_unique(
        ind_rows,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        build_registration_rows(products, checked_at),
    )
    write_csv(MANUAL_INDICATION_PATH, ind_fields, ind_rows)

    fact_fields, fact_rows = read_csv(MANUAL_FACT_PATH)
    added_facts = append_unique(fact_rows, ["fact_id"], build_fact_rows(products, checked_at))
    write_csv(MANUAL_FACT_PATH, fact_fields, fact_rows)

    summary = {
        "backup": str(backup),
        "workbook_changes": len(workbook_changes),
        "manual_official_indication_rows_added": added_indications,
        "manual_product_fact_rows_added": added_facts,
        "target_record_ids": sorted(TARGET_SPECS),
        "category_corrected_record_ids": sorted(
            record_id
            for record_id, fields in WORKBOOK_UPDATES.items()
            if "Category_L2" in fields or "Tech_Type_Std" in fields
        ),
        "changed_fields_sample": workbook_changes[:60],
    }
    out_path = AUDIT_DIR / f"user_confirmed_regulatory_indications_batch2_{stamp}.json"
    latest_path = AUDIT_DIR / "user_confirmed_regulatory_indications_batch2_latest.json"
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8", errors="replace")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
