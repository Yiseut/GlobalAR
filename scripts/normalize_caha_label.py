from __future__ import annotations

from datetime import datetime
from pathlib import Path
import json
import shutil

import openpyxl


ROOT = Path(r"E:\shared\Documents\data")
WORKBOOK = ROOT / "全球医美企业库_标准化版v4.xlsx"
AUDIT_DIR = ROOT / "global_aesthetics_dashboard" / "data" / "audits"


REPLACEMENTS = {
    "CaHA 微晶瓷/羟基磷灰石钙填充剂": "CaHA 胶原刺激剂",
    "CaHA 微晶瓷/羟基磷酸钙填充剂": "CaHA 胶原刺激剂",
    "CaHA 填充剂": "CaHA 胶原刺激剂",
    "HA + CaHA 复合胶原刺激填充剂": "HA + CaHA 复合胶原刺激剂",
    "复合注射填充剂": "复合注射剂",
    "透明质酸+羟基磷灰石钙混合注射剂": "HA + CaHA 混合注射剂",
    "透明质酸+羟基磷酸钙混合注射剂": "HA + CaHA 混合注射剂",
    "羟基磷灰石钙 (CaHA)": "CaHA",
    "羟基磷酸钙 (CaHA)": "CaHA",
    "CaHA (羟基磷灰石钙)": "CaHA",
    "CaHA (羟基磷酸钙)": "CaHA",
    "羟基磷灰石钙": "CaHA",
    "羟基磷酸钙": "CaHA",
    "CaHA 微晶瓷": "CaHA",
    "微晶瓷": "CaHA",
}


PRODUCT_LINE_USE_CASE_UPDATES = {
    "REC_0314": {
        "Feature_Tags": "ha-caha-hybrid, caha, calcium-hydroxylapatite, biostimulator, hybrid-injectable",
        "Introduction": "双效混合注射剂。结合透明质酸 (HA) 的即刻支撑效果和 CaHA 的胶原再生能力。目前主要在欧洲及部分国际市场上市，属于 HA + CaHA 复合注射/再生赛道。",
    },
    "REC_0266": {
        "Verified_Product_Type_CN": "CaHA 胶原刺激剂",
        "Feature_Tags": "caha, calcium-hydroxylapatite, collagen-stimulator, dermal-filler",
        "Introduction": "CGBIO 官方产品页将 FACETEM/FACETEM S 标注为 Calcium Hydroxylapatite (CaHA) dermal filler；材料归属按 CaHA/胶原刺激剂处理。",
    },
    "REC_0960": {
        "Verified_Product_Type_CN": "CaHA 胶原刺激剂",
        "Feature_Tags": "caha, calcium-hydroxylapatite, collagen-stimulator, cgbio-daewoong",
        "Introduction": "FACETEM 属于 CGBIO/Daewoong 体系的 Calcium Hydroxylapatite (CaHA) 产品线；源行原 HA 描述已按人工复核改正，并作为 REC_0266 的非主重复线索保留。",
    },
    "REC_0590": {
        "Verified_Product_Type_CN": "CaHA 胶原刺激剂（含稀释打法）",
        "Feature_Tags": "caha, calcium-hydroxylapatite, biostimulator, hyperdilute, skin-quality",
    },
    "REC_1023": {
        "Verified_Product_Type_CN": "HA + CaHA 复合胶原刺激剂（水光/肤质改善打法）",
        "Feature_Tags": "ha-caha-hybrid, peg-crosslinked, biostimulator, skinbooster-use, skin-quality",
        "Introduction": "Neauvia/Matex Lab 官方资料将 STIMULATE 描述为 PEG 交联 HA 并加入 Calcium Hydroxyapatite (CaHA) 的复合注射剂；人工复核补充：该线按水光/肤质改善打法理解，不按传统填充剂统称。",
    },
}


def normalize_text(value: str) -> str:
    output = value
    for old, new in REPLACEMENTS.items():
        output = output.replace(old, new)
    return output


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(ws[1])
        if cell.value is not None and str(cell.value).strip()
    }


def apply_product_line_use_case_updates(wb, changes: list[dict[str, str]]) -> None:
    if "Product_Lines" not in wb.sheetnames:
        return
    ws = wb["Product_Lines"]
    headers = header_map(ws)
    if "Record_ID" not in headers:
        return
    for row in range(2, ws.max_row + 1):
        record_id = str(ws.cell(row=row, column=headers["Record_ID"]).value or "").strip()
        updates = PRODUCT_LINE_USE_CASE_UPDATES.get(record_id)
        if not updates:
            continue
        for field, updated in updates.items():
            if field not in headers:
                continue
            cell = ws.cell(row=row, column=headers[field])
            before = cell.value
            if before == updated:
                continue
            cell.value = updated
            changes.append(
                {
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "before": "" if before is None else str(before),
                    "after": updated,
                }
            )


def main() -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = WORKBOOK.with_name(f"{WORKBOOK.stem}.backup_before_caha_label_normalize_{timestamp}{WORKBOOK.suffix}")
    shutil.copy2(WORKBOOK, backup)

    wb = openpyxl.load_workbook(WORKBOOK)
    changes: list[dict[str, str]] = []
    for ws in wb.worksheets:
      for row in ws.iter_rows():
        for cell in row:
          value = cell.value
          if not isinstance(value, str):
            continue
          updated = normalize_text(value)
          if updated != value:
            cell.value = updated
            changes.append(
              {
                "sheet": ws.title,
                "cell": cell.coordinate,
                "before": value,
                "after": updated,
              }
            )

    apply_product_line_use_case_updates(wb, changes)

    wb.save(WORKBOOK)
    wb.close()

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    summary = {
        "run_marker": "caha_label_normalize",
        "timestamp": timestamp,
        "workbook": str(WORKBOOK),
        "backup": str(backup),
        "changes": len(changes),
        "by_sheet": {},
        "sample": changes[:40],
    }
    for item in changes:
        summary["by_sheet"][item["sheet"]] = summary["by_sheet"].get(item["sheet"], 0) + 1
    out = AUDIT_DIR / "caha_label_normalize_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
