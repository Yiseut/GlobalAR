#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_DIR / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = Path(r"E:\shared\Documents\data\全球医美企业库_标准化版v4.xlsx")

SEED_RECORD_IDS = {
    "REC_0007",
    "REC_0036",
    "REC_0046",
    "REC_0267",
    "REC_0470",
    "REC_0657",
}

CSV_CLEANUP_FILES = [
    "manual_product_fact_evidence.csv",
    "manual_evidence_promotion_log.csv",
    "evidence_promotion_log.csv",
    "official_website_master.csv",
    "company_media_asset_index.csv",
    "company_official_source_plan.csv",
    "company_official_website.csv",
    "briefing_update_candidates.csv",
    "briefing_verified_update_events.csv",
    "briefing_product_gap_candidates.csv",
    "product_specification_evidence.csv",
]

JSONL_CLEANUP_FILES = [
    "company_official_source_evidence.jsonl",
]


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def compact(value: Any) -> str:
    return "".join(ch.lower() for ch in norm(value) if ch.isalnum() or "\u4e00" <= ch <= "\u9fff")


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def row_as_dict(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, colmap: dict[str, int]) -> dict[str, Any]:
    return {name: ws.cell(row=row_idx, column=col).value for name, col in colmap.items()}


def cell(row: dict[str, Any], *names: str) -> str:
    lowered = {key.lower(): value for key, value in row.items()}
    for name in names:
        value = row.get(name)
        if norm(value):
            return norm(value)
        value = lowered.get(name.lower())
        if norm(value):
            return norm(value)
    return ""


def split_ids(value: Any) -> set[str]:
    return {part.strip() for part in norm(value).replace(";", ",").split(",") if part.strip()}


def is_research_data_service_product(row: dict[str, Any]) -> bool:
    record_id = cell(row, "Record_ID", "record_id")
    if record_id in SEED_RECORD_IDS:
        return True
    path = cell(row, "Material_Taxonomy_Path_CN", "material_taxonomy_path_cn")
    if "科研/数据服务" in path or "非产品服务" in path:
        return True
    category_l1 = compact(cell(row, "Category_L1", "category_l1"))
    tech = cell(row, "Tech_Type_Std", "tech_type_std", "Tech_Type", "tech_type").lower()
    core = cell(row, "Core_Product", "core_product", "Products", "product_family").lower()
    if category_l1 in {"services", "service"} and any(term in tech or term in core for term in ("genetic", "dna", "training", "education", "conference")):
        return True
    return False


def delete_rows(ws: openpyxl.worksheet.worksheet.Worksheet, row_numbers: Iterable[int]) -> int:
    rows = sorted(set(row_numbers), reverse=True)
    for row_idx in rows:
        ws.delete_rows(row_idx, 1)
    return len(rows)


def remaining_product_counts(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[Counter[str], dict[str, set[str]], dict[str, Counter[str]]]:
    colmap = headers(ws)
    company_counts: Counter[str] = Counter()
    brand_sets: dict[str, set[str]] = defaultdict(set)
    track_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for row_idx in range(2, ws.max_row + 1):
        row = row_as_dict(ws, row_idx, colmap)
        company = cell(row, "Company", "company")
        if not company:
            continue
        company_counts[company] += 1
        brand = cell(row, "Brand", "brand")
        if brand:
            brand_sets[company].add(brand)
        track = cell(row, "Category_L1", "category_l1")
        if track:
            track_counts[company][track] += 1
    return company_counts, brand_sets, track_counts


def removed_identity_match(row: dict[str, Any], removed_products: list[dict[str, str]], removed_companies: set[str]) -> bool:
    company = cell(row, "Company", "company")
    if not company:
        return False
    if company in removed_companies:
        return True
    row_brand = compact(cell(row, "Brand", "brand"))
    row_product_text = compact(
        " ".join(
            [
                cell(row, "Core_Product", "core_product"),
                cell(row, "Products", "products"),
                cell(row, "product_family", "Product_Family"),
                cell(row, "standard_product_name", "Standard_Product_Name"),
                cell(row, "product_name", "Product_Name"),
                cell(row, "candidate_product_or_family", "Candidate_Product_Or_Family"),
            ]
        )
    )
    for item in removed_products:
        if item["company"] != company:
            continue
        if row_brand and row_brand == item["brand_key"]:
            return True
        if item["core_key"] and item["core_key"] in row_product_text:
            return True
    return False


def workbook_cleanup(stamp: str) -> dict[str, Any]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_research_data_service_removal_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)

    wb = openpyxl.load_workbook(SOURCE_BOOK)
    removed_products: list[dict[str, str]] = []
    removed_by_sheet: dict[str, int] = {}

    ws = wb["Product_Lines"]
    colmap = headers(ws)
    product_delete_rows: list[int] = []
    for row_idx in range(2, ws.max_row + 1):
        row = row_as_dict(ws, row_idx, colmap)
        if not is_research_data_service_product(row):
            continue
        product_delete_rows.append(row_idx)
        removed_products.append(
            {
                "record_id": cell(row, "Record_ID", "record_id"),
                "product_id": cell(row, "Product_UUID", "product_uuid"),
                "company": cell(row, "Company", "company"),
                "brand": cell(row, "Brand", "brand"),
                "brand_key": compact(cell(row, "Brand", "brand")),
                "core_product": cell(row, "Core_Product", "core_product"),
                "core_key": compact(cell(row, "Core_Product", "core_product")),
                "material_path": cell(row, "Material_Taxonomy_Path_CN", "material_taxonomy_path_cn"),
            }
        )
    removed_by_sheet["Product_Lines"] = delete_rows(ws, product_delete_rows)

    product_ids = {item["product_id"] for item in removed_products if item["product_id"]}
    record_ids = {item["record_id"] for item in removed_products if item["record_id"]}
    affected_companies = {item["company"] for item in removed_products if item["company"]}
    company_counts, brand_sets, track_counts = remaining_product_counts(ws)
    removed_companies = {company for company in affected_companies if company_counts.get(company, 0) == 0}

    if "Brand_Portfolio" in wb.sheetnames:
        ws = wb["Brand_Portfolio"]
        colmap = headers(ws)
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = row_as_dict(ws, row_idx, colmap)
            if removed_identity_match(row, removed_products, removed_companies):
                rows.append(row_idx)
        removed_by_sheet["Brand_Portfolio"] = delete_rows(ws, rows)

    if "Companies" in wb.sheetnames:
        ws = wb["Companies"]
        colmap = headers(ws)
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = row_as_dict(ws, row_idx, colmap)
            company = cell(row, "Company", "company")
            if company in removed_companies:
                rows.append(row_idx)
                continue
            if company in affected_companies:
                if "Product_Count" in colmap:
                    ws.cell(row=row_idx, column=colmap["Product_Count"]).value = company_counts.get(company, 0)
                if "Brand_Count" in colmap:
                    ws.cell(row=row_idx, column=colmap["Brand_Count"]).value = len(brand_sets.get(company, set()))
                if "Primary_Track" in colmap and track_counts.get(company):
                    ws.cell(row=row_idx, column=colmap["Primary_Track"]).value = track_counts[company].most_common(1)[0][0]
        removed_by_sheet["Companies"] = delete_rows(ws, rows)

    if "Product_Master" in wb.sheetnames:
        ws = wb["Product_Master"]
        colmap = headers(ws)
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = row_as_dict(ws, row_idx, colmap)
            if cell(row, "product_id", "Product_ID") in product_ids or cell(row, "seed_record_id", "Seed_Record_ID") in record_ids:
                rows.append(row_idx)
        removed_by_sheet["Product_Master"] = delete_rows(ws, rows)

    if "Product_Family_Master" in wb.sheetnames:
        ws = wb["Product_Family_Master"]
        colmap = headers(ws)
        rows = []
        skipped_mixed: list[dict[str, Any]] = []
        for row_idx in range(2, ws.max_row + 1):
            row = row_as_dict(ws, row_idx, colmap)
            ids = split_ids(cell(row, "source_record_ids", "Source_Record_IDs"))
            if ids & record_ids:
                if ids - record_ids:
                    skipped_mixed.append({"row": row_idx, "source_record_ids": sorted(ids)})
                else:
                    rows.append(row_idx)
        removed_by_sheet["Product_Family_Master"] = delete_rows(ws, rows)
    else:
        skipped_mixed = []

    if "Product_SKU_Master" in wb.sheetnames:
        ws = wb["Product_SKU_Master"]
        colmap = headers(ws)
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = row_as_dict(ws, row_idx, colmap)
            if cell(row, "sku_id", "SKU_ID") in product_ids or cell(row, "seed_record_id", "Seed_Record_ID") in record_ids:
                rows.append(row_idx)
        removed_by_sheet["Product_SKU_Master"] = delete_rows(ws, rows)

    wb.save(SOURCE_BOOK)
    wb.close()

    return {
        "source_workbook": str(SOURCE_BOOK),
        "workbook_backup": str(backup),
        "removed_products": removed_products,
        "removed_companies": sorted(removed_companies),
        "remaining_affected_companies": {
            company: {
                "product_count": company_counts.get(company, 0),
                "brand_count": len(brand_sets.get(company, set())),
            }
            for company in sorted(affected_companies - removed_companies)
        },
        "removed_by_sheet": removed_by_sheet,
        "skipped_mixed_family_rows": skipped_mixed,
    }


def data_row_is_removed(row: dict[str, Any], removed_products: list[dict[str, str]], removed_companies: set[str]) -> bool:
    product_ids = {item["product_id"] for item in removed_products if item["product_id"]}
    record_ids = {item["record_id"] for item in removed_products if item["record_id"]}
    if cell(row, "product_id", "Product_ID") in product_ids:
        return True
    if cell(row, "seed_record_id", "Seed_Record_ID", "record_id", "Record_ID") in record_ids:
        return True
    return removed_identity_match(row, removed_products, removed_companies)


def cleanup_csv(path: Path, stamp: str, removed_products: list[dict[str, str]], removed_companies: set[str]) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path), "exists": False, "removed": 0}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        rows = list(reader)
    kept: list[dict[str, Any]] = []
    removed: list[dict[str, Any]] = []
    for row in rows:
        if data_row_is_removed(row, removed_products, removed_companies):
            removed.append(row)
        else:
            kept.append(row)
    if not removed:
        return {"path": str(path), "exists": True, "rows": len(rows), "removed": 0}
    backup = path.with_name(f"{path.stem}.backup_before_research_data_service_removal_{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(kept)
    return {
        "path": str(path),
        "backup": str(backup),
        "rows_before": len(rows),
        "rows_after": len(kept),
        "removed": len(removed),
        "removed_preview": removed[:5],
    }


def cleanup_jsonl(path: Path, stamp: str, removed_products: list[dict[str, str]], removed_companies: set[str]) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path), "exists": False, "removed": 0}
    kept_lines: list[str] = []
    removed_rows: list[dict[str, Any]] = []
    total = 0
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        total += 1
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            kept_lines.append(line)
            continue
        if isinstance(payload, dict) and data_row_is_removed(payload, removed_products, removed_companies):
            removed_rows.append(payload)
        else:
            kept_lines.append(line)
    if not removed_rows:
        return {"path": str(path), "exists": True, "rows": total, "removed": 0}
    backup = path.with_name(f"{path.stem}.backup_before_research_data_service_removal_{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    path.write_text("\n".join(kept_lines) + ("\n" if kept_lines else ""), encoding="utf-8")
    return {
        "path": str(path),
        "backup": str(backup),
        "rows_before": total,
        "rows_after": len(kept_lines),
        "removed": len(removed_rows),
        "removed_preview": removed_rows[:5],
    }


def data_cleanup(stamp: str, workbook_report: dict[str, Any]) -> dict[str, Any]:
    removed_products = workbook_report["removed_products"]
    removed_companies = set(workbook_report["removed_companies"])
    csv_reports = [
        cleanup_csv(DATA_DIR / filename, stamp, removed_products, removed_companies)
        for filename in CSV_CLEANUP_FILES
    ]
    jsonl_reports = [
        cleanup_jsonl(DATA_DIR / filename, stamp, removed_products, removed_companies)
        for filename in JSONL_CLEANUP_FILES
    ]
    return {"csv": csv_reports, "jsonl": jsonl_reports}


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook_report = workbook_cleanup(stamp)
    data_report = data_cleanup(stamp, workbook_report)
    summary = {
        "timestamp": stamp,
        "decision": "Remove scientific research/data-service/training rows from product scope; keep backups for rollback.",
        "workbook": workbook_report,
        "data_files": data_report,
    }
    report_path = AUDIT_DIR / f"research_data_service_removal_{stamp}.json"
    latest_path = AUDIT_DIR / "research_data_service_removal_latest.json"
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
