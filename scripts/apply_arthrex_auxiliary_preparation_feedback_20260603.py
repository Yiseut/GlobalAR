from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TZ = timezone(timedelta(hours=8))
RUN_TS = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")

ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT.parent
WORKBOOK_PATH = DATA_ROOT / "全球医美企业库_标准化版v4.xlsx"
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"

TAG = "arthrex_auxiliary_preparation_feedback_20260603"
AUDIT_MESSAGE = (
    "User feedback 2026-06-03: Arthrex lines are auxiliary tools/devices/consumables "
    "or preparation systems for aesthetic procedures; they are not EBD, not radiation/energy "
    "sources, and not PRP/material ingredients themselves."
)


RECORD_UPDATES: dict[str, dict[str, Any]] = {
    "REC_0008": {
        "Category_L1": "Consumables",
        "Category_L2": "Consumables",
        "Tech_Type_Std": "Autologous adipose preparation system",
        "Verified_Product_Type_CN": "自体脂肪采集/处理制备系统",
        "Market_Channel": "aesthetic procedure support tool / consumable",
        "Material_Taxonomy_L1_CN": "耗材/器械",
        "Material_Taxonomy_L2_CN": "制备系统",
        "Material_Taxonomy_L3_CN": "自体组织/血液制备系统",
        "Material_Taxonomy_Path_CN": "耗材/器械 > 制备系统 > 自体组织/血液制备系统",
        "Material_Taxonomy_Source": f"manual_correction:{TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": (
            "ACA-Kit is an autologous adipose harvesting/processing kit; classified as an "
            "auxiliary preparation tool/consumable rather than PRP or regenerative material ingredient."
        ),
        "Material_Family": "自体组织/血液制备系统",
        "Backfill_Audit": AUDIT_MESSAGE,
    },
    "REC_0049": {
        "Category_L1": "Consumables",
        "Category_L2": "Consumables",
        "Tech_Type_Std": "Autologous blood/PRP preparation system",
        "Verified_Product_Type_CN": "自体血液/PRP 制备系统",
        "Market_Channel": "aesthetic procedure support tool / consumable",
        "Material_Taxonomy_L1_CN": "耗材/器械",
        "Material_Taxonomy_L2_CN": "制备系统",
        "Material_Taxonomy_L3_CN": "自体组织/血液制备系统",
        "Material_Taxonomy_Path_CN": "耗材/器械 > 制备系统 > 自体组织/血液制备系统",
        "Material_Taxonomy_Source": f"manual_correction:{TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": (
            "Arthrex ACP is a closed double-syringe preparation system. It supports PRP/PRF workflows "
            "but is not the PRP material category and is not an energy-based device."
        ),
        "Material_Family": "自体组织/血液制备系统",
        "Backfill_Audit": AUDIT_MESSAGE,
    },
    "REC_0475": {
        "Category_L1": "Surgical",
        "Category_L2": "Surgical",
        "Tech_Type_Std": "Endoscopic surgical visualization tool",
        "Verified_Product_Type_CN": "微创内窥镜/手术可视化辅助工具",
        "Market_Channel": "aesthetic surgery support tool / device",
        "Material_Taxonomy_L1_CN": "耗材/器械",
        "Material_Taxonomy_L2_CN": "手术辅助器械",
        "Material_Taxonomy_L3_CN": "内窥镜/可视化工具",
        "Material_Taxonomy_Path_CN": "耗材/器械 > 手术辅助器械 > 内窥镜/可视化工具",
        "Material_Taxonomy_Source": f"manual_correction:{TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": (
            "NanoScope is an auxiliary endoscopic visualization tool; any energy in the system is not "
            "the direct aesthetic treatment source, so it should not sit in EBD."
        ),
        "Material_Family": "手术辅助器械",
        "Inclusion_Status": "active",
        "Duplicate_Note": "reclassified_by_user_feedback_20260603: auxiliary surgical visualization tool; not EBD.",
        "V4_1_Registration_Review_Status": None,
        "Backfill_Audit": AUDIT_MESSAGE,
    },
}


def clean(value: Any) -> str:
    return "" if value is None else " ".join(str(value).replace("\u00a0", " ").split())


def backup_file(path: Path) -> Path:
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".xlsx":
        backup = path.with_name(f"{path.stem}.backup_before_{TAG}_{RUN_TS}{path.suffix}")
    else:
        backup = path.with_name(f"{path.stem}_backup_before_{TAG}_{RUN_TS}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def headers(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}


def append_audit(existing: Any, message: str) -> str:
    before = clean(existing)
    entry = f"{TAG}: {message}"
    if entry in before:
        return before
    return " | ".join(part for part in [before, entry] if part)


def update_search_blob(blob: Any, replacements: dict[str, Any]) -> str:
    text = clean(blob)
    if not text:
        return text
    pairs = {
        "EBD": replacements.get("Category_L1") or replacements.get("category_l1"),
        "Radiofrequency": replacements.get("Category_L2") or replacements.get("category_l2"),
        "Other EBD": replacements.get("Category_L2") or replacements.get("category_l2"),
        "Regenerative": replacements.get("Category_L1") or replacements.get("category_l1"),
        "PRP-PRF": replacements.get("Category_L2") or replacements.get("category_l2"),
        "注射类 > 自体来源 > PRP 富血小板血浆": replacements.get("material_taxonomy_path_cn")
        or replacements.get("Material_Taxonomy_Path_CN"),
        "注射类": replacements.get("material_taxonomy_l1_cn") or replacements.get("Material_Taxonomy_L1_CN"),
        "自体来源": replacements.get("material_taxonomy_l2_cn") or replacements.get("Material_Taxonomy_L2_CN"),
        "PRP 富血小板血浆": replacements.get("material_taxonomy_l3_cn") or replacements.get("Material_Taxonomy_L3_CN"),
        "耗材/器械 > 针具 > 注射针/钝头套管": replacements.get("material_taxonomy_path_cn")
        or replacements.get("Material_Taxonomy_Path_CN"),
        "针具": replacements.get("material_taxonomy_l2_cn") or replacements.get("Material_Taxonomy_L2_CN"),
        "注射针/钝头套管": replacements.get("material_taxonomy_l3_cn") or replacements.get("Material_Taxonomy_L3_CN"),
        "rule:autologous_prp": f"manual_correction:{TAG}",
        "rule:consumable_needle_cannula": f"manual_correction:{TAG}",
        "auto_applied": "manual_verified",
        "PRP term": "auxiliary preparation/support tool",
        "needle/cannula term": "auxiliary surgical support tool",
    }
    for old, new in pairs.items():
        if new:
            text = text.replace(old, str(new))
    return text


def build_layer(row_values: dict[str, Any], updates: dict[str, Any]) -> str:
    commercial_l1 = updates.get("Category_L1") or updates.get("category_l1")
    commercial_l2 = updates.get("Category_L2") or updates.get("category_l2")
    material_path = updates.get("Material_Taxonomy_Path_CN") or updates.get("material_taxonomy_path_cn")
    tech = updates.get("Tech_Type_Std") or updates.get("tech_type") or row_values.get("tech_type")
    layer = {
        "commercial": f"{commercial_l1} > {commercial_l2}",
        "material_taxonomy": material_path,
        "material_taxonomy_detail": {
            "l1": updates.get("Material_Taxonomy_L1_CN") or updates.get("material_taxonomy_l1_cn"),
            "l2": updates.get("Material_Taxonomy_L2_CN") or updates.get("material_taxonomy_l2_cn"),
            "l3": updates.get("Material_Taxonomy_L3_CN") or updates.get("material_taxonomy_l3_cn"),
            "path": material_path,
            "source": f"manual_correction:{TAG}",
            "confidence": "high",
            "review_status": "manual_verified",
            "note": "Auxiliary preparation/support tool; not EBD and not a material ingredient.",
            "family": updates.get("Material_Family") or updates.get("material_family") or "",
            "inclusion_status": updates.get("Inclusion_Status") or updates.get("inclusion_status") or "active",
            "backfill_audit": AUDIT_MESSAGE,
        },
        "material_family": updates.get("Material_Family") or updates.get("material_family") or "",
        "inclusion_status": updates.get("Inclusion_Status") or updates.get("inclusion_status") or "active",
        "technology": tech,
        "regulatory": "pending_registration_evidence",
        "source_review": row_values.get("verification_status") or "official_product_and_spec_cross_checked",
    }
    return json.dumps(layer, ensure_ascii=False)


def set_cell(row, header_map: dict[str, int], field: str, value: Any) -> bool:
    if field not in header_map:
        return False
    cell = row[header_map[field] - 1]
    before = cell.value
    cell.value = value
    return before != value


def update_workbook() -> dict[str, Any]:
    backup = backup_file(WORKBOOK_PATH)
    wb = load_workbook(WORKBOOK_PATH)
    changed: list[dict[str, Any]] = []
    category_deltas: dict[str, dict[str, int]] = {}

    pl = wb["Product_Lines"]
    pl_headers = headers(pl)
    by_record: dict[str, dict[str, Any]] = {}
    for row in pl.iter_rows(min_row=2):
        record_id = clean(row[pl_headers["Record_ID"] - 1].value)
        updates = RECORD_UPDATES.get(record_id)
        if not updates:
            continue
        before = {k: row[pl_headers[k] - 1].value for k in pl_headers}
        for field, value in updates.items():
            if field == "Backfill_Audit":
                value = append_audit(row[pl_headers[field] - 1].value, value)
            set_cell(row, pl_headers, field, value)
        after = {k: row[pl_headers[k] - 1].value for k in pl_headers}
        if clean(before.get("Category_L1")) != clean(after.get("Category_L1")):
            old_l1 = clean(before.get("Category_L1"))
            new_l1 = clean(after.get("Category_L1"))
            old_brand = clean(before.get("Brand"))
            new_brand = clean(after.get("Brand"))
            old_company = clean(before.get("Company"))
            new_company = clean(after.get("Company"))
            if old_l1:
                category_deltas.setdefault(old_l1, {"Companies": 0, "Brands": 0, "Products": 0})["Products"] -= 1
            if new_l1:
                category_deltas.setdefault(new_l1, {"Companies": 0, "Brands": 0, "Products": 0})["Products"] += 1
            if old_l1 and old_l1 != new_l1:
                still_old_company = any(
                    clean(r[pl_headers["Category_L1"] - 1].value) == old_l1
                    and clean(r[pl_headers["Company"] - 1].value) == old_company
                    for r in pl.iter_rows(min_row=2)
                    if r != row
                )
                still_old_brand = any(
                    clean(r[pl_headers["Category_L1"] - 1].value) == old_l1
                    and clean(r[pl_headers["Company"] - 1].value) == old_company
                    and clean(r[pl_headers["Brand"] - 1].value) == old_brand
                    for r in pl.iter_rows(min_row=2)
                    if r != row
                )
                if not still_old_company:
                    category_deltas.setdefault(old_l1, {"Companies": 0, "Brands": 0, "Products": 0})["Companies"] -= 1
                if not still_old_brand:
                    category_deltas.setdefault(old_l1, {"Companies": 0, "Brands": 0, "Products": 0})["Brands"] -= 1
                already_new_company = any(
                    clean(r[pl_headers["Category_L1"] - 1].value) == new_l1
                    and clean(r[pl_headers["Company"] - 1].value) == new_company
                    for r in pl.iter_rows(min_row=2)
                    if r != row
                )
                already_new_brand = any(
                    clean(r[pl_headers["Category_L1"] - 1].value) == new_l1
                    and clean(r[pl_headers["Company"] - 1].value) == new_company
                    and clean(r[pl_headers["Brand"] - 1].value) == new_brand
                    for r in pl.iter_rows(min_row=2)
                    if r != row
                )
                if new_l1 and not already_new_company:
                    category_deltas.setdefault(new_l1, {"Companies": 0, "Brands": 0, "Products": 0})["Companies"] += 1
                if new_l1 and not already_new_brand:
                    category_deltas.setdefault(new_l1, {"Companies": 0, "Brands": 0, "Products": 0})["Brands"] += 1
        by_record[record_id] = after
        changed.append({"sheet": "Product_Lines", "record_id": record_id, "before_l1": before.get("Category_L1"), "after_l1": after.get("Category_L1")})

    bp = wb["Brand_Portfolio"]
    bp_headers = headers(bp)
    brand_lookup = {clean(v["Brand"]): v for v in by_record.values()}
    for row in bp.iter_rows(min_row=2):
        brand = clean(row[bp_headers["Brand"] - 1].value)
        src = brand_lookup.get(brand)
        if not src:
            continue
        set_cell(row, bp_headers, "Category_L1", src["Category_L1"])
        set_cell(row, bp_headers, "Category_L2", src["Category_L2"])
        set_cell(row, bp_headers, "Tech_Type", src["Tech_Type_Std"])
        changed.append({"sheet": "Brand_Portfolio", "record_id": src["Record_ID"], "before_l1": None, "after_l1": src["Category_L1"]})

    sheet_key = {
        "Product_Master": "seed_record_id",
        "Product_Family_Master": "source_record_ids",
        "Product_SKU_Master": "seed_record_id",
    }
    for sheet_name, record_field in sheet_key.items():
        ws = wb[sheet_name]
        h = headers(ws)
        for row in ws.iter_rows(min_row=2):
            record_id = clean(row[h[record_field] - 1].value)
            if "," in record_id:
                ids = {clean(part) for part in record_id.split(",")}
                record_id = next((rid for rid in RECORD_UPDATES if rid in ids), "")
            if record_id not in by_record:
                continue
            src = by_record[record_id]
            lower_updates = {
                "commercial_path_l1": src["Category_L1"],
                "commercial_path_l2": src["Category_L2"],
                "category_l1": src["Category_L1"],
                "category_l2": src["Category_L2"],
                "material_taxonomy_l1_cn": src["Material_Taxonomy_L1_CN"],
                "material_taxonomy_l2_cn": src["Material_Taxonomy_L2_CN"],
                "material_taxonomy_l3_cn": src["Material_Taxonomy_L3_CN"],
                "material_taxonomy_path_cn": src["Material_Taxonomy_Path_CN"],
                "material_taxonomy_confidence": src["Material_Taxonomy_Confidence"],
                "material_taxonomy_review_status": src["Material_Taxonomy_Review_Status"],
                "material_family": src["Material_Family"],
                "inclusion_status": src["Inclusion_Status"],
                "technology_path_l1": src["Tech_Type_Std"],
                "technology_path_l2": src["Tech_Type_Std"],
                "material_or_energy_source": src["Tech_Type_Std"],
                "tech_type": src["Tech_Type_Std"],
                "Material_Taxonomy_L1_CN": src["Material_Taxonomy_L1_CN"],
                "Material_Taxonomy_L2_CN": src["Material_Taxonomy_L2_CN"],
                "Material_Taxonomy_L3_CN": src["Material_Taxonomy_L3_CN"],
                "Material_Taxonomy_Path_CN": src["Material_Taxonomy_Path_CN"],
                "Material_Taxonomy_Source": src["Material_Taxonomy_Source"],
                "Material_Taxonomy_Confidence": src["Material_Taxonomy_Confidence"],
                "Material_Taxonomy_Review_Status": src["Material_Taxonomy_Review_Status"],
                "Material_Taxonomy_Note": src["Material_Taxonomy_Note"],
                "Material_Family": src["Material_Family"],
            }
            row_values = {name: row[idx - 1].value for name, idx in h.items()}
            if "classification_layer" in h:
                lower_updates["classification_layer"] = build_layer(row_values, lower_updates)
            if "Backfill_Audit" in h:
                lower_updates["Backfill_Audit"] = append_audit(row[h["Backfill_Audit"] - 1].value, AUDIT_MESSAGE)
            if "search_blob" in h:
                lower_updates["search_blob"] = update_search_blob(row[h["search_blob"] - 1].value, lower_updates)
            for field, value in lower_updates.items():
                set_cell(row, h, field, value)
            changed.append({"sheet": sheet_name, "record_id": record_id, "before_l1": None, "after_l1": src["Category_L1"]})

    if "Track_Distribution" in wb.sheetnames:
        td = wb["Track_Distribution"]
        td_headers = headers(td)
        row_by_l1 = {clean(row[td_headers["Category_L1"] - 1].value): row for row in td.iter_rows(min_row=2)}
        for category, deltas in category_deltas.items():
            row = row_by_l1.get(category)
            if not row:
                continue
            for field, delta in deltas.items():
                if field not in td_headers:
                    continue
                cell = row[td_headers[field] - 1]
                cell.value = (cell.value or 0) + delta
            changed.append({"sheet": "Track_Distribution", "record_id": category, "before_l1": None, "after_l1": category})

    wb.save(WORKBOOK_PATH)
    return {"workbook_backup": str(backup), "changed": changed}


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)


def update_csv_layers() -> list[dict[str, Any]]:
    outputs: list[dict[str, Any]] = []
    for filename, record_field in {
        "product_master.csv": "seed_record_id",
        "product_family_master.csv": "source_record_ids",
        "product_sku_master.csv": "seed_record_id",
    }.items():
        path = DATA_DIR / filename
        if not path.exists():
            continue
        backup = backup_file(path)
        fields, rows = read_csv(path)
        changed = 0
        for row in rows:
            record_id = clean(row.get(record_field))
            if "," in record_id:
                ids = {clean(part) for part in record_id.split(",")}
                record_id = next((rid for rid in RECORD_UPDATES if rid in ids), "")
            updates = RECORD_UPDATES.get(record_id)
            if not updates:
                continue
            lower_updates = {
                "commercial_path_l1": updates["Category_L1"],
                "commercial_path_l2": updates["Category_L2"],
                "category_l1": updates["Category_L1"],
                "category_l2": updates["Category_L2"],
                "material_taxonomy_l1_cn": updates["Material_Taxonomy_L1_CN"],
                "material_taxonomy_l2_cn": updates["Material_Taxonomy_L2_CN"],
                "material_taxonomy_l3_cn": updates["Material_Taxonomy_L3_CN"],
                "material_taxonomy_path_cn": updates["Material_Taxonomy_Path_CN"],
                "material_taxonomy_confidence": "high",
                "material_taxonomy_review_status": "manual_verified",
                "material_family": updates["Material_Family"],
                "inclusion_status": updates.get("Inclusion_Status", row.get("inclusion_status", "active") or "active"),
                "technology_path_l1": updates["Tech_Type_Std"],
                "technology_path_l2": updates["Tech_Type_Std"],
                "material_or_energy_source": updates["Tech_Type_Std"],
                "tech_type": updates["Tech_Type_Std"],
                "Material_Taxonomy_L1_CN": updates["Material_Taxonomy_L1_CN"],
                "Material_Taxonomy_L2_CN": updates["Material_Taxonomy_L2_CN"],
                "Material_Taxonomy_L3_CN": updates["Material_Taxonomy_L3_CN"],
                "Material_Taxonomy_Path_CN": updates["Material_Taxonomy_Path_CN"],
                "Material_Taxonomy_Source": updates["Material_Taxonomy_Source"],
                "Material_Taxonomy_Confidence": "high",
                "Material_Taxonomy_Review_Status": "manual_verified",
                "Material_Taxonomy_Note": updates["Material_Taxonomy_Note"],
                "Material_Family": updates["Material_Family"],
            }
            if "classification_layer" in fields:
                lower_updates["classification_layer"] = build_layer(row, lower_updates)
            if "Backfill_Audit" in fields:
                lower_updates["Backfill_Audit"] = append_audit(row.get("Backfill_Audit"), AUDIT_MESSAGE)
            if "search_blob" in fields:
                lower_updates["search_blob"] = update_search_blob(row.get("search_blob"), lower_updates)
            for field, value in lower_updates.items():
                if field in fields:
                    row[field] = "" if value is None else str(value)
            changed += 1
        write_csv(path, fields, rows)
        outputs.append({"file": str(path), "backup": str(backup), "changed_rows": changed})
    return outputs


def write_audit(result: dict[str, Any]) -> Path:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    audit_path = AUDIT_DIR / f"{TAG}_{RUN_TS}.csv"
    latest_path = AUDIT_DIR / f"{TAG}_latest.csv"
    rows = result["workbook"]["changed"]
    with audit_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["sheet", "record_id", "before_l1", "after_l1"])
        writer.writeheader()
        writer.writerows(rows)
    shutil.copy2(audit_path, latest_path)
    summary_path = AUDIT_DIR / f"{TAG}_latest.json"
    summary_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return audit_path


def main() -> None:
    result = {
        "tag": TAG,
        "run_ts": RUN_TS,
        "workbook": update_workbook(),
        "csv_layers": update_csv_layers(),
    }
    audit_path = write_audit(result)
    print(json.dumps({"audit": str(audit_path), **result}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
