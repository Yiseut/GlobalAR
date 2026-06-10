from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TZ = timezone(timedelta(hours=8))
RUN_TS = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = REPO_ROOT.parent
WORKBOOK_PATH = DATA_ROOT / "全球医美企业库_标准化版v4.xlsx"
AUDIT_DIR = REPO_ROOT / "data" / "audits"

TARGET_RECORD_ID = "REC_0303"
TARGET_PRODUCT_ID = "prod_439b182b8315"
OLD_FAMILY_ID = "pf_ab942908f374"

TARGET_COMPANY = "Arion"
TARGET_BRAND = "Gluteal / Calf Implants"
TARGET_PRODUCT_FAMILY = "臀部/身体假体"
TARGET_CATEGORY_L1 = "Implants"
TARGET_CATEGORY_L2 = "Other Implant"
TARGET_TECH_TYPE = "Silicone Body Implant"
TARGET_MATERIAL_L1 = "植入物"
TARGET_MATERIAL_L2 = "身体假体"
TARGET_MATERIAL_L3 = "硅胶假体"
TARGET_MATERIAL_PATH = "植入物 > 身体假体 > 硅胶假体"
TARGET_MATERIAL_FAMILY = "硅胶假体"

CORRECTION_TAG = "rec0303_implant_fix_20260602"
CORRECTION_NOTE = (
    "REC_0303 is gluteal/calf silicone body implant; corrected from legacy fallback "
    "classification to implant taxonomy based on existing backfill_audit and material taxonomy contradiction."
)
BACKFILL_APPEND = (
    "rec0303_implant_fix_20260602: corrected legacy commercial classification to "
    "Implants/Other Implant; body silicone implant taxonomy applied."
)


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return " ".join(text.replace("\u00a0", " ").split())


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(norm(part).lower() for part in parts if norm(part))
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12] if raw else "0" * 12
    return f"{prefix}_{digest}"


TARGET_FAMILY_ID = stable_id(
    "pf",
    TARGET_COMPANY,
    TARGET_BRAND,
    TARGET_PRODUCT_FAMILY,
    TARGET_CATEGORY_L1,
    TARGET_CATEGORY_L2,
    TARGET_TECH_TYPE,
)


def backup_file(path: Path) -> Path:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".xlsx":
        backup = path.with_name(f"{path.stem}.backup_before_{CORRECTION_TAG}_{RUN_TS}{path.suffix}")
    else:
        backup = AUDIT_DIR / f"{path.stem}.backup_before_{CORRECTION_TAG}_{RUN_TS}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def sheet_header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = norm(cell.value)
        if key:
            headers[key] = idx
    return headers


def update_cell(row, headers: dict[str, int], field: str, value: Any) -> tuple[str, str]:
    if field not in headers:
        raise KeyError(f"Missing Product_Lines column: {field}")
    cell = row[headers[field] - 1]
    before = "" if cell.value is None else str(cell.value)
    cell.value = value
    return before, "" if value is None else str(value)


def update_workbook() -> dict[str, Any]:
    backup = backup_file(WORKBOOK_PATH)
    wb = load_workbook(WORKBOOK_PATH)
    if "Product_Lines" not in wb.sheetnames:
        raise RuntimeError("Workbook has no Product_Lines sheet")
    ws = wb["Product_Lines"]
    headers = sheet_header_map(ws)
    if "Record_ID" not in headers:
        raise RuntimeError("Product_Lines has no Record_ID column")

    target_row = None
    for row in ws.iter_rows(min_row=2):
        if norm(row[headers["Record_ID"] - 1].value) == TARGET_RECORD_ID:
            target_row = row
            break
    if target_row is None:
        raise RuntimeError(f"Could not find {TARGET_RECORD_ID} in Product_Lines")

    changes: dict[str, dict[str, str]] = {}
    updates = {
        "Category_L1": TARGET_CATEGORY_L1,
        "Category_L2": TARGET_CATEGORY_L2,
        "Tech_Type_Std": TARGET_TECH_TYPE,
        "Material_Taxonomy_L1_CN": TARGET_MATERIAL_L1,
        "Material_Taxonomy_L2_CN": TARGET_MATERIAL_L2,
        "Material_Taxonomy_L3_CN": TARGET_MATERIAL_L3,
        "Material_Taxonomy_Path_CN": TARGET_MATERIAL_PATH,
        "Material_Taxonomy_Source": f"manual_correction:{CORRECTION_TAG}",
        "Material_Taxonomy_Confidence": "high",
        "Material_Taxonomy_Review_Status": "manual_verified",
        "Material_Taxonomy_Note": CORRECTION_NOTE,
        "Material_Family": TARGET_MATERIAL_FAMILY,
    }
    for field, value in updates.items():
        before, after = update_cell(target_row, headers, field, value)
        if before != after:
            changes[field] = {"before": before, "after": after}

    if "Backfill_Audit" not in headers:
        raise KeyError("Missing Product_Lines column: Backfill_Audit")
    audit_cell = target_row[headers["Backfill_Audit"] - 1]
    before_audit = "" if audit_cell.value is None else str(audit_cell.value)
    updated_audit = before_audit.replace(
        "rec0303_implant_fix_20260602: corrected commercial classification from "
        "EBD/Other EBD to Implants/Other Implant; body silicone implant taxonomy applied.",
        BACKFILL_APPEND,
    )
    if CORRECTION_TAG not in updated_audit:
        updated_audit = " | ".join(x for x in [updated_audit, BACKFILL_APPEND] if x)
    audit_cell.value = updated_audit
    after_audit = "" if audit_cell.value is None else str(audit_cell.value)
    if before_audit != after_audit:
        changes["Backfill_Audit"] = {"before": before_audit, "after": after_audit}

    wb.save(WORKBOOK_PATH)
    return {
        "path": str(WORKBOOK_PATH),
        "backup": str(backup),
        "record_id": TARGET_RECORD_ID,
        "target_family_id": TARGET_FAMILY_ID,
        "changes": changes,
    }


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise RuntimeError(f"{path} has no header")
        return list(reader.fieldnames), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def row_matches_product(row: dict[str, str]) -> bool:
    return (
        norm(row.get("company")) == TARGET_COMPANY
        and norm(row.get("brand")) == TARGET_BRAND
        and (
            norm(row.get("product_family_id")) in {OLD_FAMILY_ID, TARGET_FAMILY_ID}
            or norm(row.get("product_id")) == TARGET_PRODUCT_ID
            or norm(row.get("seed_record_id")) == TARGET_RECORD_ID
        )
    )


def update_plan_rows(row: dict[str, str]) -> bool:
    if not row_matches_product(row):
        return False
    before = dict(row)
    row["product_family_id"] = TARGET_FAMILY_ID
    row["category_l1"] = TARGET_CATEGORY_L1
    row["category_l2"] = TARGET_CATEGORY_L2
    row["tech_type"] = TARGET_TECH_TYPE
    query = row.get("query", "")
    query = query.replace(" EBD ", f" {TARGET_CATEGORY_L1} ")
    query = query.replace(" EBD Silicone Body Implant", f" {TARGET_CATEGORY_L1} {TARGET_TECH_TYPE}")
    query = query.replace(" Other EBD ", f" {TARGET_CATEGORY_L2} ")
    row["query"] = " ".join(query.split())
    return row != before


def update_family_only(row: dict[str, str]) -> bool:
    if not row_matches_product(row):
        return False
    before = row.get("product_family_id", "")
    row["product_family_id"] = TARGET_FAMILY_ID
    return row.get("product_family_id", "") != before


def update_csv(path: Path, updater) -> dict[str, Any]:
    fieldnames, rows = read_csv(path)
    changed = 0
    for row in rows:
        if updater(row):
            changed += 1
    backup = None
    if changed:
        backup = backup_file(path)
        write_csv(path, fieldnames, rows)
    return {"path": str(path), "backup": str(backup) if backup else "", "changed_rows": changed}


def main() -> None:
    if TARGET_FAMILY_ID != "pf_f0c2e4808248":
        raise RuntimeError(f"Unexpected target family id: {TARGET_FAMILY_ID}")

    workbook_result = update_workbook()
    csv_results = [
        update_csv(REPO_ROOT / "data" / "company_official_source_plan.csv", update_plan_rows),
        update_csv(REPO_ROOT / "data" / "manual_product_fact_evidence.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "product_specification_evidence.csv", update_family_only),
        update_csv(REPO_ROOT / "data" / "company_media_asset_index.csv", update_family_only),
    ]

    audit = {
        "run_ts": RUN_TS,
        "correction": CORRECTION_TAG,
        "record_id": TARGET_RECORD_ID,
        "product_id": TARGET_PRODUCT_ID,
        "old_family_id": OLD_FAMILY_ID,
        "new_family_id": TARGET_FAMILY_ID,
        "classification": {
            "category_l1": TARGET_CATEGORY_L1,
            "category_l2": TARGET_CATEGORY_L2,
            "tech_type": TARGET_TECH_TYPE,
            "material_taxonomy_path_cn": TARGET_MATERIAL_PATH,
            "review_status": "manual_verified",
        },
        "workbook": workbook_result,
        "csv_updates": csv_results,
    }
    audit_path = AUDIT_DIR / f"{CORRECTION_TAG}_{RUN_TS}.json"
    latest_path = AUDIT_DIR / f"{CORRECTION_TAG}_latest.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    latest_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
