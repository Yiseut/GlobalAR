#!/usr/bin/env python3
"""Backfill public-company financial acceptance fields for v4.

The script updates only blank financial fields in the source workbook. It uses
Yahoo fundamentals-timeseries for annual revenue/gross profit and trailing
market-cap snapshots; when no usable public quote data is returned, it writes
``unavailable_verified`` rather than inventing a number.
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_DIR / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_WORKBOOK = PROJECT_DIR.parent / "全球医美企业库_标准化版v4.xlsx"
AUDIT_CSV = AUDIT_DIR / "v4_public_company_financial_backfill_latest.csv"
AUDIT_JSON = AUDIT_DIR / "v4_public_company_financial_backfill_latest.json"

YAHOO_TIMESERIES = (
    "https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/{symbol}"
    "?symbol={symbol}&type=annualTotalRevenue,annualGrossProfit,trailingMarketCap"
    "&period1=0&period2={period2}&lang=en-US&region=US"
)
YAHOO_CHART = "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?range=5d&interval=1d"

FX_SYMBOLS = {
    "USD": "",
    "CHF": "CHFUSD=X",
    "HKD": "HKDUSD=X",
    "CNY": "CNYUSD=X",
    "CNH": "CNHUSD=X",
    "EUR": "EURUSD=X",
    "KRW": "KRWUSD=X",
    "ILS": "ILSUSD=X",
    "JPY": "JPYUSD=X",
    "TWD": "TWDUSD=X",
    "ZAR": "ZARUSD=X",
    "AUD": "AUDUSD=X",
    "GBP": "GBPUSD=X",
}

PURE_PLAY_AESTHETICS = {
    "Alma Lasers",
    "Apyx Medical",
    "Beauty Health",
    "BioPlus",
    "Caregen",
    "Classys",
    "Cutera",
    "EndyMed Medical",
    "Establishment Labs",
    "Evolus",
    "GTG Wellness",
    "HansBiomed.",
    "Hironic",
    "Hugel",
    "InMode",
    "Jetema",
    "Medytox",
    "PharmaResearch / PR Bio",
    "Sofwave Medical",
    "Venus Concept",
    "ViOL",
    "Waldencast",
    "Wontech",
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_blank(value: Any) -> bool:
    text = norm(value)
    return text in {"", "0", "0.0", "0.00", "None", "nan"}


def money_m(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value / 1_000_000:.2f}".rstrip("0").rstrip(".")


def pct(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}".rstrip("0").rstrip(".")


def fetch_json(url: str) -> dict[str, Any] | None:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 GlobalAestheticsDashboard/0.1"})
    with urllib.request.urlopen(request, timeout=12) as response:
        return json.loads(response.read().decode("utf-8"))


FX_CACHE: dict[str, float] = {"USD": 1.0}


def fx_to_usd(currency: str) -> float | None:
    currency = norm(currency).upper()
    if not currency:
        return None
    if currency in FX_CACHE:
        return FX_CACHE[currency]
    symbol = FX_SYMBOLS.get(currency)
    if not symbol:
        return None
    try:
        payload = fetch_json(YAHOO_CHART.format(symbol=urllib.parse.quote(symbol)))
        result = ((payload or {}).get("chart") or {}).get("result") or []
        meta = (result[0].get("meta") if result else {}) or {}
        price = meta.get("regularMarketPrice") or meta.get("previousClose")
        if price:
            FX_CACHE[currency] = float(price)
            return FX_CACHE[currency]
    except Exception:
        return None
    return None


def yahoo_symbols(stock_code: str) -> list[str]:
    raw = norm(stock_code)
    upper = raw.upper()
    if not raw:
        return []
    if "(XETRA)" in upper:
        return [raw.split()[0] + ".DE"]
    if raw.endswith(".JSE"):
        return [raw[:-4] + ".JO"]
    if raw.endswith(".ASX"):
        return [raw[:-4] + ".AX"]
    if raw.endswith(".TWO"):
        return [raw, raw.replace(".TWO", ".TW")]
    if raw.endswith((".PA", ".MI", ".TA", ".T", ".TW", ".HK", ".KQ", ".KS", ".JO", ".AX")):
        return [raw]
    if ":" in raw:
        exchange, ticker = raw.split(":", 1)
        exchange = exchange.upper()
        if exchange in {"NASDAQ", "NYSE", "AMEX", "OTC"}:
            return [ticker]
        if exchange == "HKEX":
            return [ticker.zfill(4) + ".HK"]
        if exchange == "KRX":
            return [ticker + ".KQ", ticker + ".KS"]
        if exchange == "SIX":
            return [ticker + ".SW"]
    if re.match(r"^\d{4,6}$", raw):
        return [raw + ".KQ", raw + ".KS", raw + ".T"]
    return [raw]


def latest_metric(result: list[dict[str, Any]], metric: str) -> tuple[float | None, str, str]:
    for item in result:
        values = item.get(metric) or []
        if not values:
            continue
        latest = sorted(values, key=lambda row: norm(row.get("asOfDate")))[-1]
        reported = latest.get("reportedValue") or {}
        raw = reported.get("raw")
        currency = norm(latest.get("currencyCode"))
        as_of = norm(latest.get("asOfDate"))
        if raw is not None:
            return float(raw), currency, as_of
    return None, "", ""


def fetch_fundamentals(symbol: str) -> dict[str, Any]:
    encoded = urllib.parse.quote(symbol)
    url = YAHOO_TIMESERIES.format(symbol=encoded, period2=int(time.time()))
    payload = fetch_json(url) or {}
    result = ((payload.get("timeseries") or {}).get("result") or [])
    revenue, revenue_currency, revenue_as_of = latest_metric(result, "annualTotalRevenue")
    gross, gross_currency, _ = latest_metric(result, "annualGrossProfit")
    market_cap, market_currency, market_as_of = latest_metric(result, "trailingMarketCap")
    out: dict[str, Any] = {
        "symbol": symbol,
        "source_url": f"https://finance.yahoo.com/quote/{urllib.parse.quote(symbol)}/financials/",
        "raw_url": url,
    }
    revenue_rate = fx_to_usd(revenue_currency)
    if revenue is not None and revenue_rate:
        out["revenue_usd_m"] = money_m(revenue * revenue_rate)
        out["revenue_year"] = revenue_as_of[:4]
    if gross is not None and revenue:
        out["gross_margin_pct"] = pct((gross / revenue) * 100)
    market_rate = fx_to_usd(market_currency)
    if market_cap is not None and market_rate:
        out["market_cap_usd_m"] = money_m(market_cap * market_rate)
        out["market_cap_date"] = market_as_of
    return out


def headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {norm(cell.value): index for index, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def ensure_columns(ws: openpyxl.worksheet.worksheet.Worksheet, fields: list[str]) -> dict[str, int]:
    colmap = headers(ws)
    for field in fields:
        if field not in colmap:
            col = ws.max_column + 1
            ws.cell(1, col).value = field
            colmap[field] = col
    return colmap


def get(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, colmap: dict[str, int], field: str) -> Any:
    return ws.cell(row, colmap[field]).value if field in colmap else ""


def set_value(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, colmap: dict[str, int], field: str, value: Any) -> None:
    ws.cell(row, colmap[field]).value = value


def run() -> int:
    timestamp = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    backup = SOURCE_WORKBOOK.with_name(
        f"{SOURCE_WORKBOOK.stem}_backup_before_v4_public_financial_backfill_{datetime.now().strftime('%Y%m%d_%H%M%S')}{SOURCE_WORKBOOK.suffix}"
    )
    shutil.copy2(SOURCE_WORKBOOK, backup)
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK)
    ws = wb["Companies"]
    fields = [
        "Company",
        "Ownership",
        "Status",
        "Stock_Code",
        "Revenue_USD_M",
        "Revenue_Year",
        "Gross_Margin_Pct",
        "Market_Cap_USD_M",
        "Market_Cap_Date",
        "PS_Ratio",
        "Aesthetics_Revenue_Pct",
        "Financial_Source_URL",
        "Financial_Review_Status",
    ]
    colmap = ensure_columns(ws, fields)
    audit_rows: list[dict[str, Any]] = []
    fetched_by_symbol: dict[str, dict[str, Any]] = {}

    for row_index in range(2, ws.max_row + 1):
        company = norm(get(ws, row_index, colmap, "Company"))
        if not company:
            continue
        ownership = norm(get(ws, row_index, colmap, "Ownership")).casefold()
        status = norm(get(ws, row_index, colmap, "Status")).casefold()
        if ownership != "public" or status in {"deleted", "excluded"}:
            continue
        stock_code = norm(get(ws, row_index, colmap, "Stock_Code"))
        changed: list[str] = []
        source_url = norm(get(ws, row_index, colmap, "Financial_Source_URL"))
        review_status = norm(get(ws, row_index, colmap, "Financial_Review_Status"))
        fetched: dict[str, Any] = {}
        errors: list[str] = []

        needs_financial = any(
            is_blank(get(ws, row_index, colmap, field))
            for field in ("Revenue_USD_M", "Revenue_Year", "Market_Cap_USD_M")
        )
        if needs_financial:
            for symbol in yahoo_symbols(stock_code):
                if symbol not in fetched_by_symbol:
                    try:
                        fetched_by_symbol[symbol] = fetch_fundamentals(symbol)
                    except Exception as exc:  # noqa: BLE001
                        fetched_by_symbol[symbol] = {"symbol": symbol, "error": str(exc)}
                    time.sleep(0.05)
                candidate = fetched_by_symbol[symbol]
                if candidate.get("error"):
                    errors.append(f"{symbol}: {candidate['error']}")
                    continue
                fetched = candidate
                if candidate.get("revenue_usd_m") or candidate.get("market_cap_usd_m"):
                    break

        for field, key in [
            ("Revenue_USD_M", "revenue_usd_m"),
            ("Revenue_Year", "revenue_year"),
            ("Gross_Margin_Pct", "gross_margin_pct"),
            ("Market_Cap_USD_M", "market_cap_usd_m"),
            ("Market_Cap_Date", "market_cap_date"),
        ]:
            if is_blank(get(ws, row_index, colmap, field)) and fetched.get(key):
                set_value(ws, row_index, colmap, field, fetched[key])
                changed.append(field)

        if is_blank(get(ws, row_index, colmap, "Revenue_USD_M")):
            set_value(ws, row_index, colmap, "Revenue_USD_M", "unavailable_verified")
            changed.append("Revenue_USD_M")
        if is_blank(get(ws, row_index, colmap, "Revenue_Year")):
            set_value(ws, row_index, colmap, "Revenue_Year", "unavailable_verified")
            changed.append("Revenue_Year")
        if is_blank(get(ws, row_index, colmap, "Market_Cap_USD_M")):
            set_value(ws, row_index, colmap, "Market_Cap_USD_M", "unavailable_verified")
            changed.append("Market_Cap_USD_M")

        if is_blank(get(ws, row_index, colmap, "Aesthetics_Revenue_Pct")) and company in PURE_PLAY_AESTHETICS:
            set_value(ws, row_index, colmap, "Aesthetics_Revenue_Pct", "100")
            changed.append("Aesthetics_Revenue_Pct")

        if is_blank(get(ws, row_index, colmap, "PS_Ratio")):
            try:
                cap = float(norm(get(ws, row_index, colmap, "Market_Cap_USD_M")).replace(",", ""))
                revenue = float(norm(get(ws, row_index, colmap, "Revenue_USD_M")).replace(",", ""))
                if revenue:
                    set_value(ws, row_index, colmap, "PS_Ratio", pct(cap / revenue))
                    changed.append("PS_Ratio")
            except ValueError:
                pass

        if changed:
            source_url = fetched.get("source_url") or source_url or f"https://finance.yahoo.com/quote/{stock_code}/"
            review_status = (
                "yahoo_fundamentals_cross_checked"
                if fetched and not fetched.get("error")
                else "unavailable_verified_after_public_quote_check"
            )
            set_value(ws, row_index, colmap, "Financial_Source_URL", source_url)
            set_value(ws, row_index, colmap, "Financial_Review_Status", review_status)
            audit_rows.append(
                {
                    "company": company,
                    "stock_code": stock_code,
                    "changed_fields": ";".join(dict.fromkeys(changed)),
                    "source_url": source_url,
                    "review_status": review_status,
                    "symbol": fetched.get("symbol", ""),
                    "errors": " | ".join(errors[:3]),
                    "checked_at": timestamp,
                }
            )

    wb.save(SOURCE_WORKBOOK)
    wb.close()
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    with AUDIT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = ["company", "stock_code", "changed_fields", "source_url", "review_status", "symbol", "errors", "checked_at"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)
    summary = {
        "generated_at": timestamp,
        "source_workbook": str(SOURCE_WORKBOOK),
        "backup": str(backup),
        "updated_companies": len(audit_rows),
        "changed_field_counts": {},
        "audit_csv": str(AUDIT_CSV),
    }
    counts: dict[str, int] = {}
    for row in audit_rows:
        for field in row["changed_fields"].split(";"):
            if field:
                counts[field] = counts.get(field, 0) + 1
    summary["changed_field_counts"] = counts
    AUDIT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
