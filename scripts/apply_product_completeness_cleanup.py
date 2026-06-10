"""Apply targeted product-completeness fixes from the 2026-06-01 QA pass."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "Product_Master.csv"
MANUAL_FACT_PATH = DATA_DIR / "manual_product_fact_evidence.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def append_unique(rows: list[dict[str, str]], key_field: str, new_rows: list[dict[str, str]]) -> int:
    existing = {norm(row.get(key_field)) for row in rows}
    added = 0
    for row in new_rows:
        key = norm(row.get(key_field))
        if not key or key in existing:
            continue
        rows.append(row)
        existing.add(key)
        added += 1
    return added


def product_lookup() -> dict[str, dict[str, str]]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    return {norm(row.get("seed_record_id")): row for row in rows if norm(row.get("seed_record_id"))}


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_product_completeness_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)
    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    row_by_id = {
        norm(ws.cell(row=row, column=colmap["Record_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value)
    }
    changes: list[dict[str, str]] = []

    def set_cell(record_id: str, field: str, value: object) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get(field)
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        new = norm(value)
        if old == new:
            return
        ws.cell(row=row, column=col, value=value)
        changes.append({"record_id": record_id, "field": field, "old": old, "new": new})

    def append_audit(record_id: str, note: str) -> None:
        row = row_by_id.get(record_id)
        col = colmap.get("Backfill_Audit")
        if not row or not col:
            return
        old = norm(ws.cell(row=row, column=col).value)
        if note in old:
            return
        ws.cell(row=row, column=col, value=f"{old}; {note}".strip("; "))
        changes.append({"record_id": record_id, "field": "Backfill_Audit", "old": old, "new": norm(ws.cell(row=row, column=col).value)})

    set_cell("REC_0266", "Manufactured_By", "CG Bio Co., Ltd.")
    set_cell("REC_0266", "Marketing_Holder", "CG Bio Co., Ltd.")
    append_audit("REC_0266", "product_completeness_20260601: FACETEM kept as the active CGBIO product line; official use captured in indication evidence; Korea/MFDS number not confirmed.")

    set_cell("REC_0960", "Is_Primary_Record", 0)
    set_cell("REC_0960", "Duplicate_Note", "duplicate_of:REC_0266; FACETEM active product line is held under CGBIO/Cell Growth Bio; Daewoong kept only as parent/affiliate trace.")
    append_audit("REC_0960", "product_completeness_20260601: non-primary duplicate removed from dashboard/product-gap scope.")

    for record_id in ["REC_0479", "REC_1023"]:
        set_cell(record_id, "Manufactured_By", "MATEX LAB SPA")
        set_cell(record_id, "Marketing_Holder", "Matex Lab")
        set_cell(record_id, "Data_Source", "official_company_fact_override")
    set_cell("REC_0479", "CE_Status", "EU MDR Class III medical device (official Neauvia statement)")
    append_audit("REC_0479", "product_completeness_20260601: official Neauvia Intense page/factsheet captured; public certificate number not available.")
    append_audit("REC_1023", "product_completeness_20260601: official Neauvia Stimulate page captured; official use promoted from product page.")

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_fact_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    specs = {
        "REC_0479": [
            ("official_product_page", "official_product_page", "Neauvia Intense official product page", "https://www.neauvia.com/product/intense/", "INTENSE", "Official Neauvia product page for INTENSE."),
            ("official_specification_candidate", "ha_concentration", "28 mg/ml", "https://www.neauvia.com/wp-content/uploads/2025/07/Exe_ProductFactSheet_INTENSE_210x297_LR.pdf", "INTENSE factsheet", "HA concentration 28 mg/ml; syringe 1x1ml; area of injection Face; injection plane Subcutis and/or over the periosteum."),
            ("official_specification_candidate", "injection_plane", "Subcutis and/or over the periosteum", "https://www.neauvia.com/wp-content/uploads/2025/07/Exe_ProductFactSheet_INTENSE_210x297_LR.pdf", "INTENSE factsheet", "Injection plane listed as Subcutis and/or over the periosteum."),
        ],
        "REC_1023": [
            ("official_product_page", "official_product_page", "Neauvia Stimulate official product page", "https://www.neauvia.com/product/stimulate/", "STIMULATE", "Official Neauvia product page for STIMULATE."),
            ("official_specification_candidate", "ha_concentration", "26 mg/ml", "https://www.neauvia.com/product/stimulate/", "STIMULATE", "HA concentration 26 mg/ml; ingredients include L-Proline, Calcium Hydroxyapatite (CaHA) and Glycine."),
            ("official_specification_candidate", "composition", "PEG-HA + CaHA hydrogel", "https://www.neauvia.com/product/stimulate/", "STIMULATE", "The product page describes STIMULATE as PEG-HA + CaHA hydrogel with Glycine and L-Proline."),
        ],
    }
    rows: list[dict[str, str]] = []
    for record_id, items in specs.items():
        product = products.get(record_id, {})
        for fact_group, field_name, field_value, url, title, excerpt in items:
            rows.append(
                {
                    "fact_id": stable_id("pfact", record_id, fact_group, field_name, url, field_value),
                    "product_id": product.get("product_id", ""),
                    "seed_record_id": record_id,
                    "company_id": product.get("company_id", ""),
                    "company": product.get("company", ""),
                    "brand": product.get("brand", ""),
                    "product_family_id": "",
                    "standard_product_name": product.get("standard_product_name", ""),
                    "priority": "P0",
                    "fact_group": fact_group,
                    "field_name": field_name,
                    "field_value": field_value,
                    "source_url": url,
                    "evidence_title": title,
                    "evidence_excerpt": excerpt,
                    "source_type": "official_product_page" if fact_group == "official_product_page" else "official_product_factsheet",
                    "confidence": "official_site_cross_checked",
                    "captured_at": checked_at,
                    "promoted_at": checked_at,
                    "review_status": "auto_cross_checked",
                    "note": "product_completeness_20260601",
                }
            )
    return rows


def build_indication_rows(products: dict[str, dict[str, str]], checked_at: str) -> list[dict[str, str]]:
    definitions = {
        "REC_0266": {
            "jurisdiction": "Global",
            "regulator": "Official labeling",
            "pathway": "Official product page",
            "status": "Official intended use",
            "registered_name": "FACETEM / FACETEM S",
            "approved_indication": "This device is intended to be used for tissue augmentation. It is recommended to be used for shaping the contours of the face and for the correction of wrinkles and folds and mid and/or deep depression of the skin.",
            "source_url": "https://www.cgbio.co.kr/en/product/aesthetics/facetem",
            "source_type": "official_product_page",
            "evidence_title": "CGBIO FACETEM official product page",
            "excerpt": "Official CGBIO product page states intended use and indications for FACETEM/FACETEM S.",
        },
        "REC_0479": {
            "jurisdiction": "EU / Global",
            "regulator": "CE/MDR",
            "pathway": "Official product page / factsheet",
            "status": "EU MDR Class III medical device (official Neauvia statement)",
            "registered_name": "INTENSE",
            "approved_indication": "Intense is a biodegradable Hyaluronic Acid hydrogel crosslinked with PEG. It is resorbed over time and intended to restore lost volume of the soft tissue.",
            "source_url": "https://www.neauvia.com/wp-content/uploads/2025/07/Exe_ProductFactSheet_INTENSE_210x297_LR.pdf",
            "source_type": "official_product_factsheet",
            "evidence_title": "Neauvia INTENSE factsheet",
            "excerpt": "Official factsheet states Intense is intended to restore lost volume of the soft tissue.",
        },
        "REC_1023": {
            "jurisdiction": "EU / Global",
            "regulator": "CE/MDR",
            "pathway": "Official product page",
            "status": "EU MDR Class III medical device (official Neauvia statement)",
            "registered_name": "STIMULATE",
            "approved_indication": "STIMULATE is a biodegradable crosslinked hyaluronic acid hydrogel resorbed over time, intended to be injected into the subdermis and/or over the periosteum to restore lost volume of the soft tissue.",
            "source_url": "https://www.neauvia.com/product/stimulate/",
            "source_type": "official_product_page",
            "evidence_title": "Neauvia STIMULATE official product page",
            "excerpt": "Official Neauvia product page states the intended injection plane and intended restoration of lost soft-tissue volume.",
        },
    }
    rows: list[dict[str, str]] = []
    for record_id, item in definitions.items():
        product = products.get(record_id, {})
        rows.append(
            {
                "product_id": product.get("product_id", ""),
                "seed_record_id": record_id,
                "company_id": product.get("company_id", ""),
                "company": product.get("company", ""),
                "brand": product.get("brand", ""),
                "jurisdiction": item["jurisdiction"],
                "regulator": item["regulator"],
                "regulatory_pathway": item["pathway"],
                "status": item["status"],
                "registration_no": "",
                "approval_date": "",
                "expiry_date": "",
                "registered_name": item["registered_name"],
                "approved_indication": item["approved_indication"],
                "intended_use": item["approved_indication"],
                "legal_manufacturer": product.get("legal_manufacturer") or product.get("company", ""),
                "local_holder": "",
                "source_key": "product_completeness_20260601",
                "source_url": item["source_url"],
                "source_type": item["source_type"],
                "evidence_title": item["evidence_title"],
                "evidence_excerpt": item["excerpt"],
                "official_description_exact": item["approved_indication"],
                "official_description_source_field": "intended_use",
                "field_note": "Precise official wording from an official product page or official factsheet. Certificate number remains blank when not publicly disclosed.",
                "checked_at": checked_at,
                "reviewed_by": "codex_product_completeness_qa",
                "review_status": "auto_cross_checked",
                "confidence": "official_product_page_or_factsheet",
            }
        )
    return rows


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    products = product_lookup()

    fact_fields, fact_rows = read_csv(MANUAL_FACT_PATH)
    added_facts = append_unique(fact_rows, "fact_id", build_fact_rows(products, checked_at))
    write_csv(MANUAL_FACT_PATH, fact_fields, fact_rows)

    indication_fields, indication_rows = read_csv(MANUAL_INDICATION_PATH)
    new_indications = build_indication_rows(products, checked_at)
    for row in new_indications:
        row["source_key"] = stable_id("mind", row.get("seed_record_id"), row.get("source_url"), row.get("approved_indication"))
    added_indications = append_unique(indication_rows, "source_key", new_indications)
    write_csv(MANUAL_INDICATION_PATH, indication_fields, indication_rows)

    summary = {
        "source_workbook": str(SOURCE_BOOK),
        "backup": str(backup),
        "workbook_changes": workbook_changes,
        "manual_product_fact_rows_added": added_facts,
        "manual_official_indication_rows_added": added_indications,
    }
    out = AUDIT_DIR / f"product_completeness_cleanup_{stamp}.json"
    latest = AUDIT_DIR / "product_completeness_cleanup_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
