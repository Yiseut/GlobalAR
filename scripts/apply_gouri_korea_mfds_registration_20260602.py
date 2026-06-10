"""Apply user-confirmed Korea MFDS registration for Dexlevo GOURI."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audits"
SOURCE_BOOK = ROOT.parent / "全球医美企业库_标准化版v4.xlsx"
PRODUCT_MASTER_PATH = DATA_DIR / "product_master.csv"
MANUAL_INDICATION_PATH = DATA_DIR / "manual_official_indication_evidence.csv"

TARGET_RECORD_ID = "REC_0307"
TARGET_PRODUCT_ID = "prod_c019a8d54482"

SOURCE_URL = "https://www.thebell.co.kr/front/newsview.asp?key=202605130602069380102883"
SUPPORTING_URL = "https://www.hankookilbo.com/news/article/A2026041310270001876"

APPROVED_INDICATION = (
    "GOURI 是全液态 PCL 胶原刺激剂，不含微粒，注射后可在皮下自然扩散，"
    "用于全面部胶原蛋白新生、提升皮肤弹性、全面部抗衰和肤质改善。"
)


def norm(value: Any) -> str:
    return str(value or "").strip()


def stable_id(prefix: str, *parts: object) -> str:
    blob = "||".join(norm(part).casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(blob.encode('utf-8')).hexdigest()[:12]}"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def headers(ws) -> dict[str, int]:
    return {norm(cell.value): cell.column for cell in ws[1] if norm(cell.value)}


def product_lookup() -> dict[str, str]:
    _, rows = read_csv(PRODUCT_MASTER_PATH)
    for row in rows:
        if norm(row.get("seed_record_id")) == TARGET_RECORD_ID:
            return row
    raise SystemExit(f"Missing target product in product_master.csv: {TARGET_RECORD_ID}")


def append_unique(rows: list[dict[str, str]], key_fields: list[str], row: dict[str, str]) -> bool:
    key = tuple(norm(row.get(field)) for field in key_fields)
    existing = {tuple(norm(existing_row.get(field)) for field in key_fields) for existing_row in rows}
    if key in existing:
        return False
    rows.append(row)
    return True


def update_existing_gouri_notes(rows: list[dict[str, str]]) -> int:
    updated = 0
    replacement = "Korea MFDS/KFDA registration is captured separately in the KR row; CE certificate number not captured."
    stale_note = "Korea MFDS/KFDA status still requires separate confirmation."
    for row in rows:
        if norm(row.get("seed_record_id")) != TARGET_RECORD_ID:
            continue
        if norm(row.get("jurisdiction")) != "EU":
            continue
        if stale_note not in norm(row.get("field_note")):
            continue
        row["field_note"] = replacement
        updated += 1
    return updated


def normalize_existing_gouri_kr_regulator(rows: list[dict[str, str]]) -> int:
    updated = 0
    for row in rows:
        if norm(row.get("seed_record_id")) != TARGET_RECORD_ID:
            continue
        if norm(row.get("jurisdiction")) != "KR":
            continue
        if norm(row.get("regulator")) != "MFDS / KFDA":
            continue
        row["regulator"] = "MFDS"
        updated += 1
    return updated


def update_workbook(stamp: str) -> tuple[Path, list[dict[str, str]]]:
    backup = SOURCE_BOOK.with_name(f"{SOURCE_BOOK.stem}.backup_before_gouri_korea_mfds_{stamp}{SOURCE_BOOK.suffix}")
    shutil.copy2(SOURCE_BOOK, backup)

    wb = load_workbook(SOURCE_BOOK)
    ws = wb["Product_Lines"]
    colmap = headers(ws)
    target_row = None
    for row in range(2, ws.max_row + 1):
        if norm(ws.cell(row=row, column=colmap["Record_ID"]).value) == TARGET_RECORD_ID:
            target_row = row
            break
    if not target_row:
        wb.close()
        raise SystemExit(f"Missing target row in workbook: {TARGET_RECORD_ID}")

    changes: list[dict[str, str]] = []

    def set_cell(field: str, value: str) -> None:
        col = colmap.get(field)
        if not col:
            return
        old = norm(ws.cell(row=target_row, column=col).value)
        if old == value:
            return
        ws.cell(row=target_row, column=col, value=value)
        changes.append({"record_id": TARGET_RECORD_ID, "field": field, "old": old, "new": value})

    def append_audit(note: str) -> None:
        col = colmap.get("Backfill_Audit")
        if not col:
            return
        old = norm(ws.cell(row=target_row, column=col).value)
        if note in old:
            return
        new = f"{old}; {note}".strip("; ")
        ws.cell(row=target_row, column=col, value=new)
        changes.append({"record_id": TARGET_RECORD_ID, "field": "Backfill_Audit", "old": old, "new": new})

    set_cell("KFDA_Status", "KFDA/MFDS registration confirmed by user/public media; certificate number pending.")
    set_cell("KFDA_Approval_Date", "2025-09")
    set_cell("CE_Status", "KFDA/MFDS and CE confirmed; certificate numbers pending.")
    set_cell("Manufactured_By", "Dexlevo")
    set_cell("Marketing_Holder", "Dexlevo")
    append_audit(
        "gouri_korea_mfds_20260602: user/public-media confirmed Korea MFDS registration; "
        "approval month 2025-09, certificate number not captured."
    )

    wb.save(SOURCE_BOOK)
    wb.close()
    return backup, changes


def build_registration_row(product: dict[str, str], checked_at: str) -> dict[str, str]:
    return {
        "product_id": product.get("product_id") or TARGET_PRODUCT_ID,
        "seed_record_id": TARGET_RECORD_ID,
        "company_id": product.get("company_id", ""),
        "company": product.get("company", "Dexlevo"),
        "brand": product.get("brand", "GOURI"),
        "jurisdiction": "KR",
        "regulator": "MFDS",
        "regulatory_pathway": "Medical device registration",
        "status": "KFDA/MFDS registration confirmed by user and public media; certificate number not captured",
        "registration_no": "",
        "approval_date": "2025-09",
        "expiry_date": "",
        "registered_name": "GOURI",
        "approved_indication": APPROVED_INDICATION,
        "intended_use": APPROVED_INDICATION,
        "legal_manufacturer": "Dexlevo",
        "local_holder": "Dexlevo",
        "source_key": stable_id("user_confirmed_gouri_mfds", TARGET_RECORD_ID, "KR", "2025-09"),
        "source_url": SOURCE_URL,
        "source_type": "user_confirmed_public_media_mfds_claim",
        "evidence_title": "thebell report on GOURI Korea domestic approval and launch",
        "evidence_excerpt": (
            "Public media reports state that Dexlevo's GOURI obtained Korea domestic item approval "
            "from MFDS in September 2025 and started domestic commercialization afterward; "
            f"supporting media URL: {SUPPORTING_URL}"
        ),
        "official_description_exact": APPROVED_INDICATION,
        "official_description_source_field": "approved_indication",
        "field_note": "Certificate number not captured; approval month captured as 2025-09 from user-confirmed public reporting.",
        "checked_at": checked_at,
        "reviewed_by": "user_feedback_20260602",
        "review_status": "user_confirmed",
        "confidence": "user_confirmed_public_media",
    }


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    backup, workbook_changes = update_workbook(stamp)
    product = product_lookup()

    fields, rows = read_csv(MANUAL_INDICATION_PATH)
    stale_notes_updated = update_existing_gouri_notes(rows)
    kr_regulator_rows_normalized = normalize_existing_gouri_kr_regulator(rows)
    added = append_unique(
        rows,
        ["seed_record_id", "jurisdiction", "regulator", "registered_name", "source_key"],
        build_registration_row(product, checked_at),
    )
    write_csv(MANUAL_INDICATION_PATH, fields, rows)

    summary = {
        "target_record_id": TARGET_RECORD_ID,
        "target_product_id": TARGET_PRODUCT_ID,
        "backup": str(backup),
        "workbook_changes": workbook_changes,
        "manual_official_indication_rows_added": 1 if added else 0,
        "stale_gouri_ce_notes_updated": stale_notes_updated,
        "kr_regulator_rows_normalized": kr_regulator_rows_normalized,
        "source_url": SOURCE_URL,
        "supporting_url": SUPPORTING_URL,
        "approval_date": "2025-09",
        "status": "applied",
    }
    out = AUDIT_DIR / f"gouri_korea_mfds_registration_{stamp}.json"
    latest = AUDIT_DIR / "gouri_korea_mfds_registration_latest.json"
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
